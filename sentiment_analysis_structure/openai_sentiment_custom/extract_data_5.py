#!/usr/bin/env python3
# dual_output_50tabs_and_20tabs_transpose.py
"""
Creates two Excel files:

1) raw_values_6cat_50tabs_no_blanks.xlsx
   => 50 tabs total (10 sentiments x 5 measures).
      Each tab => no blank rows in each of 6 columns (one column per category).
      (Same as before.)

2) raw_values_6cat_20tabs_intv2_no_blanks_transposed.xlsx
   => 20 tabs total (2 per sentiment).
      For each sentiment S => 2 tabs:

         Tab A => "S_3Int" => 3 intensity measures * 6 categories
            => after transposing, each row is one cat_measure combination,
               and columns are consecutive data points horizontally (no blank row).
         Tab B => "S_2NonInt" => 2 non-intensity measures * 6 categories
            => transposed in the same fashion.

Hence the second file is the same data as the prior "20 tabs" approach, but transposed
so rows represent (category x measure) and columns represent consecutive data points.
"""

import json
import re
import numpy as np
import pandas as pd
from tqdm import tqdm
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# ------------------ GLOBAL CONFIG ------------------
INPUT_JSONL_FILE = "processed_all_articles_fixed_4.jsonl"

OUTPUT_50TABS = "raw_values_6cat_50tabs_no_blanks.xlsx"
OUTPUT_20TABS_TRANSPOSED = "raw_values_6cat_20tabs_intv2_no_blanks_transposed.xlsx"

# 6 categories
MEDIA_CATEGORIES = {
    "Scientific": ["nature","sciam","stat","newscientist"],
    "Left": ["theatlantic","the daily beast","the intercept","mother jones","msnbc","slate","vox","huffpost"],
    "Lean Left": ["ap","axios","cnn","guardian","business insider","nbcnews",
                  "npr","nytimes","politico","propublica","wapo","usa today"],
    "Center": ["reuters","marketwatch","financial times","newsweek","forbes"],
    "Lean Right": ["thedispatch","epochtimes","foxbusiness","wsj","national review","washtimes"],
    "Right": ["breitbart","theblaze","daily mail","dailywire","foxnews","nypost","newsmax"],
}

SENTIMENTS = [
    "joy","sadness","anger","fear",
    "surprise","disgust","trust","anticipation",
    "negative_sentiment","positive_sentiment"
]

MEASURES_5 = [  # used for the 50-tab file
    "Fulltext",
    "Quotation",
    "Fulltext_Intensity",
    "Quotation_Intensity",
    "Title_Intensity",
]

# For the 20-tab version
THREE_INTENSITIES = ["Fulltext_Intensity", "Quotation_Intensity", "Title_Intensity"]
TWO_NONINTENSITIES= ["Fulltext", "Quotation"]


# ------------------ 1) LOAD + MAP ------------------
def load_jsonl(jsonl_file):
    records=[]
    with open(jsonl_file, "r", encoding="utf-8") as f:
        for line in tqdm(f, desc="Loading JSONL"):
            rec=json.loads(line)
            records.append(rec)
    return pd.DataFrame(records)

def map_outlet_to_6cats(df):
    cat_map={}
    for cat, outls in MEDIA_CATEGORIES.items():
        for o in outls:
            cat_map[o.lower().strip()] = cat
    df2=df.copy()
    df2["media_outlet_clean"] = df2["media_outlet"].str.lower().str.strip()
    df2["media_category_6"]   = df2["media_outlet_clean"].map(cat_map).fillna("Other")
    return df2

# ------------------ 2) EXTRACT MEASURE VALUE ------------------
def extract_measure_value(row, sentiment, measure):
    """
    - Quotation => average columns ^{sentiment}_\d+$
    - Quotation_Intensity => average columns ^{sentiment}_\d+_intensity$
    - Fulltext => row[f"{sentiment}_fulltext"]
    - Fulltext_Intensity => row[f"{sentiment}_fulltext_intensity"]
    - Title_Intensity => row[f"{sentiment}_title_intensity"]
    Returns float or np.nan
    """
    try:
        if measure=="Quotation":
            pat = rf"^{re.escape(sentiment)}_\d+$"
            matched = [c for c in row.index if re.match(pat,c)]
            if not matched:
                return np.nan
            vals = row[matched].astype(float).clip(lower=0)
            return vals.mean(skipna=True)

        elif measure=="Quotation_Intensity":
            pat = rf"^{re.escape(sentiment)}_\d+_intensity$"
            matched = [c for c in row.index if re.match(pat,c)]
            if not matched:
                return np.nan
            vals = row[matched].astype(float).clip(lower=0)
            return vals.mean(skipna=True)

        elif measure=="Fulltext":
            col = f"{sentiment}_fulltext"
            if col not in row.index:
                return np.nan
            return float(row[col])

        elif measure=="Fulltext_Intensity":
            col = f"{sentiment}_fulltext_intensity"
            if col not in row.index:
                return np.nan
            return float(row[col])

        elif measure=="Title_Intensity":
            col = f"{sentiment}_title_intensity"
            if col not in row.index:
                return np.nan
            return float(row[col])

        else:
            return np.nan
    except:
        return np.nan


# ------------------ UTIL: Build no-blank col for (category, measure) ------------------
def build_no_blank_cols_1measure(df, sentiment, measure):
    """
    Returns a DataFrame of shape (max_len, 6).
    Columns => [Scientific, Left, Lean Left, Center, Lean Right, Right]
    Each column => consecutive measure values for that category, skipping NaN.
    """
    cat_lists = {
        "Scientific": [],
        "Left": [],
        "Lean Left": [],
        "Center": [],
        "Lean Right": [],
        "Right": []
    }
    for i, row in df.iterrows():
        cat_6 = row["media_category_6"]
        if cat_6 not in cat_lists:
            continue
        val = extract_measure_value(row, sentiment, measure)
        if pd.notna(val):
            cat_lists[cat_6].append(val)

    max_len = max(len(lst) for lst in cat_lists.values()) if cat_lists else 0
    out_cols = ["Scientific","Left","Lean Left","Center","Lean Right","Right"]
    data_dict = {}
    for c in out_cols:
        col_vals=cat_lists[c]
        if len(col_vals)<max_len:
            col_vals = col_vals + [np.nan]*(max_len-len(col_vals))
        data_dict[c] = col_vals
    df_out = pd.DataFrame(data_dict, columns=out_cols)
    return df_out

# ------------------ PART A: Create 50 tabs file -----------
def create_50tabs_file(df, wb):
    # remove default "Sheet"
    if "Sheet" in wb.sheetnames and len(wb.sheetnames)==1:
        if wb.sheetnames[0]=="Sheet":
            wb.remove(wb["Sheet"])

    for sentiment in SENTIMENTS:
        for measure in MEASURES_5:
            tab_name = f"{sentiment[:10]}_{measure[:10]}"
            subdf = build_no_blank_cols_1measure(df, sentiment, measure)
            # Write to sheet
            ws = wb.create_sheet(title=tab_name)
            # first row => col names
            col_names = list(subdf.columns)
            # we build row-of-rows
            rows_list = [col_names]
            for row_idx in range(len(subdf)):
                row_vals = []
                for c in col_names:
                    row_vals.append(subdf.loc[row_idx,c])
                rows_list.append(row_vals)

            for r_idx, row_vals in enumerate(rows_list,1):
                for c_idx, cell_val in enumerate(row_vals,1):
                    ws.cell(row=r_idx,column=c_idx,value=cell_val)

# ------------------ PART B: 20 tabs, transposed --------------
def build_no_blank_colset_for_multiple_measures_transposed(df, sentiment, measure_list):
    """
    For the "20-tab" version, we want to gather multiple measures for each category,
    but then TRANSPOSE so each row = (category_measure) and columns = consecutive data points.

    measure_list might be [Fulltext_Intensity, Quotation_Intensity, Title_Intensity]
    or [Fulltext, Quotation].
    We'll produce 6 x len(measure_list) rows in the final transposed DataFrame.
    """
    cat_order = ["Scientific","Left","Lean Left","Center","Lean Right","Right"]
    # We'll store lists in a dict => (cat, measure) => list of consecutive values
    # key as cat_meas = f"{cat}_{meas}"
    col_lists = {}
    for cat in cat_order:
        for meas in measure_list:
            col_lists[f"{cat}_{meas}"] = []

    # gather consecutive values
    for i, row in df.iterrows():
        cat_6 = row["media_category_6"]
        if cat_6 not in cat_order:
            continue
        for meas in measure_list:
            val = extract_measure_value(row, sentiment, meas)
            if pd.notna(val):
                col_key = f"{cat_6}_{meas}"
                col_lists[col_key].append(val)

    # find max_len among them
    max_len = max(len(lst) for lst in col_lists.values()) if col_lists else 0
    # build a DataFrame => shape (max_len, #keys)
    # then transpose
    out_keys = []
    for cat in cat_order:
        for m in measure_list:
            out_keys.append(f"{cat}_{m}")

    data_dict = {}
    for k in out_keys:
        arr = col_lists[k]
        if len(arr)<max_len:
            arr = arr + [np.nan]*(max_len-len(arr))
        data_dict[k] = arr

    df_wide = pd.DataFrame(data_dict, columns=out_keys)  # shape => (max_len, #out_keys)
    # Now transpose => shape => (#out_keys, max_len)
    df_t = df_wide.T.reset_index(drop=False)
    # The index column is e.g. "Scientific_Fulltext_Intensity", etc.
    # We'll rename that index col => "CatMeasure"
    df_t.rename(columns={"index":"CatMeasure"}, inplace=True)
    # So df_t => columns => ["CatMeasure", 0,1,2,3,... up to max_len-1]
    return df_t

def create_20tabs_file_transposed(df, wb_20):
    """
    For each sentiment => 2 tabs => "S_3Int" and "S_2NonInt"
    But transposed => each row = (cat_measure), each column => consecutive data points
    """
    if "Sheet" in wb_20.sheetnames and len(wb_20.sheetnames)==1:
        if wb_20.sheetnames[0]=="Sheet":
            wb_20.remove(wb_20["Sheet"])

    for sentiment in SENTIMENTS:
        # A) 3 intensities
        tabA = f"{sentiment[:10]}_3Int"
        df_A = build_no_blank_colset_for_multiple_measures_transposed(df, sentiment, THREE_INTENSITIES)
        wsA  = wb_20.create_sheet(title=tabA)
        # write
        col_names = list(df_A.columns)  # e.g. ["CatMeasure", 0,1,2, ...]
        row_data = [col_names]
        for row_idx in range(len(df_A)):
            row_list = []
            for c in col_names:
                row_list.append(df_A.loc[row_idx,c])
            row_data.append(row_list)
        for r_idx, row_vals in enumerate(row_data,1):
            for c_idx, val in enumerate(row_vals,1):
                wsA.cell(row=r_idx,column=c_idx,value=val)

        # B) 2 non-intensities
        tabB = f"{sentiment[:10]}_2NonInt"
        df_B = build_no_blank_colset_for_multiple_measures_transposed(df, sentiment, TWO_NONINTENSITIES)
        wsB  = wb_20.create_sheet(title=tabB)
        col_namesB = list(df_B.columns)
        row_dataB  = [col_namesB]
        for row_idx in range(len(df_B)):
            row_list=[]
            for c in col_namesB:
                row_list.append(df_B.loc[row_idx,c])
            row_dataB.append(row_list)
        for r_idx, row_vals in enumerate(row_dataB,1):
            for c_idx, val in enumerate(row_vals,1):
                wsB.cell(row=r_idx,column=c_idx,value=val)

# ------------------ MAIN ------------------
def main():
    print("Loading + mapping data...")
    df = load_jsonl(INPUT_JSONL_FILE)
    df = map_outlet_to_6cats(df)
    if "article_id" not in df.columns:
        df["article_id"] = df.index + 1

    # 1) Create the 50-tabs file => same as before
    print(f"Creating 50-tab file => {OUTPUT_50TABS}")
    wb_50 = Workbook()
    create_50tabs_file(df, wb_50)
    wb_50.save(OUTPUT_50TABS)
    print(f"Wrote => {OUTPUT_50TABS}")

    # 2) Create the 20-tabs transposed file
    print(f"Creating 20-tab transposed file => {OUTPUT_20TABS_TRANSPOSED}")
    wb_20 = Workbook()
    create_20tabs_file_transposed(df, wb_20)
    wb_20.save(OUTPUT_20TABS_TRANSPOSED)
    print(f"Wrote => {OUTPUT_20TABS_TRANSPOSED}")

    print("All done.")

if __name__=="__main__":
    main()
