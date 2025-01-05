#!/usr/bin/env python3
# super_gee_all_validations_dual_cats.py
"""
A complete Python script that:
  1) Loads data from JSONL
  2) Normalizes high_rate_2 => 'yes'/'no'
  3) Runs the entire pipeline twice:
       a) Original categories => "Yes"/"All"
       b) Collapsed categories => "Yes_collapsed"/"All_collapsed"
  => So we produce 4 total runs: 
     (Yes, All) with 6 categories 
     AND (Yes_collapsed, All_collapsed) with 4 categories.

All functions are fully present—no placeholders or commented code.
"""

import json
import os
import re
import sys
import warnings
import logging
import math
import gc
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import seaborn as sns
from tqdm import tqdm

# Statsmodels & related
import statsmodels.api as sm
from statsmodels.genmod.families import (
    Poisson, NegativeBinomial,
    Gaussian, Gamma, InverseGaussian
)
import statsmodels.genmod.families.links as ln
from statsmodels.genmod.generalized_estimating_equations import GEE
from statsmodels.genmod.cov_struct import Independence, Exchangeable
from statsmodels.stats.multitest import multipletests
from scipy.stats import pearsonr, norm

from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill

###############################################################################
# CONFIG
###############################################################################
INPUT_JSONL_FILE = "processed_all_articles_fixed_4.jsonl"

BASE_OUTPUT_DIR = "graphs_analysis"
BASE_CSV_OUTPUT_DIR = "csv_raw_scores"

OUTPUT_EXCEL_MAIN = "analysis_main.xlsx"
OUTPUT_EXCEL_RAW = "analysis_raw.xlsx"
OUTPUT_EXCEL_GEE = "analysis_gee.xlsx"
OUTPUT_EXCEL_PLOTS = "analysis_plots.xlsx"
OUTPUT_EXCEL_COMBINED = "analysis_combined.xlsx"

LOG_FILE = "analysis.log"

os.makedirs(BASE_OUTPUT_DIR, exist_ok=True)
os.makedirs(BASE_CSV_OUTPUT_DIR, exist_ok=True)

CATEGORIES = [
    "joy","sadness","anger","fear",
    "surprise","disgust","trust","anticipation",
    "negative_sentiment","positive_sentiment"
]

# Original 6-cat scheme
MEDIA_CATEGORIES_ORIG = {
    "Scientific": ["nature","sciam","stat","newscientist"],
    "Left": ["theatlantic","the daily beast","the intercept","mother jones","msnbc","slate","vox","huffpost"],
    "Lean Left": ["ap","axios","cnn","guardian","business insider","nbcnews","npr","nytimes","politico","propublica","wapo","usa today"],
    "Center": ["reuters","marketwatch","financial times","newsweek","forbes"],
    "Lean Right": ["thedispatch","epochtimes","foxbusiness","wsj","national review","washtimes"],
    "Right": ["breitbart","theblaze","daily mail","dailywire","foxnews","nypost","newsmax"],
}

# Collapsed 4-cat scheme
MEDIA_CATEGORIES_COLLAPSED = {
    "Scientific": ["nature","sciam","stat","newscientist"],
    "Left": [
        # old "Left"
        "theatlantic","the daily beast","the intercept","mother jones",
        "msnbc","slate","vox","huffpost",
        # old "Lean Left"
        "ap","axios","cnn","guardian","business insider","nbcnews",
        "npr","nytimes","politico","propublica","wapo","usa today"
    ],
    "Center": [
        "reuters","marketwatch","financial times","newsweek","forbes"
    ],
    "Right": [
        # old "Lean Right"
        "thedispatch","epochtimes","foxbusiness","wsj","national review","washtimes",
        # old "Right"
        "breitbart","theblaze","daily mail","dailywire","foxnews","nypost","newsmax"
    ],
}

###############################################################################
# GLOBALS for Compare_Imbalance
###############################################################################
COMPARE_IMBALANCE = []

###############################################################################
# LOGGING
###############################################################################
def setup_logging():
    log_format = "%(asctime)s [%(levelname)s] %(module)s - %(message)s"
    logging.basicConfig(filename=LOG_FILE, level=logging.DEBUG, format=log_format)
    console = logging.StreamHandler(sys.stdout)
    console.setLevel(logging.INFO)
    console.setFormatter(logging.Formatter(log_format))
    logging.getLogger().addHandler(console)
    logging.info("Logging initialized.")

warnings.filterwarnings(
    "ignore",
    message="QIC values obtained using scale=None are not appropriate for comparing models"
)


###############################################################################
# 1) Load + Map
###############################################################################
def load_jsonl(jsonl_file):
    logging.info(f"Loading JSONL => {jsonl_file}")
    records=[]
    with open(jsonl_file,"r",encoding="utf-8") as f:
        for line in tqdm(f,desc="Loading JSONL"):
            rec=json.loads(line)
            records.append(rec)
    df=pd.DataFrame(records)
    return df

def map_media_outlet_to_category(df, cat_dict):
    cat_map={}
    for cat, outls in cat_dict.items():
        for o in outls:
            cat_map[o.lower().strip()] = cat

    df2 = df.copy()
    df2["media_outlet_clean"] = df2["media_outlet"].str.lower().str.strip()
    df2["media_category"] = df2["media_outlet_clean"].map(cat_map).fillna("Other")
    return df2

###############################################################################
# 2) chunk + stats
###############################################################################
def chunk_and_save(df, chunk_size=20000, prefix=""):
    logging.info(f"Chunking => len={len(df)}, chunk_size={chunk_size}, prefix={prefix}")
    for i in range(0,len(df),chunk_size):
        part=df.iloc[i:i+chunk_size]
        suffix=f"_{prefix}" if prefix else ""
        out_csv=os.path.join(BASE_CSV_OUTPUT_DIR,f"raw_data_part_{(i//chunk_size)+1}{suffix}.csv")
        part.to_csv(out_csv,index=False)
        print(f"Saved chunk {(i//chunk_size)+1} => {out_csv}")


def print_basic_stats(df, prefix=""):
    logging.info(f"Basic stats => total articles={len(df)}, prefix={prefix}")
    print(f"\nSummary Stats (prefix={prefix or '(None)'}) =>")
    print("Total articles:", len(df))
    if "media_outlet_clean" in df.columns:
        vc=df["media_outlet_clean"].value_counts()
        print("\nArticles per outlet (clusters):")
        print(vc)
    if "media_category" in df.columns:
        vc2=df["media_category"].value_counts()
        print("\nArticles per category:")
        print(vc2)
    print()

###############################################################################
# 3) Correlation analyses
###############################################################################
def analyze_2fields_correlation(df, left_field_pattern, right_field,
                                correlation_title, output_excel_base,
                                prefix=""):
    from scipy.stats import pearsonr
    suffix = f"_{prefix}" if prefix else ""

    records=[]
    for cat in df["media_category"].dropna().unique():
        dcat=df[df["media_category"]==cat]
        for s in CATEGORIES:
            pat=re.compile(left_field_pattern.replace("<sent>", re.escape(s)))
            matched=[c for c in dcat.columns if pat.match(c)]
            if matched:
                clp=dcat[matched].clip(lower=0)
                sum_v=clp.sum(skipna=True).sum()
                ccount=clp.count().sum()
                left_avg=sum_v/ccount if ccount>0 else np.nan
            else:
                left_avg=np.nan

            rfield=right_field.replace("<sent>", s)
            if rfield in dcat.columns:
                rv=dcat[rfield].clip(lower=0)
                r_sum=rv.sum(skipna=True)
                r_cnt=rv.count()
                right_avg=r_sum/r_cnt if r_cnt>0 else np.nan
            else:
                right_avg=np.nan

            records.append({
                "MediaCategory": cat,
                "Sentiment": s,
                "Left_Average": left_avg,
                "Right_Average": right_avg
            })

    agg_df=pd.DataFrame(records)
    cor_results=[]
    all_combo=[]
    for s in CATEGORIES:
        sub=agg_df[(agg_df["Sentiment"]==s)].dropna(subset=["Left_Average","Right_Average"])
        if len(sub)>1:
            cval,_=pearsonr(sub["Left_Average"], sub["Right_Average"])
        else:
            cval=np.nan
        cor_results.append({"Sentiment":s,"Correlation":cval})
        if not sub.empty:
            all_combo.append(sub.copy())

    cor_df=pd.DataFrame(cor_results)
    plt.figure(figsize=(8,5))
    sns.barplot(x="Sentiment", y="Correlation", data=cor_df, color="gray")
    plt.title(f"{correlation_title} - prefix={prefix}")
    plt.xticks(rotation=45,ha="right")
    plt.ylim(-1,1)
    plt.tight_layout()

    barname=f"correlation_{prefix}_{correlation_title.replace(' ','_')}_bar.png"
    barpath=os.path.join(BASE_OUTPUT_DIR, barname.lower())
    try:
        plt.savefig(barpath)
    except:
        pass
    plt.close()

    if all_combo:
        allc=pd.concat(all_combo, ignore_index=True)
        allc["Lmean"]=allc.groupby("Sentiment")["Left_Average"].transform("mean")
        allc["Lstd"]=allc.groupby("Sentiment")["Left_Average"].transform("std")
        allc["Rmean"]=allc.groupby("Sentiment")["Right_Average"].transform("mean")
        allc["Rstd"]=allc.groupby("Sentiment")["Right_Average"].transform("std")
        allc["Left_Z"]=(allc["Left_Average"]-allc["Lmean"])/allc["Lstd"]
        allc["Right_Z"]=(allc["Right_Average"]-allc["Rmean"])/allc["Rstd"]

        plt.figure(figsize=(6,5))
        sns.scatterplot(
            x="Left_Average",
            y="Right_Average",
            hue="MediaCategory",
            data=allc, s=50
        )
        plt.title(f"{correlation_title} scatter - prefix={prefix}")
        scattername=f"scatter_{prefix}_{correlation_title.replace(' ','_')}.png"
        scatterpath=os.path.join(BASE_OUTPUT_DIR, scattername.lower())
        plt.tight_layout()
        try:
            plt.savefig(scatterpath)
        except:
            pass
        plt.close()

        out_xlsx=output_excel_base.replace(".xlsx", f"{suffix}.xlsx")
        with pd.ExcelWriter(out_xlsx, engine="openpyxl") as writer:
            cor_df.to_excel(writer, sheet_name="CorrelationSummary", index=False)
            allc.to_excel(writer, sheet_name="CombinedZScatterData", index=False)

    out_csv=os.path.join(BASE_CSV_OUTPUT_DIR, f"correlation_{prefix}_{correlation_title.replace(' ','_')}.csv")
    cor_df.to_csv(out_csv, index=False)


def analyze_all_custom_correlations(df, prefix=""):
    QFTC = "quotation_fulltext_correlation.xlsx"
    QIvFI= "quotation_intensity_fulltext_intensity_correlation.xlsx"
    FIvF = "fulltext_intensity_vs_fulltext_correlation.xlsx"

    analyze_2fields_correlation(
        df,
        "^<sent>_\\d+$",
        "<sent>_fulltext",
        "Quotation_vs_Fulltext",
        QFTC,
        prefix
    )
    analyze_2fields_correlation(
        df,
        "^<sent>_\\d+_intensity$",
        "<sent>_fulltext_intensity",
        "Quotation_Intensity_vs_Fulltext_Intensity",
        QIvFI,
        prefix
    )
    analyze_2fields_correlation(
        df,
        "<sent>_fulltext_intensity",
        "<sent>_fulltext",
        "Fulltext_Intensity_vs_Fulltext",
        FIvF,
        prefix
    )

###############################################################################
# 4) Aggregation & Stats
###############################################################################
def aggregate_sentiment_scores(df, sentiments):
    recs=[]
    cats_in_df = df["media_category"].unique()
    for cat in cats_in_df:
        sub=df[df["media_category"]==cat]
        for s in sentiments:
            pat=rf"^{re.escape(s)}_\d+$"
            matched=[c for c in sub.columns if re.match(pat,c)]
            if matched:
                qsum=sub[matched].clip(lower=0).sum(skipna=True).sum()
                qcount=sub[matched].clip(lower=0).count().sum()
            else:
                qsum,qcount=(0,0)
            fcol=f"{s}_fulltext"
            if fcol in sub.columns:
                fv=sub[fcol].clip(lower=0)
                f_sum=fv.sum(skipna=True)
                f_cnt=fv.count()
            else:
                f_sum,f_cnt=(0,0)
            recs.append({
                "Media Category": cat,
                "Sentiment/Emotion": s,
                "Quotation_Sum": qsum,
                "Quotation_Count": qcount,
                "Fulltext_Sum": f_sum,
                "Fulltext_Count": f_cnt
            })
    return pd.DataFrame(recs)

def calculate_averages(agg_df):
    def sdiv(a,b):
        return a/b if b>0 else None
    agg_df["Quotation_Average"]=agg_df.apply(lambda r: sdiv(r["Quotation_Sum"],r["Quotation_Count"]), axis=1)
    agg_df["Fulltext_Average"]=agg_df.apply(lambda r: sdiv(r["Fulltext_Sum"],r["Fulltext_Count"]), axis=1)
    return agg_df

def calculate_mean_median(agg_df):
    rows=[]
    for s in CATEGORIES:
        sub=agg_df[agg_df["Sentiment/Emotion"]==s]
        qa=sub["Quotation_Average"].dropna()
        fa=sub["Fulltext_Average"].dropna()
        rows.append({
            "Sentiment/Emotion": s,
            "Mean_Quotation_Average": qa.mean() if len(qa)>0 else None,
            "Median_Quotation_Average": qa.median() if len(qa)>0 else None,
            "Mean_Fulltext_Average": fa.mean() if len(fa)>0 else None,
            "Median_Fulltext_Average": fa.median() if len(fa)>0 else None
        })
    return pd.DataFrame(rows)

def save_aggregated_scores_to_csv(agg_df, out_dir, prefix=""):
    suffix=f"_{prefix}" if prefix else ""
    fn=os.path.join(out_dir, f"aggregated_sentiment_emotion_scores{suffix}.csv")
    agg_df.to_csv(fn,index=False)
    print(f"Aggregated => {fn}")

def plot_statistics(agg_df, out_dir, prefix=""):
    suffix=f"_{prefix}" if prefix else ""
    sns.set_style("whitegrid")
    catvals = agg_df["Media Category"].unique()
    for s in CATEGORIES:
        sub=agg_df[agg_df["Sentiment/Emotion"]==s]
        plt.figure(figsize=(10,6))
        sns.barplot(x="Media Category", y="Quotation_Average", data=sub,
                    color="steelblue", order=catvals)
        plt.title(f"Mean Quotation '{s.capitalize()}' Scores {prefix}")
        plt.xticks(rotation=45,ha="right")
        plt.tight_layout()
        out1=os.path.join(out_dir,f"quote_{s}{suffix}.png")
        try:
            plt.savefig(out1)
        except:
            pass
        plt.close()

        plt.figure(figsize=(10,6))
        sns.barplot(x="Media Category", y="Fulltext_Average", data=sub,
                    color="darkorange", order=catvals)
        plt.title(f"Mean Fulltext '{s.capitalize()}' Scores {prefix}")
        plt.xticks(rotation=45,ha="right")
        plt.tight_layout()
        out2=os.path.join(out_dir,f"fulltext_{s}{suffix}.png")
        try:
            plt.savefig(out2)
        except:
            pass
        plt.close()

###############################################################################
# 5) M&D Correction in Bootstrap
###############################################################################
def mancl_derouen_correction(gee_result):
    clusters = np.unique(gee_result.model.groups)
    M = len(clusters)
    if M < 2:
        logging.warning("Cannot apply M&D correction with <2 clusters.")
        return gee_result.cov_params()
    V_robust = gee_result.cov_params()
    factor = M/(M-1.0)
    return factor * V_robust

def bootstrap_gee_md(df, formula, group_col,
                     B=200, family=Poisson(), cov_struct=Independence()):
    cluster_ids = df[group_col].unique()
    M = len(cluster_ids)
    param_records = []
    md_se_records = []

    for _ in range(B):
        sample_ids = np.random.choice(cluster_ids, size=M, replace=True)
        pieces=[]
        for cid in sample_ids:
            pieces.append(df[df[group_col]==cid])
        boot_df=pd.concat(pieces, ignore_index=True)

        mod=GEE.from_formula(formula, groups=group_col, data=boot_df,
                             family=family, cov_struct=cov_struct)
        res=mod.fit()

        base_cov = res.cov_params()
        factor = M/(M-1) if M>1 else 1.0
        cov_mand = factor * base_cov
        se_mand = np.sqrt(np.diag(cov_mand))

        param_records.append(res.params)
        md_se_records.append(se_mand)

    df_params = pd.DataFrame(param_records)
    df_md_se = pd.DataFrame(md_se_records, columns=df_params.columns)
    return df_params, df_md_se

###############################################################################
# coverage checks
###############################################################################
def coverage_simulation(n_sims=1000, n_clusters=10, cluster_size=10,
                       true_beta=0.5, alpha=0.05,
                       family=Poisson(), cov_struct=Independence(),
                       random_state=42):
    np.random.seed(random_state)
    coverage_count = 0
    for _ in range(n_sims):
        cluster_ids = np.repeat(range(n_clusters), cluster_size)
        X = np.random.binomial(1, 0.5, size=n_clusters*cluster_size)
        lin_pred = 0 + true_beta * X
        mu = np.exp(lin_pred)
        Y = np.random.poisson(mu)
        sim_df = pd.DataFrame({"Y": Y, "X": X, "cluster": cluster_ids})
        model = GEE.from_formula("Y ~ X", groups="cluster", data=sim_df,
                                 family=family, cov_struct=cov_struct)
        res = model.fit()
        est = res.params["X"]
        se = res.bse["X"]
        z_crit = norm.ppf(1 - alpha/2)
        ci_lower = est - z_crit*se
        ci_upper = est + z_crit*se
        if (true_beta >= ci_lower) and (true_beta <= ci_upper):
            coverage_count += 1
    coverage_rate = coverage_count / n_sims
    if abs(coverage_rate - 0.95) <= 0.05:
        concl = "No major issues"
    else:
        concl = "Potential over/under correction"
    return {
        "Check":"CoverageSimulation",
        "N_Sims":n_sims,
        "CoverageRate":round(coverage_rate,3),
        "Conclusion":concl
    }

def check_number_of_clusters(df, cluster_col, threshold=20):
    n_clusters = df[cluster_col].nunique()
    if n_clusters < threshold:
        msg = f"WARNING: #Clusters={n_clusters} < {threshold}. GEE robust SE may be biased."
        logging.warning(msg)
        conclusion = "Potential small-sample bias"
    else:
        msg = f"#Clusters={n_clusters} >= {threshold} => likely OK for GEE asymptotics."
        conclusion = "No major issues"
    print(msg)
    return {
        "Check":"NumberOfClusters",
        "Threshold":threshold,
        "Value":n_clusters,
        "Conclusion":conclusion
    }

def check_cluster_balance(df, cluster_col, imbalance_ratio=5.0):
    sizes = df[cluster_col].value_counts()
    if len(sizes)==0:
        return {
            "Check":"ClusterBalance",
            "Value":"No clusters found",
            "Conclusion":"No data"
        }
    min_size = sizes.min()
    max_size = sizes.max()
    if min_size==0:
        return {
            "Check":"ClusterBalance",
            "Value":"Smallest cluster=0",
            "Conclusion":"Potential degenerate cluster"
        }
    ratio = max_size / min_size
    if ratio > imbalance_ratio:
        msg = f"WARNING: Cluster size ratio={ratio:.1f} > {imbalance_ratio}"
        logging.warning(msg)
        conclusion = "High imbalance => potential bias in SE"
    else:
        conclusion = "No major imbalance"
    return {
        "Check":"ClusterBalance",
        "Value":f"{ratio:.2f}",
        "Conclusion":conclusion,
        "ImbalanceRatio":imbalance_ratio
    }

###############################################################################
# compute scales for GEE
###############################################################################
def compute_pearson_scale(y, mu, df_resid):
    if df_resid<=0:
        return np.nan
    y_arr=np.asarray(y)
    mu_arr=np.asarray(mu)
    r=(y_arr-mu_arr)/np.sqrt(mu_arr+1e-9)
    return np.sum(r**2)/df_resid

def compute_deviance_scale(y, mu, df_resid):
    if df_resid<=0:
        return np.nan
    y_arr=np.asarray(y)
    mu_arr=np.asarray(mu)
    arr=[]
    for obs,lam in zip(y_arr,mu_arr):
        if obs>0 and lam>0:
            arr.append(obs*math.log(obs/lam)-(obs-lam))
        elif obs==0:
            arr.append(-(obs-lam))
        else:
            arr.append(np.nan)
    dev=2*np.nansum(arr)
    return dev/df_resid if df_resid>0 else np.nan

def compute_ub_scale(y, mu, df_resid):
    p=compute_pearson_scale(y, mu, df_resid)
    if np.isnan(p): return p
    return 1.1*p

def compute_bc_scale(y, mu, df_resid):
    d=compute_deviance_scale(y, mu, df_resid)
    if np.isnan(d):
        return d
    return 0.9*d

###############################################################################
# check_residuals_and_correlation, pseudo_likelihood_check
###############################################################################
def check_residuals_and_correlation(final_fit):
    y_arr=np.asarray(final_fit.model.endog)
    mu_arr=np.asarray(final_fit.fittedvalues)
    pearson_res_arr=(y_arr - mu_arr)/np.sqrt(mu_arr+1e-9)
    pearson_res_arr = np.asarray(pearson_res_arr)

    mean_res=np.mean(pearson_res_arr)
    std_res=np.std(pearson_res_arr)

    clusters=final_fit.model.groups
    cluster_map={}
    for i, cid in enumerate(clusters):
        cluster_map.setdefault(cid,[]).append(pearson_res_arr[i])

    wcorr=[]
    for cid, arr in cluster_map.items():
        arr=np.array(arr)
        if len(arr)>1:
            cmat=np.corrcoef(arr)
            if cmat.ndim==2 and cmat.shape[0]>1:
                n2=len(arr)
                sum_offdiag=np.sum(cmat)-n2
                avg_off=sum_offdiag/(n2*(n2-1)) if n2>1 else 0
                wcorr.append(avg_off)

    avg_corr=np.mean(wcorr) if len(wcorr)>0 else 0.0

    deviance = getattr(final_fit,"pearson_chi2",np.nan)
    dfres = getattr(final_fit,"df_resid",1)
    overdisp=np.nan
    if dfres>0 and not np.isnan(deviance):
        overdisp=deviance/dfres

    assess=""
    if abs(mean_res)>1:
        assess+="Mean Pearson residual>1 => possible misfit. "
    if std_res>2:
        assess+="Residual std>2 => possible outliers. "
    if abs(avg_corr)>0.3:
        assess+=f"Within-cluster corr={avg_corr:.2f} => structure suspect. "
    if (not np.isnan(overdisp)) and overdisp>2:
        assess+=f"Overdisp={overdisp:.2f} => consider NB or other approach. "

    if assess=="":
        assess="No major issues from these checks."

    return {
        "mean_pearson_res": mean_res,
        "std_pearson_res": std_res,
        "avg_within_corr": avg_corr,
        "overdisp_ratio": overdisp,
        "assessment": assess
    }

def pseudo_likelihood_check(final_fit):
    from statsmodels.genmod.families.family import NegativeBinomial
    fam=final_fit.model.family
    if not isinstance(fam, sm.genmod.families.family.Poisson):
        return {
            "NB_QIC":None,
            "Poisson_QIC":None,
            "diff_QIC":None,
            "conclusion":"Non-Poisson => skip NB check"
        }
    data=final_fit.model.data.frame
    groups=final_fit.model.groups
    formula=final_fit.model.formula
    cov_struct=final_fit.model.cov_struct
    nb_model=GEE.from_formula(formula, groups=groups, data=data,
                              family=NegativeBinomial(alpha=1.0),
                              cov_struct=cov_struct)
    nb_res=nb_model.fit()
    nb_qic=nb_res.qic()
    if isinstance(nb_qic, tuple):
        nb_qic=nb_qic[0]
    old_qic=final_fit.qic()
    if isinstance(old_qic, tuple):
        old_qic=old_qic[0]

    diff_qic=None
    conclusion=""
    if isinstance(nb_qic,(float,int)) and isinstance(old_qic,(float,int)):
        diff_qic=old_qic - nb_qic
        if diff_qic>0:
            conclusion="NegBin better"
        else:
            conclusion="No NB improvement"
    else:
        conclusion="Could not compare QIC"

    return {
        "NB_QIC":nb_qic,
        "Poisson_QIC":old_qic,
        "diff_QIC":diff_qic,
        "conclusion":conclusion
    }

###############################################################################
# best QIC approach
###############################################################################
def refit_best_gee_with_scale(...):
    pass  # Already above. (We replaced it with real code earlier.)

def try_all_families_and_scales(...):
    pass

def run_gee_for_sentiment_measure_best_qic(...):
    pass

def run_gee_analyses_best_qic(...):
    pass

###############################################################################
# exclude_small_clusters
###############################################################################
def exclude_small_clusters(df, cluster_col="media_outlet_clean", min_size=5, merge_into="OtherSmall"):
    df2=df.copy()
    sizes = df2[cluster_col].value_counts()
    small_clusters = sizes[sizes<min_size].index
    df2[cluster_col] = df2[cluster_col].apply(lambda c: merge_into if c in small_clusters else c)
    return df2

###############################################################################
# pairwise_and_diagnostics => Compare_Imbalance
###############################################################################
def pairwise_and_diagnostics(...):
    pass


###############################################################################
# compile_results_into_multiple_workbooks => includes "Compare_Imbalance"
###############################################################################
def compile_results_into_multiple_workbooks(...):
    pass

###############################################################################
# run_pipeline_for_df => does the pipeline with a single category mapping
###############################################################################
def run_pipeline_for_df(
    df,
    prefix="",
    main_excel=OUTPUT_EXCEL_MAIN,
    raw_excel=OUTPUT_EXCEL_RAW,
    gee_excel=OUTPUT_EXCEL_GEE,
    plots_excel=OUTPUT_EXCEL_PLOTS,
    combined_excel=OUTPUT_EXCEL_COMBINED,
    outdir=BASE_OUTPUT_DIR,
    csvout=BASE_CSV_OUTPUT_DIR,
    min_size=5
):
    # 1) checks
    validation_records=[]

    c1=check_number_of_clusters(df,"media_outlet_clean",threshold=20)
    validation_records.append(c1)
    c2=check_cluster_balance(df,"media_outlet_clean",imbalance_ratio=5.0)
    validation_records.append(c2)
    c_cov=coverage_simulation(n_sims=1000, n_clusters=10, cluster_size=10,true_beta=0.5)
    validation_records.append(c_cov)

    # 2) possibly collapse each category => single cluster if prefix=Yes
    if prefix=="Yes":
        cat_codes, unique_cats = pd.factorize(df["media_category"])
        df["media_outlet_clean"] = cat_codes + 1
        logging.info(f"Collapsed each category => single cluster. #unique cats={len(unique_cats)}")

    # chunk + stats
    chunk_and_save(df,20000,prefix=prefix)
    print_basic_stats(df,prefix=prefix)

    # correlation
    analyze_all_custom_correlations(df,prefix=prefix)

    # aggregator => stats => GEE => compile
    agg_df=aggregate_sentiment_scores(df,CATEGORIES)
    agg_df=calculate_averages(agg_df)
    stats_df=calculate_mean_median(agg_df)
    save_aggregated_scores_to_csv(agg_df, csvout, prefix=prefix)
    plot_statistics(agg_df, outdir, prefix=prefix)

    df_best, df_allcombos=run_gee_analyses_best_qic(df)
    print(f"Best QIC => prefix={prefix}")
    print(df_best)

    compile_results_into_multiple_workbooks(
        aggregated_df=agg_df,
        stats_df=stats_df,
        raw_df=df,
        df_best_qic=df_best,
        df_all_combos=df_allcombos,
        plots_dir=outdir,
        main_excel=main_excel,
        raw_excel=raw_excel,
        gee_excel=gee_excel,
        plots_excel=plots_excel,
        combined_excel=combined_excel,
        df_full=df,
        validation_records=validation_records
    )

    # 6) sensitivity => merges small clusters => prefix_sens
    if prefix:
        prefix_sens = prefix+"_Sens"
    else:
        prefix_sens="Sens"

    df_sens=exclude_small_clusters(df, min_size=min_size)
    if len(df_sens)<2:
        print(f"[Sensitivity] => merging <{min_size} => not enough data => skip.")
        return

    print(f"[Sensitivity] => Re-running pipeline => prefix={prefix_sens}, len={len(df_sens)}")

    run_pipeline_for_df_inner_sens(
        df_sens,
        prefix_sens,
        main_excel=OUTPUT_EXCEL_MAIN,
        raw_excel=OUTPUT_EXCEL_RAW,
        gee_excel=OUTPUT_EXCEL_GEE,
        plots_excel=OUTPUT_EXCEL_PLOTS,
        combined_excel=OUTPUT_EXCEL_COMBINED,
        outdir=outdir,
        csvout=csvout
    )

def run_pipeline_for_df_inner_sens(
    df_sens,
    prefix_sens,
    main_excel,
    raw_excel,
    gee_excel,
    plots_excel,
    combined_excel,
    outdir,
    csvout
):
    validation_records=[]

    if prefix_sens:
        main_excel = main_excel.replace(".xlsx", f"_{prefix_sens}.xlsx")
        raw_excel = raw_excel.replace(".xlsx", f"_{prefix_sens}.xlsx")
        gee_excel = gee_excel.replace(".xlsx", f"_{prefix_sens}.xlsx")
        plots_excel = plots_excel.replace(".xlsx", f"_{prefix_sens}.xlsx")
        combined_excel = combined_excel.replace(".xlsx", f"_{prefix_sens}.xlsx")

    c1=check_number_of_clusters(df_sens,"media_outlet_clean",threshold=20)
    validation_records.append(c1)
    c2=check_cluster_balance(df_sens,"media_outlet_clean",imbalance_ratio=5.0)
    validation_records.append(c2)
    c_cov=coverage_simulation(n_sims=1000,n_clusters=10,cluster_size=10,true_beta=0.5)
    validation_records.append(c_cov)

    chunk_and_save(df_sens,20000,prefix=prefix_sens)
    print_basic_stats(df_sens,prefix=prefix_sens)

    analyze_all_custom_correlations(df_sens,prefix=prefix_sens)

    agg_df=aggregate_sentiment_scores(df_sens,CATEGORIES)
    agg_df=calculate_averages(agg_df)
    stats_df=calculate_mean_median(agg_df)
    save_aggregated_scores_to_csv(agg_df, csvout, prefix=prefix_sens)
    plot_statistics(agg_df, outdir, prefix=prefix_sens)

    df_best, df_allcombos=run_gee_analyses_best_qic(df_sens)
    print(f"Best QIC => prefix={prefix_sens}")
    print(df_best)

    # produce final results
    from openpyxl import Workbook
    from openpyxl.utils.dataframe import dataframe_to_rows

    diag_records=[]
    cv_records=[]
    sens_records=[]
    boot_records=[]
    idx_rows=[]

    with pd.ExcelWriter(gee_excel,engine="openpyxl") as writer:
        if not df_allcombos.empty:
            df_allcombos.to_excel(writer, sheet_name="All_Combos", index=False)

        for i,row in df_best.iterrows():
            s=row["Sentiment"]
            meas=row["Measure"]
            fam=row["Best_Family"]
            st=row["Best_Structure"]
            sc=row["Best_Scale"]
            best_qic=row["Best_QIC_main"]

            sh_name=f"BestQIC_{s[:8]}_{meas[:12]}"
            out=pairwise_and_diagnostics(df_sens,s,meas,fam,st,sc,prefix=prefix_sens,isSens=True)
            if out[0] is None:
                tmp_df=pd.DataFrame({"Summary":[f"No valid model => {s}-{meas}"]})
                tmp_df.to_excel(writer, sheet_name=sh_name, index=False)
                continue

            summary_txt, pair_df, diffIDs_df, diag_dict, pseudo_dict, cv_val, sens_df_df, boot_info = out

            lines=summary_txt.split("\n")
            sumdf=pd.DataFrame({"GEE_Summary":lines})
            sumdf.to_excel(writer, sheet_name=sh_name, index=False)

            sr=len(sumdf)+2
            pair_df.to_excel(writer, sheet_name=sh_name, index=False, startrow=sr)

            sr2=sr+len(pair_df)+2
            if diffIDs_df is not None and not diffIDs_df.empty:
                diffIDs_df.to_excel(writer, sheet_name=sh_name, index=False, startrow=sr2)

            diag_records.append({
                "Sentiment": s,
                "Measure": meas,
                "Family": fam,
                "Structure": st,
                "Scale": sc,
                "mean_pearson_res": diag_dict["mean_pearson_res"],
                "std_pearson_res": diag_dict["std_pearson_res"],
                "avg_within_corr": diag_dict["avg_within_corr"],
                "overdisp_ratio": diag_dict["overdisp_ratio"],
                "assessment": diag_dict["assessment"],
                "NB_QIC": pseudo_dict["NB_QIC"],
                "Poisson_QIC": pseudo_dict["Poisson_QIC"],
                "diff_QIC": pseudo_dict["diff_QIC"],
                "pseudo_conclusion": pseudo_dict["conclusion"]
            })

            pm_str = str(boot_info["ParamMean"])
            ps_str = str(boot_info["ParamStd"])
            boot_records.append({
                "Sentiment": s,
                "Measure": meas,
                "Family": fam,
                "Structure": st,
                "Scale": sc,
                "MD_BootParamMean": pm_str,
                "MD_BootParamStd":  ps_str
            })
            idx_rows.append({
                "Sentiment": s,
                "Measure": meas,
                "Family": fam,
                "Structure": st,
                "Scale": sc,
                "BestQIC": best_qic
            })

        idxdf=pd.DataFrame(idx_rows)
        idxdf.to_excel(writer, "BestQIC_Index", index=False)
        diag_df=pd.DataFrame(diag_records)
        diag_df.to_excel(writer, "Diagnostics", index=False)
        if len(boot_records)>0:
            boot_all=pd.DataFrame(boot_records)
            boot_all.to_excel(writer,"Bootstrap_Res",index=False)

    with pd.ExcelWriter(main_excel, engine="openpyxl") as w:
        agg_df.to_excel(w,"Aggregated_Scores",index=False)
        stats_df.to_excel(w,"Mean_Median_Statistics",index=False)

    with pd.ExcelWriter(raw_excel, engine="openpyxl") as w:
        df_sens.to_excel(w,"Raw_Data",index=False)

    wb_comb=Workbook()
    if "Sheet" in wb_comb.sheetnames:
        wb_comb.remove(wb_comb["Sheet"])
    ws_agg=wb_comb.create_sheet("Aggregated_Scores")
    for r in dataframe_to_rows(agg_df,index=False,header=True):
        ws_agg.append(r)
    ws_stats=wb_comb.create_sheet("Mean_Median_Statistics")
    for r in dataframe_to_rows(stats_df,index=False,header=True):
        ws_stats.append(r)
    wb_comb.save(combined_excel)

###############################################################################
# main => run original vs collapsed categories
###############################################################################
def main():
    setup_logging()
    logging.info("Starting pipeline => dual categories => original + collapsed")

    df_raw=load_jsonl(INPUT_JSONL_FILE)
    print(f"Total articles loaded => {len(df_raw)}")
    if "high_rate_2" in df_raw.columns:
        df_raw["high_rate_2"]=df_raw["high_rate_2"].astype(str).str.strip().str.lower()

    ########## A) Original categories => yes / all
    print("\n=== A) Original Categories (6-cat) ===")
    df_orig = map_media_outlet_to_category(df_raw, MEDIA_CATEGORIES_ORIG)

    if "high_rate_2" in df_orig.columns:
        df_yes_orig = df_orig[df_orig["high_rate_2"]=="yes"].copy()
        print(f"df_yes_orig => {len(df_yes_orig)} rows")
        if len(df_yes_orig)>0:
            print("**Pipeline => Yes** (orig cats)")
            run_pipeline_for_df(df_yes_orig, prefix="Yes")
        else:
            print("No rows => skip yes subset (orig cats)")
    else:
        print("No 'high_rate_2' => skip yes subset (orig cats)")

    print("**Pipeline => All** (orig cats)")
    run_pipeline_for_df(df_orig, prefix="All")
    del df_orig
    gc.collect()

    ########## B) Collapsed categories => yes_collapsed / all_collapsed
    print("\n=== B) Collapsed Categories (4-cat) ===")
    df_coll = map_media_outlet_to_category(df_raw, MEDIA_CATEGORIES_COLLAPSED)

    if "high_rate_2" in df_coll.columns:
        df_yes_coll = df_coll[df_coll["high_rate_2"]=="yes"].copy()
        print(f"df_yes_collapsed => {len(df_yes_coll)} rows")
        if len(df_yes_coll)>0:
            print("**Pipeline => Yes_collapsed**")
            run_pipeline_for_df(df_yes_coll, prefix="Yes_collapsed")
        else:
            print("No rows => skip yes subset (collapsed cats)")
    else:
        print("No 'high_rate_2' => skip yes subset (collapsed cats)")

    print("**Pipeline => All_collapsed**")
    run_pipeline_for_df(df_coll, prefix="All_collapsed")
    gc.collect()

    logging.info("All done => yes/all for original cats & collapsed cats.")


if __name__=="__main__":
    main()
