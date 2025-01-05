import json
import pandas as pd
import os
from datetime import datetime

# ------------------------------ #
#        Configuration           #
# ------------------------------ #

# Input JSON file
INPUT_JSON_FILE = 'sentiment_openai.json'

# Output Excel file
OUTPUT_EXCEL_FILE = 'sentiment_openai.xlsx'

# ------------------------------ #
#         Helper Functions       #
# ------------------------------ #

def load_json_to_dataframe(json_file):
    """
    Loads JSON data from a file and converts it to a pandas DataFrame.

    Args:
        json_file (str): Path to the JSON file.

    Returns:
        pd.DataFrame: DataFrame containing the JSON data.
    """
    try:
        with open(json_file, 'r', encoding='utf-8') as f:
            data = json.load(f)
        df = pd.DataFrame(data)
        print(f"Successfully loaded data from {json_file}")
        return df
    except FileNotFoundError:
        print(f"Error: The file {json_file} does not exist.")
        return None
    except json.JSONDecodeError as e:
        print(f"Error decoding JSON: {e}")
        return None
    except Exception as e:
        print(f"An unexpected error occurred while loading JSON: {e}")
        return None

def save_dataframe_to_excel(df, excel_file):
    """
    Saves a pandas DataFrame to an Excel file.

    Args:
        df (pd.DataFrame): The DataFrame to save.
        excel_file (str): Path to the output Excel file.
    """
    try:
        # Create a Pandas Excel writer using openpyxl as the engine
        with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Sentiment Analysis')
        
        print(f"Data successfully saved to {excel_file}")
    except Exception as e:
        print(f"Error saving DataFrame to Excel: {e}")

def apply_basic_styling(excel_file):
    """
    Applies basic styling to the Excel file for better readability.

    Args:
        excel_file (str): Path to the Excel file to style.
    """
    try:
        from openpyxl import load_workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        from openpyxl.utils import get_column_letter

        wb = load_workbook(excel_file)
        ws = wb.active

        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        alignment = Alignment(horizontal="center", vertical="center")

        # Apply styles to header
        for col in ws.iter_cols(min_row=1, max_row=1):
            for cell in col:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = alignment

        # Adjust column widths
        for column_cells in ws.columns:
            length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
            adjusted_width = (length + 2)
            ws.column_dimensions[get_column_letter(column_cells[0].column)].width = adjusted_width

        # Center align all cells
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = alignment

        # Save the workbook with styles
        wb.save(excel_file)
        print(f"Basic styling applied to {excel_file}")
    except ImportError:
        print("openpyxl is not installed. Please install it to apply styling.")
    except Exception as e:
        print(f"Error applying styling to Excel file: {e}")

def add_timestamp_to_excel(excel_file):
    """
    Adds a timestamp sheet to the Excel file indicating when it was created.

    Args:
        excel_file (str): Path to the Excel file.
    """
    try:
        from openpyxl import load_workbook
        from openpyxl.styles import Font, Alignment

        wb = load_workbook(excel_file)
        ws = wb.create_sheet(title='Timestamp')

        current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        ws['A1'] = "Report Generated On:"
        ws['A1'].font = Font(bold=True)
        ws['B1'] = current_time
        ws['B1'].alignment = Alignment(horizontal="left")

        # Adjust column widths
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 25

        wb.save(excel_file)
        print(f"Timestamp added to {excel_file}")
    except ImportError:
        print("openpyxl is not installed. Please install it to add a timestamp.")
    except Exception as e:
        print(f"Error adding timestamp to Excel file: {e}")

# ------------------------------ #
#             Main               #
# ------------------------------ #

def main():
    # Check if input file exists
    if not os.path.isfile(INPUT_JSON_FILE):
        print(f"Error: The input file '{INPUT_JSON_FILE}' does not exist.")
        return

    # Load JSON data into DataFrame
    df = load_json_to_dataframe(INPUT_JSON_FILE)
    if df is None:
        print("Failed to load data. Exiting.")
        return

    # Save DataFrame to Excel
    save_dataframe_to_excel(df, OUTPUT_EXCEL_FILE)

    # Apply basic styling (optional)
    apply_basic_styling(OUTPUT_EXCEL_FILE)

    # Add a timestamp sheet (optional)
    add_timestamp_to_excel(OUTPUT_EXCEL_FILE)

    print("JSON to Excel conversion completed successfully.")

if __name__ == "__main__":
    main()
