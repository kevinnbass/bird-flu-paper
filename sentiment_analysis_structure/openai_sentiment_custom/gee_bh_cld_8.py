#!/usr/bin/env python3
# gee_final_best_qic_with_pairwise_cld.py
"""
Script that:
  1) Loads data from JSONL
  2) Chunks & saves CSV
  3) Performs correlation analyses & scatterplots for Quotation vs Fulltext
  4) Aggregates sentiment/emotion scores + bar plots
  5) Finds best QIC combination (Ind/Exch × scale=[none,pearson,deviance,ub,bc])
  6) For each best QIC model:
     - Refit & produce GEE summary
     - Compute pairwise comparisons (BH correction)
     - Build a CLD table with 4 columns:
         [MediaCategory, Mean, UserCLD (blank), AutoCLD (auto-assigned letters)]
  7) Writes summary, pairwise, & CLD into analysis_gee.xlsx (one sheet per sentiment×measure),
     highlighting rows where reject_H0=True in red
  8) Also saves analysis_main, analysis_raw, analysis_plots, analysis_combined
"""

import json
import os
import re
import sys
import warnings
import logging
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import seaborn as sns
from tqdm import tqdm

from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill

import statsmodels.api as sm
from statsmodels.genmod.generalized_estimating_equations import GEE
from statsmodels.genmod.families import Poisson
from statsmodels.genmod.cov_struct import Independence, Exchangeable
from statsmodels.stats.multitest import multipletests
from scipy.stats import pearsonr, norm

# --------------------- #
# Configuration
# --------------------- #
INPUT_JSONL_FILE = "processed_all_articles_fixed_2.jsonl"

OUTPUT_DIR = "graphs_analysis"
os.makedirs(OUTPUT_DIR, exist_ok=True)

CSV_OUTPUT_DIR = "csv_raw_scores"
os.makedirs(CSV_OUTPUT_DIR, exist_ok=True)

OUTPUT_EXCEL_MAIN = "analysis_main.xlsx"
OUTPUT_EXCEL_RAW = "analysis_raw.xlsx"
OUTPUT_EXCEL_GEE = "analysis_gee.xlsx"   # best QIC approach w/ pairwise + CLD
OUTPUT_EXCEL_PLOTS = "analysis_plots.xlsx"
OUTPUT_EXCEL_COMBINED = "analysis_combined.xlsx"

LOG_FILE = "analysis.log"

CATEGORIES = [
    "joy","sadness","anger","fear",
    "surprise","disgust","trust","anticipation",
    "negative_sentiment","positive_sentiment"
]

# The final output order for the CLD table:
CLD_ORDER = ["Scientific","Left","Lean Left","Center","Lean Right","Right"]

MEDIA_CATEGORIES = {
    "Scientific": ["nature","sciam","stat","newscientist"],
    "Left": ["theatlantic","the daily beast","the intercept","mother jones","msnbc","slate","vox","huffpost"],
    "Lean Left": ["ap","axios","cnn","guardian","business insider","nbcnews","npr","nytimes","politico","propublica","wapo","usa today"],
    "Center": ["reuters","marketwatch","financial times","newsweek","forbes"],
    "Lean Right": ["thedispatch","epochtimes","foxbusiness","wsj","national review","washtimes"],
    "Right": ["breitbart","theblaze","daily mail","dailywire","foxnews","nypost","newsmax"],
}


def setup_logging():
    """Initialize logging to both file and console."""
    log_format = "%(asctime)s [%(levelname)s] %(module)s - %(message)s"
    logging.basicConfig(filename=LOG_FILE, level=logging.DEBUG, format=log_format)
    console = logging.StreamHandler(sys.stdout)
    console.setLevel(logging.INFO)
    console.setFormatter(logging.Formatter(log_format))
    logging.getLogger().addHandler(console)
    logging.info("Logging initialized (file+console).")


# Silence the "QIC scale=None" warning from statsmodels
warnings.filterwarnings(
    "ignore",
    message="QIC values obtained using scale=None are not appropriate for comparing models"
)


# ------------------------------
# 1) Load, chunk, basic stats
# ------------------------------
def load_jsonl(jsonl_file):
    logging.info(f"Loading JSONL from {jsonl_file}")
    records=[]
    with open(jsonl_file, "r", encoding="utf-8") as f:
        for line in tqdm(f, desc="Loading JSONL data"):
            rec=json.loads(line)
            records.append(rec)
    df=pd.DataFrame(records)
    logging.debug(f"Loaded DataFrame shape={df.shape}")
    return df

def map_media_outlet_to_category(df):
    """Map media outlets to the categories in MEDIA_CATEGORIES."""
    cat_map={}
    for cat, outls in MEDIA_CATEGORIES.items():
        for o in outls:
            cat_map[o.lower().strip()]=cat

    if "media_outlet" not in df.columns:
        raise KeyError("'media_outlet' not found in data")

    df["media_outlet_clean"]=df["media_outlet"].str.lower().str.strip()
    df["media_category"]=df["media_outlet_clean"].map(cat_map).fillna("Other")

    unmapped=df[df["media_category"]=="Other"]["media_outlet"].unique()
    if len(unmapped)>0:
        logging.warning(f"Unmapped => {unmapped}")
        print(f"Warning: Not mapped => {unmapped}")
    return df

def chunk_and_save(df, chunk_size=20000):
    logging.info(f"Chunking DataFrame => len={len(df)}, chunk_size={chunk_size}")
    for i in range(0,len(df),chunk_size):
        part=df.iloc[i:i+chunk_size]
        out_csv=os.path.join(CSV_OUTPUT_DIR,f"raw_data_part_{(i//chunk_size)+1}.csv")
        part.to_csv(out_csv,index=False)
        print(f"Saved chunk {(i//chunk_size)+1} to {out_csv}")

def print_basic_stats(df):
    logging.info(f"Basic stats => total articles = {len(df)}")
    print("\nSummary Statistics:")
    print("Total number of articles:", len(df))
    if "media_outlet_clean" in df.columns:
        vc=df["media_outlet_clean"].value_counts()
        print("\nArticles per outlet:")
        print(vc)
    if "media_category" in df.columns:
        vc2=df["media_category"].value_counts()
        print("\nArticles per category:")
        print(vc2)
    print()


# ------------------------------
# 2) Quotation vs. Fulltext correlation
# ------------------------------
def analyze_quotation_fulltext_correlation(df):
    """
    1) For each (media_category, sentiment), compute Quotation_Average & Fulltext_Average
    2) Correlation, scatter, bar
    3) Combined z-scatter
    """
    logging.info("Analyzing Quotation vs Fulltext correlation.")

    records=[]
    for cat in df["media_category"].dropna().unique():
        dcat=df[df["media_category"]==cat]
        for s in CATEGORIES:
            pat=rf"^{s}_\d+$"
            matched=[c for c in dcat.columns if re.match(pat,c)]
            if matched:
                clp=dcat[matched].clip(lower=0)
                qsum=clp.sum(skipna=True).sum()
                qcount=clp.count().sum()
                qavg=qsum/qcount if qcount>0 else np.nan
            else:
                qavg=np.nan

            fcol=f"{s}_fulltext"
            if fcol in dcat.columns:
                fv=dcat[fcol].clip(lower=0)
                f_sum=fv.sum(skipna=True)
                f_cnt=fv.count()
                favg=f_sum/f_cnt if f_cnt>0 else np.nan
            else:
                favg=np.nan

            records.append({
                "MediaCategory": cat,
                "Sentiment": s,
                "Quotation_Average": qavg,
                "Fulltext_Average": favg
            })
    agg_df=pd.DataFrame(records)

    cor_results=[]
    scatter_map={}
    for s in CATEGORIES:
        sub=agg_df[(agg_df["Sentiment"]==s)].dropna(subset=["Quotation_Average","Fulltext_Average"])
        if len(sub)>1:
            cor_val,_=pearsonr(sub["Quotation_Average"],sub["Fulltext_Average"])
        else:
            cor_val=np.nan
        cor_results.append({"Sentiment":s,"Correlation":cor_val})

        if not sub.empty:
            # scatter
            plt.figure(figsize=(6,5))
            sns.scatterplot(x="Quotation_Average",y="Fulltext_Average",data=sub,hue="MediaCategory",s=50)
            plt.title(f"{s.capitalize()} (Quotation vs Fulltext)\nr={cor_val:.3f}")
            plt.tight_layout()
            outpath=os.path.join(OUTPUT_DIR,f"scatter_{s}.png")
            try:
                plt.savefig(outpath)
            except:
                pass
            plt.close()
        scatter_map[s]=sub.copy()

    cor_df=pd.DataFrame(cor_results)
    # bar
    plt.figure(figsize=(8,5))
    sns.barplot(x="Sentiment",y="Correlation",data=cor_df,color="gray")
    plt.title("Correlation (Quotation vs. Fulltext) per Sentiment")
    plt.xticks(rotation=45,ha="right")
    plt.ylim(-1,1)
    plt.tight_layout()
    cbar=os.path.join(OUTPUT_DIR,"correlation_quotation_fulltext_bar.png")
    try:
        plt.savefig(cbar)
    except:
        pass
    plt.close()

    # combined z-scatter
    combo=[]
    for s, sd in scatter_map.items():
        if not sd.empty:
            cpy=sd.copy()
            cpy["Sentiment"]=s
            combo.append(cpy)
    if combo:
        allc=pd.concat(combo,ignore_index=True)
        allc["Qmean"]=allc.groupby("Sentiment")["Quotation_Average"].transform("mean")
        allc["Qstd"]=allc.groupby("Sentiment")["Quotation_Average"].transform("std")
        allc["Fmean"]=allc.groupby("Sentiment")["Fulltext_Average"].transform("mean")
        allc["Fstd"]=allc.groupby("Sentiment")["Fulltext_Average"].transform("std")
        allc["Quotation_Z"]=(allc["Quotation_Average"]-allc["Qmean"])/allc["Qstd"]
        allc["Fulltext_Z"]=(allc["Fulltext_Average"]-allc["Fmean"])/allc["Fstd"]
        v=allc.dropna(subset=["Quotation_Z","Fulltext_Z"])
        if len(v)>1:
            r_val,_=pearsonr(v["Quotation_Z"],v["Fulltext_Z"])
        else:
            r_val=np.nan
        plt.figure(figsize=(7,5))
        sns.regplot(x="Quotation_Z",y="Fulltext_Z",data=v,
                    scatter_kws={"color":"black","alpha":0.6},line_kws={"color":"red"})
        plt.title(f"All Sentiments Combined (Z-scores)\nr={r_val:.3f}")
        plt.tight_layout()
        csc=os.path.join(OUTPUT_DIR,"combined_normalized_scatter.png")
        try:
            plt.savefig(csc)
        except:
            pass
        plt.close()

    out_csv=os.path.join(CSV_OUTPUT_DIR,"quotation_fulltext_correlation.csv")
    cor_df.to_csv(out_csv,index=False)
    logging.info(f"Correlation data => {out_csv}")


# ------------------------------
# 3) Aggregation & Stats
# ------------------------------
def aggregate_sentiment_scores(df, sentiments):
    logging.info("Aggregating sentiment/emotion scores by category + sentiment.")
    recs=[]
    for cat in MEDIA_CATEGORIES.keys():
        sub=df[df["media_category"]==cat]
        for s in sentiments:
            pat=rf"^{re.escape(s)}_\d+$"
            matched=[c for c in sub.columns if re.match(pat,c)]
            if matched:
                qsum=sub[matched].clip(lower=0).sum(skipna=True).sum()
                qcount=sub[matched].clip(lower=0).count().sum()
            else:
                qsum,qcount=(0,0)

            fcol=f"{s}_fulltext"
            if fcol in sub.columns:
                fv=sub[fcol].clip(lower=0)
                f_sum=fv.sum(skipna=True)
                f_cnt=fv.count()
            else:
                f_sum,f_cnt=(0,0)

            recs.append({
                "Media Category":cat,
                "Sentiment/Emotion": s,
                "Quotation_Sum": qsum,
                "Quotation_Count": qcount,
                "Fulltext_Sum": f_sum,
                "Fulltext_Count": f_cnt
            })
    return pd.DataFrame(recs)

def calculate_averages(agg_df):
    logging.info("Calculating Quotation_Average & Fulltext_Average.")
    def sdiv(a,b):
        return a/b if b>0 else None
    agg_df["Quotation_Average"]=agg_df.apply(lambda r: sdiv(r["Quotation_Sum"],r["Quotation_Count"]),axis=1)
    agg_df["Fulltext_Average"]=agg_df.apply(lambda r: sdiv(r["Fulltext_Sum"],r["Fulltext_Count"]),axis=1)
    return agg_df

def calculate_mean_median(agg_df):
    logging.info("Computing global mean/median of Quotation/Fulltext averages.")
    rows=[]
    for s in CATEGORIES:
        sub=agg_df[agg_df["Sentiment/Emotion"]==s]
        qa=sub["Quotation_Average"].dropna()
        fa=sub["Fulltext_Average"].dropna()
        mean_q=qa.mean() if len(qa)>0 else None
        med_q=qa.median() if len(qa)>0 else None
        mean_f=fa.mean() if len(fa)>0 else None
        med_f=fa.median() if len(fa)>0 else None
        rows.append({
            "Sentiment/Emotion": s,
            "Mean_Quotation_Average": mean_q,
            "Median_Quotation_Average": med_q,
            "Mean_Fulltext_Average": mean_f,
            "Median_Fulltext_Average": med_f
        })
    return pd.DataFrame(rows)

def save_aggregated_scores_to_csv(agg_df, out_dir):
    fn=os.path.join(out_dir,"aggregated_sentiment_emotion_scores.csv")
    agg_df.to_csv(fn,index=False)
    print(f"Aggregated sentiment/emotion scores => {fn}")
    logging.info(f"Aggregated => {fn}")

def plot_statistics(agg_df, out_dir):
    logging.info("Plotting bar charts for Quotation/Fulltext across categories.")
    sns.set_style("whitegrid")
    for s in CATEGORIES:
        sub=agg_df[agg_df["Sentiment/Emotion"]==s]
        # Quotation
        plt.figure(figsize=(10,6))
        sns.barplot(x="Media Category",y="Quotation_Average",data=sub,color="steelblue")
        plt.title(f"Mean Quotation-Based '{s.capitalize()}' Scores")
        plt.xticks(rotation=45,ha="right")
        plt.tight_layout()
        out1=os.path.join(out_dir,f"quote_{s}.png")
        try:
            plt.savefig(out1)
        except:
            pass
        plt.close()

        # Fulltext
        plt.figure(figsize=(10,6))
        sns.barplot(x="Media Category",y="Fulltext_Average",data=sub,color="darkorange")
        plt.title(f"Mean Fulltext-Based '{s.capitalize()}' Scores")
        plt.xticks(rotation=45,ha="right")
        plt.tight_layout()
        out2=os.path.join(out_dir,f"fulltext_{s}.png")
        try:
            plt.savefig(out2)
        except:
            pass
        plt.close()


# ------------------------------
# 4) GEE scale computations
# ------------------------------
def compute_pearson_scale(y,mu,df_resid):
    if df_resid<=0: return np.nan
    y=np.asarray(y)
    mu=np.asarray(mu)
    r=(y-mu)/np.sqrt(mu)
    return np.sum(r**2)/df_resid

def compute_deviance_scale(y,mu,df_resid):
    if df_resid<=0: return np.nan
    y=np.asarray(y)
    mu=np.asarray(mu)
    arr=np.zeros_like(y,dtype=float)
    for i in range(len(y)):
        if y[i]>0 and mu[i]>0:
            arr[i]=y[i]*np.log(y[i]/mu[i])-(y[i]-mu[i])
        elif y[i]==0:
            arr[i]=-(y[i]-mu[i])
        else:
            arr[i]=np.nan
    return 2*np.nansum(arr)/df_resid

def compute_ub_scale(y,mu,df_resid):
    p=compute_pearson_scale(y,mu,df_resid)
    if np.isnan(p): return p
    return 1.1*p

def compute_bc_scale(y,mu,df_resid):
    d=compute_deviance_scale(y,mu,df_resid)
    if np.isnan(d): return d
    return 0.9*d

def fit_and_compute_scales(model):
    base_res=model.fit(scale=None)
    base_qic=base_res.qic()
    y=base_res.model.endog
    mu=base_res.fittedvalues
    n=len(y)
    p=len(base_res.params)
    df_resid=n-p

    pear=compute_pearson_scale(y,mu,df_resid)
    dev=compute_deviance_scale(y,mu,df_resid)
    ubv=compute_ub_scale(y,mu,df_resid)
    bcv=compute_bc_scale(y,mu,df_resid)

    results={}
    if isinstance(base_qic, tuple):
        q_m,q_a=base_qic
    else:
        q_m,q_a=base_qic,None
    results["none"]=(q_m,q_a,None)

    from math import isnan
    for (nm,val) in [("pearson",pear),("deviance",dev),("ub",ubv),("bc",bcv)]:
        if (not isnan(val)) and (df_resid>0):
            re2=model.fit(scale=val)
            r_qic=re2.qic()
            if isinstance(r_qic,tuple):
                rm,ra=r_qic
            else:
                rm,ra=r_qic,None
            results[nm]=(rm,ra,val)
        else:
            results[nm]=(np.nan,None,val)

    return results


# ------------------------------
# 5) Best QIC approach
# ------------------------------
from statsmodels.genmod.generalized_estimating_equations import GEE

def run_gee_for_sentiment_measure_best_qic(df, sentiment, measure):
    d2=df.copy()
    if measure=="Quotation":
        pat=rf"^{re.escape(sentiment)}_\d+$"
        matched=[c for c in d2.columns if re.match(pat,c)]
        if not matched:
            return None
        d2["_score_col"]=d2[matched].clip(lower=0).mean(axis=1)
    else:
        fcol=f"{sentiment}_fulltext"
        if fcol not in d2.columns:
            return None
        d2["_score_col"]=d2[fcol].clip(lower=0)

    needed=["_score_col","media_outlet_clean","media_category"]
    d2=d2.dropna(subset=needed)
    if len(d2)<2 or d2["media_category"].nunique()<2:
        return None

    best_qic_main=np.inf
    best_tuple=None
    combos=[]
    structures=[Independence(),Exchangeable()]
    for cov_obj in structures:
        st_name=cov_obj.__class__.__name__
        model=GEE.from_formula("_score_col ~ media_category",
            groups="media_outlet_clean",data=d2,
            family=Poisson(),cov_struct=cov_obj)
        sc_map=fit_and_compute_scales(model)
        for sm,(qm,qa,sc_val) in sc_map.items():
            combos.append({
                "Sentiment":sentiment,
                "Measure":measure,
                "Structure":st_name,
                "ScaleMethod":sm,
                "NumericScale":sc_val,
                "QIC_main":qm,
                "QIC_alt":qa
            })
            if (qm<best_qic_main) and (not np.isnan(qm)):
                best_qic_main=qm
                best_tuple=(st_name,sm,qm,sc_val)

    if best_tuple is None:
        return None
    stn, sm, qicv, scale_num=best_tuple
    return {
        "Sentiment":sentiment,
        "Measure":measure,
        "Best_Structure":stn,
        "Best_Scale":sm,
        "Best_QIC_main":qicv,
        "AllCombos":pd.DataFrame(combos),
        "Summary":f"Best scale={sm}, numeric={scale_num}"
    }

def run_gee_analyses_best_qic(df):
    logging.info("Running best QIC approach for each sentiment–measure.")
    best_list=[]
    all_combos=[]
    for s in CATEGORIES:
        for meas in ["Quotation","Fulltext"]:
            info=run_gee_for_sentiment_measure_best_qic(df,s,meas)
            if info is not None:
                best_list.append({
                    "Sentiment": info["Sentiment"],
                    "Measure": info["Measure"],
                    "Best_Structure": info["Best_Structure"],
                    "Best_Scale": info["Best_Scale"],
                    "Best_QIC_main": info["Best_QIC_main"],
                    "Summary": info["Summary"]
                })
                all_combos.append(info["AllCombos"])
    df_best=pd.DataFrame(best_list)
    df_all=pd.concat(all_combos,ignore_index=True) if all_combos else pd.DataFrame()
    return df_best, df_all


# ------------------------------
# 6) Pairwise + CLD
# ------------------------------
def refit_best_gee_with_scale(df, sentiment, measure, struct, scale_name):
    """
    Build model => apply scale => return fitted result or None
    """
    d2=df.copy()
    if measure=="Quotation":
        pat=rf"^{re.escape(sentiment)}_\d+$"
        matched=[c for c in d2.columns if re.match(pat,c)]
        if not matched:
            return None
        d2["_score_col"]=d2[matched].clip(lower=0).mean(axis=1)
    else:
        fcol=f"{sentiment}_fulltext"
        if fcol not in d2.columns:
            return None
        d2["_score_col"]=d2[fcol].clip(lower=0)

    needed=["_score_col","media_outlet_clean","media_category"]
    d2=d2.dropna(subset=needed)
    if len(d2)<2 or d2["media_category"].nunique()<2:
        return None

    d2["media_category"]=d2["media_category"].astype("category")
    if struct=="Independence":
        cov_obj=Independence()
    else:
        cov_obj=Exchangeable()

    model=GEE.from_formula("_score_col ~ media_category",
                           groups="media_outlet_clean",data=d2,
                           family=Poisson(),cov_struct=cov_obj)
    bres=model.fit(scale=None)
    if scale_name=="none":
        return bres

    y=bres.model.endog
    mu=bres.fittedvalues
    n=len(y)
    p=len(bres.params)
    dfresid=n-p
    if dfresid<=0:
        return bres

    if scale_name=="pearson":
        scv=compute_pearson_scale(y,mu,dfresid)
    elif scale_name=="deviance":
        scv=compute_deviance_scale(y,mu,dfresid)
    elif scale_name=="ub":
        scv=compute_ub_scale(y,mu,dfresid)
    elif scale_name=="bc":
        scv=compute_bc_scale(y,mu,dfresid)
    else:
        scv=None
    if scv is None or np.isnan(scv):
        return bres
    return model.fit(scale=scv)


def pairwise_and_cld(df, sentiment, measure, struct, scale_name):
    """
    1) Refit best GEE => final_fit
    2) GEE summary
    3) Pairwise p-values (BH)
    4) Build CLD => 4 columns: [MediaCategory, Mean, UserCLD (blank), AutoCLD (auto letters)]
    """
    final_fit=refit_best_gee_with_scale(df, sentiment, measure, struct, scale_name)
    if final_fit is None:
        return None,None,None
    summary_txt=final_fit.summary().as_text()

    # Pairwise table
    params=final_fit.params
    cov=final_fit.cov_params()
    mdf=final_fit.model.data.frame
    cats=mdf["media_category"].cat.categories
    ref=cats[0]
    idx_map={ref:0}
    for c in cats[1:]:
        nm=f"media_category[T.{c}]"
        idx_map[c]=final_fit.model.exog_names.index(nm)

    pair_list=[]
    for i in range(len(cats)):
        for j in range(i+1,len(cats)):
            ca,cb=cats[i],cats[j]
            con=np.zeros(len(params))
            if ca==ref and cb!=ref:
                con[idx_map[cb]]=-1.0
            elif cb==ref and ca!=ref:
                con[idx_map[ca]]=1.0
            else:
                con[idx_map[ca]]=1.0
                con[idx_map[cb]]=-1.0
            diff_est=con@params
            diff_var=con@cov@con
            diff_se=np.sqrt(diff_var)
            z=diff_est/diff_se
            pval=2*(1-norm.cdf(abs(z)))
            pair_list.append((ca,cb,diff_est,diff_se,z,pval))

    pair_df=pd.DataFrame(pair_list,columns=["CategoryA","CategoryB","Difference","SE","Z","p_value"])

    # BH correction
    rej,p_adj,_,_ = multipletests(pair_df["p_value"],method="fdr_bh")
    pair_df["p_value_adj"]=p_adj
    pair_df["reject_H0"]=rej

    # Build the final CLD with 4 columns
    cld_df=build_cld_4_columns(final_fit, pair_df)

    return summary_txt, pair_df, cld_df


# ------------------------------
# 7) Build 4-column CLD
# ------------------------------
def build_cld_4_columns(final_fit, pair_df):
    """
    Returns a 4-column DataFrame with columns:
      [MediaCategory, Mean, UserCLD, AutoCLD]

    'UserCLD' is left blank (for manual assignment).
    'AutoCLD' is assigned automatically using a multi-letter approach:
      - If two cats differ => no letter overlap
      - If not differ => share >=1 letter
    """
    # 1) Gather predicted means
    mdf=final_fit.model.data.frame
    cats=mdf["media_category"].cat.categories
    intercept=final_fit.params[0]
    cat_effect={cats[0]: 0.0}
    for c in cats[1:]:
        nm=f"media_category[T.{c}]"
        cat_effect[c]=final_fit.params.get(nm, 0.0)

    mean_map={}
    for c in cats:
        mean_map[c]=np.exp(intercept + cat_effect[c])

    # 2) Sort categories by ascending mean
    sorted_cats=sorted(cats, key=lambda x: mean_map[x])

    # 3) Build differ dictionary from pair_df
    #    differ[a][b] = True if "they differ" => can't share letters
    #                  False if "they do not differ" => must share letters
    differ={c:{} for c in sorted_cats}
    for row in pair_df.itertuples(index=False):
        A,B=row.CategoryA, row.CategoryB
        differ[A][B]=row.reject_H0
        differ[B][A]=row.reject_H0

    # 4) Assign letters automatically
    max_letters=26
    letter_pool=[chr(i) for i in range(ord('a'), ord('a')+max_letters)]
    cat_letters={c:set() for c in sorted_cats}
    used_letters=0

    def ensure_share_letter(cA,cB):
        nonlocal used_letters
        # if cA and cB have no intersection => add a letter
        common=cat_letters[cA].intersection(cat_letters[cB])
        if not common:
            # union
            possible=cat_letters[cA].union(cat_letters[cB])
            if possible:
                # pick any letter from possible
                letter_to_add=next(iter(possible))
                cat_letters[cA].add(letter_to_add)
                cat_letters[cB].add(letter_to_add)
            else:
                if used_letters<max_letters:
                    new_letter=letter_pool[used_letters]
                    used_letters+=1
                    cat_letters[cA].add(new_letter)
                    cat_letters[cB].add(new_letter)

    def ensure_no_shared_letter(cA,cB):
        intersection=cat_letters[cA].intersection(cat_letters[cB])
        if intersection:
            cat_letters[cA]=cat_letters[cA].difference(intersection)
            cat_letters[cB]=cat_letters[cB].difference(intersection)

    changed=True
    iteration=0
    max_iter=100
    while changed and iteration<max_iter:
        changed=False
        iteration+=1
        for i in range(len(sorted_cats)):
            for j in range(i+1, len(sorted_cats)):
                cA=sorted_cats[i]
                cB=sorted_cats[j]
                they_differ=differ[cA].get(cB,False)  # True => differ => can't share
                beforeA=set(cat_letters[cA])
                beforeB=set(cat_letters[cB])
                if they_differ:
                    ensure_no_shared_letter(cA,cB)
                else:
                    ensure_share_letter(cA,cB)
                if cat_letters[cA]!=beforeA or cat_letters[cB]!=beforeB:
                    changed=True

    # 5) Build final table with columns => [MediaCategory, Mean, UserCLD, AutoCLD]
    rows=[]
    for c in CLD_ORDER:
        if c in cats:
            # sort letters
            letter_list=sorted(cat_letters[c])
            auto_str=''.join(letter_list)
            rows.append((c, mean_map[c], "", auto_str))
        else:
            rows.append((c, None, "", ""))

    df_out=pd.DataFrame(rows, columns=["MediaCategory","Mean","UserCLD","AutoCLD"])
    return df_out


# ------------------------------
# 8) Compile results
# ------------------------------
def compile_results_into_multiple_workbooks(
    aggregated_df, stats_df, raw_df,
    df_best_qic, df_all_combos,
    plots_dir,
    main_excel, raw_excel, gee_excel, plots_excel, combined_excel,
    df_full
):
    # 1) analysis_main
    with pd.ExcelWriter(main_excel, engine="openpyxl") as w:
        aggregated_df.to_excel(w,"Aggregated_Scores",index=False)
        stats_df.to_excel(w,"Mean_Median_Statistics",index=False)

    # 2) analysis_raw
    with pd.ExcelWriter(raw_excel, engine="openpyxl") as w:
        raw_df.to_excel(w,"Raw_Data",index=False)
        for s in CATEGORIES:
            s_cols=[c for c in raw_df.columns if c.startswith(s+"_")]
            s_df=raw_df[["media_category","media_outlet"]+s_cols].copy()
            sheet_name=f"Raw_{s[:29]}"
            s_df.to_excel(w,sheet_name,index=False)

    # 3) analysis_gee => best QIC w/ pairwise + new CLD
    with pd.ExcelWriter(gee_excel, engine="openpyxl") as writer:
        idx_rows=[]
        for i, row in df_best_qic.iterrows():
            s=row["Sentiment"]
            meas=row["Measure"]
            st=row["Best_Structure"]
            sc=row["Best_Scale"]
            sh_name=f"BestQIC_{s[:10]}_{meas[:8]}"
            summary_txt, pair_df, cld_df = pairwise_and_cld(df_full, s, meas, st, sc)
            if summary_txt is None:
                tmp_df=pd.DataFrame({"Summary":["No valid model or not enough data."]})
                tmp_df.to_excel(writer,sh_name,index=False)
                continue

            # Summaries
            lines=summary_txt.split("\n")
            sumdf=pd.DataFrame({"GEE_Summary":lines})
            sumdf.to_excel(writer,sh_name,index=False)

            sr=len(sumdf)+2
            pair_df.to_excel(writer,sh_name,index=False,startrow=sr)

            # highlight rows in red if reject_H0 = True
            ws=writer.sheets[sh_name]
            for row_idx in range(len(pair_df)):
                if pair_df.loc[row_idx,"reject_H0"] == True:
                    excel_row = sr+1 + row_idx
                    for col_idx in range(1, pair_df.shape[1] + 1):
                        cell=ws.cell(row=excel_row+1, column=col_idx)
                        cell.fill=PatternFill(fill_type="solid", start_color="FFFF0000", end_color="FFFF0000")

            sr2=sr+len(pair_df)+2
            if cld_df is not None and not cld_df.empty:
                cld_df.to_excel(writer,sh_name,index=False,startrow=sr2)

            idx_rows.append({
                "Sentiment": s,
                "Measure": meas,
                "SheetName": sh_name,
                "Structure": st,
                "Scale": sc,
                "BestQIC": row["Best_QIC_main"]
            })

        idxdf=pd.DataFrame(idx_rows)
        idxdf.to_excel(writer,"BestQIC_Index",index=False)

        if not df_all_combos.empty:
            df_all_combos.to_excel(writer,"All_Combos",index=False)

    # 4) analysis_plots
    wb_plots=Workbook()
    if "Sheet" in wb_plots.sheetnames:
        wb_plots.remove(wb_plots["Sheet"])
    for s in CATEGORIES:
        q_path=os.path.join(plots_dir,f"quote_{s}.png")
        if os.path.exists(q_path):
            st=f"Quote_{s[:28]}"
            ws=wb_plots.create_sheet(title=st)
            try:
                img=ExcelImage(q_path)
                img.anchor="A1"
                ws.add_image(img)
            except:
                pass
        f_path=os.path.join(plots_dir,f"fulltext_{s}.png")
        if os.path.exists(f_path):
            st2=f"Fulltext_{s[:25]}"
            ws2=wb_plots.create_sheet(title=st2)
            try:
                img2=ExcelImage(f_path)
                img2.anchor="A1"
                ws2.add_image(img2)
            except:
                pass

    cbar=os.path.join(plots_dir,"correlation_quotation_fulltext_bar.png")
    if os.path.exists(cbar):
        ws3=wb_plots.create_sheet("Correlation_Bar")
        try:
            ig3=ExcelImage(cbar)
            ig3.anchor="A1"
            ws3.add_image(ig3)
        except:
            pass

    combp=os.path.join(plots_dir,"combined_normalized_scatter.png")
    if os.path.exists(combp):
        ws4=wb_plots.create_sheet("Combined_ZScatter")
        try:
            ig4=ExcelImage(combp)
            ig4.anchor="A1"
            ws4.add_image(ig4)
        except:
            pass

    wb_plots.save(plots_excel)

    # 5) analysis_combined
    raw_clean=raw_df.copy()
    raw_clean=raw_clean.applymap(lambda x:", ".join(x) if isinstance(x,list) else x)
    wb_comb=Workbook()
    if "Sheet" in wb_comb.sheetnames:
        wb_comb.remove(wb_comb["Sheet"])

    ws_agg=wb_comb.create_sheet("Aggregated_Scores")
    for r in dataframe_to_rows(aggregated_df,index=False,header=True):
        ws_agg.append(r)

    ws_stats=wb_comb.create_sheet("Mean_Median_Statistics")
    for r in dataframe_to_rows(stats_df,index=False,header=True):
        ws_stats.append(r)

    ws_raw=wb_comb.create_sheet("Raw_Data")
    for r in dataframe_to_rows(raw_clean,index=False,header=True):
        ws_raw.append(r)

    ws_best=wb_comb.create_sheet("BestQIC_Table")
    for r in dataframe_to_rows(df_best_qic,index=False,header=True):
        ws_best.append(r)

    wb_comb.save(combined_excel)


# ------------------------------
def main():
    setup_logging()
    logging.info("Starting best QIC GEE approach with BH post-hoc & 4-column CLD table (UserCLD blank, AutoCLD assigned).")

    # 1) Load
    df=load_jsonl(INPUT_JSONL_FILE)
    print(f"Total articles loaded: {len(df)}")

    # 2) map + chunk + stats
    df=map_media_outlet_to_category(df)
    chunk_and_save(df,20000)
    print_basic_stats(df)

    # 3) correlation
    print("Performing Quotation vs Fulltext correlation analysis...")
    analyze_quotation_fulltext_correlation(df)

    # 4) aggregation
    print("Aggregating sentiment/emotion scores per media category...")
    agg_df=aggregate_sentiment_scores(df,CATEGORIES)
    agg_df=calculate_averages(agg_df)
    stats_df=calculate_mean_median(agg_df)
    save_aggregated_scores_to_csv(agg_df,CSV_OUTPUT_DIR)
    plot_statistics(agg_df,OUTPUT_DIR)

    # 5) best QIC approach
    print("Fitting best QIC approach for each sentiment–measure (with BH pairwise + 4-col CLD).")
    df_best, df_allcombos = run_gee_analyses_best_qic(df)
    print("Best QIC approach completed.\n")

    # 6) compile
    compile_results_into_multiple_workbooks(
        aggregated_df=agg_df,
        stats_df=stats_df,
        raw_df=df,
        df_best_qic=df_best,
        df_all_combos=df_allcombos,
        plots_dir=OUTPUT_DIR,
        main_excel=OUTPUT_EXCEL_MAIN,
        raw_excel=OUTPUT_EXCEL_RAW,
        gee_excel=OUTPUT_EXCEL_GEE,
        plots_excel=OUTPUT_EXCEL_PLOTS,
        combined_excel=OUTPUT_EXCEL_COMBINED,
        df_full=df
    )

    print("Analysis completed successfully.")
    logging.info("Analysis completed successfully.")


if __name__=="__main__":
    main()
