import json
import pandas as pd
import matplotlib.pyplot as plt
import seaborn as sns
import os
import logging
import sys
from tqdm import tqdm
from openpyxl.drawing.image import Image as ExcelImage
import re
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import numpy as np
import patsy

import statsmodels.api as sm
from statsmodels.genmod.generalized_estimating_equations import GEE
from statsmodels.genmod.families import Poisson
from statsmodels.genmod.cov_struct import Independence, Exchangeable, Unstructured, Autoregressive
from statsmodels.stats.multitest import multipletests
from scipy.stats import norm, pearsonr

# ------------------------------ #
#          Configuration         #
# ------------------------------ #

INPUT_JSONL_FILE = 'processed_all_articles_fixed.jsonl'  # Updated per your request
OUTPUT_DIR = 'graphs_analysis'
os.makedirs(OUTPUT_DIR, exist_ok=True)

CSV_OUTPUT_DIR = 'csv_raw_scores'
os.makedirs(CSV_OUTPUT_DIR, exist_ok=True)

OUTPUT_EXCEL_MAIN = 'analysis_main.xlsx'
OUTPUT_EXCEL_RAW = 'analysis_raw.xlsx'
OUTPUT_EXCEL_LMM = 'analysis_gee.xlsx'
OUTPUT_EXCEL_PLOTS = 'analysis_plots.xlsx'
OUTPUT_EXCEL_COMBINED = 'analysis_combined.xlsx'
OUTPUT_EXCEL_QIC = 'analysis_gee_qic.xlsx'  # file to store QIC results if desired

CATEGORIES = [
    'joy', 'sadness', 'anger', 'fear',
    'surprise', 'disgust', 'trust', 'anticipation',
    'negative_sentiment', 'positive_sentiment'
]

MEDIA_CATEGORIES = {
    'Scientific': ['nature', 'sciam', 'stat', 'newscientist'],
    'Left': ['theatlantic', 'the daily beast', 'the intercept', 'mother jones', 'msnbc', 'slate', 'vox', 'huffpost'],
    'Lean Left': ['ap', 'axios', 'cnn', 'guardian', 'business insider', 'nbcnews', 'npr', 'nytimes', 'politico', 'propublica', 'wapo', 'usa today'],
    'Center': ['reuters', 'marketwatch', 'financial times', 'newsweek', 'forbes'],
    'Lean Right': ['thedispatch', 'epochtimes', 'foxbusiness', 'wsj', 'national review', 'washtimes'],
    'Right': ['breitbart', 'theblaze', 'daily mail', 'dailywire', 'foxnews', 'nypost', 'newsmax'],
}


def setup_logging(log_file='analysis.log'):
    """Sets up logging (file + console) and suppresses font-manager debug logs."""
    logging.getLogger('matplotlib.font_manager').setLevel(logging.WARNING)
    logging.basicConfig(
        filename=log_file,
        level=logging.DEBUG,
        format='%(asctime)s [%(levelname)s] %(module)s - %(message)s'
    )
    console = logging.StreamHandler(sys.stdout)
    console.setLevel(logging.DEBUG)
    console.setFormatter(logging.Formatter('[%(levelname)s] %(module)s: %(message)s'))
    logging.getLogger('').addHandler(console)

    logging.info("Logging initialized. File + console output.")


def load_jsonl(jsonl_file):
    """Load a JSONL file into a pandas DataFrame."""
    records = []
    with open(jsonl_file, 'r', encoding='utf-8') as f:
        for line in tqdm(f, desc="Loading JSONL data"):
            try:
                record = json.loads(line)
                records.append(record)
            except json.JSONDecodeError as e:
                logging.error(f"JSON decode error: {e}")
    return pd.DataFrame(records)


def map_media_outlet_to_category(df, media_categories):
    """Map each media_outlet to a known media_category or 'Other' if not recognized."""
    if 'media_outlet' not in df.columns:
        msg = "'media_outlet' column not found in DataFrame."
        logging.error(msg)
        raise KeyError(msg)

    # Build a lookup dictionary
    outlet_map = {}
    for cat, outlets in media_categories.items():
        for o in outlets:
            outlet_map[o.lower().strip()] = cat

    df['media_outlet_clean'] = df['media_outlet'].str.lower().str.strip()
    df['media_category'] = df['media_outlet_clean'].map(outlet_map).fillna('Other')
    unmapped = df.loc[df['media_category']=='Other','media_outlet'].unique()
    if len(unmapped)>0:
        logging.warning(f"Unmapped outlets -> 'Other': {unmapped}")
        print(f"Warning: Some media outlets mapped to 'Other': {unmapped}")
    return df


def aggregate_sentiment_scores(df, sentiment_categories):
    """Aggregate sentiment sums and counts across categories for both Quotation and Fulltext."""
    aggregation = []
    for media_cat in MEDIA_CATEGORIES.keys():
        sub = df[df['media_category']==media_cat]
        for sentiment in sentiment_categories:
            # Use a raw string approach to avoid invalid escape sequences: 
            # or build with re.escape:
            pattern = r'^' + re.escape(sentiment) + r'_\d+$'
            matched_cols = [c for c in df.columns if re.match(pattern,c)]
            if matched_cols:
                qsum = sub[matched_cols].clip(lower=0).sum(skipna=True).sum()
                qcount = sub[matched_cols].clip(lower=0).count().sum()
            else:
                qsum,qcount=0,0

            fcol = f"{sentiment}_fulltext"
            if fcol in df.columns:
                fsum = sub[fcol].clip(lower=0).sum(skipna=True)
                fcount = sub[fcol].clip(lower=0).count()
            else:
                fsum,fcount=0,0

            aggregation.append({
                'Media Category': media_cat,
                'Sentiment/Emotion': sentiment,
                'Quotation_Sum': qsum,
                'Quotation_Count': qcount,
                'Fulltext_Sum': fsum,
                'Fulltext_Count': fcount
            })
    return pd.DataFrame(aggregation)


def calculate_averages(agg):
    """Compute Quotation_Average and Fulltext_Average from sums & counts."""
    agg['Quotation_Average'] = agg.apply(
        lambda r: r['Quotation_Sum']/r['Quotation_Count'] if r['Quotation_Count']>0 else None,
        axis=1
    )
    agg['Fulltext_Average'] = agg.apply(
        lambda r: r['Fulltext_Sum']/r['Fulltext_Count'] if r['Fulltext_Count']>0 else None,
        axis=1
    )
    return agg


def calculate_mean_median(aggdf):
    """Compute overall mean and median across categories for each sentiment."""
    rows=[]
    for s in CATEGORIES:
        sub=aggdf[aggdf['Sentiment/Emotion']==s]
        qa= sub['Quotation_Average'].dropna()
        fa= sub['Fulltext_Average'].dropna()
        rows.append({
            'Sentiment/Emotion': s,
            'Mean_Quotation_Average': qa.mean() if len(qa)>0 else None,
            'Median_Quotation_Average': qa.median() if len(qa)>0 else None,
            'Mean_Fulltext_Average': fa.mean() if len(fa)>0 else None,
            'Median_Fulltext_Average': fa.median() if len(fa)>0 else None
        })
    return pd.DataFrame(rows)


def save_aggregated_scores_to_csv(aggdf, outdir, prefix='aggregated_sentiment_emotion_scores.csv'):
    csv_path= os.path.join(outdir,prefix)
    try:
        aggdf.to_csv(csv_path, index=False)
        print(f"Aggregated sentiment/emotion scores saved to '{csv_path}'.")
        logging.info(f"Aggregated sentiment/emotion scores saved to '{csv_path}'.")
    except Exception as e:
        logging.error(f"Error saving aggregated: {e}")


def plot_statistics(aggdf, outdir):
    """Plot bar charts of Quotation_Average & Fulltext_Average by media_category for each sentiment."""
    sns.set_style("whitegrid")
    for s in CATEGORIES:
        subdf= aggdf.loc[aggdf['Sentiment/Emotion']==s].copy()

        # Quotation
        plt.figure(figsize=(10,6))
        sns.barplot(x='Media Category',y='Quotation_Average',data=subdf,color='steelblue')
        plt.title(f"Mean Quotation-Based '{s.capitalize()}' Scores")
        plt.xlabel('Media Category')
        plt.ylabel('Mean Quotation-Based Average Score')
        plt.xticks(rotation=45,ha='right')
        plt.tight_layout()
        out_quote = os.path.join(outdir,f"quote_{s}.png")
        try:
            plt.savefig(out_quote)
            plt.close()
            print(f"Quotation-Based '{s}' scores plot saved to '{out_quote}'.")
        except Exception as e:
            logging.error(f"Error saving Q-plot for {s}: {e}")

        # Fulltext
        plt.figure(figsize=(10,6))
        sns.barplot(x='Media Category',y='Fulltext_Average',data=subdf,color='darkorange')
        plt.title(f"Mean Fulltext-Based '{s.capitalize()}' Scores")
        plt.xlabel('Media Category')
        plt.ylabel('Mean Fulltext-Based Average Score')
        plt.xticks(rotation=45,ha='right')
        plt.tight_layout()
        out_full = os.path.join(outdir,f"fulltext_{s}.png")
        try:
            plt.savefig(out_full)
            plt.close()
            print(f"Fulltext-Based '{s}' scores plot saved to '{out_full}'.")
        except Exception as e:
            logging.error(f"Error saving F-plot for {s}: {e}")


# --------------------------------------------------
#   Multi-Structure GEE with QIC (Poisson)
# --------------------------------------------------
def fit_gee_model(df, formula, groups, cov_struct, family=Poisson(), time=None):
    """Try to fit GEE with the specified covariance structure, return results or None."""
    model= GEE.from_formula(formula, groups=groups, data=df, cov_struct=cov_struct, family=family, time=time)
    try:
        res = model.fit()
    except Exception as e:
        logging.error(f"GEE fit failed: {e}")
        return None
    return res

def run_gee_and_qic(df, sentiment, measure_type, structures, min_date=None):
    """
    Attempt GEE with each correlation structure in 'structures', compute QIC, return dict of {struct_name: (results, qic)}.
    If measure_type='Quotation', we average across columns sentiment_\d+.
    If measure_type='Fulltext', we use sentiment_fulltext. 
    If AR(1) is tested and date is present, we define int_date.

    Return e.g.: {'Exchangeable': (res, qic_val), 'Independence': (res2, qic_val2), ...}
    """
    # Build the outcome column
    if measure_type=='Quotation':
        # use pattern to find columns
        pattern = r'^' + re.escape(sentiment) + r'_\d+$'
        matched = [c for c in df.columns if re.match(pattern,c)]
        if not matched:
            return {}
        temp = df.copy()
        col_mean= f"{sentiment}_quotation_mean"
        temp[col_mean] = temp[matched].clip(lower=0).mean(axis=1)
        score_col= col_mean
    else:
        fcol= f"{sentiment}_fulltext"
        if fcol not in df.columns:
            return {}
        temp=df.copy()
        clipped= f"{sentiment}_fulltext_clipped"
        temp[clipped] = temp[fcol].clip(lower=0)
        score_col= clipped

    mdf= temp.dropna(subset=[score_col,'media_category','media_outlet']).copy()
    # Must have >=2 categories
    if mdf['media_category'].nunique()<2:
        return {}

    # If we want AR(1) with date => ensure valid datetime
    # Then create 'int_date' if date is parseable
    if 'date' in mdf.columns and mdf['date'].dtype == 'O':
        # try to coerce
        mdf['date'] = pd.to_datetime(mdf['date'], errors='coerce')

    # We'll only do AR(1) if min_date is known and date is valid
    if min_date is not None and 'date' in mdf.columns:
        # drop rows that are still invalid date
        mdf = mdf.dropna(subset=['date'])

    mdf['media_category'] = mdf['media_category'].astype('category')
    formula = f"{score_col} ~ media_category"
    groups= "media_outlet"

    # If we can do AR(1), define int_date
    if min_date is not None and 'date' in mdf.columns and mdf['date'].notna().sum()>0:
        mdf = mdf.sort_values(['media_outlet','date'])
        # strictly ensure date
        mdf['int_date'] = (mdf['date'] - min_date).dt.days + 1

    result_map={}
    for struct_name, struct_obj in structures.items():
        time_arr=None
        if struct_name=='AR1':
            # If no int_date => skip
            if 'int_date' not in mdf.columns:
                continue
            time_arr = mdf['int_date'].values
            # sort must be stable => we did above
        res = fit_gee_model(mdf, formula, groups=groups, cov_struct=struct_obj, time=time_arr)
        if res is None:
            continue
        try:
            qic_val= res.qic()
        except:
            qic_val= np.nan
        result_map[struct_name] = (res, qic_val)
    return result_map

def fit_and_pick_best_structure(df, sentiment, measure_type, min_date=None):
    """
    Fit multiple structures => pick best by QIC => build pairwise for that best => return result dict.
    """
    structures_test= {
        'Independence': Independence(),
        'Exchangeable': Exchangeable(),
        'Unstructured': Unstructured(),
        'AR1': Autoregressive()
    }
    res_map= run_gee_and_qic(df, sentiment, measure_type, structures_test, min_date)
    if not res_map:
        return None

    # pick best by QIC
    best_struct=None
    best_qic=np.inf
    all_info=[]
    for nm, (res_obj,qic_val) in res_map.items():
        all_info.append((nm,qic_val))
        if (qic_val is not None) and (qic_val < best_qic):
            best_qic= qic_val
            best_struct= (nm,res_obj)

    if best_struct is None:
        return None
    chosen_struct, best_res= best_struct

    # Build summary
    summary_txt= best_res.summary().as_text()

    # Pairwise
    params= best_res.params
    cov= best_res.cov_params()
    # We can get the actual DF from best_res.model.data.frame if needed:
    model_df= best_res.model.data.frame
    cat_var= model_df['media_category'].astype('category')
    categories= cat_var.cat.categories
    ref_cat= categories[0]

    param_names= best_res.model.exog_names
    cat_to_idx= {ref_cat:0}
    for cat in categories[1:]:
        pname= f"media_category[T.{cat}]"
        cat_to_idx[cat]= param_names.index(pname)

    pairwise_list=[]
    cats= list(categories)
    for i in range(len(cats)):
        for j in range(i+1,len(cats)):
            catA,catB = cats[i],cats[j]
            cvec= np.zeros(len(params))
            if catA==ref_cat and catB!=ref_cat:
                cvec[cat_to_idx[catB]]=-1.
            elif catB==ref_cat and catA!=ref_cat:
                cvec[cat_to_idx[catA]]=1.
            else:
                cvec[cat_to_idx[catA]]=1.
                cvec[cat_to_idx[catB]]= -1.
            diff_est= cvec@params
            diff_var= cvec@cov@cvec
            diff_se= np.sqrt(diff_var) if diff_var>0 else 0
            z= diff_est/diff_se if diff_se>0 else np.inf
            pval= 2*(1- norm.cdf(abs(z)))
            pairwise_list.append((catA,catB,diff_est,diff_se,z,pval))

    pairwise_df= pd.DataFrame(pairwise_list, columns=["CategoryA","CategoryB","Difference","SE","Z","p_value"])
    reject, p_adj, _, _ = multipletests(pairwise_df['p_value'], method='holm')
    pairwise_df['p_value_adj']= p_adj
    pairwise_df['reject_H0']= reject

    # Show in terminal:
    print(f"    => Best structure for sentiment='{sentiment}', measure='{measure_type}' is: {chosen_struct} (QIC={best_qic:.2f})")
    print("       Other structures and QIC =>")
    for (n,q) in all_info:
        print(f"         {n}: QIC={q}")

    return {
        'BestStructure': chosen_struct,
        'QIC_All': all_info,
        'GEE_Summary': summary_txt,
        'Pairwise': pairwise_df
    }

def run_gee_analyses(df):
    # If date is present and parseable, define min_date
    if 'date' in df.columns:
        # coerce to datetime
        df['date']= pd.to_datetime(df['date'], errors='coerce')
        # min_date
        valid_dates= df.dropna(subset=['date'])
        if len(valid_dates)>0:
            min_date= valid_dates['date'].min()
        else:
            min_date= None
    else:
        min_date= None

    # For each sentiment & measure
    results={}
    for s in CATEGORIES:
        results[s]={}
        for measure in ['Quotation','Fulltext']:
            best_res= fit_and_pick_best_structure(df, s, measure, min_date)
            if best_res is not None:
                results[s][measure]= best_res
    return results

def compile_results_into_multiple_workbooks(aggregated_df, stats_df, raw_df, gee_results, plots_dir,
                                            main_excel, raw_excel, lmm_excel, plots_excel, combined_excel):
    # main
    with pd.ExcelWriter(main_excel, engine='openpyxl') as w:
        aggregated_df.to_excel(w, sheet_name='Aggregated_Scores', index=False)
        stats_df.to_excel(w, sheet_name='Mean_Median_Statistics', index=False)

    # raw
    raw_df2 = raw_df.copy()
    raw_df2['media_category'] = raw_df2['media_category'].astype('category')
    raw_df2.sort_values(['media_category','media_outlet'], inplace=True)
    with pd.ExcelWriter(raw_excel, engine='openpyxl') as w:
        raw_df2.to_excel(w, sheet_name='Raw_Data', index=False)
        for sentiment in CATEGORIES:
            scols = [c for c in raw_df2.columns if c.startswith(sentiment+'_')]
            ssub = raw_df2[['media_category','media_outlet']+scols].copy()
            sheetnm= f"Raw_{sentiment[:29]}"
            ssub.to_excel(w, sheet_name=sheetnm, index=False)

    # lmm
    with pd.ExcelWriter(lmm_excel, engine='openpyxl') as w:
        summary_rows=[]
        for sentiment in gee_results:
            for measure_type in gee_results[sentiment]:
                sheet_name = f"GEE_{sentiment[:20]}_{measure_type[:8]}"
                best_struct= gee_results[sentiment][measure_type]['BestStructure']
                qic_list=gee_results[sentiment][measure_type]['QIC_All']
                summary_text= gee_results[sentiment][measure_type]['GEE_Summary']
                pairwise_df= gee_results[sentiment][measure_type]['Pairwise']

                # store summary
                summary_df= pd.DataFrame({'GEE_Summary': summary_text.split('\n')})
                summary_df.to_excel(w, sheet_name=sheet_name, index=False)
                row_start= len(summary_df)+2
                # store best struct & QIC
                info_df= pd.DataFrame({
                    'Info':[
                        f"BestStructure={best_struct}",
                        f"AllStructuresAndQIC={qic_list}"
                    ]
                })
                info_df.to_excel(w, sheet_name=sheet_name, index=False, startrow=row_start)
                row_start+= (len(info_df)+2)
                # store pairwise
                pairwise_df.to_excel(w, sheet_name=sheet_name, index=False, startrow=row_start)

                summary_rows.append({
                    'Sentiment':sentiment,
                    'Measure':measure_type,
                    'BestStructure':best_struct,
                    'SheetName':sheet_name
                })
        indexdf= pd.DataFrame(summary_rows)
        indexdf.to_excel(w, sheet_name='GEE_Results_Index', index=False)

    # plots
    wb= Workbook()
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])

    for s in CATEGORIES:
        qplot= os.path.join(plots_dir,f"quote_{s}.png")
        if os.path.exists(qplot):
            stitle= f"Quote_{s[:28]}"
            ws= wb.create_sheet(title=stitle)
            try:
                im= ExcelImage(qplot)
                im.anchor='A1'
                ws.add_image(im)
            except Exception as e:
                logging.error(f"Error embedding quote {qplot}: {e}")

        fplot= os.path.join(plots_dir,f"fulltext_{s}.png")
        if os.path.exists(fplot):
            stitle2= f"Fulltext_{s[:25]}"
            ws2= wb.create_sheet(title=stitle2)
            try:
                im2=ExcelImage(fplot)
                im2.anchor='A1'
                ws2.add_image(im2)
            except Exception as e:
                logging.error(f"Error embedding fulltext {fplot}: {e}")

    wb.save(plots_excel)

    # combined
    raw_df_clean= raw_df2.copy()
    for col in raw_df_clean.columns:
        if raw_df_clean[col].dtype==object:
            raw_df_clean[col] = raw_df_clean[col].apply(lambda x: ", ".join(x) if isinstance(x,list) else x)

    wb_comb= Workbook()
    if 'Sheet' in wb_comb.sheetnames:
        wb_comb.remove(wb_comb['Sheet'])

    ws_agg= wb_comb.create_sheet("Aggregated_Scores")
    for r in dataframe_to_rows(aggregated_df, index=False, header=True):
        ws_agg.append(r)

    ws_stats= wb_comb.create_sheet("Mean_Median_Statistics")
    for r in dataframe_to_rows(stats_df, index=False, header=True):
        ws_stats.append(r)

    ws_raw= wb_comb.create_sheet("Raw_Data")
    for r in dataframe_to_rows(raw_df_clean, index=False, header=True):
        ws_raw.append(r)

    # Summaries
    ws_gee= wb_comb.create_sheet("GEE_Summaries")
    ws_gee.append(["Sentiment","Measure","BestStructure","SummarySnippet"])
    for s in gee_results:
        for measure in gee_results[s]:
            best_st= gee_results[s][measure]['BestStructure']
            summ = gee_results[s][measure]['GEE_Summary'].split('\n')
            snippet= "\n".join(summ[:6])  # truncated
            ws_gee.append([s, measure, best_st, snippet])

    wb_comb.save(combined_excel)

def main():
    setup_logging()
    print("Single run with multi-structure GEE + QIC comparison + Poisson + clustering by media_outlet.")
    df= load_jsonl(INPUT_JSONL_FILE)
    print(f"Total articles loaded: {len(df)}")

    df= map_media_outlet_to_category(df, MEDIA_CATEGORIES)

    # Split raw
    chunk_size=20000
    for i in range(0, len(df), chunk_size):
        chunk= df.iloc[i:i+chunk_size]
        out_chunk= os.path.join(CSV_OUTPUT_DIR, f"raw_data_part_{i//chunk_size+1}.csv")
        chunk.to_csv(out_chunk, index=False)
        print(f"Saved chunk {i//chunk_size+1} to {out_chunk}")

    print("\nSummary Statistics:")
    print("Total number of articles:", len(df))
    print("\nNumber of articles per outlet:")
    print(df['media_outlet_clean'].value_counts())
    print("\nNumber of articles per category:")
    print(df['media_category'].value_counts())
    print()

    print("Aggregating sentiment/emotion scores per media category...")
    aggregated_df = aggregate_sentiment_scores(df, CATEGORIES)
    aggregated_df = calculate_averages(aggregated_df)
    stats_df= calculate_mean_median(aggregated_df)

    save_aggregated_scores_to_csv(aggregated_df, CSV_OUTPUT_DIR)
    plot_statistics(aggregated_df, OUTPUT_DIR)

    print("Fitting GEE models with multiple correlation structures (QIC) => best structure selection.")
    gee_results= run_gee_analyses(df)

    print("\n** GEE Model + Sensitivity Analysis **")
    for s in CATEGORIES:
        if s in gee_results:
            for measure in gee_results[s]:
                best_s= gee_results[s][measure]['BestStructure']
                print(f"For sentiment='{s}', measure='{measure}', best structure => {best_s}")

    compile_results_into_multiple_workbooks(
        aggregated_df, stats_df, df, gee_results, OUTPUT_DIR,
        OUTPUT_EXCEL_MAIN, OUTPUT_EXCEL_RAW, OUTPUT_EXCEL_LMM, OUTPUT_EXCEL_PLOTS, OUTPUT_EXCEL_COMBINED
    )

    print("Analysis completed successfully.")


if __name__=="__main__":
    main()

