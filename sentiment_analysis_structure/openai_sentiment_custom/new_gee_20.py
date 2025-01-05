import json
import pandas as pd
import matplotlib.pyplot as plt
import seaborn as sns
import os
import logging
from tqdm import tqdm
from openpyxl.drawing.image import Image as ExcelImage
import re
import sys
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import numpy as np
import patsy

import statsmodels.api as sm
from statsmodels.genmod.generalized_estimating_equations import GEE
from statsmodels.genmod.families import Poisson
from statsmodels.genmod.cov_struct import Exchangeable, Independence, Unstructured, Autoregressive
from statsmodels.stats.multitest import multipletests
from scipy.stats import norm, pearsonr

# ------------------------------ #
#        Configuration           #
# ------------------------------ #

INPUT_JSONL_FILE = 'processed_all_articles_merged.jsonl'
OUTPUT_DIR = 'graphs_analysis'
os.makedirs(OUTPUT_DIR, exist_ok=True)

CSV_OUTPUT_DIR = 'csv_raw_scores'
os.makedirs(CSV_OUTPUT_DIR, exist_ok=True)

OUTPUT_EXCEL_MAIN = 'analysis_main.xlsx'
OUTPUT_EXCEL_RAW = 'analysis_raw.xlsx'
OUTPUT_EXCEL_LMM = 'analysis_gee.xlsx'
OUTPUT_EXCEL_PLOTS = 'analysis_plots.xlsx'
OUTPUT_EXCEL_COMBINED = 'analysis_combined.xlsx'
OUTPUT_EXCEL_GRAPHING = 'analysis_for_graphing.xlsx'
OUTPUT_EXCEL_CORRELATION = 'analysis_correlation_data.xlsx'
OUTPUT_EXCEL_COMBINED_ALL = 'analysis_combined_all.xlsx'
OUTPUT_EXCEL_QIC = 'analysis_gee_qic.xlsx'

CATEGORIES = [
    'joy', 'sadness', 'anger', 'fear',
    'surprise', 'disgust', 'trust', 'anticipation',
    'negative_sentiment', 'positive_sentiment'
]

MEDIA_CATEGORIES = {
    'Scientific': ['nature', 'sciam', 'stat', 'newscientist'],
    'Left': ['theatlantic', 'the daily beast', 'the intercept', 'mother jones', 'msnbc', 'slate', 'vox', 'huffpost'],
    'Lean Left': ['ap', 'axios', 'cnn', 'guardian', 'business insider', 'nbcnews', 'npr', 'nytimes', 'politico', 'propublica', 'wapo', 'usa today'],
    'Center': ['reuters', 'marketwatch', 'financial times', 'newsweek', 'forbes'],
    'Lean Right': ['thedispatch', 'epochtimes', 'foxbusiness', 'wsj', 'national review', 'washtimes'],
    'Right': ['breitbart', 'theblaze', 'daily mail', 'dailywire', 'foxnews', 'nypost', 'newsmax'],
}

CATEGORY_ORDER = ["Scientific", "Left", "Lean Left", "Center", "Lean Right", "Right"]

def setup_logging(log_file='analysis.log'):
    logging.basicConfig(
        filename=log_file,
        level=logging.DEBUG,
        format='%(asctime)s - %(levelname)s - %(message)s'
    )
    logging.info("Logging initialized.")

def load_jsonl(jsonl_file):
    records = []
    with open(jsonl_file, 'r', encoding='utf-8') as f:
        for line in tqdm(f, desc="Loading JSONL data"):
            try:
                record = json.loads(line)
                records.append(record)
            except json.JSONDecodeError as e:
                logging.error(f"JSON decoding error: {e}")
    df = pd.DataFrame(records)
    return df

def map_media_outlet_to_category(df, media_categories):
    outlet_to_category = {}
    for category, outlets in media_categories.items():
        for outlet in outlets:
            outlet_to_category[outlet.lower().strip()] = category

    if 'media_outlet' not in df.columns:
        logging.error("'media_outlet' column not found in DataFrame.")
        raise KeyError("'media_outlet' column not found in DataFrame.")

    df['media_outlet_clean'] = df['media_outlet'].str.lower().str.strip()
    df['media_category'] = df['media_outlet_clean'].map(outlet_to_category)
    df['media_category'] = df['media_category'].fillna('Other')
    unmapped_outlets = df[df['media_category'] == 'Other']['media_outlet'].unique()
    if len(unmapped_outlets) > 0:
        logging.warning(f"Unmapped media outlets found: {unmapped_outlets}")
        print(f"Warning: The following media outlets were not mapped and categorized as 'Other': {unmapped_outlets}")
    return df

def compute_article_level_scores(df, sentiment_categories):
    for sentiment in sentiment_categories:
        pattern = f"^{re.escape(sentiment)}_\\d+$"
        matched_cols = [col for col in df.columns if re.match(pattern, col)]
        if matched_cols:
            df[f"{sentiment}_quotation_mean_article"] = df[matched_cols].clip(lower=0).mean(axis=1)
        else:
            df[f"{sentiment}_quotation_mean_article"] = np.nan

        fulltext_col = f"{sentiment}_fulltext"
        if fulltext_col in df.columns:
            df[f"{sentiment}_fulltext_article"] = df[fulltext_col].clip(lower=0)
        else:
            df[f"{sentiment}_fulltext_article"] = np.nan
    return df

def aggregate_sentiment_scores(df, sentiment_categories):
    aggregation = []
    for media_category in MEDIA_CATEGORIES.keys():
        mask = (df['media_category'] == media_category)
        subset = df[mask]

        for sentiment in sentiment_categories:
            pattern = f"^{re.escape(sentiment)}_\\d+$"
            matched_cols = [col for col in df.columns if re.match(pattern, col)]

            if matched_cols:
                quotation_sum = subset[matched_cols].clip(lower=0).sum(skipna=True).sum()
                quotation_count = subset[matched_cols].clip(lower=0).count().sum()
            else:
                quotation_sum = 0
                quotation_count = 0

            fulltext_col = f"{sentiment}_fulltext"
            if fulltext_col in df.columns:
                fulltext_sum = subset[fulltext_col].clip(lower=0).sum(skipna=True)
                fulltext_count = subset[fulltext_col].clip(lower=0).count()
            else:
                fulltext_sum = 0
                fulltext_count = 0

            aggregation.append({
                'Media Category': media_category,
                'Sentiment/Emotion': sentiment,
                'Quotation_Sum': quotation_sum,
                'Quotation_Count': quotation_count,
                'Fulltext_Sum': fulltext_sum,
                'Fulltext_Count': fulltext_count
            })

    aggregated_df = pd.DataFrame(aggregation)
    return aggregated_df

def calculate_averages(aggregated_df):
    aggregated_df['Quotation_Average'] = aggregated_df.apply(
        lambda row: row['Quotation_Sum'] / row['Quotation_Count'] if row['Quotation_Count'] > 0 else None, axis=1
    )
    aggregated_df['Fulltext_Average'] = aggregated_df.apply(
        lambda row: row['Fulltext_Sum'] / row['Fulltext_Count'] if row['Fulltext_Count'] > 0 else None, axis=1
    )
    return aggregated_df

def calculate_mean_median(aggregated_df):
    stats = []
    for sentiment in CATEGORIES:
        sentiment_data = aggregated_df[aggregated_df['Sentiment/Emotion'] == sentiment]

        quotation_avg = sentiment_data['Quotation_Average'].dropna()
        if not quotation_avg.empty:
            mean_quotation = quotation_avg.mean()
        else:
            mean_quotation = None

        fulltext_avg = sentiment_data['Fulltext_Average'].dropna()
        if not fulltext_avg.empty:
            mean_fulltext = fulltext_avg.mean()
        else:
            mean_fulltext = None

        stats.append({
            'Sentiment/Emotion': sentiment,
            'Mean_Quotation_Average': mean_quotation,
            'Mean_Fulltext_Average': mean_fulltext
        })

    stats_df = pd.DataFrame(stats)
    return stats_df

def save_aggregated_scores_to_csv(aggregated_df, csv_output_dir, prefix='aggregated_sentiment_emotion_scores.csv'):
    csv_file = os.path.join(csv_output_dir, prefix)
    try:
        aggregated_df.to_csv(csv_file, index=False)
        print(f"Aggregated sentiment/emotion scores saved to '{csv_file}'.")
        logging.info(f"Aggregated sentiment/emotion scores saved to '{csv_file}'.")
    except Exception as e:
        print(f"Error saving aggregated scores to CSV: {e}")
        logging.error(f"Error saving aggregated scores to CSV: {e}")

def plot_statistics(aggregated_df, output_dir):
    sns.set_style("whitegrid")
    for sentiment in CATEGORIES:
        sentiment_data = aggregated_df[aggregated_df['Sentiment/Emotion'] == sentiment]

        # Mean Quotation-Based Plot
        plt.figure(figsize=(10, 6))
        sns.barplot(
            x='Media Category',
            y='Quotation_Average',
            data=sentiment_data,
            color='steelblue'
        )
        plt.title(f"Mean Quotation-Based '{sentiment.capitalize()}' Scores", fontsize=14)
        plt.xlabel('Media Category', fontsize=12)
        plt.ylabel('Mean Quotation-Based Average Score', fontsize=12)
        plt.xticks(rotation=45, ha='right')
        plt.tight_layout()
        quote_plot_filename = f"quote_{sentiment}.png"
        quote_plot_path = os.path.join(output_dir, quote_plot_filename)
        try:
            plt.savefig(quote_plot_path)
            plt.close()
            print(f"Quotation-Based '{sentiment}' scores plot saved to '{quote_plot_path}'.")
        except Exception as e:
            print(f"Error saving quotation-based '{sentiment}' scores plot: {e}")

        # Mean Fulltext-Based Plot
        plt.figure(figsize=(10, 6))
        sns.barplot(
            x='Media Category',
            y='Fulltext_Average',
            data=sentiment_data,
            color='darkorange'
        )
        plt.title(f"Mean Fulltext-Based '{sentiment.capitalize()}' Scores", fontsize=14)
        plt.xlabel('Media Category', fontsize=12)
        plt.ylabel('Mean Fulltext-Based Average Score', fontsize=12)
        plt.xticks(rotation=45, ha='right')
        plt.tight_layout()
        fulltext_plot_filename = f"fulltext_{sentiment}.png"
        fulltext_plot_path = os.path.join(output_dir, fulltext_plot_filename)
        try:
            plt.savefig(fulltext_plot_path)
            plt.close()
            print(f"Fulltext-Based '{sentiment}' scores plot saved to '{fulltext_plot_path}'.")
        except Exception as e:
            print(f"Error saving fulltext-based '{sentiment}' scores plot: {e}")

def fit_gee_and_pairwise(df, sentiment, measure_type='Quotation', cov_struct_type=Exchangeable):
    # Prepare data
    if measure_type == 'Quotation':
        pattern = f"^{re.escape(sentiment)}_\\d+$"
        matched_cols = [col for col in df.columns if re.match(pattern, col)]
        if not matched_cols:
            return None
        temp_df = df.copy()
        temp_df[f'{sentiment}_quotation_mean'] = temp_df[matched_cols].clip(lower=0).mean(axis=1)
        score_col = f'{sentiment}_quotation_mean'
    else:
        fulltext_col = f"{sentiment}_fulltext"
        if fulltext_col not in df.columns:
            return None
        temp_df = df.copy()
        temp_df[f'{sentiment}_fulltext_clipped'] = temp_df[fulltext_col].clip(lower=0)
        score_col = f'{sentiment}_fulltext_clipped'

    model_df = temp_df.dropna(subset=[score_col, 'media_category', 'media_outlet']).copy()
    if model_df['media_category'].nunique() < 2:
        return None

    if 'date' in model_df.columns:
        if not np.issubdtype(model_df['date'].dtype, np.datetime64):
            model_df['date'] = pd.to_datetime(model_df['date'], errors='coerce')
    model_df = model_df.dropna(subset=['date'])
    if len(model_df) < 2:
        return None

    model_df['media_category'] = model_df['media_category'].astype('category')

    # Only set grid=True for Autoregressive structure
    if cov_struct_type == Autoregressive:
        cov_struct = cov_struct_type(grid=True)
    else:
        cov_struct = cov_struct_type()

    time = None
    if isinstance(cov_struct, Autoregressive):
        model_df = model_df.sort_values(['media_outlet', 'date'])
        model_df['time_index'] = model_df.groupby('media_outlet').cumcount()
        time = model_df['time_index']

    family = Poisson()

    # Construct endog/exog
    y, X = patsy.dmatrices(f"{score_col} ~ media_category", data=model_df, return_type='dataframe')
    # Create GEE model with scale=1
    model = GEE(y, X, groups=model_df["media_outlet"], time=time, family=family, cov_struct=cov_struct, scale=1)
    try:
        results = model.fit()
    except Exception as e:
        logging.error(f"Model failed to fit with {cov_struct_type.__name__} for sentiment {sentiment} ({measure_type}): {e}")
        return None

    summary_text = results.summary().as_text()
    params = results.params
    cov = results.cov_params()
    all_categories = model_df['media_category'].cat.categories
    reference_category = all_categories[0]

    param_names = results.model.exog_names
    cat_to_param_index = {reference_category: 0}
    for cat in all_categories[1:]:
        pname = f"media_category[T.{cat}]"
        cat_to_param_index[cat] = param_names.index(pname)

    pairwise_results = []
    cats = list(all_categories)
    for i in range(len(cats)):
        for j in range(i+1, len(cats)):
            catA = cats[i]
            catB = cats[j]

            c = np.zeros(len(params))
            if catA == reference_category and catB != reference_category:
                c[cat_to_param_index[catB]] = -1.0
            elif catB == reference_category and catA != reference_category:
                c[cat_to_param_index[catA]] = 1.0
            else:
                c[cat_to_param_index[catA]] = 1.0
                c[cat_to_param_index[catB]] = -1.0

            diff_est = c @ params
            diff_var = c @ cov @ c
            diff_se = np.sqrt(diff_var)
            z = diff_est / diff_se
            p_value = 2 * (1 - norm.cdf(abs(z)))

            pairwise_results.append((catA, catB, diff_est, diff_se, z, p_value))

    pairwise_df = pd.DataFrame(pairwise_results, columns=["CategoryA", "CategoryB", "Difference", "SE", "Z", "p_value"])
    reject, p_adj, _, _ = multipletests(pairwise_df['p_value'], method='holm')
    pairwise_df['p_value_adj'] = p_adj
    pairwise_df['reject_H0'] = reject

    # Attempt to compute QIC
    try:
        qic_value = results.qic()
    except:
        qic_value = np.nan

    return {
        'GEE_Summary': summary_text,
        'Pairwise': pairwise_df,
        'QIC': qic_value,
        'Results': results
    }

def try_all_cov_structures(df, sentiment, measure_type):
    structures = [Independence, Exchangeable, Unstructured, Autoregressive]
    results_list = []
    for struct in structures:
        res = fit_gee_and_pairwise(df, sentiment, measure_type=measure_type, cov_struct_type=struct)
        if res is not None:
            qic_val = res['QIC']
            results_list.append((struct.__name__, qic_val, res))
    return results_list

def run_gee_analyses(df):
    all_results = {}
    all_qic_results = []
    for sentiment in CATEGORIES:
        all_results[sentiment] = {}
        for measure_type in ['Quotation', 'Fulltext']:
            results_list = try_all_cov_structures(df, sentiment, measure_type)
            if len(results_list) == 0:
                continue
            results_list = [(sn, (qic if np.issubdtype(type(qic), np.number) else np.nan), r) for (sn, qic, r) in results_list]
            results_list = sorted(results_list, key=lambda x: (np.nan if x[1] is None else x[1]))
            best_struct = results_list[0]
            all_results[sentiment][measure_type] = best_struct[2]

            for struct_name, qic_val, res in results_list:
                all_qic_results.append({
                    'Sentiment': sentiment,
                    'Measure': measure_type,
                    'Structure': struct_name,
                    'QIC': qic_val
                })

    qic_df = pd.DataFrame(all_qic_results)
    return all_results, qic_df

def compile_results_into_multiple_workbooks(aggregated_df, stats_df, raw_df, gee_results, qic_df, plots_dir,
                                            main_excel, raw_excel, lmm_excel, plots_excel, combined_excel, qic_excel):
    with pd.ExcelWriter(main_excel, engine='openpyxl') as writer:
        aggregated_df.to_excel(writer, sheet_name='Aggregated_Scores', index=False)
        stats_df.to_excel(writer, sheet_name='Mean_Median_Statistics', index=False)

    raw_df = raw_df.copy()
    raw_df['media_category'] = pd.Categorical(raw_df['media_category'], categories=CATEGORY_ORDER + ['Other'], ordered=True)
    raw_df = raw_df.sort_values(['media_category', 'media_outlet'])

    with pd.ExcelWriter(raw_excel, engine='openpyxl') as writer:
        raw_df.to_excel(writer, sheet_name='Raw_Data', index=False)
        for sentiment in CATEGORIES:
            sentiment_cols = [c for c in raw_df.columns if c.startswith(sentiment+'_')]
            sentiment_df = raw_df[['media_category', 'media_outlet'] + sentiment_cols].copy()
            sheet_name = f"Raw_{sentiment[:29]}"
            sentiment_df.to_excel(writer, sheet_name=sheet_name, index=False)

    with pd.ExcelWriter(lmm_excel, engine='openpyxl') as writer:
        summary_rows = []
        for sentiment in gee_results:
            for measure_type in gee_results[sentiment]:
                sheet_name = f"GEE_{sentiment[:20]}_{measure_type[:8]}"
                gee_summary = gee_results[sentiment][measure_type]['GEE_Summary']
                pairwise_df = gee_results[sentiment][measure_type]['Pairwise']

                summary_df = pd.DataFrame({'GEE_Summary': gee_summary.split('\n')})
                summary_df.to_excel(writer, sheet_name=sheet_name, index=False)
                startrow = len(summary_df) + 2
                pairwise_df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=startrow)

                summary_rows.append({
                    'Sentiment': sentiment,
                    'Measure': measure_type,
                    'SheetName': sheet_name
                })

        summary_index_df = pd.DataFrame(summary_rows)
        summary_index_df.to_excel(writer, sheet_name='GEE_Results_Index', index=False)

    with pd.ExcelWriter(qic_excel, engine='openpyxl') as writer:
        qic_df.to_excel(writer, sheet_name='QIC_Results', index=False)

    wb = Workbook()
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])

    for sentiment in CATEGORIES:
        quote_plot_path = os.path.join(plots_dir, f"quote_{sentiment}.png")
        if os.path.exists(quote_plot_path):
            sheet_title = f"Quote_{sentiment[:28]}"
            worksheet = wb.create_sheet(title=sheet_title)
            try:
                img = ExcelImage(quote_plot_path)
                img.anchor = 'A1'
                worksheet.add_image(img)
            except Exception as e:
                logging.error(f"Error embedding image '{quote_plot_path}' into Excel: {e}")

        fulltext_plot_path = os.path.join(plots_dir, f"fulltext_{sentiment}.png")
        if os.path.exists(fulltext_plot_path):
            sheet_title = f"Fulltext_{sentiment[:25]}"
            worksheet = wb.create_sheet(title=sheet_title)
            try:
                img = ExcelImage(fulltext_plot_path)
                img.anchor = 'A1'
                worksheet.add_image(img)
            except Exception as e:
                logging.error(f"Error embedding image '{fulltext_plot_path}' into Excel: {e}")

    wb.save(OUTPUT_EXCEL_PLOTS)

    raw_df_cleaned = raw_df.copy()
    # Avoid deprecated applymap usage:
    for col in raw_df_cleaned.columns:
        raw_df_cleaned[col] = raw_df_cleaned[col].apply(lambda x: ", ".join(x) if isinstance(x, list) else x)

    wb_combined = Workbook()
    if 'Sheet' in wb_combined.sheetnames:
        wb_combined.remove(wb_combined['Sheet'])

    ws_agg = wb_combined.create_sheet("Aggregated_Scores")
    for r in dataframe_to_rows(aggregated_df, index=False, header=True):
        ws_agg.append(r)

    ws_stats = wb_combined.create_sheet("Mean_Median_Statistics")
    for r in dataframe_to_rows(stats_df, index=False, header=True):
        ws_stats.append(r)

    ws_raw = wb_combined.create_sheet("Raw_Data")
    for r in dataframe_to_rows(raw_df_cleaned, index=False, header=True):
        ws_raw.append(r)

    ws_gee = wb_combined.create_sheet("GEE_Summaries")
    ws_gee.append(["Sentiment", "Measure", "GEE_Summary"])
    for sentiment in gee_results:
        for measure_type in gee_results[sentiment]:
            summary = gee_results[sentiment][measure_type]['GEE_Summary']
            ws_gee.append([sentiment, measure_type, summary])

    wb_combined.save(OUTPUT_EXCEL_COMBINED)

def create_analysis_for_graphing_file(df, sentiments, category_order, output_file):
    df = df.copy()
    df['media_category'] = pd.Categorical(df['media_category'], categories=category_order + ['Other'], ordered=True)
    df = df[df['media_category'].isin(category_order)]

    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        for sentiment in sentiments:
            q_col = f"{sentiment}_quotation_mean_article"
            f_col = f"{sentiment}_fulltext_article"

            cat_data_quotation = {cat: [] for cat in category_order}
            for idx, row in df.iterrows():
                val = row[q_col] if q_col in df.columns else np.nan
                row_values = {cat: np.nan for cat in category_order}
                if row['media_category'] in category_order:
                    row_values[row['media_category']] = val
                for cat in category_order:
                    cat_data_quotation[cat].append(row_values[cat])
            quotation_df = pd.DataFrame(cat_data_quotation, columns=category_order)

            cat_data_fulltext = {cat: [] for cat in category_order}
            for idx, row in df.iterrows():
                val = row[f_col] if f_col in df.columns else np.nan
                row_values = {cat: np.nan for cat in category_order}
                if row['media_category'] in category_order:
                    row_values[row['media_category']] = val
                for cat in category_order:
                    cat_data_fulltext[cat].append(row_values[cat])
            fulltext_df = pd.DataFrame(cat_data_fulltext, columns=category_order)

            quotation_df.to_excel(writer, sheet_name=f"{sentiment}_Quotation", index=False)
            fulltext_df.to_excel(writer, sheet_name=f"{sentiment}_Fulltext", index=False)

def main():
    setup_logging()
    print("Single run with GEE and pairwise comparisons (Poisson family), testing multiple covariance structures including AR(1) using 'date'.")
    df = load_jsonl(INPUT_JSONL_FILE)
    print(f"Total articles loaded: {len(df)}")
    df = map_media_outlet_to_category(df, MEDIA_CATEGORIES)

    if 'date' not in df.columns:
        print("Warning: 'date' field not found. AR(1) structure may not be applicable.")
    else:
        df['date'] = pd.to_datetime(df['date'], errors='coerce')

    df = compute_article_level_scores(df, CATEGORIES)

    chunk_size = 20000
    for i in range(0, len(df), chunk_size):
        chunk = df.iloc[i:i+chunk_size]
        chunk_csv_path = os.path.join(CSV_OUTPUT_DIR, f"raw_data_part_{i//chunk_size+1}.csv")
        chunk.to_csv(chunk_csv_path, index=False)
        print(f"Saved chunk {i//chunk_size+1} to {chunk_csv_path}")

    print("\nSummary Statistics:")
    print("Total number of articles:", len(df))
    outlet_counts = df['media_outlet_clean'].value_counts()
    print("\nNumber of articles per outlet:")
    print(outlet_counts)
    category_counts = df['media_category'].value_counts()
    print("\nNumber of articles per category:")
    print(category_counts)
    print()

    print("Aggregating sentiment/emotion scores per media category...")
    aggregated_df = aggregate_sentiment_scores(df, CATEGORIES)
    aggregated_df = calculate_averages(aggregated_df)
    stats_df = calculate_mean_median(aggregated_df)

    save_aggregated_scores_to_csv(aggregated_df, CSV_OUTPUT_DIR)

    plot_statistics(aggregated_df, OUTPUT_DIR)

    correlation_results = []
    correlation_scatter_data = {}
    for sentiment in CATEGORIES:
        sentiment_data = aggregated_df[aggregated_df['Sentiment/Emotion'] == sentiment]
        sentiment_data = sentiment_data[sentiment_data['Media Category'].isin(MEDIA_CATEGORIES.keys())]
        sentiment_data = sentiment_data.dropna(subset=['Quotation_Average', 'Fulltext_Average'])
        if len(sentiment_data) > 1:
            corr, _ = pearsonr(sentiment_data['Quotation_Average'], sentiment_data['Fulltext_Average'])
        else:
            corr = np.nan
        correlation_results.append({'Sentiment/Emotion': sentiment, 'Correlation': corr})
        correlation_scatter_data[sentiment] = sentiment_data[['Media Category', 'Quotation_Average', 'Fulltext_Average']].copy()

        plt.figure(figsize=(6, 6))
        sns.scatterplot(x='Quotation_Average', y='Fulltext_Average', data=sentiment_data, hue='Media Category', s=50)
        plt.title(f"Scatter: {sentiment.capitalize()} (Quotation vs Fulltext)")
        plt.xlabel('Quotation_Average')
        plt.ylabel('Fulltext_Average')
        plt.legend(bbox_to_anchor=(1.05, 1), loc='upper left')
        plt.tight_layout()
        scatter_plot_path = os.path.join(OUTPUT_DIR, f"scatter_{sentiment}.png")
        plt.savefig(scatter_plot_path)
        plt.close()
        print(f"Scatter plot for {sentiment} saved to '{scatter_plot_path}'.")

    correlation_df = pd.DataFrame(correlation_results)
    plt.figure(figsize=(10,6))
    sns.barplot(x='Sentiment/Emotion', y='Correlation', data=correlation_df, color='gray')
    plt.title("Correlation Between Quotation and Fulltext Averages")
    plt.xticks(rotation=45, ha='right')
    plt.ylim(-1, 1)
    plt.tight_layout()
    correlation_plot_path = os.path.join(OUTPUT_DIR, "correlation_quotation_fulltext.png")
    plt.savefig(correlation_plot_path)
    plt.close()
    print(f"Correlation plot saved to '{correlation_plot_path}'.")

    with pd.ExcelWriter(OUTPUT_EXCEL_CORRELATION, engine='openpyxl') as writer:
        correlation_df.to_excel(writer, sheet_name='Correlation', index=False)
        for sentiment, sdata in correlation_scatter_data.items():
            sdata.to_excel(writer, sheet_name=f"{sentiment}_Data", index=False)

    combined_list = []
    for sentiment, sdata in correlation_scatter_data.items():
        df_temp = sdata.copy()
        df_temp['Sentiment/Emotion'] = sentiment
        combined_list.append(df_temp)
    combined_all = pd.concat(combined_list, ignore_index=True)

    combined_all['Quotation_Z'] = combined_all.groupby('Sentiment/Emotion')['Quotation_Average'].transform(
        lambda x: (x - x.mean())/x.std() if x.std() != 0 else x - x.mean())
    combined_all['Fulltext_Z'] = combined_all.groupby('Sentiment/Emotion')['Fulltext_Average'].transform(
        lambda x: (x - x.mean())/x.std() if x.std() != 0 else x - x.mean())

    combined_all.to_excel(OUTPUT_EXCEL_COMBINED_ALL, index=False)

    combined_nonan = combined_all.dropna(subset=['Quotation_Z', 'Fulltext_Z'])
    if len(combined_nonan) > 1:
        r_value, _ = pearsonr(combined_nonan['Quotation_Z'], combined_nonan['Fulltext_Z'])
    else:
        r_value = np.nan

    plt.figure(figsize=(10, 8))
    sns.regplot(x='Quotation_Z', y='Fulltext_Z', data=combined_nonan, scatter_kws={'color':'black'}, line_kws={'color':'red'})
    plt.title(f"All Sentiments Combined (Normalized)\nR = {r_value:.2f}")
    plt.xlabel('Normalized Quotation_Average (Z-score)')
    plt.ylabel('Normalized Fulltext_Average (Z-score)')
    plt.tight_layout()
    combined_scatter_plot_path = os.path.join(OUTPUT_DIR, "combined_normalized_scatter.png")
    plt.savefig(combined_scatter_plot_path)
    plt.close()
    print(f"Combined normalized scatter plot saved to '{combined_scatter_plot_path}'.")

    print("No predictor columns specified for collinearity check.")
    print("Fitting GEE models and performing pairwise comparisons with multiple covariance structures (including AR(1))...")
    gee_results, qic_df = run_gee_analyses(df)
    print("GEE analysis and pairwise tests completed with sensitivity analysis and QIC comparison.\n")

    compile_results_into_multiple_workbooks(
        aggregated_df, stats_df, df, gee_results, qic_df, OUTPUT_DIR,
        OUTPUT_EXCEL_MAIN, OUTPUT_EXCEL_RAW, OUTPUT_EXCEL_LMM, OUTPUT_EXCEL_PLOTS, OUTPUT_EXCEL_COMBINED, OUTPUT_EXCEL_QIC
    )

    create_analysis_for_graphing_file(df, CATEGORIES, CATEGORY_ORDER, OUTPUT_EXCEL_GRAPHING)

    print("Analysis completed successfully.")

if __name__ == "__main__":
    main()
