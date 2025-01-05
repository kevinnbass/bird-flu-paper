#!/usr/bin/env python3
# gee_final_best_qic_only.py
"""
Script that:
  1) Loads data from JSONL
  2) Chunks & saves CSV
  3) Performs correlation analyses, aggregates sentiment scores
  4) Finds best QIC combination (Ind/Exch + scale) for each sentiment–measure
  5) Refits each best QIC model => writes summary to analysis_gee.xlsx
  6) No more old 'pairwise' GEE; only best-QIC approach is used.

Excel files generated:
  - analysis_main.xlsx
  - analysis_raw.xlsx
  - analysis_gee.xlsx    (but no Pairwise tabs, only best-QIC summaries)
  - analysis_plots.xlsx
  - analysis_combined.xlsx
"""

import json
import os
import re
import sys
import warnings
import logging
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import seaborn as sns
from tqdm import tqdm

from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.utils.dataframe import dataframe_to_rows

import statsmodels.api as sm
from statsmodels.genmod.generalized_estimating_equations import GEE
from statsmodels.genmod.families import Poisson
from statsmodels.genmod.cov_struct import Independence, Exchangeable
from statsmodels.stats.multitest import multipletests
from scipy.stats import pearsonr, norm

# ------------------------------ #
# Configuration
# ------------------------------ #
INPUT_JSONL_FILE = "processed_all_articles_fixed_2.jsonl"

OUTPUT_DIR = "graphs_analysis"
os.makedirs(OUTPUT_DIR, exist_ok=True)

CSV_OUTPUT_DIR = "csv_raw_scores"
os.makedirs(CSV_OUTPUT_DIR, exist_ok=True)

OUTPUT_EXCEL_MAIN = "analysis_main.xlsx"
OUTPUT_EXCEL_RAW = "analysis_raw.xlsx"
OUTPUT_EXCEL_GEE = "analysis_gee.xlsx"  # We'll use this solely for best-QIC model summaries
OUTPUT_EXCEL_PLOTS = "analysis_plots.xlsx"
OUTPUT_EXCEL_COMBINED = "analysis_combined.xlsx"

LOG_FILE = "analysis.log"

CATEGORIES = [
    "joy","sadness","anger","fear",
    "surprise","disgust","trust","anticipation",
    "negative_sentiment","positive_sentiment"
]

MEDIA_CATEGORIES = {
    "Scientific": ["nature","sciam","stat","newscientist"],
    "Left": ["theatlantic","the daily beast","the intercept","mother jones","msnbc","slate","vox","huffpost"],
    "Lean Left": ["ap","axios","cnn","guardian","business insider","nbcnews","npr","nytimes","politico","propublica","wapo","usa today"],
    "Center": ["reuters","marketwatch","financial times","newsweek","forbes"],
    "Lean Right": ["thedispatch","epochtimes","foxbusiness","wsj","national review","washtimes"],
    "Right": ["breitbart","theblaze","daily mail","dailywire","foxnews","nypost","newsmax"],
}


def setup_logging(log_file=LOG_FILE):
    log_format = "%(asctime)s [%(levelname)s] %(module)s - %(message)s"
    logging.basicConfig(filename=log_file, level=logging.DEBUG, format=log_format)
    console = logging.StreamHandler(sys.stdout)
    console.setLevel(logging.INFO)
    console.setFormatter(logging.Formatter(log_format))
    logging.getLogger().addHandler(console)
    logging.info("Logging initialized (file+console).")

# Silence the "QIC scale=None" warning
warnings.filterwarnings(
    "ignore",
    message="QIC values obtained using scale=None are not appropriate for comparing models"
)


# ------------------------------
# 1) Loading + chunk
# ------------------------------
def load_jsonl(jsonl_file):
    logging.info(f"Loading JSONL from {jsonl_file}")
    records=[]
    with open(jsonl_file,"r",encoding="utf-8") as f:
        for line in tqdm(f, desc="Loading JSONL data"):
            try:
                rec=json.loads(line)
                records.append(rec)
            except json.JSONDecodeError as e:
                logging.error(f"JSON decode error: {e}")
    df=pd.DataFrame(records)
    logging.debug(f"Loaded DataFrame shape={df.shape}")
    return df

def map_media_outlet_to_category(df, media_dict):
    logging.info("Mapping media_outlet -> category")
    out_map={}
    for cat, outls in media_dict.items():
        for o in outls:
            out_map[o.lower().strip()]=cat
    if "media_outlet" not in df.columns:
        raise KeyError("'media_outlet' column not found")

    df["media_outlet_clean"] = df["media_outlet"].str.lower().str.strip()
    df["media_category"] = df["media_outlet_clean"].map(out_map).fillna("Other")
    unmapped = df[df["media_category"]=="Other"]["media_outlet"].unique()
    if len(unmapped)>0:
        logging.warning(f"Unmapped => {unmapped}")
        print(f"Warning: Not mapped => {unmapped}")
    return df

def chunk_and_save(df, chunk_size=20000):
    logging.info(f"Chunking DataFrame len={len(df)} in size={chunk_size}")
    for i in range(0,len(df),chunk_size):
        part=df.iloc[i:i+chunk_size]
        out_csv=os.path.join(CSV_OUTPUT_DIR,f"raw_data_part_{i//chunk_size+1}.csv")
        part.to_csv(out_csv,index=False)
        logging.info(f"Saved chunk {i//chunk_size+1} => {out_csv}")
        print(f"Saved chunk {i//chunk_size+1} to {out_csv}")

def print_basic_stats(df):
    logging.info(f"Basic stats: total articles = {len(df)}")
    print("\nSummary Statistics:")
    print("Total number of articles:", len(df))
    if "media_outlet_clean" in df.columns:
        oc=df["media_outlet_clean"].value_counts()
        logging.debug(f"Articles per outlet:\n{oc}")
        print("\nArticles per outlet:")
        print(oc)

    if "media_category" in df.columns:
        catvc=df["media_category"].value_counts()
        logging.debug(f"Articles per category:\n{catvc}")
        print("\nArticles per category:")
        print(catvc)
    print()

# ------------------------------
# 2) Correlation analysis
# ------------------------------
def analyze_quotation_fulltext_correlation(df):
    """
    Build aggregator for Quotation vs. Fulltext averages -> correlation, scatter, bar.
    """
    logging.info("Analyzing Quotation vs Fulltext correlation.")
    records=[]
    for cat in df["media_category"].dropna().unique():
        dcat=df[df["media_category"]==cat]
        if dcat.empty:
            continue
        for s in CATEGORIES:
            pat=rf"^{s}_\d+$"
            matched=[c for c in dcat.columns if re.match(pat,c)]
            if matched:
                clipped=dcat[matched].clip(lower=0)
                qsum=clipped.sum(skipna=True).sum()
                qcount=clipped.count().sum()
                qavg=qsum/qcount if qcount>0 else np.nan
            else:
                qavg=np.nan

            fcol=f"{s}_fulltext"
            if fcol in dcat.columns:
                fvals=dcat[fcol].clip(lower=0)
                f_sum=fvals.sum(skipna=True)
                f_count=fvals.count()
                favg=f_sum/f_count if f_count>0 else np.nan
            else:
                favg=np.nan

            records.append({
                "MediaCategory":cat,
                "Sentiment":s,
                "Quotation_Average":qavg,
                "Fulltext_Average":favg
            })
    agg_df=pd.DataFrame(records)
    correlation_results=[]
    scatter_map={}

    for s in CATEGORIES:
        sdf=agg_df[(agg_df["Sentiment"]==s)].dropna(subset=["Quotation_Average","Fulltext_Average"])
        if len(sdf)>1:
            corr_val,_=pearsonr(sdf["Quotation_Average"], sdf["Fulltext_Average"])
        else:
            corr_val=np.nan
        correlation_results.append({"Sentiment":s,"Correlation":corr_val})

        if not sdf.empty:
            # scatter
            plt.figure(figsize=(6,5))
            sns.scatterplot(x="Quotation_Average",y="Fulltext_Average",
                            data=sdf,hue="MediaCategory",s=50)
            plt.title(f"{s.capitalize()} (Quotation vs. Fulltext)\nr={corr_val:.3f}")
            plt.tight_layout()
            out_scatter=os.path.join(OUTPUT_DIR,f"scatter_{s}.png")
            try:
                plt.savefig(out_scatter)
            except Exception as e:
                logging.error(f"Err saving scatter {s}: {e}")
            plt.close()

        scatter_map[s]=sdf.copy()

    corr_df=pd.DataFrame(correlation_results)
    # bar
    plt.figure(figsize=(8,5))
    sns.barplot(x="Sentiment",y="Correlation",data=corr_df,color="gray")
    plt.title("Correlation (Quotation vs Fulltext) per Sentiment")
    plt.xticks(rotation=45,ha="right")
    plt.ylim(-1,1)
    plt.tight_layout()
    bar_path=os.path.join(OUTPUT_DIR,"correlation_quotation_fulltext_bar.png")
    try:
        plt.savefig(bar_path)
    except Exception as e:
        logging.error(f"Err saving correlation bar: {e}")
    plt.close()

    # combined z-scatter
    combo_list=[]
    for s,sdf in scatter_map.items():
        if not sdf.empty:
            cpy=sdf.copy()
            cpy["Sentiment"]=s
            combo_list.append(cpy)
    if combo_list:
        cat_all=pd.concat(combo_list,ignore_index=True)
        cat_all["Qmean"]=cat_all.groupby("Sentiment")["Quotation_Average"].transform("mean")
        cat_all["Qstd"]=cat_all.groupby("Sentiment")["Quotation_Average"].transform("std")
        cat_all["Fmean"]=cat_all.groupby("Sentiment")["Fulltext_Average"].transform("mean")
        cat_all["Fstd"]=cat_all.groupby("Sentiment")["Fulltext_Average"].transform("std")
        cat_all["Quotation_Z"]=(cat_all["Quotation_Average"]-cat_all["Qmean"])/cat_all["Qstd"]
        cat_all["Fulltext_Z"]=(cat_all["Fulltext_Average"]-cat_all["Fmean"])/cat_all["Fstd"]
        valid=cat_all.dropna(subset=["Quotation_Z","Fulltext_Z"])
        if len(valid)>1:
            r_val,_=pearsonr(valid["Quotation_Z"], valid["Fulltext_Z"])
        else:
            r_val=np.nan

        plt.figure(figsize=(7,5))
        sns.regplot(x="Quotation_Z",y="Fulltext_Z",data=valid,
                    scatter_kws={"color":"black","alpha":0.6}, line_kws={"color":"red"})
        plt.title(f"All Sentiments Combined (Z-scores)\nr={r_val:.3f}")
        plt.tight_layout()
        comb_path=os.path.join(OUTPUT_DIR,"combined_normalized_scatter.png")
        try:
            plt.savefig(comb_path)
        except Exception as e:
            logging.error(f"Err saving combined z-scatter: {e}")
        plt.close()

    out_csv=os.path.join(CSV_OUTPUT_DIR,"quotation_fulltext_correlation.csv")
    try:
        corr_df.to_csv(out_csv,index=False)
        logging.info(f"Correlation data => {out_csv}")
    except Exception as e:
        logging.error(f"Err saving correlation CSV: {e}")


# ------------------------------
# 3) Aggregation & Stats
# ------------------------------
def aggregate_sentiment_scores(df,sentiment_cats):
    logging.info("Aggregating sentiment/emotion scores by category + sentiment.")
    out=[]
    for mc in MEDIA_CATEGORIES.keys():
        dcat=df[df["media_category"]==mc]
        for s in sentiment_cats:
            pat=rf"^{re.escape(s)}_\d+$"
            matched=[c for c in dcat.columns if re.match(pat,c)]
            if matched:
                qsum=dcat[matched].clip(lower=0).sum(skipna=True).sum()
                qcount=dcat[matched].clip(lower=0).count().sum()
            else:
                qsum,qcount=(0,0)

            fcol=f"{s}_fulltext"
            if fcol in dcat.columns:
                ft_clipped=dcat[fcol].clip(lower=0)
                f_sum=ft_clipped.sum(skipna=True)
                f_count=ft_clipped.count()
            else:
                f_sum,f_count=(0,0)

            out.append({
                "Media Category":mc,
                "Sentiment/Emotion":s,
                "Quotation_Sum":qsum,
                "Quotation_Count":qcount,
                "Fulltext_Sum":f_sum,
                "Fulltext_Count":f_count
            })
    return pd.DataFrame(out)

def calculate_averages(agg_df):
    logging.info("Calculating Quotation_Average, Fulltext_Average per row.")
    def safe_div(a,b):
        return a/b if b>0 else None
    agg_df["Quotation_Average"]=agg_df.apply(lambda r: safe_div(r["Quotation_Sum"],r["Quotation_Count"]), axis=1)
    agg_df["Fulltext_Average"]=agg_df.apply(lambda r: safe_div(r["Fulltext_Sum"],r["Fulltext_Count"]), axis=1)
    return agg_df

def calculate_mean_median(agg_df):
    logging.info("Computing global mean/median of Quotation/Fulltext averages.")
    stats=[]
    for s in CATEGORIES:
        sub=agg_df[agg_df["Sentiment/Emotion"]==s]
        qa=sub["Quotation_Average"].dropna()
        fa=sub["Fulltext_Average"].dropna()
        mean_q=qa.mean() if len(qa)>0 else None
        med_q=qa.median() if len(qa)>0 else None
        mean_f=fa.mean() if len(fa)>0 else None
        med_f=fa.median() if len(fa)>0 else None
        stats.append({
            "Sentiment/Emotion":s,
            "Mean_Quotation_Average":mean_q,
            "Median_Quotation_Average":med_q,
            "Mean_Fulltext_Average":mean_f,
            "Median_Fulltext_Average":med_f
        })
    return pd.DataFrame(stats)

def save_aggregated_scores_to_csv(aggregated_df,out_dir,prefix="aggregated_sentiment_emotion_scores.csv"):
    csv_file=os.path.join(out_dir,prefix)
    try:
        aggregated_df.to_csv(csv_file,index=False)
        print(f"Aggregated sentiment/emotion scores => {csv_file}")
        logging.info(f"Aggregated => {csv_file}")
    except Exception as e:
        logging.error(f"Error saving aggregated scores => {e}")

def plot_statistics(agg_df, out_dir):
    sns.set_style("whitegrid")
    for s in CATEGORIES:
        sdata=agg_df[agg_df["Sentiment/Emotion"]==s]
        # Quotation
        plt.figure(figsize=(10,6))
        sns.barplot(x="Media Category",y="Quotation_Average",data=sdata,color="steelblue")
        plt.title(f"Mean Quotation-Based '{s.capitalize()}' Scores")
        plt.xticks(rotation=45,ha="right")
        plt.tight_layout()
        q_png=os.path.join(out_dir,f"quote_{s}.png")
        try:
            plt.savefig(q_png)
            plt.close()
        except Exception as e:
            logging.error(f"Err saving quote {s}: {e}")

        # Fulltext
        plt.figure(figsize=(10,6))
        sns.barplot(x="Media Category",y="Fulltext_Average",data=sdata,color="darkorange")
        plt.title(f"Mean Fulltext-Based '{s.capitalize()}' Scores")
        plt.xticks(rotation=45,ha="right")
        plt.tight_layout()
        f_png=os.path.join(out_dir,f"fulltext_{s}.png")
        try:
            plt.savefig(f_png)
            plt.close()
        except Exception as e:
            logging.error(f"Err saving fulltext {s}: {e}")


# ------------------------------
# 4) GEE scale computations
# ------------------------------
def compute_pearson_scale(y,mu,df_resid):
    if df_resid<=0:
        return np.nan
    y = np.asarray(y)
    mu = np.asarray(mu)
    resid=(y-mu)/np.sqrt(mu)
    chi2=np.sum(resid**2)
    return chi2/df_resid

def compute_deviance_scale(y,mu,df_resid):
    if df_resid<=0:
        return np.nan
    y=np.asarray(y)
    mu=np.asarray(mu)
    dev_array=np.zeros_like(y,dtype=float)
    for i in range(len(y)):
        if y[i]>0:
            if mu[i]>0:
                val=y[i]*np.log(y[i]/mu[i])
                dev_array[i] = val - (y[i]-mu[i])
            else:
                dev_array[i] = np.nan
        else:
            dev_array[i] = -(0 - mu[i])
    dev=2*np.nansum(dev_array)
    return dev/df_resid

def compute_ub_scale(y,mu,df_resid):
    if df_resid<=0:
        return np.nan
    val=compute_pearson_scale(y,mu,df_resid)
    return 1.1*val if not np.isnan(val) else np.nan

def compute_bc_scale(y,mu,df_resid):
    if df_resid<=0:
        return np.nan
    val=compute_deviance_scale(y,mu,df_resid)
    return 0.9*val if not np.isnan(val) else np.nan

def fit_and_compute_scales(model):
    base_res=model.fit(scale=None)
    base_qic=base_res.qic()
    y=base_res.model.endog
    mu=base_res.fittedvalues
    n=len(y)
    p=len(base_res.params)
    df_resid=n-p

    pear=compute_pearson_scale(y,mu,df_resid)
    dev=compute_deviance_scale(y,mu,df_resid)
    ubv=compute_ub_scale(y,mu,df_resid)
    bcv=compute_bc_scale(y,mu,df_resid)

    results={}
    if isinstance(base_qic,tuple):
        q_m,q_a=base_qic
    else:
        q_m,q_a=base_qic,None
    results["none"]=(q_m,q_a,None)

    from math import isnan
    for (nm,val) in [("pearson",pear),("deviance",dev),("ub",ubv),("bc",bcv)]:
        if (not isnan(val)) and (df_resid>0):
            re_res=model.fit(scale=val)
            re_qic=re_res.qic()
            if isinstance(re_qic,tuple):
                qq_m,qq_a=re_qic
            else:
                qq_m,qq_a=re_qic,None
            results[nm]=(qq_m,qq_a,val)
        else:
            results[nm]=(np.nan,None,val)
    return results

# ------------------------------
# 5) Run best QIC approach
# ------------------------------
def run_gee_for_sentiment_measure_best_qic(df, sentiment, measure):
    d2=df.copy()
    if measure=="Quotation":
        pat=rf"^{re.escape(sentiment)}_\d+$"
        matched=[c for c in d2.columns if re.match(pat,c)]
        if not matched:
            return None
        d2["_score_col"]=d2[matched].clip(lower=0).mean(axis=1)
    else:
        fcol=f"{sentiment}_fulltext"
        if fcol not in d2.columns:
            return None
        d2["_score_col"]=d2[fcol].clip(lower=0)

    needed=["_score_col","media_outlet_clean","media_category"]
    d2=d2.dropna(subset=needed)
    if len(d2)<2 or d2["media_category"].nunique()<2:
        return None

    best_qic_main=np.inf
    best_combo=None
    all_records=[]

    structures=[Independence(), Exchangeable()]
    for cov_obj in structures:
        st_name=cov_obj.__class__.__name__
        model=GEE.from_formula("_score_col ~ media_category",
            groups="media_outlet_clean",data=d2,
            family=Poisson(),cov_struct=cov_obj)
        scale_res=fit_and_compute_scales(model)
        for sm,(q_m,q_a,sc) in scale_res.items():
            all_records.append({
                "Sentiment":sentiment,
                "Measure":measure,
                "Structure":st_name,
                "ScaleMethod":sm,
                "NumericScale":sc,
                "QIC_main":q_m,
                "QIC_alt":q_a
            })
            if (q_m<best_qic_main) and (not np.isnan(q_m)):
                best_qic_main=q_m
                best_combo=(st_name,sm,q_m,sc)

    if best_combo is None:
        return None
    best_struct,best_scale,best_qic_val,best_scale_num=best_combo
    return {
        "Sentiment":sentiment,
        "Measure":measure,
        "Best_Structure":best_struct,
        "Best_Scale":best_scale,
        "Best_QIC_main":best_qic_val,
        "AllCombos":pd.DataFrame(all_records),
        "Summary":f"Best scale={best_scale}, numeric={best_scale_num}"
    }

def run_gee_analyses_best_qic(df):
    logging.info("Running GEE approach => find best QIC among Ind/Exch + scale=[none,pearson,deviance,ub,bc].")
    best_rows=[]
    allcombo_list=[]
    for s in CATEGORIES:
        for meas in ["Quotation","Fulltext"]:
            info=run_gee_for_sentiment_measure_best_qic(df,s,meas)
            if info is not None:
                best_rows.append({
                    "Sentiment":info["Sentiment"],
                    "Measure":info["Measure"],
                    "Best_Structure":info["Best_Structure"],
                    "Best_Scale":info["Best_Scale"],
                    "Best_QIC_main":info["Best_QIC_main"],
                    "Summary":info["Summary"]
                })
                allcombo_list.append(info["AllCombos"])
    best_df=pd.DataFrame(best_rows)
    combos_df=pd.concat(allcombo_list,ignore_index=True) if allcombo_list else pd.DataFrame()
    return best_df, combos_df


# ------------------------------
# 6) Refit best QIC => Summaries
# ------------------------------
def refit_best_gee(df, sentiment, measure, structure_name, scale_method):
    """
    For the best structure & scale => refit model => return .summary().as_text()
    """
    d2=df.copy()
    if measure=="Quotation":
        pat=rf"^{re.escape(sentiment)}_\d+$"
        matched=[c for c in d2.columns if re.match(pat,c)]
        if not matched:
            return None
        d2["_score_col"]=d2[matched].clip(lower=0).mean(axis=1)
    else:
        fcol=f"{sentiment}_fulltext"
        if fcol not in d2.columns:
            return None
        d2["_score_col"]=d2[fcol].clip(lower=0)

    needed=["_score_col","media_outlet_clean","media_category"]
    d2=d2.dropna(subset=needed)
    if len(d2)<2 or d2["media_category"].nunique()<2:
        return None

    # pick correlation structure
    if structure_name=="Independence":
        cov_obj=Independence()
    else:
        cov_obj=Exchangeable()

    model=GEE.from_formula(
        "_score_col ~ media_category",
        groups="media_outlet_clean",
        data=d2,
        family=Poisson(),
        cov_struct=cov_obj
    )
    base_res=model.fit(scale=None)
    if scale_method=="none":
        final_res=base_res
    else:
        # must compute numeric scale for that method
        y=base_res.model.endog
        mu=base_res.fittedvalues
        n=len(y)
        p=len(base_res.params)
        df_resid=n-p
        if scale_method=="pearson":
            scale_val=compute_pearson_scale(y,mu,df_resid)
        elif scale_method=="deviance":
            scale_val=compute_deviance_scale(y,mu,df_resid)
        elif scale_method=="ub":
            scale_val=compute_ub_scale(y,mu,df_resid)
        elif scale_method=="bc":
            scale_val=compute_bc_scale(y,mu,df_resid)
        else:
            scale_val=None

        if (scale_val is not None) and (not np.isnan(scale_val)):
            final_res=model.fit(scale=scale_val)
        else:
            # fallback if scale_val is invalid
            final_res=base_res

    return final_res.summary().as_text()


# ------------------------------
# 7) Excel compilation
# ------------------------------
def compile_results_into_multiple_workbooks(
    aggregated_df, stats_df, raw_df,
    df_best_qic,  # we keep the best QIC results
    df_all_combos,
    # No old pairwise
    plots_dir,
    main_excel, raw_excel, gee_excel, plots_excel, combined_excel,
    df_full # the entire data for refits
):
    logging.info("Compiling results into multiple Excel files (WITHOUT old pairwise).")

    # main
    with pd.ExcelWriter(main_excel,engine="openpyxl") as w:
        aggregated_df.to_excel(w,"Aggregated_Scores",index=False)
        stats_df.to_excel(w,"Mean_Median_Statistics",index=False)

    # raw
    with pd.ExcelWriter(raw_excel,engine="openpyxl") as w:
        raw_df.to_excel(w,"Raw_Data",index=False)
        for s in CATEGORIES:
            s_cols=[c for c in raw_df.columns if c.startswith(s+"_")]
            s_df=raw_df[["media_category","media_outlet"]+s_cols].copy()
            sheet_name=f"Raw_{s[:29]}"
            s_df.to_excel(w,sheet_name,index=False)

    # analysis_gee => only best QIC + new refitted summary
    with pd.ExcelWriter(gee_excel,engine="openpyxl") as w:
        # Summaries for each sentiment–measure from best QIC approach
        summary_rows=[]
        for idx,row in df_best_qic.iterrows():
            s=row["Sentiment"]
            meas=row["Measure"]
            st=row["Best_Structure"]
            sc_meth=row["Best_Scale"]
            sheet_name=f"BestQIC_{s[:10]}_{meas[:8]}"

            # Refit
            refit_summary=refit_best_gee(df_full,s,meas,st,sc_meth)
            if refit_summary is None:
                # skip if can't refit
                summdf=pd.DataFrame({"Summary":["No valid model"]})
            else:
                splitted=refit_summary.split("\n")
                summdf=pd.DataFrame({"Summary":splitted})

            summdf.to_excel(w,sheet_name=sheet_name,index=False)
            summary_rows.append({
                "Sentiment": s,
                "Measure": meas,
                "TabName": sheet_name,
                "BestStructure": st,
                "BestScale": sc_meth,
                "QIC": row["Best_QIC_main"]
            })

        # Then store the best combos themselves
        df_idx=pd.DataFrame(summary_rows)
        df_idx.to_excel(w,"BestQIC_Index",index=False)

        # Also store the full combos if you want
        if not df_all_combos.empty:
            df_all_combos.to_excel(w,"All_Combos",index=False)

    # plots
    wb_plots=Workbook()
    if "Sheet" in wb_plots.sheetnames:
        wb_plots.remove(wb_plots["Sheet"])
    for s in CATEGORIES:
        q_png=os.path.join(plots_dir,f"quote_{s}.png")
        if os.path.exists(q_png):
            st=f"Quote_{s[:28]}"
            ws=wb_plots.create_sheet(title=st)
            try:
                img=ExcelImage(q_png)
                img.anchor="A1"
                ws.add_image(img)
            except Exception as e:
                logging.error(f"Embed {q_png}: {e}")

        f_png=os.path.join(plots_dir,f"fulltext_{s}.png")
        if os.path.exists(f_png):
            st=f"Fulltext_{s[:25]}"
            ws=wb_plots.create_sheet(title=st)
            try:
                img=ExcelImage(f_png)
                img.anchor="A1"
                ws.add_image(img)
            except Exception as e:
                logging.error(f"Embed {f_png}: {e}")

    corr_bar=os.path.join(plots_dir,"correlation_quotation_fulltext_bar.png")
    if os.path.exists(corr_bar):
        ws=wb_plots.create_sheet("Correlation_Bar")
        try:
            img=ExcelImage(corr_bar)
            img.anchor="A1"
            ws.add_image(img)
        except Exception as e:
            logging.error(f"Embed {corr_bar}: {e}")

    comb_path=os.path.join(plots_dir,"combined_normalized_scatter.png")
    if os.path.exists(comb_path):
        ws=wb_plots.create_sheet("Combined_ZScatter")
        try:
            img=ExcelImage(comb_path)
            img.anchor="A1"
            ws.add_image(img)
        except Exception as e:
            logging.error(f"Embed {comb_path}: {e}")

    wb_plots.save(plots_excel)

    # combined
    raw_clean=raw_df.copy()
    # if you want to map lists to string
    raw_clean=raw_clean.applymap(lambda x:", ".join(x) if isinstance(x,list) else x)
    wb_combined=Workbook()
    if "Sheet" in wb_combined.sheetnames:
        wb_combined.remove(wb_combined["Sheet"])

    ws_agg=wb_combined.create_sheet("Aggregated_Scores")
    for r in dataframe_to_rows(aggregated_df,index=False,header=True):
        ws_agg.append(r)

    ws_stats=wb_combined.create_sheet("Mean_Median_Statistics")
    for r in dataframe_to_rows(stats_df,index=False,header=True):
        ws_stats.append(r)

    ws_raw=wb_combined.create_sheet("Raw_Data")
    for r in dataframe_to_rows(raw_clean,index=False,header=True):
        ws_raw.append(r)

    ws_bestqic=wb_combined.create_sheet("BestQIC_Index")
    for r in dataframe_to_rows(df_best_qic,index=False,header=True):
        ws_bestqic.append(r)

    wb_combined.save(combined_excel)


# ------------------------------
def main():
    setup_logging()
    logging.info("Starting analysis: only best-QIC GEE approach, no old pairwise method.")

    df=load_jsonl(INPUT_JSONL_FILE)
    print(f"Total articles loaded: {len(df)}")
    df=map_media_outlet_to_category(df, MEDIA_CATEGORIES)
    chunk_and_save(df,20000)
    print_basic_stats(df)

    print("Performing Quotation vs Fulltext correlation analysis...")
    analyze_quotation_fulltext_correlation(df)

    print("Aggregating sentiment/emotion scores per media category...")
    agg_df=aggregate_sentiment_scores(df,CATEGORIES)
    agg_df=calculate_averages(agg_df)
    stats_df=calculate_mean_median(agg_df)
    save_aggregated_scores_to_csv(agg_df,CSV_OUTPUT_DIR)
    plot_statistics(agg_df,OUTPUT_DIR)

    # NO old GEE approach
    # Directly run new best QIC approach
    print("Fitting GEE approach => best QIC for each sentiment–measure...")
    best_qic_df, all_combos_df=run_gee_analyses_best_qic(df)
    print("Best QIC approach completed.\n")

    # compile results
    compile_results_into_multiple_workbooks(
        aggregated_df=agg_df,
        stats_df=stats_df,
        raw_df=df,
        df_best_qic=best_qic_df,
        df_all_combos=all_combos_df,
        plots_dir=OUTPUT_DIR,
        main_excel=OUTPUT_EXCEL_MAIN,
        raw_excel=OUTPUT_EXCEL_RAW,
        gee_excel=OUTPUT_EXCEL_GEE,
        plots_excel=OUTPUT_EXCEL_PLOTS,
        combined_excel=OUTPUT_EXCEL_COMBINED,
        df_full=df  # needed for refitting the best model
    )

    print("Analysis completed successfully.")
    logging.info("Analysis completed successfully.")

if __name__=="__main__":
    main()
