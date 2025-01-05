from openpyxl import load_workbook, Workbook

wb = load_workbook("analysis_results.xlsx")

# Create a new workbook and copy a few sheets
wb_part1 = Workbook()
for sheet_name in ["Aggregated_Scores", "Mean_Median_Statistics"]:
    ws = wb[sheet_name]
    new_ws = wb_part1.create_sheet(sheet_name)
    for row in ws.iter_rows(values_only=True):
        new_ws.append(row)

# Remove the default empty sheet created by Workbook()
if "Sheet" in wb_part1.sheetnames:
    std = wb_part1["Sheet"]
    wb_part1.remove(std)

wb_part1.save("analysis_results_part1.xlsx")

# Repeat as needed for other sheets
