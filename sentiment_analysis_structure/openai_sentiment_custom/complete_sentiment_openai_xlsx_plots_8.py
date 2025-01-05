import json
import pandas as pd
import matplotlib.pyplot as plt
import seaborn as sns
import os
import logging
from tqdm import tqdm
from openpyxl.drawing.image import Image as ExcelImage

# ------------------------------ #
#        Configuration           #
# ------------------------------ #

# Input JSONL file produced by the first script
INPUT_JSONL_FILE = 'sentiment_emotion_analysis_openai.jsonl'

# Output directories and files
OUTPUT_DIR = 'graphs_analysis'
os.makedirs(OUTPUT_DIR, exist_ok=True)

CSV_OUTPUT_DIR = 'csv_raw_scores'
os.makedirs(CSV_OUTPUT_DIR, exist_ok=True)

OUTPUT_EXCEL_FILE = 'analysis_results.xlsx'

# Define the ten sentiment/emotion categories
CATEGORIES = [
    'joy', 'sadness', 'anger', 'fear',
    'surprise', 'disgust', 'trust', 'anticipation',
    'negative_sentiment', 'positive_sentiment'
]

# Define the six media categories and their corresponding outlets
MEDIA_CATEGORIES = {
    'Scientific': ['Nature', 'SciAm', 'STAT', 'NewScientist'],
    'Left': ['TheAtlantic', 'The Daily Beast', 'The Intercept', 'Mother Jones', 'MSNBC', 'Slate', 'Vox', 'HuffPost'],
    'Lean Left': ['AP', 'Axios', 'CNN', 'Guardian', 'Business Insider', 'NBCNews', 'NPR', 'NYTimes', 'Politico', 'ProPublica', 'WaPo', 'USA Today'],
    'Center': ['Reuters', 'MarketWatch', 'Financial Times', 'Newsweek', 'Forbes'],
    'Lean Right': ['TheDispatch', 'EpochTimes', 'FoxBusiness', 'WSJ', 'National Review', 'WashTimes'],
    'Right': ['Breitbart', 'TheBlaze', 'Daily Mail', 'DailyWire', 'FoxNews', 'NYPost', 'Newsmax'],
}

# ------------------------------ #
#         Helper Functions       #
# ------------------------------ #

def load_jsonl(jsonl_file):
    """
    Loads data from a JSON Lines (JSONL) file into a pandas DataFrame.

    Args:
        jsonl_file (str): Path to the JSONL file.

    Returns:
        pd.DataFrame: DataFrame containing all articles with sentiment/emotion fields.
    """
    records = []
    with open(jsonl_file, 'r', encoding='utf-8') as f:
        for line in tqdm(f, desc="Loading JSONL data"):
            try:
                record = json.loads(line)
                records.append(record)
            except json.JSONDecodeError as e:
                logging.error(f"JSON decoding error: {e}")
    df = pd.DataFrame(records)
    return df

def map_media_outlet_to_category(df, media_categories):
    """
    Maps each media outlet to its corresponding media category and extracts the media outlet name.

    Args:
        df (pd.DataFrame): DataFrame containing the data.
        media_categories (dict): Dictionary mapping categories to outlets.

    Returns:
        pd.DataFrame: Updated DataFrame with new 'media_category' and 'media_outlet_name' columns.
    """
    # Create a reverse mapping from outlet to category
    outlet_to_category = {}
    for category, outlets in media_categories.items():
        for outlet in outlets:
            outlet_to_category[outlet.lower().strip()] = category

    # Check if 'media_outlet' field exists
    if 'media_outlet' not in df.columns:
        logging.error("'media_outlet' column not found in DataFrame.")
        raise KeyError("'media_outlet' column not found in DataFrame.")
    
    # Clean 'media_outlet' field
    df['media_outlet_clean'] = df['media_outlet'].str.lower().str.strip()
    
    # Map 'media_outlet_clean' to 'media_category'
    df['media_category'] = df['media_outlet_clean'].map(outlet_to_category)

    # Handle unmapped outlets by assigning 'Other'
    df['media_category'] = df['media_category'].fillna('Other')  # Direct assignment to avoid FutureWarning
    unmapped_outlets = df[df['media_category'] == 'Other']['media_outlet'].unique()
    if len(unmapped_outlets) > 0:
        logging.warning(f"Unmapped media outlets found: {unmapped_outlets}")
        print(f"Warning: The following media outlets were not mapped and categorized as 'Other': {unmapped_outlets}")
    
    # Extract the original media outlet name for clarity
    df['media_outlet_name'] = df['media_outlet'].str.strip()

    return df

def add_four_to_scores(df, sentiment_categories):
    """
    Adds +4 to every sentiment/emotion score in the DataFrame.

    Args:
        df (pd.DataFrame): DataFrame containing the data.
        sentiment_categories (list): List of sentiment/emotion categories.

    Returns:
        pd.DataFrame: Updated DataFrame with adjusted sentiment/emotion scores.
    """
    for sentiment in sentiment_categories:
        if sentiment in df.columns:
            df[sentiment] = df[sentiment] + 4
        else:
            logging.warning(f"Sentiment/Emotion category '{sentiment}' not found in DataFrame. Skipping addition.")
    return df

def create_combined_columns(df, media_categories, sentiment_categories):
    """
    Creates new columns for each "Category: Media Outlet" combination and populates them with sentiment/emotion scores.

    Args:
        df (pd.DataFrame): DataFrame containing the data.
        media_categories (dict): Dictionary mapping categories to outlets.
        sentiment_categories (list): List of sentiment/emotion categories.

    Returns:
        pd.DataFrame: Updated DataFrame with new combined columns.
    """
    # Iterate through each media category and its outlets
    for category, outlets in media_categories.items():
        for outlet in outlets:
            column_label = f"{category}: {outlet}"
            # Initialize the new column with NaN
            df[column_label] = None
            # Filter rows matching the current category and outlet
            mask = (df['media_category'] == category) & (df['media_outlet_name'].str.lower() == outlet.lower())
            # Assign the sentiment/emotion scores to the new column
            for sentiment in sentiment_categories:
                sentiment_column = f"{sentiment}"
                if sentiment_column in df.columns:
                    df.loc[mask, column_label] = df.loc[mask, sentiment_column]
                else:
                    logging.warning(f"Sentiment/Emotion category '{sentiment}' not found in DataFrame.")
    return df

def calculate_mean_median(df, combined_columns, sentiment_categories):
    """
    Calculates the mean and median for each sentiment/emotion category per combined media category and outlet.

    Args:
        df (pd.DataFrame): DataFrame containing the data.
        combined_columns (list): List of combined "Category: Media Outlet" column names.
        sentiment_categories (list): List of sentiment/emotion categories.

    Returns:
        pd.DataFrame: DataFrame containing mean and median for each combined category and sentiment/emotion.
    """
    stats = []

    for column in combined_columns:
        for sentiment in sentiment_categories:
            if column in df.columns and df[sentiment].notna().any():
                mean_val = df.loc[df[column].notna(), sentiment].mean()
                median_val = df.loc[df[column].notna(), sentiment].median()
            else:
                logging.warning(f"Sentiment/Emotion category '{sentiment}' not found or no data for '{column}'.")
                mean_val = None
                median_val = None
            stats.append({
                'Combined Category': column,
                'Sentiment/Emotion': sentiment,
                'Mean': mean_val,
                'Median': median_val
            })

    stats_df = pd.DataFrame(stats)
    return stats_df

def save_raw_scores_to_csv(df, combined_columns, sentiment_categories, csv_output_dir):
    """
    Saves raw scores for each sentiment/emotion category into separate CSV files based on combined columns.

    Args:
        df (pd.DataFrame): DataFrame containing the data.
        combined_columns (list): List of combined "Category: Media Outlet" column names.
        sentiment_categories (list): List of sentiment/emotion categories.
        csv_output_dir (str): Directory to save the CSV files.

    Returns:
        None
    """
    for sentiment in sentiment_categories:
        data = {}
        # Determine the maximum number of entries among combined columns for the current sentiment
        max_length = max(len(df[df[col].notna()]) for col in combined_columns)
        for col in combined_columns:
            subset = df[df[col].notna()][sentiment].tolist()
            # Pad with NaN to ensure equal length
            if len(subset) < max_length:
                subset += [None] * (max_length - len(subset))
            data[col] = subset
        csv_df = pd.DataFrame(data)
        csv_file = os.path.join(csv_output_dir, f"{sentiment}_raw_scores.csv")
        try:
            csv_df.to_csv(csv_file, index=False)
            print(f"Raw scores for '{sentiment}' saved to '{csv_file}'.")
            logging.info(f"Raw scores for '{sentiment}' saved to '{csv_file}'.")
        except Exception as e:
            print(f"Error saving raw scores to CSV for '{sentiment}': {e}")
            logging.error(f"Error saving raw scores to CSV for '{sentiment}': {e}")

def plot_median_scores(stats_df, sentiment_categories, output_dir):
    """
    Generates bar plots for the median scores of each sentiment/emotion category across combined media categories and outlets.

    Args:
        stats_df (pd.DataFrame): DataFrame containing mean and median statistics.
        sentiment_categories (list): List of sentiment/emotion categories.
        output_dir (str): Directory to save the plots.

    Returns:
        None
    """
    sns.set_style("whitegrid")
    palette = sns.color_palette("viridis", len(stats_df['Combined Category'].unique()))

    for sentiment in sentiment_categories:
        subset = stats_df[stats_df['Sentiment/Emotion'] == sentiment]
        if subset.empty:
            logging.warning(f"No data available for sentiment/emotion category '{sentiment}'. Skipping plot.")
            continue

        plt.figure(figsize=(12, 8))
        sns.barplot(x='Combined Category', y='Median', data=subset, palette=palette)
        plt.title(f"Median Scores for '{sentiment.capitalize()}' Across Media Categories and Outlets", fontsize=16)
        plt.xlabel('Media Category: Media Outlet', fontsize=14)
        plt.ylabel('Median Score', fontsize=14)
        plt.ylim(-4, 8)  # Adjusted to accommodate +4 addition
        plt.xticks(rotation=90, ha='right')
        plt.tight_layout()
        plot_file = os.path.join(output_dir, f"{sentiment}_median_plot.png")
        try:
            plt.savefig(plot_file)
            plt.close()
            print(f"Median plot for '{sentiment}' saved to '{plot_file}'.")
            logging.info(f"Median plot for '{sentiment}' saved to '{plot_file}'.")
        except Exception as e:
            print(f"Error saving plot for '{sentiment}': {e}")
            logging.error(f"Error saving plot for '{sentiment}': {e}")

def plot_mean_scores(stats_df, sentiment_categories, output_dir):
    """
    Generates bar plots for the mean scores of each sentiment/emotion category across combined media categories and outlets.

    Args:
        stats_df (pd.DataFrame): DataFrame containing mean and median statistics.
        sentiment_categories (list): List of sentiment/emotion categories.
        output_dir (str): Directory to save the plots.

    Returns:
        None
    """
    sns.set_style("whitegrid")
    palette = sns.color_palette("magma", len(stats_df['Combined Category'].unique()))

    for sentiment in sentiment_categories:
        subset = stats_df[stats_df['Sentiment/Emotion'] == sentiment]
        if subset.empty:
            logging.warning(f"No data available for sentiment/emotion category '{sentiment}'. Skipping plot.")
            continue

        plt.figure(figsize=(12, 8))
        sns.barplot(x='Combined Category', y='Mean', data=subset, palette=palette)
        plt.title(f"Mean Scores for '{sentiment.capitalize()}' Across Media Categories and Outlets", fontsize=16)
        plt.xlabel('Media Category: Media Outlet', fontsize=14)
        plt.ylabel('Mean Score', fontsize=14)
        plt.ylim(-4, 8)  # Adjusted to accommodate +4 addition
        plt.xticks(rotation=90, ha='right')
        plt.tight_layout()
        plot_file = os.path.join(output_dir, f"{sentiment}_mean_plot.png")
        try:
            plt.savefig(plot_file)
            plt.close()
            print(f"Mean plot for '{sentiment}' saved to '{plot_file}'.")
            logging.info(f"Mean plot for '{sentiment}' saved to '{plot_file}'.")
        except Exception as e:
            print(f"Error saving plot for '{sentiment}': {e}")
            logging.error(f"Error saving plot for '{sentiment}': {e}")

def compile_statistics_to_excel(stats_df, plots_dir, output_excel_file, sentiment_categories):
    """
    Compiles the statistics and plots into an Excel file with multiple sheets.

    Args:
        stats_df (pd.DataFrame): DataFrame containing mean and median statistics.
        plots_dir (str): Directory containing the plot images.
        output_excel_file (str): Path to the output Excel file.
        sentiment_categories (list): List of sentiment/emotion categories.

    Returns:
        None
    """
    with pd.ExcelWriter(output_excel_file, engine='openpyxl') as writer:
        # Write statistics to the first sheet
        stats_df.to_excel(writer, sheet_name='Mean_Median_Statistics', index=False)

        workbook = writer.book

        # Insert median plots into separate sheets
        for sentiment in sentiment_categories:
            median_plot_file = os.path.join(plots_dir, f"{sentiment}_median_plot.png")
            if os.path.exists(median_plot_file):
                # Create a new sheet for each median plot
                sheet_name = f"{sentiment.capitalize()} Median Plot"
                # Limit sheet name to 31 characters
                sheet_name = sheet_name[:31]
                if sheet_name in workbook.sheetnames:
                    # Avoid duplicate sheet names
                    sheet_name = f"{sentiment.capitalize()} Median Plot_1"
                worksheet = workbook.create_sheet(title=sheet_name)

                # Add image to the sheet
                try:
                    img = ExcelImage(median_plot_file)
                    img.anchor = 'A1'
                    worksheet.add_image(img)
                except Exception as e:
                    logging.error(f"Error embedding image '{median_plot_file}' into Excel: {e}")
            else:
                logging.warning(f"Median plot file '{median_plot_file}' not found. Skipping embedding in Excel.")

        # Insert mean plots into separate sheets
        for sentiment in sentiment_categories:
            mean_plot_file = os.path.join(plots_dir, f"{sentiment}_mean_plot.png")
            if os.path.exists(mean_plot_file):
                # Create a new sheet for each mean plot
                sheet_name = f"{sentiment.capitalize()} Mean Plot"
                # Limit sheet name to 31 characters
                sheet_name = sheet_name[:31]
                if sheet_name in workbook.sheetnames:
                    # Avoid duplicate sheet names
                    sheet_name = f"{sentiment.capitalize()} Mean Plot_1"
                worksheet = workbook.create_sheet(title=sheet_name)

                # Add image to the sheet
                try:
                    img = ExcelImage(mean_plot_file)
                    img.anchor = 'A1'
                    worksheet.add_image(img)
                except Exception as e:
                    logging.error(f"Error embedding image '{mean_plot_file}' into Excel: {e}")
            else:
                logging.warning(f"Mean plot file '{mean_plot_file}' not found. Skipping embedding in Excel.")

        # Save the Excel file
        writer.save()
    print(f"All statistics and plots have been compiled into '{output_excel_file}'.")
    logging.info(f"All statistics and plots have been compiled into '{output_excel_file}'.")

def setup_logging(log_file='analysis.log'):
    """
    Configures the logging settings.

    Args:
        log_file (str): Path to the log file.

    Returns:
        None
    """
    logging.basicConfig(
        filename=log_file,
        level=logging.INFO,
        format='%(asctime)s - %(levelname)s - %(message)s'
    )
    logging.info("Logging initialized.")

# ------------------------------ #
#             Main               #
# ------------------------------ #

def main():
    # Setup logging
    setup_logging()

    # Load data
    print("Loading data from JSONL file...")
    try:
        df = load_jsonl(INPUT_JSONL_FILE)
        print(f"Total articles loaded: {len(df)}")
        logging.info(f"Total articles loaded: {len(df)}")
    except Exception as e:
        print(f"Error loading data: {e}")
        logging.error(f"Error loading data: {e}")
        return

    if df.empty:
        print("The input file is empty. Exiting.")
        logging.info("The input file is empty. Exiting.")
        return

    # Map media outlets to categories and extract media outlet names
    print("Mapping media outlets to categories and extracting media outlet names...")
    try:
        df = map_media_outlet_to_category(df, MEDIA_CATEGORIES)
        print("Media outlets mapped to categories and names extracted successfully.\n")
        logging.info("Media outlets mapped to categories and names extracted successfully.")
    except Exception as e:
        print(f"Error mapping media outlets to categories: {e}")
        logging.error(f"Error mapping media outlets to categories: {e}")
        return

    # Define media categories and sentiment/emotion categories
    media_categories = MEDIA_CATEGORIES  # Dictionary
    sentiment_categories = CATEGORIES.copy()  # Including 'fear'

    # Add +4 to all sentiment/emotion scores
    print("Adding +4 to all sentiment/emotion scores...")
    try:
        df = add_four_to_scores(df, sentiment_categories)
        print("+4 added to all sentiment/emotion scores successfully.\n")
        logging.info("+4 added to all sentiment/emotion scores successfully.")
    except Exception as e:
        print(f"Error adding +4 to sentiment/emotion scores: {e}")
        logging.error(f"Error adding +4 to sentiment/emotion scores: {e}")
        return

    # Create combined "Category: Media Outlet" columns
    print("Creating combined 'Category: Media Outlet' columns...")
    try:
        df = create_combined_columns(df, media_categories, sentiment_categories)
        print("Combined columns created successfully.\n")
        logging.info("Combined columns created successfully.")
    except Exception as e:
        print(f"Error creating combined columns: {e}")
        logging.error(f"Error creating combined columns: {e}")
        return

    # Gather all combined column names
    combined_columns = []
    for category, outlets in media_categories.items():
        for outlet in outlets:
            combined_columns.append(f"{category}: {outlet}")

    # Calculate mean and median
    print("Calculating mean and median for each sentiment/emotion category per combined media category and outlet...")
    try:
        stats_df = calculate_mean_median(df, combined_columns, sentiment_categories)
        print("Mean and median calculations completed successfully.\n")
        logging.info("Mean and median calculations completed successfully.")
    except Exception as e:
        print(f"Error calculating statistics: {e}")
        logging.error(f"Error calculating statistics: {e}")
        return

    # Save raw scores to CSVs
    print("Saving raw scores to CSV files...")
    try:
        save_raw_scores_to_csv(df, combined_columns, sentiment_categories, CSV_OUTPUT_DIR)
        print("Raw scores saved to CSV files successfully.\n")
        logging.info("Raw scores saved to CSV files successfully.")
    except Exception as e:
        print(f"Error saving raw scores to CSV: {e}")
        logging.error(f"Error saving raw scores to CSV: {e}")
        return

    # Plot median scores
    print("Generating median score plots...")
    try:
        plot_median_scores(stats_df, sentiment_categories, OUTPUT_DIR)
        print("Median score plots generated successfully.\n")
        logging.info("Median score plots generated successfully.")
    except Exception as e:
        print(f"Error generating median plots: {e}")
        logging.error(f"Error generating median plots: {e}")
        return

    # Plot mean scores
    print("Generating mean score plots...")
    try:
        plot_mean_scores(stats_df, sentiment_categories, OUTPUT_DIR)
        print("Mean score plots generated successfully.\n")
        logging.info("Mean score plots generated successfully.")
    except Exception as e:
        print(f"Error generating mean plots: {e}")
        logging.error(f"Error generating mean plots: {e}")
        return

    # Compile statistics and plots into Excel
    print("Compiling statistics and plots into Excel file...")
    try:
        compile_statistics_to_excel(stats_df, OUTPUT_DIR, OUTPUT_EXCEL_FILE, sentiment_categories)
        print("Statistics and plots compiled into Excel file successfully.\n")
        logging.info("Statistics and plots compiled into Excel file successfully.")
    except Exception as e:
        print(f"Error compiling statistics and plots into Excel: {e}")
        logging.error(f"Error compiling statistics and plots into Excel: {e}")
        return

    print("\nAnalysis completed successfully.")
    logging.info("Analysis completed successfully.")

if __name__ == "__main__":
    main()
