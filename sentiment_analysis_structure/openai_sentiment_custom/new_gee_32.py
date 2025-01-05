import json
import pandas as pd
import matplotlib.pyplot as plt
import seaborn as sns
import os
import logging
import sys
from tqdm import tqdm
from openpyxl.drawing.image import Image as ExcelImage
import re
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import numpy as np
import patsy

import statsmodels.api as sm
from statsmodels.genmod.generalized_estimating_equations import GEE
from statsmodels.genmod.families import Poisson
from statsmodels.genmod.cov_struct import Independence, Exchangeable, Unstructured, Autoregressive
from statsmodels.stats.multitest import multipletests
from scipy.stats import norm, pearsonr

# ------------------------------ #
#          Configuration         #
# ------------------------------ #
INPUT_JSONL_FILE = 'processed_all_articles_with_fulltext_sentiment_analysis.jsonl'
OUTPUT_DIR = 'graphs_analysis'
os.makedirs(OUTPUT_DIR, exist_ok=True)

CSV_OUTPUT_DIR = 'csv_raw_scores'
os.makedirs(CSV_OUTPUT_DIR, exist_ok=True)

OUTPUT_EXCEL_MAIN = 'analysis_main.xlsx'
OUTPUT_EXCEL_RAW = 'analysis_raw.xlsx'
OUTPUT_EXCEL_LMM = 'analysis_gee.xlsx'
OUTPUT_EXCEL_PLOTS = 'analysis_plots.xlsx'
OUTPUT_EXCEL_COMBINED = 'analysis_combined.xlsx'
OUTPUT_EXCEL_QIC = 'analysis_gee_qic.xlsx'  # new file to store QIC results

CATEGORIES = [
    'joy', 'sadness', 'anger', 'fear',
    'surprise', 'disgust', 'trust', 'anticipation',
    'negative_sentiment', 'positive_sentiment'
]

MEDIA_CATEGORIES = {
    'Scientific': ['nature', 'sciam', 'stat', 'newscientist'],
    'Left': ['theatlantic', 'the daily beast', 'the intercept', 'mother jones', 'msnbc', 'slate', 'vox', 'huffpost'],
    'Lean Left': ['ap', 'axios', 'cnn', 'guardian', 'business insider', 'nbcnews', 'npr', 'nytimes', 'politico', 'propublica', 'wapo', 'usa today'],
    'Center': ['reuters', 'marketwatch', 'financial times', 'newsweek', 'forbes'],
    'Lean Right': ['thedispatch', 'epochtimes', 'foxbusiness', 'wsj', 'national review', 'washtimes'],
    'Right': ['breitbart', 'theblaze', 'daily mail', 'dailywire', 'foxnews', 'nypost', 'newsmax'],
}

def setup_logging(log_file='analysis.log'):
    # Suppress excessive font-manager debug logs
    logging.getLogger('matplotlib.font_manager').setLevel(logging.WARNING)

    logging.basicConfig(
        filename=log_file,
        level=logging.DEBUG,
        format='%(asctime)s [%(levelname)s] %(module)s - %(message)s'
    )
    # Also output to console
    console = logging.StreamHandler(sys.stdout)
    console.setLevel(logging.DEBUG)
    formatter = logging.Formatter('[%(levelname)s] %(module)s: %(message)s')
    console.setFormatter(formatter)
    logging.getLogger('').addHandler(console)

    logging.info("Logging initialized. File + console output.")


def load_jsonl(jsonl_file):
    records = []
    with open(jsonl_file, 'r', encoding='utf-8') as f:
        for line in tqdm(f, desc="Loading JSONL data"):
            try:
                record = json.loads(line)
                records.append(record)
            except json.JSONDecodeError as e:
                logging.error(f"JSON decode error: {e}")
    df = pd.DataFrame(records)
    return df

def map_media_outlet_to_category(df, media_categories):
    outlet_to_category = {}
    for category, outlets in media_categories.items():
        for outlet in outlets:
            outlet_to_category[outlet.lower().strip()] = category

    if 'media_outlet' not in df.columns:
        logging.error("'media_outlet' column not found in DataFrame.")
        raise KeyError("'media_outlet' column not found in DataFrame.")

    df['media_outlet_clean'] = df['media_outlet'].str.lower().str.strip()
    df['media_category'] = df['media_outlet_clean'].map(outlet_to_category).fillna('Other')
    unmapped = df.loc[df['media_category']=='Other','media_outlet'].unique()
    if len(unmapped)>0:
        logging.warning(f"Unmapped outlets mapped to 'Other': {unmapped}")
        print(f"Warning: These media outlets were not recognized: {unmapped}")
    return df

def aggregate_sentiment_scores(df, sentiment_categories):
    # Same aggregator as before
    aggregation = []
    for media_cat in MEDIA_CATEGORIES.keys():
        sub = df[df['media_category']==media_cat]
        for sentiment in sentiment_categories:
            pattern = f"^{re.escape(sentiment)}_\\d+$"
            matched_cols = [c for c in df.columns if re.match(pattern,c)]
            if matched_cols:
                qsum = sub[matched_cols].clip(lower=0).sum(skipna=True).sum()
                qcount = sub[matched_cols].clip(lower=0).count().sum()
            else:
                qsum,qcount=0,0

            fcol = f"{sentiment}_fulltext"
            if fcol in df.columns:
                fsum = sub[fcol].clip(lower=0).sum(skipna=True)
                fcount = sub[fcol].clip(lower=0).count()
            else:
                fsum,fcount=0,0

            aggregation.append({
                'Media Category':media_cat,
                'Sentiment/Emotion':sentiment,
                'Quotation_Sum':qsum,
                'Quotation_Count':qcount,
                'Fulltext_Sum':fsum,
                'Fulltext_Count':fcount
            })
    return pd.DataFrame(aggregation)

def calculate_averages(agg):
    agg['Quotation_Average']=agg.apply(
        lambda r: r['Quotation_Sum']/r['Quotation_Count'] if r['Quotation_Count']>0 else None, axis=1
    )
    agg['Fulltext_Average']=agg.apply(
        lambda r: r['Fulltext_Sum']/r['Fulltext_Count'] if r['Fulltext_Count']>0 else None, axis=1
    )
    return agg

def calculate_mean_median(agg):
    rows=[]
    for s in CATEGORIES:
        sub = agg.loc[agg['Sentiment/Emotion']==s]
        qa=sub['Quotation_Average'].dropna()
        fa=sub['Fulltext_Average'].dropna()
        rows.append({
            'Sentiment/Emotion': s,
            'Mean_Quotation_Average': qa.mean() if len(qa)>0 else None,
            'Median_Quotation_Average': qa.median() if len(qa)>0 else None,
            'Mean_Fulltext_Average': fa.mean() if len(fa)>0 else None,
            'Median_Fulltext_Average': fa.median() if len(fa)>0 else None
        })
    return pd.DataFrame(rows)

def save_aggregated_scores_to_csv(aggdf, outdir, prefix='aggregated_sentiment_emotion_scores.csv'):
    csvp = os.path.join(outdir, prefix)
    try:
        aggdf.to_csv(csvp,index=False)
        print(f"Aggregated sentiment/emotion scores saved to '{csvp}'.")
        logging.info(f"Aggregated sentiment/emotion scores saved to '{csvp}'.")
    except Exception as e:
        logging.error(f"Error saving aggregated: {e}")


def plot_statistics(aggdf, outdir):
    sns.set_style("whitegrid")
    for s in CATEGORIES:
        sub = aggdf.loc[aggdf['Sentiment/Emotion']==s].copy()

        # Quotation
        plt.figure(figsize=(10,6))
        sns.barplot(x='Media Category',y='Quotation_Average',data=sub,color='steelblue')
        plt.title(f"Mean Quotation-Based '{s.capitalize()}' Scores")
        plt.xlabel('Media Category')
        plt.ylabel('Mean Quotation-Based Average Score')
        plt.xticks(rotation=45,ha='right')
        plt.tight_layout()
        out_quote = os.path.join(outdir,f"quote_{s}.png")
        try:
            plt.savefig(out_quote)
            plt.close()
            print(f"Quotation-Based '{s}' scores plot saved to '{out_quote}'.")
        except Exception as e:
            logging.error(f"Error saving Q plot for {s}: {e}")

        # Fulltext
        plt.figure(figsize=(10,6))
        sns.barplot(x='Media Category', y='Fulltext_Average',data=sub,color='darkorange')
        plt.title(f"Mean Fulltext-Based '{s.capitalize()}' Scores")
        plt.xlabel('Media Category')
        plt.ylabel('Mean Fulltext-Based Average Score')
        plt.xticks(rotation=45,ha='right')
        plt.tight_layout()
        out_full = os.path.join(outdir,f"fulltext_{s}.png")
        try:
            plt.savefig(out_full)
            plt.close()
            print(f"Fulltext-Based '{s}' scores plot saved to '{out_full}'.")
        except Exception as e:
            logging.error(f"Error saving F plot for {s}: {e}")


# --------------------------------------------------
#    GEE Fitting with Multiple Structures + QIC
# --------------------------------------------------
def fit_gee_model(df, formula, groups, cov_struct, family=Poisson(), time=None):
    """Helper function that attempts to fit a GEE with the given cov_struct, returns results or None."""
    model = GEE.from_formula(formula, groups=groups, data=df, cov_struct=cov_struct, family=family, time=time)
    try:
        res = model.fit()
    except Exception as e:
        logging.error(f"GEE fit failed: {e}")
        return None
    return res

def run_gee_and_qic(df, sentiment, measure_type, structures, min_date=None):
    """
    Attempts GEE with each correlation structure in 'structures', computes QIC, 
    returns a dictionary of results keyed by structure name, or None if no success.
    If measure_type='Quotation', we average over columns matching sentiment_\d+.
    If measure_type='Fulltext', we use sentiment_fulltext column.

    If 'date' is present and structure=Autoregressive, we define a time index per outlet for AR(1).
    """
    if measure_type=='Quotation':
        pattern = f"^{re.escape(sentiment)}_\\d+$"
        matched_cols = [c for c in df.columns if re.match(pattern,c)]
        if not matched_cols:
            return {}
        temp = df.copy()
        col_mean = f"{sentiment}_quotation_mean"
        temp[col_mean] = temp[matched_cols].clip(lower=0).mean(axis=1)
        score_col = col_mean
    else:
        fcol = f"{sentiment}_fulltext"
        if fcol not in df.columns:
            return {}
        temp = df.copy()
        clipped = f"{sentiment}_fulltext_clipped"
        temp[clipped] = temp[fcol].clip(lower=0)
        score_col = clipped

    mdf = temp.dropna(subset=[score_col, 'media_category', 'media_outlet']).copy()
    # Must have at least 2 categories to compare
    if mdf['media_category'].nunique()<2:
        return {}

    # If date is present, we might do AR(1)
    time_col = None
    if ('date' in mdf.columns) and (mdf['date'].notna().sum()>0) and min_date is not None:
        # convert date -> int day index
        mdf['int_date'] = (mdf['date'] - min_date).dt.days + 1
        # might use it if structure=Autoregressive

    mdf['media_category'] = mdf['media_category'].astype('category')

    formula = f"{score_col} ~ media_category"
    groups = "media_outlet"
    result_dict = {}
    for struct_name, struct_obj in structures.items():
        # If AR(1), we need to define time if date is present
        if isinstance(struct_obj, Autoregressive):
            if 'int_date' not in mdf.columns:
                # No date => skip AR(1)
                continue
            # Sort by outlet, int_date
            mdf = mdf.sort_values(['media_outlet','int_date'])
            # AR(1) time array => we pass time=mdf['int_date']
            res = fit_gee_model(mdf, formula, groups=groups, cov_struct=struct_obj, time=mdf['int_date'].values)
        else:
            # No time param needed
            res = fit_gee_model(mdf, formula, groups=groups, cov_struct=struct_obj)
        if res is None:
            continue
        # get QIC
        try:
            qic_val = res.qic()
        except:
            qic_val = np.nan
        result_dict[struct_name] = (res, qic_val)
    return result_dict


def fit_and_pick_best_structure(df, sentiment, measure_type, min_date=None):
    """
    Fits multiple structures => picks best by QIC => returns best results + summary of all.
    """
    # We define the correlation structures we want to test
    structures = {
        "Independence": Independence(),
        "Exchangeable": Exchangeable(),
        "Unstructured": Unstructured(),
        "AR1": Autoregressive()
    }

    # We'll run GEE for each structure, store QIC, pick best
    results_map = run_gee_and_qic(df, sentiment, measure_type, structures, min_date)
    if not results_map:
        return None

    # Now pick best by QIC
    best_struct = None
    best_qic = np.inf
    all_info = []
    for struct_name, (res_obj, qic_val) in results_map.items():
        all_info.append((struct_name, qic_val))
        if (qic_val is not None) and (qic_val < best_qic):
            best_qic = qic_val
            best_struct = (struct_name, res_obj)

    if best_struct is None:
        return None

    # We do pairwise comparisons with the best structure
    chosen_struct, best_res = best_struct
    summary_text = best_res.summary().as_text()

    # Build pairwise
    params = best_res.params
    cov = best_res.cov_params()
    model_df = best_res.model.data.frame
    cat_var = model_df['media_category'].astype('category')
    all_categories = cat_var.cat.categories
    reference_category = all_categories[0]

    param_names = best_res.model.exog_names
    cat_to_param_index = {reference_category: 0}
    for cat in all_categories[1:]:
        pname = f"media_category[T.{cat}]"
        cat_to_param_index[cat] = param_names.index(pname)

    pairwise_list=[]
    cats=list(all_categories)
    for i in range(len(cats)):
        for j in range(i+1,len(cats)):
            catA, catB = cats[i], cats[j]
            cvec = np.zeros(len(params))
            if catA==reference_category and catB!=reference_category:
                cvec[cat_to_param_index[catB]]=-1.
            elif catB==reference_category and catA!=reference_category:
                cvec[cat_to_param_index[catA]]=1.
            else:
                cvec[cat_to_param_index[catA]]=1.
                cvec[cat_to_param_index[catB]]=-1.
            diff_est = cvec@params
            diff_var = cvec@cov@cvec
            diff_se = np.sqrt(diff_var)
            z = diff_est/diff_se if diff_se>0 else np.inf
            p = 2*(1-norm.cdf(abs(z)))
            pairwise_list.append((catA,catB,diff_est,diff_se,z,p))

    pairwise_df= pd.DataFrame(pairwise_list, columns=["CategoryA","CategoryB","Difference","SE","Z","p_value"])
    reject, p_adj, _, _ = multipletests(pairwise_df['p_value'], method='holm')
    pairwise_df['p_value_adj']=p_adj
    pairwise_df['reject_H0']=reject

    # print to console which structure is best
    print(f"    => Best structure for sentiment='{sentiment}', measure='{measure_type}' is: {chosen_struct} (QIC={best_qic:.2f})")
    # Also show the other QIC's
    for (n,q) in all_info:
        print(f"       {n} => QIC={q}")

    return {
        'BestStructure': chosen_struct,
        'QIC_All': all_info, 
        'GEE_Summary': summary_text,
        'Pairwise': pairwise_df
    }


def run_gee_analyses(df):
    """
    We'll attempt multiple correlation structures for each sentiment & measure,
    pick the best by QIC, store results.
    If date is present, we allow AR(1). Otherwise, AR(1) is effectively skipped.
    """
    # If date is present, we define min_date
    if 'date' in df.columns and df['date'].notna().sum()>0:
        min_date = df['date'].min()
    else:
        min_date = None

    all_results={}
    for sentiment in CATEGORIES:
        all_results[sentiment]={}
        for measure_type in ['Quotation','Fulltext']:
            best_result= fit_and_pick_best_structure(df, sentiment, measure_type, min_date)
            if best_result is not None:
                all_results[sentiment][measure_type]= best_result
    return all_results


def compile_results_into_multiple_workbooks(aggregated_df, stats_df, raw_df, gee_results, plots_dir,
                                            main_excel, raw_excel, lmm_excel, plots_excel, combined_excel, qic_excel=None):
    """
    Modified to store QIC results into a new workbook if we want. 
    But here we'll simply store them within lmm_excel or combined if you prefer.
    """
    # main
    with pd.ExcelWriter(main_excel, engine='openpyxl') as w:
        aggregated_df.to_excel(w, sheet_name='Aggregated_Scores', index=False)
        stats_df.to_excel(w, sheet_name='Mean_Median_Statistics', index=False)

    # raw
    raw_df2= raw_df.copy()
    raw_df2['media_category'] = raw_df2['media_category'].astype('category')
    raw_df2.sort_values(['media_category','media_outlet'], inplace=True)

    with pd.ExcelWriter(raw_excel, engine='openpyxl') as w:
        raw_df2.to_excel(w, sheet_name='Raw_Data',index=False)
        for s in CATEGORIES:
            scols = [c for c in raw_df2.columns if c.startswith(s+'_')]
            partdf = raw_df2[['media_category','media_outlet']+scols]
            sheetnm=f"Raw_{s[:29]}"
            partdf.to_excel(w, sheet_name=sheetnm, index=False)

    # lmm (the GEE results)
    with pd.ExcelWriter(lmm_excel, engine='openpyxl') as w:
        summary_index=[]
        for sentiment in gee_results:
            for measure_type in gee_results[sentiment]:
                sheetnm = f"GEE_{sentiment[:20]}_{measure_type[:8]}"
                best_struct= gee_results[sentiment][measure_type]['BestStructure']
                qic_all=gee_results[sentiment][measure_type]['QIC_All']
                summary_txt= gee_results[sentiment][measure_type]['GEE_Summary']
                pairwise_df= gee_results[sentiment][measure_type]['Pairwise']

                # Store summary
                sumdf= pd.DataFrame({'GEE_Summary': summary_txt.split('\n')})
                sumdf.to_excel(w, sheet_name=sheetnm, index=False)
                # Possibly store best structure & QIC info just below
                row_cur = len(sumdf)+2
                best_info_df = pd.DataFrame({
                    'Info':[
                        f"BestStructure={best_struct}",
                        f"AllStructures={qic_all}"
                    ]
                })
                best_info_df.to_excel(w, sheet_name=sheetnm, index=False, startrow=row_cur)
                row_cur += (len(best_info_df)+2)

                pairwise_df.to_excel(w, sheet_name=sheetnm, index=False, startrow=row_cur)

                summary_index.append({
                    'Sentiment': sentiment,
                    'Measure': measure_type,
                    'BestStructure': best_struct,
                    'SheetName': sheetnm
                })
        idxdf= pd.DataFrame(summary_index)
        idxdf.to_excel(w, sheet_name='GEE_Results_Index', index=False)

    # If you want a separate QIC, you can store a condensed version there, or skip

    # plots
    wb=Workbook()
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])

    for s in CATEGORIES:
        qplot=os.path.join(plots_dir,f"quote_{s}.png")
        if os.path.exists(qplot):
            stitle= f"Quote_{s[:28]}"
            ws=wb.create_sheet(title=stitle)
            try:
                img= ExcelImage(qplot)
                img.anchor='A1'
                ws.add_image(img)
            except Exception as e:
                logging.error(f"Error embedding {qplot}: {e}")

        fplot=os.path.join(plots_dir,f"fulltext_{s}.png")
        if os.path.exists(fplot):
            stitle2= f"Fulltext_{s[:25]}"
            ws2= wb.create_sheet(title=stitle2)
            try:
                im2= ExcelImage(fplot)
                im2.anchor='A1'
                ws2.add_image(im2)
            except Exception as e:
                logging.error(f"Error embedding {fplot}: {e}")

    wb.save(plots_excel)

    # combined
    raw_df_clean= raw_df2.copy()
    # convert any list columns
    for col in raw_df_clean.columns:
        raw_df_clean[col] = raw_df_clean[col].apply(lambda x: ", ".join(x) if isinstance(x,list) else x)

    wb_comb= Workbook()
    if 'Sheet' in wb_comb.sheetnames:
        wb_comb.remove(wb_comb['Sheet'])

    ws_agg= wb_comb.create_sheet("Aggregated_Scores")
    for r in dataframe_to_rows(aggregated_df, index=False, header=True):
        ws_agg.append(r)

    ws_stats= wb_comb.create_sheet("Mean_Median_Statistics")
    for r in dataframe_to_rows(stats_df, index=False, header=True):
        ws_stats.append(r)

    ws_raw= wb_comb.create_sheet("Raw_Data")
    for r in dataframe_to_rows(raw_df_clean, index=False, header=True):
        ws_raw.append(r)

    ws_gee= wb_comb.create_sheet("GEE_Summaries")
    ws_gee.append(["Sentiment","Measure","Best_Structure","GEE_Summary(Truncated)"])
    for sentiment in gee_results:
        for measure_type in gee_results[sentiment]:
            best_struct= gee_results[sentiment][measure_type]['BestStructure']
            sum_lines= gee_results[sentiment][measure_type]['GEE_Summary'].split('\n')
            truncated= "\n".join(sum_lines[:6])  # just a snippet
            ws_gee.append([sentiment, measure_type, best_struct, truncated])

    wb_comb.save(combined_excel)


# --------------------------------------------------
#                       main
# --------------------------------------------------
def main():
    setup_logging()
    print("Single run with multi-structure GEE + QIC comparison + Poisson + clustering by media_outlet.")
    df = load_jsonl(INPUT_JSONL_FILE)
    print(f"Total articles loaded: {len(df)}")

    df = map_media_outlet_to_category(df, MEDIA_CATEGORIES)

    # Split raw data
    chunk_size=20000
    for i in range(0, len(df), chunk_size):
        chunk=df.iloc[i:i+chunk_size]
        cfile=os.path.join(CSV_OUTPUT_DIR,f"raw_data_part_{i//chunk_size+1}.csv")
        chunk.to_csv(cfile,index=False)
        print(f"Saved chunk {i//chunk_size+1} to {cfile}")

    print("\nSummary Statistics:")
    print("Total number of articles:", len(df))
    print("\nNumber of articles per outlet:")
    print(df['media_outlet_clean'].value_counts())
    print("\nNumber of articles per category:")
    print(df['media_category'].value_counts())
    print()

    # Aggregation
    print("Aggregating sentiment/emotion scores per media category...")
    aggregated_df = aggregate_sentiment_scores(df, CATEGORIES)
    aggregated_df = calculate_averages(aggregated_df)
    stats_df = calculate_mean_median(aggregated_df)

    save_aggregated_scores_to_csv(aggregated_df, CSV_OUTPUT_DIR)
    plot_statistics(aggregated_df, OUTPUT_DIR)

    print("Fitting GEE models with multiple correlation structures (QIC) => best structure selection.")
    gee_results = run_gee_analyses(df)
    # If you want to see the final chosen structures in the console
    print("\n** GEE Model Results + Sensitivity Analysis **")
    for s in CATEGORIES:
        if s in gee_results:
            for measure in gee_results[s]:
                best_struct= gee_results[s][measure]['BestStructure']
                print(f"For sentiment='{s}', measure='{measure}', best structure => {best_struct}")

    compile_results_into_multiple_workbooks(
        aggregated_df, stats_df, df, gee_results, OUTPUT_DIR,
        OUTPUT_EXCEL_MAIN, OUTPUT_EXCEL_RAW, OUTPUT_EXCEL_LMM,
        OUTPUT_EXCEL_PLOTS, OUTPUT_EXCEL_COMBINED
    )

    print("Analysis completed successfully. Check the console logs for QIC-based structure comparisons.")


if __name__=="__main__":
    main()
