#!/usr/bin/env python3
# dual_output_50tabs_and_20tabs.py
"""
Creates two Excel files:

1) raw_values_6cat_50tabs_no_blanks.xlsx
   => 50 tabs total (10 sentiments × 5 measures).
      Each tab => no blank rows in each of the 6 columns.
      Measures:
        - Fulltext
        - Quotation (mean across matching quote columns)
        - Fulltext_Intensity
        - Quotation_Intensity (mean across matching quote-intensity columns)
        - Title_Intensity

2) raw_values_6cat_20tabs_intv2_no_blanks.xlsx
   => 20 tabs total (2 for each sentiment).
      For each sentiment S:
        A) "S_3Int" => 18 columns = 6 categories × 3 intensities
           [Fulltext_Intensity, Quotation_Intensity, Title_Intensity]
        B) "S_2NonInt" => 12 columns = 6 categories × 2 measures
           [Fulltext, Quotation]
      Each column is a vertical list of measure values, no blank rows inserted.

Categories:
   Scientific, Left, Lean Left, Center, Lean Right, Right

Sentiments:
   joy, sadness, anger, fear, surprise, disgust, trust, anticipation,
   negative_sentiment, positive_sentiment
"""

import json
import re
import numpy as np
import pandas as pd
from tqdm import tqdm
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# ------------------ GLOBAL CONFIG ------------------
INPUT_JSONL_FILE = "processed_all_articles_fixed_4.jsonl"

OUTPUT_50TABS = "raw_values_6cat_50tabs_no_blanks.xlsx"
OUTPUT_20TABS = "raw_values_6cat_20tabs_intv2_no_blanks.xlsx"

# 6 categories
MEDIA_CATEGORIES = {
    "Scientific": ["nature","sciam","stat","newscientist"],
    "Left": ["theatlantic","the daily beast","the intercept","mother jones","msnbc","slate","vox","huffpost"],
    "Lean Left": ["ap","axios","cnn","guardian","business insider","nbcnews",
                  "npr","nytimes","politico","propublica","wapo","usa today"],
    "Center": ["reuters","marketwatch","financial times","newsweek","forbes"],
    "Lean Right": ["thedispatch","epochtimes","foxbusiness","wsj","national review","washtimes"],
    "Right": ["breitbart","theblaze","daily mail","dailywire","foxnews","nypost","newsmax"],
}

SENTIMENTS = [
    "joy","sadness","anger","fear",
    "surprise","disgust","trust","anticipation",
    "negative_sentiment","positive_sentiment"
]

MEASURES_5 = [  # used for the 50-tab file
    "Fulltext",
    "Quotation",
    "Fulltext_Intensity",
    "Quotation_Intensity",
    "Title_Intensity",
]

# We'll define "3 intensities" and "2 non-intensities" for the 20-tab file
THREE_INTENSITIES = ["Fulltext_Intensity", "Quotation_Intensity", "Title_Intensity"]
TWO_NONINTENSITIES= ["Fulltext", "Quotation"]


# ------------------ 1) LOAD + MAP ------------------
def load_jsonl(jsonl_file):
    records=[]
    with open(jsonl_file, "r", encoding="utf-8") as f:
        for line in tqdm(f, desc="Loading JSONL"):
            rec=json.loads(line)
            records.append(rec)
    return pd.DataFrame(records)

def map_outlet_to_6cats(df):
    cat_map={}
    for cat, outls in MEDIA_CATEGORIES.items():
        for o in outls:
            cat_map[o.lower().strip()] = cat
    df2=df.copy()
    df2["media_outlet_clean"] = df2["media_outlet"].str.lower().str.strip()
    df2["media_category_6"]   = df2["media_outlet_clean"].map(cat_map).fillna("Other")
    return df2

# ------------------ 2) EXTRACT MEASURE VALUE ------------------
def extract_measure_value(row, sentiment, measure):
    """
    - Quotation => average columns ^{sentiment}_\d+$
    - Quotation_Intensity => average columns ^{sentiment}_\d+_intensity$
    - Fulltext => row[f"{sentiment}_fulltext"]
    - Fulltext_Intensity => row[f"{sentiment}_fulltext_intensity"]
    - Title_Intensity => row[f"{sentiment}_title_intensity"]
    Returns float or np.nan
    """
    try:
        if measure=="Quotation":
            pat = rf"^{re.escape(sentiment)}_\d+$"
            matched = [c for c in row.index if re.match(pat,c)]
            if not matched:
                return np.nan
            vals = row[matched].astype(float).clip(lower=0)
            return vals.mean(skipna=True)

        elif measure=="Quotation_Intensity":
            pat = rf"^{re.escape(sentiment)}_\d+_intensity$"
            matched = [c for c in row.index if re.match(pat,c)]
            if not matched:
                return np.nan
            vals = row[matched].astype(float).clip(lower=0)
            return vals.mean(skipna=True)

        elif measure=="Fulltext":
            col = f"{sentiment}_fulltext"
            if col not in row.index:
                return np.nan
            return float(row[col])

        elif measure=="Fulltext_Intensity":
            col = f"{sentiment}_fulltext_intensity"
            if col not in row.index:
                return np.nan
            return float(row[col])

        elif measure=="Title_Intensity":
            col = f"{sentiment}_title_intensity"
            if col not in row.index:
                return np.nan
            return float(row[col])

        else:
            return np.nan
    except:
        return np.nan


# ------------------ UTILITY: Build "no-blank-row" columns for each cat X measure -------------
def build_no_blank_cols_for_measure(df, sentiment, measure_list):
    """
    Returns a dict of lists, where each key = f"{cat}_{measure}" and we store consecutive values.
    e.g. "Scientific_Fulltext_Intensity", "Scientific_Quotation_Intensity", "Scientific_Title_Intensity"
    We'll skip rows if measure is np.nan. => no blank row in that column.
    """
    # We'll define col_names = [f"{cat}_{m}" for cat in 6 cats for m in measure_list]
    # Then we'll store them in a dict => col_lists = { "Scientific_Fulltext_Intensity": [] , ... }
    out_cols = []
    for cat in MEDIA_CATEGORIES.keys():  # e.g. "Scientific","Left","Lean Left", ...
        for meas in measure_list:
            col_key = f"{cat}_{meas}"
            out_cols.append(col_key)

    col_lists = {c:[] for c in out_cols}

    for i, row in df.iterrows():
        cat_6 = row["media_category_6"]
        if cat_6 not in MEDIA_CATEGORIES.keys():
            continue  # skip if "Other"
        # for each measure in measure_list => extract value
        for meas in measure_list:
            measure_val = extract_measure_value(row, sentiment, meas)
            if pd.notna(measure_val):
                col_key = f"{cat_6}_{meas}"
                # append measure_val
                col_lists[col_key].append(measure_val)
            # else => skip => no blank

    # Now each col_lists[key] is a consecutive list of values => we must find max_len
    max_len = max(len(v) for v in col_lists.values())
    # We'll build a DataFrame with columns in out_cols order
    data_dict = {}
    for c in out_cols:
        vals = col_lists[c]
        if len(vals)<max_len:
            vals = vals + [np.nan]*(max_len - len(vals))
        data_dict[c] = vals

    df_result = pd.DataFrame(data_dict, columns=out_cols)
    return df_result


# ------------------ PART A: 50 tabs (the original approach) ------------------
def create_50tabs_file(df, wb):
    """
    For each of 10 sentiments, for each of 5 measures => 50 tabs.
    In each tab, 6 columns => each col is just consecutive measure values for that cat.
    """
    from openpyxl.utils.dataframe import dataframe_to_rows

    # remove default if present
    if "Sheet" in wb.sheetnames and len(wb.sheetnames)==1:
        if wb.sheetnames[0]=="Sheet":
            wb.remove(wb["Sheet"])

    for sentiment in SENTIMENTS:
        for measure in MEASURES_5:
            tab_name = f"{sentiment[:10]}_{measure[:10]}"
            # We'll create 6 columns => one measure for each cat => no blank rows
            # We'll do a simpler approach => dictionary col => consecutive values
            cat_lists = {
                "Scientific": [],
                "Left": [],
                "Lean Left": [],
                "Center": [],
                "Lean Right": [],
                "Right": []
            }
            for i, row in df.iterrows():
                cat_6 = row["media_category_6"]
                if cat_6 not in cat_lists: 
                    continue
                val = extract_measure_value(row, sentiment, measure)
                if pd.notna(val):
                    cat_lists[cat_6].append(val)

            # find max_len
            max_len = max(len(lst) for lst in cat_lists.values())
            out_cols = ["Scientific","Left","Lean Left","Center","Lean Right","Right"]
            data_dict = {}
            for c in out_cols:
                col_vals = cat_lists[c]
                if len(col_vals)<max_len:
                    col_vals = col_vals + [np.nan]*(max_len - len(col_vals))
                data_dict[c] = col_vals
            result_df = pd.DataFrame(data_dict, columns=out_cols)

            ws = wb.create_sheet(title=tab_name)
            # add a header row => out_cols
            df_rows = [out_cols]
            for row_idx in range(max_len):
                row_list=[]
                for c in out_cols:
                    row_list.append(result_df.loc[row_idx,c])
                df_rows.append(row_list)

            for r_idx, row_vals in enumerate(df_rows,1):
                for c_idx, cell_val in enumerate(row_vals,1):
                    ws.cell(row=r_idx,column=c_idx,value=cell_val)


# ------------------ PART B: 20 tabs => 2 tabs/sentiment => intensities vs non-intensities -------
def create_20tabs_file(df, wb_20):
    """
    For each sentiment => 2 tabs => 
      1) 3 intensities => Fulltext_Intensity, Quotation_Intensity, Title_Intensity
         => 6 categories * 3 = 18 columns
      2) 2 non-intensities => Fulltext, Quotation
         => 6 categories * 2 = 12 columns
    No blank rows in each column.
    """
    if "Sheet" in wb_20.sheetnames and len(wb_20.sheetnames)==1:
        if wb_20.sheetnames[0]=="Sheet":
            wb_20.remove(wb_20["Sheet"])

    for sentiment in SENTIMENTS:
        # A) intensities => 3 measures
        tab_nameA = f"{sentiment[:10]}_3Int"
        df_int = build_no_blank_cols_for_measure(df, sentiment, THREE_INTENSITIES)
        wsA = wb_20.create_sheet(title=tab_nameA)
        # We'll write the DataFrame
        # first row => column names
        col_names = list(df_int.columns)
        df_rows = [col_names]
        max_len = len(df_int)
        for row_idx in range(max_len):
            row_list=[]
            for c in col_names:
                row_list.append(df_int.loc[row_idx, c])
            df_rows.append(row_list)
        for r_idx, row_vals in enumerate(df_rows,1):
            for c_idx, val in enumerate(row_vals,1):
                wsA.cell(row=r_idx, column=c_idx, value=val)

        # B) non-intensities => 2 measures
        tab_nameB = f"{sentiment[:10]}_2NonInt"
        df_non = build_no_blank_cols_for_measure(df, sentiment, TWO_NONINTENSITIES)
        wsB = wb_20.create_sheet(title=tab_nameB)
        col_names2 = list(df_non.columns)
        df_rows2 = [col_names2]
        max_len2 = len(df_non)
        for row_idx in range(max_len2):
            row_list=[]
            for c in col_names2:
                row_list.append(df_non.loc[row_idx, c])
            df_rows2.append(row_list)
        for r_idx, row_vals in enumerate(df_rows2,1):
            for c_idx, val in enumerate(row_vals,1):
                wsB.cell(row=r_idx, column=c_idx, value=val)


# ------------------ UTILITY build_no_blank_cols_for_measure -----------
def build_no_blank_cols_for_measure(df, sentiment, measure_list):
    """
    For each measure in measure_list (e.g. [Fulltext_Intensity, Quotation_Intensity, Title_Intensity]),
    and for each of the 6 categories, build one column => consecutive measure values (no blank row).
    The result => N rows = max col length, 6*N_measures columns.
    """
    # e.g. for 3 intensities => each category has 3 columns => total 18
    # We'll produce col keys like "Scientific_Fulltext_Intensity"
    cat_order = ["Scientific","Left","Lean Left","Center","Lean Right","Right"]
    out_cols=[]
    for cat in cat_order:
        for meas in measure_list:
            out_cols.append(f"{cat}_{meas}")

    col_lists = {c:[] for c in out_cols}
    for i, row in df.iterrows():
        cat_6 = row["media_category_6"]
        if cat_6 not in cat_order:
            continue
        # For each measure in measure_list => extract => if not NaN => store
        for meas in measure_list:
            val = extract_measure_value(row, sentiment, meas)
            if pd.notna(val):
                col_key=f"{cat_6}_{meas}"
                col_lists[col_key].append(val)

    # Now find max len
    max_len = max(len(v) for v in col_lists.values()) if col_lists else 0
    data_dict = {}
    for c in out_cols:
        col_vals = col_lists[c]
        if len(col_vals)<max_len:
            col_vals = col_vals + [np.nan]*(max_len - len(col_vals))
        data_dict[c] = col_vals

    df_result = pd.DataFrame(data_dict, columns=out_cols)
    return df_result


# ------------------ MAIN -----------------------------------------------------
def main():
    print("Loading + mapping data...")
    df = load_jsonl(INPUT_JSONL_FILE)
    df = map_outlet_to_6cats(df)
    if "article_id" not in df.columns:
        df["article_id"] = df.index + 1

    # 1) Create the 50-tabs file
    print("Creating 50-tab file =>", OUTPUT_50TABS)
    wb_50 = Workbook()
    create_50tabs_file(df, wb_50)
    wb_50.save(OUTPUT_50TABS)
    print(f"Wrote => {OUTPUT_50TABS}")

    # 2) Create the 20-tabs file
    print("Creating 20-tab file =>", OUTPUT_20TABS)
    wb_20 = Workbook()
    create_20tabs_file(df, wb_20)
    wb_20.save(OUTPUT_20TABS)
    print(f"Wrote => {OUTPUT_20TABS}")

    print("All done.")

if __name__=="__main__":
    main()
