#!/usr/bin/env python3
"""
Script to output raw measure values into 6 columns:
    Scientific, Left, Lean Left, Center, Lean Right, Right
... for each sentiment/emotion, for each measure:
    - Fulltext
    - Quotation (mean across all quote columns for that sentiment)
    - Fulltext_Intensity
    - Quotation_Intensity (mean across all quote-intensity columns)
    - Title_Intensity
We produce an XLSX with one sheet per measure.

Usage:
  python3 generate_raw_sixcats.py
"""

import json
import os
import re
import numpy as np
import pandas as pd
from tqdm import tqdm
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# ----------------------------------------------------------------------------------
# 1) CONFIG
# ----------------------------------------------------------------------------------
INPUT_JSONL_FILE = "processed_all_articles_fixed_4.jsonl"
OUTPUT_EXCEL = "raw_values_six_cats.xlsx"

# 6 categories
MEDIA_CATEGORIES = {
    "Scientific": ["nature","sciam","stat","newscientist"],
    "Left": ["theatlantic","the daily beast","the intercept","mother jones","msnbc","slate","vox","huffpost"],
    "Lean Left": ["ap","axios","cnn","guardian","business insider","nbcnews","npr","nytimes","politico","propublica","wapo","usa today"],
    "Center": ["reuters","marketwatch","financial times","newsweek","forbes"],
    "Lean Right": ["thedispatch","epochtimes","foxbusiness","wsj","national review","washtimes"],
    "Right": ["breitbart","theblaze","daily mail","dailywire","foxnews","nypost","newsmax"],
}

# Sentiments/emotions
SENTIMENTS = [
    "joy", "sadness", "anger", "fear",
    "surprise", "disgust", "trust", "anticipation",
    "negative_sentiment", "positive_sentiment"
]

# Measures
MEASURES = [
    "Fulltext",
    "Quotation",
    "Fulltext_Intensity",
    "Quotation_Intensity",
    "Title_Intensity",
]

# ----------------------------------------------------------------------------------
# 2) HELPER: LOAD + MAP
# ----------------------------------------------------------------------------------
def load_jsonl(jsonl_file):
    """Load records from JSONL => DataFrame"""
    records = []
    with open(jsonl_file, "r", encoding="utf-8") as f:
        for line in tqdm(f, desc="Loading JSONL"):
            rec=json.loads(line)
            records.append(rec)
    df = pd.DataFrame(records)
    return df

def map_outlet_to_6cats(df):
    """Map each media_outlet to one of the 6 categories, store in 'media_category_6'."""
    cat_map = {}
    for cat, outls in MEDIA_CATEGORIES.items():
        for o in outls:
            cat_map[o.lower().strip()] = cat

    df2 = df.copy()
    df2["media_outlet_clean"] = df2["media_outlet"].str.lower().str.strip()
    df2["media_category_6"]   = df2["media_outlet_clean"].map(cat_map).fillna("Other")
    return df2

# ----------------------------------------------------------------------------------
# 3) HELPER: EXTRACT A MEASURE's VALUE FOR A SINGLE (ARTICLE, SENTIMENT)
# ----------------------------------------------------------------------------------
def extract_measure_value(df_row, sentiment, measure):
    """
    - If measure=="Quotation": average across all columns matching r"^{sentiment}_\d+$"
    - If measure=="Quotation_Intensity": average across columns matching r"^{sentiment}_\d+_intensity$"
    - If measure=="Fulltext": get df_row[f"{sentiment}_fulltext"]
    - If measure=="Fulltext_Intensity": get df_row[f"{sentiment}_fulltext_intensity"]
    - If measure=="Title_Intensity": get df_row[f"{sentiment}_title_intensity"]
    Return float or np.nan
    """
    try:
        if measure=="Quotation":
            pat = rf"^{re.escape(sentiment)}_\d+$"
            matched_cols = [c for c in df_row.index if re.match(pat,c)]
            if not matched_cols:
                return np.nan
            vals = df_row[matched_cols].astype(float).clip(lower=0)
            return vals.mean()

        elif measure=="Quotation_Intensity":
            pat = rf"^{re.escape(sentiment)}_\d+_intensity$"
            matched_cols = [c for c in df_row.index if re.match(pat,c)]
            if not matched_cols:
                return np.nan
            vals = df_row[matched_cols].astype(float).clip(lower=0)
            return vals.mean()

        elif measure=="Fulltext":
            col = f"{sentiment}_fulltext"
            if col not in df_row.index:
                return np.nan
            return float(df_row[col])

        elif measure=="Fulltext_Intensity":
            col = f"{sentiment}_fulltext_intensity"
            if col not in df_row.index:
                return np.nan
            return float(df_row[col])

        elif measure=="Title_Intensity":
            col = f"{sentiment}_title_intensity"
            if col not in df_row.index:
                return np.nan
            return float(df_row[col])

        else:
            return np.nan

    except:
        return np.nan

# ----------------------------------------------------------------------------------
# 4) MAIN: LOAD -> MAP -> BUILD => OUTPUT
# ----------------------------------------------------------------------------------
def main():
    # 1) LOAD
    df = load_jsonl(INPUT_JSONL_FILE)

    # 2) MAP => 6 cats
    df = map_outlet_to_6cats(df)

    # We'll define a unique ArticleID (if not present):
    if "article_id" not in df.columns:
        df["article_id"] = df.index + 1  # or some other unique ID

    # For convenience: Subset columns so we can do .loc indexing
    # We'll keep everything, but note the new columns:
    #   article_id, media_category_6, plus all sentiment columns.

    # 3) We build a wide table for each measure, but a smaller approach is:
    # We'll build a single big DataFrame with columns:
    #   ["ArticleID", "Sentiment", "MediaCategory6", "MeasureValue"]
    # and from that, pivot so we have 6 columns for each category.

    # Then for each measure, we'll dump that pivot -> an Excel sheet.

    wb = Workbook()
    # remove default "Sheet"
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    # We'll do one sheet per measure
    for measure in MEASURES:
        rows_for_measure = []
        for i, row in df.iterrows():
            article_id = row["article_id"]
            cat_6      = row["media_category_6"]
            # We'll gather measure value for each sentiment:
            for sent in SENTIMENTS:
                val = extract_measure_value(row, sent, measure)
                # We'll store row => (articleID, sentiment, cat_6, val)
                rows_for_measure.append({
                    "ArticleID": article_id,
                    "Sentiment": sent,
                    "Category6": cat_6,
                    "MeasureValue": val
                })

        measure_df = pd.DataFrame(rows_for_measure)

        # Now we pivot => columns = 6 cat, index= (ArticleID, Sentiment)
        # measure_df => pivot
        pivoted = measure_df.pivot_table(
            index=["ArticleID","Sentiment"],
            columns="Category6",
            values="MeasureValue",
            aggfunc="mean"   # typically not needed, but just in case
        )

        # Ensure the 6 cat columns exist in correct order => fill missing with NaN
        col_order = ["Scientific","Left","Lean Left","Center","Lean Right","Right"]
        for c in col_order:
            if c not in pivoted.columns:
                pivoted[c] = np.nan
        pivoted = pivoted[col_order]

        # We'll flatten the index => pivoted is now multi-index
        pivoted.reset_index(drop=False, inplace=True)

        # Now pivoted columns => [ArticleID, Sentiment, Scientific, Left, Lean Left, Center, Lean Right, Right]
        # We'll store this in an Excel sheet
        ws = wb.create_sheet(title=measure[:29])  # limit sheet name length

        # Convert pivoted to row-of-rows
        pivot_rows = dataframe_to_rows(pivoted, index=False, header=True)
        for r_idx, row_vals in enumerate(pivot_rows, 1):
            for c_idx, cell_val in enumerate(row_vals, 1):
                ws.cell(row=r_idx, column=c_idx, value=cell_val)

    # Finally save
    wb.save(OUTPUT_EXCEL)
    print(f"Wrote => {OUTPUT_EXCEL}")

if __name__=="__main__":
    main()
