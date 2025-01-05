import json
import pandas as pd
import matplotlib.pyplot as plt
import seaborn as sns
import os
import logging
import sys
from tqdm import tqdm
from openpyxl.drawing.image import Image as ExcelImage
import re
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import numpy as np
import patsy

import statsmodels.api as sm
from statsmodels.genmod.generalized_estimating_equations import GEE
from statsmodels.genmod.families import Poisson
from statsmodels.genmod.cov_struct import Independence, Exchangeable, Unstructured, Autoregressive
from statsmodels.stats.multitest import multipletests
from scipy.stats import norm, pearsonr

#####################################
# Configuration (same as before)
#####################################

INPUT_JSONL_FILE = 'processed_all_articles_fixed_2.jsonl'
OUTPUT_DIR = 'graphs_analysis'
os.makedirs(OUTPUT_DIR, exist_ok=True)

CSV_OUTPUT_DIR = 'csv_raw_scores'
os.makedirs(CSV_OUTPUT_DIR, exist_ok=True)

OUTPUT_EXCEL_MAIN = 'analysis_main.xlsx'
OUTPUT_EXCEL_RAW = 'analysis_raw.xlsx'
OUTPUT_EXCEL_LMM = 'analysis_gee.xlsx'
OUTPUT_EXCEL_PLOTS = 'analysis_plots.xlsx'
OUTPUT_EXCEL_COMBINED = 'analysis_combined.xlsx'
OUTPUT_EXCEL_QIC = 'analysis_gee_qic.xlsx'

CATEGORIES = [
    'joy', 'sadness', 'anger', 'fear',
    'surprise', 'disgust', 'trust', 'anticipation',
    'negative_sentiment', 'positive_sentiment'
]

MEDIA_CATEGORIES = {
    'Scientific': ['nature', 'sciam', 'stat', 'newscientist'],
    'Left': ['theatlantic', 'the daily beast', 'the intercept', 'mother jones', 'msnbc', 'slate', 'vox', 'huffpost'],
    'Lean Left': ['ap', 'axios', 'cnn', 'guardian', 'business insider', 'nbcnews', 'npr', 'nytimes', 'politico', 'propublica', 'wapo', 'usa today'],
    'Center': ['reuters', 'marketwatch', 'financial times', 'newsweek', 'forbes'],
    'Lean Right': ['thedispatch', 'epochtimes', 'foxbusiness', 'wsj', 'national review', 'washtimes'],
    'Right': ['breitbart', 'theblaze', 'daily mail', 'dailywire', 'foxnews', 'nypost', 'newsmax'],
}

def setup_logging(log_file='analysis.log'):
    """Sets up file + console logging for debugging; suppress font-manager logs."""
    logging.getLogger('matplotlib.font_manager').setLevel(logging.WARNING)
    logging.basicConfig(
        filename=log_file,
        level=logging.DEBUG,
        format='%(asctime)s [%(levelname)s] %(module)s - %(message)s'
    )
    console = logging.StreamHandler(sys.stdout)
    console.setLevel(logging.DEBUG)
    console.setFormatter(logging.Formatter('[%(levelname)s] %(module)s: %(message)s'))
    logging.getLogger('').addHandler(console)
    logging.info("Logging initialized (file + console).")


def load_jsonl(jsonl_file):
    """Load JSONL into DataFrame."""
    records = []
    with open(jsonl_file, 'r', encoding='utf-8') as f:
        for line in tqdm(f, desc="Loading JSONL data"):
            try:
                record = json.loads(line)
                records.append(record)
            except json.JSONDecodeError as e:
                logging.error(f"JSON decode error: {e}")
    return pd.DataFrame(records)


def map_media_outlet_to_category(df, media_categories):
    """Map each outlet to known category or 'Other'."""
    if 'media_outlet' not in df.columns:
        raise KeyError("'media_outlet' column not found in DataFrame.")
    outlet_map = {}
    for cat, outs in media_categories.items():
        for o in outs:
            outlet_map[o.lower().strip()] = cat
    df['media_outlet_clean'] = df['media_outlet'].str.lower().str.strip()
    df['media_category'] = df['media_outlet_clean'].map(outlet_map).fillna('Other')
    unmapped = df.loc[df['media_category']=='Other','media_outlet'].unique()
    if len(unmapped)>0:
        logging.warning(f"Unmapped -> 'Other': {unmapped}")
        print(f"Warning: Some outlets mapped to 'Other': {unmapped}")
    return df


def aggregate_sentiment_scores(df, sentiment_categories):
    """Aggregate sums + counts for Quotation & Fulltext for each category."""
    rows=[]
    for cat in MEDIA_CATEGORIES.keys():
        sub = df[df['media_category']==cat]
        for s in sentiment_categories:
            pat = r'^' + re.escape(s) + r'_\d+$'
            matched = [c for c in df.columns if re.match(pat,c)]
            if matched:
                qsum = sub[matched].clip(lower=0).sum(skipna=True).sum()
                qcount = sub[matched].clip(lower=0).count().sum()
            else:
                qsum,qcount=0,0
            fcol = f"{s}_fulltext"
            if fcol in df.columns:
                fsum = sub[fcol].clip(lower=0).sum(skipna=True)
                fcount = sub[fcol].clip(lower=0).count()
            else:
                fsum,fcount=0,0
            rows.append({
                'Media Category':cat,
                'Sentiment/Emotion': s,
                'Quotation_Sum': qsum,
                'Quotation_Count': qcount,
                'Fulltext_Sum': fsum,
                'Fulltext_Count': fcount
            })
    return pd.DataFrame(rows)


def calculate_averages(aggdf):
    """Compute Quotation_Average and Fulltext_Average from sums/counts."""
    aggdf['Quotation_Average'] = aggdf.apply(
        lambda r: r['Quotation_Sum']/r['Quotation_Count'] if r['Quotation_Count']>0 else None, axis=1
    )
    aggdf['Fulltext_Average'] = aggdf.apply(
        lambda r: r['Fulltext_Sum']/r['Fulltext_Count'] if r['Fulltext_Count']>0 else None, axis=1
    )
    return aggdf


def calculate_mean_median(aggdf):
    """Compute overall mean & median across categories for each sentiment."""
    rows=[]
    for s in CATEGORIES:
        sub= aggdf[aggdf['Sentiment/Emotion']==s]
        qa= sub['Quotation_Average'].dropna()
        fa= sub['Fulltext_Average'].dropna()
        rows.append({
            'Sentiment/Emotion': s,
            'Mean_Quotation_Average': qa.mean() if len(qa)>0 else None,
            'Median_Quotation_Average': qa.median() if len(qa)>0 else None,
            'Mean_Fulltext_Average': fa.mean() if len(fa)>0 else None,
            'Median_Fulltext_Average': fa.median() if len(fa)>0 else None
        })
    return pd.DataFrame(rows)


def save_aggregated_scores_to_csv(aggdf, outdir, prefix='aggregated_sentiment_emotion_scores.csv'):
    path= os.path.join(outdir,prefix)
    try:
        aggdf.to_csv(path, index=False)
        msg=f"Aggregated sentiment/emotion scores saved to '{path}'."
        print(msg)
        logging.info(msg)
    except Exception as e:
        logging.error(f"Error saving aggregated: {e}")


def plot_statistics(aggdf, outdir):
    """Bar plots for Quotation & Fulltext averages."""
    sns.set_style('whitegrid')
    for s in CATEGORIES:
        sub= aggdf[aggdf['Sentiment/Emotion']==s]

        # Quotation
        plt.figure(figsize=(10,6))
        sns.barplot(x='Media Category',y='Quotation_Average',data=sub,color='steelblue')
        plt.title(f"Mean Quotation-Based '{s.capitalize()}' Scores")
        plt.xlabel('Media Category')
        plt.ylabel('Mean Quotation-Based Average Score')
        plt.xticks(rotation=45, ha='right')
        plt.tight_layout()
        out_q= os.path.join(outdir, f"quote_{s}.png")
        try:
            plt.savefig(out_q)
            plt.close()
            print(f"Quotation-Based '{s}' scores plot saved to '{out_q}'.")
        except Exception as e:
            logging.error(f"Error saving QPlot {s}: {e}")

        # Fulltext
        plt.figure(figsize=(10,6))
        sns.barplot(x='Media Category',y='Fulltext_Average',data=sub,color='darkorange')
        plt.title(f"Mean Fulltext-Based '{s.capitalize()}' Scores")
        plt.xlabel('Media Category')
        plt.ylabel('Mean Fulltext-Based Average Score')
        plt.xticks(rotation=45, ha='right')
        plt.tight_layout()
        out_f= os.path.join(outdir, f"fulltext_{s}.png")
        try:
            plt.savefig(out_f)
            plt.close()
            print(f"Fulltext-Based '{s}' scores plot saved to '{out_f}'.")
        except Exception as e:
            logging.error(f"Error saving FPlot {s}: {e}")


##############################
# GEE + QIC with extra steps
##############################

def remove_zero_variance_groups(df, score_col):
    """
    Remove any groups (media_outlet) that have zero variance in the outcome.
    This can reduce NaN/Inf problems for some structures.
    """
    group_var = df.groupby('media_outlet')[score_col].var()
    zero_var_groups = group_var[group_var==0].index
    if len(zero_var_groups)>0:
        logging.debug(f"Removing {len(zero_var_groups)} groups with zero variance for {score_col}: {list(zero_var_groups)}")
        df = df[~df['media_outlet'].isin(zero_var_groups)].copy()
    return df


def fit_gee_model(df, formula, groups, cov_struct, family=Poisson(), time=None):
    """Try to fit GEE, returning results or None. Keep QIC parse separate."""
    logging.debug(f"fit_gee_model => cov_struct={cov_struct.__class__.__name__}, time={'YES' if time is not None else 'NO'}")
    try:
        model = GEE.from_formula(formula, groups=groups, data=df, cov_struct=cov_struct, family=family, time=time)
    except Exception as e:
        logging.error(f"Model creation failed: {e}")
        return None

    try:
        res = model.fit()
        logging.debug("GEE fit successful.")
        return res
    except Exception as e:
        logging.error(f"GEE fit failed for {cov_struct.__class__.__name__}: {e}")
        return None


def run_gee_and_qic(df, sentiment, measure_type, structures, min_date=None):
    """
    Fit multiple structures => return {struct_name: (res, qic_val)}. We remove zero-variance groups
    to reduce numeric errors. Also handle QIC being tuple => use the first element if so.
    """
    logging.debug(f"run_gee_and_qic => sentiment={sentiment}, measure={measure_type}, min_date={min_date}")
    if measure_type=='Quotation':
        pat = r'^' + re.escape(sentiment) + r'_\d+$'
        matched = [c for c in df.columns if re.match(pat,c)]
        if not matched:
            logging.debug("No matched columns => {}")
            return {}
        tmp = df.copy()
        col_mean = f"{sentiment}_quotation_mean"
        tmp[col_mean] = tmp[matched].clip(lower=0).mean(axis=1)
        score_col = col_mean
    else:
        fullcol = f"{sentiment}_fulltext"
        if fullcol not in df.columns:
            logging.debug("No fulltext col => {}")
            return {}
        tmp= df.copy()
        clipped= f"{sentiment}_fulltext_clipped"
        tmp[clipped] = tmp[fullcol].clip(lower=0)
        score_col = clipped

    # must have >=2 categories
    mdf= tmp.dropna(subset=[score_col,'media_category','media_outlet']).copy()
    logging.debug(f"After dropna => {len(mdf)} rows for {score_col}.")
    if mdf['media_category'].nunique()<2:
        logging.debug("Not enough categories => {}")
        return {}

    # parse date
    if 'date' in mdf.columns:
        mdf['date']= pd.to_datetime(mdf['date'], errors='coerce')
        pre= len(mdf)
        mdf = mdf.dropna(subset=['date'])
        logging.debug(f"Dropped {pre - len(mdf)} invalid-date rows => {len(mdf)} remain.")
    mdf['media_category']= mdf['media_category'].astype('category')

    # remove zero-variance groups
    mdf = remove_zero_variance_groups(mdf, score_col)
    if len(mdf)<2:
        logging.debug("After removing zero-variance groups => <2 rows => no GEE possible.")
        return {}

    # create int_date if possible
    time_arr = None
    if min_date is not None and 'date' in mdf.columns and mdf['date'].notna().sum()>0:
        mdf= mdf.sort_values(['media_outlet','date'])
        mdf['int_date'] = (mdf['date'] - min_date).dt.days + 1
        mdf['int_date'] = mdf['int_date'].astype(np.int32)
        time_arr = mdf['int_date'].values

    formula= f"{score_col} ~ media_category"
    groups= "media_outlet"

    result_map= {}
    for s_name, s_obj in structures.items():
        used_time= None
        if time_arr is not None and len(mdf)>1:
            used_time= time_arr

        logging.debug(f"Fitting {s_name} => #rows={len(mdf)} time={'YES' if used_time is not None else 'NO'}")
        res = fit_gee_model(mdf, formula, groups=groups, cov_struct=s_obj, time=used_time)
        if res is None:
            logging.debug(f"Fit returned None => skipping {s_name}")
            continue

        # get QIC => might be tuple => just use the first float
        try:
            qic_ = res.qic()
            if isinstance(qic_, (list, tuple)) and len(qic_)>0:
                qic_val= qic_[0]
            else:
                qic_val= qic_
        except:
            qic_val= np.nan

        logging.debug(f"{s_name} => QIC={qic_val}")
        result_map[s_name]= (res, qic_val)
    return result_map


def fit_and_pick_best_structure(df, sentiment, measure_type, min_date=None):
    """Fit Ind, Exch, Unstr, AR1 => pick best by QIC => pairwise => dict."""
    structures_test= {
        'Independence': Independence(),
        'Exchangeable': Exchangeable(),
        'Unstructured': Unstructured(),
        'AR1': Autoregressive()
    }
    logging.debug(f"fit_and_pick_best_structure => s={sentiment}, measure={measure_type}")
    rmap= run_gee_and_qic(df, sentiment, measure_type, structures_test, min_date)
    if not rmap:
        logging.debug("No results => None.")
        return None

    best_s= None
    best_q= np.inf
    all_list= []
    for nm, (res_obj, qv) in rmap.items():
        all_list.append((nm,qv))
        if np.isscalar(qv) and qv<best_q:
            best_q= qv
            best_s= (nm, res_obj)

    if best_s is None:
        return None
    chosen_struct, best_res= best_s
    summary_txt= best_res.summary().as_text()

    params= best_res.params
    cov= best_res.cov_params()
    mod_df= best_res.model.data.frame
    cat_var= mod_df['media_category'].astype('category')
    if cat_var.nunique()<2:
        return None
    ref_cat= cat_var.cat.categories[0]
    param_names= best_res.model.exog_names
    cat_to_idx= {ref_cat:0}
    for c in cat_var.cat.categories[1:]:
        pname= f"media_category[T.{c}]"
        if pname in param_names:
            cat_to_idx[c]= param_names.index(pname)

    p_list= []
    cats= list(cat_var.cat.categories)
    for i in range(len(cats)):
        for j in range(i+1, len(cats)):
            catA,catB= cats[i], cats[j]
            cvec= np.zeros(len(params))
            if catA==ref_cat and catB!=ref_cat:
                cvec[cat_to_idx[catB]]= -1.
            elif catB==ref_cat and catA!=ref_cat:
                cvec[cat_to_idx[catA]]= 1.
            else:
                cvec[cat_to_idx[catA]]=1.
                cvec[cat_to_idx[catB]]= -1.
            diff_est= cvec@params
            diff_var= cvec@cov@cvec
            diff_se= np.sqrt(diff_var) if diff_var>0 else 0
            z= diff_est/diff_se if diff_se>0 else np.inf
            pval= 2*(1-norm.cdf(abs(z)))
            p_list.append((catA,catB,diff_est,diff_se,z,pval))

    pwdf= pd.DataFrame(p_list, columns=["CategoryA","CategoryB","Difference","SE","Z","p_value"])
    reject, p_adj, _, _= multipletests(pwdf['p_value'], method='holm')
    pwdf['p_value_adj']= p_adj
    pwdf['reject_H0']= reject

    print(f"    => Best structure for sentiment='{sentiment}', measure='{measure_type}' is: {chosen_struct} (QIC={best_q})")
    print("       Other structures and QIC =>")
    for (n,q) in all_list:
        print(f"         {n}: QIC={q}")

    return {
        'BestStructure': chosen_struct,
        'QIC_All': all_list,
        'GEE_Summary': summary_txt,
        'Pairwise': pwdf
    }


def run_gee_analyses(df):
    # min_date
    if 'date' in df.columns:
        df['date']= pd.to_datetime(df['date'], errors='coerce')
        valid= df.dropna(subset=['date'])
        min_date= valid['date'].min() if len(valid)>0 else None
    else:
        min_date= None
    logging.debug(f"run_gee_analyses => min_date={min_date}")

    results={}
    for s in CATEGORIES:
        results[s]={}
        for m in ['Quotation','Fulltext']:
            logging.debug(f"Analyzing {s}, {m}")
            br= fit_and_pick_best_structure(df, s, m, min_date)
            if br is not None:
                results[s][m]= br
    return results


###################################
#   Compile + main
###################################

def compile_results_into_multiple_workbooks(aggdf, statsdf, rawdf, gee_results,
                                           plots_dir, main_excel, raw_excel,
                                           lmm_excel, plots_excel, combined_excel):
    # main
    with pd.ExcelWriter(main_excel, engine='openpyxl') as w:
        aggdf.to_excel(w, sheet_name='Aggregated_Scores', index=False)
        statsdf.to_excel(w, sheet_name='Mean_Median_Statistics', index=False)

    # raw
    raw_df2= rawdf.copy()
    raw_df2['media_category']= raw_df2['media_category'].astype('category')
    raw_df2.sort_values(['media_category','media_outlet'], inplace=True)
    with pd.ExcelWriter(raw_excel, engine='openpyxl') as w:
        raw_df2.to_excel(w, sheet_name='Raw_Data', index=False)
        for s in CATEGORIES:
            scols= [c for c in raw_df2.columns if c.startswith(s+'_')]
            ssub= raw_df2[['media_category','media_outlet']+scols].copy()
            nm= f"Raw_{s[:29]}"
            ssub.to_excel(w, sheet_name=nm, index=False)

    # lmm
    with pd.ExcelWriter(lmm_excel, engine='openpyxl') as w:
        sum_rows=[]
        for s in gee_results:
            for m in gee_results[s]:
                sheetnm = f"GEE_{s[:20]}_{m[:8]}"
                bestst = gee_results[s][m]['BestStructure']
                qic_list= gee_results[s][m]['QIC_All']
                summ= gee_results[s][m]['GEE_Summary']
                pwdf= gee_results[s][m]['Pairwise']

                summ_df= pd.DataFrame({'GEE_Summary': summ.split('\n')})
                summ_df.to_excel(w, sheet_name=sheetnm, index=False)
                rstart= len(summ_df)+2
                info_df= pd.DataFrame({
                    'Info':[f"BestStructure={bestst}", f"AllStructuresAndQIC={qic_list}"]
                })
                info_df.to_excel(w, sheet_name=sheetnm, index=False, startrow=rstart)
                rstart+= (len(info_df)+2)
                pwdf.to_excel(w, sheet_name=sheetnm, index=False, startrow=rstart)

                sum_rows.append({
                    'Sentiment': s,
                    'Measure': m,
                    'BestStructure': bestst,
                    'SheetName': sheetnm
                })
        indexdf= pd.DataFrame(sum_rows)
        indexdf.to_excel(w, sheet_name='GEE_Results_Index', index=False)

    # plots
    wb= Workbook()
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])

    for s in CATEGORIES:
        qp= os.path.join(plots_dir,f"quote_{s}.png")
        if os.path.exists(qp):
            st= f"Quote_{s[:28]}"
            ws= wb.create_sheet(st)
            try:
                im= ExcelImage(qp)
                im.anchor='A1'
                ws.add_image(im)
            except Exception as e:
                logging.error(f"Error embedding {qp}: {e}")

        fp= os.path.join(plots_dir,f"fulltext_{s}.png")
        if os.path.exists(fp):
            st2= f"Fulltext_{s[:25]}"
            ws2= wb.create_sheet(st2)
            try:
                im2= ExcelImage(fp)
                im2.anchor='A1'
                ws2.add_image(im2)
            except Exception as e:
                logging.error(f"Error embed {fp}: {e}")

    wb.save(plots_excel)

    # combined
    raw_clean= raw_df2.copy()
    for c in raw_clean.columns:
        if raw_clean[c].dtype==object:
            raw_clean[c] = raw_clean[c].apply(lambda x: ", ".join(x) if isinstance(x,list) else x)

    wb2= Workbook()
    if 'Sheet' in wb2.sheetnames:
        wb2.remove(wb2['Sheet'])

    ws_ag= wb2.create_sheet("Aggregated_Scores")
    for r in dataframe_to_rows(aggdf, index=False, header=True):
        ws_ag.append(r)
    ws_st= wb2.create_sheet("Mean_Median_Statistics")
    for r in dataframe_to_rows(statsdf, index=False, header=True):
        ws_st.append(r)
    ws_raw= wb2.create_sheet("Raw_Data")
    for r in dataframe_to_rows(raw_clean, index=False, header=True):
        ws_raw.append(r)

    ws_g= wb2.create_sheet("GEE_Summaries")
    ws_g.append(["Sentiment","Measure","BestStructure","SummarySnippet"])
    for s in gee_results:
        for m in gee_results[s]:
            bests= gee_results[s][m]['BestStructure']
            summ= gee_results[s][m]['GEE_Summary'].split('\n')
            snip= "\n".join(summ[:6])
            ws_g.append([s, m, bests, snip])
    wb2.save(combined_excel)


def main():
    setup_logging()
    print("Single run with multi-structure GEE + QIC + clustering, skipping zero-variance groups to reduce errors.")
    df= load_jsonl(INPUT_JSONL_FILE)
    print(f"Total articles loaded: {len(df)}")
    df= map_media_outlet_to_category(df, MEDIA_CATEGORIES)

    # split raw
    chunk_size=20000
    for i in range(0, len(df), chunk_size):
        chunk= df.iloc[i:i+chunk_size]
        cpath= os.path.join(CSV_OUTPUT_DIR,f"raw_data_part_{(i//chunk_size)+1}.csv")
        chunk.to_csv(cpath, index=False)
        print(f"Saved chunk {(i//chunk_size)+1} to {cpath}")

    print("\nSummary Statistics:")
    print("Total number of articles:", len(df))
    print("\nNumber of articles per outlet:")
    print(df['media_outlet_clean'].value_counts())
    print("\nNumber of articles per category:")
    print(df['media_category'].value_counts())
    print()

    print("Aggregating sentiment/emotion scores per media category...")
    aggdf= aggregate_sentiment_scores(df, CATEGORIES)
    aggdf= calculate_averages(aggdf)
    statsdf= calculate_mean_median(aggdf)
    save_aggregated_scores_to_csv(aggdf, CSV_OUTPUT_DIR)
    plot_statistics(aggdf, OUTPUT_DIR)

    print("Fitting GEE models with multiple correlation structures (QIC) => best structure selection.")
    gee_results= run_gee_analyses(df)

    print("\n** GEE Model + Sensitivity Analysis **")
    for s in CATEGORIES:
        if s in gee_results:
            for m in gee_results[s]:
                best_s= gee_results[s][m]['BestStructure']
                print(f"For sentiment='{s}', measure='{m}', best structure => {best_s}")

    compile_results_into_multiple_workbooks(
        aggdf, statsdf, df, gee_results,
        OUTPUT_DIR, OUTPUT_EXCEL_MAIN, OUTPUT_EXCEL_RAW,
        OUTPUT_EXCEL_LMM, OUTPUT_EXCEL_PLOTS, OUTPUT_EXCEL_COMBINED
    )
    print("Analysis completed successfully.")


if __name__=="__main__":
    main()
