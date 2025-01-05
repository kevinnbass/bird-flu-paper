#!/usr/bin/env python3
# extract_data_10_fixed.py
"""
Creates two Excel files, non-transposed, with alignment across Quotation-based measures:

File A) raw_values_6cat_60tabs_no_blanks_scatterplot.xlsx
   => 60 tabs => (previous 50 tabs) + (10 new "Title" tabs)
   => 10 sentiments × [Quotation, Fulltext, Title_Intensity, Quotation_Intensity, Fulltext_Intensity, Title]
   => We unify rows in groups [Quotation, Fulltext, Title_Intensity],
                               [Quotation_Intensity, Fulltext_Intensity, Title_Intensity],
                               [Title].
   => If Quotation is missing for an article => Fulltext, Title_Intensity become blank in that row.
   => If Quotation_Intensity is missing => Fulltext_Intensity, Title_Intensity become blank in that row.
   => For Title (new measure), we do NOT forcibly blank it based on Quotation or Quotation_Intensity.

File B) raw_values_6cat_20tabs_intv2_no_blanks_scatterplot.xlsx
   => 20 tabs => 2 per sentiment: "3Int" and "2NonInt"
   => "3Int": [Fulltext_Intensity, Quotation_Intensity, Title_Intensity]
      if Quotation_Intensity is missing => we blank Fulltext_Intensity, Title_Intensity in that row.
   => "2NonInt": [Fulltext, Quotation]
      if Quotation is missing => we blank Fulltext in that row.

Author: ChatGPT
"""

import json
import re
import numpy as np
import pandas as pd
from tqdm import tqdm
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# ------------------ GLOBAL CONFIG ------------------
INPUT_JSONL_FILE = "processed_all_articles_fixed_5.jsonl"

OUTPUT_50TABS_SCATTER = "raw_values_6cat_60tabs_no_blanks_scatterplot.xlsx"
OUTPUT_20TABS_SCATTER = "raw_values_6cat_20tabs_intv2_no_blanks_scatterplot.xlsx"

# 6 categories
MEDIA_CATEGORIES = {
    "Scientific": ["nature","sciam","stat","newscientist"],
    "Left": ["theatlantic","the daily beast","the intercept","mother jones",
             "msnbc","slate","vox","huffpost"],
    "Lean Left": ["ap","axios","cnn","guardian","business insider","nbcnews",
                  "npr","nytimes","politico","propublica","wapo","usa today"],
    "Center": ["reuters","marketwatch","financial times","newsweek","forbes"],
    "Lean Right": ["thedispatch","epochtimes","foxbusiness","wsj","national review","washtimes"],
    "Right": ["breitbart","theblaze","daily mail","dailywire","foxnews","nypost","newsmax"],
}

SENTIMENTS = [
    "joy","sadness","anger","fear",
    "surprise","disgust","trust","anticipation",
    "negative_sentiment","positive_sentiment"
]

MEASURES_5 = [
    "Fulltext",
    "Quotation",
    "Fulltext_Intensity",
    "Quotation_Intensity",
    "Title_Intensity"
]

# We'll also add "Title" as a new measure, not in the original MEASURES_5
# but recognized in extract_measure_value() and placed in the extra measure group
THREE_INTENSITIES = ["Fulltext_Intensity","Quotation_Intensity","Title_Intensity"]
TWO_NONINTS       = ["Fulltext","Quotation"]


# --------------------------------------------------------------------------------
# 1) Load + map
# --------------------------------------------------------------------------------
def load_jsonl(jsonl_file):
    records = []
    with open(jsonl_file,"r",encoding="utf-8") as f:
        for line in tqdm(f,desc="Loading JSONL"):
            rec = json.loads(line)
            records.append(rec)
    return pd.DataFrame(records)

def map_outlet_to_6cats(df):
    cat_map={}
    for cat, outls in MEDIA_CATEGORIES.items():
        for o in outls:
            cat_map[o.lower().strip()] = cat
    df2 = df.copy()
    df2["media_outlet_clean"] = df2["media_outlet"].str.lower().str.strip()
    df2["media_category_6"]   = df2["media_outlet_clean"].map(cat_map).fillna("Other")
    return df2


# --------------------------------------------------------------------------------
# 2) extract_measure_value
# --------------------------------------------------------------------------------
def extract_measure_value(row, sentiment, measure):
    """
    Quotation => average columns ^{sentiment}_\\d+$
    Quotation_Intensity => average columns ^{sentiment}_\\d+_intensity$
    Fulltext => row[f"{sentiment}_fulltext"]
    Fulltext_Intensity => row[f"{sentiment}_fulltext_intensity"]
    Title_Intensity => row[f"{sentiment}_title_intensity"]
    Title => row[f"{sentiment}_title"]  (NEW measure)
    """
    try:
        if measure=="Quotation":
            pat = rf"^{re.escape(sentiment)}_\d+$"
            matched = [c for c in row.index if re.match(pat,c)]
            if not matched:
                return np.nan
            vals = row[matched].astype(float).clip(lower=0)
            return vals.mean(skipna=True)

        elif measure=="Quotation_Intensity":
            pat = rf"^{re.escape(sentiment)}_\d+_intensity$"
            matched = [c for c in row.index if re.match(pat,c)]
            if not matched:
                return np.nan
            vals = row[matched].astype(float).clip(lower=0)
            return vals.mean(skipna=True)

        elif measure=="Fulltext":
            col = f"{sentiment}_fulltext"
            if col not in row.index:
                return np.nan
            return float(row[col])

        elif measure=="Fulltext_Intensity":
            col = f"{sentiment}_fulltext_intensity"
            if col not in row.index:
                return np.nan
            return float(row[col])

        elif measure=="Title_Intensity":
            col = f"{sentiment}_title_intensity"
            if col not in row.index:
                return np.nan
            return float(row[col])

        elif measure=="Title":
            # The new measure we add
            col = f"{sentiment}_title"
            if col not in row.index:
                return np.nan
            return float(row[col])

        else:
            return np.nan
    except:
        return np.nan


# --------------------------------------------------------------------------------
# remove_default_sheet_if_present
# --------------------------------------------------------------------------------
def remove_default_sheet_if_present(wb):
    if "Sheet" in wb.sheetnames:
        default_ws = wb["Sheet"]
        wb.remove(default_ws)

# --------------------------------------------------------------------------------
# 3) unify_rows_across_measures => gather data, then forcibly blank Fulltext/Title_Intensity if Quotation=NaN
#    forcibly blank Fulltext_Intensity/Title_Intensity if Quotation_Intensity=NaN
#    (Note: If measure group == ["Title"] only, no forced blanking)
# --------------------------------------------------------------------------------
def gather_article_vals_for_measure(df, sentiment, measure):
    """Return {cat: list of (article_id, measure_val)} for that measure, storing articles even if val=NaN."""
    cat_names = ["Scientific","Left","Lean Left","Center","Lean Right","Right"]
    cat_dict = {c: [] for c in cat_names}
    for i, row in df.iterrows():
        cat_6 = row["media_category_6"]
        if cat_6 not in cat_dict:
            continue
        a_id = row["article_id"]
        val  = extract_measure_value(row, sentiment, measure)
        cat_dict[cat_6].append((a_id,val))
    return cat_dict


def unify_rows_across_measures(df, sentiment, measure_group):
    """
    measure_group could be e.g. ["Quotation","Fulltext","Title_Intensity"], 
    or ["Quotation_Intensity","Fulltext_Intensity","Title_Intensity"], 
    or the single ["Title"] (new).
    1) Build measure_data[measure][cat][art_id] = val
    2) forcibly blank if Quotation=NaN => Fulltext=NaN, Title_Intensity=NaN
       forcibly blank if Quotation_Intensity=NaN => Fulltext_Intensity=NaN, Title_Intensity=NaN
    3) unify => cat_out[cat] => list of row_dict => each => {article_id, measureX...}
    """
    cat_names = ["Scientific","Left","Lean Left","Center","Lean Right","Right"]
    measure_data={}
    for meas in measure_group:
        measure_data[meas] = {cat:{} for cat in cat_names}

    # fill
    for meas in measure_group:
        cat_dict = gather_article_vals_for_measure(df, sentiment, meas)
        for cat in cat_names:
            for (a_id,v) in cat_dict[cat]:
                measure_data[meas][cat][a_id] = v

    # forcibly blank if Quotation=NaN (only if measure_group contains Quotation)
    if "Quotation" in measure_group:
        for cat in cat_names:
            for a_id,qval in measure_data["Quotation"][cat].items():
                if pd.isna(qval):
                    if "Fulltext" in measure_group:
                        measure_data["Fulltext"][cat][a_id] = np.nan
                    if "Title_Intensity" in measure_group:
                        measure_data["Title_Intensity"][cat][a_id] = np.nan

    # forcibly blank if Quotation_Intensity=NaN (only if measure_group contains Quotation_Intensity)
    if "Quotation_Intensity" in measure_group:
        for cat in cat_names:
            for a_id,qi_val in measure_data["Quotation_Intensity"][cat].items():
                if pd.isna(qi_val):
                    if "Fulltext_Intensity" in measure_group:
                        measure_data["Fulltext_Intensity"][cat][a_id] = np.nan
                    if "Title_Intensity" in measure_group:
                        measure_data["Title_Intensity"][cat][a_id] = np.nan

    # unify => cat_out
    cat_out={}
    for cat in cat_names:
        # gather all article IDs
        id_set = set()
        for meas in measure_group:
            id_set.update(measure_data[meas][cat].keys())
        sorted_ids = sorted(list(id_set))
        cat_out[cat] = []
        for a_id in sorted_ids:
            row_dict={"article_id": a_id}
            for meas in measure_group:
                val = measure_data[meas][cat].get(a_id, np.nan)
                row_dict[meas] = val
            cat_out[cat].append(row_dict)
    return cat_out

# --------------------------------------------------------------------------------
# 4) CREATE THE 60TAB => measure_alignment_groups => produce separate tabs
# --------------------------------------------------------------------------------
def create_50tabs_scatter(df, wb):
    """
    We keep the function name for consistency, but now it will produce 60 tabs total:
    - 2 'alignment groups' from before
    - plus 1 new group for 'Title' only
    """
    remove_default_sheet_if_present(wb)

    cat_names = ["Scientific","Left","Lean Left","Center","Lean Right","Right"]
    # Now we define 3 measure groups instead of 2
    measure_alignment_groups = [
        ["Quotation","Fulltext","Title_Intensity"],
        ["Quotation_Intensity","Fulltext_Intensity","Title_Intensity"],
        ["Title"],  # NEW measure group for raw Title
    ]
    for sent in SENTIMENTS:
        for group in measure_alignment_groups:
            cat_out = unify_rows_across_measures(df, sent, group)
            for meas in group:
                # We create a new tab for each measure in this group
                # The old code excluded any measure not in MEASURES_5;
                # but now "Title" won't be in MEASURES_5, so let's not skip it.
                # We'll produce a tab for "Title" anyway.

                tab_name=f"{sent[:10]}_{meas[:10]}"
                # Avoid duplicates if it already exists
                if tab_name in wb.sheetnames:
                    continue

                # build DataFrame => shape (max_len, 6)
                max_len = max(len(cat_out[c]) for c in cat_names) if len(cat_names) > 0 else 0
                data_dict={}
                for c in cat_names:
                    col_vals=[]
                    row_list = cat_out[c]
                    for r_i in range(max_len):
                        if r_i<len(row_list):
                            val= row_list[r_i].get(meas,np.nan)
                            col_vals.append(val)
                        else:
                            col_vals.append(np.nan)
                    data_dict[c]=col_vals
                df_tab = pd.DataFrame(data_dict, columns=cat_names)

                ws = wb.create_sheet(title=tab_name)
                row_data = [cat_names]  # first row => col names
                for r_i in range(len(df_tab)):
                    rowvals=[]
                    for c_ in cat_names:
                        rowvals.append(df_tab.loc[r_i,c_])
                    row_data.append(rowvals)
                for R_i, rowv in enumerate(row_data,1):
                    for C_i, valv in enumerate(rowv,1):
                        ws.cell(row=R_i,column=C_i,value=valv)


# --------------------------------------------------------------------------------
# 5) CREATE THE 20TAB => 2 tabs / sentiment => "3Int" & "2NonInt"
# --------------------------------------------------------------------------------
def create_20tabs_scatter(df, wb_20):
    remove_default_sheet_if_present(wb_20)

    cat_names = ["Scientific","Left","Lean Left","Center","Lean Right","Right"]

    for sent in SENTIMENTS:
        # (A) 3Int
        measure_grpA = ["Fulltext_Intensity","Quotation_Intensity","Title_Intensity"]
        cat_outA= unify_rows_across_measures(df, sent, measure_grpA)
        out_colsA=[]
        for cat in cat_names:
            for m in measure_grpA:
                out_colsA.append(f"{cat}_{m}")
        max_lenA = max(len(cat_outA[c]) for c in cat_names)
        dataA={}
        for colN in out_colsA:
            dataA[colN]=[]
        for cat in cat_names:
            row_list = cat_outA[cat]
            for m in measure_grpA:
                col_key=f"{cat}_{m}"
                col_vals=[]
                for r_i in range(max_lenA):
                    if r_i<len(row_list):
                        val = row_list[r_i].get(m,np.nan)
                        col_vals.append(val)
                    else:
                        col_vals.append(np.nan)
                dataA[col_key] = col_vals
        dfA = pd.DataFrame(dataA, columns=out_colsA)
        tabA_name= f"{sent[:10]}_3Int"
        wsA=wb_20.create_sheet(title=tabA_name)
        row_dataA = [out_colsA]
        for r_i in range(len(dfA)):
            rowvals=[]
            for c_ in out_colsA:
                rowvals.append(dfA.loc[r_i,c_])
            row_dataA.append(rowvals)
        for R_i, rowv in enumerate(row_dataA,1):
            for C_i, valv in enumerate(rowv,1):
                wsA.cell(row=R_i,column=C_i,value=valv)

        # (B) 2NonInt
        measure_grpB=["Fulltext","Quotation"]
        cat_outB=unify_rows_across_measures(df,sent, measure_grpB)
        out_colsB=[]
        for cat in cat_names:
            for m in measure_grpB:
                out_colsB.append(f"{cat}_{m}")
        max_lenB= max(len(cat_outB[c]) for c in cat_names)
        dataB={}
        for colN in out_colsB:
            dataB[colN]=[]
        for cat in cat_names:
            row_list = cat_outB[cat]
            for m in measure_grpB:
                col_key=f"{cat}_{m}"
                col_vals=[]
                for r_i in range(max_lenB):
                    if r_i< len(row_list):
                        val = row_list[r_i].get(m, np.nan)
                        col_vals.append(val)
                    else:
                        col_vals.append(np.nan)
                dataB[col_key]= col_vals
        dfB= pd.DataFrame(dataB, columns=out_colsB)
        tabB_name= f"{sent[:10]}_2NonInt"
        wsB= wb_20.create_sheet(title=tabB_name)
        row_dataB=[out_colsB]
        for r_i in range(len(dfB)):
            rowvals=[]
            for c_ in out_colsB:
                rowvals.append(dfB.loc[r_i,c_])
            row_dataB.append(rowvals)
        for R_i, rowv in enumerate(row_dataB,1):
            for C_i, valv in enumerate(rowv,1):
                wsB.cell(row=R_i,column=C_i,value=valv)


# --------------------------------------------------------------------------------
# 6) main
# --------------------------------------------------------------------------------
def main():
    print("Loading + mapping data...")
    df = load_jsonl(INPUT_JSONL_FILE)
    df = map_outlet_to_6cats(df)
    if "article_id" not in df.columns:
        df["article_id"] = df.index + 1

    print(f"Creating 60-tab scatter file => {OUTPUT_50TABS_SCATTER}")
    wb50 = Workbook()
    create_50tabs_scatter(df, wb50)
    wb50.save(OUTPUT_50TABS_SCATTER)
    print(f"Wrote => {OUTPUT_50TABS_SCATTER}")

    print(f"Creating 20-tab scatter file => {OUTPUT_20TABS_SCATTER}")
    wb20 = Workbook()
    create_20tabs_scatter(df, wb20)
    wb20.save(OUTPUT_20TABS_SCATTER)
    print(f"Wrote => {OUTPUT_20TABS_SCATTER}")

    print("All done.")


if __name__=="__main__":
    main()
