#!/usr/bin/env python3
# generate_50_tabs_six_cats.py
"""
Creates 50 tabs (10 sentiments × 5 measures) in a single XLSX:
   - Each tab => [ArticleID, Scientific, Left, Lean Left, Center, Lean Right, Right]
   - Rows => one row per article
   - The cell for e.g. "Left" is the raw measure value for that article if its
     media_category_6 is "Left". Others are NaN.
   - Quotation & Quotation_Intensity => mean across matched columns
   - Fulltext, Fulltext_Intensity, Title_Intensity => direct from single column
"""

import json
import re
import numpy as np
import pandas as pd
from tqdm import tqdm
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# 1) Config
INPUT_JSONL_FILE = "processed_all_articles_fixed_4.jsonl"
OUTPUT_XLSX_FILE = "raw_values_50tabs_sixcats.xlsx"

# 6 categories
MEDIA_CATEGORIES = {
    "Scientific": ["nature","sciam","stat","newscientist"],
    "Left": ["theatlantic","the daily beast","the intercept","mother jones","msnbc","slate","vox","huffpost"],
    "Lean Left": ["ap","axios","cnn","guardian","business insider","nbcnews","npr","nytimes","politico","propublica","wapo","usa today"],
    "Center": ["reuters","marketwatch","financial times","newsweek","forbes"],
    "Lean Right": ["thedispatch","epochtimes","foxbusiness","wsj","national review","washtimes"],
    "Right": ["breitbart","theblaze","daily mail","dailywire","foxnews","nypost","newsmax"],
}

SENTIMENTS = [
    "joy","sadness","anger","fear",
    "surprise","disgust","trust","anticipation",
    "negative_sentiment","positive_sentiment"
]

MEASURES = [
    "Fulltext",
    "Quotation",
    "Fulltext_Intensity",
    "Quotation_Intensity",
    "Title_Intensity",
]

# --------------------------------------------------------------------------------
# HELPER: Load + map
# --------------------------------------------------------------------------------
def load_jsonl(jsonl_file):
    records = []
    with open(jsonl_file,"r",encoding="utf-8") as f:
        for line in tqdm(f, desc="Loading JSONL"):
            rec=json.loads(line)
            records.append(rec)
    df = pd.DataFrame(records)
    return df

def map_outlet_to_6cats(df):
    cat_map = {}
    for cat, outls in MEDIA_CATEGORIES.items():
        for o in outls:
            cat_map[o.lower().strip()] = cat

    df2 = df.copy()
    df2["media_outlet_clean"] = df2["media_outlet"].str.lower().str.strip()
    df2["media_category_6"]   = df2["media_outlet_clean"].map(cat_map).fillna("Other")
    return df2

# --------------------------------------------------------------------------------
# HELPER: Extract single measure for a single sentiment
# --------------------------------------------------------------------------------
def extract_measure_value(row, sentiment, measure):
    """Return numeric measure for this row, for the given sentiment."""
    try:
        if measure=="Quotation":
            # average columns like joy_0, joy_1, ...
            pat = rf"^{re.escape(sentiment)}_\d+$"
            matched = [c for c in row.index if re.match(pat,c)]
            if not matched:
                return np.nan
            vals = row[matched].astype(float).clip(lower=0)
            return vals.mean(skipna=True)

        elif measure=="Quotation_Intensity":
            # average columns like joy_0_intensity, joy_1_intensity, ...
            pat = rf"^{re.escape(sentiment)}_\d+_intensity$"
            matched = [c for c in row.index if re.match(pat,c)]
            if not matched:
                return np.nan
            vals = row[matched].astype(float).clip(lower=0)
            return vals.mean(skipna=True)

        elif measure=="Fulltext":
            col = f"{sentiment}_fulltext"
            if col not in row.index:
                return np.nan
            return float(row[col])

        elif measure=="Fulltext_Intensity":
            col = f"{sentiment}_fulltext_intensity"
            if col not in row.index:
                return np.nan
            return float(row[col])

        elif measure=="Title_Intensity":
            col = f"{sentiment}_title_intensity"
            if col not in row.index:
                return np.nan
            return float(row[col])

        else:
            return np.nan

    except:
        return np.nan

# --------------------------------------------------------------------------------
# MAIN
# --------------------------------------------------------------------------------
def main():
    df = load_jsonl(INPUT_JSONL_FILE)
    df = map_outlet_to_6cats(df)

    # If no article_id, define
    if "article_id" not in df.columns:
        df["article_id"] = df.index + 1

    wb = Workbook()
    # Remove default
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    # We'll have 50 tabs => for each sentiment, for each measure
    for sent in SENTIMENTS:
        for meas in MEASURES:
            tab_name = f"{sent[:10]}_{meas[:10]}"  # shorten if needed for Excel sheet name limit

            # Build a data frame with columns:
            #   [ArticleID, Scientific, Left, Lean Left, Center, Lean Right, Right]
            # Each row => one article
            # The cell => measure value if that article is that category, else NaN
            # Approach:
            #    1) For each article, compute measure_value
            #    2) Put measure_value into the column for that article's category_6
            #    3) Others are NaN

            rows = []
            for i, row in df.iterrows():
                article_id = row["article_id"]
                cat_6      = row["media_category_6"]
                val        = extract_measure_value(row, sent, meas)

                # We'll build a dict with all 6 columns = np.nan
                row_dict = {
                    "ArticleID": article_id,
                    "Scientific": np.nan,
                    "Left": np.nan,
                    "Lean Left": np.nan,
                    "Center": np.nan,
                    "Lean Right": np.nan,
                    "Right": np.nan,
                }
                # If cat_6 is among the 6 known categories => fill val
                if cat_6 in row_dict.keys():
                    row_dict[cat_6] = val

                rows.append(row_dict)

            out_df = pd.DataFrame(rows)

            # We might want to sort by ArticleID, optionally
            out_df.sort_values("ArticleID", inplace=True)

            # Now add a sheet
            ws = wb.create_sheet(title=tab_name)
            # Convert to rows
            df_rows = dataframe_to_rows(out_df, index=False, header=True)
            for r_idx, row_vals in enumerate(df_rows, 1):
                for c_idx, cell_val in enumerate(row_vals, 1):
                    ws.cell(row=r_idx, column=c_idx, value=cell_val)

    wb.save(OUTPUT_XLSX_FILE)
    print(f"Wrote => {OUTPUT_XLSX_FILE}, with {len(SENTIMENTS)*len(MEASURES)} tabs")

if __name__=="__main__":
    main()
