#!/usr/bin/env python3
# gee_w_validation_4.py
"""
Script that:
  1) Loads data from JSONL
  2) Chunks & saves CSV
  3) Performs correlation analyses & scatterplots for Quotation vs Fulltext
     => Also writes correlation summary + combined Z-scatter data into
        'quotation_fulltext_correlation.xlsx'
  4) Aggregates sentiment/emotion scores + bar plots
  5) Finds best QIC combination (Ind/Exch × scale=[none,pearson,deviance,ub,bc])
  6) For each best QIC model:
     - Refit & produce GEE summary
     - Compute pairwise comparisons (BH correction)
     - Build a [category, diffIDs] table => if category i differs from j => j's index => "2,4,5"
  7) Writes summary, pairwise, & "DiffIDs" table into analysis_gee.xlsx
  8) Also saves analysis_main, analysis_raw, analysis_plots, analysis_combined
  9) Adds validations: Negative Binomial check, residual checks, cross-validation,
     sensitivity analysis, bootstrap, plus storing the combined Z-scatter data in a second sheet.

Fixes the "TypeError: unsupported operand type(s) for -: 'tuple' and 'tuple'" by extracting [0] from `qic()` outputs if they are tuples.
"""

import json
import os
import re
import sys
import warnings
import logging
import math
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import seaborn as sns
from tqdm import tqdm

from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill

import statsmodels.api as sm
from statsmodels.genmod.generalized_estimating_equations import GEE
from statsmodels.genmod.families import Poisson, NegativeBinomial
from statsmodels.genmod.cov_struct import Independence, Exchangeable
from statsmodels.stats.multitest import multipletests
from scipy.stats import pearsonr, norm

# --------------------- #
# Configuration
# --------------------- #
INPUT_JSONL_FILE = "processed_all_articles_fixed_2.jsonl"

OUTPUT_DIR = "graphs_analysis"
os.makedirs(OUTPUT_DIR, exist_ok=True)

CSV_OUTPUT_DIR = "csv_raw_scores"
os.makedirs(CSV_OUTPUT_DIR, exist_ok=True)

OUTPUT_EXCEL_MAIN = "analysis_main.xlsx"
OUTPUT_EXCEL_RAW = "analysis_raw.xlsx"
OUTPUT_EXCEL_GEE = "analysis_gee.xlsx"
OUTPUT_EXCEL_PLOTS = "analysis_plots.xlsx"
OUTPUT_EXCEL_COMBINED = "analysis_combined.xlsx"

LOG_FILE = "analysis.log"

CATEGORIES = [
    "joy","sadness","anger","fear",
    "surprise","disgust","trust","anticipation",
    "negative_sentiment","positive_sentiment"
]

MEDIA_CATEGORIES = {
    "Scientific": ["nature","sciam","stat","newscientist"],
    "Left": ["theatlantic","the daily beast","the intercept","mother jones","msnbc","slate","vox","huffpost"],
    "Lean Left": ["ap","axios","cnn","guardian","business insider","nbcnews","npr","nytimes","politico","propublica","wapo","usa today"],
    "Center": ["reuters","marketwatch","financial times","newsweek","forbes"],
    "Lean Right": ["thedispatch","epochtimes","foxbusiness","wsj","national review","washtimes"],
    "Right": ["breitbart","theblaze","daily mail","dailywire","foxnews","nypost","newsmax"],
}


def setup_logging():
    """Initialize logging to both file and console."""
    log_format = "%(asctime)s [%(levelname)s] %(module)s - %(message)s"
    logging.basicConfig(filename=LOG_FILE, level=logging.DEBUG, format=log_format)
    console = logging.StreamHandler(sys.stdout)
    console.setLevel(logging.INFO)
    console.setFormatter(logging.Formatter(log_format))
    logging.getLogger().addHandler(console)
    logging.info("Logging initialized (file+console).")


warnings.filterwarnings(
    "ignore",
    message="QIC values obtained using scale=None are not appropriate for comparing models"
)

###############################################################################
# 1) Load, chunk, basic stats
###############################################################################
def load_jsonl(jsonl_file):
    logging.info(f"Loading JSONL from {jsonl_file}")
    records=[]
    with open(jsonl_file, "r", encoding="utf-8") as f:
        for line in tqdm(f, desc="Loading JSONL data"):
            rec=json.loads(line)
            records.append(rec)
    df=pd.DataFrame(records)
    logging.debug(f"Loaded DataFrame shape={df.shape}")
    return df

def map_media_outlet_to_category(df):
    cat_map={}
    for cat, outls in MEDIA_CATEGORIES.items():
        for o in outls:
            cat_map[o.lower().strip()] = cat

    if "media_outlet" not in df.columns:
        raise KeyError("'media_outlet' not found in data")

    df["media_outlet_clean"] = df["media_outlet"].str.lower().str.strip()
    df["media_category"] = df["media_outlet_clean"].map(cat_map).fillna("Other")

    unmapped = df[df["media_category"]=="Other"]["media_outlet"].unique()
    if len(unmapped)>0:
        logging.warning(f"Unmapped => {unmapped}")
        print(f"Warning: Not mapped => {unmapped}")
    return df

def chunk_and_save(df, chunk_size=20000):
    logging.info(f"Chunking DataFrame => len={len(df)}, chunk_size={chunk_size}")
    for i in range(0,len(df),chunk_size):
        part=df.iloc[i:i+chunk_size]
        out_csv=os.path.join("csv_raw_scores",f"raw_data_part_{(i//chunk_size)+1}.csv")
        part.to_csv(out_csv,index=False)
        print(f"Saved chunk {(i//chunk_size)+1} to {out_csv}")

def print_basic_stats(df):
    logging.info(f"Basic stats => total articles = {len(df)}")
    print("\nSummary Statistics:")
    print("Total number of articles:", len(df))
    if "media_outlet_clean" in df.columns:
        vc=df["media_outlet_clean"].value_counts()
        print("\nArticles per outlet:")
        print(vc)
    if "media_category" in df.columns:
        vc2=df["media_category"].value_counts()
        print("\nArticles per category:")
        print(vc2)
    print()

###############################################################################
# 2) Quotation vs. Fulltext correlation
###############################################################################
def analyze_quotation_fulltext_correlation(df):
    """
    1) For each (media_category, sentiment), compute Quotation_Average & Fulltext_Average
    2) Show correlation in a bar chart
    3) Combined z-scatter => "quotation_fulltext_correlation.xlsx", second sheet
    """
    from scipy.stats import pearsonr
    logging.info("Analyzing Quotation vs Fulltext correlation.")
    records=[]
    for cat in df["media_category"].dropna().unique():
        dcat=df[df["media_category"]==cat]
        for s in CATEGORIES:
            pat=rf"^{s}_\d+$"
            matched=[c for c in dcat.columns if re.match(pat,c)]
            if matched:
                clp=dcat[matched].clip(lower=0)
                qsum=clp.sum(skipna=True).sum()
                qcount=clp.count().sum()
                qavg=qsum/qcount if qcount>0 else np.nan
            else:
                qavg=np.nan

            fcol=f"{s}_fulltext"
            if fcol in dcat.columns:
                fv=dcat[fcol].clip(lower=0)
                f_sum=fv.sum(skipna=True)
                f_cnt=fv.count()
                favg=f_sum/f_cnt if f_cnt>0 else np.nan
            else:
                favg=np.nan

            records.append({
                "MediaCategory": cat,
                "Sentiment": s,
                "Quotation_Average": qavg,
                "Fulltext_Average": favg
            })

    agg_df=pd.DataFrame(records)
    cor_results=[]
    scatter_map={}

    for s in CATEGORIES:
        sub=agg_df[(agg_df["Sentiment"]==s)].dropna(subset=["Quotation_Average","Fulltext_Average"])
        if len(sub)>1:
            cor_val,_=pearsonr(sub["Quotation_Average"],sub["Fulltext_Average"])
        else:
            cor_val=np.nan
        cor_results.append({"Sentiment":s,"Correlation":cor_val})
        scatter_map[s]=sub.copy()

    cor_df=pd.DataFrame(cor_results)
    # bar
    plt.figure(figsize=(8,5))
    sns.barplot(data=cor_df, x="Sentiment", y="Correlation", color="gray")
    plt.title("Correlation (Quotation vs. Fulltext) per Sentiment")
    plt.xticks(rotation=45,ha="right")
    plt.ylim(-1,1)
    plt.tight_layout()
    cbar=os.path.join(OUTPUT_DIR,"correlation_quotation_fulltext_bar.png")
    try:
        plt.savefig(cbar)
    except:
        pass
    plt.close()

    # Write correlation + combined z to xlsx
    with pd.ExcelWriter("quotation_fulltext_correlation.xlsx", engine="openpyxl") as w:
        cor_df.to_excel(w, sheet_name="CorrelationSummary", index=False)

        combo=[]
        for s, sd in scatter_map.items():
            if not sd.empty:
                cpy=sd.copy()
                cpy["Sentiment"]=s
                combo.append(cpy)
        if combo:
            allc=pd.concat(combo, ignore_index=True)
            allc["Qmean"]=allc.groupby("Sentiment")["Quotation_Average"].transform("mean")
            allc["Qstd"]=allc.groupby("Sentiment")["Quotation_Average"].transform("std")
            allc["Fmean"]=allc.groupby("Sentiment")["Fulltext_Average"].transform("mean")
            allc["Fstd"]=allc.groupby("Sentiment")["Fulltext_Average"].transform("std")
            allc["Quotation_Z"]=(allc["Quotation_Average"]-allc["Qmean"])/allc["Qstd"]
            allc["Fulltext_Z"]=(allc["Fulltext_Average"]-allc["Fmean"])/allc["Fstd"]
            v=allc.dropna(subset=["Quotation_Z","Fulltext_Z"])
            v.to_excel(w, sheet_name="CombinedZScatterData", index=False)

            if len(v)>1:
                r_val2,_=pearsonr(v["Quotation_Z"], v["Fulltext_Z"])
            else:
                r_val2=np.nan

            plt.figure(figsize=(7,5))
            sns.regplot(data=v, x="Quotation_Z", y="Fulltext_Z",
                        scatter_kws={"color":"black","alpha":0.6}, line_kws={"color":"red"})
            plt.title(f"All Sentiments Combined (Z-scores)\nr={r_val2:.3f}")
            plt.tight_layout()
            csc=os.path.join(OUTPUT_DIR,"combined_normalized_scatter.png")
            try:
                plt.savefig(csc)
            except:
                pass
            plt.close()

    # also store cor_df in CSV
    out_csv=os.path.join(CSV_OUTPUT_DIR,"quotation_fulltext_correlation.csv")
    cor_df.to_csv(out_csv, index=False)
    logging.info(f"Correlation data => {out_csv}")


###############################################################################
# 3) Aggregation & Stats
###############################################################################
def aggregate_sentiment_scores(df, sentiments):
    logging.info("Aggregating sentiment/emotion scores by category + sentiment.")
    recs=[]
    for cat in MEDIA_CATEGORIES.keys():
        sub=df[df["media_category"]==cat]
        for s in sentiments:
            pat=rf"^{re.escape(s)}_\d+$"
            matched=[c for c in sub.columns if re.match(pat,c)]
            if matched:
                qsum=sub[matched].clip(lower=0).sum(skipna=True).sum()
                qcount=sub[matched].clip(lower=0).count().sum()
            else:
                qsum,qcount=(0,0)

            fcol=f"{s}_fulltext"
            if fcol in sub.columns:
                fv=sub[fcol].clip(lower=0)
                f_sum=fv.sum(skipna=True)
                f_cnt=fv.count()
            else:
                f_sum,f_cnt=(0,0)

            recs.append({
                "Media Category":cat,
                "Sentiment/Emotion": s,
                "Quotation_Sum": qsum,
                "Quotation_Count": qcount,
                "Fulltext_Sum": f_sum,
                "Fulltext_Count": f_cnt
            })
    return pd.DataFrame(recs)

def calculate_averages(agg_df):
    logging.info("Calculating Quotation_Average & Fulltext_Average.")
    def sdiv(a,b): return a/b if b>0 else None
    agg_df["Quotation_Average"]=agg_df.apply(lambda r: sdiv(r["Quotation_Sum"],r["Quotation_Count"]),axis=1)
    agg_df["Fulltext_Average"]=agg_df.apply(lambda r: sdiv(r["Fulltext_Sum"],r["Fulltext_Count"]),axis=1)
    return agg_df

def calculate_mean_median(agg_df):
    logging.info("Computing global mean/median of Quotation/Fulltext averages.")
    rows=[]
    for s in CATEGORIES:
        sub=agg_df[agg_df["Sentiment/Emotion"]==s]
        qa=sub["Quotation_Average"].dropna()
        fa=sub["Fulltext_Average"].dropna()
        rows.append({
            "Sentiment/Emotion": s,
            "Mean_Quotation_Average": qa.mean() if len(qa)>0 else None,
            "Median_Quotation_Average": qa.median() if len(qa)>0 else None,
            "Mean_Fulltext_Average": fa.mean() if len(fa)>0 else None,
            "Median_Fulltext_Average": fa.median() if len(fa)>0 else None
        })
    return pd.DataFrame(rows)

def save_aggregated_scores_to_csv(agg_df, out_dir):
    fn=os.path.join(out_dir,"aggregated_sentiment_emotion_scores.csv")
    agg_df.to_csv(fn,index=False)
    print(f"Aggregated sentiment/emotion scores => {fn}")
    logging.info(f"Aggregated => {fn}")

def plot_statistics(agg_df, out_dir):
    logging.info("Plotting bar charts for Quotation/Fulltext across categories.")
    sns.set_style("whitegrid")
    for s in CATEGORIES:
        sub=agg_df[agg_df["Sentiment/Emotion"]==s]
        # Quotation
        plt.figure(figsize=(10,6))
        sns.barplot(data=sub, x="Media Category", y="Quotation_Average", color="steelblue")
        plt.title(f"Mean Quotation-Based '{s.capitalize()}' Scores")
        plt.xticks(rotation=45,ha="right")
        plt.tight_layout()
        out1=os.path.join(out_dir,f"quote_{s}.png")
        try:
            plt.savefig(out1)
        except:
            pass
        plt.close()

        # Fulltext
        plt.figure(figsize=(10,6))
        sns.barplot(data=sub, x="Media Category", y="Fulltext_Average", color="darkorange")
        plt.title(f"Mean Fulltext-Based '{s.capitalize()}' Scores")
        plt.xticks(rotation=45,ha="right")
        plt.tight_layout()
        out2=os.path.join(out_dir,f"fulltext_{s}.png")
        try:
            plt.savefig(out2)
        except:
            pass
        plt.close()

###############################################################################
# 4) GEE scale computations
###############################################################################
def compute_pearson_scale(y,mu,df_resid):
    if df_resid<=0: return np.nan
    y=np.asarray(y)
    mu=np.asarray(mu)
    r=(y-mu)/np.sqrt(mu+1e-9)
    return np.sum(r**2)/df_resid

def compute_deviance_scale(y,mu,df_resid):
    if df_resid<=0: return np.nan
    arr=[]
    for obs,lam in zip(y,mu):
        if obs>0 and lam>0:
            arr.append(obs*math.log(obs/lam)-(obs-lam))
        elif obs==0:
            arr.append(-(obs-lam))
        else:
            arr.append(np.nan)
    dev=2*np.nansum(arr)
    return dev/df_resid if df_resid>0 else np.nan

def compute_ub_scale(y,mu,df_resid):
    p=compute_pearson_scale(y,mu,df_resid)
    if np.isnan(p): return p
    return 1.1*p

def compute_bc_scale(y,mu,df_resid):
    d=compute_deviance_scale(y,mu,df_resid)
    if np.isnan(d):
        return d
    return 0.9*d

def fit_and_compute_scales(model):
    base_res=model.fit(scale=None)
    base_qic=base_res.qic()
    y=base_res.model.endog
    mu=base_res.fittedvalues
    n=len(y)
    p=len(base_res.params)
    df_resid=n-p

    pear=compute_pearson_scale(y,mu,df_resid)
    dev=compute_deviance_scale(y,mu,df_resid)
    ubv=compute_ub_scale(y,mu,df_resid)
    bcv=compute_bc_scale(y,mu,df_resid)

    results={}
    if isinstance(base_qic, tuple):
        q_m,q_a=base_qic
    else:
        q_m,q_a=base_qic,None
    results["none"]=(q_m,q_a,None)

    from math import isnan
    for (nm,val) in [("pearson",pear),("deviance",dev),("ub",ubv),("bc",bcv)]:
        if (not isnan(val)) and (df_resid>0):
            re2=model.fit(scale=val)
            r_qic=re2.qic()
            if isinstance(r_qic,tuple):
                rm,ra=r_qic
            else:
                rm,ra=r_qic,None
            results[nm]=(rm,ra,val)
        else:
            results[nm]=(np.nan,None,val)
    return results

###############################################################################
# 5) Best QIC approach
###############################################################################
def run_gee_for_sentiment_measure_best_qic(df, sentiment, measure):
    d2=df.copy()
    if measure=="Quotation":
        pat=rf"^{re.escape(sentiment)}_\d+$"
        matched=[c for c in d2.columns if re.match(pat,c)]
        if not matched:
            return None
        d2["_score_col"]=d2[matched].clip(lower=0).mean(axis=1)
    else:
        fcol=f"{sentiment}_fulltext"
        if fcol not in d2.columns:
            return None
        d2["_score_col"]=d2[fcol].clip(lower=0)

    needed=["_score_col","media_outlet_clean","media_category"]
    d2=d2.dropna(subset=needed)
    if len(d2)<2 or d2["media_category"].nunique()<2:
        return None

    from statsmodels.genmod.generalized_estimating_equations import GEE
    best_qic_main=np.inf
    best_tuple=None
    combos=[]
    structures=[Independence(),Exchangeable()]
    for cov_obj in structures:
        st_name=cov_obj.__class__.__name__
        model=GEE.from_formula("_score_col ~ media_category",
            groups="media_outlet_clean",data=d2,
            family=Poisson(),cov_struct=cov_obj)
        sc_map=fit_and_compute_scales(model)
        for sm,(qm,qa,sc_val) in sc_map.items():
            combos.append({
                "Sentiment":sentiment,
                "Measure":measure,
                "Structure":st_name,
                "ScaleMethod":sm,
                "NumericScale":sc_val,
                "QIC_main":qm,
                "QIC_alt":qa
            })
            if (qm<best_qic_main) and (not np.isnan(qm)):
                best_qic_main=qm
                best_tuple=(st_name,sm,qm,sc_val)

    if best_tuple is None:
        return None
    stn, sm, qicv, scale_num=best_tuple
    return {
        "Sentiment":sentiment,
        "Measure":measure,
        "Best_Structure":stn,
        "Best_Scale":sm,
        "Best_QIC_main":qicv,
        "AllCombos":pd.DataFrame(combos),
        "Summary":f"Best scale={sm}, numeric={scale_num}"
    }

def run_gee_analyses_best_qic(df):
    logging.info("Running best QIC approach for each sentiment–measure.")
    best_list=[]
    all_combos=[]
    for s in CATEGORIES:
        for meas in ["Quotation","Fulltext"]:
            info=run_gee_for_sentiment_measure_best_qic(df,s,meas)
            if info is not None:
                best_list.append({
                    "Sentiment": info["Sentiment"],
                    "Measure": info["Measure"],
                    "Best_Structure": info["Best_Structure"],
                    "Best_Scale": info["Best_Scale"],
                    "Best_QIC_main": info["Best_QIC_main"],
                    "Summary": info["Summary"]
                })
                all_combos.append(info["AllCombos"])
    df_best=pd.DataFrame(best_list)
    df_all=pd.concat(all_combos,ignore_index=True) if len(all_combos)>0 else pd.DataFrame()
    return df_best, df_all

###############################################################################
# Additional checks: residuals, cross-val, sensitivity, bootstrap
###############################################################################
def check_residuals_and_correlation(final_fit):
    # same approach as before
    y=final_fit.model.endog
    mu=final_fit.fittedvalues
    pearson_res_arr = (y - mu) / np.sqrt(mu + 1e-9)
    pearson_res_arr = np.asarray(pearson_res_arr)

    mean_res=np.mean(pearson_res_arr)
    std_res=np.std(pearson_res_arr)

    clusters=final_fit.model.groups
    cluster_map={}
    for i, cid in enumerate(clusters):
        cluster_map.setdefault(cid,[]).append(pearson_res_arr[i])

    wcorr=[]
    for cid, arr in cluster_map.items():
        arr=np.array(arr)
        if len(arr)==1:
            continue
        elif len(arr)==2:
            cmat=np.corrcoef(arr)
            if cmat.shape==(2,2):
                wcorr.append(cmat[0,1])
        else:
            cmat=np.corrcoef(arr)
            n2=len(arr)
            sum_offdiag=np.sum(cmat)-n2
            avg_off=sum_offdiag/(n2*(n2-1))
            wcorr.append(avg_off)

    avg_corr=np.mean(wcorr) if len(wcorr)>0 else 0.0
    deviance=getattr(final_fit,"pearson_chi2",np.nan)
    dfres=getattr(final_fit,"df_resid",1)
    overdisp=np.nan
    if dfres>0 and not np.isnan(deviance):
        overdisp=deviance/dfres

    assess=""
    if abs(mean_res)>1:
        assess+="Mean Pearson residual>1 => misfit? "
    if std_res>2:
        assess+="Residual std>2 => outliers or overdisp. "
    if abs(avg_corr)>0.3:
        assess+=f"Within-cluster corr={avg_corr:.2f} => structure suspect. "
    if (not np.isnan(overdisp)) and overdisp>2:
        assess+=f"Overdisp={overdisp:.2f} => consider NB. "
    if assess=="":
        assess="No major issues from these checks."

    return {
        "mean_pearson_res": mean_res,
        "std_pearson_res": std_res,
        "avg_within_corr": avg_corr,
        "deviance": deviance,
        "df_resid": dfres,
        "overdisp_ratio": overdisp,
        "assessment": assess
    }


def pseudo_likelihood_check(final_fit):
    fam=final_fit.model.family
    if not isinstance(fam, Poisson):
        return {
            "NB_QIC": None,
            "Poisson_QIC": None,
            "diff_QIC": None,
            "conclusion": "Non-Poisson => skip NB"
        }
    data=final_fit.model.data.frame
    groups=final_fit.model.groups
    formula=final_fit.model.formula
    cov_struct=final_fit.model.cov_struct

    # Fit negative binomial
    nb_model=GEE.from_formula(formula, groups=groups, data=data,
                              family=NegativeBinomial(), cov_struct=cov_struct)
    nb_res=nb_model.fit()
    nb_qic=nb_res.qic()
    if isinstance(nb_qic, tuple):
        nb_qic=nb_qic[0]   # take first if tuple

    old_qic=final_fit.qic()
    if isinstance(old_qic, tuple):
        old_qic=old_qic[0] # take first if tuple

    diff_qic=None
    conclusion=""
    if (isinstance(nb_qic,(float,int))) and (isinstance(old_qic,(float,int))):
        diff_qic=old_qic-nb_qic
        conclusion="NegBin better" if diff_qic>0 else "No NB improvement"
    else:
        conclusion="Could not compare QIC (non-float)"

    return {
        "NB_QIC": nb_qic,
        "Poisson_QIC": old_qic,
        "diff_QIC": diff_qic,
        "conclusion": conclusion
    }

def cross_validation_gee(df, formula, group_col, family, cov_struct, n_folds=5):
    from statsmodels.genmod.generalized_estimating_equations import GEE
    cluster_ids=df[group_col].unique()
    np.random.shuffle(cluster_ids)
    folds=np.array_split(cluster_ids,n_folds)
    metrics=[]
    for i in range(n_folds):
        testc=set(folds[i])
        train_df=df[~df[group_col].isin(testc)]
        test_df=df[df[group_col].isin(testc)]
        if len(train_df)==0 or len(test_df)==0:
            continue
        mod=GEE.from_formula(formula, groups=group_col, data=train_df, family=family, cov_struct=cov_struct)
        res=mod.fit()
        pred=res.predict(test_df)
        obs=test_df[res.model.endog_names]
        mse=np.mean((obs-pred)**2)
        metrics.append(mse)
    avg_mse=np.mean(metrics) if len(metrics)>0 else np.nan
    return avg_mse

def sensitivity_analysis_correlation(df, formula, group_col,
                                     families=[Poisson(),NegativeBinomial()],
                                     structures=[Independence(),Exchangeable()]):
    from statsmodels.genmod.generalized_estimating_equations import GEE
    results=[]
    for fam in families:
        for covs in structures:
            mod=GEE.from_formula(formula, groups=group_col, data=df, family=fam, cov_struct=covs)
            res=mod.fit()
            qic=res.qic()
            if isinstance(qic, tuple):
                qic_val=qic[0]
            else:
                qic_val=qic
            sample_params=str(res.params.head(4).round(3).to_dict())
            results.append({
                "Family":fam.__class__.__name__,
                "Structure":covs.__class__.__name__,
                "QIC":qic_val,
                "ParamSample":sample_params
            })
    return pd.DataFrame(results)

def bootstrap_gee(df, formula, group_col, B=5, family=Poisson(), cov_struct=Independence()):
    from statsmodels.genmod.generalized_estimating_equations import GEE
    clusters=df[group_col].unique()
    param_records=[]
    for b in range(B):
        sample_c=np.random.choice(clusters,len(clusters),replace=True)
        pieces=[]
        for cid in sample_c:
            pieces.append(df[df[group_col]==cid])
        boot_df=pd.concat(pieces, ignore_index=True)
        mod=GEE.from_formula(formula, groups=group_col, data=boot_df,
                             family=family, cov_struct=cov_struct)
        fit_res=mod.fit()
        param_records.append(fit_res.params)
    param_df=pd.DataFrame(param_records)
    means=param_df.mean()
    sds=param_df.std()
    return pd.DataFrame({
        "Param": means.index,
        "BootMean": means.values,
        "BootStd": sds.values
    })

###############################################################################
# 6) Pairwise + "DiffIDs"
###############################################################################
def pairwise_and_diffIDs(df, sentiment, measure, struct, scale_name):
    from statsmodels.genmod.generalized_estimating_equations import GEE
    d2=df.copy()
    if measure=="Quotation":
        pat=rf"^{re.escape(sentiment)}_\d+$"
        matched=[c for c in d2.columns if re.match(pat,c)]
        if not matched:
            return None,None,None
        d2["_score_col"]=d2[matched].clip(lower=0).mean(axis=1)
    else:
        fcol=f"{sentiment}_fulltext"
        if fcol not in d2.columns:
            return None,None,None
        d2["_score_col"]=d2[fcol].clip(lower=0)

    needed=["_score_col","media_outlet_clean","media_category"]
    d2=d2.dropna(subset=needed)
    if len(d2)<2 or d2["media_category"].nunique()<2:
        return None,None,None

    d2["media_category"]=d2["media_category"].astype("category")
    if struct=="Independence":
        cov_obj=Independence()
    else:
        cov_obj=Exchangeable()

    model=GEE.from_formula("_score_col ~ media_category",
        groups="media_outlet_clean",data=d2,
        family=Poisson(),cov_struct=cov_obj)
    bres=model.fit(scale=None)

    y=bres.model.endog
    mu=bres.fittedvalues
    n=len(y)
    p=len(bres.params)
    dfresid=n-p
    from math import isnan
    if scale_name=="none":
        final_fit=bres
    else:
        if dfresid<=0:
            final_fit=bres
        else:
            if scale_name=="pearson":
                scv=compute_pearson_scale(y,mu,dfresid)
            elif scale_name=="deviance":
                scv=compute_deviance_scale(y,mu,dfresid)
            elif scale_name=="ub":
                scv=1.1*compute_pearson_scale(y,mu,dfresid)
            elif scale_name=="bc":
                scv=0.9*compute_deviance_scale(y,mu,dfresid)
            else:
                scv=None
            if scv is None or isnan(scv):
                final_fit=bres
            else:
                final_fit=model.fit(scale=scv)

    summary_txt=final_fit.summary().as_text()

    # Pairwise
    params=final_fit.params
    cov=final_fit.cov_params()
    mdf=final_fit.model.data.frame
    cats=mdf["media_category"].cat.categories
    ref=cats[0]
    idx_map={ref:0}
    for c in cats[1:]:
        nm=f"media_category[T.{c}]"
        idx_map[c]=final_fit.model.exog_names.index(nm)

    pair_list=[]
    for i in range(len(cats)):
        for j in range(i+1,len(cats)):
            ca,cb=cats[i], cats[j]
            con=np.zeros(len(params))
            if ca==ref and cb!=ref:
                con[idx_map[cb]]=-1.0
            elif cb==ref and ca!=ref:
                con[idx_map[ca]]=1.0
            else:
                con[idx_map[ca]]=1.0
                con[idx_map[cb]]= -1.0
            diff_est=con@params
            diff_var=con@cov@con
            diff_se=np.sqrt(diff_var)
            z=diff_est/diff_se
            pval=2*(1-norm.cdf(abs(z)))
            pair_list.append((ca,cb,diff_est,diff_se,z,pval))

    pair_df=pd.DataFrame(pair_list,columns=["CategoryA","CategoryB","Difference","SE","Z","p_value"])
    rej,p_adj,_,_=multipletests(pair_df["p_value"],method="fdr_bh")
    pair_df["p_value_adj"]=p_adj
    pair_df["reject_H0"]=rej

    # DiffIDs
    cat_list=list(cats)
    cat_index_map={cat_list[i]: i+1 for i in range(len(cat_list))}
    diff_map={c: set() for c in cat_list}
    for i,row in pair_df.iterrows():
        A=row["CategoryA"]
        B=row["CategoryB"]
        if row["reject_H0"]:
            diff_map[A].add(cat_index_map[B])
            diff_map[B].add(cat_index_map[A])

    rows=[]
    for c in cat_list:
        diffs=sorted(list(diff_map[c]))
        diffs_str=",".join(str(x) for x in diffs)
        rows.append((c,diffs_str))
    diffIDs_df=pd.DataFrame(rows,columns=["Category","DiffIDs"])

    return summary_txt, pair_df, diffIDs_df

###############################################################################
# 7) Compile
###############################################################################
def compile_results_into_multiple_workbooks(
    aggregated_df, stats_df, raw_df,
    df_best_qic, df_all_combos,
    plots_dir,
    main_excel, raw_excel, gee_excel, plots_excel, combined_excel,
    df_full
):
    with pd.ExcelWriter(main_excel, engine="openpyxl") as w:
        aggregated_df.to_excel(w, "Aggregated_Scores", index=False)
        stats_df.to_excel(w, "Mean_Median_Statistics", index=False)

    with pd.ExcelWriter(raw_excel, engine="openpyxl") as w:
        raw_df.to_excel(w, "Raw_Data", index=False)
        for s in CATEGORIES:
            s_cols=[c for c in raw_df.columns if c.startswith(s+"_")]
            s_df=raw_df[["media_category","media_outlet"]+s_cols].copy()
            sheet_name=f"Raw_{s[:29]}"
            s_df.to_excel(w, sheet_name, index=False)

    diag_records=[]
    cv_records=[]
    sens_records=[]
    boot_records=[]

    with pd.ExcelWriter(gee_excel, engine="openpyxl") as writer:
        idx_rows=[]
        for i, row in df_best_qic.iterrows():
            s=row["Sentiment"]
            meas=row["Measure"]
            st=row["Best_Structure"]
            sc=row["Best_Scale"]
            sh_name=f"BestQIC_{s[:10]}_{meas[:8]}"

            summary_txt, pair_df, diffIDs_df = pairwise_and_diffIDs(df_full, s, meas, st, sc)
            if summary_txt is None:
                tmp_df=pd.DataFrame({"Summary":[f"No valid model => {s}-{meas}"]})
                tmp_df.to_excel(writer, sh_name, index=False)
                continue

            lines=summary_txt.split("\n")
            sumdf=pd.DataFrame({"GEE_Summary":lines})
            sumdf.to_excel(writer, sh_name, index=False)

            sr=len(sumdf)+2
            pair_df.to_excel(writer, sh_name, index=False, startrow=sr)

            ws=writer.sheets[sh_name]
            for row_idx in range(len(pair_df)):
                if pair_df.loc[row_idx,"reject_H0"]:
                    excel_row=sr+1+row_idx
                    for col_idx in range(1, pair_df.shape[1]+1):
                        cell=ws.cell(row=excel_row+1, column=col_idx)
                        cell.fill=PatternFill(fill_type="solid", start_color="FFFF0000", end_color="FFFF0000")

            sr2=sr+len(pair_df)+2
            if diffIDs_df is not None and not diffIDs_df.empty:
                diffIDs_df.to_excel(writer, sh_name, index=False, startrow=sr2)

            # re-fit for diagnostics
            from statsmodels.genmod.generalized_estimating_equations import GEE
            d2=df_full.copy()
            if meas=="Quotation":
                pat=rf"^{re.escape(s)}_\d+$"
                matched=[c for c in d2.columns if re.match(pat,c)]
                if not matched:
                    idx_rows.append({"Sentiment":s,"Measure":meas,"SheetName":sh_name,"Structure":st,"Scale":sc,"BestQIC":row["Best_QIC_main"]})
                    continue
                d2["_score_col"]=d2[matched].clip(lower=0).mean(axis=1)
            else:
                fcol=f"{s}_fulltext"
                if fcol not in d2.columns:
                    idx_rows.append({"Sentiment":s,"Measure":meas,"SheetName":sh_name,"Structure":st,"Scale":sc,"BestQIC":row["Best_QIC_main"]})
                    continue
                d2["_score_col"]=d2[fcol].clip(lower=0)
            needed=["_score_col","media_outlet_clean","media_category"]
            d2=d2.dropna(subset=needed)
            if len(d2)<2 or d2["media_category"].nunique()<2:
                idx_rows.append({"Sentiment":s,"Measure":meas,"SheetName":sh_name,"Structure":st,"Scale":sc,"BestQIC":row["Best_QIC_main"]})
                continue
            d2["media_category"]=d2["media_category"].astype("category")

            if st=="Independence":
                cov_obj=Independence()
            else:
                cov_obj=Exchangeable()

            base_model=GEE.from_formula("_score_col ~ media_category",
                                        groups="media_outlet_clean",data=d2,
                                        family=Poisson(),cov_struct=cov_obj)
            bres=base_model.fit(scale=None)
            y=bres.model.endog
            mu=bres.fittedvalues
            n=len(y)
            p=len(bres.params)
            dfresid=n-p
            from math import isnan
            if sc=="none":
                final_fit=bres
            else:
                if dfresid<=0:
                    final_fit=bres
                else:
                    if sc=="pearson":
                        scv=compute_pearson_scale(y,mu,dfresid)
                    elif sc=="deviance":
                        scv=compute_deviance_scale(y,mu,dfresid)
                    elif sc=="ub":
                        scv=1.1*compute_pearson_scale(y,mu,dfresid)
                    elif sc=="bc":
                        scv=0.9*compute_deviance_scale(y,mu,dfresid)
                    else:
                        scv=None
                    if scv is None or isnan(scv):
                        final_fit=bres
                    else:
                        final_fit=base_model.fit(scale=scv)

            # gather checks
            diag=check_residuals_and_correlation(final_fit)
            ps=pseudo_likelihood_check(final_fit)

            diag_records.append({
                "Sentiment": s,
                "Measure": meas,
                "Structure": st,
                "Scale": sc,
                "mean_pearson_res": diag["mean_pearson_res"],
                "std_pearson_res": diag["std_pearson_res"],
                "avg_within_corr": diag["avg_within_corr"],
                "overdisp_ratio": diag["overdisp_ratio"],
                "assessment": diag["assessment"],
                "NB_QIC": ps["NB_QIC"],
                "Poisson_QIC": ps["Poisson_QIC"],
                "diff_QIC": ps["diff_QIC"],
                "pseudo_conclusion": ps["conclusion"]
            })

            # cross-val
            cv_val=cross_validation_gee(d2, "_score_col ~ media_category",
                                        "media_outlet_clean",
                                        Poisson(), cov_obj, n_folds=3)
            cv_records.append({
                "Sentiment": s,
                "Measure": meas,
                "Structure": st,
                "Scale": sc,
                "CV_MSE": cv_val
            })

            # sensitivity
            sens_df=sensitivity_analysis_correlation(d2, "_score_col ~ media_category","media_outlet_clean")
            sens_df["Sentiment"]=s
            sens_df["Measure"]=meas
            sens_records.append(sens_df)

            # bootstrap
            boot_df=bootstrap_gee(d2, "_score_col ~ media_category",
                                  "media_outlet_clean",B=5,
                                  family=Poisson(),cov_struct=cov_obj)
            boot_df["Sentiment"]=s
            boot_df["Measure"]=meas
            boot_records.append(boot_df)

            idx_rows.append({
                "Sentiment": s,
                "Measure": meas,
                "SheetName": sh_name,
                "Structure": st,
                "Scale": sc,
                "BestQIC": row["Best_QIC_main"]
            })

        idxdf=pd.DataFrame(idx_rows)
        idxdf.to_excel(writer,"BestQIC_Index",index=False)

        if not df_all_combos.empty:
            df_all_combos.to_excel(writer,"All_Combos",index=False)

        diag_df=pd.DataFrame(diag_records)
        diag_df.to_excel(writer,"Diagnostics",index=False)

        cv_df=pd.DataFrame(cv_records)
        cv_df.to_excel(writer,"CrossValidation_Res",index=False)

        if len(sens_records)>0:
            sens_all=pd.concat(sens_records,ignore_index=True)
            sens_all.to_excel(writer,"Sensitivity_Analysis",index=False)

        if len(boot_records)>0:
            boot_all=pd.concat(boot_records,ignore_index=True)
            boot_all.to_excel(writer,"Bootstrap_Res",index=False)

    # analysis_plots
    wb_plots=Workbook()
    if "Sheet" in wb_plots.sheetnames:
        wb_plots.remove(wb_plots["Sheet"])
    any_sheets=False
    for s in CATEGORIES:
        q_path=os.path.join(plots_dir,f"quote_{s}.png")
        if os.path.exists(q_path):
            st=f"Quote_{s[:28]}"
            ws=wb_plots.create_sheet(title=st)
            try:
                img=ExcelImage(q_path)
                img.anchor="A1"
                ws.add_image(img)
            except:
                pass
            any_sheets=True

        f_path=os.path.join(plots_dir,f"fulltext_{s}.png")
        if os.path.exists(f_path):
            st2=f"Fulltext_{s[:25]}"
            ws2=wb_plots.create_sheet(title=st2)
            try:
                img2=ExcelImage(f_path)
                img2.anchor="A1"
                ws2.add_image(img2)
            except:
                pass
            any_sheets=True

    cbar=os.path.join(plots_dir,"correlation_quotation_fulltext_bar.png")
    if os.path.exists(cbar):
        ws3=wb_plots.create_sheet("Correlation_Bar")
        try:
            ig3=ExcelImage(cbar)
            ig3.anchor="A1"
            ws3.add_image(ig3)
        except:
            pass
        any_sheets=True

    combp=os.path.join(plots_dir,"combined_normalized_scatter.png")
    if os.path.exists(combp):
        ws4=wb_plots.create_sheet("Combined_ZScatter")
        try:
            ig4=ExcelImage(combp)
            ig4.anchor="A1"
            ws4.add_image(ig4)
        except:
            pass
        any_sheets=True

    if not any_sheets:
        wb_plots.create_sheet("DummySheet")
    wb_plots.save(plots_excel)

    # analysis_combined
    raw_clean=raw_df.copy()
    raw_clean=raw_clean.applymap(lambda x:", ".join(x) if isinstance(x,list) else x)
    wb_comb=Workbook()
    if "Sheet" in wb_comb.sheetnames:
        wb_comb.remove(wb_comb["Sheet"])

    ws_agg=wb_comb.create_sheet("Aggregated_Scores")
    for r in dataframe_to_rows(aggregated_df,index=False,header=True):
        ws_agg.append(r)

    ws_stats=wb_comb.create_sheet("Mean_Median_Statistics")
    for r in dataframe_to_rows(stats_df,index=False,header=True):
        ws_stats.append(r)

    ws_raw=wb_comb.create_sheet("Raw_Data")
    for r in dataframe_to_rows(raw_clean,index=False,header=True):
        ws_raw.append(r)

    ws_best=wb_comb.create_sheet("BestQIC_Table")
    for r in dataframe_to_rows(df_best_qic,index=False,header=True):
        ws_best.append(r)

    wb_comb.save(combined_excel)


def main():
    setup_logging()
    logging.info("Starting best QIC GEE approach + 'DiffIDs' + extended checks + CombinedZScatter data.")

    df=load_jsonl(INPUT_JSONL_FILE)
    print(f"Total articles loaded: {len(df)}")

    df=map_media_outlet_to_category(df)
    chunk_and_save(df,20000)
    print_basic_stats(df)

    print("Performing Quotation vs Fulltext correlation analysis...")
    analyze_quotation_fulltext_correlation(df)

    print("Aggregating sentiment/emotion scores per media category...")
    agg_df=aggregate_sentiment_scores(df,CATEGORIES)
    agg_df=calculate_averages(agg_df)
    stats_df=calculate_mean_median(agg_df)
    save_aggregated_scores_to_csv(agg_df,CSV_OUTPUT_DIR)
    plot_statistics(agg_df,OUTPUT_DIR)

    print("Fitting best QIC approach for each sentiment–measure (with BH pairwise + DiffIDs).")
    df_best, df_allcombos = run_gee_analyses_best_qic(df)
    print("Best QIC approach completed.\n")

    compile_results_into_multiple_workbooks(
        aggregated_df=agg_df,
        stats_df=stats_df,
        raw_df=df,
        df_best_qic=df_best,
        df_all_combos=df_allcombos,
        plots_dir=OUTPUT_DIR,
        main_excel=OUTPUT_EXCEL_MAIN,
        raw_excel=OUTPUT_EXCEL_RAW,
        gee_excel=OUTPUT_EXCEL_GEE,
        plots_excel=OUTPUT_EXCEL_PLOTS,
        combined_excel=OUTPUT_EXCEL_COMBINED,
        df_full=df
    )

    print("Analysis completed successfully with extended methods + data for CombinedZScatter.")
    logging.info("Analysis completed successfully.")


if __name__=="__main__":
    main()
