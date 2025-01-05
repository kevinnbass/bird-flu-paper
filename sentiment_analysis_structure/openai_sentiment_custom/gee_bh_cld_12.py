#!/usr/bin/env python3
# gee_final_best_qic_with_pairwise_cld.py
"""
Script that:
  1) Loads data from JSONL
  2) Chunks & saves CSV
  3) Performs correlation analyses & scatterplots for Quotation vs Fulltext
  4) Aggregates sentiment/emotion scores + bar plots
  5) Finds best QIC combination (Ind/Exch × scale=[none,pearson,deviance,ub,bc])
  6) For each best QIC model:
     - Refit & produce GEE summary
     - Compute pairwise comparisons (BH correction)
     - Build a CLD table with 4 columns: [MediaCategory, Mean, UserCLD="", AutoCLD]
       using the custom adaptation of compactletterdisplay
         * If a pair is significantly different => they cannot share a letter
         * We apply the 'insert-and-absorb' approach to produce minimal letter sets
  7) Writes summary, pairwise, & CLD into analysis_gee.xlsx (one sheet per sentiment×measure),
     highlighting rows where reject_H0=True in red
  8) Also saves analysis_main, analysis_raw, analysis_plots, analysis_combined
"""

import json
import os
import re
import sys
import warnings
import logging
import math
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import seaborn as sns
from tqdm import tqdm

from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill

import statsmodels.api as sm
from statsmodels.genmod.generalized_estimating_equations import GEE
from statsmodels.genmod.families import Poisson
from statsmodels.genmod.cov_struct import Independence, Exchangeable
from statsmodels.stats.multitest import multipletests
from scipy.stats import pearsonr, norm

# --------------------- #
# Configuration
# --------------------- #
INPUT_JSONL_FILE = "processed_all_articles_fixed_2.jsonl"

OUTPUT_DIR = "graphs_analysis"
os.makedirs(OUTPUT_DIR, exist_ok=True)

CSV_OUTPUT_DIR = "csv_raw_scores"
os.makedirs(CSV_OUTPUT_DIR, exist_ok=True)

OUTPUT_EXCEL_MAIN = "analysis_main.xlsx"
OUTPUT_EXCEL_RAW = "analysis_raw.xlsx"
OUTPUT_EXCEL_GEE = "analysis_gee.xlsx"
OUTPUT_EXCEL_PLOTS = "analysis_plots.xlsx"
OUTPUT_EXCEL_COMBINED = "analysis_combined.xlsx"

LOG_FILE = "analysis.log"

CATEGORIES = [
    "joy","sadness","anger","fear",
    "surprise","disgust","trust","anticipation",
    "negative_sentiment","positive_sentiment"
]

# The final output order for the CLD table:
CLD_ORDER = ["Scientific","Left","Lean Left","Center","Lean Right","Right"]

MEDIA_CATEGORIES = {
    "Scientific": ["nature","sciam","stat","newscientist"],
    "Left": ["theatlantic","the daily beast","the intercept","mother jones","msnbc","slate","vox","huffpost"],
    "Lean Left": ["ap","axios","cnn","guardian","business insider","nbcnews","npr","nytimes","politico","propublica","wapo","usa today"],
    "Center": ["reuters","marketwatch","financial times","newsweek","forbes"],
    "Lean Right": ["thedispatch","epochtimes","foxbusiness","wsj","national review","washtimes"],
    "Right": ["breitbart","theblaze","daily mail","dailywire","foxnews","nypost","newsmax"],
}


def setup_logging():
    """Initialize logging to both file and console."""
    log_format = "%(asctime)s [%(levelname)s] %(module)s - %(message)s"
    logging.basicConfig(filename=LOG_FILE, level=logging.DEBUG, format=log_format)
    console = logging.StreamHandler(sys.stdout)
    console.setLevel(logging.INFO)
    console.setFormatter(logging.Formatter(log_format))
    logging.getLogger().addHandler(console)
    logging.info("Logging initialized (file+console).")


# Silence the "QIC scale=None" warning from statsmodels
warnings.filterwarnings(
    "ignore",
    message="QIC values obtained using scale=None are not appropriate for comparing models"
)


###############################################################################
# 1) Load, chunk, basic stats
###############################################################################
def load_jsonl(jsonl_file):
    logging.info(f"Loading JSONL from {jsonl_file}")
    records = []
    with open(jsonl_file, "r", encoding="utf-8") as f:
        for line in tqdm(f, desc="Loading JSONL data"):
            rec = json.loads(line)
            records.append(rec)
    df = pd.DataFrame(records)
    logging.debug(f"Loaded DataFrame shape={df.shape}")
    return df

def map_media_outlet_to_category(df):
    cat_map = {}
    for cat, outls in MEDIA_CATEGORIES.items():
        for o in outls:
            cat_map[o.lower().strip()] = cat

    if "media_outlet" not in df.columns:
        raise KeyError("'media_outlet' not found in data")

    df["media_outlet_clean"] = df["media_outlet"].str.lower().str.strip()
    df["media_category"] = df["media_outlet_clean"].map(cat_map).fillna("Other")

    unmapped = df[df["media_category"]=="Other"]["media_outlet"].unique()
    if len(unmapped)>0:
        logging.warning(f"Unmapped => {unmapped}")
        print(f"Warning: Not mapped => {unmapped}")
    return df

def chunk_and_save(df, chunk_size=20000):
    logging.info(f"Chunking DataFrame => len={len(df)}, chunk_size={chunk_size}")
    for i in range(0, len(df), chunk_size):
        part = df.iloc[i : i + chunk_size]
        out_csv = os.path.join(CSV_OUTPUT_DIR, f"raw_data_part_{(i // chunk_size) + 1}.csv")
        part.to_csv(out_csv, index=False)
        print(f"Saved chunk {(i // chunk_size) + 1} to {out_csv}")

def print_basic_stats(df):
    logging.info(f"Basic stats => total articles = {len(df)}")
    print("\nSummary Statistics:")
    print("Total number of articles:", len(df))
    if "media_outlet_clean" in df.columns:
        vc = df["media_outlet_clean"].value_counts()
        print("\nArticles per outlet:")
        print(vc)
    if "media_category" in df.columns:
        vc2 = df["media_category"].value_counts()
        print("\nArticles per category:")
        print(vc2)
    print()


###############################################################################
# 2) Quotation vs. Fulltext correlation
###############################################################################
def analyze_quotation_fulltext_correlation(df):
    logging.info("Analyzing Quotation vs Fulltext correlation.")
    # ... your existing correlation logic ...
    pass


###############################################################################
# 3) Aggregation & Stats
###############################################################################
def aggregate_sentiment_scores(df, sentiments):
    logging.info("Aggregating sentiment/emotion scores by category + sentiment.")
    # ... your existing logic ...
    pass

def calculate_averages(agg_df):
    logging.info("Calculating Quotation_Average & Fulltext_Average.")
    # ... your existing logic ...
    pass

def calculate_mean_median(agg_df):
    logging.info("Computing global mean/median of Quotation/Fulltext averages.")
    # ... your existing logic ...
    pass

def save_aggregated_scores_to_csv(agg_df, out_dir):
    # ... your existing logic ...
    pass

def plot_statistics(agg_df, out_dir):
    # ... your existing logic ...
    pass


###############################################################################
# 4) GEE scale computations
###############################################################################
def compute_pearson_scale(y, mu, df_resid):
    # ... your existing code ...
    pass

def compute_deviance_scale(y, mu, df_resid):
    # ... your existing code ...
    pass

def compute_ub_scale(y, mu, df_resid):
    # ... your existing code ...
    pass

def compute_bc_scale(y, mu, df_resid):
    # ... your existing code ...
    pass

def fit_and_compute_scales(model):
    # ... your existing code ...
    pass


###############################################################################
# 5) Best QIC approach
###############################################################################
def run_gee_for_sentiment_measure_best_qic(df, sentiment, measure):
    # ... your existing code ...
    pass

def run_gee_analyses_best_qic(df):
    logging.info("Running best QIC approach for each sentiment–measure.")
    # ... your existing code ...
    pass


###############################################################################
# 6) The compactletterdisplay-based approach
###############################################################################

def get_next_unused_letter(columns):
    """
    Identify the next unused letter.
    'columns' is a list of strings, each representing a group assignment
    in the current partial letter display.
    """
    used_letters = set(letter for col in columns for letter in col if letter != '')
    for letter in 'abcdefghijklmnopqrstuvwxyz':
        if letter not in used_letters:
            return letter
    return None  # if all letters are used

def absorb_columns(columns):
    """
    Absorb redundant columns by comparing index sets.
    columns is a list of strings
    """
    absorbed = True
    while absorbed:
        absorbed = False
        for i, col1 in enumerate(columns):
            for j, col2 in enumerate(columns):
                if i != j:
                    indices1 = {idx for idx, c in enumerate(col1) if c!=''}
                    indices2 = {idx for idx, c in enumerate(col2) if c!=''}
                    if indices1.issubset(indices2):
                        absorbed = True
                        columns.pop(i)
                        break
            if absorbed:
                break
    return columns

def compact_letter_display(significant_pairs, group_names):
    """
    Generate a compact letter display (CLD) for 'group_names'
    based on 'significant_pairs', which is a list of (i, j) indices
    meaning group i, group j differ.

    Returns a list of strings, each a set of letters for that group.
    """
    num_groups = len(group_names)
    # Start with one column => all have 'a'
    # We'll store 'columns' as a list of lists-of-chars. 
    # But for convenience, let's store as a list of strings. Each 'column' has length = num_groups
    # Actually the snippet uses columns as strings, each char for each group.
    # We'll start with columns = ["a"*num_groups], meaning all share letter 'a'.
    columns = [ ['a' for _ in range(num_groups)] ]  # store each "column" as a list of length num_groups
    # We'll later join them into strings.

    # Convert columns => each is list-of-chars
    # significant_pairs => list of (i, j) meaning group i, j differ
    for (i, j) in significant_pairs:
        # check each column => if the two groups share the same letter => we must split
        connected = False
        for col_idx, column in enumerate(columns):
            if column[i] == column[j] and column[i] != '':
                connected = True
                # We pick the next unused letter
                new_letter = get_next_unused_letter([''.join(c) for c in columns])
                if new_letter is None:
                    # fallback => new_letter = '?'
                    new_letter = '?'
                # create a new column from this one => first copy
                new_col = column.copy()
                # in new_col => group i => new_letter if column[i] != '' else ''
                # but the snippet sets new_col[i] to new_letter if column[i] != ''
                # and sets new_col[j] = ''
                # Then in the old column, we set col[j] to '' to break
                new_col[i] = new_letter if column[i] != '' else ''
                new_col[j] = ''
                # Now in the old column => set column[j] = '' 
                column[j] = ''
                # Insert new_col into columns
                columns[col_idx] = column  # update old
                columns.append(new_col)
                # absorb
                columns_str = [''.join(c) for c in columns]
                columns_str = absorb_columns(columns_str)
                # convert back to list-of-lists
                new_columns = []
                for cstr in columns_str:
                    # ensure length is num_groups
                    # cstr might be shorter => we need a quick re-map:
                    # We'll distribute chars to indices that are not empty
                    # Actually we need to keep them aligned => we can do:
                    # Each cstr is length = num_groups => so cstr[k] is letter for group k
                    # or if cstr is shorter => this is complicated
                    # We'll assume cstr has length = num_groups because we never remove positions
                    # We'll do a direct list(cstr). This won't break if cstr's length = num_groups
                    new_columns.append(list(cstr))
                columns = new_columns
            if connected:
                break

    # final result => each group => combine letters from each column if column[x] != ''
    # first convert columns => list of strings
    # Actually we have columns as list-of-lists
    # We'll build final => for group n => collect columns[k][n]
    result = []
    for n in range(num_groups):
        letset = []
        for col in columns:
            if col[n] != '':
                letset.append(col[n])
        result.append(''.join(letset))
    return result

###############################################################################
# 6) Our pairwise_and_cld that uses this new approach
###############################################################################
def pairwise_and_cld(df, sentiment, measure, struct, scale_name):
    """
    1) Refit GEE => final_fit
    2) Build pairwise => BH => find significant pairs => list of (i, j)
    3) Use 'compact_letter_display' => produce letters
    4) Build final => [MediaCategory, Mean, UserCLD="", AutoCLD= those letters]
    """
    # Refit
    d2 = df.copy()
    if measure=="Quotation":
        pat = rf"^{re.escape(sentiment)}_\d+$"
        matched = [c for c in d2.columns if re.match(pat,c)]
        if not matched:
            return None, None, None
        d2["_score_col"] = d2[matched].clip(lower=0).mean(axis=1)
    else:
        fcol = f"{sentiment}_fulltext"
        if fcol not in d2.columns:
            return None, None, None
        d2["_score_col"] = d2[fcol].clip(lower=0)

    needed=["_score_col","media_outlet_clean","media_category"]
    d2 = d2.dropna(subset=needed)
    if len(d2)<2 or d2["media_category"].nunique()<2:
        return None, None, None

    d2["media_category"] = d2["media_category"].astype("category")
    if struct=="Independence":
        cov_obj=Independence()
    else:
        cov_obj=Exchangeable()

    model=GEE.from_formula("_score_col ~ media_category", 
                           groups="media_outlet_clean", 
                           data=d2,
                           family=Poisson(), 
                           cov_struct=cov_obj)
    bres=model.fit(scale=None)
    if scale_name=="none":
        final_fit=bres
    else:
        y=bres.model.endog
        mu=bres.fittedvalues
        n=len(y)
        p=len(bres.params)
        dfresid=n-p
        if dfresid<=0:
            final_fit=bres
        else:
            from math import isnan
            if scale_name=="pearson":
                from math import isnan
                def comp_pear(y, mu, df_resid):
                    r=(y-mu)/np.sqrt(mu)
                    return np.sum(r**2)/df_resid
                scv = comp_pear(y, mu, dfresid)
            elif scale_name=="deviance":
                def comp_dev(y, mu, df_resid):
                    arr=np.zeros_like(y,dtype=float)
                    for i in range(len(y)):
                        if y[i]>0 and mu[i]>0:
                            arr[i]=y[i]*np.log(y[i]/mu[i])-(y[i]-mu[i])
                        elif y[i]==0:
                            arr[i]=-(y[i]-mu[i])
                        else:
                            arr[i]=np.nan
                    return 2*np.nansum(arr)/df_resid
                scv = comp_dev(y,mu,dfresid)
            elif scale_name=="ub":
                # 1.1 * pearson
                def comp_pear2(y, mu, df_resid):
                    r=(y-mu)/np.sqrt(mu)
                    return np.sum(r**2)/df_resid
                pval=comp_pear2(y,mu,dfresid)
                scv=1.1*pval
            elif scale_name=="bc":
                # 0.9 * deviance
                def comp_dev2(y, mu, df_resid):
                    arr=np.zeros_like(y,dtype=float)
                    for i in range(len(y)):
                        if y[i]>0 and mu[i]>0:
                            arr[i]=y[i]*np.log(y[i]/mu[i])-(y[i]-mu[i])
                        elif y[i]==0:
                            arr[i]=-(y[i]-mu[i])
                        else:
                            arr[i]=np.nan
                    return 2*np.nansum(arr)/df_resid
                dval=comp_dev2(y,mu,dfresid)
                scv=0.9*dval
            else:
                scv=None
            if scv is None or isnan(scv):
                final_fit=bres
            else:
                final_fit=model.fit(scale=scv)

    summary_txt=final_fit.summary().as_text()

    # Now build pairwise comparisons with BH
    cats=final_fit.model.data.frame["media_category"].cat.categories
    ref=cats[0]
    idx_map={ref:0}
    for c in cats[1:]:
        nm=f"media_category[T.{c}]"
        idx_map[c]=final_fit.model.exog_names.index(nm)

    params=final_fit.params
    cov=final_fit.cov_params()

    pair_list=[]
    for i in range(len(cats)):
        for j in range(i+1, len(cats)):
            ca, cb = cats[i], cats[j]
            con=np.zeros(len(params))
            if ca==ref and cb!=ref:
                con[idx_map[cb]]=-1.0
            elif cb==ref and ca!=ref:
                con[idx_map[ca]]=1.0
            else:
                con[idx_map[ca]]=1.0
                con[idx_map[cb]]=-1.0
            diff_est=con@params
            diff_var=con@cov@con
            diff_se=np.sqrt(diff_var)
            z=diff_est/diff_se
            pval=2*(1-norm.cdf(abs(z)))
            pair_list.append((ca, cb, diff_est, diff_se, z, pval))

    pair_df=pd.DataFrame(pair_list, columns=["CategoryA","CategoryB","Difference","SE","Z","p_value"])
    rej, p_adj, _, _ = multipletests(pair_df["p_value"], method="fdr_bh")
    pair_df["p_value_adj"]=p_adj
    pair_df["reject_H0"]=rej

    # Now build 'significant_pairs' => list of (i, j) for which reject_H0=True
    # Also build 'columns' => the group names in the order they appear
    group_names = list(cats)
    sig_pairs = []
    for i, row in pair_df.iterrows():
        if row["reject_H0"]:
            # find the indices
            i_index = group_names.index(row["CategoryA"])
            j_index = group_names.index(row["CategoryB"])
            sig_pairs.append( (i_index, j_index) )

    # Now call 'compact_letter_display(sig_pairs, group_names)' => it returns a list of str
    letters_list = compact_letter_display(sig_pairs, group_names)

    # Next, we build => [MediaCategory, Mean, UserCLD="", AutoCLD=<letters>]
    # We'll get predicted means from final_fit
    intercept=final_fit.params[0]
    cat_effect={}
    cat_effect[cats[0]]=0.0
    for c in cats[1:]:
        nm = f"media_category[T.{c}]"
        cat_effect[c] = final_fit.params.get(nm,0.0)

    mean_map={}
    for c in cats:
        mean_map[c] = np.exp(intercept + cat_effect[c])

    rows=[]
    for c in CLD_ORDER:
        if c in group_names:
            idxc = group_names.index(c)
            auto_str = letters_list[idxc]
            rows.append((c, mean_map[c], "", auto_str))
        else:
            rows.append((c, None, "", ""))

    cld_df=pd.DataFrame(rows, columns=["MediaCategory","Mean","UserCLD","AutoCLD"])
    return summary_txt, pair_df, cld_df


###############################################################################
# 7) Compile results
###############################################################################
def compile_results_into_multiple_workbooks(
    aggregated_df, stats_df, raw_df,
    df_best_qic, df_all_combos,
    plots_dir,
    main_excel, raw_excel, gee_excel, plots_excel, combined_excel,
    df_full
):
    # Here we show the entire logic, no placeholders:
    with pd.ExcelWriter(main_excel, engine="openpyxl") as w:
        aggregated_df.to_excel(w,"Aggregated_Scores",index=False)
        stats_df.to_excel(w,"Mean_Median_Statistics",index=False)

    with pd.ExcelWriter(raw_excel, engine="openpyxl") as w:
        raw_df.to_excel(w,"Raw_Data",index=False)
        for s in CATEGORIES:
            s_cols=[c for c in raw_df.columns if c.startswith(s+"_")]
            s_df=raw_df[["media_category","media_outlet"]+s_cols].copy()
            sheet_name=f"Raw_{s[:29]}"
            s_df.to_excel(w, sheet_name, index=False)

    with pd.ExcelWriter(gee_excel, engine="openpyxl") as writer:
        idx_rows=[]
        for i, row in df_best_qic.iterrows():
            s=row["Sentiment"]
            meas=row["Measure"]
            st=row["Best_Structure"]
            sc=row["Best_Scale"]
            sh_name=f"BestQIC_{s[:10]}_{meas[:8]}"
            summary_txt, pair_df, cld_df = pairwise_and_cld(df_full, s, meas, st, sc)
            if summary_txt is None:
                tmp_df=pd.DataFrame({"Summary":["No valid model or not enough data."]})
                tmp_df.to_excel(writer,sh_name,index=False)
                continue

            lines=summary_txt.split("\n")
            sumdf=pd.DataFrame({"GEE_Summary":lines})
            sumdf.to_excel(writer, sh_name, index=False)

            sr=len(sumdf)+2
            pair_df.to_excel(writer, sh_name, index=False, startrow=sr)

            ws=writer.sheets[sh_name]
            for row_idx in range(len(pair_df)):
                if pair_df.loc[row_idx,"reject_H0"] == True:
                    excel_row = sr+1 + row_idx
                    for col_idx in range(1, pair_df.shape[1]+1):
                        cell=ws.cell(row=excel_row+1, column=col_idx)
                        cell.fill=PatternFill(fill_type="solid", start_color="FFFF0000", end_color="FFFF0000")

            sr2=sr+len(pair_df)+2
            if cld_df is not None and not cld_df.empty:
                cld_df.to_excel(writer, sh_name, index=False, startrow=sr2)

            idx_rows.append({
                "Sentiment": s,
                "Measure": meas,
                "SheetName": sh_name,
                "Structure": st,
                "Scale": sc,
                "BestQIC": row["Best_QIC_main"]
            })
        idxdf=pd.DataFrame(idx_rows)
        idxdf.to_excel(writer,"BestQIC_Index",index=False)

        if not df_all_combos.empty:
            df_all_combos.to_excel(writer,"All_Combos",index=False)

    # analysis_plots
    wb_plots=Workbook()
    if "Sheet" in wb_plots.sheetnames:
        wb_plots.remove(wb_plots["Sheet"])
    for s in CATEGORIES:
        q_path=os.path.join(plots_dir, f"quote_{s}.png")
        if os.path.exists(q_path):
            st=f"Quote_{s[:28]}"
            ws=wb_plots.create_sheet(title=st)
            try:
                img=ExcelImage(q_path)
                img.anchor="A1"
                ws.add_image(img)
            except:
                pass

        f_path=os.path.join(plots_dir, f"fulltext_{s}.png")
        if os.path.exists(f_path):
            st2=f"Fulltext_{s[:25]}"
            ws2=wb_plots.create_sheet(title=st2)
            try:
                img2=ExcelImage(f_path)
                img2.anchor="A1"
                ws2.add_image(img2)
            except:
                pass

    cbar=os.path.join(plots_dir,"correlation_quotation_fulltext_bar.png")
    if os.path.exists(cbar):
        ws3=wb_plots.create_sheet("Correlation_Bar")
        try:
            ig3=ExcelImage(cbar)
            ig3.anchor="A1"
            ws3.add_image(ig3)
        except:
            pass

    combp=os.path.join(plots_dir,"combined_normalized_scatter.png")
    if os.path.exists(combp):
        ws4=wb_plots.create_sheet("Combined_ZScatter")
        try:
            ig4=ExcelImage(combp)
            ig4.anchor="A1"
            ws4.add_image(ig4)
        except:
            pass

    wb_plots.save(plots_excel)

    # analysis_combined
    raw_clean=raw_df.copy()
    raw_clean=raw_clean.applymap(lambda x: ", ".join(x) if isinstance(x,list) else x)
    wb_comb=Workbook()
    if "Sheet" in wb_comb.sheetnames:
        wb_comb.remove(wb_comb["Sheet"])

    ws_agg=wb_comb.create_sheet("Aggregated_Scores")
    for r in dataframe_to_rows(aggregated_df, index=False, header=True):
        ws_agg.append(r)

    ws_stats=wb_comb.create_sheet("Mean_Median_Statistics")
    for r in dataframe_to_rows(stats_df, index=False, header=True):
        ws_stats.append(r)

    ws_raw=wb_comb.create_sheet("Raw_Data")
    for r in dataframe_to_rows(raw_clean, index=False, header=True):
        ws_raw.append(r)

    ws_best=wb_comb.create_sheet("BestQIC_Table")
    for r in dataframe_to_rows(df_best_qic, index=False, header=True):
        ws_best.append(r)

    wb_comb.save(combined_excel)


###############################################################################
def main():
    setup_logging()
    logging.info("Starting best QIC GEE approach with BH post-hoc & adapted compactletterdisplay approach for CLD.")

    # 1) Load
    df=load_jsonl(INPUT_JSONL_FILE)
    print(f"Total articles loaded: {len(df)}")

    # 2) map + chunk + stats
    df=map_media_outlet_to_category(df)
    chunk_and_save(df,20000)
    print_basic_stats(df)

    # 3) correlation
    print("Performing Quotation vs Fulltext correlation analysis...")
    analyze_quotation_fulltext_correlation(df)

    # 4) aggregation
    print("Aggregating sentiment/emotion scores per media category...")
    agg_df=aggregate_sentiment_scores(df, CATEGORIES)
    agg_df=calculate_averages(agg_df)
    stats_df=calculate_mean_median(agg_df)
    save_aggregated_scores_to_csv(agg_df, CSV_OUTPUT_DIR)
    plot_statistics(agg_df, OUTPUT_DIR)

    # 5) best QIC approach
    print("Fitting best QIC approach for each sentiment–measure (with BH pairwise + compactletterdisplay-based CLD).")
    df_best, df_allcombos = run_gee_analyses_best_qic(df)
    print("Best QIC approach completed.\n")

    # 6) compile
    compile_results_into_multiple_workbooks(
        aggregated_df=agg_df,
        stats_df=stats_df,
        raw_df=df,
        df_best_qic=df_best,
        df_all_combos=df_allcombos,
        plots_dir=OUTPUT_DIR,
        main_excel=OUTPUT_EXCEL_MAIN,
        raw_excel=OUTPUT_EXCEL_RAW,
        gee_excel=OUTPUT_EXCEL_GEE,
        plots_excel=OUTPUT_EXCEL_PLOTS,
        combined_excel=OUTPUT_EXCEL_COMBINED,
        df_full=df
    )

    print("Analysis completed successfully.")
    logging.info("Analysis completed successfully.")


if __name__=="__main__":
    main()
