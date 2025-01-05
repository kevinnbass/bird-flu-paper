#!/usr/bin/env python3
# generate_50tabs_6cols_no_blanks.py
"""
Creates 50 tabs (10 sentiments x 5 measures) in a single XLSX.
Each tab has 6 columns: [Scientific, Left, Lean Left, Center, Lean Right, Right].
No blank rows in each column => each column is a consecutive vertical list of measure values.

Quotation & Quotation_Intensity => average across matched columns (e.g. joy_0, joy_1).
Fulltext, Fulltext_Intensity, Title_Intensity => direct from single relevant column.
"""

import json
import re
import numpy as np
import pandas as pd
from tqdm import tqdm
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# ------------------ CONFIG ------------------
INPUT_JSONL_FILE = "processed_all_articles_fixed_4.jsonl"
OUTPUT_XLSX_FILE = "raw_values_6cat_50tabs_no_blanks.xlsx"

# 6 categories
MEDIA_CATEGORIES = {
    "Scientific": ["nature","sciam","stat","newscientist"],
    "Left": ["theatlantic","the daily beast","the intercept","mother jones","msnbc","slate","vox","huffpost"],
    "Lean Left": ["ap","axios","cnn","guardian","business insider","nbcnews",
                  "npr","nytimes","politico","propublica","wapo","usa today"],
    "Center": ["reuters","marketwatch","financial times","newsweek","forbes"],
    "Lean Right": ["thedispatch","epochtimes","foxbusiness","wsj","national review","washtimes"],
    "Right": ["breitbart","theblaze","daily mail","dailywire","foxnews","nypost","newsmax"],
}

SENTIMENTS = [
    "joy","sadness","anger","fear",
    "surprise","disgust","trust","anticipation",
    "negative_sentiment","positive_sentiment"
]

MEASURES = [
    "Fulltext",
    "Quotation",
    "Fulltext_Intensity",
    "Quotation_Intensity",
    "Title_Intensity",
]

# ------------------ STEP 1: LOAD + MAP ------------------
def load_jsonl(jsonl_file):
    """Load records from JSONL => DataFrame."""
    records = []
    with open(jsonl_file, "r", encoding="utf-8") as f:
        for line in tqdm(f, desc="Loading JSONL"):
            rec = json.loads(line)
            records.append(rec)
    return pd.DataFrame(records)

def map_outlet_to_6cats(df):
    """Map each media_outlet to one of the 6 categories => 'media_category_6' column."""
    cat_map = {}
    for cat, outls in MEDIA_CATEGORIES.items():
        for o in outls:
            cat_map[o.lower().strip()] = cat

    df2 = df.copy()
    df2["media_outlet_clean"] = df2["media_outlet"].str.lower().str.strip()
    df2["media_category_6"]   = df2["media_outlet_clean"].map(cat_map).fillna("Other")
    return df2

# ------------------ STEP 2: EXTRACT MEASURE VALUES ------------------
def extract_measure_value(row, sentiment, measure):
    """
    Return numeric measure for this article row, for the given sentiment.
    - Quotation => average across columns ^{sentiment}_\d+$
    - Quotation_Intensity => average across columns ^{sentiment}_\d+_intensity$
    - Fulltext => row[f"{sentiment}_fulltext"]
    - Fulltext_Intensity => row[f"{sentiment}_fulltext_intensity"]
    - Title_Intensity => row[f"{sentiment}_title_intensity"]
    """
    try:
        if measure == "Quotation":
            pat = rf"^{re.escape(sentiment)}_\d+$"
            matched = [c for c in row.index if re.match(pat, c)]
            if not matched:
                return np.nan
            vals = row[matched].astype(float).clip(lower=0)
            return vals.mean(skipna=True)

        elif measure == "Quotation_Intensity":
            pat = rf"^{re.escape(sentiment)}_\d+_intensity$"
            matched = [c for c in row.index if re.match(pat, c)]
            if not matched:
                return np.nan
            vals = row[matched].astype(float).clip(lower=0)
            return vals.mean(skipna=True)

        elif measure == "Fulltext":
            col = f"{sentiment}_fulltext"
            if col not in row.index:
                return np.nan
            return float(row[col])

        elif measure == "Fulltext_Intensity":
            col = f"{sentiment}_fulltext_intensity"
            if col not in row.index:
                return np.nan
            return float(row[col])

        elif measure == "Title_Intensity":
            col = f"{sentiment}_title_intensity"
            if col not in row.index:
                return np.nan
            return float(row[col])

        else:
            return np.nan

    except:
        return np.nan

# ------------------ STEP 3: MAIN => 50 TABS => NO BLANK ROWS ------------------
def main():
    df = load_jsonl(INPUT_JSONL_FILE)
    df = map_outlet_to_6cats(df)

    # If no article_id, define
    if "article_id" not in df.columns:
        df["article_id"] = df.index + 1

    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    # We'll create 50 tabs => 10 sentiments × 5 measures
    for sentiment in SENTIMENTS:
        for measure in MEASURES:
            # short tab name to fit Excel's 31-char limit
            tab_name = f"{sentiment[:10]}_{measure[:10]}"

            # 6 columns => each is "a vertical stream of numbers" for that category
            # We'll gather the measure values for each category => no blanks in that col
            # E.g. col "Scientific" has measure values for all articles that are scientific
            cat_lists = {
                "Scientific": [],
                "Left": [],
                "Lean Left": [],
                "Center": [],
                "Lean Right": [],
                "Right": []
            }

            # Loop over all articles
            for i, row in df.iterrows():
                cat_6 = row["media_category_6"]
                if cat_6 not in cat_lists:
                    # not in the 6 => skip
                    continue
                val = extract_measure_value(row, sentiment, measure)
                # If we want to skip NaNs, we can still store them, or skip them:
                if pd.notna(val):
                    cat_lists[cat_6].append(val)
                # else => skip => no blank row

            # Now we'll build a DataFrame with 6 columns, each being consecutive values
            # We'll find the max length among them to see how many rows to store
            max_len = max(len(lst) for lst in cat_lists.values())

            out_cols = ["Scientific","Left","Lean Left","Center","Lean Right","Right"]
            data_dict = {}
            for c in out_cols:
                # pad with np.nan if shorter
                col_vals = cat_lists[c]
                # if we want exactly no blank rows, just store them at the top
                # we'll pad the bottom with np.nan
                # so each column has length = max_len
                if len(col_vals) < max_len:
                    col_vals = col_vals + [np.nan]*(max_len - len(col_vals))
                data_dict[c] = col_vals

            result_df = pd.DataFrame(data_dict, columns=out_cols)

            # Write into an Excel sheet
            ws = wb.create_sheet(title=tab_name)
            # We'll add a header row => same as out_cols
            df_rows = [out_cols]  # first row => col names
            # Then add each row
            for row_idx in range(max_len):
                row_list = []
                for c in out_cols:
                    row_list.append(result_df.loc[row_idx, c])
                df_rows.append(row_list)

            # Now write these rows into ws
            for r_idx, row_vals in enumerate(df_rows, 1):
                for c_idx, val in enumerate(row_vals, 1):
                    ws.cell(row=r_idx, column=c_idx, value=val)

    wb.save(OUTPUT_XLSX_FILE)
    print(f"Wrote => {OUTPUT_XLSX_FILE} with {len(SENTIMENTS)*len(MEASURES)} tabs")

if __name__=="__main__":
    main()
