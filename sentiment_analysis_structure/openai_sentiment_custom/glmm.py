import json
import pandas as pd
import matplotlib.pyplot as plt
import seaborn as sns
import os
import logging
from tqdm import tqdm
from openpyxl.drawing.image import Image as ExcelImage
import re
import sys
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import numpy as np

import bambi as bmb
import arviz as az

# ------------------------------ #
#        Configuration           #
# ------------------------------ #

INPUT_JSONL_FILE = 'processed_all_articles_with_fulltext_sentiment_analysis.jsonl'
OUTPUT_DIR = 'graphs_analysis'
os.makedirs(OUTPUT_DIR, exist_ok=True)

CSV_OUTPUT_DIR = 'csv_raw_scores'
os.makedirs(CSV_OUTPUT_DIR, exist_ok=True)

OUTPUT_EXCEL_MAIN = 'analysis_main.xlsx'
OUTPUT_EXCEL_RAW = 'analysis_raw.xlsx'
OUTPUT_EXCEL_LMM = 'analysis_bambi.xlsx'  # Using Bambi for GLMM now
OUTPUT_EXCEL_PLOTS = 'analysis_plots.xlsx'
OUTPUT_EXCEL_COMBINED = 'analysis_combined.xlsx'

CATEGORIES = [
    'joy', 'sadness', 'anger', 'fear',
    'surprise', 'disgust', 'trust', 'anticipation',
    'negative_sentiment', 'positive_sentiment'
]

MEDIA_CATEGORIES = {
    'Scientific': ['nature', 'sciam', 'stat', 'newscientist'],
    'Left': ['theatlantic', 'the daily beast', 'the intercept', 'mother jones', 'msnbc', 'slate', 'vox', 'huffpost'],
    'Lean Left': ['ap', 'axios', 'cnn', 'guardian', 'business insider', 'nbcnews', 'npr', 'nytimes', 'politico', 'propublica', 'wapo', 'usa today'],
    'Center': ['reuters', 'marketwatch', 'financial times', 'newsweek', 'forbes'],
    'Lean Right': ['thedispatch', 'epochtimes', 'foxbusiness', 'wsj', 'national review', 'washtimes'],
    'Right': ['breitbart', 'theblaze', 'daily mail', 'dailywire', 'foxnews', 'nypost', 'newsmax'],
}

def setup_logging(log_file='analysis.log'):
    logging.basicConfig(
        filename=log_file,
        level=logging.DEBUG,
        format='%(asctime)s - %(levelname)s - %(message)s'
    )
    logging.info("Logging initialized.")

def load_jsonl(jsonl_file):
    records = []
    with open(jsonl_file, 'r', encoding='utf-8') as f:
        for line in tqdm(f, desc="Loading JSONL data"):
            try:
                record = json.loads(line)
                records.append(record)
            except json.JSONDecodeError as e:
                logging.error(f"JSON decoding error: {e}")
    df = pd.DataFrame(records)
    return df

def map_media_outlet_to_category(df, media_categories):
    outlet_to_category = {}
    for category, outlets in media_categories.items():
        for outlet in outlets:
            outlet_to_category[outlet.lower().strip()] = category

    if 'media_outlet' not in df.columns:
        logging.error("'media_outlet' column not found in DataFrame.")
        raise KeyError("'media_outlet' column not found in DataFrame.")

    df['media_outlet_clean'] = df['media_outlet'].str.lower().str.strip()
    df['media_category'] = df['media_outlet_clean'].map(outlet_to_category)
    df['media_category'] = df['media_category'].fillna('Other')
    unmapped_outlets = df[df['media_category'] == 'Other']['media_outlet'].unique()
    if len(unmapped_outlets) > 0:
        logging.warning(f"Unmapped media outlets found: {unmapped_outlets}")
        print(f"Warning: The following media outlets were not mapped and categorized as 'Other': {unmapped_outlets}")
    return df

def aggregate_sentiment_scores(df, sentiment_categories):
    # We'll only do quotation-based aggregation now, ignoring fulltext.
    # The code identifies quotation-based sentiment columns by pattern: sentiment_\\d+
    # We'll sum them and count them for each sentiment per media category (like before).
    aggregation = []
    for media_category in MEDIA_CATEGORIES.keys():
        mask = (df['media_category'] == media_category)
        subset = df[mask]

        for sentiment in sentiment_categories:
            pattern = f"^{re.escape(sentiment)}_\\d+$"
            matched_cols = [col for col in df.columns if re.match(pattern, col)]

            if matched_cols:
                quotation_sum = subset[matched_cols].clip(lower=0).sum(skipna=True).sum()
                quotation_count = subset[matched_cols].clip(lower=0).count().sum()
            else:
                quotation_sum = 0
                quotation_count = 0

            # No fulltext computations now
            aggregation.append({
                'Media Category': media_category,
                'Sentiment/Emotion': sentiment,
                'Quotation_Sum': quotation_sum,
                'Quotation_Count': quotation_count
            })

    aggregated_df = pd.DataFrame(aggregation)
    return aggregated_df

def calculate_averages(aggregated_df):
    # Only quotation-based averages
    aggregated_df['Quotation_Average'] = aggregated_df.apply(
        lambda row: row['Quotation_Sum'] / row['Quotation_Count'] if row['Quotation_Count'] > 0 else None, axis=1
    )
    return aggregated_df

def calculate_mean_median(aggregated_df):
    stats = []
    for sentiment in CATEGORIES:
        sentiment_data = aggregated_df[aggregated_df['Sentiment/Emotion'] == sentiment]

        quotation_avg = sentiment_data['Quotation_Average'].dropna()
        if not quotation_avg.empty:
            mean_quotation = quotation_avg.mean()
            median_quotation = quotation_avg.median()
        else:
            mean_quotation = None
            median_quotation = None

        stats.append({
            'Sentiment/Emotion': sentiment,
            'Mean_Quotation_Average': mean_quotation,
            'Median_Quotation_Average': median_quotation
        })

    stats_df = pd.DataFrame(stats)
    return stats_df

def save_aggregated_scores_to_csv(aggregated_df, csv_output_dir, prefix='aggregated_sentiment_emotion_scores.csv'):
    csv_file = os.path.join(csv_output_dir, prefix)
    try:
        aggregated_df.to_csv(csv_file, index=False)
        print(f"Aggregated sentiment/emotion scores saved to '{csv_file}'.")
        logging.info(f"Aggregated sentiment/emotion scores saved to '{csv_file}'.")
    except Exception as e:
        print(f"Error saving aggregated scores to CSV: {e}")
        logging.error(f"Error saving aggregated scores to CSV: {e}")

def plot_statistics(aggregated_df, output_dir):
    sns.set_style("whitegrid")
    for sentiment in CATEGORIES:
        sentiment_data = aggregated_df[aggregated_df['Sentiment/Emotion'] == sentiment]

        # Quotation-Based Plot only
        plt.figure(figsize=(10, 6))
        sns.barplot(
            x='Media Category',
            y='Quotation_Average',
            data=sentiment_data,
            color='steelblue'
        )
        plt.title(f"Mean Quotation-Based '{sentiment.capitalize()}' Scores", fontsize=14)
        plt.xlabel('Media Category', fontsize=12)
        plt.ylabel('Mean Quotation-Based Average Score', fontsize=12)
        plt.xticks(rotation=45, ha='right')
        plt.tight_layout()
        quote_plot_filename = f"quote_{sentiment}.png"
        quote_plot_path = os.path.join(output_dir, quote_plot_filename)
        try:
            plt.savefig(quote_plot_path)
            plt.close()
            print(f"Quotation-Based '{sentiment}' scores plot saved to '{quote_plot_path}'.")
        except Exception as e:
            print(f"Error saving quotation-based '{sentiment}' scores plot: {e}")

def fit_bambi_model_for_sentiment(df, sentiment):
    # We will build a Poisson GLMM on the quotation counts for this sentiment
    # We'll model Quotation_Sum as response (treating it as a count),
    # media_category as a fixed effect and outlet as a random effect.
    # Note: if your data isn't strictly counts, this is a modeling assumption.
    # Ensure Quotation_Sum is integer.
    # We'll filter aggregated_df rows for the given sentiment and pivot back to original data if needed.
    # Actually, we need per-article data, not aggregated by category. The aggregated data loses per-article info.
    # We must work at article level.

    # Let's reconstruct a per-article dataframe with Quotation_Sum for that sentiment:
    # Quotation_Sum per article can be computed by summing matched_cols for each article.
    pattern = f"^{re.escape(sentiment)}_\\d+$"
    matched_cols = [col for col in df.columns if re.match(pattern, col)]

    if not matched_cols:
        return None, None

    # Compute total quotation sum per article
    df = df.copy()
    df['Quotation_Sum'] = df[matched_cols].clip(lower=0).sum(axis=1)
    # Filter out rows with NaN in media_category or Quotation_Sum
    df = df.dropna(subset=['media_category', 'media_outlet_clean', 'Quotation_Sum'])

    # Convert types
    df['media_category'] = df['media_category'].astype('category')
    df['outlet'] = df['media_outlet_clean'].astype('category')

    # Fit a Bayesian GLMM with Bambi
    # Formula: Quotation_Sum ~ media_category + (1|outlet)
    # If you had article-level clustering, you'd add (1|article) if article column existed.
    import bambi as bmb
    model = bmb.Model("Quotation_Sum ~ media_category + (1|outlet)", data=df, family='poisson')
    # Fit the model using MCMC
    idata = model.fit(draws=1000, tune=1000, chains=2, cores=2)
    return model, idata

def run_bambi_analyses(df):
    # For each sentiment, fit the Bambi model and do pairwise comparisons
    results = {}
    for sentiment in CATEGORIES:
        print(f"Fitting Bambi model for sentiment: {sentiment}")
        model, idata = fit_bambi_model_for_sentiment(df, sentiment)
        if model is None:
            continue
        # Summarize model
        summary_str = model.summary(idata=idata).to_string()

        # Pairwise comparisons between categories
        # Bambi has a comparisons method:
        comps = model.compare(idata=idata, factor='media_category')
        # comps is a DataFrame of comparisons, posterior estimates, intervals, etc.

        # Posterior predictive checks
        ppc = model.predict(idata=idata, kind='pps', inplace=False)
        # Plot PPC
        sns.histplot(ppc.mean(dim=("chain","draw")).values, kde=True)
        plt.title(f"Posterior Predictive Check - {sentiment}")
        ppc_plot_path = os.path.join(OUTPUT_DIR, f"ppc_{sentiment}.png")
        plt.savefig(ppc_plot_path)
        plt.close()

        results[sentiment] = {
            'Model_Summary': summary_str,
            'Pairwise': comps
        }
    return results

def compile_results_into_multiple_workbooks(aggregated_df, stats_df, raw_df, bambi_results, plots_dir,
                                            main_excel, raw_excel, lmm_excel, plots_excel, combined_excel):
    # Similar to previous code but for Bambi results
    with pd.ExcelWriter(main_excel, engine='openpyxl') as writer:
        aggregated_df.to_excel(writer, sheet_name='Aggregated_Scores', index=False)
        stats_df.to_excel(writer, sheet_name='Mean_Median_Statistics', index=False)

    with pd.ExcelWriter(raw_excel, engine='openpyxl') as writer:
        raw_df.to_excel(writer, sheet_name='Raw_Data', index=False)
        for sentiment in CATEGORIES:
            sentiment_cols = [c for c in raw_df.columns if c.startswith(sentiment+'_')]
            sentiment_df = raw_df[['media_category', 'media_outlet'] + sentiment_cols].copy()
            sheet_name = f"Raw_{sentiment[:29]}"
            sentiment_df.to_excel(writer, sheet_name=sheet_name, index=False)

    with pd.ExcelWriter(lmm_excel, engine='openpyxl') as writer:
        summary_rows = []
        for sentiment in bambi_results:
            sheet_name = f"BAMBI_{sentiment[:20]}"
            model_summary = bambi_results[sentiment]['Model_Summary']
            pairwise_df = bambi_results[sentiment]['Pairwise']

            summary_df = pd.DataFrame({'Model_Summary': model_summary.split('\n')})
            summary_df.to_excel(writer, sheet_name=sheet_name, index=False)
            startrow = len(summary_df) + 2
            pairwise_df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=startrow)

            summary_rows.append({
                'Sentiment': sentiment,
                'SheetName': sheet_name
            })

        summary_index_df = pd.DataFrame(summary_rows)
        summary_index_df.to_excel(writer, sheet_name='BAMBI_Results_Index', index=False)

    wb = Workbook()
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])

    # Add plots if needed (quotation plots already saved)
    for sentiment in CATEGORIES:
        quote_plot_path = os.path.join(plots_dir, f"quote_{sentiment}.png")
        if os.path.exists(quote_plot_path):
            sheet_title = f"Quote_{sentiment[:28]}"
            worksheet = wb.create_sheet(title=sheet_title)
            try:
                img = ExcelImage(quote_plot_path)
                img.anchor = 'A1'
                worksheet.add_image(img)
            except Exception as e:
                logging.error(f"Error embedding image '{quote_plot_path}' into Excel: {e}")

    wb.save(plots_excel)

    raw_df_cleaned = raw_df.copy()
    raw_df_cleaned = raw_df_cleaned.applymap(lambda x: ", ".join(x) if isinstance(x, list) else x)

    wb_combined = Workbook()
    if 'Sheet' in wb_combined.sheetnames:
        wb_combined.remove(wb_combined['Sheet'])

    ws_agg = wb_combined.create_sheet("Aggregated_Scores")
    for r in dataframe_to_rows(aggregated_df, index=False, header=True):
        ws_agg.append(r)

    ws_stats = wb_combined.create_sheet("Mean_Median_Statistics")
    for r in dataframe_to_rows(stats_df, index=False, header=True):
        ws_stats.append(r)

    ws_raw = wb_combined.create_sheet("Raw_Data")
    for r in dataframe_to_rows(raw_df_cleaned, index=False, header=True):
        ws_raw.append(r)

    ws_bambi = wb_combined.create_sheet("BAMBI_Summaries")
    ws_bambi.append(["Sentiment", "Model_Summary"])
    for sentiment in bambi_results:
        summary = bambi_results[sentiment]['Model_Summary']
        ws_bambi.append([sentiment, summary])

    wb_combined.save(combined_excel)

def main():
    setup_logging()
    print("Single run with Bambi/PyMC GLMM and pairwise comparisons (Poisson family), quotation-based only.")
    df = load_jsonl(INPUT_JSONL_FILE)
    print(f"Total articles loaded: {len(df)}")
    df = map_media_outlet_to_category(df, MEDIA_CATEGORIES)

    # Print counts
    print("\nSummary Statistics:")
    print("Total number of articles:", len(df))
    outlet_counts = df['media_outlet_clean'].value_counts()
    print("\nNumber of articles per outlet:")
    print(outlet_counts)
    category_counts = df['media_category'].value_counts()
    print("\nNumber of articles per category:")
    print(category_counts)
    print()  # Blank line

    print("Aggregating sentiment/emotion scores per media category...")
    aggregated_df = aggregate_sentiment_scores(df, CATEGORIES)
    aggregated_df = calculate_averages(aggregated_df)
    stats_df = calculate_mean_median(aggregated_df)

    save_aggregated_scores_to_csv(aggregated_df, CSV_OUTPUT_DIR)
    plot_statistics(aggregated_df, OUTPUT_DIR)

    print("No predictor columns specified for collinearity check.")
    print("Fitting Bambi models (Poisson GLMM) and performing pairwise comparisons for quotations only...")
    bambi_results = run_bambi_analyses(df)
    print("Bambi GLMM analysis and pairwise tests completed successfully.\n")

    compile_results_into_multiple_workbooks(aggregated_df, stats_df, df, bambi_results, OUTPUT_DIR,
                                            OUTPUT_EXCEL_MAIN, OUTPUT_EXCEL_RAW, OUTPUT_EXCEL_LMM, OUTPUT_EXCEL_PLOTS, OUTPUT_EXCEL_COMBINED)
    print("Analysis completed successfully.")

if __name__ == "__main__":
    main()
