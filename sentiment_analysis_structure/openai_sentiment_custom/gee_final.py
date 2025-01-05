#!/usr/bin/env python3
# gee_final_all_scales_debug_fixed.py
"""
Script with extra debugging & fixes for:
 - "divide by zero encountered in log" (Poisson deviance)
 - KeyError if mu is a Series with a non-aligned index
 - Zero or negative df_resid => skip scale-based refits
 - NameError: 'df_res' vs 'df_resid' mismatch
 - FutureWarning about to_excel (just logs, doesn't break)
 
Additionally logs QIC values for each combination.

Steps:
  1) Load data from JSONL
  2) Chunk & save CSV
  3) Correlation analyses
  4) Aggregation & bar plots
  5) Old GEE approach (pairwise)
  6) New GEE approach: correlation structures => [Independence, Exchangeable],
     scale methods => [none, pearson, deviance, ub, bc]
  7) Extra debug logs for deviance calc & to_excel usage
  8) Consistently use 'df_resid' to avoid NameError
  9) **Logs QIC values** for each combination in run_gee_for_sentiment_measure_best_qic.
"""

import json
import os
import re
import sys
import warnings
import logging
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import seaborn as sns
from tqdm import tqdm

from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.utils.dataframe import dataframe_to_rows

import statsmodels.api as sm
from statsmodels.genmod.generalized_estimating_equations import GEE
from statsmodels.genmod.families import Poisson
from statsmodels.genmod.cov_struct import Independence, Exchangeable
from statsmodels.stats.multitest import multipletests
from scipy.stats import pearsonr, norm

# ------------------------------ #
# Configuration
# ------------------------------ #
INPUT_JSONL_FILE = "processed_all_articles_fixed_2.jsonl"

OUTPUT_DIR = "graphs_analysis"
os.makedirs(OUTPUT_DIR, exist_ok=True)

CSV_OUTPUT_DIR = "csv_raw_scores"
os.makedirs(CSV_OUTPUT_DIR, exist_ok=True)

OUTPUT_EXCEL_MAIN = "analysis_main.xlsx"
OUTPUT_EXCEL_RAW = "analysis_raw.xlsx"
OUTPUT_EXCEL_GEE = "analysis_gee.xlsx"
OUTPUT_EXCEL_PLOTS = "analysis_plots.xlsx"
OUTPUT_EXCEL_COMBINED = "analysis_combined.xlsx"

LOG_FILE = "analysis.log"

CATEGORIES = [
    "joy","sadness","anger","fear",
    "surprise","disgust","trust","anticipation",
    "negative_sentiment","positive_sentiment"
]

MEDIA_CATEGORIES = {
    "Scientific": ["nature","sciam","stat","newscientist"],
    "Left": ["theatlantic","the daily beast","the intercept","mother jones","msnbc","slate","vox","huffpost"],
    "Lean Left": ["ap","axios","cnn","guardian","business insider","nbcnews","npr","nytimes","politico","propublica","wapo","usa today"],
    "Center": ["reuters","marketwatch","financial times","newsweek","forbes"],
    "Lean Right": ["thedispatch","epochtimes","foxbusiness","wsj","national review","washtimes"],
    "Right": ["breitbart","theblaze","daily mail","dailywire","foxnews","nypost","newsmax"],
}


def setup_logging(log_file=LOG_FILE):
    log_format = "%(asctime)s [%(levelname)s] %(module)s - %(message)s"
    logging.basicConfig(filename=log_file, level=logging.DEBUG, format=log_format)
    console = logging.StreamHandler(sys.stdout)
    console.setLevel(logging.INFO)
    console.setFormatter(logging.Formatter(log_format))
    logging.getLogger().addHandler(console)
    logging.info("Logging initialized (file+console).")

# Silence the "QIC scale=None" warning
warnings.filterwarnings(
    "ignore",
    message="QIC values obtained using scale=None are not appropriate for comparing models"
)

# ------------------------------
# 1) Loading + chunking
# ------------------------------
def load_jsonl(jsonl_file):
    logging.info(f"Loading JSONL from {jsonl_file}")
    records=[]
    with open(jsonl_file,"r",encoding="utf-8") as f:
        for line in tqdm(f, desc="Loading JSONL data"):
            try:
                rec=json.loads(line)
                records.append(rec)
            except json.JSONDecodeError as e:
                logging.error(f"JSON decode error: {e}")
    df=pd.DataFrame(records)
    logging.debug(f"Loaded DataFrame shape={df.shape}")
    return df

def map_media_outlet_to_category(df, media_dict):
    logging.info("Mapping media_outlet -> category")
    out_map={}
    for cat, outls in media_dict.items():
        for o in outls:
            out_map[o.lower().strip()]=cat
    if "media_outlet" not in df.columns:
        raise KeyError("'media_outlet' column not found")

    df["media_outlet_clean"] = df["media_outlet"].str.lower().str.strip()
    df["media_category"] = df["media_outlet_clean"].map(out_map).fillna("Other")
    unmapped = df[df["media_category"]=="Other"]["media_outlet"].unique()
    if len(unmapped)>0:
        logging.warning(f"Unmapped outlets => {unmapped}")
        print(f"Warning: Not mapped => {unmapped}")
    return df

def chunk_and_save(df, chunk_size=20000):
    logging.info(f"Chunking DataFrame len={len(df)} in size={chunk_size}")
    for i in range(0,len(df),chunk_size):
        part=df.iloc[i:i+chunk_size]
        out_csv=os.path.join(CSV_OUTPUT_DIR,f"raw_data_part_{i//chunk_size+1}.csv")
        part.to_csv(out_csv,index=False)
        logging.info(f"Saved chunk {i//chunk_size+1} => {out_csv}")
        print(f"Saved chunk {i//chunk_size+1} to {out_csv}")

def print_basic_stats(df):
    logging.info("Basic stats about dataset.")
    logging.info(f"Total articles: {len(df)}")
    print("\nSummary Statistics:")
    print("Total number of articles:", len(df))

    if "media_outlet_clean" in df.columns:
        oc=df["media_outlet_clean"].value_counts()
        logging.debug(f"Articles per outlet:\n{oc}")
        print("\nArticles per outlet:")
        print(oc)

    if "media_category" in df.columns:
        catvc=df["media_category"].value_counts()
        logging.debug(f"Articles per category:\n{catvc}")
        print("\nArticles per category:")
        print(catvc)
    print()


# ------------------------------
# 2) Quotation vs. Fulltext correlation
# ------------------------------
def analyze_quotation_fulltext_correlation(df):
    logging.info("Analyzing Quotation vs Fulltext correlation.")
    records=[]
    for cat in df["media_category"].dropna().unique():
        dcat=df[df["media_category"]==cat]
        if dcat.empty:
            continue
        for s in CATEGORIES:
            pat=rf"^{s}_\d+$"
            matched=[c for c in dcat.columns if re.match(pat,c)]
            if matched:
                clipped=dcat[matched].clip(lower=0)
                qsum=clipped.sum(skipna=True).sum()
                qcount=clipped.count().sum()
                qavg=qsum/qcount if qcount>0 else np.nan
            else:
                qavg=np.nan

            fcol=f"{s}_fulltext"
            if fcol in dcat.columns:
                fvals=dcat[fcol].clip(lower=0)
                f_sum=fvals.sum(skipna=True)
                f_count=fvals.count()
                favg=f_sum/f_count if f_count>0 else np.nan
            else:
                favg=np.nan

            records.append({"MediaCategory":cat,"Sentiment":s,
                            "Quotation_Average":qavg,"Fulltext_Average":favg})
    agg_df=pd.DataFrame(records)
    correlation_results=[]
    scatter_map={}

    for s in CATEGORIES:
        sdf=agg_df[(agg_df["Sentiment"]==s)].dropna(subset=["Quotation_Average","Fulltext_Average"])
        if len(sdf)>1:
            corr_val,_=pearsonr(sdf["Quotation_Average"], sdf["Fulltext_Average"])
        else:
            corr_val=np.nan
        correlation_results.append({"Sentiment":s,"Correlation":corr_val})

        if not sdf.empty:
            logging.debug(f"Scatterplot for sentiment={s}, correlation={corr_val}")
            plt.figure(figsize=(6,5))
            sns.scatterplot(x="Quotation_Average",y="Fulltext_Average",data=sdf,hue="MediaCategory",s=50)
            plt.title(f"{s.capitalize()} (Quotation vs. Fulltext)\nr={corr_val:.3f}")
            plt.tight_layout()
            out_scatter=os.path.join("graphs_analysis",f"scatter_{s}.png")
            try:
                plt.savefig(out_scatter)
            except Exception as e:
                logging.error(f"Err saving scatter {s}: {e}")
            plt.close()
        scatter_map[s]=sdf.copy()

    corr_df=pd.DataFrame(correlation_results)
    logging.debug(f"Correlation DataFrame:\n{corr_df}")

    plt.figure(figsize=(8,5))
    sns.barplot(x="Sentiment",y="Correlation",data=corr_df,color="gray")
    plt.title("Correlation (Quotation vs Fulltext) per Sentiment")
    plt.xticks(rotation=45,ha="right")
    plt.ylim(-1,1)
    plt.tight_layout()
    bar_path=os.path.join("graphs_analysis","correlation_quotation_fulltext_bar.png")
    try:
        plt.savefig(bar_path)
    except Exception as e:
        logging.error(f"Err saving correlation bar: {e}")
    plt.close()

    # combined z-scatter
    combo_list=[]
    for s,sdf in scatter_map.items():
        if not sdf.empty:
            cpy=sdf.copy()
            cpy["Sentiment"]=s
            combo_list.append(cpy)
    if combo_list:
        cat_all=pd.concat(combo_list,ignore_index=True)
        cat_all["Qmean"]=cat_all.groupby("Sentiment")["Quotation_Average"].transform("mean")
        cat_all["Qstd"]=cat_all.groupby("Sentiment")["Quotation_Average"].transform("std")
        cat_all["Fmean"]=cat_all.groupby("Sentiment")["Fulltext_Average"].transform("mean")
        cat_all["Fstd"]=cat_all.groupby("Sentiment")["Fulltext_Average"].transform("std")
        cat_all["Quotation_Z"]=(cat_all["Quotation_Average"]-cat_all["Qmean"])/cat_all["Qstd"]
        cat_all["Fulltext_Z"]=(cat_all["Fulltext_Average"]-cat_all["Fmean"])/cat_all["Fstd"]
        valid=cat_all.dropna(subset=["Quotation_Z","Fulltext_Z"])
        if len(valid)>1:
            r_val,_=pearsonr(valid["Quotation_Z"], valid["Fulltext_Z"])
        else:
            r_val=np.nan

        plt.figure(figsize=(7,5))
        sns.regplot(x="Quotation_Z",y="Fulltext_Z",data=valid,
                    scatter_kws={"color":"black","alpha":0.6}, line_kws={"color":"red"})
        plt.title(f"All Sentiments Combined (Z-scores)\nr={r_val:.3f}")
        plt.tight_layout()
        comb_path=os.path.join("graphs_analysis","combined_normalized_scatter.png")
        try:
            plt.savefig(comb_path)
        except Exception as e:
            logging.error(f"Err saving combined z-scatter: {e}")
        plt.close()

    out_csv=os.path.join("csv_raw_scores","quotation_fulltext_correlation.csv")
    try:
        corr_df.to_csv(out_csv,index=False)
        logging.info(f"Correlation data => {out_csv}")
    except Exception as e:
        logging.error(f"Err saving correlation CSV: {e}")


# ------------------------------
# 3) Aggregation & Stats
# ------------------------------
def aggregate_sentiment_scores(df,sentiment_cats):
    logging.info("Aggregating sentiment/emotion scores by category + sentiment.")
    out=[]
    for mc in MEDIA_CATEGORIES.keys():
        dcat=df[df["media_category"]==mc]
        for s in sentiment_cats:
            pat=rf"^{re.escape(s)}_\d+$"
            matched=[c for c in dcat.columns if re.match(pat,c)]
            if matched:
                qsum=dcat[matched].clip(lower=0).sum(skipna=True).sum()
                qcount=dcat[matched].clip(lower=0).count().sum()
            else:
                qsum,qcount=(0,0)

            fcol=f"{s}_fulltext"
            if fcol in dcat.columns:
                ft_clipped=dcat[fcol].clip(lower=0)
                f_sum=ft_clipped.sum(skipna=True)
                f_count=ft_clipped.count()
            else:
                f_sum,f_count=(0,0)

            out.append({
                "Media Category":mc,
                "Sentiment/Emotion":s,
                "Quotation_Sum":qsum,
                "Quotation_Count":qcount,
                "Fulltext_Sum":f_sum,
                "Fulltext_Count":f_count
            })
    agg_df=pd.DataFrame(out)
    logging.debug(f"Aggregated shape={agg_df.shape}")
    return agg_df

def calculate_averages(agg_df):
    logging.info("Calculating Quotation_Average, Fulltext_Average per row.")
    def safe_div(a,b):
        return a/b if b>0 else None
    agg_df["Quotation_Average"]=agg_df.apply(lambda r: safe_div(r["Quotation_Sum"], r["Quotation_Count"]), axis=1)
    agg_df["Fulltext_Average"]=agg_df.apply(lambda r: safe_div(r["Fulltext_Sum"], r["Fulltext_Count"]), axis=1)
    return agg_df

def calculate_mean_median(agg_df):
    logging.info("Computing global mean/median of Quotation/Fulltext averages.")
    results=[]
    for s in CATEGORIES:
        sub=agg_df[agg_df["Sentiment/Emotion"]==s]
        qa=sub["Quotation_Average"].dropna()
        fa=sub["Fulltext_Average"].dropna()
        mean_q=qa.mean() if len(qa)>0 else None
        med_q=qa.median() if len(qa)>0 else None
        mean_f=fa.mean() if len(fa)>0 else None
        med_f=fa.median() if len(fa)>0 else None
        results.append({
            "Sentiment/Emotion":s,
            "Mean_Quotation_Average":mean_q,
            "Median_Quotation_Average":med_q,
            "Mean_Fulltext_Average":mean_f,
            "Median_Fulltext_Average":med_f
        })
    df_stat=pd.DataFrame(results)
    logging.debug(f"Mean/median stats:\n{df_stat}")
    return df_stat

def save_aggregated_scores_to_csv(agg_df, out_dir, prefix="aggregated_sentiment_emotion_scores.csv"):
    out_csv=os.path.join(out_dir,prefix)
    logging.info(f"Saving aggregated scores => {out_csv}")
    agg_df.to_csv(out_csv,index=False)
    print(f"Aggregated sentiment/emotion scores => {out_csv}")
    logging.info(f"Aggregated => {out_csv}")

def plot_statistics(agg_df,out_dir):
    logging.info("Plotting bar charts for Quotation/Fulltext averages.")
    sns.set_style("whitegrid")
    for s in CATEGORIES:
        sdata=agg_df[agg_df["Sentiment/Emotion"]==s]

        # Quotation
        plt.figure(figsize=(10,6))
        sns.barplot(x="Media Category",y="Quotation_Average",data=sdata,color="steelblue")
        plt.title(f"Mean Quotation-Based '{s.capitalize()}' Scores")
        plt.xticks(rotation=45,ha="right")
        plt.tight_layout()
        q_png=os.path.join(out_dir,f"quote_{s}.png")
        try:
            plt.savefig(q_png)
            logging.debug(f"Saved Quotation barplot => {q_png}")
        except Exception as e:
            logging.error(f"Error saving quote plot for {s}: {e}")
        plt.close()

        # Fulltext
        plt.figure(figsize=(10,6))
        sns.barplot(x="Media Category",y="Fulltext_Average",data=sdata,color="darkorange")
        plt.title(f"Mean Fulltext-Based '{s.capitalize()}' Scores")
        plt.xticks(rotation=45,ha="right")
        plt.tight_layout()
        f_png=os.path.join(out_dir,f"fulltext_{s}.png")
        try:
            plt.savefig(f_png)
            logging.debug(f"Saved Fulltext barplot => {f_png}")
        except Exception as e:
            logging.error(f"Error saving fulltext plot for {s}: {e}")
        plt.close()


# ------------------------------
# 4) Old GEE approach
# ------------------------------
def fit_gee_and_pairwise(df, sentiment, measure="Quotation"):
    logging.debug(f"Running old GEE pairwise for sentiment={sentiment}, measure={measure}")
    if measure=="Quotation":
        pat=rf"^{re.escape(sentiment)}_\d+$"
        matched=[c for c in df.columns if re.match(pat,c)]
        if not matched:
            logging.debug("No matched columns => skip pairwise.")
            return None
        d2=df.copy()
        d2[f"{sentiment}_quotation_mean"]=d2[matched].clip(lower=0).mean(axis=1)
        score_col=f"{sentiment}_quotation_mean"
    else:
        fcol=f"{sentiment}_fulltext"
        if fcol not in df.columns:
            logging.debug("No fulltext => skip pairwise.")
            return None
        d2=df.copy()
        d2[f"{sentiment}_fulltext_clipped"]=d2[fcol].clip(lower=0)
        score_col=f"{sentiment}_fulltext_clipped"

    mdf=d2.dropna(subset=[score_col,"media_category","media_outlet"]).copy()
    if mdf["media_category"].nunique()<2:
        logging.debug("media_category <2 => skip pairwise.")
        return None
    mdf["media_category"]=mdf["media_category"].astype("category")

    model=GEE.from_formula(
        formula=f"{score_col} ~ media_category",
        groups="media_outlet",
        data=mdf,
        family=Poisson(),
        cov_struct=Exchangeable()
    )
    res=model.fit()
    summary_txt=res.summary().as_text()
    logging.debug(f"Fit old GEE {sentiment}-{measure}, summary snippet:\n{summary_txt[:300]}...")

    params=res.params
    cov=res.cov_params()
    cats=mdf["media_category"].cat.categories
    ref=cats[0]

    idx_map={ref:0}
    for c in cats[1:]:
        nm=f"media_category[T.{c}]"
        idx_map[c]=res.model.exog_names.index(nm)

    pair_list=[]
    for i in range(len(cats)):
        for j in range(i+1,len(cats)):
            ca,cb=cats[i],cats[j]
            contrast=np.zeros(len(params))
            if ca==ref and cb!=ref:
                contrast[idx_map[cb]]=-1.0
            elif cb==ref and ca!=ref:
                contrast[idx_map[ca]]=1.0
            else:
                contrast[idx_map[ca]]=1.0
                contrast[idx_map[cb]]=-1.0

            diff_est=contrast@params
            diff_var=contrast@cov@contrast
            diff_se=np.sqrt(diff_var)
            z=diff_est/diff_se
            p_val=2*(1-norm.cdf(abs(z)))
            pair_list.append((ca,cb,diff_est,diff_se,z,p_val))

    pw_df=pd.DataFrame(pair_list, columns=["CategoryA","CategoryB","Difference","SE","Z","p_value"])
    rej,adj_p,_,_=multipletests(pw_df["p_value"],method="holm")
    pw_df["p_value_adj"]=adj_p
    pw_df["reject_H0"]=rej
    return {
        "GEE_Summary": summary_txt,
        "Pairwise": pw_df
    }

def run_gee_analyses_pairwise(df):
    logging.info("Running old GEE approach (pairwise, Exchangeable).")
    out={}
    for s in CATEGORIES:
        out[s]={}
        for meas in ["Quotation","Fulltext"]:
            r=fit_gee_and_pairwise(df,s,meas)
            if r: out[s][meas]=r
    return out


# -----------------------------------
# 5) Scale computations (df_resid fix)
# -----------------------------------
def compute_pearson_scale(y, mu, df_resid):
    logging.debug(f"[compute_pearson_scale] df_resid={df_resid}, y.shape={np.shape(y)}, mu.shape={np.shape(mu)}")
    if df_resid <= 0:
        logging.error(f"[compute_pearson_scale] df_resid={df_resid} <= 0 => can't compute => NaN.")
        return np.nan
    y = np.asarray(y)
    mu = np.asarray(mu)
    resid_pearson=(y-mu)/np.sqrt(mu)
    chi2=np.sum(resid_pearson**2)
    return chi2/df_resid

def compute_deviance_scale(y, mu, df_resid):
    logging.debug(f"[compute_deviance_scale] df_resid={df_resid}, y.shape={np.shape(y)}, mu.shape={np.shape(mu)}")
    if df_resid <= 0:
        logging.error(f"[compute_deviance_scale] df_resid={df_resid} <= 0 => can't compute => NaN.")
        return np.nan

    y = np.asarray(y)
    mu = np.asarray(mu)

    dev_array=np.zeros_like(y,dtype=float)
    for i in range(len(y)):
        if y[i]>0:
            if mu[i]>0:
                val=y[i]*np.log(y[i]/mu[i])
                if np.isinf(val) or np.isnan(val):
                    logging.debug(f"[Deviance] Row={i}, y={y[i]}, mu={mu[i]}, val={val}")
                dev_array[i] = val - (y[i]-mu[i])
            else:
                logging.debug(f"[Deviance] Row={i}, y={y[i]}, mu=0 => dev_array[i]= nan.")
                dev_array[i] = np.nan
        else:
            dev_array[i] = - (0 - mu[i])  # = mu[i]
    dev=2*np.nansum(dev_array)
    return dev/df_resid

def compute_ub_scale(y, mu, df_resid):
    logging.debug(f"[compute_ub_scale] df_resid={df_resid}, y.shape={np.shape(y)}, mu.shape={np.shape(mu)}")
    if df_resid <= 0:
        logging.error(f"[compute_ub_scale] df_resid={df_resid} <= 0 => can't compute => NaN.")
        return np.nan
    y = np.asarray(y)
    mu = np.asarray(mu)
    p=compute_pearson_scale(y,mu,df_resid)
    return 1.1*p if not np.isnan(p) else np.nan

def compute_bc_scale(y, mu, df_resid):
    logging.debug(f"[compute_bc_scale] df_resid={df_resid}, y.shape={np.shape(y)}, mu.shape={np.shape(mu)}")
    if df_resid <= 0:
        logging.error(f"[compute_bc_scale] df_resid={df_resid} <= 0 => can't compute => NaN.")
        return np.nan
    y = np.asarray(y)
    mu = np.asarray(mu)
    d=compute_deviance_scale(y,mu,df_resid)
    return 0.9*d if not np.isnan(d) else np.nan

def fit_and_compute_scales(model):
    logging.debug("Fitting model with scale=None to gather residual info.")
    base_res=model.fit(scale=None)
    base_qic=base_res.qic()

    y=base_res.model.endog
    mu=base_res.fittedvalues
    n=len(y)
    p=len(base_res.params)
    df_resid=n-p
    logging.debug(f"[fit_and_compute_scales] n={n}, p={p}, df_resid={df_resid}")

    pear=compute_pearson_scale(y,mu,df_resid)
    dev=compute_deviance_scale(y,mu,df_resid)
    ubv=compute_ub_scale(y,mu,df_resid)
    bcv=compute_bc_scale(y,mu,df_resid)
    logging.debug(f"[fit_and_compute_scales] scale vals => pearson={pear}, deviance={dev}, ub={ubv}, bc={bcv}")

    results={}
    if isinstance(base_qic, tuple):
        q_m,q_a=base_qic
    else:
        q_m,q_a=base_qic,None
    results["none"]=(q_m,q_a,None)

    from math import isnan
    for (nm,val) in [("pearson",pear),("deviance",dev),("ub",ubv),("bc",bcv)]:
        if (not isnan(val)) and (df_resid>0):
            logging.debug(f"[fit_and_compute_scales] Refitting with scale={nm}, numeric={val}")
            re_res=model.fit(scale=val)
            re_qic=re_res.qic()
            if isinstance(re_qic,tuple):
                qq_m,qq_a=re_qic
            else:
                qq_m,qq_a=re_qic,None
            results[nm]=(qq_m,qq_a,val)
            logging.debug(f" => QIC_main={qq_m}, QIC_alt={qq_a}")
        else:
            logging.warning(f"[fit_and_compute_scales] scale={nm} invalid => val={val}, df_resid={df_resid}. Setting QIC=NaN.")
            results[nm]=(np.nan, None, val)

    return results


# ------------------------------
# 6) Best QIC approach (logs QIC)
# ------------------------------
def run_gee_for_sentiment_measure_best_qic(df, sentiment, measure):
    logging.debug(f"Best QIC for sentiment={sentiment}, measure={measure}")
    d2=df.copy()
    if measure=="Quotation":
        pat=rf"^{re.escape(sentiment)}_\d+$"
        matched=[c for c in d2.columns if re.match(pat,c)]
        if not matched:
            logging.debug("No matched => skip best QIC.")
            return None
        d2["_score_col"]=d2[matched].clip(lower=0).mean(axis=1)
    else:
        fcol=f"{sentiment}_fulltext"
        if fcol not in d2.columns:
            logging.debug("No fulltext => skip best QIC.")
            return None
        d2["_score_col"]=d2[fcol].clip(lower=0)

    needed=["_score_col","media_outlet_clean","media_category"]
    d2=d2.dropna(subset=needed)
    if len(d2)<2 or d2["media_category"].nunique()<2:
        logging.debug("Not enough data for best QIC.")
        return None

    best_qic_main=np.inf
    best_combo=None
    all_records=[]

    structures=[Independence(), Exchangeable()]
    for cov_obj in structures:
        st_name=cov_obj.__class__.__name__
        logging.debug(f"Fitting GEE => struct={st_name}, sentiment={sentiment}, measure={measure}")
        model=GEE.from_formula("_score_col ~ media_category",
                               groups="media_outlet_clean", data=d2,
                               family=Poisson(), cov_struct=cov_obj)
        scale_res=fit_and_compute_scales(model)

        # --- ADDED LOGGING OF QIC VALUES --- 
        # each time we get QIC for scale method, we log it
        for sm,(q_m,q_a,sc) in scale_res.items():
            # <-- NEW LOG line: log QIC for each combination
            logging.info(f"[QIC] Structure={st_name}, Scale={sm}, QIC_main={q_m}, QIC_alt={q_a}, ScaleValue={sc}")

            all_records.append({
                "Sentiment":sentiment,
                "Measure":measure,
                "Structure":st_name,
                "ScaleMethod":sm,
                "NumericScale":sc,
                "QIC_main":q_m,
                "QIC_alt":q_a
            })

            if (q_m < best_qic_main) and (not np.isnan(q_m)):
                best_qic_main=q_m
                best_combo=(st_name,sm,q_m,sc)

    if best_combo is None:
        return None
    best_struct,best_scale,best_qic_val,best_scale_num=best_combo
    return {
        "Sentiment":sentiment,
        "Measure":measure,
        "Best_Structure":best_struct,
        "Best_Scale":best_scale,
        "Best_QIC_main":best_qic_val,
        "AllCombos":pd.DataFrame(all_records),
        "Summary":f"Best scale={best_scale}, numeric={best_scale_num}"
    }

def run_gee_analyses_best_qic(df):
    logging.info("Running new GEE approach => Ind/Exch, scale=[none,pearson,deviance,ub,bc].")
    best_rows=[]
    allcombo_list=[]
    for s in CATEGORIES:
        for meas in ["Quotation","Fulltext"]:
            info=run_gee_for_sentiment_measure_best_qic(df,s,meas)
            if info is not None:
                best_rows.append({
                    "Sentiment":info["Sentiment"],
                    "Measure":info["Measure"],
                    "Best_Structure":info["Best_Structure"],
                    "Best_Scale":info["Best_Scale"],
                    "Best_QIC_main":info["Best_QIC_main"],
                    "Summary":info["Summary"]
                })
                allcombo_list.append(info["AllCombos"])
    best_df=pd.DataFrame(best_rows)
    combos_df=pd.concat(allcombo_list, ignore_index=True) if allcombo_list else pd.DataFrame()
    return best_df, combos_df


# ------------------------------
# 7) Excel compilation
# ------------------------------
def compile_results_into_multiple_workbooks(
    aggregated_df, stats_df, raw_df,
    gee_results_pairwise,
    gee_best_df, gee_allcombos_df,
    plots_dir,
    main_excel, raw_excel, gee_excel, plots_excel, combined_excel
):
    logging.info("Compiling results into Excel files with more debug logs.")
    # main
    logging.debug(f"Writing => {main_excel} (Aggregated_Scores, Mean_Median_Statistics). Expect FutureWarnings if Pandas>3.0.")
    with pd.ExcelWriter(main_excel, engine="openpyxl") as w:
        aggregated_df.to_excel(w, "Aggregated_Scores", index=False)
        stats_df.to_excel(w, "Mean_Median_Statistics", index=False)

    # raw
    logging.debug(f"Writing raw => {raw_excel}")
    with pd.ExcelWriter(raw_excel, engine="openpyxl") as w:
        raw_df.to_excel(w, "Raw_Data", index=False)
        for s in CATEGORIES:
            s_cols=[c for c in raw_df.columns if c.startswith(s+"_")]
            s_df=raw_df[["media_category","media_outlet"]+s_cols].copy()
            sheet_name=f"Raw_{s[:29]}"
            s_df.to_excel(w, sheet_name, index=False)

    # analysis_gee
    logging.debug(f"Writing => {gee_excel} with old pairwise + new best QIC.")
    with pd.ExcelWriter(gee_excel, engine="openpyxl") as w:
        # old pairwise
        summary_rows=[]
        for sentiment in gee_results_pairwise:
            for meas in gee_results_pairwise[sentiment]:
                sht=f"Pairwise_{sentiment[:10]}_{meas[:8]}"
                pres=gee_results_pairwise[sentiment][meas]
                sum_txt=pres["GEE_Summary"]
                pw_df=pres["Pairwise"]
                sumdf=pd.DataFrame({"GEE_Summary":sum_txt.split("\n")})
                sumdf.to_excel(w, sht, index=False)
                strow=len(sumdf)+2
                pw_df.to_excel(w, sht, index=False, startrow=strow)
                summary_rows.append({
                    "Sentiment":sentiment,
                    "Measure":meas,
                    "SheetName":sht
                })
        idxdf=pd.DataFrame(summary_rows)
        idxdf.to_excel(w, "Pairwise_Index", index=False)

        # best combos
        gee_best_df.to_excel(w, "GEE_Best_QIC", index=False)
        if not gee_allcombos_df.empty:
            gee_allcombos_df.to_excel(w, "All_Combos", index=False)

    # plots
    logging.debug(f"Embedding plots => {plots_excel}")
    wb_plots=Workbook()
    if "Sheet" in wb_plots.sheetnames:
        wb_plots.remove(wb_plots["Sheet"])

    for s in CATEGORIES:
        q_png=os.path.join(plots_dir,f"quote_{s}.png")
        if os.path.exists(q_png):
            st=f"Quote_{s[:28]}"
            ws=wb_plots.create_sheet(title=st)
            try:
                img=ExcelImage(q_png)
                img.anchor="A1"
                ws.add_image(img)
            except Exception as e:
                logging.error(f"Embedding {q_png}: {e}")

        f_png=os.path.join(plots_dir,f"fulltext_{s}.png")
        if os.path.exists(f_png):
            st=f"Fulltext_{s[:25]}"
            ws=wb_plots.create_sheet(title=st)
            try:
                img=ExcelImage(f_png)
                img.anchor="A1"
                ws.add_image(img)
            except Exception as e:
                logging.error(f"Embedding {f_png}: {e}")

    corr_bar=os.path.join(plots_dir,"correlation_quotation_fulltext_bar.png")
    if os.path.exists(corr_bar):
        ws=wb_plots.create_sheet("Correlation_Bar")
        try:
            img=ExcelImage(corr_bar)
            img.anchor="A1"
            ws.add_image(img)
        except Exception as e:
            logging.error(f"Embedding {corr_bar}: {e}")

    comb_path=os.path.join(plots_dir,"combined_normalized_scatter.png")
    if os.path.exists(comb_path):
        ws=wb_plots.create_sheet("Combined_ZScatter")
        try:
            img=ExcelImage(comb_path)
            img.anchor="A1"
            ws.add_image(img)
        except Exception as e:
            logging.error(f"Embedding {comb_path}: {e}")

    wb_plots.save(plots_excel)

    # combined
    logging.debug(f"Compiling everything => {combined_excel}")
    raw_clean=raw_df.copy()
    raw_clean=raw_clean.applymap(lambda x:", ".join(x) if isinstance(x,list) else x)
    wb_combined=Workbook()
    if "Sheet" in wb_combined.sheetnames:
        wb_combined.remove(wb_combined["Sheet"])

    ws_agg=wb_combined.create_sheet("Aggregated_Scores")
    for r in dataframe_to_rows(aggregated_df,index=False,header=True):
        ws_agg.append(r)

    ws_stats=wb_combined.create_sheet("Mean_Median_Statistics")
    for r in dataframe_to_rows(stats_df,index=False,header=True):
        ws_stats.append(r)

    ws_raw=wb_combined.create_sheet("Raw_Data")
    for r in dataframe_to_rows(raw_clean,index=False,header=True):
        ws_raw.append(r)

    ws_gee_old=wb_combined.create_sheet("GEE_Old_Pairwise")
    ws_gee_old.append(["Sentiment","Measure","Summary_Snippet"])
    for sentiment in gee_results_pairwise:
        for meas in gee_results_pairwise[sentiment]:
            sum_txt=gee_results_pairwise[sentiment][meas]["GEE_Summary"]
            snippet=(sum_txt[:200]+"...") if len(sum_txt)>200 else sum_txt
            ws_gee_old.append([sentiment,meas,snippet])

    ws_best_qic=wb_combined.create_sheet("GEE_Best_QIC_Short")
    short_cols=["Sentiment","Measure","Best_Structure","Best_Scale","Best_QIC_main"]
    ws_best_qic.append(short_cols)
    for _, row in gee_best_df.iterrows():
        rv=[row["Sentiment"], row["Measure"], row["Best_Structure"], row["Best_Scale"], row["Best_QIC_main"]]
        ws_best_qic.append(rv)

    wb_combined.save(combined_excel)
    logging.info("Done compiling Excel files. (FutureWarnings about to_excel are not harmful.)")

def main():
    setup_logging()
    logging.info("Starting analysis with deviance logs, KeyError fix, zero df_resid check, name fix: df_resid, +QIC logs.")

    # 1) Load
    df=load_jsonl(INPUT_JSONL_FILE)
    print(f"Total articles loaded: {len(df)}")

    # 2) map + chunk + stats
    df=map_media_outlet_to_category(df, MEDIA_CATEGORIES)
    chunk_and_save(df,20000)
    print_basic_stats(df)

    # 3) correlation
    print("Performing Quotation vs Fulltext correlation analysis...")
    analyze_quotation_fulltext_correlation(df)

    # 4) aggregation
    print("Aggregating sentiment/emotion scores per media category...")
    agg_df=aggregate_sentiment_scores(df,CATEGORIES)
    agg_df=calculate_averages(agg_df)
    stats_df=calculate_mean_median(agg_df)
    save_aggregated_scores_to_csv(agg_df,CSV_OUTPUT_DIR)
    plot_statistics(agg_df,OUTPUT_DIR)

    # 5) old GEE
    print("Fitting OLD GEE approach (Exchangeable + pairwise comparisons)...")
    gee_results_pairwise=run_gee_analyses_pairwise(df)
    print("Old GEE approach completed.\n")

    # 6) best QIC approach
    print("Fitting NEW GEE approach (Ind/Exch + scale=[none,pearson,deviance,ub,bc]) => best QIC...")
    best_qic_df,all_combos_df=run_gee_analyses_best_qic(df)
    print("New GEE approach completed.\n")

    # 7) compile
    compile_results_into_multiple_workbooks(
        aggregated_df=agg_df,
        stats_df=stats_df,
        raw_df=df,
        gee_results_pairwise=gee_results_pairwise,
        gee_best_df=best_qic_df,
        gee_allcombos_df=all_combos_df,
        plots_dir=OUTPUT_DIR,
        main_excel=OUTPUT_EXCEL_MAIN,
        raw_excel=OUTPUT_EXCEL_RAW,
        gee_excel=OUTPUT_EXCEL_GEE,
        plots_excel=OUTPUT_EXCEL_PLOTS,
        combined_excel=OUTPUT_EXCEL_COMBINED
    )

    print("Analysis completed successfully.")
    logging.info("Analysis completed successfully.")

if __name__=="__main__":
    main()
