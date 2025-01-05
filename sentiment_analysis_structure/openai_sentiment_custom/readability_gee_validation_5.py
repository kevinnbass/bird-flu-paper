#!/usr/bin/env python3
# gee_w_validation_10_fixed_zscatter_fog_fk.py
"""
Script that:
  1) Loads data, does correlation analyses (REMOVED the quotation vs. fulltext correlation)
  2) Aggregates stats for measures: flesch_kincaid_grade_global, gunning_fog_global
  3) Finds best QIC combination (multi-family) for each measure
  4) For each best QIC model:
     - Refit & produce GEE summary
     - Advanced checks => diagnostics, cross-validation, etc.
     - Pairwise => "DiffIDs"
  5) Writes summary, pairwise, & advanced checks into analysis_gee.xlsx
     with BestQIC_Index, All_Combos, Diagnostics, CrossValidation_Res, etc.
  6) Also saves analysis_main, analysis_raw, analysis_plots, analysis_combined
"""

import json
import os
import re
import sys
import warnings
import logging
import math
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import seaborn as sns
from tqdm import tqdm

import statsmodels.api as sm
from statsmodels.genmod.families import (
    Poisson, NegativeBinomial,
    Gaussian, Gamma, InverseGaussian
)
import statsmodels.genmod.families.links as ln
from statsmodels.genmod.generalized_estimating_equations import GEE
from statsmodels.genmod.cov_struct import Independence, Exchangeable
from statsmodels.stats.multitest import multipletests
from scipy.stats import pearsonr, norm

from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill

# Configuration
INPUT_JSONL_FILE = "processed_all_articles_fixed_3.jsonl"

OUTPUT_DIR = "graphs_analysis"
os.makedirs(OUTPUT_DIR, exist_ok=True)

CSV_OUTPUT_DIR = "csv_raw_scores"
os.makedirs(CSV_OUTPUT_DIR, exist_ok=True)

OUTPUT_EXCEL_MAIN = "analysis_main.xlsx"
OUTPUT_EXCEL_RAW = "analysis_raw.xlsx"
OUTPUT_EXCEL_GEE = "analysis_gee.xlsx"
OUTPUT_EXCEL_PLOTS = "analysis_plots.xlsx"
OUTPUT_EXCEL_COMBINED = "analysis_combined.xlsx"

LOG_FILE = "analysis.log"

# We now focus on two measures:
MEASURES = [
    "flesch_kincaid_grade_global",
    "gunning_fog_global"
]

MEDIA_CATEGORIES = {
    "Scientific": ["nature","sciam","stat","newscientist"],
    "Left": ["theatlantic","the daily beast","the intercept","mother jones","msnbc","slate","vox","huffpost"],
    "Lean Left": ["ap","axios","cnn","guardian","business insider","nbcnews","npr","nytimes","politico","propublica","wapo","usa today"],
    "Center": ["reuters","marketwatch","financial times","newsweek","forbes"],
    "Lean Right": ["thedispatch","epochtimes","foxbusiness","wsj","national review","washtimes"],
    "Right": ["breitbart","theblaze","daily mail","dailywire","foxnews","nypost","newsmax"],
}

def setup_logging():
    log_format = "%(asctime)s [%(levelname)s] %(module)s - %(message)s"
    logging.basicConfig(filename=LOG_FILE, level=logging.DEBUG, format=log_format)
    console = logging.StreamHandler(sys.stdout)
    console.setLevel(logging.INFO)
    console.setFormatter(logging.Formatter(log_format))
    logging.getLogger().addHandler(console)
    logging.info("Logging initialized (file+console).")

warnings.filterwarnings(
    "ignore",
    message="QIC values obtained using scale=None are not appropriate for comparing models"
)

###############################################################################
# 1) Load, chunk, basic stats
###############################################################################
def load_jsonl(jsonl_file):
    logging.info(f"Loading JSONL from {jsonl_file}")
    records=[]
    with open(jsonl_file,"r",encoding="utf-8") as f:
        for line in tqdm(f, desc="Loading JSONL data"):
            rec=json.loads(line)
            records.append(rec)
    df=pd.DataFrame(records)
    logging.debug(f"Loaded DataFrame shape={df.shape}")
    return df

def map_media_outlet_to_category(df):
    cat_map={}
    for cat, outls in MEDIA_CATEGORIES.items():
        for o in outls:
            cat_map[o.lower().strip()] = cat

    if "media_outlet" not in df.columns:
        raise KeyError("'media_outlet' not found in data")

    df["media_outlet_clean"] = df["media_outlet"].str.lower().str.strip()
    df["media_category"] = df["media_outlet_clean"].map(cat_map).fillna("Other")

    unmapped = df[df["media_category"]=="Other"]["media_outlet"].unique()
    if len(unmapped)>0:
        logging.warning(f"Unmapped => {unmapped}")
        print(f"Warning: Not mapped => {unmapped}")
    return df

def chunk_and_save(df, chunk_size=20000):
    logging.info(f"Chunking DataFrame => len={len(df)}, chunk_size={chunk_size}")
    for i in range(0,len(df),chunk_size):
        part=df.iloc[i:i+chunk_size]
        out_csv=os.path.join(CSV_OUTPUT_DIR,f"raw_data_part_{(i//chunk_size)+1}.csv")
        part.to_csv(out_csv,index=False)
        print(f"Saved chunk {(i//chunk_size)+1} to {out_csv}")

def print_basic_stats(df):
    logging.info(f"Basic stats => total articles = {len(df)}")
    print("\nSummary Statistics:")
    print("Total number of articles:", len(df))
    if "media_outlet_clean" in df.columns:
        vc=df["media_outlet_clean"].value_counts()
        print("\nArticles per outlet:")
        print(vc)
    if "media_category" in df.columns:
        vc2=df["media_category"].value_counts()
        print("\nArticles per category:")
        print(vc2)
    print()

###############################################################################
# REMOVED: Quotation vs. Fulltext correlation analysis
###############################################################################

###############################################################################
# 2) Aggregation & Stats (for our two measures)
###############################################################################
def aggregate_measures(df, measures):
    """
    Aggregates each measure by media_category.
    We'll compute sum, count, and then an average for each measure 
    across each media category.
    """
    logging.info("Aggregating measures by category.")
    recs = []
    for cat in MEDIA_CATEGORIES.keys():
        sub = df[df["media_category"] == cat]
        for m in measures:
            if m in sub.columns:
                # Force values to numeric, setting invalid ones to NaN, then drop them
                numeric_vals = pd.to_numeric(sub[m], errors="coerce").dropna()
                the_sum = numeric_vals.sum()
                the_count = numeric_vals.count()
            else:
                the_sum, the_count = (0, 0)

            recs.append({
                "Media Category": cat,
                "Measure": m,
                "Sum": the_sum,
                "Count": the_count,
            })
    return pd.DataFrame(recs)

def calculate_averages(agg_df):
    logging.info("Calculating measure Average.")
    def sdiv(a,b): return a/b if b>0 else None
    agg_df["Average"] = agg_df.apply(lambda r: sdiv(r["Sum"], r["Count"]), axis=1)
    return agg_df

def calculate_mean_median(agg_df, measures):
    """
    Compute global mean/median of each measure's Average.
    """
    logging.info("Computing global mean/median of measure averages.")
    rows=[]
    for m in measures:
        sub = agg_df[agg_df["Measure"] == m]
        avg_col = sub["Average"].dropna()
        rows.append({
            "Measure": m,
            "Mean_of_Averages": avg_col.mean() if len(avg_col)>0 else None,
            "Median_of_Averages": avg_col.median() if len(avg_col)>0 else None
        })
    return pd.DataFrame(rows)

def save_aggregated_scores_to_csv(agg_df, out_dir):
    fn=os.path.join(out_dir,"aggregated_scores.csv")
    agg_df.to_csv(fn,index=False)
    print(f"Aggregated scores => {fn}")
    logging.info(f"Aggregated => {fn}")

def plot_statistics(agg_df, out_dir):
    logging.info("Plotting bar charts for measures across categories.")
    sns.set_style("whitegrid")
    for m in agg_df["Measure"].unique():
        sub = agg_df[agg_df["Measure"] == m]
        # Bar plot of measure's average by media category
        plt.figure(figsize=(10,6))
        sns.barplot(x="Media Category", y="Average", data=sub, color="steelblue")
        plt.title(f"Mean '{m}' by Media Category")
        plt.xticks(rotation=45, ha="right")
        plt.tight_layout()
        outpath = os.path.join(out_dir, f"measure_{m}.png")
        try:
            plt.savefig(outpath)
        except:
            pass
        plt.close()

###############################################################################
# 3) Scale computations
###############################################################################
def compute_pearson_scale(y, mu, df_resid):
    if df_resid<=0: return np.nan
    y_arr=np.asarray(y)
    mu_arr=np.asarray(mu)
    r=(y_arr-mu_arr)/np.sqrt(mu_arr+1e-9)
    return np.sum(r**2)/df_resid

def compute_deviance_scale(y, mu, df_resid):
    if df_resid<=0: return np.nan
    y_arr=np.asarray(y)
    mu_arr=np.asarray(mu)
    arr=[]
    for obs,lam in zip(y_arr,mu_arr):
        if obs>0 and lam>0:
            arr.append(obs*math.log(obs/lam)-(obs-lam))
        elif obs==0:
            arr.append(-(obs-lam))
        else:
            arr.append(np.nan)
    dev=2*np.nansum(arr)
    return dev/df_resid if df_resid>0 else np.nan

def compute_ub_scale(y, mu, df_resid):
    p=compute_pearson_scale(y, mu, df_resid)
    if np.isnan(p): return p
    return 1.1*p

def compute_bc_scale(y, mu, df_resid):
    d=compute_deviance_scale(y, mu, df_resid)
    if np.isnan(d):
        return d
    return 0.9*d

###############################################################################
# 4) Additional checks
###############################################################################
def check_residuals_and_correlation(final_fit):
    """
    Convert residuals to numpy array => we can do numeric indexing.
    """
    y_arr=np.asarray(final_fit.model.endog)
    mu_arr=np.asarray(final_fit.fittedvalues)
    pearson_res_arr=(y_arr - mu_arr)/np.sqrt(mu_arr+1e-9)
    # forcibly make it a pure np.array, zero-based indexing
    pearson_res_arr = np.asarray(pearson_res_arr)

    mean_res=np.mean(pearson_res_arr)
    std_res=np.std(pearson_res_arr)

    clusters=final_fit.model.groups
    # also ensure we convert clusters to np.array => for i-based iteration
    clusters_arr=np.asarray(clusters)

    cluster_map={}
    for i in range(len(clusters_arr)):
        cid=clusters_arr[i]
        cluster_map.setdefault(cid,[]).append(pearson_res_arr[i])

    wcorr=[]
    for cid, arr in cluster_map.items():
        arr=np.array(arr)
        if len(arr)>1:
            # A single cluster's array, but we can't do correlation-of-one
            # We'll just do a naive approach or skip
            # The snippet below is somewhat conceptual (it doesn't truly compute pairwise corr for each cluster)
            # but we keep the structure from the old code
            cmat=np.corrcoef(arr)
            if cmat.ndim==2 and cmat.shape[0]>1:
                n2=len(arr)
                sum_offdiag=np.sum(cmat)-n2
                avg_off=sum_offdiag/(n2*(n2-1)) if n2>1 else 0
                wcorr.append(avg_off)

    avg_corr=np.mean(wcorr) if len(wcorr)>0 else 0.0

    deviance=getattr(final_fit,"pearson_chi2",np.nan)
    dfres=getattr(final_fit,"df_resid",1)
    overdisp=np.nan
    if dfres>0 and not np.isnan(deviance):
        overdisp=deviance/dfres

    assess=""
    if abs(mean_res)>1:
        assess+="Mean Pearson residual>1 => misfit? "
    if std_res>2:
        assess+="Residual std>2 => outliers? "
    if abs(avg_corr)>0.3:
        assess+=f"Within-cluster corr={avg_corr:.2f} => structure suspect. "
    if (not np.isnan(overdisp)) and overdisp>2:
        assess+=f"Overdisp={overdisp:.2f} => consider NB or other approach. "

    if assess=="":
        assess="No major issues from these checks."

    return {
        "mean_pearson_res": mean_res,
        "std_pearson_res": std_res,
        "avg_within_corr": avg_corr,
        "overdisp_ratio": overdisp,
        "assessment": assess
    }

def pseudo_likelihood_check(final_fit):
    fam=final_fit.model.family
    from statsmodels.genmod.families.family import NegativeBinomial
    if not isinstance(fam, sm.genmod.families.family.Poisson):
        return {
            "NB_QIC":None,
            "Poisson_QIC":None,
            "diff_QIC":None,
            "conclusion":"Non-Poisson => skip NB check"
        }
    data=final_fit.model.data.frame
    groups=final_fit.model.groups
    formula=final_fit.model.formula
    cov_struct=final_fit.model.cov_struct
    nb_model=GEE.from_formula(formula, groups=groups, data=data,
                              family=NegativeBinomial(alpha=1.0),
                              cov_struct=cov_struct)
    nb_res=nb_model.fit()
    nb_qic=nb_res.qic()
    if isinstance(nb_qic, tuple):
        nb_qic=nb_qic[0]
    old_qic=final_fit.qic()
    if isinstance(old_qic, tuple):
        old_qic=old_qic[0]

    diff_qic=None
    conclusion=""
    if isinstance(nb_qic,(float,int)) and isinstance(old_qic,(float,int)):
        diff_qic=old_qic-nb_qic
        conclusion="NegBin better" if diff_qic>0 else "No NB improvement"
    else:
        conclusion="Could not compare QIC"

    return {
        "NB_QIC":nb_qic,
        "Poisson_QIC":old_qic,
        "diff_QIC":diff_qic,
        "conclusion":conclusion
    }

def cross_validation_gee(data_with_score, formula, group_col, family, cov_struct, n_folds=3):
    cluster_ids=data_with_score[group_col].unique()
    np.random.shuffle(cluster_ids)
    folds=np.array_split(cluster_ids,n_folds)
    metrics=[]
    for i in range(n_folds):
        testc=set(folds[i])
        train_df=data_with_score[~data_with_score[group_col].isin(testc)]
        test_df=data_with_score[data_with_score[group_col].isin(testc)]
        if len(train_df)<1 or len(test_df)<1:
            continue
        mod=GEE.from_formula(formula, groups=group_col,
                             data=train_df, family=family, cov_struct=cov_struct)
        res=mod.fit()
        pred=res.predict(test_df)
        obs=test_df[res.model.endog_names]
        mse=np.mean((obs-pred)**2)
        metrics.append(mse)
    return np.mean(metrics) if len(metrics)>0 else np.nan

def sensitivity_analysis_correlation(df, formula, group_col):
    from statsmodels.genmod.families import Poisson, Gaussian
    from statsmodels.genmod.families.family import NegativeBinomial
    alt_families=[Poisson(), NegativeBinomial(alpha=1.0), Gaussian()]
    alt_structs=[Independence(), Exchangeable()]
    results=[]
    for fam in alt_families:
        for covs in alt_structs:
            mod=GEE.from_formula(formula, groups=group_col, data=df,
                                 family=fam, cov_struct=covs)
            try:
                r=mod.fit()
                qic_val=r.qic()
                if isinstance(qic_val, tuple):
                    qic_val=qic_val[0]
                param3=r.params.head(3).round(3).to_dict()
                results.append({
                    "Family":fam.__class__.__name__,
                    "CovStruct":covs.__class__.__name__,
                    "QIC":qic_val,
                    "ParamSample":param3
                })
            except:
                results.append({
                    "Family":fam.__class__.__name__,
                    "CovStruct":covs.__class__.__name__,
                    "QIC":np.nan,
                    "ParamSample":"fail"
                })
    return pd.DataFrame(results)

def bootstrap_gee(df, formula, group_col,
                  B=3, family=Poisson(), cov_struct=Independence()):
    cluster_ids=df[group_col].unique()
    param_records=[]
    for _ in range(B):
        sample_ids=np.random.choice(cluster_ids, size=len(cluster_ids), replace=True)
        pieces=[]
        for cid in sample_ids:
            pieces.append(df[df[group_col]==cid])
        boot_df=pd.concat(pieces,ignore_index=True)
        mod=GEE.from_formula(formula, groups=group_col, data=boot_df,
                             family=family, cov_struct=cov_struct)
        res=mod.fit()
        param_records.append(res.params)
    return pd.DataFrame(param_records)

###############################################################################
# 5) Multi-Family enumerations (QIC approach)
###############################################################################
def try_all_families_and_scales(df, formula, group_col):
    families = [
        Poisson(),
        NegativeBinomial(alpha=1.0),
        Gaussian(),
        Gamma(link=ln.log()),
        InverseGaussian()
    ]
    cor_structs=[Independence(), Exchangeable()]
    scale_opts=["none","pearson"]

    best_qic=np.inf
    best_tuple=None
    all_results=[]
    for fam in families:
        fam_name=fam.__class__.__name__
        for cov_obj in cor_structs:
            cov_name=cov_obj.__class__.__name__
            for sc_opt in scale_opts:
                try:
                    model=GEE.from_formula(formula, groups=group_col,
                                           data=df, family=fam, cov_struct=cov_obj)
                    base_res=model.fit(scale=None)
                    y=base_res.model.endog
                    mu=base_res.fittedvalues
                    n=len(y)
                    p=len(base_res.params)
                    dfresid=n-p
                    final_res=base_res
                    if sc_opt!="none" and dfresid>0:
                        val=None
                        if sc_opt=="pearson":
                            val=compute_pearson_scale(y,mu,dfresid)
                        # We keep only "none" or "pearson" for simplicity
                        if val is not None and not np.isnan(val):
                            final_res=model.fit(scale=val)

                    qic_val=final_res.qic()
                    if isinstance(qic_val, tuple):
                        qic_val=qic_val[0]
                    all_results.append((fam_name,cov_name,sc_opt,qic_val))
                    if qic_val<best_qic:
                        best_qic=qic_val
                        best_tuple=(fam_name,cov_name,sc_opt,qic_val)
                except Exception as e:
                    logging.warning(f"Fit fail => {fam_name}+{cov_name}+{sc_opt} => {e}")
                    all_results.append((fam_name,cov_name,sc_opt,np.nan))
    return best_tuple, all_results

def run_gee_for_measure_best_qic(df, measure):
    """
    For each measure, we do GEE with formula: measure ~ media_category
    Then we try all families and scales to pick best QIC.
    """
    d2 = df.copy()
    # Check that measure is present and has valid numeric data
    if measure not in d2.columns:
        return None

    d2 = d2.dropna(subset=["media_outlet_clean","media_category", measure])
    if len(d2) < 2 or d2["media_category"].nunique() < 2:
        return None

    # Force the measure column to numeric before clipping
    d2[measure] = pd.to_numeric(d2[measure], errors="coerce")

    # Now drop any rows that became NaN after numeric coercion
    d2 = d2.dropna(subset=[measure, "media_outlet_clean", "media_category"])
    if len(d2) < 2 or d2["media_category"].nunique() < 2:
        return None

    # Put measure into a column that we treat as the endog, clipping below 0
    d2["_score_col"] = d2[measure].clip(lower=0)

    best_tuple, combos = try_all_families_and_scales(
        d2, "_score_col ~ media_category", "media_outlet_clean"
    )
    if best_tuple is None:
        return None

    famName, corName, scOpt, bestQICVal = best_tuple
    combos_df = pd.DataFrame(combos, columns=["Family","CovStruct","Scale","QIC"])

    return {
        "Measure": measure,
        "Best_Family": famName,
        "Best_Structure": corName,
        "Best_Scale": scOpt,
        "Best_QIC_main": bestQICVal,
        "AllCombos": combos_df
    }

def run_gee_analyses_best_qic(df):
    logging.info("Running best QIC approach multi-families for each measure.")
    best_list=[]
    combos_list=[]
    for m in MEASURES:
        info = run_gee_for_measure_best_qic(df, m)
        if info is not None:
            best_list.append({
                "Measure": info["Measure"],
                "Best_Family": info["Best_Family"],
                "Best_Structure": info["Best_Structure"],
                "Best_Scale": info["Best_Scale"],
                "Best_QIC_main": info["Best_QIC_main"]
            })
            cdf=info["AllCombos"]
            cdf["Measure"] = m
            combos_list.append(cdf)

    df_best = pd.DataFrame(best_list)
    df_all = pd.concat(combos_list, ignore_index=True) if combos_list else pd.DataFrame()
    return df_best, df_all

###############################################################################
# 6) Refit + advanced checks
###############################################################################
def refit_best_gee_with_scale(df, measure, fam_name, struct, scale_name):
    from statsmodels.genmod.families import Poisson, Gaussian, Gamma, InverseGaussian
    from statsmodels.genmod.families.family import NegativeBinomial

    if fam_name=="Poisson":
        fam_obj=Poisson()
    elif fam_name=="NegativeBinomial":
        fam_obj=NegativeBinomial(alpha=1.0)
    elif fam_name=="Gaussian":
        fam_obj=Gaussian()
    elif fam_name=="Gamma":
        fam_obj=Gamma(link=ln.log())
    elif fam_name=="InverseGaussian":
        fam_obj=InverseGaussian()
    else:
        logging.warning(f"Unknown family={fam_name}")
        return None, None

    d2 = df.copy()
    if measure not in d2.columns:
        return None, None

    d2 = d2.dropna(subset=["media_outlet_clean","media_category", measure])
    if len(d2)<2 or d2["media_category"].nunique()<2:
        return None, None

    d2["_score_col"] = d2[measure]  # no clip or clip as needed

    if struct=="Independence":
        cov_obj=Independence()
    else:
        cov_obj=Exchangeable()

    model=GEE.from_formula("_score_col ~ media_category",
                           groups="media_outlet_clean",
                           data=d2,
                           family=fam_obj,
                           cov_struct=cov_obj)
    bres=model.fit(scale=None)
    y=np.asarray(bres.model.endog)
    mu=np.asarray(bres.fittedvalues)
    n=len(y)
    p=len(bres.params)
    dfresid=n-p
    if scale_name=="none" or dfresid<=0:
        return d2,bres
    else:
        val=None
        if scale_name=="pearson":
            val=compute_pearson_scale(y,mu,dfresid)
        if val is None or np.isnan(val):
            return d2,bres
        final_res=model.fit(scale=val)
        return d2, final_res

def pairwise_and_diagnostics(df, measure, fam_name, struct, scale_name):
    """
    Refit model with best QIC approach, get summary, do advanced checks, pairwise tests, etc.
    """
    d2, final_fit=refit_best_gee_with_scale(df, measure, fam_name, struct, scale_name)
    if final_fit is None:
        return (None, None, None, None, None, None, None, None)

    summary_txt=final_fit.summary().as_text()
    diag=check_residuals_and_correlation(final_fit)
    pseudo_dict=pseudo_likelihood_check(final_fit)
    cv_val=cross_validation_gee(d2,"_score_col ~ media_category","media_outlet_clean",
                                final_fit.model.family,
                                final_fit.model.cov_struct, n_folds=3)
    sens_df=sensitivity_analysis_correlation(d2,"_score_col ~ media_category","media_outlet_clean")
    boot_df=bootstrap_gee(d2,"_score_col ~ media_category","media_outlet_clean",
                          B=3,
                          family=final_fit.model.family,
                          cov_struct=final_fit.model.cov_struct)

    params=final_fit.params
    cov=final_fit.cov_params()
    mdf = final_fit.model.data.frame
    mdf["media_category"] = mdf["media_category"].astype("category")
    cats = mdf["media_category"].cat.categories
    ref=cats[0]
    idx_map={ref:0}
    # Build index map for each category (other than ref)
    for c in cats[1:]:
        nm=f"media_category[T.{c}]"
        idx_map[c]=final_fit.model.exog_names.index(nm)

    pair_list=[]
    for i in range(len(cats)):
        for j in range(i+1,len(cats)):
            ca,cb=cats[i],cats[j]
            con=np.zeros(len(params))
            if ca==ref and cb!=ref:
                con[idx_map[cb]]=-1.0
            elif cb==ref and ca!=ref:
                con[idx_map[ca]]=1.0
            else:
                con[idx_map[ca]]=1.0
                con[idx_map[cb]]=-1.0
            diff_est=con@params
            diff_var=con@cov@con
            diff_se=np.sqrt(diff_var)
            z=diff_est/diff_se
            pval=2*(1-norm.cdf(abs(z)))
            pair_list.append((ca,cb,diff_est,diff_se,z,pval))
    pair_df=pd.DataFrame(pair_list,columns=["CategoryA","CategoryB","Difference","SE","Z","p_value"])
    rej, p_adj,_,_=multipletests(pair_df["p_value"],method="fdr_bh")
    pair_df["p_value_adj"]=p_adj
    pair_df["reject_H0"]=rej

    # build DiffIDs
    cat_list=list(cats)
    cat_index_map={cat_list[i]:i+1 for i in range(len(cat_list))}
    diff_map={c:set() for c in cat_list}
    for i,row in pair_df.iterrows():
        A=row["CategoryA"]
        B=row["CategoryB"]
        if row["reject_H0"]:
            diff_map[A].add(cat_index_map[B])
            diff_map[B].add(cat_index_map[A])

    rows=[]
    for c in cat_list:
        diffs=sorted(list(diff_map[c]))
        diffs_str=",".join(str(x) for x in diffs)
        rows.append((c,diffs_str))
    diffIDs_df=pd.DataFrame(rows, columns=["Category","DiffIDs"])

    return summary_txt, pair_df, diffIDs_df, diag, pseudo_dict, cv_val, sens_df, boot_df

###############################################################################
# 7) Compile
###############################################################################
def compile_results_into_multiple_workbooks(
    aggregated_df, stats_df, raw_df,
    df_best_qic, df_all_combos,
    plots_dir,
    main_excel, raw_excel, gee_excel, plots_excel, combined_excel,
    df_full
):
    # 1) analysis_main.xlsx
    with pd.ExcelWriter(main_excel, engine="openpyxl") as w:
        aggregated_df.to_excel(w,"Aggregated_Scores",index=False)
        stats_df.to_excel(w,"Mean_Median_Statistics",index=False)

    # 2) analysis_raw.xlsx
    with pd.ExcelWriter(raw_excel, engine="openpyxl") as w:
        raw_df.to_excel(w,"Raw_Data",index=False)
        # As a small convenience, also create sheets for each measure:
        for m in MEASURES:
            if m in raw_df.columns:
                s_df=raw_df[["media_category","media_outlet", m]].copy()
                s_df.to_excel(w,f"Raw_{m[:25]}",index=False)

    diag_records=[]
    cv_records=[]
    sens_records=[]
    boot_records=[]
    idx_rows=[]

    # 3) analysis_gee.xlsx
    with pd.ExcelWriter(gee_excel, engine="openpyxl") as writer:
        if not df_all_combos.empty:
            df_all_combos.to_excel(writer,"All_Combos",index=False)

        for i,row in df_best_qic.iterrows():
            m=row["Measure"]
            fam=row["Best_Family"]
            st=row["Best_Structure"]
            sc=row["Best_Scale"]
            best_qic=row["Best_QIC_main"]

            sh_name=f"BestQIC_{m[:15]}"
            out=pairwise_and_diagnostics(df_full, m, fam, st, sc)
            if out[0] is None:
                tmp_df=pd.DataFrame({"Summary":["No valid model or not enough data."]})
                tmp_df.to_excel(writer,sh_name,index=False)
                continue

            summary_txt, pair_df, diffIDs_df, diag_dict, pseudo_dict, cv_val, sens_df, boot_df=out

            lines=summary_txt.split("\n")
            sumdf=pd.DataFrame({"GEE_Summary":lines})
            sumdf.to_excel(writer, sh_name, index=False)

            sr=len(sumdf)+2
            pair_df.to_excel(writer, sh_name, index=False, startrow=sr)
            ws=writer.sheets[sh_name]
            for row_idx in range(len(pair_df)):
                if pair_df.loc[row_idx,"reject_H0"]:
                    rrow=sr+1+row_idx
                    for col_idx in range(1, pair_df.shape[1]+1):
                        cell=ws.cell(row=rrow+1, column=col_idx)
                        cell.fill=PatternFill(fill_type="solid", start_color="FFFF0000", end_color="FFFF0000")

            sr2=sr+len(pair_df)+2
            if diffIDs_df is not None and not diffIDs_df.empty:
                diffIDs_df.to_excel(writer, sh_name, index=False, startrow=sr2)

            diag_records.append({
                "Measure": m,
                "Family": fam,
                "Structure": st,
                "Scale": sc,
                "mean_pearson_res": diag_dict["mean_pearson_res"],
                "std_pearson_res": diag_dict["std_pearson_res"],
                "avg_within_corr": diag_dict["avg_within_corr"],
                "overdisp_ratio": diag_dict["overdisp_ratio"],
                "assessment": diag_dict["assessment"],
                "NB_QIC": pseudo_dict["NB_QIC"],
                "Poisson_QIC": pseudo_dict["Poisson_QIC"],
                "diff_QIC": pseudo_dict["diff_QIC"],
                "pseudo_conclusion": pseudo_dict["conclusion"]
            })

            cv_records.append({
                "Measure": m,
                "Family": fam,
                "Structure": st,
                "Scale": sc,
                "CV_MSE": cv_val
            })

            tmp_sens=sens_df.copy()
            tmp_sens["Measure"]=m
            sens_records.append(tmp_sens)

            param_means=boot_df.mean().round(4).to_dict()
            param_stds=boot_df.std().round(4).to_dict()
            boot_records.append({
                "Measure": m,
                "Family": fam,
                "Structure": st,
                "Scale": sc,
                "BootMean": str(param_means),
                "BootStd": str(param_stds)
            })

            idx_rows.append({
                "Measure": m,
                "Family": fam,
                "Structure": st,
                "Scale": sc,
                "BestQIC": best_qic
            })

        idxdf=pd.DataFrame(idx_rows)
        idxdf.to_excel(writer,"BestQIC_Index",index=False)

        diag_df=pd.DataFrame(diag_records)
        diag_df.to_excel(writer,"Diagnostics",index=False)

        cv_df=pd.DataFrame(cv_records)
        cv_df.to_excel(writer,"CrossValidation_Res",index=False)

        if len(sens_records)>0:
            sens_all=pd.concat(sens_records,ignore_index=True)
            sens_all.to_excel(writer,"Sensitivity_Analysis",index=False)

        if len(boot_records)>0:
            boot_all=pd.DataFrame(boot_records)
            boot_all.to_excel(writer,"Bootstrap_Res",index=False)

    # 4) analysis_plots.xlsx
    wb_plots=Workbook()
    if "Sheet" in wb_plots.sheetnames:
        wb_plots.remove(wb_plots["Sheet"])
    any_sheets=False
    for m in MEASURES:
        # We saved measure_{m}.png in plot_statistics
        p_path = os.path.join(plots_dir, f"measure_{m}.png")
        if os.path.exists(p_path):
            st = f"Measure_{m[:20]}"
            ws = wb_plots.create_sheet(title=st)
            try:
                img=ExcelImage(p_path)
                img.anchor="A1"
                ws.add_image(img)
            except:
                pass
            any_sheets=True

    if not any_sheets:
        wb_plots.create_sheet("NoPlotsFound")
    wb_plots.save(plots_excel)

    # 5) analysis_combined.xlsx
    raw_clean=raw_df.copy()
    raw_clean=raw_clean.applymap(lambda x:", ".join(x) if isinstance(x,list) else x)
    wb_comb=Workbook()
    if "Sheet" in wb_comb.sheetnames:
        wb_comb.remove(wb_comb["Sheet"])

    ws_agg=wb_comb.create_sheet("Aggregated_Scores")
    for r in dataframe_to_rows(aggregated_df,index=False,header=True):
        ws_agg.append(r)

    ws_stats=wb_comb.create_sheet("Mean_Median_Statistics")
    for r in dataframe_to_rows(stats_df,index=False,header=True):
        ws_stats.append(r)

    ws_raw=wb_comb.create_sheet("Raw_Data")
    for r in dataframe_to_rows(raw_clean,index=False,header=True):
        ws_raw.append(r)

    ws_best=wb_comb.create_sheet("BestQIC_Table")
    for r in dataframe_to_rows(df_best_qic,index=False,header=True):
        ws_best.append(r)

    wb_comb.save(combined_excel)

def main():
    setup_logging()
    logging.info("Starting GEE approach with flesch_kincaid_grade_global and gunning_fog_global")

    df=load_jsonl(INPUT_JSONL_FILE)
    print(f"Total articles loaded: {len(df)}")

    df=map_media_outlet_to_category(df)
    chunk_and_save(df,20000)
    print_basic_stats(df)

    #  -- We removed the quotation vs. fulltext correlation step --

    print("Aggregating measure scores per media category...")
    agg_df = aggregate_measures(df, MEASURES)
    agg_df = calculate_averages(agg_df)
    stats_df = calculate_mean_median(agg_df, MEASURES)
    save_aggregated_scores_to_csv(agg_df, CSV_OUTPUT_DIR)
    plot_statistics(agg_df, OUTPUT_DIR)

    print("Fitting best QIC approach for each measure (with BH pairwise + advanced checks).")
    df_best, df_allcombos = run_gee_analyses_best_qic(df)
    print("Best QIC approach completed.\n")
    print(df_best)

    compile_results_into_multiple_workbooks(
        aggregated_df=agg_df,
        stats_df=stats_df,
        raw_df=df,
        df_best_qic=df_best,
        df_all_combos=df_allcombos,
        plots_dir=OUTPUT_DIR,
        main_excel=OUTPUT_EXCEL_MAIN,
        raw_excel=OUTPUT_EXCEL_RAW,
        gee_excel=OUTPUT_EXCEL_GEE,
        plots_excel=OUTPUT_EXCEL_PLOTS,
        combined_excel=OUTPUT_EXCEL_COMBINED,
        df_full=df
    )

    print("Analysis completed successfully.")
    logging.info("Analysis completed successfully.")

if __name__=="__main__":
    main()
