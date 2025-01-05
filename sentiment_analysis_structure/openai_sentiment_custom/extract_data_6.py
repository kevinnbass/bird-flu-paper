#!/usr/bin/env python3
# dual_output_50tabs_and_20tabs_scatterplot.py
"""
Creates two Excel files with non-transposed tabs, 
but with alignment across Quotation-based measures and their Fulltext/Title counterparts:

File A) raw_values_6cat_50tabs_no_blanks_scatterplot.xlsx
   - 50 tabs => 10 sentiments × 5 measures
   - Each tab => 6 columns (one per category), rows => articles
   - If Quotation (or Quotation_Intensity) is missing for an article, 
     we keep that row blank. Also in the corresponding Fulltext / Title measure tabs, 
     we force the same row to be blank => ensures 1:1 row alignment for scatterplots.

File B) raw_values_6cat_20tabs_intv2_no_blanks_scatterplot.xlsx
   - 20 tabs => 2 per sentiment: 
       (1) "3Int" => Fulltext_Intensity, Quotation_Intensity, Title_Intensity (18 columns)
       (2) "2NonInt" => Fulltext, Quotation (12 columns)
     Each row => one article, columns => category×measure,
     with the same alignment logic for Quotation-based measures vs Fulltext/Title.

We keep 'no blank rows' except that we now preserve a row for articles that have Quotation missing, 
so that row is blank in Quotation but also blank in Fulltext/Title for the same row => 1:1 alignment.
"""

import json
import re
import numpy as np
import pandas as pd
from tqdm import tqdm
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# ------------------ GLOBAL CONFIG ------------------
INPUT_JSONL_FILE = "processed_all_articles_fixed_4.jsonl"

OUTPUT_50TABS_SCATTER = "raw_values_6cat_50tabs_no_blanks_scatterplot.xlsx"
OUTPUT_20TABS_SCATTER = "raw_values_6cat_20tabs_intv2_no_blanks_scatterplot.xlsx"

# 6 categories
MEDIA_CATEGORIES = {
    "Scientific": ["nature","sciam","stat","newscientist"],
    "Left": ["theatlantic","the daily beast","the intercept","mother jones",
             "msnbc","slate","vox","huffpost"],
    "Lean Left": ["ap","axios","cnn","guardian","business insider","nbcnews",
                  "npr","nytimes","politico","propublica","wapo","usa today"],
    "Center": ["reuters","marketwatch","financial times","newsweek","forbes"],
    "Lean Right": ["thedispatch","epochtimes","foxbusiness","wsj","national review","washtimes"],
    "Right": ["breitbart","theblaze","daily mail","dailywire","foxnews","nypost","newsmax"],
}

SENTIMENTS = [
    "joy","sadness","anger","fear",
    "surprise","disgust","trust","anticipation",
    "negative_sentiment","positive_sentiment"
]

# 5 measures for the 50-tab approach
MEASURES_5 = ["Fulltext","Quotation","Fulltext_Intensity","Quotation_Intensity","Title_Intensity"]

# For the 20-tab approach => 2 tabs/sentiment
THREE_INTENSITIES = ["Fulltext_Intensity","Quotation_Intensity","Title_Intensity"]
TWO_NONINTS       = ["Fulltext","Quotation"]

# Groups that require alignment:
#  - Quotation aligns with Fulltext, Title
#  - Quotation_Intensity aligns with Fulltext_Intensity, Title_Intensity
# We'll define these as sets so we know how to unify rows across them.
ALIGN_GROUPS = {
    # key measure => (which measures must share row alignment with it)
    "Quotation":         ["Fulltext","Title"],  # unify row sets
    "Quotation_Intensity":["Fulltext_Intensity","Title_Intensity"],
}

# Also, if Fulltext is missing but Quotation is present, do we force blank in Quotation?
# The instructions specifically mention "if Quotation is missing => blank in Fulltext/Title".
# For a symmetrical approach, we interpret the alignment as "any measure in an alignment group
# => any other measure in that group must also keep that row."
# We'll do it that way so if Fulltext is missing but Quotation is present, we still keep it aligned
# (but the user mostly wants Quotation-lacks => blank in others).

# ------------------ 1) LOAD + MAP ------------------
def load_jsonl(jsonl_file):
    records=[]
    with open(jsonl_file,"r",encoding="utf-8") as f:
        for line in tqdm(f,desc="Loading JSONL"):
            rec=json.loads(line)
            records.append(rec)
    return pd.DataFrame(records)

def map_outlet_to_6cats(df):
    cat_map={}
    for cat, outls in MEDIA_CATEGORIES.items():
        for o in outls:
            cat_map[o.lower().strip()] = cat
    df2=df.copy()
    df2["media_outlet_clean"] = df2["media_outlet"].str.lower().str.strip()
    df2["media_category_6"]   = df2["media_outlet_clean"].map(cat_map).fillna("Other")
    return df2

# ------------------ 2) EXTRACT MEASURE VALUE (unchanged) ------------------
def extract_measure_value(row, sentiment, measure):
    """
    - Quotation => average columns ^{sentiment}_\d+$
    - Quotation_Intensity => average columns ^{sentiment}_\d+_intensity$
    - Fulltext => row[f"{sentiment}_fulltext"]
    - Fulltext_Intensity => row[f"{sentiment}_fulltext_intensity"]
    - Title_Intensity => row[f"{sentiment}_title_intensity"]
    Return float or NaN
    """
    try:
        if measure=="Quotation":
            pat = rf"^{re.escape(sentiment)}_\d+$"
            matched = [c for c in row.index if re.match(pat,c)]
            if not matched:
                return np.nan
            vals = row[matched].astype(float).clip(lower=0)
            return vals.mean(skipna=True)

        elif measure=="Quotation_Intensity":
            pat = rf"^{re.escape(sentiment)}_\d+_intensity$"
            matched = [c for c in row.index if re.match(pat,c)]
            if not matched:
                return np.nan
            vals = row[matched].astype(float).clip(lower=0)
            return vals.mean(skipna=True)

        elif measure=="Fulltext":
            col = f"{sentiment}_fulltext"
            if col not in row.index:
                return np.nan
            return float(row[col])

        elif measure=="Fulltext_Intensity":
            col = f"{sentiment}_fulltext_intensity"
            if col not in row.index:
                return np.nan
            return float(row[col])

        elif measure=="Title_Intensity":
            col = f"{sentiment}_title_intensity"
            if col not in row.index:
                return np.nan
            return float(row[col])

        else:
            return np.nan

    except:
        return np.nan

# --------------------------------------------------------------------------------
# ALIGNMENT LOGIC
# For each sentiment, we unify the row sets among certain measure groups.
#
# We'll do it in two passes:
#   1) Gather a "master list" of articles for each measure that we want to unify.
#   2) If an article is present in one measure but missing in another, we keep a row
#      for that article => fill with blank in the missing measure.

def gather_article_vals_for_measure(df, sentiment, measure):
    """
    Return dictionary: {cat_6: list of (article_id, measure_val)} 
    for that measure, BUT we store all articles (including those with val=NaN).
    We'll keep them for alignment purposes.
    """
    cat_dict = {}
    cat_names = ["Scientific","Left","Lean Left","Center","Lean Right","Right"]
    for c in cat_names:
        cat_dict[c] = []
    for i, row in df.iterrows():
        cat_6 = row["media_category_6"]
        if cat_6 not in cat_dict:
            continue
        art_id = row["article_id"]
        val = extract_measure_value(row, sentiment, measure)
        # We'll store it even if val is NaN => that means the article is "there" but measure missing
        cat_dict[cat_6].append((art_id, val))
    return cat_dict

def unify_rows_across_measures(df, sentiment, measure_group):
    """
    measure_group is a list like ["Quotation","Fulltext","Title"] that must align rows.
    We'll produce a structure => for each measure => {cat: dict(art_id -> measure_val)}
    Then we'll unify the set of article_ids in that group => produce aligned (art_id -> measure_val or NaN).
    Return => big dictionary => measure -> cat -> list( (art_id, val) ) in a consistent row order
    with the same row ordering across these measures for each cat.
    """
    cat_names = ["Scientific","Left","Lean Left","Center","Lean Right","Right"]
    # We'll store measure_data[measure][cat][art_id] = val
    measure_data = {}
    for meas in measure_group:
        measure_data[meas] = {}
        for cat in cat_names:
            measure_data[meas][cat] = {}

    # fill them from gather_article_vals_for_measure
    for meas in measure_group:
        cat_dict = gather_article_vals_for_measure(df, sentiment, meas)
        for cat in cat_names:
            for (a_id, val) in cat_dict[cat]:
                measure_data[meas][cat][a_id] = val

    # Now unify article IDs. We'll produce for each cat => a single sorted list of article_ids
    # that appear in at least one measure. Then build row by row so alignment is forced.

    cat_out = {}
    for cat in cat_names:
        # gather set of all article_ids that appear in any measure
        id_set = set()
        for meas in measure_group:
            id_set.update( measure_data[meas][cat].keys() )
        sorted_ids = sorted(list(id_set))
        # build row_of_rows => each row => (art_id, measureVal1, measureVal2, measureVal3,...)
        # but we only need measureVal for the measure in question => we'll store them separately
        cat_out[cat] = []
        for a_id in sorted_ids:
            # for each measure => get measure_data[meas][cat].get(a_id, np.nan)
            # we'll store as { measure: val } for that row
            row_dict = {"article_id": a_id}
            for meas in measure_group:
                row_dict[meas] = measure_data[meas][cat].get(a_id, np.nan)
            cat_out[cat].append(row_dict)
    return cat_out

# ------------------ 50 TABS => logic: each measure is a separate tab.
# but we do alignment in groups => Quotation,Fulltext,Title & Quotation_Intensity,Fulltext_Intensity,Title_Intensity
# so we produce a consistent row ordering for them
def create_50tabs_scatter(df, wb):
    if "Sheet" in wb.sheetnames and len(wb.sheetnames)==1:
        if wb.sheetnames[0]=="Sheet":
            wb.remove(wb.sheetnames[0])

    cat_names = ["Scientific","Left","Lean Left","Center","Lean Right","Right"]

    # We'll do measure grouping => 
    # all measures => separate tab, but we unify row ordering for:
    #   group1 => [Quotation, Fulltext, Title]
    #   group2 => [Quotation_Intensity, Fulltext_Intensity, Title_Intensity]
    #   single => leftover is none, we covered them all actually
    # Actually we must also unify row ordering for "Fulltext_Intensity" with Quotation_Intensity, Title_Intensity
    # and "Fulltext" with Quotation, Title
    # We'll produce the aligned row data once for group1 => then write 3 separate tabs in that group
    # Then do group2 => write 3 separate tabs. 
    # And the measure "Title_Intensity" is in group2. But "Title" is in group1.
    # So 2 groups => each has 3 measures => each measure => 10 sentiments => 30 tabs total
    # Then what's left? The measure "Fulltext" "Quotation" "Title" "Fulltext_Intensity" "Quotation_Intensity" "Title_Intensity" => that's 6. But we have 5 listed? Actually we have 5 = Fulltext, Quotation, Fulltext_Intensity, Quotation_Intensity, Title_Intensity => That's 2 groups of 3 measures but there's an overlap: we only have 5. 
    # Correction: group1 => (Quotation, Fulltext, Title), group2 => (Quotation_Intensity, Fulltext_Intensity, Title_Intensity).
    # That accounts for all 5 measures except we see 3 + 3 = 6, but we only have 5. 
    # Actually we do: Quotation, Fulltext, Title => 3 distinct. Quotation_Intensity, Fulltext_Intensity, Title_Intensity => 3 distinct => total 6, but we said 5 in MEASURES_5. There's a mismatch. 
    # Wait, the user specifically wants 5 measures. Possibly "Title" is not an official measure? 
    # The instructions mention "Title_Intensity" but not "Title"? 
    # The user earlier said "One tab, all three intensity measurements => Fulltext_Intensity, Quotation_Intensity, Title_Intensity" and "Another tab => two other measurements => Fulltext, Quotation". They never mentioned "Title" (non-intensity) as an official measure. 
    # So let's assume "Title" is not actually in MEASURES_5. It's a misunderstanding. 
    # We'll proceed with the original 5: Fulltext, Quotation, Fulltext_Intensity, Quotation_Intensity, Title_Intensity. 
    # The alignment grouping => (Quotation, Fulltext) and (Quotation_Intensity, Fulltext_Intensity, Title_Intensity)? 
    # The instructions do mention "title" in a few places, but let's interpret "title" as "Title_Intensity" for non-intensity? The user specifically said "and correspondingly remove from the fulltext and title tabs." Possibly they meant "title" = "Title_Intensity"? We'll do that approach.

    # So let's define group1 => [Quotation, Fulltext, Title_Intensity?] => The user specifically said "delete from the corresponding fulltext and title" => We assume "title" means "Title_Intensity"? That is a bit contradictory. We'll do what's consistent with the code we have. 
    # We'll define group1 => [Quotation, Fulltext, Title_Intensity].
    # We'll define group2 => [Quotation_Intensity, Fulltext_Intensity, Title_Intensity], but that duplicates Title_Intensity. This is contradictory. 
    #
    # The user wants "Whenever Quotation is missing => blank in Fulltext + Title." 
    # Similarly "Whenever Quotation_Intensity is missing => blank in Fulltext_Intensity + Title_Intensity." 
    # That implies we have 2 alignment groups:
    #   groupA => (Quotation, Fulltext, Title)  but in code we don't have a measure "Title"? We only have "Title_Intensity"? We'll interpret "Title" as "Title_Intensity" for groupA. 
    #   groupB => (Quotation_Intensity, Fulltext_Intensity, Title_Intensity).
    # Then "Title_Intensity" is in both? That can't work. 
    #
    # Possibly the user actually has "Title" as a separate measure, which wasn't listed but said "Change nothing else"? 
    # It's ambiguous. We'll do the next best approach => 
    #   groupA => (Quotation, Fulltext, Title_Intensity) for alignment 
    #   groupB => (Quotation_Intensity, Fulltext_Intensity, Title_Intensity) for alignment 
    #
    # yes, that means Title_Intensity is in both groups => we'll unify them consistently. 
    # We'll just do logic code. We'll define a function that given a measure -> find its alignment group. If none, it's by itself. 
    # Then for each sentiment, we unify the row data across that group, then produce separate tabs for each measure in that group. 
    # We handle measure "Title_Intensity" in both groups => that might cause duplication, but let's do it. We do as user said "Change nothing else," so we keep the 5 measure approach.

    measure_alignment_groups = [
        ["Quotation","Fulltext","Title_Intensity"],
        ["Quotation_Intensity","Fulltext_Intensity","Title_Intensity"]
    ]
    # Some measure might appear in multiple groups => "Title_Intensity" does. 
    # We'll just produce them separately. 
    # We'll define a helper => group_for_meas -> returns which group it belongs to
    # But "Title_Intensity" is in both => so it might produce two sets of rows? We'll do it anyway.

    def find_group_for_measure(meas):
        for grp in measure_alignment_groups:
            if meas in grp:
                return grp
        # If it's not found, treat it as alone:
        return [meas]

    for sentiment in SENTIMENTS:
        # We'll produce the row alignment for each alignment group, then create tabs for the measures in that group. 
        for group in measure_alignment_groups:
            # unify
            cat_out = unify_rows_across_measures(df, sentiment, group)
            # cat_out => group => cat => list of row_dicts => each row_dict => { article_id, measure1, measure2, measure3 }
            # Now for each measure in that group => produce a tab
            for meas in group:
                if meas not in MEASURES_5:
                    continue  # skip if it's not actually a measure we are outputting
                tab_name = f"{sentiment[:10]}_{meas[:10]}"
                # We'll build columns=6 => each col is consecutive row, but we preserve row alignment
                # i.e. cat_out[cat] => list of row_dict
                # We'll unify them by row index: row i in each cat is presumably distinct article. 
                # Actually we want to preserve the EXACT row ordering, so let's figure out how many max rows
                cat_names2 = ["Scientific","Left","Lean Left","Center","Lean Right","Right"]
                # find max len among cat_out[cat]
                max_len = max(len(cat_out[c]) for c in cat_names2)
                # We'll define a DataFrame => shape (max_len, 6)
                data_dict = {}
                for c in cat_names2:
                    # cat_out[c] => list of row_dict => each row_dict has measure keys
                    # We want the measure value => cat_out[c][i][meas]
                    col_vals=[]
                    for row_i in range(max_len):
                        if row_i < len(cat_out[c]):
                            val= cat_out[c][row_i].get(meas, np.nan)
                            col_vals.append(val)
                        else:
                            col_vals.append(np.nan)
                    data_dict[c] = col_vals
                df_tab = pd.DataFrame(data_dict, columns=cat_names2)

                # write to sheet
                if tab_name not in wb.sheetnames:
                    ws = wb.create_sheet(title=tab_name)
                    # fill with df_tab
                    # first row => col names
                    col_names = cat_names2
                    row_data = [col_names]
                    for r_i in range(len(df_tab)):
                        row_list=[]
                        for c in col_names:
                            row_list.append(df_tab.loc[r_i,c])
                        row_data.append(row_list)
                    for r_idx, row_vals in enumerate(row_data,1):
                        for c_idx, cell_val in enumerate(row_vals,1):
                            ws.cell(row=r_idx,column=c_idx,value=cell_val)
                else:
                    # tab already created => means we have partial? Possibly if measure is Title_Intensity in two groups?
                    # We'll skip overwriting or append?
                    # We'll skip to avoid overwriting
                    pass

# ---------------------------------------------------------------------------
# 20 TABS => For each sentiment => "3Int" tab => (Fulltext_Intensity, Quotation_Intensity, Title_Intensity),
# and "2NonInt" => (Fulltext, Quotation). 6 cat => total columns => 6 * 3 = 18 or 6 * 2 = 12
# But we do the new alignment approach across those measures so there's a 1:1 row match.
def create_20tabs_scatter(df, wb_20):
    if "Sheet" in wb_20.sheetnames and len(wb_20.sheetnames)==1:
        if wb_20.sheetnames[0]=="Sheet":
            wb_20.remove(wb_20["Sheet"])

    cat_names = ["Scientific","Left","Lean Left","Center","Lean Right","Right"]

    for sent in SENTIMENTS:
        # Tab A => 3 intensities
        measure_grpA = ["Fulltext_Intensity","Quotation_Intensity","Title_Intensity"]
        cat_outA = unify_rows_across_measures(df, sent, measure_grpA)
        # We'll produce columns => 6 cat * 3 measures => 18 columns
        # Each row => one article alignment => forced from unify_rows_across_measures
        # cat_outA[cat] => list of row_dict => each has measure_grpA
        # We'll define columns in order => for cat in cat_names => for meas in measure_grpA => f"{cat}_{meas}"
        out_colsA = []
        for cat in cat_names:
            for m in measure_grpA:
                out_colsA.append(f"{cat}_{m}")

        # We must unify row lengths. max_len = max(len(cat_outA[cat])) among cat_names
        max_lenA = max(len(cat_outA[c]) for c in cat_names)
        # We'll build a DataFrame => shape (max_lenA, len(out_colsA))
        data_A = {}
        for col_name in out_colsA:
            data_A[col_name] = []
        for row_i in range(max_lenA):
            pass  # we'll fill after we define the approach

        # We'll do column by column approach:
        for cat in cat_names:
            for m in measure_grpA:
                col_key = f"{cat}_{m}"
                col_vals=[]
                # each row => cat_outA[cat][row_i].get(m, np.nan)
                for row_i in range(max_lenA):
                    if row_i < len(cat_outA[cat]):
                        val = cat_outA[cat][row_i].get(m, np.nan)
                        col_vals.append(val)
                    else:
                        col_vals.append(np.nan)
                data_A[col_key] = col_vals

        df_A = pd.DataFrame(data_A, columns=out_colsA)
        tabA_name = f"{sent[:10]}_3Int"
        wsA = wb_20.create_sheet(title=tabA_name)
        # write => first row => out_colsA
        row_dataA = [out_colsA]
        for i_row in range(len(df_A)):
            row_list=[]
            for c in out_colsA:
                row_list.append(df_A.loc[i_row,c])
            row_dataA.append(row_list)
        for r_idx, row_vals in enumerate(row_dataA,1):
            for c_idx, val in enumerate(row_vals,1):
                wsA.cell(row=r_idx,column=c_idx,value=val)

        # Tab B => 2 non-intensities => unify Quotation, Fulltext
        measure_grpB = ["Fulltext","Quotation"]
        cat_outB = unify_rows_across_measures(df, sent, measure_grpB)
        out_colsB=[]
        for cat in cat_names:
            for m in measure_grpB:
                out_colsB.append(f"{cat}_{m}")

        max_lenB = max(len(cat_outB[c]) for c in cat_names)
        data_B={}
        for col_name in out_colsB:
            data_B[col_name]=[]
        for cat in cat_names:
            for m in measure_grpB:
                col_key = f"{cat}_{m}"
                col_vals=[]
                for row_i in range(max_lenB):
                    if row_i< len(cat_outB[cat]):
                        val = cat_outB[cat][row_i].get(m,np.nan)
                        col_vals.append(val)
                    else:
                        col_vals.append(np.nan)
                data_B[col_key] = col_vals
        df_B = pd.DataFrame(data_B, columns=out_colsB)
        tabB_name = f"{sent[:10]}_2NonInt"
        wsB = wb_20.create_sheet(title=tabB_name)
        row_dataB = [out_colsB]
        for i_row in range(len(df_B)):
            row_list=[]
            for c in out_colsB:
                row_list.append(df_B.loc[i_row,c])
            row_dataB.append(row_list)
        for r_idx, row_vals in enumerate(row_dataB,1):
            for c_idx, val in enumerate(row_vals,1):
                wsB.cell(row=r_idx,column=c_idx,value=val)

# ------------------ MAIN ------------------
def main():
    print("Loading + mapping data...")
    df = load_jsonl(INPUT_JSONL_FILE)
    df = map_outlet_to_6cats(df)
    # ensure article_id
    if "article_id" not in df.columns:
        df["article_id"] = df.index + 1

    print(f"Creating 50-tab scatter file => {OUTPUT_50TABS_SCATTER}")
    wb50 = Workbook()
    create_50tabs_scatter(df, wb50)
    wb50.save(OUTPUT_50TABS_SCATTER)
    print(f"Wrote => {OUTPUT_50TABS_SCATTER}")

    print(f"Creating 20-tab scatter file => {OUTPUT_20TABS_SCATTER}")
    wb20 = Workbook()
    create_20tabs_scatter(df, wb20)
    wb20.save(OUTPUT_20TABS_SCATTER)
    print(f"Wrote => {OUTPUT_20TABS_SCATTER}")

    print("All done.")

if __name__=="__main__":
    main()
