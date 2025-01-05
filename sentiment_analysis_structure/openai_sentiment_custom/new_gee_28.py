import json
import pandas as pd
import matplotlib.pyplot as plt
import seaborn as sns
import os
import logging
from tqdm import tqdm
from openpyxl.drawing.image import Image as ExcelImage
import re
import sys
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import numpy as np
import patsy

import statsmodels.api as sm
from statsmodels.genmod.generalized_estimating_equations import GEE
from statsmodels.genmod.families import Poisson
from statsmodels.genmod.cov_struct import Exchangeable, Independence, Unstructured, Autoregressive
from statsmodels.stats.multitest import multipletests
from scipy.stats import norm, pearsonr

# ------------------------------ #
#        Configuration           #
# ------------------------------ #

INPUT_JSONL_FILE = 'processed_all_articles_merged.jsonl'
OUTPUT_DIR = 'graphs_analysis'
os.makedirs(OUTPUT_DIR, exist_ok=True)

CSV_OUTPUT_DIR = 'csv_raw_scores'
os.makedirs(CSV_OUTPUT_DIR, exist_ok=True)

OUTPUT_EXCEL_MAIN = 'analysis_main.xlsx'
OUTPUT_EXCEL_RAW = 'analysis_raw.xlsx'
OUTPUT_EXCEL_LMM = 'analysis_gee.xlsx'
OUTPUT_EXCEL_PLOTS = 'analysis_plots.xlsx'
OUTPUT_EXCEL_COMBINED = 'analysis_combined.xlsx'
OUTPUT_EXCEL_GRAPHING = 'analysis_for_graphing.xlsx'
OUTPUT_EXCEL_CORRELATION = 'analysis_correlation_data.xlsx'
OUTPUT_EXCEL_COMBINED_ALL = 'analysis_combined_all.xlsx'
OUTPUT_EXCEL_QIC = 'analysis_gee_qic.xlsx'

CATEGORIES = [
    'joy', 'sadness', 'anger', 'fear',
    'surprise', 'disgust', 'trust', 'anticipation',
    'negative_sentiment', 'positive_sentiment'
]

MEDIA_CATEGORIES = {
    'Scientific': ['nature', 'sciam', 'stat', 'newscientist'],
    'Left': ['theatlantic', 'the daily beast', 'the intercept', 'mother jones', 'msnbc', 'slate', 'vox', 'huffpost'],
    'Lean Left': ['ap', 'axios', 'cnn', 'guardian', 'business insider', 'nbcnews', 'npr', 'nytimes', 'politico', 'propublica', 'wapo', 'usa today'],
    'Center': ['reuters', 'marketwatch', 'financial times', 'newsweek', 'forbes'],
    'Lean Right': ['thedispatch', 'epochtimes', 'foxbusiness', 'wsj', 'national review', 'washtimes'],
    'Right': ['breitbart', 'theblaze', 'daily mail', 'dailywire', 'foxnews', 'nypost', 'newsmax'],
}

CATEGORY_ORDER = ["Scientific", "Left", "Lean Left", "Center", "Lean Right", "Right"]

def setup_logging(log_file='analysis.log'):
    logging.basicConfig(
        filename=log_file,
        level=logging.DEBUG,
        format='%(asctime)s - %(levelname)s - %(message)s'
    )
    logging.info("Logging initialized.")

def load_jsonl(jsonl_file):
    records = []
    with open(jsonl_file, 'r', encoding='utf-8') as f:
        for line in tqdm(f, desc="Loading JSONL data"):
            try:
                record = json.loads(line)
                records.append(record)
            except json.JSONDecodeError as e:
                logging.error(f"JSON decoding error: {e}")
    df = pd.DataFrame(records)
    return df

def map_media_outlet_to_category(df, media_categories):
    outlet_to_category = {}
    for category, outlets in media_categories.items():
        for outlet in outlets:
            outlet_to_category[outlet.lower().strip()] = category

    if 'media_outlet' not in df.columns:
        logging.error("'media_outlet' column not found in DataFrame.")
        raise KeyError("'media_outlet' column not found in DataFrame.")

    df['media_outlet_clean'] = df['media_outlet'].str.lower().str.strip()
    df['media_category'] = df['media_outlet_clean'].map(outlet_to_category)
    df['media_category'] = df['media_category'].fillna('Other')
    unmapped_outlets = df[df['media_category'] == 'Other']['media_outlet'].unique()
    if len(unmapped_outlets) > 0:
        logging.warning(f"Unmapped media outlets found: {unmapped_outlets}")
        print(f"Warning: The following media outlets were not mapped and categorized as 'Other': {unmapped_outlets}")
    return df

def compute_article_level_scores(df, sentiment_categories):
    for sentiment in sentiment_categories:
        pattern = f"^{re.escape(sentiment)}_\\d+$"
        matched_cols = [col for col in df.columns if re.match(pattern, col)]
        if matched_cols:
            df[f"{sentiment}_quotation_mean_article"] = df[matched_cols].clip(lower=0).mean(axis=1)
        else:
            df[f"{sentiment}_quotation_mean_article"] = np.nan

        fulltext_col = f"{sentiment}_fulltext"
        if fulltext_col in df.columns:
            df[f"{sentiment}_fulltext_article"] = df[fulltext_col].clip(lower=0)
        else:
            df[f"{sentiment}_fulltext_article"] = np.nan
    return df

def aggregate_sentiment_scores(df, sentiment_categories):
    aggregation = []
    for media_category in MEDIA_CATEGORIES.keys():
        mask = (df['media_category'] == media_category)
        subset = df[mask]

        for sentiment in sentiment_categories:
            pattern = f"^{re.escape(sentiment)}_\\d+$"
            matched_cols = [col for col in df.columns if re.match(pattern, col)]

            if matched_cols:
                quotation_sum = subset[matched_cols].clip(lower=0).sum(skipna=True).sum()
                quotation_count = subset[matched_cols].clip(lower=0).count().sum()
            else:
                quotation_sum = 0
                quotation_count = 0

            fulltext_col = f"{sentiment}_fulltext"
            if fulltext_col in df.columns:
                fulltext_sum = subset[fulltext_col].clip(lower=0).sum(skipna=True)
                fulltext_count = subset[fulltext_col].clip(lower=0).count()
            else:
                fulltext_sum = 0
                fulltext_count = 0

            aggregation.append({
                'Media Category': media_category,
                'Sentiment/Emotion': sentiment,
                'Quotation_Sum': quotation_sum,
                'Quotation_Count': quotation_count,
                'Fulltext_Sum': fulltext_sum,
                'Fulltext_Count': fulltext_count
            })

    aggregated_df = pd.DataFrame(aggregation)
    return aggregated_df

def calculate_averages(aggregated_df):
    aggregated_df['Quotation_Average'] = aggregated_df.apply(
        lambda row: row['Quotation_Sum'] / row['Quotation_Count'] if row['Quotation_Count'] > 0 else None, axis=1
    )
    aggregated_df['Fulltext_Average'] = aggregated_df.apply(
        lambda row: row['Fulltext_Sum'] / row['Fulltext_Count'] if row['Fulltext_Count'] > 0 else None, axis=1
    )
    return aggregated_df

def calculate_mean_median(aggregated_df):
    stats = []
    for sentiment in CATEGORIES:
        sentiment_data = aggregated_df[aggregated_df['Sentiment/Emotion'] == sentiment]

        quotation_avg = sentiment_data['Quotation_Average'].dropna()
        mean_quotation = quotation_avg.mean() if not quotation_avg.empty else None

        fulltext_avg = sentiment_data['Fulltext_Average'].dropna()
        mean_fulltext = fulltext_avg.mean() if not fulltext_avg.empty else None

        stats.append({
            'Sentiment/Emotion': sentiment,
            'Mean_Quotation_Average': mean_quotation,
            'Mean_Fulltext_Average': mean_fulltext
        })

    stats_df = pd.DataFrame(stats)
    return stats_df

def save_aggregated_scores_to_csv(aggregated_df, csv_output_dir, prefix='aggregated_sentiment_emotion_scores.csv'):
    csv_file = os.path.join(csv_output_dir, prefix)
    try:
        aggregated_df.to_csv(csv_file, index=False)
        print(f"Aggregated sentiment/emotion scores saved to '{csv_file}'.")
        logging.info(f"Aggregated sentiment/emotion scores saved to '{csv_file}'.")
    except Exception as e:
        print(f"Error saving aggregated scores to CSV: {e}")
        logging.error(f"Error saving aggregated scores to CSV: {e}")

def plot_statistics(aggregated_df, output_dir):
    sns.set_style("whitegrid")
    for sentiment in CATEGORIES:
        sentiment_data = aggregated_df[aggregated_df['Sentiment/Emotion'] == sentiment]

        # Mean Quotation-Based Plot
        plt.figure(figsize=(10, 6))
        sns.barplot(
            x='Media Category',
            y='Quotation_Average',
            data=sentiment_data,
            color='steelblue'
        )
        plt.title(f"Mean Quotation-Based '{sentiment.capitalize()}' Scores", fontsize=14)
        plt.xlabel('Media Category', fontsize=12)
        plt.ylabel('Mean Quotation-Based Average Score', fontsize=12)
        plt.xticks(rotation=45, ha='right')
        plt.tight_layout()
        quote_plot_filename = f"quote_{sentiment}.png"
        quote_plot_path = os.path.join(output_dir, quote_plot_filename)
        try:
            plt.savefig(quote_plot_path)
            plt.close()
            print(f"Quotation-Based '{sentiment}' scores plot saved to '{quote_plot_path}'.")
        except Exception as e:
            print(f"Error saving quotation-based '{sentiment}' scores plot: {e}")

        # Mean Fulltext-Based Plot
        plt.figure(figsize=(10, 6))
        sns.barplot(
            x='Media Category',
            y='Fulltext_Average',
            data=sentiment_data,
            color='darkorange'
        )
        plt.title(f"Mean Fulltext-Based '{sentiment.capitalize()}' Scores", fontsize=14)
        plt.xlabel('Media Category', fontsize=12)
        plt.ylabel('Mean Fulltext-Based Average Score', fontsize=12)
        plt.xticks(rotation=45, ha='right')
        plt.tight_layout()
        fulltext_plot_filename = f"fulltext_{sentiment}.png"
        fulltext_plot_path = os.path.join(output_dir, fulltext_plot_filename)
        try:
            plt.savefig(fulltext_plot_path)
            plt.close()
            print(f"Fulltext-Based '{sentiment}' scores plot saved to '{fulltext_plot_path}'.")
        except Exception as e:
            print(f"Error saving fulltext-based '{sentiment}' scores plot: {e}")

def filter_clusters_for_ar(model_df, group_col='media_outlet'):
    # AR(1) requires at least 2 observations per cluster
    counts = model_df.groupby(group_col).size()
    valid_clusters = counts[counts >= 2].index
    return model_df[model_df[group_col].isin(valid_clusters)].copy()

def fit_gee_and_pairwise(df, sentiment, measure_type='Quotation', cov_struct_type=Exchangeable,
                         global_min_date=None):
    if measure_type == 'Quotation':
        pattern = f"^{re.escape(sentiment)}_\\d+$"
        matched_cols = [col for col in df.columns if re.match(pattern, col)]
        if not matched_cols:
            return None
        temp_df = df.copy()
        temp_df[f'{sentiment}_quotation_mean'] = temp_df[matched_cols].clip(lower=0).mean(axis=1)
        score_col = f'{sentiment}_quotation_mean'
    else:
        fulltext_col = f"{sentiment}_fulltext"
        if fulltext_col not in df.columns:
            return None
        temp_df = df.copy()
        temp_df[f'{sentiment}_fulltext_clipped'] = temp_df[fulltext_col].clip(lower=0)
        score_col = f'{sentiment}_fulltext_clipped'

    columns_needed = [score_col, 'media_category', 'media_outlet', 'date']
    model_df = temp_df.dropna(subset=columns_needed).copy()
    if model_df['media_category'].nunique() < 2:
        return None

    # Convert date -> integer day index from earliest date
    model_df['int_date'] = (model_df['date'] - global_min_date).dt.days + 1
    # Ensure int_date is strictly int32
    model_df['int_date'] = model_df['int_date'].astype(np.int32)

    if len(model_df) < 2:
        return None

    model_df['media_category'] = model_df['media_category'].astype('category')
    time = None

    if cov_struct_type == Autoregressive:
        # Filter to clusters with >=2 obs
        model_df = filter_clusters_for_ar(model_df, 'media_outlet')
        if len(model_df) < 2:
            return None
        model_df = model_df.sort_values(['media_outlet', 'int_date'])

        # Re-check cluster sizes
        cts = model_df.groupby('media_outlet').size()
        if (cts < 2).any():
            return None

        # time is int_date, already int
        time = model_df['int_date'].values

        # Additional logging check:
        # We'll verify time is integer dtype and log if not.
        if not np.issubdtype(time.dtype, np.integer):
            # Log every row that's not int
            for i, val in enumerate(time):
                if not isinstance(val, (int, np.integer)):
                    logging.error(f"Non-integer time value found at index {i}: {val}")
            return None  # skip AR(1) if any invalid

        cov_struct = Autoregressive()
    else:
        cov_struct = cov_struct_type()

    family = Poisson()
    y, X = patsy.dmatrices(f"{score_col} ~ media_category", data=model_df, return_type='dataframe')
    groups = model_df["media_outlet"].values

    model = GEE(y, X, groups=groups, time=time, family=family, cov_struct=cov_struct)
    try:
        results = model.fit()
    except Exception as e:
        logging.error(f"Model failed with {cov_struct_type.__name__} for sentiment={sentiment}, measure={measure_type}: {e}")
        return None

    summary_text = results.summary().as_text()
    params = results.params
    cov = results.cov_params()

    all_categories = model_df['media_category'].cat.categories
    reference_category = all_categories[0]
    param_names = results.model.exog_names

    # Pairwise comps
    cat_to_param_index = {reference_category: 0}
    for cat in all_categories[1:]:
        pname = f"media_category[T.{cat}]"
        cat_to_param_index[cat] = param_names.index(pname)

    pairwise_results = []
    cats = list(all_categories)
    for i in range(len(cats)):
        for j in range(i+1, len(cats)):
            catA = cats[i]
            catB = cats[j]
            c = np.zeros(len(params))
            if catA == reference_category and catB != reference_category:
                c[cat_to_param_index[catB]] = -1.0
            elif catB == reference_category and catA != reference_category:
                c[cat_to_param_index[catA]] = 1.0
            else:
                c[cat_to_param_index[catA]] = 1.0
                c[cat_to_param_index[catB]] = -1.0

            diff_est = c @ params
            diff_var = c @ cov @ c
            diff_se = np.sqrt(diff_var)
            z = diff_est / diff_se
            p_value = 2 * (1 - norm.cdf(abs(z)))
            pairwise_results.append((catA, catB, diff_est, diff_se, z, p_value))

    pairwise_df = pd.DataFrame(pairwise_results,
                               columns=["CategoryA", "CategoryB", "Difference", "SE", "Z", "p_value"])
    reject, p_adj, _, _ = multipletests(pairwise_df['p_value'], method='holm')
    pairwise_df['p_value_adj'] = p_adj
    pairwise_df['reject_H0'] = reject

    try:
        qic_val = results.qic()
    except:
        qic_val = np.nan

    return {
        'GEE_Summary': summary_text,
        'Pairwise': pairwise_df,
        'QIC': qic_val,
        'Results': results
    }

def try_all_cov_structures(df, sentiment, measure_type, global_min_date):
    structures = [Independence, Exchangeable, Unstructured, Autoregressive]
    results_list = []
    for struct in structures:
        res = fit_gee_and_pairwise(df, sentiment, measure_type=measure_type,
                                   cov_struct_type=struct, global_min_date=global_min_date)
        if res is not None:
            results_list.append((struct.__name__, res['QIC'], res))
    return results_list

def run_gee_analyses(df):
    min_date = df['date'].dropna().min()
    if pd.isnull(min_date):
        print("No valid dates found; skipping GEE AR(1).")
    all_results = {}
    all_qic_results = []
    for sentiment in CATEGORIES:
        all_results[sentiment] = {}
        for measure_type in ['Quotation', 'Fulltext']:
            results_list = try_all_cov_structures(df, sentiment, measure_type, min_date)
            if len(results_list) == 0:
                continue
            final_list = []
            for (name, qicv, ress) in results_list:
                if not isinstance(qicv, (int, float, np.number)) or np.isnan(qicv):
                    qicv = np.nan
                final_list.append((name, qicv, ress))
            final_list = sorted(final_list, key=lambda x: (np.inf if pd.isna(x[1]) else x[1]))
            best = final_list[0]
            all_results[sentiment][measure_type] = best[2]

            for struct_name, qic_val, resobj in final_list:
                all_qic_results.append({
                    'Sentiment': sentiment,
                    'Measure': measure_type,
                    'Structure': struct_name,
                    'QIC': qic_val
                })

    qic_df = pd.DataFrame(all_qic_results)
    return all_results, qic_df

def compile_results_into_multiple_workbooks(aggregated_df, stats_df, raw_df, gee_results, qic_df, plots_dir,
                                            main_excel, raw_excel, lmm_excel, plots_excel, combined_excel, qic_excel):
    with pd.ExcelWriter(main_excel, engine='openpyxl') as writer:
        aggregated_df.to_excel(writer, sheet_name='Aggregated_Scores', index=False)
        stats_df.to_excel(writer, sheet_name='Mean_Median_Statistics', index=False)

    raw_df = raw_df.copy()
    raw_df['media_category'] = pd.Categorical(raw_df['media_category'],
                                              categories=CATEGORY_ORDER + ['Other'], ordered=True)
    raw_df = raw_df.sort_values(['media_category', 'media_outlet'])

    with pd.ExcelWriter(raw_excel, engine='openpyxl') as writer:
        raw_df.to_excel(writer, sheet_name='Raw_Data', index=False)
        for sentiment in CATEGORIES:
            sentiment_cols = [c for c in raw_df.columns if c.startswith(sentiment+'_')]
            sentiment_df = raw_df[['media_category', 'media_outlet'] + sentiment_cols].copy()
            sheet_name = f"Raw_{sentiment[:29]}"
            sentiment_df.to_excel(writer, sheet_name=sheet_name, index=False)

    with pd.ExcelWriter(lmm_excel, engine='openpyxl') as writer:
        summary_rows = []
        for sentiment in gee_results:
            for measure_type in gee_results[sentiment]:
                sheet_name = f"GEE_{sentiment[:20]}_{measure_type[:8]}"
                gee_summary = gee_results[sentiment][measure_type]['GEE_Summary']
                pairwise_df = gee_results[sentiment][measure_type]['Pairwise']

                summary_df = pd.DataFrame({'GEE_Summary': gee_summary.split('\n')})
                summary_df.to_excel(writer, sheet_name=sheet_name, index=False)
                startrow = len(summary_df) + 2
                pairwise_df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=startrow)

                summary_rows.append({
                    'Sentiment': sentiment,
                    'Measure': measure_type,
                    'SheetName': sheet_name
                })

        summary_index_df = pd.DataFrame(summary_rows)
        summary_index_df.to_excel(writer, sheet_name='GEE_Results_Index', index=False)

    with pd.ExcelWriter(qic_excel, engine='openpyxl') as writer:
        qic_df.to_excel(writer, sheet_name='QIC_Results', index=False)

    wb = Workbook()
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])

    # Insert plots
    for sentiment in CATEGORIES:
        quote_plot_path = os.path.join(plots_dir, f"quote_{sentiment}.png")
        if os.path.exists(quote_plot_path):
            stitle = f"Quote_{sentiment[:28]}"
            worksheet = wb.create_sheet(title=stitle)
            try:
                img = ExcelImage(quote_plot_path)
                img.anchor = 'A1'
                worksheet.add_image(img)
            except Exception as e:
                logging.error(f"Error embedding image '{quote_plot_path}' into Excel: {e}")

        fulltext_plot_path = os.path.join(plots_dir, f"fulltext_{sentiment}.png")
        if os.path.exists(fulltext_plot_path):
            stitle = f"Fulltext_{sentiment[:25]}"
            worksheet = wb.create_sheet(title=stitle)
            try:
                img = ExcelImage(fulltext_plot_path)
                img.anchor = 'A1'
                worksheet.add_image(img)
            except Exception as e:
                logging.error(f"Error embedding image '{fulltext_plot_path}' into Excel: {e}")

    wb.save(plots_excel)

    raw_df_cleaned = raw_df.copy()
    for col in raw_df_cleaned.columns:
        raw_df_cleaned[col] = raw_df_cleaned[col].apply(lambda x: ", ".join(x) if isinstance(x, list) else x)

    wb_combined = Workbook()
    if 'Sheet' in wb_combined.sheetnames:
        wb_combined.remove(wb_combined['Sheet'])

    ws_agg = wb_combined.create_sheet("Aggregated_Scores")
    for r in dataframe_to_rows(aggregated_df, index=False, header=True):
        ws_agg.append(r)

    ws_stats = wb_combined.create_sheet("Mean_Median_Statistics")
    for r in dataframe_to_rows(stats_df, index=False, header=True):
        ws_stats.append(r)

    ws_raw = wb_combined.create_sheet("Raw_Data")
    for r in dataframe_to_rows(raw_df_cleaned, index=False, header=True):
        ws_raw.append(r)

    ws_gee = wb_combined.create_sheet("GEE_Summaries")
    ws_gee.append(["Sentiment", "Measure", "GEE_Summary"])
    for sentiment in gee_results:
        for measure_type in gee_results[sentiment]:
            summary = gee_results[sentiment][measure_type]['GEE_Summary']
            ws_gee.append([sentiment, measure_type, summary])

    wb_combined.save(OUTPUT_EXCEL_COMBINED)

def create_analysis_for_graphing_file(df, sentiments, category_order, output_file):
    df = df.copy()
    df['media_category'] = pd.Categorical(df['media_category'],
                                          categories=category_order+['Other'], ordered=True)
    df = df[df['media_category'].isin(category_order)]

    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        for sentiment in sentiments:
            q_col = f"{sentiment}_quotation_mean_article"
            f_col = f"{sentiment}_fulltext_article"

            cat_data_quotation = {cat: [] for cat in category_order}
            for idx, row in df.iterrows():
                val = row[q_col] if q_col in df.columns else np.nan
                row_vals = {c: np.nan for c in category_order}
                if row['media_category'] in category_order:
                    row_vals[row['media_category']] = val
                for c in category_order:
                    cat_data_quotation[c].append(row_vals[c])
            quotation_df = pd.DataFrame(cat_data_quotation, columns=category_order)

            cat_data_fulltext = {cat: [] for cat in category_order}
            for idx, row in df.iterrows():
                val = row[f_col] if f_col in df.columns else np.nan
                row_vals = {c: np.nan for c in category_order}
                if row['media_category'] in category_order:
                    row_vals[row['media_category']] = val
                for c in category_order:
                    cat_data_fulltext[c].append(row_vals[c])
            fulltext_df = pd.DataFrame(cat_data_fulltext, columns=category_order)

            quotation_df.to_excel(writer, sheet_name=f"{sentiment}_Quotation", index=False)
            fulltext_df.to_excel(writer, sheet_name=f"{sentiment}_Fulltext", index=False)

def main():
    setup_logging()
    print("Single run with GEE and pairwise comparisons (Poisson family), plus AR(1) with int day index.")
    df = load_jsonl(INPUT_JSONL_FILE)
    print(f"Total articles loaded: {len(df)}")
    df = map_media_outlet_to_category(df, MEDIA_CATEGORIES)

    if 'date' not in df.columns:
        print("Warning: 'date' field not found. Can't do AR(1).")
    else:
        df['date'] = pd.to_datetime(df['date'], format='%Y-%m-%d', errors='coerce')

    df = compute_article_level_scores(df, CATEGORIES)

    chunk_size = 20000
    for i in range(0, len(df), chunk_size):
        chunk = df.iloc[i:i+chunk_size]
        out_name = f"raw_data_part_{i//chunk_size+1}.csv"
        out_path = os.path.join(CSV_OUTPUT_DIR, out_name)
        chunk.to_csv(out_path, index=False)
        print(f"Saved chunk {i//chunk_size+1} to {out_path}")

    print("\nSummary Statistics:")
    print("Total number of articles:", len(df))
    print("\nNumber of articles per outlet:")
    print(df['media_outlet_clean'].value_counts())
    print("\nNumber of articles per category:")
    print(df['media_category'].value_counts())
    print()

    print("Aggregating sentiment/emotion scores per media category...")
    aggregated_df = aggregate_sentiment_scores(df, CATEGORIES)
    aggregated_df = calculate_averages(aggregated_df)
    stats_df = calculate_mean_median(aggregated_df)

    save_aggregated_scores_to_csv(aggregated_df, CSV_OUTPUT_DIR)
    plot_statistics(aggregated_df, OUTPUT_DIR)

    # Correlation
    correlation_results = []
    correlation_scatter_data = {}
    for sentiment in CATEGORIES:
        sub_df = aggregated_df.loc[aggregated_df['Sentiment/Emotion']==sentiment].copy()
        sub_df = sub_df.loc[sub_df['Media Category'].isin(MEDIA_CATEGORIES.keys())]
        sub_df.dropna(subset=['Quotation_Average','Fulltext_Average'], inplace=True)
        if len(sub_df)>1:
            valcorr,_ = pearsonr(sub_df['Quotation_Average'], sub_df['Fulltext_Average'])
        else:
            valcorr = np.nan
        correlation_results.append({'Sentiment/Emotion':sentiment,'Correlation':valcorr})
        correlation_scatter_data[sentiment] = sub_df[['Media Category','Quotation_Average','Fulltext_Average']].copy()

        # Scatter
        plt.figure(figsize=(6,6))
        sns.scatterplot(x='Quotation_Average',y='Fulltext_Average', data=sub_df,
                        hue='Media Category', s=50)
        plt.title(f"Scatter: {sentiment.capitalize()} (Quotation vs Fulltext)")
        plt.xlabel("Quotation_Average")
        plt.ylabel("Fulltext_Average")
        plt.legend(bbox_to_anchor=(1.05,1), loc='upper left')
        plt.tight_layout()
        scat_out = os.path.join(OUTPUT_DIR,f"scatter_{sentiment}.png")
        plt.savefig(scat_out)
        plt.close()
        print(f"Scatter plot for {sentiment} saved to '{scat_out}'.")

    corr_df = pd.DataFrame(correlation_results)
    plt.figure(figsize=(10,6))
    sns.barplot(x='Sentiment/Emotion',y='Correlation',data=corr_df,color='gray')
    plt.title("Correlation Between Quotation and Fulltext Averages")
    plt.xticks(rotation=45,ha='right')
    plt.ylim(-1,1)
    plt.tight_layout()
    corr_out = os.path.join(OUTPUT_DIR,"correlation_quotation_fulltext.png")
    plt.savefig(corr_out)
    plt.close()
    print(f"Correlation plot saved to '{corr_out}'.")

    with pd.ExcelWriter(OUTPUT_EXCEL_CORRELATION, engine='openpyxl') as writer:
        corr_df.to_excel(writer, sheet_name='Correlation', index=False)
        for sentiment, sdata in correlation_scatter_data.items():
            sdata.to_excel(writer, sheet_name=f"{sentiment}_Data", index=False)

    # Combined all-sentiment
    all_list = []
    for st, sdata in correlation_scatter_data.items():
        dtemp = sdata.copy()
        dtemp['Sentiment/Emotion'] = st
        all_list.append(dtemp)
    combined_all = pd.concat(all_list, ignore_index=True)

    combined_all['Quotation_Z'] = combined_all.groupby('Sentiment/Emotion')['Quotation_Average'].transform(
        lambda x: (x - x.mean())/x.std() if x.std()!=0 else x - x.mean()
    )
    combined_all['Fulltext_Z'] = combined_all.groupby('Sentiment/Emotion')['Fulltext_Average'].transform(
        lambda x: (x - x.mean())/x.std() if x.std()!=0 else x - x.mean()
    )
    combined_all.to_excel(OUTPUT_EXCEL_COMBINED_ALL, index=False)

    cnon = combined_all.dropna(subset=['Quotation_Z','Fulltext_Z'])
    if len(cnon)>1:
        r_val, _ = pearsonr(cnon['Quotation_Z'], cnon['Fulltext_Z'])
    else:
        r_val = np.nan

    plt.figure(figsize=(10,8))
    sns.regplot(x='Quotation_Z',y='Fulltext_Z', data=cnon,
                scatter_kws={'color':'black'}, line_kws={'color':'red'})
    plt.title(f"All Sentiments Combined (Normalized)\nR = {r_val:.2f}")
    plt.xlabel("Normalized Quotation_Average (Z-score)")
    plt.ylabel("Normalized Fulltext_Average (Z-score)")
    plt.tight_layout()
    combo_out = os.path.join(OUTPUT_DIR,"combined_normalized_scatter.png")
    plt.savefig(combo_out)
    plt.close()
    print(f"Combined normalized scatter plot saved to '{combo_out}'.")

    print("No predictor columns. Fitting GEE models w/ AR(1) integer day index.")
    gee_results, qic_df = run_gee_analyses(df)
    print("GEE analysis completed.\n")

    compile_results_into_multiple_workbooks(
        aggregated_df, stats_df, df, gee_results, qic_df, OUTPUT_DIR,
        OUTPUT_EXCEL_MAIN, OUTPUT_EXCEL_RAW, OUTPUT_EXCEL_LMM, OUTPUT_EXCEL_PLOTS,
        OUTPUT_EXCEL_COMBINED, OUTPUT_EXCEL_QIC
    )

    create_analysis_for_graphing_file(df, CATEGORIES, CATEGORY_ORDER, OUTPUT_EXCEL_GRAPHING)
    print("Analysis completed successfully.")

if __name__=="__main__":
    main()
