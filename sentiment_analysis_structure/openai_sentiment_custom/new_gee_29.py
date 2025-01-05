import json
import pandas as pd
import matplotlib.pyplot as plt
import seaborn as sns
import os
import logging
from tqdm import tqdm
from openpyxl.drawing.image import Image as ExcelImage
import re
import sys
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import numpy as np
import patsy

import statsmodels.api as sm
from statsmodels.genmod.generalized_estimating_equations import GEE
from statsmodels.genmod.families import Poisson
from statsmodels.genmod.cov_struct import Exchangeable, Independence, Unstructured, Autoregressive
from statsmodels.stats.multitest import multipletests
from scipy.stats import norm, pearsonr

# ------------------------------ #
#        Configuration           #
# ------------------------------ #

INPUT_JSONL_FILE = 'processed_all_articles_merged.jsonl'
OUTPUT_DIR = 'graphs_analysis'
os.makedirs(OUTPUT_DIR, exist_ok=True)

CSV_OUTPUT_DIR = 'csv_raw_scores'
os.makedirs(CSV_OUTPUT_DIR, exist_ok=True)

OUTPUT_EXCEL_MAIN = 'analysis_main.xlsx'
OUTPUT_EXCEL_RAW = 'analysis_raw.xlsx'
OUTPUT_EXCEL_LMM = 'analysis_gee.xlsx'
OUTPUT_EXCEL_PLOTS = 'analysis_plots.xlsx'
OUTPUT_EXCEL_COMBINED = 'analysis_combined.xlsx'
OUTPUT_EXCEL_GRAPHING = 'analysis_for_graphing.xlsx'
OUTPUT_EXCEL_CORRELATION = 'analysis_correlation_data.xlsx'
OUTPUT_EXCEL_COMBINED_ALL = 'analysis_combined_all.xlsx'
OUTPUT_EXCEL_QIC = 'analysis_gee_qic.xlsx'

CATEGORIES = [
    'joy', 'sadness', 'anger', 'fear',
    'surprise', 'disgust', 'trust', 'anticipation',
    'negative_sentiment', 'positive_sentiment'
]

MEDIA_CATEGORIES = {
    'Scientific': ['nature', 'sciam', 'stat', 'newscientist'],
    'Left': ['theatlantic', 'the daily beast', 'the intercept', 'mother jones', 'msnbc', 'slate', 'vox', 'huffpost'],
    'Lean Left': ['ap', 'axios', 'cnn', 'guardian', 'business insider', 'nbcnews', 'npr', 'nytimes', 'politico', 'propublica', 'wapo', 'usa today'],
    'Center': ['reuters', 'marketwatch', 'financial times', 'newsweek', 'forbes'],
    'Lean Right': ['thedispatch', 'epochtimes', 'foxbusiness', 'wsj', 'national review', 'washtimes'],
    'Right': ['breitbart', 'theblaze', 'daily mail', 'dailywire', 'foxnews', 'nypost', 'newsmax'],
}

CATEGORY_ORDER = ["Scientific", "Left", "Lean Left", "Center", "Lean Right", "Right"]

def setup_logging(log_file='analysis.log'):
    # More verbose logging format
    logging.basicConfig(
        filename=log_file,
        level=logging.DEBUG,
        format='%(asctime)s [%(levelname)s] %(module)s - %(message)s'
    )
    console = logging.StreamHandler(sys.stdout)
    console.setLevel(logging.DEBUG)
    formatter = logging.Formatter('[%(levelname)s] %(module)s: %(message)s')
    console.setFormatter(formatter)
    logging.getLogger('').addHandler(console)
    logging.info("Logging initialized, including console output.")

def load_jsonl(jsonl_file):
    records = []
    with open(jsonl_file, 'r', encoding='utf-8') as f:
        for line in tqdm(f, desc="Loading JSONL data"):
            try:
                record = json.loads(line)
                records.append(record)
            except json.JSONDecodeError as e:
                logging.error(f"JSON decoding error: {e}")
    df = pd.DataFrame(records)
    return df

def map_media_outlet_to_category(df, media_categories):
    logging.info("Mapping media outlets to categories.")
    outlet_to_category = {}
    for category, outlets in media_categories.items():
        for outlet in outlets:
            outlet_to_category[outlet.lower().strip()] = category

    if 'media_outlet' not in df.columns:
        msg = "'media_outlet' column not found in DataFrame."
        logging.error(msg)
        raise KeyError(msg)

    df['media_outlet_clean'] = df['media_outlet'].str.lower().str.strip()
    df['media_category'] = df['media_outlet_clean'].map(outlet_to_category)
    df['media_category'] = df['media_category'].fillna('Other')
    unmapped_outlets = df[df['media_category'] == 'Other']['media_outlet'].unique()
    if len(unmapped_outlets) > 0:
        logging.warning(f"Unmapped media outlets found: {unmapped_outlets}")
        print(f"Warning: The following media outlets were not mapped and categorized as 'Other': {unmapped_outlets}")
    return df

def compute_article_level_scores(df, sentiment_categories):
    logging.info("Computing article-level sentiment scores for Quotation & Fulltext.")
    for sentiment in sentiment_categories:
        pattern = f"^{re.escape(sentiment)}_\\d+$"
        matched_cols = [c for c in df.columns if re.match(pattern, c)]
        if matched_cols:
            df[f"{sentiment}_quotation_mean_article"] = df[matched_cols].clip(lower=0).mean(axis=1)
        else:
            df[f"{sentiment}_quotation_mean_article"] = np.nan

        fulltext_col = f"{sentiment}_fulltext"
        if fulltext_col in df.columns:
            df[f"{sentiment}_fulltext_article"] = df[fulltext_col].clip(lower=0)
        else:
            df[f"{sentiment}_fulltext_article"] = np.nan
    return df

def aggregate_sentiment_scores(df, sentiment_categories):
    logging.info("Aggregating sentiment/emotion scores by media_category.")
    aggregated = []
    for cat in MEDIA_CATEGORIES.keys():
        subset = df[df['media_category']==cat].copy()
        for sentiment in sentiment_categories:
            pattern = f"^{re.escape(sentiment)}_\\d+$"
            matched_cols = [col for col in df.columns if re.match(pattern, col)]
            if matched_cols:
                qsum = subset[matched_cols].clip(lower=0).sum(skipna=True).sum()
                qcnt = subset[matched_cols].clip(lower=0).count().sum()
            else:
                qsum, qcnt = 0,0

            fcol = f"{sentiment}_fulltext"
            if fcol in df.columns:
                fsum = subset[fcol].clip(lower=0).sum(skipna=True)
                fcnt = subset[fcol].clip(lower=0).count()
            else:
                fsum,fcnt=0,0
            aggregated.append({
                'Media Category':cat,
                'Sentiment/Emotion':sentiment,
                'Quotation_Sum':qsum,
                'Quotation_Count':qcnt,
                'Fulltext_Sum':fsum,
                'Fulltext_Count':fcnt
            })
    return pd.DataFrame(aggregated)

def calculate_averages(aggdf):
    logging.info("Calculating Quotation and Fulltext averages.")
    aggdf['Quotation_Average'] = aggdf.apply(
        lambda row: row['Quotation_Sum']/row['Quotation_Count'] if row['Quotation_Count']>0 else None,
        axis=1
    )
    aggdf['Fulltext_Average'] = aggdf.apply(
        lambda row: row['Fulltext_Sum']/row['Fulltext_Count'] if row['Fulltext_Count']>0 else None,
        axis=1
    )
    return aggdf

def calculate_mean_median(aggdf):
    logging.info("Computing overall mean of Quotation/Fulltext for each sentiment.")
    rows=[]
    for sentiment in CATEGORIES:
        tmp=aggdf[aggdf['Sentiment/Emotion']==sentiment].copy()
        qavg=tmp['Quotation_Average'].dropna()
        favg=tmp['Fulltext_Average'].dropna()
        rows.append({
            'Sentiment/Emotion':sentiment,
            'Mean_Quotation_Average': qavg.mean() if len(qavg)>0 else None,
            'Mean_Fulltext_Average': favg.mean() if len(favg)>0 else None
        })
    return pd.DataFrame(rows)

def save_aggregated_scores_to_csv(aggdf, csvdir, prefix='aggregated_sentiment_emotion_scores.csv'):
    outpath=os.path.join(csvdir,prefix)
    try:
        aggdf.to_csv(outpath,index=False)
        msg = f"Aggregated sentiment/emotion scores saved to '{outpath}'."
        print(msg)
        logging.info(msg)
    except Exception as e:
        logging.error(f"Error saving aggregated scores to CSV: {e}")

def plot_statistics(aggdf, outdir):
    sns.set_style("whitegrid")
    for sentiment in CATEGORIES:
        tmp = aggdf[aggdf['Sentiment/Emotion']==sentiment].copy()

        plt.figure(figsize=(10,6))
        sns.barplot(x='Media Category',y='Quotation_Average',data=tmp,color='steelblue')
        plt.title(f"Mean Quotation-Based '{sentiment.capitalize()}' Scores")
        plt.xlabel('Media Category')
        plt.ylabel('Mean Quotation-Based Average Score')
        plt.xticks(rotation=45,ha='right')
        plt.tight_layout()
        outf_q = os.path.join(outdir, f"quote_{sentiment}.png")
        try:
            plt.savefig(outf_q)
            plt.close()
            print(f"Quotation-Based '{sentiment}' scores plot saved to '{outf_q}'.")
        except Exception as e:
            logging.error(f"Error saving Quotation plot for {sentiment}: {e}")

        plt.figure(figsize=(10,6))
        sns.barplot(x='Media Category',y='Fulltext_Average',data=tmp,color='darkorange')
        plt.title(f"Mean Fulltext-Based '{sentiment.capitalize()}' Scores")
        plt.xlabel('Media Category')
        plt.ylabel('Mean Fulltext-Based Average Score')
        plt.xticks(rotation=45,ha='right')
        plt.tight_layout()
        outf_f = os.path.join(outdir, f"fulltext_{sentiment}.png")
        try:
            plt.savefig(outf_f)
            plt.close()
            print(f"Fulltext-Based '{sentiment}' scores plot saved to '{outf_f}'.")
        except Exception as e:
            logging.error(f"Error saving Fulltext plot for {sentiment}: {e}")

def filter_clusters_for_ar(model_df):
    """Keep only clusters with >=2 obs."""
    counts = model_df.groupby('media_outlet').size()
    valid_outlets = counts[counts>=2].index
    newdf = model_df[model_df['media_outlet'].isin(valid_outlets)].copy()
    logging.debug(f"filter_clusters_for_ar: before={len(model_df)}, after={len(newdf)}, valid_outlets={len(valid_outlets)}")
    return newdf

def fit_gee_and_pairwise(df, sentiment, measure_type='Quotation', cov_struct_type=Exchangeable, global_min_date=None):
    """
    Convert 'date' -> integer day index from earliest date => (date - global_min_date).dt.days +1
    Ensure time is int, log invalid elements, skip if found.
    """
    logging.debug(f"fit_gee_and_pairwise: sentiment={sentiment}, measure_type={measure_type}, struct={cov_struct_type.__name__}")
    if measure_type=='Quotation':
        pattern = f"^{re.escape(sentiment)}_\\d+$"
        matched_cols = [c for c in df.columns if re.match(pattern,c)]
        if not matched_cols:
            logging.debug("No matched quotation columns found, returning None.")
            return None
        temp = df.copy()
        temp[f'{sentiment}_quotation_mean'] = temp[matched_cols].clip(lower=0).mean(axis=1)
        score_col = f'{sentiment}_quotation_mean'
    else:
        fc = f"{sentiment}_fulltext"
        if fc not in df.columns:
            logging.debug(f"Fulltext col {fc} not found, returning None.")
            return None
        temp = df.copy()
        temp[f'{sentiment}_fulltext_clipped'] = temp[fc].clip(lower=0)
        score_col = f'{sentiment}_fulltext_clipped'

    needed = [score_col,'media_category','media_outlet','date']
    tdf = temp.dropna(subset=needed).copy()
    if tdf['media_category'].nunique()<2:
        logging.debug("Not enough categories, returning None.")
        return None

    # Convert date to integer day index
    logging.debug(f"Using global_min_date={global_min_date}")
    tdf['int_date'] = (tdf['date'] - global_min_date).dt.days + 1
    tdf['int_date'] = tdf['int_date'].astype(np.int32)

    logging.debug(f"Created int_date for sentiment={sentiment}, measure={measure_type}. Example:\n{tdf[['date','int_date']].head()}")

    if len(tdf)<2:
        logging.debug("Data has <2 rows after subsetting, returning None.")
        return None

    tdf['media_category'] = tdf['media_category'].astype('category')

    if cov_struct_type == Autoregressive:
        # Filter & sort
        logging.debug("Applying AR(1) filter_clusters_for_ar.")
        tdf = filter_clusters_for_ar(tdf)
        if len(tdf)<2:
            logging.debug("After AR(1) cluster filter, <2 rows, returning None.")
            return None

        tdf.sort_values(['media_outlet','int_date'],inplace=True)
        cts = tdf.groupby('media_outlet').size()
        if any(cts<2):
            logging.debug("Some cluster still has <2 obs, returning None.")
            return None

        time_arr = tdf['int_date'].values
        logging.debug(f"AR(1): final time_arr has shape={time_arr.shape}. Example first 5: {time_arr[:5]}")
        cov_struct = Autoregressive()
    else:
        time_arr = None
        cov_struct = cov_struct_type()

    family = Poisson()
    y, X = patsy.dmatrices(f"{score_col} ~ media_category", data=tdf, return_type='dataframe')
    groups = tdf['media_outlet'].values

    # -------------- TIME VALIDATION --------------
    if time_arr is not None:
        logging.debug("Validating time array for AR(1).")
        # Confirm integer dtype
        if not np.issubdtype(time_arr.dtype, np.integer):
            logging.error("Time array is NOT integer dtype!")
            for idx, val in enumerate(time_arr):
                if not isinstance(val,(int,np.integer)):
                    logging.error(f"time-arr invalid element idx={idx}, val={val}")
            return None
        # Check invalid
        invalid_mask = (time_arr<1) | np.isnan(time_arr)
        if invalid_mask.any():
            badidx = np.where(invalid_mask)[0]
            for i in badidx:
                logging.error(f"Invalid time value at row {i}, time={time_arr[i]}")
            return None

    model = GEE(y, X, groups=groups, time=time_arr, family=family, cov_struct=cov_struct)
    try:
        logging.debug(f"Fitting GEE now: sentiment={sentiment}, measure={measure_type}, struct={cov_struct_type.__name__}. y-shape={y.shape}, X-shape={X.shape}")
        results = model.fit()
        logging.debug("GEE fit successful.")
    except Exception as e:
        logging.error(f"GEE fit failed for struct={cov_struct_type.__name__}, sentiment={sentiment}, measure={measure_type}: {e}")
        return None

    summ = results.summary().as_text()
    params = results.params
    cov_ = results.cov_params()

    cat_cats = tdf['media_category'].cat.categories
    ref_cat = cat_cats[0]
    param_names = results.model.exog_names

    cat2idx = {ref_cat:0}
    for cat in cat_cats[1:]:
        nm = f"media_category[T.{cat}]"
        cat2idx[cat]=param_names.index(nm)

    pairwise_list=[]
    cats = list(cat_cats)
    for i in range(len(cats)):
        for j in range(i+1,len(cats)):
            catA,catB = cats[i],cats[j]
            c_ = np.zeros(len(params))
            if catA==ref_cat and catB!=ref_cat:
                c_[cat2idx[catB]]=-1.
            elif catB==ref_cat and catA!=ref_cat:
                c_[cat2idx[catA]]=1.
            else:
                c_[cat2idx[catA]]=1.
                c_[cat2idx[catB]]=-1.
            diff_est = c_ @ params
            diff_var = c_ @ cov_ @ c_
            diff_se = np.sqrt(diff_var) if diff_var>=0 else 0.
            z_ = diff_est/diff_se if diff_se>0 else np.inf
            p_ = 2*(1-norm.cdf(abs(z_)))
            pairwise_list.append((catA, catB, diff_est, diff_se, z_, p_))

    pdf = pd.DataFrame(pairwise_list, columns=["CategoryA","CategoryB","Difference","SE","Z","p_value"])
    reject, p_adj, _, _ = multipletests(pdf['p_value'], method='holm')
    pdf['p_value_adj'] = p_adj
    pdf['reject_H0'] = reject

    try:
        qicv = results.qic()
    except:
        qicv = np.nan

    return {
        'GEE_Summary': summ,
        'Pairwise': pdf,
        'QIC': qicv,
        'Results': results
    }

def try_all_cov_structures(df, sentiment, measure_type, min_date):
    structures = [Independence, Exchangeable, Unstructured, Autoregressive]
    outlist=[]
    for struct in structures:
        ret = fit_gee_and_pairwise(df, sentiment, measure_type, struct, global_min_date=min_date)
        if ret is not None:
            outlist.append((struct.__name__, ret['QIC'], ret))
    return outlist

def run_gee_analyses(df):
    logging.info("run_gee_analyses: Starting GEE analyses with multiple covariance structures.")
    min_date = df['date'].dropna().min()
    logging.debug(f"Global min_date={min_date}.")
    if pd.isnull(min_date):
        logging.debug("No valid min_date found => AR(1) not possible.")
    all_results={}
    all_qic=[]
    for sentiment in CATEGORIES:
        all_results[sentiment]={}
        for measure in ['Quotation','Fulltext']:
            logging.debug(f"Analyzing sentiment={sentiment}, measure={measure}.")
            rst_list = try_all_cov_structures(df, sentiment, measure, min_date)
            if len(rst_list)==0:
                logging.debug(f"No successful GEE fits for {sentiment}-{measure}.")
                continue
            # Sort
            final=[]
            for (nm,qv,res_) in rst_list:
                if not isinstance(qv,(int,float,np.number)) or pd.isna(qv):
                    qv=np.nan
                final.append((nm,qv,res_))
            final.sort(key=lambda x: (np.inf if np.isnan(x[1]) else x[1]))
            best_struct=final[0]
            all_results[sentiment][measure]=best_struct[2]
            for sn_,qicv_,rs_ in final:
                all_qic.append({
                    'Sentiment':sentiment,
                    'Measure':measure,
                    'Structure':sn_,
                    'QIC':qicv_
                })
    qicdf = pd.DataFrame(all_qic)
    return all_results, qicdf

def compile_results_into_multiple_workbooks(aggdf, statsdf, rawdf, gee_res, qicdf, plotsdir,
                                            main_xlsx, raw_xlsx, lmm_xlsx, plots_xlsx,
                                            combined_xlsx, qic_xlsx):

    logging.info("Compiling results into multiple Excel workbooks.")
    with pd.ExcelWriter(main_xlsx, engine='openpyxl') as w:
        aggdf.to_excel(w, sheet_name='Aggregated_Scores',index=False)
        statsdf.to_excel(w, sheet_name='Mean_Median_Statistics',index=False)

    rawdf2 = rawdf.copy()
    rawdf2['media_category'] = pd.Categorical(rawdf2['media_category'],
                                              categories=CATEGORY_ORDER+['Other'],
                                              ordered=True)
    rawdf2.sort_values(['media_category','media_outlet'], inplace=True)
    with pd.ExcelWriter(raw_xlsx, engine='openpyxl') as w:
        rawdf2.to_excel(w, sheet_name='Raw_Data', index=False)
        for sentiment in CATEGORIES:
            scols = [c for c in rawdf2.columns if c.startswith(sentiment+'_')]
            sdf=rawdf2[['media_category','media_outlet']+scols].copy()
            sheetnm = f"Raw_{sentiment[:29]}"
            sdf.to_excel(w, sheet_name=sheetnm, index=False)

    with pd.ExcelWriter(lmm_xlsx, engine='openpyxl') as w:
        sumrows=[]
        for sentiment in gee_res:
            for measure in gee_res[sentiment]:
                sheetnm = f"GEE_{sentiment[:20]}_{measure[:8]}"
                summ=gee_res[sentiment][measure]['GEE_Summary']
                pwise=gee_res[sentiment][measure]['Pairwise']
                sdf = pd.DataFrame({'GEE_Summary':summ.split('\n')})
                sdf.to_excel(w, sheet_name=sheetnm, index=False)
                startrow=len(sdf)+2
                pwise.to_excel(w, sheet_name=sheetnm, index=False, startrow=startrow)

                sumrows.append({
                    'Sentiment':sentiment,
                    'Measure':measure,
                    'SheetName':sheetnm
                })
        df_idx=pd.DataFrame(sumrows)
        df_idx.to_excel(w, sheet_name='GEE_Results_Index',index=False)

    with pd.ExcelWriter(qic_xlsx, engine='openpyxl') as w:
        qicdf.to_excel(w, sheet_name='QIC_Results',index=False)

    wb = Workbook()
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])

    # embed plots
    for sentiment in CATEGORIES:
        qplot=os.path.join(plotsdir,f"quote_{sentiment}.png")
        if os.path.exists(qplot):
            stitle=f"Quote_{sentiment[:28]}"
            ws=wb.create_sheet(title=stitle)
            try:
                img = ExcelImage(qplot)
                img.anchor='A1'
                ws.add_image(img)
            except Exception as e:
                logging.error(f"Error embedding quote plot {qplot} : {e}")

        fplot=os.path.join(plotsdir,f"fulltext_{sentiment}.png")
        if os.path.exists(fplot):
            stitle2=f"Fulltext_{sentiment[:25]}"
            ws2=wb.create_sheet(title=stitle2)
            try:
                img2=ExcelImage(fplot)
                img2.anchor='A1'
                ws2.add_image(img2)
            except Exception as e:
                logging.error(f"Error embedding fulltext plot {fplot} : {e}")

    wb.save(plots_xlsx)

    rawdf_clean = rawdf2.copy()
    for col in rawdf_clean.columns:
        rawdf_clean[col] = rawdf_clean[col].apply(lambda x:", ".join(x) if isinstance(x,list) else x)

    wbcomb=Workbook()
    if 'Sheet' in wbcomb.sheetnames:
        wbcomb.remove(wbcomb['Sheet'])

    ws_agg=wbcomb.create_sheet("Aggregated_Scores")
    for r in dataframe_to_rows(aggdf, index=False, header=True):
        ws_agg.append(r)

    ws_stats=wbcomb.create_sheet("Mean_Median_Statistics")
    for r in dataframe_to_rows(statsdf, index=False, header=True):
        ws_stats.append(r)

    ws_raw=wbcomb.create_sheet("Raw_Data")
    for r in dataframe_to_rows(rawdf_clean, index=False, header=True):
        ws_raw.append(r)

    ws_gee=wbcomb.create_sheet("GEE_Summaries")
    ws_gee.append(["Sentiment","Measure","GEE_Summary"])
    for sentiment in gee_res:
        for measure in gee_res[sentiment]:
            summ_ = gee_res[sentiment][measure]['GEE_Summary']
            ws_gee.append([sentiment, measure, summ_])

    wbcomb.save(combined_xlsx)

def create_analysis_for_graphing_file(df, sentiments, category_order, outxlsx):
    df_ = df.copy()
    df_['media_category'] = pd.Categorical(df_['media_category'],
                                           categories=category_order+['Other'],
                                           ordered=True)
    df_ = df_[df_['media_category'].isin(category_order)]

    with pd.ExcelWriter(outxlsx, engine='openpyxl') as w:
        for sentiment in sentiments:
            qcol = f"{sentiment}_quotation_mean_article"
            fcol = f"{sentiment}_fulltext_article"

            catq={c:[] for c in category_order}
            for idx, row in df_.iterrows():
                val = row[qcol] if qcol in df_.columns else np.nan
                rowd={c:np.nan for c in category_order}
                if row['media_category'] in category_order:
                    rowd[row['media_category']]=val
                for c in category_order:
                    catq[c].append(rowd[c])
            qdf = pd.DataFrame(catq, columns=category_order)

            catf={c:[] for c in category_order}
            for idx, row in df_.iterrows():
                val= row[fcol] if fcol in df_.columns else np.nan
                rowd={c:np.nan for c in category_order}
                if row['media_category'] in category_order:
                    rowd[row['media_category']]=val
                for c in category_order:
                    catf[c].append(rowd[c])
            fdf=pd.DataFrame(catf, columns=category_order)

            qdf.to_excel(w, sheet_name=f"{sentiment}_Quotation",index=False)
            fdf.to_excel(w, sheet_name=f"{sentiment}_Fulltext",index=False)

def main():
    setup_logging()
    logging.info("Starting main() for GEE analyses.")
    print("Single run with GEE and AR(1), with integer day indices and extensive logging.")
    df = load_jsonl(INPUT_JSONL_FILE)
    print(f"Total articles loaded: {len(df)}")
    df = map_media_outlet_to_category(df, MEDIA_CATEGORIES)

    if 'date' not in df.columns:
        logging.warning("No 'date' column => AR(1) invalid.")
    else:
        df['date']=pd.to_datetime(df['date'],format='%Y-%m-%d', errors='coerce')
        logging.info("Converted 'date' to datetime.")

    # compute article-level
    df = compute_article_level_scores(df, CATEGORIES)

    # chunk raw
    chunk_size=20000
    for i in range(0, len(df), chunk_size):
        chunk = df.iloc[i:i+chunk_size]
        chunk_file = os.path.join(CSV_OUTPUT_DIR, f"raw_data_part_{i//chunk_size+1}.csv")
        chunk.to_csv(chunk_file, index=False)
        print(f"Saved chunk {i//chunk_size+1} to {chunk_file}")

    print("\nSummary Statistics:")
    print("Total number of articles:", len(df))
    print("\nNumber of articles per outlet:")
    print(df['media_outlet_clean'].value_counts())
    print("\nNumber of articles per category:")
    print(df['media_category'].value_counts())
    print()

    logging.info("Aggregating sentiment/emotion scores per media category.")
    aggdf= aggregate_sentiment_scores(df, CATEGORIES)
    aggdf= calculate_averages(aggdf)
    statsdf= calculate_mean_median(aggdf)

    save_aggregated_scores_to_csv(aggdf, CSV_OUTPUT_DIR)
    plot_statistics(aggdf, OUTPUT_DIR)

    correlation_data = []
    correlation_scatter_data = {}
    for sentiment in CATEGORIES:
        ssub = aggdf.loc[aggdf['Sentiment/Emotion']==sentiment].copy()
        ssub = ssub.loc[ssub['Media Category'].isin(MEDIA_CATEGORIES.keys())]
        ssub.dropna(subset=['Quotation_Average','Fulltext_Average'], inplace=True)

        if len(ssub)>1:
            cval, _ = pearsonr(ssub['Quotation_Average'], ssub['Fulltext_Average'])
        else:
            cval=np.nan
        correlation_data.append({'Sentiment/Emotion':sentiment,'Correlation':cval})
        correlation_scatter_data[sentiment]= ssub[['Media Category','Quotation_Average','Fulltext_Average']].copy()

        plt.figure(figsize=(6,6))
        sns.scatterplot(x='Quotation_Average',y='Fulltext_Average',data=ssub,hue='Media Category', s=50)
        plt.title(f"Scatter: {sentiment.capitalize()} (Quotation vs Fulltext)")
        plt.xlabel('Quotation_Average')
        plt.ylabel('Fulltext_Average')
        plt.legend(bbox_to_anchor=(1.05,1), loc='upper left')
        plt.tight_layout()
        scatter_file = os.path.join(OUTPUT_DIR, f"scatter_{sentiment}.png")
        plt.savefig(scatter_file)
        plt.close()
        print(f"Scatter plot for {sentiment} saved to '{scatter_file}'.")

    corr_df = pd.DataFrame(correlation_data)
    plt.figure(figsize=(10,6))
    sns.barplot(x='Sentiment/Emotion', y='Correlation', data=corr_df, color='gray')
    plt.title("Correlation Between Quotation and Fulltext Averages")
    plt.xticks(rotation=45, ha='right')
    plt.ylim(-1,1)
    plt.tight_layout()
    correlation_plot = os.path.join(OUTPUT_DIR, "correlation_quotation_fulltext.png")
    plt.savefig(correlation_plot)
    plt.close()
    print(f"Correlation plot saved to '{correlation_plot}'.")

    corr_xlsx = pd.ExcelWriter(OUTPUT_EXCEL_CORRELATION, engine='openpyxl')
    corr_df.to_excel(corr_xlsx, sheet_name='Correlation', index=False)
    for sentiment, scatdf in correlation_scatter_data.items():
        scatdf.to_excel(corr_xlsx, sheet_name=f"{sentiment}_Data", index=False)
    corr_xlsx.save()

    combined_list=[]
    for sentiment, data_ in correlation_scatter_data.items():
        tmp_= data_.copy()
        tmp_['Sentiment/Emotion']=sentiment
        combined_list.append(tmp_)
    combined_all = pd.concat(combined_list, ignore_index=True)

    combined_all['Quotation_Z']=combined_all.groupby('Sentiment/Emotion')['Quotation_Average'].transform(
        lambda x:(x-x.mean())/x.std() if x.std()!=0 else x - x.mean()
    )
    combined_all['Fulltext_Z']=combined_all.groupby('Sentiment/Emotion')['Fulltext_Average'].transform(
        lambda x:(x-x.mean())/x.std() if x.std()!=0 else x - x.mean()
    )
    combined_all.to_excel(OUTPUT_EXCEL_COMBINED_ALL, index=False)

    comb_nonan= combined_all.dropna(subset=['Quotation_Z','Fulltext_Z'])
    if len(comb_nonan)>1:
        r_,_= pearsonr(comb_nonan['Quotation_Z'], comb_nonan['Fulltext_Z'])
    else:
        r_=np.nan

    plt.figure(figsize=(10,8))
    sns.regplot(x='Quotation_Z', y='Fulltext_Z', data=comb_nonan, scatter_kws={'color':'black'}, line_kws={'color':'red'})
    plt.title(f"All Sentiments Combined (Normalized)\nR = {r_:.2f}")
    plt.xlabel('Normalized Quotation_Average (Z-score)')
    plt.ylabel('Normalized Fulltext_Average (Z-score)')
    plt.tight_layout()
    comb_outf= os.path.join(OUTPUT_DIR, "combined_normalized_scatter.png")
    plt.savefig(comb_outf)
    plt.close()
    print(f"Combined normalized scatter plot saved to '{comb_outf}'.")

    logging.info("Fitting GEE models with multiple covariance structures (including AR(1)).")
    gee_results, qic_df = run_gee_analyses(df)
    logging.info("GEE analysis complete.")
    compile_results_into_multiple_workbooks(
        aggdf, statsdf, df, gee_results, qic_df, OUTPUT_DIR,
        OUTPUT_EXCEL_MAIN, OUTPUT_EXCEL_RAW, OUTPUT_EXCEL_LMM,
        OUTPUT_EXCEL_PLOTS, OUTPUT_EXCEL_COMBINED, OUTPUT_EXCEL_QIC
    )

    create_analysis_for_graphing_file(df, CATEGORIES, CATEGORY_ORDER, OUTPUT_EXCEL_GRAPHING)
    logging.info("All analyses completed successfully.")
    print("Analysis completed successfully.")

if __name__=="__main__":
    setup_logging()
    main()
