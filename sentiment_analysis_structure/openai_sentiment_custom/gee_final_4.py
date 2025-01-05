#!/usr/bin/env python3
# gee_final_best_qic_with_pairwise_cld.py
"""
Script that:
  1) Loads data from JSONL
  2) Chunks & saves CSV
  3) Correlation analyses, aggregates sentiment scores
  4) Finds best QIC combination for each sentiment–measure
  5) For each best QIC model:
     - Refit & produce GEE summary
     - Compute pairwise comparisons (Holm correction)
     - Build a compact letter display (CLD) across categories:
       [Scientific, Left, Lean Left, Center, Lean Right, Right]
  6) Writes these (summary + pairwise + CLD) into analysis_gee.xlsx
  7) Also writes analysis_main.xlsx, analysis_raw.xlsx, analysis_plots.xlsx,
     analysis_combined.xlsx as usual.
"""

import json
import os
import re
import sys
import warnings
import logging
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import seaborn as sns
from tqdm import tqdm

from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.utils.dataframe import dataframe_to_rows

import statsmodels.api as sm
from statsmodels.genmod.generalized_estimating_equations import GEE
from statsmodels.genmod.families import Poisson
from statsmodels.genmod.cov_struct import Independence, Exchangeable
from statsmodels.stats.multitest import multipletests
from scipy.stats import pearsonr, norm

# ------------------------------------------------
# CONFIG
# ------------------------------------------------
INPUT_JSONL_FILE = "processed_all_articles_fixed_2.jsonl"

OUTPUT_DIR = "graphs_analysis"
os.makedirs(OUTPUT_DIR, exist_ok=True)

CSV_OUTPUT_DIR = "csv_raw_scores"
os.makedirs(CSV_OUTPUT_DIR, exist_ok=True)

OUTPUT_EXCEL_MAIN = "analysis_main.xlsx"
OUTPUT_EXCEL_RAW = "analysis_raw.xlsx"
OUTPUT_EXCEL_GEE = "analysis_gee.xlsx"   # Will store best QIC approach with pairwise + CLD
OUTPUT_EXCEL_PLOTS = "analysis_plots.xlsx"
OUTPUT_EXCEL_COMBINED = "analysis_combined.xlsx"

LOG_FILE = "analysis.log"

CATEGORIES = [
    "joy","sadness","anger","fear",
    "surprise","disgust","trust","anticipation",
    "negative_sentiment","positive_sentiment"
]

# The order requested for CLD:
CLD_ORDER = ["Scientific","Left","Lean Left","Center","Lean Right","Right"]

MEDIA_CATEGORIES = {
    "Scientific": ["nature","sciam","stat","newscientist"],
    "Left": ["theatlantic","the daily beast","the intercept","mother jones","msnbc","slate","vox","huffpost"],
    "Lean Left": ["ap","axios","cnn","guardian","business insider","nbcnews","npr","nytimes","politico","propublica","wapo","usa today"],
    "Center": ["reuters","marketwatch","financial times","newsweek","forbes"],
    "Lean Right": ["thedispatch","epochtimes","foxbusiness","wsj","national review","washtimes"],
    "Right": ["breitbart","theblaze","daily mail","dailywire","foxnews","nypost","newsmax"],
}

def setup_logging():
    log_format = "%(asctime)s [%(levelname)s] %(module)s - %(message)s"
    logging.basicConfig(filename=LOG_FILE, level=logging.DEBUG, format=log_format)
    console = logging.StreamHandler(sys.stdout)
    console.setLevel(logging.INFO)
    console.setFormatter(logging.Formatter(log_format))
    logging.getLogger().addHandler(console)
    logging.info("Logging initialized (file+console).")

warnings.filterwarnings("ignore", message="QIC values obtained using scale=None")

# ------------------------------------------------
# 1) Load data, chunk, basic stats
# ------------------------------------------------
def load_jsonl(path):
    logging.info(f"Loading JSONL from {path}")
    recs=[]
    with open(path, "r", encoding="utf-8") as f:
        for line in tqdm(f, desc="Loading JSONL data"):
            rec=json.loads(line)
            recs.append(rec)
    df=pd.DataFrame(recs)
    logging.debug(f"Loaded shape={df.shape}")
    return df

def map_media_outlet_to_category(df):
    logging.info("Mapping media_outlet -> category from MEDIA_CATEGORIES dict")
    # build a single dict
    cat_map={}
    for cat, outls in MEDIA_CATEGORIES.items():
        for o in outls:
            cat_map[o.lower().strip()]=cat
    if "media_outlet" not in df.columns:
        raise KeyError("'media_outlet' missing in data")

    df["media_outlet_clean"] = df["media_outlet"].str.lower().str.strip()
    df["media_category"] = df["media_outlet_clean"].map(cat_map).fillna("Other")
    unm=df[df["media_category"]=="Other"]["media_outlet"].unique()
    if len(unm)>0:
        logging.warning(f"Unmapped => {unm}")
        print(f"Warning: Not mapped => {unm}")
    return df

def chunk_and_save(df, sz=20000):
    logging.info(f"Chunking len={len(df)}, size={sz}")
    for i in range(0,len(df),sz):
        cpart=df.iloc[i:i+sz]
        fn=os.path.join(CSV_OUTPUT_DIR,f"raw_data_part_{(i//sz)+1}.csv")
        cpart.to_csv(fn,index=False)
        print(f"Saved chunk {(i//sz)+1} to {fn}")

def print_basic_stats(df):
    logging.info(f"Stats => total articles={len(df)}")
    print("\nSummary Statistics:")
    print("Total number of articles:", len(df))
    if "media_outlet_clean" in df.columns:
        vc=df["media_outlet_clean"].value_counts()
        print("\nArticles per outlet:")
        print(vc)
    if "media_category" in df.columns:
        vcc=df["media_category"].value_counts()
        print("\nArticles per category:")
        print(vcc)
    print()

# ------------------------------------------------
# 2) Quotation vs Fulltext correlation
# ------------------------------------------------
def analyze_quotation_fulltext_correlation(df):
    logging.info("Analyzing Quotation vs Fulltext correlation")
    # Similar aggregator approach
    recs=[]
    for cat in df["media_category"].dropna().unique():
        subset=df[df["media_category"]==cat]
        for s in CATEGORIES:
            pat=rf"^{s}_\d+$"
            matched=[c for c in subset.columns if re.match(pat,c)]
            if matched:
                clp=subset[matched].clip(lower=0)
                qsum=clp.sum(skipna=True).sum()
                qcount=clp.count().sum()
                qavg=qsum/qcount if qcount>0 else np.nan
            else:
                qavg=np.nan

            fcol=f"{s}_fulltext"
            if fcol in subset.columns:
                fv=subset[fcol].clip(lower=0)
                f_sum=fv.sum(skipna=True)
                f_cnt=fv.count()
                favg=f_sum/f_cnt if f_cnt>0 else np.nan
            else:
                favg=np.nan

            recs.append({
                "MediaCategory":cat,
                "Sentiment":s,
                "Quotation_Average":qavg,
                "Fulltext_Average":favg
            })
    agg_df=pd.DataFrame(recs)
    corr_list=[]
    scatter_map={}
    for s in CATEGORIES:
        sdf=agg_df[(agg_df["Sentiment"]==s)].dropna(subset=["Quotation_Average","Fulltext_Average"])
        if len(sdf)>1:
            cor_val,_=pearsonr(sdf["Quotation_Average"], sdf["Fulltext_Average"])
        else:
            cor_val=np.nan
        corr_list.append({"Sentiment":s, "Correlation":cor_val})
        if not sdf.empty:
            plt.figure(figsize=(6,5))
            sns.scatterplot(x="Quotation_Average",y="Fulltext_Average",data=sdf,hue="MediaCategory",s=50)
            plt.title(f"{s.capitalize()} (Quotation vs Fulltext)\nr={cor_val:.3f}")
            plt.tight_layout()
            out_scatter=os.path.join(OUTPUT_DIR,f"scatter_{s}.png")
            try:
                plt.savefig(out_scatter)
            except:
                pass
            plt.close()
        scatter_map[s]=sdf.copy()

    cor_df=pd.DataFrame(corr_list)
    plt.figure(figsize=(8,5))
    sns.barplot(x="Sentiment",y="Correlation",data=cor_df,color="gray")
    plt.title("Correlation (Quotation vs Fulltext) per Sentiment")
    plt.xticks(rotation=45,ha="right")
    plt.ylim(-1,1)
    plt.tight_layout()
    out_bar=os.path.join(OUTPUT_DIR,"correlation_quotation_fulltext_bar.png")
    try:
        plt.savefig(out_bar)
    except:
        pass
    plt.close()

    # combined z scatter
    combo=[]
    for s,sdf in scatter_map.items():
        if not sdf.empty:
            cpy=sdf.copy()
            cpy["Sentiment"]=s
            combo.append(cpy)
    if combo:
        comb=pd.concat(combo,ignore_index=True)
        comb["Qmean"]=comb.groupby("Sentiment")["Quotation_Average"].transform("mean")
        comb["Qstd"]=comb.groupby("Sentiment")["Quotation_Average"].transform("std")
        comb["Fmean"]=comb.groupby("Sentiment")["Fulltext_Average"].transform("mean")
        comb["Fstd"]=comb.groupby("Sentiment")["Fulltext_Average"].transform("std")
        comb["Quotation_Z"]=(comb["Quotation_Average"]-comb["Qmean"])/comb["Qstd"]
        comb["Fulltext_Z"]=(comb["Fulltext_Average"]-comb["Fmean"])/comb["Fstd"]
        val=comb.dropna(subset=["Quotation_Z","Fulltext_Z"])
        if len(val)>1:
            r_val,_=pearsonr(val["Quotation_Z"], val["Fulltext_Z"])
        else:
            r_val=np.nan
        plt.figure(figsize=(7,5))
        sns.regplot(x="Quotation_Z",y="Fulltext_Z",data=val,
                    scatter_kws={"color":"black","alpha":0.6},line_kws={"color":"red"})
        plt.title(f"All Sentiments Combined (Z-scores)\nr={r_val:.3f}")
        plt.tight_layout()
        comb_out=os.path.join(OUTPUT_DIR,"combined_normalized_scatter.png")
        try:
            plt.savefig(comb_out)
        except:
            pass
        plt.close()

    out_corr_csv=os.path.join(CSV_OUTPUT_DIR,"quotation_fulltext_correlation.csv")
    cor_df.to_csv(out_corr_csv,index=False)
    logging.info(f"Correlation data => {out_corr_csv}")

# ------------------------------------------------
# 3) Aggregation & stats
# ------------------------------------------------
def aggregate_sentiment_scores(df, sentiment_list):
    logging.info("Aggregating sentiment/emotion scores by category+sentiment.")
    recs=[]
    for cat in MEDIA_CATEGORIES.keys():
        sub=df[df["media_category"]==cat]
        for s in sentiment_list:
            pat=rf"^{re.escape(s)}_\d+$"
            matched=[c for c in sub.columns if re.match(pat,c)]
            if matched:
                qsum=sub[matched].clip(lower=0).sum(skipna=True).sum()
                qcount=sub[matched].clip(lower=0).count().sum()
            else:
                qsum,qcount=(0,0)

            fcol=f"{s}_fulltext"
            if fcol in sub.columns:
                fv=sub[fcol].clip(lower=0)
                f_sum=fv.sum(skipna=True)
                f_count=fv.count()
            else:
                f_sum,f_count=(0,0)

            recs.append({
                "Media Category":cat,
                "Sentiment/Emotion":s,
                "Quotation_Sum":qsum,
                "Quotation_Count":qcount,
                "Fulltext_Sum":f_sum,
                "Fulltext_Count":f_count
            })
    return pd.DataFrame(recs)

def calculate_averages(agg_df):
    logging.info("Calculating Quotation_Average, Fulltext_Average.")
    def sdiv(a,b):
        return a/b if b>0 else None
    agg_df["Quotation_Average"]=agg_df.apply(lambda r: sdiv(r["Quotation_Sum"],r["Quotation_Count"]),axis=1)
    agg_df["Fulltext_Average"]=agg_df.apply(lambda r: sdiv(r["Fulltext_Sum"],r["Fulltext_Count"]),axis=1)
    return agg_df

def calculate_mean_median(agg_df):
    logging.info("Computing global mean/median of Quotation/Fulltext.")
    out=[]
    for s in CATEGORIES:
        sub=agg_df[agg_df["Sentiment/Emotion"]==s]
        qa=sub["Quotation_Average"].dropna()
        fa=sub["Fulltext_Average"].dropna()
        mqa=qa.mean() if len(qa)>0 else None
        medqa=qa.median() if len(qa)>0 else None
        mfa=fa.mean() if len(fa)>0 else None
        medfa=fa.median() if len(fa)>0 else None
        out.append({
            "Sentiment/Emotion": s,
            "Mean_Quotation_Average": mqa,
            "Median_Quotation_Average": medqa,
            "Mean_Fulltext_Average": mfa,
            "Median_Fulltext_Average": medfa
        })
    return pd.DataFrame(out)

def save_aggregated_scores_to_csv(agg_df, out_dir):
    fn=os.path.join(out_dir,"aggregated_sentiment_emotion_scores.csv")
    agg_df.to_csv(fn,index=False)
    print(f"Aggregated sentiment/emotion scores => {fn}")
    logging.info(f"Aggregated => {fn}")

def plot_statistics(agg_df, out_dir):
    sns.set_style("whitegrid")
    for s in CATEGORIES:
        sub=agg_df[agg_df["Sentiment/Emotion"]==s]
        # Quotation
        plt.figure(figsize=(10,6))
        sns.barplot(x="Media Category",y="Quotation_Average",data=sub,color="steelblue")
        plt.title(f"Mean Quotation-Based '{s.capitalize()}' Scores")
        plt.xticks(rotation=45,ha="right")
        plt.tight_layout()
        out1=os.path.join(out_dir,f"quote_{s}.png")
        try:
            plt.savefig(out1)
            plt.close()
        except:
            pass

        # Fulltext
        plt.figure(figsize=(10,6))
        sns.barplot(x="Media Category",y="Fulltext_Average",data=sub,color="darkorange")
        plt.title(f"Mean Fulltext-Based '{s.capitalize()}' Scores")
        plt.xticks(rotation=45,ha="right")
        plt.tight_layout()
        out2=os.path.join(out_dir,f"fulltext_{s}.png")
        try:
            plt.savefig(out2)
            plt.close()
        except:
            pass

# ------------------------------------------------
# 4) GEE scale computations (same approach)
# ------------------------------------------------
def compute_pearson_scale(y,mu,df_resid):
    if df_resid<=0: return np.nan
    y=np.asarray(y)
    mu=np.asarray(mu)
    r=(y-mu)/np.sqrt(mu)
    return np.sum(r**2)/df_resid

def compute_deviance_scale(y,mu,df_resid):
    if df_resid<=0: return np.nan
    y=np.asarray(y)
    mu=np.asarray(mu)
    arr=np.zeros_like(y,dtype=float)
    for i in range(len(y)):
        if y[i]>0 and mu[i]>0:
            arr[i]=y[i]*np.log(y[i]/mu[i])-(y[i]-mu[i])
        elif y[i]==0:
            arr[i]=-(y[i]-mu[i])
        else:
            arr[i]=np.nan
    return 2*np.nansum(arr)/df_resid

def compute_ub_scale(y,mu,df_resid):
    p=compute_pearson_scale(y,mu,df_resid)
    if np.isnan(p): return p
    return 1.1*p

def compute_bc_scale(y,mu,df_resid):
    d=compute_deviance_scale(y,mu,df_resid)
    if np.isnan(d): return d
    return 0.9*d

def fit_and_compute_scales(model):
    base_res=model.fit(scale=None)
    base_qic=base_res.qic()
    y=base_res.model.endog
    mu=base_res.fittedvalues
    n=len(y)
    p=len(base_res.params)
    df_resid=n-p

    pear=compute_pearson_scale(y,mu,df_resid)
    dev=compute_deviance_scale(y,mu,df_resid)
    ubv=compute_ub_scale(y,mu,df_resid)
    bcv=compute_bc_scale(y,mu,df_resid)

    out_map={}
    if isinstance(base_qic, tuple):
        q_m,q_a=base_qic
    else:
        q_m,q_a=base_qic,None
    out_map["none"]=(q_m,q_a,None)

    from math import isnan
    for (nm, val) in [("pearson",pear),("deviance",dev),("ub",ubv),("bc",bcv)]:
        if (not isnan(val)) and (df_resid>0):
            re2=model.fit(scale=val)
            rq=re2.qic()
            if isinstance(rq,tuple):
                rm,ra=rq
            else:
                rm,ra=rq,None
            out_map[nm]=(rm,ra,val)
        else:
            out_map[nm]=(np.nan,None,val)

    return out_map

# ------------------------------------------------
# 5) best QIC approach
# ------------------------------------------------
def run_gee_for_sentiment_measure_best_qic(df, sentiment, measure):
    d2=df.copy()
    if measure=="Quotation":
        pat=rf"^{re.escape(sentiment)}_\d+$"
        matched=[c for c in d2.columns if re.match(pat,c)]
        if not matched:
            return None
        d2["_score_col"]=d2[matched].clip(lower=0).mean(axis=1)
    else:
        fcol=f"{sentiment}_fulltext"
        if fcol not in d2.columns:
            return None
        d2["_score_col"]=d2[fcol].clip(lower=0)

    needed=["_score_col","media_outlet_clean","media_category"]
    d2=d2.dropna(subset=needed)
    if len(d2)<2 or d2["media_category"].nunique()<2:
        return None

    best_qic_main=np.inf
    best_tuple=None
    combos=[]
    structures=[Independence(), Exchangeable()]
    for cov_obj in structures:
        st_name=cov_obj.__class__.__name__
        model=GEE.from_formula("_score_col ~ media_category",
            groups="media_outlet_clean",data=d2,
            family=Poisson(),cov_struct=cov_obj)
        sc_map=fit_and_compute_scales(model)
        for sm,(qm,qa,sc_val) in sc_map.items():
            combos.append({
                "Sentiment":sentiment,
                "Measure":measure,
                "Structure":st_name,
                "ScaleMethod":sm,
                "NumericScale":sc_val,
                "QIC_main":qm,
                "QIC_alt":qa
            })
            if (qm<best_qic_main) and (not np.isnan(qm)):
                best_qic_main=qm
                best_tuple=(st_name, sm, qm, sc_val)

    if best_tuple is None:
        return None
    stn, sm, qicv, scale_num=best_tuple
    return {
        "Sentiment":sentiment,
        "Measure":measure,
        "Best_Structure":stn,
        "Best_Scale":sm,
        "Best_QIC_main":qicv,
        "AllCombos": pd.DataFrame(combos),
        "Summary": f"Best scale={sm}, numeric={scale_num}"
    }

def run_gee_analyses_best_qic(df):
    logging.info("Running best QIC approach for each sentiment–measure.")
    best_list=[]
    combos_list=[]
    for s in CATEGORIES:
        for meas in ["Quotation","Fulltext"]:
            info=run_gee_for_sentiment_measure_best_qic(df,s,meas)
            if info is not None:
                best_list.append({
                    "Sentiment": info["Sentiment"],
                    "Measure": info["Measure"],
                    "Best_Structure": info["Best_Structure"],
                    "Best_Scale": info["Best_Scale"],
                    "Best_QIC_main": info["Best_QIC_main"],
                    "Summary": info["Summary"]
                })
                combos_list.append(info["AllCombos"])
    df_best=pd.DataFrame(best_list)
    if combos_list:
        df_all=pd.concat(combos_list, ignore_index=True)
    else:
        df_all=pd.DataFrame()
    return df_best, df_all

# ------------------------------------------------
# 6) Pairwise with p-values + CLD
# ------------------------------------------------
def pairwise_comparisons(df, sentiment, measure, structure_name, scale_name):
    """
    Refit best GEE => summary + pairwise + CLD table
    """
    # 1) Build + refit
    model_df=df.copy()
    if measure=="Quotation":
        pat=rf"^{re.escape(sentiment)}_\d+$"
        matched=[c for c in model_df.columns if re.match(pat,c)]
        if not matched:
            return None, None, None
        model_df["_score_col"]=model_df[matched].clip(lower=0).mean(axis=1)
    else:
        fcol=f"{sentiment}_fulltext"
        if fcol not in model_df.columns:
            return None, None, None
        model_df["_score_col"]=model_df[fcol].clip(lower=0)

    needed=["_score_col","media_outlet_clean","media_category"]
    model_df=model_df.dropna(subset=needed)
    if (len(model_df)<2) or (model_df["media_category"].nunique()<2):
        return None, None, None

    model_df["media_category"]=model_df["media_category"].astype("category")
    if structure_name=="Independence":
        cov_struct=Independence()
    else:
        cov_struct=Exchangeable()

    base_model=GEE.from_formula(
        "_score_col ~ media_category",
        groups="media_outlet_clean",
        data=model_df,
        family=Poisson(),
        cov_struct=cov_struct
    )
    base_res=base_model.fit(scale=None)
    # scale fix
    if scale_name=="none":
        final_res=base_res
    else:
        y=base_res.model.endog
        mu=base_res.fittedvalues
        n=len(y)
        p=len(base_res.params)
        df_resid=n-p
        if df_resid<=0:
            final_res=base_res
        else:
            if scale_name=="pearson":
                scv=compute_pearson_scale(y,mu,df_resid)
            elif scale_name=="deviance":
                scv=compute_deviance_scale(y,mu,df_resid)
            elif scale_name=="ub":
                scv=compute_ub_scale(y,mu,df_resid)
            elif scale_name=="bc":
                scv=compute_bc_scale(y,mu,df_resid)
            else:
                scv=None
            if (scv is not None) and (not np.isnan(scv)):
                final_res=base_model.fit(scale=scv)
            else:
                final_res=base_res

    summary_text=final_res.summary().as_text()

    # 2) Pairwise
    params=final_res.params
    cov=final_res.cov_params()
    cats=model_df["media_category"].cat.categories
    ref=cats[0]
    idx_map={ref:0}
    for c in cats[1:]:
        nm=f"media_category[T.{c}]"
        idx_map[c]=final_res.model.exog_names.index(nm)

    pair_list=[]
    for i in range(len(cats)):
        for j in range(i+1,len(cats)):
            ca,cb=cats[i],cats[j]
            contrast=np.zeros(len(params))
            if ca==ref and cb!=ref:
                contrast[idx_map[cb]]=-1.0
            elif cb==ref and ca!=ref:
                contrast[idx_map[ca]]=1.0
            else:
                contrast[idx_map[ca]]=1.0
                contrast[idx_map[cb]]=-1.0

            diff_est=contrast@params
            diff_var=contrast@cov@contrast
            diff_se=np.sqrt(diff_var)
            z=diff_est/diff_se
            pval=2*(1-norm.cdf(abs(z)))
            pair_list.append((ca,cb,diff_est,diff_se,z,pval))

    pair_df=pd.DataFrame(pair_list, columns=["CategoryA","CategoryB","Difference","SE","Z","p_value"])
    # holm
    rej, p_adj, _, _ = multipletests(pair_df["p_value"],method="holm")
    pair_df["p_value_adj"]=p_adj
    pair_df["reject_H0"]=rej

    # 3) Compact Letter Display
    cld_df=build_compact_letter_display(pair_df)

    return summary_text, pair_df, cld_df


def build_compact_letter_display(pair_df):
    """
    Generate a 2-column DataFrame:
      1) MediaCategory in order = CLD_ORDER
      2) CLD letter assignment
    Logic:
      - "reject_H0"=True => categories differ
      - "reject_H0"=False => categories are not significantly different => can share letter
    We'll do a standard iterative CLD approach.

    If a category does not appear in pair_df at all, we place it or skip?
    We'll place it anyway. 
    """
    # The categories we want in top-down order:
    cat_order = CLD_ORDER

    # But let's see which categories actually appear in the pairwise
    # The pair_df has CategoryA, CategoryB. We'll unify them
    present_cats=set(pair_df["CategoryA"].unique()).union(pair_df["CategoryB"].unique())

    # We'll only show CLD for categories that actually appear
    # but keep the requested order
    final_cats=[c for c in cat_order if c in present_cats]
    if len(final_cats)<2:
        # if only one cat or none => trivial
        df_empty=pd.DataFrame({
            "MediaCategory":final_cats,
            "CLD":["a"]*len(final_cats)
        })
        return df_empty

    # We'll build an adjacency matrix "diff[catA][catB] = True if reject_H0 => they differ"
    diff_map={}
    for ca, cb, diff_est, diff_se, z, p_val, p_val_adj, rej in pair_df[["CategoryA","CategoryB","Difference","SE","Z","p_value","p_value_adj","reject_H0"]].itertuples(index=False):
        # fill both ways
        diff_map.setdefault(ca,{})[cb]=rej
        diff_map.setdefault(cb,{})[ca]=rej

    # for any cat not in map, fill empty
    for c in final_cats:
        diff_map.setdefault(c,{})

    # We'll do a standard letter assignment approach 
    # If two categories are "not different" => they can share letter
    # so "diff_map[ca][cb] = True => differ => cannot share letter"
    # we want "not differ => false => can share letter"

    letters=[]
    # we'll store letter assignments in a dictionary cat->list_of_letters
    cat_letters={c:[] for c in final_cats}
    letter_pool=[chr(i) for i in range(ord('a'), ord('z')+1)]  # hopefully enough

    for c in final_cats:
        # try each letter in letters
        assigned=False
        for let in letters:
            # check if we can put cat c in letter let
            # means c is "not different" from all cats that have letter let
            # so for each cat x that has letter let => diff_map[c][x] must be false => do not differ
            conflict=False
            # find the cats that share let
            cats_with_let=[cx for cx in final_cats if let in cat_letters[cx]]
            for cx in cats_with_let:
                # if diff => can't share
                rej=diff_map.get(c,{}).get(cx,False) # if True => they differ
                if rej:
                    conflict=True
                    break
            if not conflict:
                # we can share letter let
                cat_letters[c].append(let)
                assigned=True
                break
        if not assigned:
            # new letter
            new_let=letter_pool[len(letters)]
            letters.append(new_let)
            cat_letters[c].append(new_let)

    # now build CLD string => e.g. "a", or "ab" if multiple
    cld_list=[]
    for c in final_cats:
        cat_letters[c].sort()  # lexical
        cld_str=''.join(cat_letters[c])
        cld_list.append((c,cld_str))

    cld_df=pd.DataFrame(cld_list, columns=["MediaCategory","CLD"])
    return cld_df


# ------------------------------
# 7) Compile + final
# ------------------------------
def compile_results_into_multiple_workbooks(
    aggregated_df, stats_df, raw_df,
    df_best_qic, df_all_combos,
    plots_dir,
    main_excel, raw_excel, gee_excel, plots_excel, combined_excel,
    df_full
):
    # main
    with pd.ExcelWriter(main_excel,engine="openpyxl") as w:
        aggregated_df.to_excel(w,"Aggregated_Scores",index=False)
        stats_df.to_excel(w,"Mean_Median_Statistics",index=False)

    # raw
    with pd.ExcelWriter(raw_excel,engine="openpyxl") as w:
        raw_df.to_excel(w,"Raw_Data",index=False)
        for s in CATEGORIES:
            s_cols=[c for c in raw_df.columns if c.startswith(s+"_")]
            s_df=raw_df[["media_category","media_outlet"]+s_cols].copy()
            sh=f"Raw_{s[:29]}"
            s_df.to_excel(w,sh,index=False)

    # analysis_gee => best QIC + pairwise + CLD
    with pd.ExcelWriter(gee_excel, engine="openpyxl") as w:
        index_rows=[]
        for idx, row in df_best_qic.iterrows():
            sentiment=row["Sentiment"]
            measure=row["Measure"]
            struct=row["Best_Structure"]
            scale=row["Best_Scale"]
            sheet_name=f"BestQIC_{sentiment[:10]}_{measure[:8]}"
            sum_txt, pair_df, cld_df = pairwise_comparisons(df_full, sentiment, measure, struct, scale)
            if sum_txt is None:
                # no data
                no_df=pd.DataFrame({"Summary":["No valid model or not enough data."]})
                no_df.to_excel(w, sheet_name=sheet_name, index=False)
                continue

            # write summary
            summary_lines=sum_txt.split("\n")
            sumdf=pd.DataFrame({"GEE_Summary":summary_lines})
            sumdf.to_excel(w, sheet_name=sheet_name,index=False)
            startrow=len(sumdf)+2
            # write pair_df
            pair_df.to_excel(w, sheet_name=sheet_name, index=False, startrow=startrow)
            startrow2=startrow+len(pair_df)+2

            # write CLD
            if cld_df is not None and not cld_df.empty:
                cld_df.to_excel(w, sheet_name=sheet_name, index=False, startrow=startrow2)

            index_rows.append({
                "Sentiment": sentiment,
                "Measure": measure,
                "SheetName": sheet_name,
                "BestStructure": struct,
                "BestScale": scale,
                "BestQIC": row["Best_QIC_main"]
            })

        idxdf=pd.DataFrame(index_rows)
        idxdf.to_excel(w,"BestQIC_Index",index=False)
        if not df_all_combos.empty:
            df_all_combos.to_excel(w,"All_Combos",index=False)

    # plots
    wb_plots=Workbook()
    if "Sheet" in wb_plots.sheetnames:
        wb_plots.remove(wb_plots["Sheet"])

    for s in CATEGORIES:
        q_png=os.path.join(plots_dir,f"quote_{s}.png")
        if os.path.exists(q_png):
            st=f"Quote_{s[:28]}"
            ws=wb_plots.create_sheet(title=st)
            try:
                img=ExcelImage(q_png)
                img.anchor="A1"
                ws.add_image(img)
            except:
                pass
        f_png=os.path.join(plots_dir,f"fulltext_{s}.png")
        if os.path.exists(f_png):
            st2=f"Fulltext_{s[:25]}"
            ws2=wb_plots.create_sheet(title=st2)
            try:
                img2=ExcelImage(f_png)
                img2.anchor="A1"
                ws2.add_image(img2)
            except:
                pass
    cbar=os.path.join(plots_dir,"correlation_quotation_fulltext_bar.png")
    if os.path.exists(cbar):
        ws3=wb_plots.create_sheet("Correlation_Bar")
        try:
            img3=ExcelImage(cbar)
            img3.anchor="A1"
            ws3.add_image(img3)
        except:
            pass

    combp=os.path.join(plots_dir,"combined_normalized_scatter.png")
    if os.path.exists(combp):
        ws4=wb_plots.create_sheet("Combined_ZScatter")
        try:
            img4=ExcelImage(combp)
            img4.anchor="A1"
            ws4.add_image(img4)
        except:
            pass

    wb_plots.save(plots_excel)

    # combined
    raw_clean=raw_df.copy()
    raw_clean=raw_clean.applymap(lambda x: ", ".join(x) if isinstance(x,list) else x)
    wb_comb=Workbook()
    if "Sheet" in wb_comb.sheetnames:
        wb_comb.remove(wb_comb["Sheet"])

    ws_agg=wb_comb.create_sheet("Aggregated_Scores")
    for r in dataframe_to_rows(aggregated_df,index=False,header=True):
        ws_agg.append(r)

    ws_stats=wb_comb.create_sheet("Mean_Median_Statistics")
    for r in dataframe_to_rows(stats_df,index=False,header=True):
        ws_stats.append(r)

    ws_raw=wb_comb.create_sheet("Raw_Data")
    for r in dataframe_to_rows(raw_clean,index=False,header=True):
        ws_raw.append(r)

    ws_best=wb_comb.create_sheet("BestQIC_Table")
    for r in dataframe_to_rows(df_best_qic,index=False,header=True):
        ws_best.append(r)

    wb_comb.save(combined_excel)

# ------------------------------------------------
def main():
    setup_logging()
    logging.info("Starting best QIC approach => pairwise + CLD per tab in analysis_gee.xlsx")

    df=load_jsonl(INPUT_JSONL_FILE)
    print(f"Total articles loaded: {len(df)}")
    df=map_media_outlet_to_category(df)
    chunk_and_save(df,20000)
    print_basic_stats(df)

    print("Performing Quotation vs Fulltext correlation analysis...")
    analyze_quotation_fulltext_correlation(df)

    print("Aggregating sentiment/emotion scores per media category...")
    agg_df=aggregate_sentiment_scores(df,CATEGORIES)
    agg_df=calculate_averages(agg_df)
    stats_df=calculate_mean_median(agg_df)
    save_aggregated_scores_to_csv(agg_df,CSV_OUTPUT_DIR)
    plot_statistics(agg_df,OUTPUT_DIR)

    print("Fitting best QIC approach for each sentiment–measure (with pairwise + CLD).")
    df_best, df_allcombos = run_gee_analyses_best_qic(df)
    print("Best QIC approach completed.\n")

    compile_results_into_multiple_workbooks(
        aggregated_df=agg_df,
        stats_df=stats_df,
        raw_df=df,
        df_best_qic=df_best,
        df_all_combos=df_allcombos,
        plots_dir=OUTPUT_DIR,
        main_excel=OUTPUT_EXCEL_MAIN,
        raw_excel=OUTPUT_EXCEL_RAW,
        gee_excel=OUTPUT_EXCEL_GEE,
        plots_excel=OUTPUT_EXCEL_PLOTS,
        combined_excel=OUTPUT_EXCEL_COMBINED,
        df_full=df
    )

    print("Analysis completed successfully.")
    logging.info("Analysis completed successfully.")

if __name__=="__main__":
    main()
