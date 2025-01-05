#!/usr/bin/env python3
# readability_gee_validation_dual.py

"""
Script that runs GEE-based analyses TWICE:
  1) Using original categories: Scientific, Left, Lean Left, Center, Lean Right, Right
  2) Using collapsed categories: 
       - Scientific (same as before),
       - Left (old Left + Lean Left),
       - Center (unchanged),
       - Right (old Right + Lean Right)
Generates two separate sets of outputs.

All other functionality remains the same:
  - NegativeBinomial alpha search in [0..30],
  - Multiple scale options,
  - Cross-validation,
  - Cluster bootstrap,
  - Coverage simulation,
  - Excel outputs with BestQIC, Diagnostics, CrossValidation, Sensitivity, etc.
"""

import json
import os
import sys
import warnings
import logging
import math
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import seaborn as sns
from tqdm import tqdm

import statsmodels.api as sm
from statsmodels.genmod.families import (
    Poisson, NegativeBinomial,
    Gaussian, Gamma, InverseGaussian
)
import statsmodels.genmod.families.links as ln
from statsmodels.genmod.generalized_estimating_equations import GEE
from statsmodels.genmod.cov_struct import Independence, Exchangeable
from statsmodels.stats.multitest import multipletests
from scipy.stats import norm

from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill

###############################################################################
# 0) Definitions
###############################################################################

# Original categories
MEDIA_CATEGORIES_ORIGINAL = {
    "Scientific": ["nature","sciam","stat","newscientist"],
    "Left": ["theatlantic","the daily beast","the intercept","mother jones","msnbc","slate","vox","huffpost"],
    "Lean Left": ["ap","axios","cnn","guardian","business insider","nbcnews","npr","nytimes","politico","propublica","wapo","usa today"],
    "Center": ["reuters","marketwatch","financial times","newsweek","forbes"],
    "Lean Right": ["thedispatch","epochtimes","foxbusiness","wsj","national review","washtimes"],
    "Right": ["breitbart","theblaze","daily mail","dailywire","foxnews","nypost","newsmax"],
}

# Collapsed categories
#  - "Left" = old "Left" + "Lean Left"
#  - "Right" = old "Right" + "Lean Right"
#  - "Center" = same as original "Center"
#  - "Scientific" = same as original "Scientific"
MEDIA_CATEGORIES_COLLAPSED = {
    "Scientific": ["nature","sciam","stat","newscientist"],
    "Left": [  # all from "Left" + "Lean Left"
        "theatlantic","the daily beast","the intercept","mother jones","msnbc","slate","vox","huffpost",
        "ap","axios","cnn","guardian","business insider","nbcnews","npr","nytimes","politico","propublica","wapo","usa today"
    ],
    "Center": ["reuters","marketwatch","financial times","newsweek","forbes"],
    "Right": [  # all from "Right" + "Lean Right"
        "thedispatch","epochtimes","foxbusiness","wsj","national review","washtimes",
        "breitbart","theblaze","daily mail","dailywire","foxnews","nypost","newsmax"
    ],
}

###############################################################################
# Configuration for two runs
###############################################################################
# We'll produce distinct output filenames for the original vs. collapsed run
# so that you have "twice as much output."
OUTPUT_DIR_ORIGINAL = "graphs_analysis_original"
OUTPUT_DIR_COLLAPSED = "graphs_analysis_collapsed"

CSV_OUTPUT_DIR_ORIGINAL = "csv_raw_scores_original"
CSV_OUTPUT_DIR_COLLAPSED = "csv_raw_scores_collapsed"

# We'll name the Excel files differently for the two runs:
OUTPUT_EXCEL_MAIN_ORIG = "analysis_readability_main_original.xlsx"
OUTPUT_EXCEL_RAW_ORIG = "analysis_readability_raw_original.xlsx"
OUTPUT_EXCEL_GEE_ORIG = "analysis_readability_gee_original.xlsx"
OUTPUT_EXCEL_PLOTS_ORIG = "analysis_readability_plots_original.xlsx"
OUTPUT_EXCEL_COMBINED_ORIG = "analysis_readability_combined_original.xlsx"

OUTPUT_EXCEL_MAIN_COLL = "analysis_readability_main_collapsed.xlsx"
OUTPUT_EXCEL_RAW_COLL = "analysis_readability_raw_collapsed.xlsx"
OUTPUT_EXCEL_GEE_COLL = "analysis_readability_gee_collapsed.xlsx"
OUTPUT_EXCEL_PLOTS_COLL = "analysis_readability_plots_collapsed.xlsx"
OUTPUT_EXCEL_COMBINED_COLL = "analysis_readability_combined_collapsed.xlsx"

LOG_FILE = "analysis.log"

# We analyze two measures:
MEASURES = [
    "flesch_kincaid_grade_global",
    "gunning_fog_global"
]

###############################################################################
# 1) Setup Logging
###############################################################################
def setup_logging():
    log_format = "%(asctime)s [%(levelname)s] %(module)s - %(message)s"
    logging.basicConfig(filename=LOG_FILE, level=logging.DEBUG, format=log_format)
    console = logging.StreamHandler(sys.stdout)
    console.setLevel(logging.INFO)
    console.setFormatter(logging.Formatter(log_format))
    logging.getLogger().addHandler(console)
    logging.info("Logging initialized (file+console).")

warnings.filterwarnings(
    "ignore",
    message="QIC values obtained using scale=None are not appropriate for comparing models"
)
warnings.filterwarnings("ignore", "The log link alias is deprecated. Use Log instead.")

###############################################################################
# 2) Functions to run the pipeline
###############################################################################
def load_jsonl(jsonl_file):
    logging.info(f"Loading JSONL from {jsonl_file}")
    records=[]
    with open(jsonl_file,"r",encoding="utf-8") as f:
        for line in tqdm(f, desc="Loading JSONL data"):
            rec=json.loads(line)
            records.append(rec)
    df=pd.DataFrame(records)
    logging.debug(f"Loaded DataFrame shape={df.shape}")
    return df

def map_media_outlet_to_category(df, category_map):
    """
    category_map is a dict, e.g. MEDIA_CATEGORIES_ORIGINAL or MEDIA_CATEGORIES_COLLAPSED
    """
    cat_map={}
    for cat, outls in category_map.items():
        for o in outls:
            cat_map[o.lower().strip()] = cat

    if "media_outlet" not in df.columns:
        raise KeyError("'media_outlet' not found in data")

    df["media_outlet_clean"] = df["media_outlet"].str.lower().str.strip()
    df["media_category"] = df["media_outlet_clean"].map(cat_map).fillna("Other")

    unmapped = df[df["media_category"]=="Other"]["media_outlet"].unique()
    if len(unmapped)>0:
        logging.warning(f"Unmapped => {unmapped}")
        print(f"Warning: Not mapped => {unmapped}")
    return df

def chunk_and_save(df, chunk_size, csv_dir):
    logging.info(f"Chunking DataFrame => len={len(df)}, chunk_size={chunk_size}")
    os.makedirs(csv_dir, exist_ok=True)
    for i in range(0,len(df),chunk_size):
        part=df.iloc[i:i+chunk_size]
        out_csv=os.path.join(csv_dir,f"raw_data_part_{(i//chunk_size)+1}.csv")
        part.to_csv(out_csv,index=False)
        print(f"Saved chunk {(i//chunk_size)+1} to {out_csv}")

def print_basic_stats(df):
    logging.info(f"Basic stats => total articles = {len(df)}")
    print("\nSummary Statistics:")
    print("Total number of articles:", len(df))
    if "media_outlet_clean" in df.columns:
        vc=df["media_outlet_clean"].value_counts()
        print("\nArticles per outlet:")
        print(vc)
    if "media_category" in df.columns:
        vc2=df["media_category"].value_counts()
        print("\nArticles per category:")
        print(vc2)
    print()

def aggregate_measures(df, measures):
    logging.info("Aggregating measures by category.")
    # Because we might not know in advance the category keys:
    cat_vals = df["media_category"].dropna().unique()
    recs = []
    for cat in cat_vals:
        sub = df[df["media_category"] == cat]
        for m in measures:
            if m in sub.columns:
                numeric_vals = pd.to_numeric(sub[m], errors="coerce").dropna()
                the_sum = numeric_vals.sum()
                the_count = numeric_vals.count()
            else:
                the_sum, the_count = (0, 0)

            recs.append({
                "Media Category": cat,
                "Measure": m,
                "Sum": the_sum,
                "Count": the_count,
            })
    return pd.DataFrame(recs)

def calculate_averages(agg_df):
    logging.info("Calculating measure Average.")
    def sdiv(a,b): return a/b if b>0 else None
    agg_df["Average"] = agg_df.apply(lambda r: sdiv(r["Sum"], r["Count"]), axis=1)
    return agg_df

def calculate_mean_median(agg_df, measures):
    logging.info("Computing global mean/median of measure averages.")
    rows=[]
    for m in measures:
        sub = agg_df[agg_df["Measure"] == m]
        avg_col = sub["Average"].dropna()
        rows.append({
            "Measure": m,
            "Mean_of_Averages": avg_col.mean() if len(avg_col)>0 else None,
            "Median_of_Averages": avg_col.median() if len(avg_col)>0 else None
        })
    return pd.DataFrame(rows)

def save_aggregated_scores_to_csv(agg_df, out_dir):
    os.makedirs(out_dir, exist_ok=True)
    fn=os.path.join(out_dir,"aggregated_scores.csv")
    agg_df.to_csv(fn,index=False)
    print(f"Aggregated scores => {fn}")
    logging.info(f"Aggregated => {fn}")

def plot_statistics(agg_df, out_dir):
    logging.info("Plotting bar charts for measures across categories.")
    os.makedirs(out_dir, exist_ok=True)
    sns.set_style("whitegrid")
    for m in agg_df["Measure"].unique():
        sub = agg_df[agg_df["Measure"] == m]
        plt.figure(figsize=(10,6))
        sns.barplot(x="Media Category", y="Average", data=sub, color="steelblue")
        plt.title(f"Mean '{m}' by Media Category")
        plt.xticks(rotation=45, ha="right")
        plt.tight_layout()
        outpath = os.path.join(out_dir, f"measure_{m}.png")
        try:
            plt.savefig(outpath)
        except:
            pass
        plt.close()

###############################################################################
# Scale computations (same as before)
###############################################################################
def compute_pearson_scale(y, mu, df_resid):
    if df_resid<=0: return np.nan
    y_arr=np.asarray(y)
    mu_arr=np.asarray(mu)
    r=(y_arr-mu_arr)/np.sqrt(mu_arr+1e-9)
    return np.sum(r**2)/df_resid

def compute_deviance_scale(y, mu, df_resid):
    if df_resid<=0: return np.nan
    y_arr=np.asarray(y)
    mu_arr=np.asarray(mu)
    arr=[]
    for obs,lam in zip(y_arr,mu_arr):
        if obs>0 and lam>0:
            arr.append(obs*math.log(obs/lam)-(obs-lam))
        elif obs==0:
            arr.append(-(obs-lam))
        else:
            arr.append(np.nan)
    dev=2*np.nansum(arr)
    return dev/df_resid if df_resid>0 else np.nan

def compute_ub_scale(y, mu, df_resid):
    p=compute_pearson_scale(y, mu, df_resid)
    if np.isnan(p): return p
    return 1.1*p

def compute_bc_scale(y, mu, df_resid):
    d=compute_deviance_scale(y, mu, df_resid)
    if np.isnan(d):
        return d
    return 0.9*d

###############################################################################
# Additional checks
###############################################################################
def check_residuals_and_correlation(final_fit):
    y_arr=np.asarray(final_fit.model.endog)
    mu_arr=np.asarray(final_fit.fittedvalues)
    pearson_res_arr=(y_arr - mu_arr)/np.sqrt(mu_arr+1e-9)
    pearson_res_arr = np.asarray(pearson_res_arr)

    mean_res=np.mean(pearson_res_arr)
    std_res=np.std(pearson_res_arr)

    clusters=final_fit.model.groups
    clusters_arr=np.asarray(clusters)

    cluster_map={}
    for i in range(len(clusters_arr)):
        cid=clusters_arr[i]
        cluster_map.setdefault(cid,[]).append(pearson_res_arr[i])

    wcorr=[]
    for cid, arr in cluster_map.items():
        arr=np.array(arr)
        if len(arr)>1:
            cmat=np.corrcoef(arr)
            if cmat.ndim==2 and cmat.shape[0]>1:
                n2=len(arr)
                sum_offdiag=np.sum(cmat)-n2
                avg_off=sum_offdiag/(n2*(n2-1)) if n2>1 else 0
                wcorr.append(avg_off)

    avg_corr=np.mean(wcorr) if len(wcorr)>0 else 0.0

    deviance=getattr(final_fit,"pearson_chi2",np.nan)
    dfres=getattr(final_fit,"df_resid",1)
    overdisp=np.nan
    if dfres>0 and not np.isnan(deviance):
        overdisp=deviance/dfres

    assess=""
    if abs(mean_res)>1:
        assess+="Mean Pearson residual>1 => misfit? "
    if std_res>2:
        assess+="Residual std>2 => outliers? "
    if abs(avg_corr)>0.3:
        assess+=f"Within-cluster corr={avg_corr:.2f} => structure suspect. "
    if (not np.isnan(overdisp)) and overdisp>2:
        assess+=f"Overdisp={overdisp:.2f} => consider NB or other approach. "

    if assess=="":
        assess="No major issues from these checks."

    return {
        "mean_pearson_res": mean_res,
        "std_pearson_res": std_res,
        "avg_within_corr": avg_corr,
        "overdisp_ratio": overdisp,
        "assessment": assess
    }

def pseudo_likelihood_check(final_fit):
    from statsmodels.genmod.families.family import NegativeBinomial
    fam=final_fit.model.family
    if not isinstance(fam, sm.genmod.families.family.Poisson):
        return {
            "NB_QIC":None,
            "Poisson_QIC":None,
            "diff_QIC":None,
            "conclusion":"Non-Poisson => skip NB check"
        }
    data=final_fit.model.data.frame
    cov_struct=final_fit.model.cov_struct
    nb_model=GEE.from_formula(final_fit.model.formula,
                              groups=final_fit.model.groups,
                              data=data,
                              family=NegativeBinomial(alpha=1.0),
                              cov_struct=cov_struct,
                              maxiter=200)
    nb_res=nb_model.fit()
    nb_qic=nb_res.qic()
    if isinstance(nb_qic, tuple):
        nb_qic=nb_qic[0]
    old_qic=final_fit.qic()
    if isinstance(old_qic, tuple):
        old_qic=old_qic[0]

    diff_qic=None
    conclusion=""
    if isinstance(nb_qic,(float,int)) and isinstance(old_qic,(float,int)):
        diff_qic=old_qic-nb_qic
        conclusion="NegBin better" if diff_qic>0 else "No NB improvement"
    else:
        conclusion="Could not compare QIC"

    return {
        "NB_QIC":nb_qic,
        "Poisson_QIC":old_qic,
        "diff_QIC":diff_qic,
        "conclusion":conclusion
    }

def cross_validation_gee(data_with_score, formula, group_col, family, cov_struct, n_folds=3):
    cluster_ids=data_with_score[group_col].unique()
    np.random.shuffle(cluster_ids)
    folds=np.array_split(cluster_ids,n_folds)
    metrics=[]
    for i in range(n_folds):
        testc=set(folds[i])
        train_df=data_with_score[~data_with_score[group_col].isin(testc)]
        test_df=data_with_score[data_with_score[group_col].isin(testc)]
        if len(train_df)<1 or len(test_df)<1:
            continue
        mod=GEE.from_formula(formula, groups=group_col,
                             data=train_df, family=family, cov_struct=cov_struct, maxiter=200)
        res=mod.fit()
        pred=res.predict(test_df)
        obs=test_df[res.model.endog_names]
        mse=np.mean((obs-pred)**2)
        metrics.append(mse)
    return np.mean(metrics) if len(metrics)>0 else np.nan

def sensitivity_analysis_correlation(df, formula, group_col):
    from statsmodels.genmod.families import Poisson, Gaussian
    from statsmodels.genmod.families.family import NegativeBinomial
    alt_families=[Poisson(), NegativeBinomial(alpha=1.0), Gaussian()]
    alt_structs=[Independence(), Exchangeable()]
    results=[]
    for fam in alt_families:
        for covs in alt_structs:
            mod=GEE.from_formula(formula, groups=group_col, data=df,
                                 family=fam, cov_struct=covs, maxiter=200)
            try:
                r=mod.fit()
                qic_val=r.qic()
                if isinstance(qic_val, tuple):
                    qic_val=qic_val[0]
                param3=r.params.head(3).round(3).to_dict()
                results.append({
                    "Family":fam.__class__.__name__,
                    "CovStruct":covs.__class__.__name__,
                    "QIC":qic_val,
                    "ParamSample":param3
                })
            except:
                results.append({
                    "Family":fam.__class__.__name__,
                    "CovStruct":covs.__class__.__name__,
                    "QIC":np.nan,
                    "ParamSample":"fail"
                })
    return pd.DataFrame(results)

def bootstrap_gee(df, formula, group_col,
                  B=1000, family=Poisson(), cov_struct=Independence()):
    cluster_ids=df[group_col].unique()
    param_records=[]
    for _ in range(B):
        sample_ids=np.random.choice(cluster_ids, size=len(cluster_ids), replace=True)
        pieces=[]
        for cid in sample_ids:
            pieces.append(df[df[group_col]==cid])
        boot_df=pd.concat(pieces,ignore_index=True)
        mod=GEE.from_formula(formula, groups=group_col, data=boot_df,
                             family=family, cov_struct=cov_struct, maxiter=200)
        res=mod.fit()
        param_records.append(res.params)
    return pd.DataFrame(param_records)

def coverage_simulation_gee(final_fit, n_sims=1000, alpha=0.05):
    model = final_fit.model
    family = model.family
    cov_struct = model.cov_struct

    data = model.data.frame.copy()
    clusters = final_fit.model.groups
    data["_cluster_id"] = pd.Series(clusters, index=data.index)
    data["_cluster_id"] = data["_cluster_id"].astype("category").cat.codes

    beta_true = final_fit.params.values

    mu = final_fit.fittedvalues
    sigma_est = np.sqrt(final_fit.scale) if hasattr(final_fit, "scale") else 1.0
    alpha_nb = getattr(family, "alpha", 1.0)

    z_val = abs(norm.ppf(alpha/2))
    coverage_counts = np.zeros(len(beta_true))

    for _ in range(n_sims):
        sim_data = data.copy()

        if family.__class__.__name__ == "Poisson":
            sim_data[model.endog_names] = np.random.poisson(lam=mu)
        elif family.__class__.__name__ == "NegativeBinomial":
            sim_draws = []
            for val in mu:
                r_ = 1.0/alpha_nb if alpha_nb!=0 else 10
                p_ = r_/(r_+val) if (r_+val)!=0 else 0.99
                draw = np.random.negative_binomial(r_, p_)
                sim_draws.append(draw)
            sim_data[model.endog_names] = sim_draws
        elif family.__class__.__name__ == "Gaussian":
            sim_data[model.endog_names] = np.random.normal(loc=mu, scale=sigma_est)
        else:
            sim_data[model.endog_names] = mu

        sim_model = GEE.from_formula(model.formula,
                                     groups=sim_data["_cluster_id"],
                                     data=sim_data,
                                     family=family, cov_struct=cov_struct, maxiter=200)
        sim_res = sim_model.fit()

        sim_params = sim_res.params.values
        sim_cov = sim_res.cov_params()
        sim_se = np.sqrt(np.diag(sim_cov))

        ci_lower = sim_params - z_val*sim_se
        ci_upper = sim_params + z_val*sim_se

        inside = (beta_true >= ci_lower) & (beta_true <= ci_upper)
        coverage_counts += inside.astype(int)

    coverage = coverage_counts / n_sims

    return pd.DataFrame({
        "Parameter": final_fit.params.index,
        "OriginalParam": beta_true,
        "Coverage": coverage
    })

###############################################################################
# GEE alpha search in [0..30]
###############################################################################
def try_all_families_and_scales(df, formula, group_col):
    from statsmodels.genmod.families.family import NegativeBinomial
    families_no_nb = [
        Poisson(),
        Gaussian(),
        Gamma(link=ln.log()),
        InverseGaussian()
    ]

    cor_structs=[Independence(), Exchangeable()]
    scale_opts=["none","pearson","deviance","ub","bc"]

    best_qic=np.inf
    best_tuple=None
    all_results=[]

    # Non-NB families
    for fam in families_no_nb:
        fam_name = fam.__class__.__name__
        for cov_obj in cor_structs:
            cov_name = cov_obj.__class__.__name__
            for sc_opt in scale_opts:
                try:
                    model=GEE.from_formula(formula, groups=group_col,
                                           data=df, family=fam, cov_struct=cov_obj, maxiter=200)
                    base_res=model.fit(scale=None)
                    y=base_res.model.endog
                    mu=base_res.fittedvalues
                    n=len(y)
                    p=len(base_res.params)
                    dfresid=n-p
                    final_res=base_res

                    if sc_opt!="none" and dfresid>0:
                        val=None
                        if sc_opt=="pearson":
                            val=compute_pearson_scale(y,mu,dfresid)
                        elif sc_opt=="deviance":
                            val=compute_deviance_scale(y,mu,dfresid)
                        elif sc_opt=="ub":
                            val=compute_ub_scale(y,mu,dfresid)
                        elif sc_opt=="bc":
                            val=compute_bc_scale(y,mu,dfresid)
                        if val is not None and not np.isnan(val):
                            final_res = model.fit(scale=val)

                    qic_val=final_res.qic()
                    if isinstance(qic_val, tuple):
                        qic_val=qic_val[0]

                    alpha_label = "alpha=None"
                    all_results.append((fam_name, cov_name, sc_opt, alpha_label, qic_val))

                    if qic_val<best_qic:
                        best_qic=qic_val
                        best_tuple=(fam_name,cov_name,sc_opt,None,qic_val)

                except Exception as e:
                    logging.warning(f"Fit fail => {fam_name}+{cov_name}+{sc_opt} => {e}")
                    all_results.append((fam_name,cov_name,sc_opt,"alpha=None",np.nan))

    # NegativeBinomial alpha search
    alpha_range = range(0, 31)  # 0..30
    for cov_obj in cor_structs:
        cov_name = cov_obj.__class__.__name__
        for sc_opt in scale_opts:
            for alpha_candidate in alpha_range:
                try:
                    nb_fam = NegativeBinomial(alpha=float(alpha_candidate))
                    model=GEE.from_formula(formula, groups=group_col,
                                           data=df, family=nb_fam, cov_struct=cov_obj, maxiter=200)
                    base_res=model.fit(scale=None)
                    y=base_res.model.endog
                    mu=base_res.fittedvalues
                    n=len(y)
                    p=len(base_res.params)
                    dfresid=n-p
                    final_res=base_res

                    if sc_opt!="none" and dfresid>0:
                        val=None
                        if sc_opt=="pearson":
                            val=compute_pearson_scale(y,mu,dfresid)
                        elif sc_opt=="deviance":
                            val=compute_deviance_scale(y,mu,dfresid)
                        elif sc_opt=="ub":
                            val=compute_ub_scale(y,mu,dfresid)
                        elif sc_opt=="bc":
                            val=compute_bc_scale(y,mu,dfresid)
                        if val is not None and not np.isnan(val):
                            final_res = model.fit(scale=val)

                    qic_val=final_res.qic()
                    if isinstance(qic_val, tuple):
                        qic_val=qic_val[0]

                    alpha_label = f"alpha={alpha_candidate}"
                    all_results.append(("NegativeBinomial", cov_name, sc_opt, alpha_label, qic_val))

                    if qic_val<best_qic:
                        best_qic=qic_val
                        best_tuple=("NegativeBinomial",cov_name,sc_opt,float(alpha_candidate),qic_val)

                except Exception as e:
                    logging.warning(f"NB Fit fail => alpha={alpha_candidate}+{cov_name}+{sc_opt} => {e}")
                    all_results.append(("NegativeBinomial", cov_name, sc_opt,
                                        f"alpha={alpha_candidate}", np.nan))

    return best_tuple, all_results

def run_gee_for_measure_best_qic(df, measure):
    d2 = df.copy()
    if measure not in d2.columns:
        return None

    d2[measure] = pd.to_numeric(d2[measure], errors="coerce")
    d2 = d2.dropna(subset=["media_outlet_clean","media_category", measure])
    if len(d2) < 2 or d2["media_category"].nunique() < 2:
        return None

    d2["media_category"] = d2["media_category"].astype("category")
    d2["_score_col"] = d2[measure].clip(lower=0)

    best_tuple, combos = try_all_families_and_scales(
        d2, "_score_col ~ media_category", "media_outlet_clean"
    )
    if best_tuple is None:
        return None

    famName, corName, scOpt, alphaVal, bestQICVal = best_tuple

    recs=[]
    for row in combos:
        fam_, cov_, sc_, alpha_str, q_ = row
        recs.append({
            "Family": fam_,
            "CovStruct": cov_,
            "Scale": sc_,
            "Alpha": alpha_str,
            "QIC": q_
        })
    combos_df = pd.DataFrame(recs)

    return {
        "Measure": measure,
        "Best_Family": famName,
        "Best_Structure": corName,
        "Best_Scale": scOpt,
        "Best_Alpha": alphaVal,
        "Best_QIC_main": bestQICVal,
        "AllCombos": combos_df
    }

def run_gee_analyses_best_qic(df):
    logging.info("Running best QIC approach multi-families for each measure.")
    best_list=[]
    combos_list=[]
    for m in MEASURES:
        info = run_gee_for_measure_best_qic(df, m)
        if info is not None:
            best_list.append({
                "Measure": info["Measure"],
                "Best_Family": info["Best_Family"],
                "Best_Structure": info["Best_Structure"],
                "Best_Scale": info["Best_Scale"],
                "Best_Alpha": info["Best_Alpha"],
                "Best_QIC_main": info["Best_QIC_main"]
            })
            cdf=info["AllCombos"]
            cdf["Measure"] = m
            combos_list.append(cdf)

    df_best = pd.DataFrame(best_list)
    df_all = pd.concat(combos_list, ignore_index=True) if combos_list else pd.DataFrame()
    return df_best, df_all

###############################################################################
# Refit + advanced checks
###############################################################################
def refit_best_gee_with_scale(df, measure, fam_name, struct, scale_name, alpha_val=None):
    from statsmodels.genmod.families import Poisson, Gaussian, Gamma, InverseGaussian
    from statsmodels.genmod.families.family import NegativeBinomial

    if fam_name=="Poisson":
        fam_obj=Poisson()
    elif fam_name=="NegativeBinomial":
        if alpha_val is None:
            alpha_val=1.0
        fam_obj=NegativeBinomial(alpha=alpha_val)
    elif fam_name=="Gaussian":
        fam_obj=Gaussian()
    elif fam_name=="Gamma":
        fam_obj=Gamma(link=ln.log())
    elif fam_name=="InverseGaussian":
        fam_obj=InverseGaussian()
    else:
        logging.warning(f"Unknown family={fam_name}")
        return None, None

    d2 = df.copy()
    if measure not in d2.columns:
        return None, None
    d2[measure] = pd.to_numeric(d2[measure], errors="coerce")
    d2 = d2.dropna(subset=["media_outlet_clean","media_category", measure])
    if len(d2)<2 or d2["media_category"].nunique()<2:
        return None, None

    d2["media_category"] = d2["media_category"].astype("category")
    d2["_score_col"] = d2[measure]

    if struct=="Independence":
        cov_obj=Independence()
    else:
        cov_obj=Exchangeable()

    if fam_name=="NegativeBinomial":
        logging.info(f"Re-fitting NB GEE with alpha={alpha_val}")
        print(f"DEBUG: Re-fitting NB GEE with alpha={alpha_val}")
    else:
        logging.info(f"Re-fitting {fam_name} GEE (no alpha param)")

    model=GEE.from_formula("_score_col ~ media_category",
                           groups="media_outlet_clean",
                           data=d2,
                           family=fam_obj,
                           cov_struct=cov_obj,
                           maxiter=200)
    bres=model.fit(scale=None)
    y=np.asarray(bres.model.endog)
    mu=np.asarray(bres.fittedvalues)
    n=len(y)
    p=len(bres.params)
    dfresid=n-p

    if scale_name=="none" or dfresid<=0:
        return d2,bres
    else:
        val=None
        if scale_name=="pearson":
            val=compute_pearson_scale(y,mu,dfresid)
        elif scale_name=="deviance":
            val=compute_deviance_scale(y,mu,dfresid)
        elif scale_name=="ub":
            val=compute_ub_scale(y,mu,dfresid)
        elif scale_name=="bc":
            val=compute_bc_scale(y,mu,dfresid)
        if val is None or np.isnan(val):
            return d2,bres
        final_res=model.fit(scale=val)
        return d2, final_res

def pairwise_and_diagnostics(df, measure, fam_name, struct, scale_name, alpha_val=None):
    d2, final_fit=refit_best_gee_with_scale(df, measure, fam_name, struct, scale_name, alpha_val=alpha_val)
    if final_fit is None:
        return (None,)*9

    summary_txt=final_fit.summary().as_text()
    if fam_name=="NegativeBinomial":
        lines = summary_txt.split("\n")
        lines.append(f"(Alpha used: {alpha_val})")
        summary_txt = "\n".join(lines)

    mdf=final_fit.model.data.frame
    mdf["media_category"] = mdf["media_category"].astype("category")

    diag=check_residuals_and_correlation(final_fit)
    pseudo_dict=pseudo_likelihood_check(final_fit)

    cv_mse=cross_validation_gee(d2,"_score_col ~ media_category","media_outlet_clean",
                                final_fit.model.family,
                                final_fit.model.cov_struct, n_folds=3)
    sens_df=sensitivity_analysis_correlation(d2,"_score_col ~ media_category","media_outlet_clean")

    boot_df=bootstrap_gee(d2,"_score_col ~ media_category",
                          "media_outlet_clean", B=1000,
                          family=final_fit.model.family,
                          cov_struct=final_fit.model.cov_struct)

    coverage_df = coverage_simulation_gee(final_fit, n_sims=1000, alpha=0.05)
    coverage_df["Measure"] = measure
    coverage_df["Family"] = fam_name
    coverage_df["Structure"] = struct
    coverage_df["Scale"] = scale_name
    coverage_df["Alpha"] = alpha_val

    params=final_fit.params
    cov=final_fit.cov_params()

    cats=mdf["media_category"].cat.categories
    ref=cats[0]
    idx_map={ref:0}
    for c in cats[1:]:
        nm=f"media_category[T.{c}]"
        idx_map[c]=final_fit.model.exog_names.index(nm)

    pair_list=[]
    for i in range(len(cats)):
        for j in range(i+1,len(cats)):
            ca,cb=cats[i],cats[j]
            con=np.zeros(len(params))
            if ca==ref and cb!=ref:
                con[idx_map[cb]]=-1.0
            elif cb==ref and ca!=ref:
                con[idx_map[ca]]=1.0
            else:
                con[idx_map[ca]]=1.0
                con[idx_map[cb]]=-1.0
            diff_est=con@params
            diff_var=con@cov@con
            diff_se=np.sqrt(diff_var)
            z=diff_est/diff_se
            pval=2*(1-norm.cdf(abs(z)))
            pair_list.append((ca,cb,diff_est,diff_se,z,pval))

    pair_df=pd.DataFrame(pair_list,columns=["CategoryA","CategoryB","Difference","SE","Z","p_value"])
    rej, p_adj,_,_=multipletests(pair_df["p_value"],method="fdr_bh")
    pair_df["p_value_adj"]=p_adj
    pair_df["reject_H0"]=rej

    cat_list=list(cats)
    cat_index_map={cat_list[i]:i+1 for i in range(len(cat_list))}
    diff_map={c:set() for c in cat_list}
    for i,row in pair_df.iterrows():
        A=row["CategoryA"]
        B=row["CategoryB"]
        if row["reject_H0"]:
            diff_map[A].add(cat_index_map[B])
            diff_map[B].add(cat_index_map[A])

    rows=[]
    for c in cat_list:
        diffs=sorted(list(diff_map[c]))
        diffs_str=",".join(str(x) for x in diffs)
        rows.append((c,diffs_str))
    diffIDs_df=pd.DataFrame(rows, columns=["Category","DiffIDs"])

    return summary_txt, pair_df, diffIDs_df, diag, pseudo_dict, cv_mse, sens_df, boot_df, coverage_df

def compile_results_into_multiple_workbooks(
    aggregated_df, stats_df, raw_df,
    df_best_qic, df_all_combos,
    plots_dir,
    main_excel, raw_excel, gee_excel, plots_excel, combined_excel,
    df_full
):
    """
    Writes results across multiple Excel files for one run 
    (either original or collapsed).
    """
    from openpyxl import Workbook
    from openpyxl.utils.dataframe import dataframe_to_rows

    with pd.ExcelWriter(main_excel, engine="openpyxl") as w:
        aggregated_df.to_excel(w, sheet_name="Aggregated_Scores", index=False)
        stats_df.to_excel(w, sheet_name="Mean_Median_Statistics", index=False)

    with pd.ExcelWriter(raw_excel, engine="openpyxl") as w:
        raw_df.to_excel(w, sheet_name="Raw_Data", index=False)
        for m in MEASURES:
            if m in raw_df.columns:
                s_df=raw_df[["media_category","media_outlet", m]].copy()
                sname = f"Raw_{m[:25]}"
                s_df.to_excel(w, sheet_name=sname, index=False)

    diag_records=[]
    cv_records=[]
    sens_records=[]
    coverage_records=[]
    idx_rows=[]
    bootstrap_param_records=[]

    with pd.ExcelWriter(gee_excel, engine="openpyxl") as writer:
        if not df_all_combos.empty:
            df_all_combos.to_excel(writer, sheet_name="All_Combos", index=False)

        for i,row in df_best_qic.iterrows():
            m=row["Measure"]
            fam=row["Best_Family"]
            st=row["Best_Structure"]
            sc=row["Best_Scale"]
            alpha_val=row.get("Best_Alpha", None)
            best_qic=row["Best_QIC_main"]

            sh_name=f"BestQIC_{m[:15]}"

            out=pairwise_and_diagnostics(df_full, m, fam, st, sc, alpha_val=alpha_val)
            if out[0] is None:
                tmp_df=pd.DataFrame({"Summary":["No valid model or not enough data."]})
                tmp_df.to_excel(writer, sheet_name=sh_name, index=False)
                continue

            (
                summary_txt, pair_df, diffIDs_df, diag_dict, pseudo_dict,
                cv_mse, sens_df_, boot_df, coverage_df
            ) = out

            lines=summary_txt.split("\n")
            sumdf=pd.DataFrame({"GEE_Summary":lines})
            sumdf.to_excel(writer, sheet_name=sh_name, index=False)

            sr=len(sumdf)+2
            pair_df.to_excel(writer, sheet_name=sh_name, index=False, startrow=sr)
            ws=writer.sheets[sh_name]
            for row_idx in range(len(pair_df)):
                if pair_df.loc[row_idx,"reject_H0"]:
                    rrow=sr+1+row_idx
                    for col_idx in range(1, pair_df.shape[1]+1):
                        cell=ws.cell(row=rrow+1, column=col_idx)
                        cell.fill=PatternFill(
                            fill_type="solid",
                            start_color="FFFF0000",
                            end_color="FFFF0000"
                        )

            sr2=sr+len(pair_df)+2
            if diffIDs_df is not None and not diffIDs_df.empty:
                diffIDs_df.to_excel(writer, sheet_name=sh_name, index=False, startrow=sr2)

            diag_records.append({
                "Measure": m,
                "Family": fam,
                "Structure": st,
                "Scale": sc,
                "Alpha": alpha_val,
                "mean_pearson_res": diag_dict["mean_pearson_res"],
                "std_pearson_res": diag_dict["std_pearson_res"],
                "avg_within_corr": diag_dict["avg_within_corr"],
                "overdisp_ratio": diag_dict["overdisp_ratio"],
                "assessment": diag_dict["assessment"],
                "NB_QIC": pseudo_dict["NB_QIC"],
                "Poisson_QIC": pseudo_dict["Poisson_QIC"],
                "diff_QIC": pseudo_dict["diff_QIC"],
                "pseudo_conclusion": pseudo_dict["conclusion"]
            })

            cv_records.append({
                "Measure": m,
                "Family": fam,
                "Structure": st,
                "Scale": sc,
                "Alpha": alpha_val,
                "CV_MSE": cv_mse
            })

            tmp_sens=sens_df_.copy()
            tmp_sens["Measure"]=m
            tmp_sens["Alpha"]=alpha_val
            sens_records.append(tmp_sens)

            for param_name in boot_df.columns:
                param_vals = boot_df[param_name]
                pm = param_vals.mean()
                ps = param_vals.std()
                bootstrap_param_records.append({
                    "Measure": m,
                    "Family": fam,
                    "Structure": st,
                    "Scale": sc,
                    "Alpha": alpha_val,
                    "Parameter": param_name,
                    "BootMean": round(pm,4),
                    "BootStd": round(ps,4)
                })

            coverage_records.append(coverage_df)

            idx_rows.append({
                "Measure": m,
                "Family": fam,
                "Structure": st,
                "Scale": sc,
                "Alpha": alpha_val,
                "BestQIC": best_qic
            })

        idxdf=pd.DataFrame(idx_rows)
        idxdf.to_excel(writer, sheet_name="BestQIC_Index", index=False)

        diag_df=pd.DataFrame(diag_records)
        diag_df.to_excel(writer, sheet_name="Diagnostics", index=False)

        cv_df=pd.DataFrame(cv_records)
        cv_df.to_excel(writer, sheet_name="CrossValidation_Res", index=False)

        if len(sens_records)>0:
            sens_all=pd.concat(sens_records,ignore_index=True)
            sens_all.to_excel(writer, sheet_name="Sensitivity_Analysis", index=False)

        if len(bootstrap_param_records)>0:
            boot_all=pd.DataFrame(bootstrap_param_records)
            boot_all.to_excel(writer, sheet_name="ClusterBootstrap_Validation", index=False)

        if len(coverage_records)>0:
            coverage_all=pd.concat(coverage_records, ignore_index=True)
            coverage_all.to_excel(writer, sheet_name="CoverageSimulation_Res", index=False)

    wb_plots=Workbook()
    if "Sheet" in wb_plots.sheetnames:
        wb_plots.remove(wb_plots["Sheet"])
    any_sheets=False
    for m in MEASURES:
        p_path = os.path.join(plots_dir, f"measure_{m}.png")
        if os.path.exists(p_path):
            st = f"Measure_{m[:20]}"
            ws = wb_plots.create_sheet(title=st)
            try:
                img=ExcelImage(p_path)
                img.anchor="A1"
                ws.add_image(img)
            except:
                pass
            any_sheets=True

    if not any_sheets:
        wb_plots.create_sheet("NoPlotsFound")
    wb_plots.save(plots_excel)

    # Combined
    raw_clean=raw_df.copy()
    raw_clean=raw_clean.applymap(lambda x:", ".join(x) if isinstance(x,list) else x)
    wb_comb=Workbook()
    if "Sheet" in wb_comb.sheetnames:
        wb_comb.remove(wb_comb["Sheet"])

    ws_agg=wb_comb.create_sheet("Aggregated_Scores")
    for r in dataframe_to_rows(aggregated_df,index=False,header=True):
        ws_agg.append(r)

    ws_stats=wb_comb.create_sheet("Mean_Median_Statistics")
    for r in dataframe_to_rows(stats_df,index=False,header=True):
        ws_stats.append(r)

    ws_raw=wb_comb.create_sheet("Raw_Data")
    for r in dataframe_to_rows(raw_clean,index=False,header=True):
        ws_raw.append(r)

    ws_best=wb_comb.create_sheet("BestQIC_Table")
    for r in dataframe_to_rows(df_best_qic,index=False,header=True):
        ws_best.append(r)

    wb_comb.save(combined_excel)


###############################################################################
# Main pipeline run function
###############################################################################
def run_pipeline_once(
    df_in,
    category_map,
    graphs_dir,
    csv_dir,
    main_excel,
    raw_excel,
    gee_excel,
    plots_excel,
    combined_excel
):
    """
    1) Map categories
    2) Chunk + Basic Stats
    3) Aggregate + Plot
    4) QIC approach
    5) Compile
    """
    # Copy the original to avoid messing
    df=df_in.copy()

    # Map categories
    df=map_media_outlet_to_category(df, category_map)

    # chunk
    chunk_and_save(df, 20000, csv_dir)
    print_basic_stats(df)

    # aggregate
    agg_df = aggregate_measures(df, MEASURES)
    agg_df = calculate_averages(agg_df)
    stats_df = calculate_mean_median(agg_df, MEASURES)
    save_aggregated_scores_to_csv(agg_df, csv_dir)
    plot_statistics(agg_df, graphs_dir)

    # QIC
    print("Fitting best QIC approach for each measure (with NB alpha in [0..30])...")
    df_best, df_allcombos = run_gee_analyses_best_qic(df)
    print("Best QIC approach completed.\n")
    print(df_best)

    # compile
    compile_results_into_multiple_workbooks(
        aggregated_df=agg_df,
        stats_df=stats_df,
        raw_df=df,
        df_best_qic=df_best,
        df_all_combos=df_allcombos,
        plots_dir=graphs_dir,
        main_excel=main_excel,
        raw_excel=raw_excel,
        gee_excel=gee_excel,
        plots_excel=plots_excel,
        combined_excel=combined_excel,
        df_full=df
    )

def main():
    setup_logging()
    logging.info("Starting dual-run GEE approach with NB alpha grid search, coverage, etc.")

    INPUT_JSONL_FILE = "processed_all_articles_fixed_3.jsonl"

    df=load_jsonl(INPUT_JSONL_FILE)
    print(f"Total articles loaded: {len(df)}")

    # --- 1) Run with original categories ---
    print("\n===== RUN 1: Original Categories =====")
    os.makedirs(OUTPUT_DIR_ORIGINAL, exist_ok=True)
    run_pipeline_once(
        df_in=df,
        category_map=MEDIA_CATEGORIES_ORIGINAL,
        graphs_dir=OUTPUT_DIR_ORIGINAL,
        csv_dir=CSV_OUTPUT_DIR_ORIGINAL,
        main_excel=OUTPUT_EXCEL_MAIN_ORIG,
        raw_excel=OUTPUT_EXCEL_RAW_ORIG,
        gee_excel=OUTPUT_EXCEL_GEE_ORIG,
        plots_excel=OUTPUT_EXCEL_PLOTS_ORIG,
        combined_excel=OUTPUT_EXCEL_COMBINED_ORIG
    )

    # --- 2) Run with collapsed categories ---
    print("\n===== RUN 2: Collapsed Categories =====")
    os.makedirs(OUTPUT_DIR_COLLAPSED, exist_ok=True)
    run_pipeline_once(
        df_in=df,
        category_map=MEDIA_CATEGORIES_COLLAPSED,
        graphs_dir=OUTPUT_DIR_COLLAPSED,
        csv_dir=CSV_OUTPUT_DIR_COLLAPSED,
        main_excel=OUTPUT_EXCEL_MAIN_COLL,
        raw_excel=OUTPUT_EXCEL_RAW_COLL,
        gee_excel=OUTPUT_EXCEL_GEE_COLL,
        plots_excel=OUTPUT_EXCEL_PLOTS_COLL,
        combined_excel=OUTPUT_EXCEL_COMBINED_COLL
    )

    print("Analysis completed successfully for both original & collapsed categories.")
    logging.info("Analysis completed successfully for both original & collapsed categories.")

if __name__=="__main__":
    main()
