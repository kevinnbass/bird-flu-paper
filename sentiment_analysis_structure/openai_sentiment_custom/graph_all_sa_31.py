import json
import pandas as pd
import matplotlib.pyplot as plt
import seaborn as sns
import os
import logging
from tqdm import tqdm
from openpyxl.drawing.image import Image as ExcelImage
import re
import statsmodels.api as sm
from statsmodels.formula.api import mixedlm
from statsmodels.stats.multicomp import pairwise_tukeyhsd
import sys
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import numpy as np

# ------------------------------ #
#        Configuration           #
# ------------------------------ #

INPUT_JSONL_FILE = 'processed_all_articles_with_fulltext_sentiment_analysis.jsonl'
OUTPUT_DIR = 'graphs_analysis'
os.makedirs(OUTPUT_DIR, exist_ok=True)

CSV_OUTPUT_DIR = 'csv_raw_scores'
os.makedirs(CSV_OUTPUT_DIR, exist_ok=True)

OUTPUT_EXCEL_MAIN = 'analysis_main.xlsx'
OUTPUT_EXCEL_RAW = 'analysis_raw.xlsx'
OUTPUT_EXCEL_LMM = 'analysis_lmm.xlsx'
OUTPUT_EXCEL_PLOTS = 'analysis_plots.xlsx'
OUTPUT_EXCEL_COMBINED = 'analysis_combined.xlsx'

# For the filtered version
OUTPUT_DIR_FILTERED = 'graphs_analysis_filtered'
os.makedirs(OUTPUT_DIR_FILTERED, exist_ok=True)

CSV_OUTPUT_DIR_FILTERED = 'csv_raw_scores_filtered'
os.makedirs(CSV_OUTPUT_DIR_FILTERED, exist_ok=True)

OUTPUT_EXCEL_MAIN_FILTERED = 'analysis_main_filtered.xlsx'
OUTPUT_EXCEL_RAW_FILTERED = 'analysis_raw_filtered.xlsx'
OUTPUT_EXCEL_LMM_FILTERED = 'analysis_lmm_filtered.xlsx'
OUTPUT_EXCEL_PLOTS_FILTERED = 'analysis_plots_filtered.xlsx'
OUTPUT_EXCEL_COMBINED_FILTERED = 'analysis_combined_filtered.xlsx'

CATEGORIES = [
    'joy', 'sadness', 'anger', 'fear',
    'surprise', 'disgust', 'trust', 'anticipation',
    'negative_sentiment', 'positive_sentiment'
]

MEDIA_CATEGORIES = {
    'Scientific': ['nature', 'sciam', 'stat', 'newscientist'],
    'Left': ['theatlantic', 'the daily beast', 'the intercept', 'mother jones', 'msnbc', 'slate', 'vox', 'huffpost'],
    'Lean Left': ['ap', 'axios', 'cnn', 'guardian', 'business insider', 'nbcnews', 'npr', 'nytimes', 'politico', 'propublica', 'wapo', 'usa today'],
    'Center': ['reuters', 'marketwatch', 'financial times', 'newsweek', 'forbes'],
    'Lean Right': ['thedispatch', 'epochtimes', 'foxbusiness', 'wsj', 'national review', 'washtimes'],
    'Right': ['breitbart', 'theblaze', 'daily mail', 'dailywire', 'foxnews', 'nypost', 'newsmax'],
}

def setup_logging(log_file='analysis.log'):
    logging.basicConfig(
        filename=log_file,
        level=logging.DEBUG,
        format='%(asctime)s - %(levelname)s - %(message)s'
    )
    logging.info("Logging initialized.")

def load_jsonl(jsonl_file):
    records = []
    with open(jsonl_file, 'r', encoding='utf-8') as f:
        for line in tqdm(f, desc="Loading JSONL data"):
            try:
                record = json.loads(line)
                records.append(record)
            except json.JSONDecodeError as e:
                logging.error(f"JSON decoding error: {e}")
    df = pd.DataFrame(records)
    return df

def map_media_outlet_to_category(df, media_categories):
    outlet_to_category = {}
    for category, outlets in media_categories.items():
        for outlet in outlets:
            outlet_to_category[outlet.lower().strip()] = category

    if 'media_outlet' not in df.columns:
        logging.error("'media_outlet' column not found in DataFrame.")
        raise KeyError("'media_outlet' column not found in DataFrame.")

    df['media_outlet_clean'] = df['media_outlet'].str.lower().str.strip()
    df['media_category'] = df['media_outlet_clean'].map(outlet_to_category)
    df['media_category'] = df['media_category'].fillna('Other')
    unmapped_outlets = df[df['media_category'] == 'Other']['media_outlet'].unique()
    if len(unmapped_outlets) > 0:
        logging.warning(f"Unmapped media outlets found: {unmapped_outlets}")
        print(f"Warning: The following media outlets were not mapped and categorized as 'Other': {unmapped_outlets}")
    return df

def aggregate_sentiment_scores(df, sentiment_categories):
    aggregation = []
    for media_category in MEDIA_CATEGORIES.keys():
        mask = (df['media_category'] == media_category)
        subset = df[mask]

        for sentiment in sentiment_categories:
            pattern = f"^{re.escape(sentiment)}_\\d+$"
            matched_cols = [col for col in df.columns if re.match(pattern, col)]

            if matched_cols:
                quotation_sum = subset[matched_cols].clip(lower=0).sum(skipna=True).sum()
                quotation_count = subset[matched_cols].clip(lower=0).count().sum()
            else:
                quotation_sum = 0
                quotation_count = 0

            fulltext_col = f"{sentiment}_fulltext"
            if fulltext_col in df.columns:
                fulltext_sum = subset[fulltext_col].clip(lower=0).sum(skipna=True)
                fulltext_count = subset[fulltext_col].clip(lower=0).count()
            else:
                fulltext_sum = 0
                fulltext_count = 0

            aggregation.append({
                'Media Category': media_category,
                'Sentiment/Emotion': sentiment,
                'Quotation_Sum': quotation_sum,
                'Quotation_Count': quotation_count,
                'Fulltext_Sum': fulltext_sum,
                'Fulltext_Count': fulltext_count
            })

    aggregated_df = pd.DataFrame(aggregation)
    return aggregated_df

def calculate_averages(aggregated_df):
    aggregated_df['Quotation_Average'] = aggregated_df.apply(
        lambda row: row['Quotation_Sum'] / row['Quotation_Count'] if row['Quotation_Count'] > 0 else None, axis=1
    )
    aggregated_df['Fulltext_Average'] = aggregated_df.apply(
        lambda row: row['Fulltext_Sum'] / row['Fulltext_Count'] if row['Fulltext_Count'] > 0 else None, axis=1
    )
    return aggregated_df

def calculate_mean_median(aggregated_df):
    stats = []
    for sentiment in CATEGORIES:
        sentiment_data = aggregated_df[aggregated_df['Sentiment/Emotion'] == sentiment]

        quotation_avg = sentiment_data['Quotation_Average'].dropna()
        if not quotation_avg.empty:
            mean_quotation = quotation_avg.mean()
            median_quotation = quotation_avg.median()
        else:
            mean_quotation = None
            median_quotation = None

        fulltext_avg = sentiment_data['Fulltext_Average'].dropna()
        if not fulltext_avg.empty:
            mean_fulltext = fulltext_avg.mean()
            median_fulltext = fulltext_avg.median()
        else:
            mean_fulltext = None
            median_fulltext = None

        stats.append({
            'Sentiment/Emotion': sentiment,
            'Mean_Quotation_Average': mean_quotation,
            'Median_Quotation_Average': median_quotation,
            'Mean_Fulltext_Average': mean_fulltext,
            'Median_Fulltext_Average': median_fulltext
        })

    stats_df = pd.DataFrame(stats)
    return stats_df

def save_aggregated_scores_to_csv(aggregated_df, csv_output_dir, prefix='aggregated_sentiment_emotion_scores.csv'):
    csv_file = os.path.join(csv_output_dir, prefix)
    try:
        aggregated_df.to_csv(csv_file, index=False)
        print(f"Aggregated sentiment/emotion scores saved to '{csv_file}'.")
        logging.info(f"Aggregated sentiment/emotion scores saved to '{csv_file}'.")
    except Exception as e:
        print(f"Error saving aggregated scores to CSV: {e}")
        logging.error(f"Error saving aggregated scores to CSV: {e}")

def plot_statistics(aggregated_df, output_dir):
    sns.set_style("whitegrid")
    for sentiment in CATEGORIES:
        sentiment_data = aggregated_df[aggregated_df['Sentiment/Emotion'] == sentiment]

        # Quotation-Based Plot
        plt.figure(figsize=(10, 6))
        sns.barplot(
            x='Media Category',
            y='Quotation_Average',
            data=sentiment_data,
            color='steelblue'
        )
        plt.title(f"Mean Quotation-Based '{sentiment.capitalize()}' Scores", fontsize=14)
        plt.xlabel('Media Category', fontsize=12)
        plt.ylabel('Mean Quotation-Based Average Score', fontsize=12)
        plt.xticks(rotation=45, ha='right')
        plt.tight_layout()
        quote_plot_filename = f"quote_{sentiment}.png"
        quote_plot_path = os.path.join(output_dir, quote_plot_filename)
        try:
            plt.savefig(quote_plot_path)
            plt.close()
            print(f"Quotation-Based '{sentiment}' scores plot saved to '{quote_plot_path}'.")
        except Exception as e:
            print(f"Error saving quotation-based '{sentiment}' scores plot: {e}")

        # Fulltext-Based Plot
        plt.figure(figsize=(10, 6))
        sns.barplot(
            x='Media Category',
            y='Fulltext_Average',
            data=sentiment_data,
            color='darkorange'
        )
        plt.title(f"Mean Fulltext-Based '{sentiment.capitalize()}' Scores", fontsize=14)
        plt.xlabel('Media Category', fontsize=12)
        plt.ylabel('Mean Fulltext-Based Average Score', fontsize=12)
        plt.xticks(rotation=45, ha='right')
        plt.tight_layout()
        fulltext_plot_filename = f"fulltext_{sentiment}.png"
        fulltext_plot_path = os.path.join(output_dir, fulltext_plot_filename)
        try:
            plt.savefig(fulltext_plot_path)
            plt.close()
            print(f"Fulltext-Based '{sentiment}' scores plot saved to '{fulltext_plot_path}'.")
        except Exception as e:
            print(f"Error saving fulltext-based '{sentiment}' scores plot: {e}")

def debug_singular_matrix(model_df, formula, sentiment, measure_type):
    print(f"[DEBUG] Checking model design for {sentiment}-{measure_type}")
    from patsy import dmatrices
    y, X = dmatrices(formula, data=model_df, return_type='dataframe')
    y = y.values.ravel()
    X_rank = np.linalg.matrix_rank(X)
    print(f"[DEBUG] X design matrix shape: {X.shape}")
    print(f"[DEBUG] y vector shape: {y.shape}")
    print(f"[DEBUG] Rank of X: {X_rank}")

def fit_lmm_and_posthoc(df, sentiment, measure_type='Quotation'):
    logger = logging.getLogger(__name__)

    if measure_type == 'Quotation':
        pattern = f"^{re.escape(sentiment)}_\\d+$"
        matched_cols = [col for col in df.columns if re.match(pattern, col)]
        if not matched_cols:
            return None
        df = df.copy()
        df[f'{sentiment}_quotation_mean'] = df[matched_cols].clip(lower=0).mean(axis=1)
        score_col = f'{sentiment}_quotation_mean'
    else:
        fulltext_col = f"{sentiment}_fulltext"
        if fulltext_col not in df.columns:
            return None
        df = df.copy()
        df[f'{sentiment}_fulltext_clipped'] = df[fulltext_col].clip(lower=0)
        score_col = f'{sentiment}_fulltext_clipped'

    model_df = df.dropna(subset=[score_col, 'media_category', 'media_outlet']).copy()
    row_count = len(model_df)
    category_count = model_df['media_category'].nunique()

    if row_count < 2 or category_count < 2:
        return None

    formula = f"{score_col} ~ media_category"
    try:
        md = mixedlm(formula, data=model_df, groups=model_df["media_outlet"])
        mdf = md.fit(reml=True, method='lbfgs')
    except Exception as e:
        print(f"Model failed for {sentiment}-{measure_type}: {e}")
        debug_singular_matrix(model_df, formula, sentiment, measure_type)
        return None

    if not mdf.converged:
        print(f"LMM did not converge for {sentiment}-{measure_type}.")
        debug_singular_matrix(model_df, formula, sentiment, measure_type)
        return None

    summ_text = mdf.summary().as_text()
    if "Random effects covariance is singular" in summ_text:
        print("Random effects covariance singular message found.")
        debug_singular_matrix(model_df, formula, sentiment, measure_type)

    cat_means = model_df.groupby('media_category', observed=True)[score_col].mean().reset_index()
    if cat_means['media_category'].nunique() < 2:
        return {
            'LMM_Summary': mdf.summary().as_text(),
            'PostHoc': pd.DataFrame()
        }

    try:
        tukey = pairwise_tukeyhsd(cat_means[score_col], cat_means['media_category'], alpha=0.05)
        posthoc_df = pd.DataFrame(data=tukey._results_table.data[1:], columns=tukey._results_table.data[0])
    except Exception:
        posthoc_df = pd.DataFrame()

    return {
        'LMM_Summary': mdf.summary().as_text(),
        'PostHoc': posthoc_df
    }

def run_lmm_analyses(df):
    all_results = {}
    for sentiment in CATEGORIES:
        all_results[sentiment] = {}
        for measure_type in ['Quotation', 'Fulltext']:
            res = fit_lmm_and_posthoc(df, sentiment, measure_type=measure_type)
            if res is not None:
                all_results[sentiment][measure_type] = res
    return all_results

def compile_results_into_multiple_workbooks(aggregated_df, stats_df, raw_df, lmm_results, plots_dir,
                                            main_excel, raw_excel, lmm_excel, plots_excel, combined_excel):
    with pd.ExcelWriter(main_excel, engine='openpyxl') as writer:
        aggregated_df.to_excel(writer, sheet_name='Aggregated_Scores', index=False)
        stats_df.to_excel(writer, sheet_name='Mean_Median_Statistics', index=False)

    with pd.ExcelWriter(raw_excel, engine='openpyxl') as writer:
        raw_df.to_excel(writer, sheet_name='Raw_Data', index=False)
        for sentiment in CATEGORIES:
            sentiment_cols = [c for c in raw_df.columns if c.startswith(sentiment+'_')]
            sentiment_df = raw_df[['media_category', 'media_outlet'] + sentiment_cols].copy()
            sheet_name = f"Raw_{sentiment[:29]}"
            sentiment_df.to_excel(writer, sheet_name=sheet_name, index=False)

    with pd.ExcelWriter(lmm_excel, engine='openpyxl') as writer:
        summary_rows = []
        for sentiment in lmm_results:
            for measure_type in lmm_results[sentiment]:
                sheet_name = f"LMM_{sentiment[:20]}_{measure_type[:8]}"
                lmm_summary = lmm_results[sentiment][measure_type]['LMM_Summary']
                posthoc_df = lmm_results[sentiment][measure_type]['PostHoc']

                summary_df = pd.DataFrame({'LMM_Summary': lmm_summary.split('\n')})
                summary_df.to_excel(writer, sheet_name=sheet_name, index=False)
                startrow = len(summary_df) + 2
                posthoc_df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=startrow)

                summary_rows.append({
                    'Sentiment': sentiment,
                    'Measure': measure_type,
                    'SheetName': sheet_name
                })

        summary_index_df = pd.DataFrame(summary_rows)
        summary_index_df.to_excel(writer, sheet_name='LMM_Results_Index', index=False)

    wb = Workbook()
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])

    for sentiment in CATEGORIES:
        quote_plot_path = os.path.join(plots_dir, f"quote_{sentiment}.png")
        if os.path.exists(quote_plot_path):
            sheet_title = f"Quote_{sentiment[:28]}"
            worksheet = wb.create_sheet(title=sheet_title)
            try:
                img = ExcelImage(quote_plot_path)
                img.anchor = 'A1'
                worksheet.add_image(img)
            except Exception as e:
                logging.error(f"Error embedding image '{quote_plot_path}' into Excel: {e}")

        fulltext_plot_path = os.path.join(plots_dir, f"fulltext_{sentiment}.png")
        if os.path.exists(fulltext_plot_path):
            sheet_title = f"Fulltext_{sentiment[:25]}"
            worksheet = wb.create_sheet(title=sheet_title)
            try:
                img = ExcelImage(fulltext_plot_path)
                img.anchor = 'A1'
                worksheet.add_image(img)
            except Exception as e:
                logging.error(f"Error embedding image '{fulltext_plot_path}' into Excel: {e}")

    wb.save(plots_excel)

    raw_df_cleaned = raw_df.copy()
    raw_df_cleaned = raw_df_cleaned.applymap(lambda x: ", ".join(x) if isinstance(x, list) else x)

    wb_combined = Workbook()
    if 'Sheet' in wb_combined.sheetnames:
        wb_combined.remove(wb_combined['Sheet'])

    ws_agg = wb_combined.create_sheet("Aggregated_Scores")
    for r in dataframe_to_rows(aggregated_df, index=False, header=True):
        ws_agg.append(r)

    ws_stats = wb_combined.create_sheet("Mean_Median_Statistics")
    for r in dataframe_to_rows(stats_df, index=False, header=True):
        ws_stats.append(r)

    ws_raw = wb_combined.create_sheet("Raw_Data")
    for r in dataframe_to_rows(raw_df_cleaned, index=False, header=True):
        ws_raw.append(r)

    ws_lmm = wb_combined.create_sheet("LMM_Summaries")
    ws_lmm.append(["Sentiment", "Measure", "LMM_Summary"])
    for sentiment in lmm_results:
        for measure_type in lmm_results[sentiment]:
            summary = lmm_results[sentiment][measure_type]['LMM_Summary']
            ws_lmm.append([sentiment, measure_type, summary])

    wb_combined.save(combined_excel)

def main():
    setup_logging()

    # First run: no filtering at all
    print("First run: No filtering at all.")
    df = load_jsonl(INPUT_JSONL_FILE)
    print(f"Total articles loaded: {len(df)}")
    df = map_media_outlet_to_category(df, MEDIA_CATEGORIES)

    print("Aggregating sentiment/emotion scores per media category...")
    aggregated_df = aggregate_sentiment_scores(df, CATEGORIES)
    aggregated_df = calculate_averages(aggregated_df)
    stats_df = calculate_mean_median(aggregated_df)

    save_aggregated_scores_to_csv(aggregated_df, CSV_OUTPUT_DIR)
    plot_statistics(aggregated_df, OUTPUT_DIR)

    print("No predictor columns specified for collinearity check.")
    print("Fitting linear mixed models and running post-hoc tests...")
    lmm_results_1 = run_lmm_analyses(df)
    print("LMM analysis and post-hoc tests completed successfully.\n")

    compile_results_into_multiple_workbooks(aggregated_df, stats_df, df, lmm_results_1, OUTPUT_DIR,
                                            OUTPUT_EXCEL_MAIN, OUTPUT_EXCEL_RAW, OUTPUT_EXCEL_LMM, OUTPUT_EXCEL_PLOTS, OUTPUT_EXCEL_COMBINED)
    print("First run analysis completed successfully.")

    # Second run: Filter out Mother Jones, DailyWire, and National Review
    print("\nSecond run: Filter out Mother Jones, DailyWire, and National Review.\n")
    df_filtered = load_jsonl(INPUT_JSONL_FILE)
    df_filtered = map_media_outlet_to_category(df_filtered, MEDIA_CATEGORIES)

    # Filter these outlets
    outlets_to_remove = ['mother jones', 'dailywire', 'national review']
    print("Filtering out specified outlets:", outlets_to_remove)
    df_filtered = df_filtered[~df_filtered['media_outlet_clean'].isin(outlets_to_remove)]

    print("Aggregating sentiment/emotion scores per media category (filtered)...")
    aggregated_df_filtered = aggregate_sentiment_scores(df_filtered, CATEGORIES)
    aggregated_df_filtered = calculate_averages(aggregated_df_filtered)
    stats_df_filtered = calculate_mean_median(aggregated_df_filtered)

    save_aggregated_scores_to_csv(aggregated_df_filtered, CSV_OUTPUT_DIR_FILTERED)
    plot_statistics(aggregated_df_filtered, OUTPUT_DIR_FILTERED)

    print("Fitting linear mixed models and running post-hoc tests on filtered data...")
    lmm_results_2 = run_lmm_analyses(df_filtered)
    print("Filtered LMM analysis and post-hoc tests completed successfully.\n")

    compile_results_into_multiple_workbooks(aggregated_df_filtered, stats_df_filtered, df_filtered,
                                            lmm_results_2, OUTPUT_DIR_FILTERED,
                                            OUTPUT_EXCEL_MAIN_FILTERED, OUTPUT_EXCEL_RAW_FILTERED,
                                            OUTPUT_EXCEL_LMM_FILTERED, OUTPUT_EXCEL_PLOTS_FILTERED,
                                            OUTPUT_EXCEL_COMBINED_FILTERED)

    print("Filtered analysis completed successfully. Compare the filtered outputs with the original outputs to assess differences.")

if __name__ == "__main__":
    main()
