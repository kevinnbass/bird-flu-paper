#!/usr/bin/env python3
# dual_output_50tabs_and_20tabs_scatterplot_fixed.py
"""
Creates two Excel files with non-transposed tabs, with alignment across Quotation-based measures:

File A) raw_values_6cat_50tabs_no_blanks_scatterplot.xlsx
   - 50 tabs => 10 sentiments × 5 measures
   - Each tab => 6 columns (categories), rows => articles
   - Quotation-based measures keep row alignment with Fulltext/Title-based measures

File B) raw_values_6cat_20tabs_intv2_no_blanks_scatterplot.xlsx
   - 20 tabs => 2 per sentiment
   - Tab "3Int" => Fulltext_Intensity, Quotation_Intensity, Title_Intensity (18 columns)
   - Tab "2NonInt" => Fulltext, Quotation (12 columns)
   - Also keeps row alignment for Quotation <-> Fulltext (and Quotation_Intensity <-> Fulltext_Intensity, Title_Intensity).
"""

import json
import re
import numpy as np
import pandas as pd
from tqdm import tqdm
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# ------------------ GLOBAL CONFIG ------------------
INPUT_JSONL_FILE = "processed_all_articles_fixed_4.jsonl"

OUTPUT_50TABS_SCATTER = "raw_values_6cat_50tabs_no_blanks_scatterplot.xlsx"
OUTPUT_20TABS_SCATTER = "raw_values_6cat_20tabs_intv2_no_blanks_scatterplot.xlsx"

# 6 categories
MEDIA_CATEGORIES = {
    "Scientific": ["nature","sciam","stat","newscientist"],
    "Left": ["theatlantic","the daily beast","the intercept","mother jones",
             "msnbc","slate","vox","huffpost"],
    "Lean Left": ["ap","axios","cnn","guardian","business insider","nbcnews",
                  "npr","nytimes","politico","propublica","wapo","usa today"],
    "Center": ["reuters","marketwatch","financial times","newsweek","forbes"],
    "Lean Right": ["thedispatch","epochtimes","foxbusiness","wsj","national review","washtimes"],
    "Right": ["breitbart","theblaze","daily mail","dailywire","foxnews","nypost","newsmax"],
}

SENTIMENTS = [
    "joy","sadness","anger","fear",
    "surprise","disgust","trust","anticipation",
    "negative_sentiment","positive_sentiment"
]

# 5 measures for the 50-tab approach
MEASURES_5 = ["Fulltext","Quotation","Fulltext_Intensity","Quotation_Intensity","Title_Intensity"]

# For the 20-tab approach => 2 tabs/sentiment
THREE_INTENSITIES = ["Fulltext_Intensity","Quotation_Intensity","Title_Intensity"]
TWO_NONINTS       = ["Fulltext","Quotation"]

# ------------------ 1) LOAD + MAP ------------------
def load_jsonl(jsonl_file):
    records=[]
    with open(jsonl_file,"r",encoding="utf-8") as f:
        for line in tqdm(f,desc="Loading JSONL"):
            rec=json.loads(line)
            records.append(rec)
    return pd.DataFrame(records)

def map_outlet_to_6cats(df):
    cat_map={}
    for cat, outls in MEDIA_CATEGORIES.items():
        for o in outls:
            cat_map[o.lower().strip()] = cat
    df2=df.copy()
    df2["media_outlet_clean"] = df2["media_outlet"].str.lower().str.strip()
    df2["media_category_6"]   = df2["media_outlet_clean"].map(cat_map).fillna("Other")
    return df2

# ------------------ 2) EXTRACT MEASURE VALUE ------------------
def extract_measure_value(row, sentiment, measure):
    """
    - Quotation => average columns ^{sentiment}_\\d+$
    - Quotation_Intensity => average columns ^{sentiment}_\\d+_intensity$
    - Fulltext => row[f"{sentiment}_fulltext"]
    - Fulltext_Intensity => row[f"{sentiment}_fulltext_intensity"]
    - Title_Intensity => row[f"{sentiment}_title_intensity"]
    Return float or NaN
    """
    try:
        if measure=="Quotation":
            pat = rf"^{re.escape(sentiment)}_\d+$"
            matched = [c for c in row.index if re.match(pat,c)]
            if not matched:
                return np.nan
            vals = row[matched].astype(float).clip(lower=0)
            return vals.mean(skipna=True)

        elif measure=="Quotation_Intensity":
            pat = rf"^{re.escape(sentiment)}_\d+_intensity$"
            matched = [c for c in row.index if re.match(pat,c)]
            if not matched:
                return np.nan
            vals = row[matched].astype(float).clip(lower=0)
            return vals.mean(skipna=True)

        elif measure=="Fulltext":
            col = f"{sentiment}_fulltext"
            if col not in row.index:
                return np.nan
            return float(row[col])

        elif measure=="Fulltext_Intensity":
            col = f"{sentiment}_fulltext_intensity"
            if col not in row.index:
                return np.nan
            return float(row[col])

        elif measure=="Title_Intensity":
            col = f"{sentiment}_title_intensity"
            if col not in row.index:
                return np.nan
            return float(row[col])

        else:
            return np.nan

    except:
        return np.nan

# ------------------ HELPER: remove default sheet if present ------------------
def remove_default_sheet_if_present(wb):
    """Remove the default 'Sheet' if it exists in this workbook."""
    if "Sheet" in wb.sheetnames:
        default_ws = wb["Sheet"]   # the Worksheet object
        wb.remove(default_ws)

# ---------------------------------------------------------------------------
# 3) "BLANK ALIGNMENT" LOGIC: unify rows across Quotation-based measures
#
# We'll define unify_rows_across_measures() => merges article sets so that each article
# is present in all measures in that group. If it's missing in measure X, that cell => NaN.
#
def gather_article_vals_for_measure(df, sentiment, measure):
    """
    Return dictionary: {cat_6: list of (article_id, measure_val)}, 
    storing articles even if measure_val=NaN
    """
    cat_dict = {}
    cat_names = ["Scientific","Left","Lean Left","Center","Lean Right","Right"]
    for c in cat_names:
        cat_dict[c] = []
    for i, row in df.iterrows():
        cat_6 = row["media_category_6"]
        if cat_6 not in cat_dict:
            continue
        art_id = row["article_id"]
        val = extract_measure_value(row, sentiment, measure)
        cat_dict[cat_6].append((art_id, val))
    return cat_dict

def unify_rows_across_measures(df, sentiment, measure_group):
    """
    measure_group: list like ["Quotation","Fulltext","Title_Intensity"]
    => unify so that each article that appears in any measure => 1 row in all
    => cat_out[cat] => list of row_dict => {article_id, measure1, measure2, ...}
    """
    cat_names = ["Scientific","Left","Lean Left","Center","Lean Right","Right"]
    # measure_data[meas][cat][art_id] = val
    measure_data = {}
    for meas in measure_group:
        measure_data[meas] = {}
        for cat in cat_names:
            measure_data[meas][cat] = {}

    # fill them
    for meas in measure_group:
        cat_dict = gather_article_vals_for_measure(df, sentiment, meas)
        for cat in cat_names:
            for (a_id, val) in cat_dict[cat]:
                measure_data[meas][cat][a_id] = val

    cat_out = {}
    for cat in cat_names:
        # gather all article_ids that appear in any measure
        id_set = set()
        for meas in measure_group:
            id_set.update( measure_data[meas][cat].keys() )
        sorted_ids = sorted(list(id_set))

        cat_out[cat] = []
        for a_id in sorted_ids:
            row_dict = {"article_id": a_id}
            for meas in measure_group:
                row_dict[meas] = measure_data[meas][cat].get(a_id, np.nan)
            cat_out[cat].append(row_dict)
    return cat_out

# ---------------------------------------------------------------------------
# 4) CREATE THE 50-TAB FILE => uses 2 alignment groups: 
#    (Quotation,Fulltext,Title_Intensity) + (Quotation_Intensity,Fulltext_Intensity,Title_Intensity)
def create_50tabs_scatter(df, wb):
    remove_default_sheet_if_present(wb)

    cat_names = ["Scientific","Left","Lean Left","Center","Lean Right","Right"]
    # define alignment groups
    measure_alignment_groups = [
        ["Quotation","Fulltext","Title_Intensity"],
        ["Quotation_Intensity","Fulltext_Intensity","Title_Intensity"]
    ]

    for sentiment in SENTIMENTS:
        for group in measure_alignment_groups:
            cat_out = unify_rows_across_measures(df, sentiment, group)
            for meas in group:
                # only produce a tab if meas is in MEASURES_5
                if meas not in MEASURES_5:
                    continue
                tab_name = f"{sentiment[:10]}_{meas[:10]}"
                # check if tab_name already in wb => skip if so
                if tab_name in wb.sheetnames:
                    continue

                # build a DataFrame => rows=the max # of entries among categories
                # but we keep them separate => 6 columns
                # cat_out[cat] => list of row_dict => each row_dict => measure keys
                max_len = max(len(cat_out[c]) for c in cat_names)
                data_dict = {}
                for c in cat_names:
                    col_vals=[]
                    row_list = cat_out[c]
                    for row_idx in range(max_len):
                        if row_idx < len(row_list):
                            val = row_list[row_idx].get(meas, np.nan)
                            col_vals.append(val)
                        else:
                            col_vals.append(np.nan)
                    data_dict[c] = col_vals
                df_tab = pd.DataFrame(data_dict, columns=cat_names)

                ws = wb.create_sheet(title=tab_name)
                # first row => cat_names
                row_data = [cat_names]
                for r_i in range(len(df_tab)):
                    row_list2=[]
                    for c in cat_names:
                        row_list2.append(df_tab.loc[r_i,c])
                    row_data.append(row_list2)
                for r_idx, row_vals in enumerate(row_data,1):
                    for c_idx, cell_val in enumerate(row_vals,1):
                        ws.cell(row=r_idx,column=c_idx,value=cell_val)


# ---------------------------------------------------------------------------
# 5) CREATE THE 20-TAB FILE => each sentiment => tab "3Int" (Fulltext_Intensity, Quotation_Intensity, Title_Intensity),
#                              tab "2NonInt" (Fulltext, Quotation).
#    => non-transposed => columns => cat×measure => rows => articles
def create_20tabs_scatter(df, wb_20):
    remove_default_sheet_if_present(wb_20)

    cat_names = ["Scientific","Left","Lean Left","Center","Lean Right","Right"]

    for sent in SENTIMENTS:
        # A) "3Int" => unify rows for Fulltext_Intensity, Quotation_Intensity, Title_Intensity
        measure_grpA = ["Fulltext_Intensity","Quotation_Intensity","Title_Intensity"]
        cat_outA = unify_rows_across_measures(df, sent, measure_grpA)

        out_colsA = []
        for cat in cat_names:
            for m in measure_grpA:
                out_colsA.append(f"{cat}_{m}")
        max_lenA = max(len(cat_outA[c]) for c in cat_names)
        data_A = {}
        for coln in out_colsA:
            data_A[coln] = []
        # fill
        for cat in cat_names:
            for m in measure_grpA:
                col_key = f"{cat}_{m}"
                col_vals=[]
                row_list = cat_outA[cat]
                for row_i in range(max_lenA):
                    if row_i < len(row_list):
                        val = row_list[row_i].get(m, np.nan)
                        col_vals.append(val)
                    else:
                        col_vals.append(np.nan)
                data_A[col_key] = col_vals
        dfA = pd.DataFrame(data_A, columns=out_colsA)
        tabA_name = f"{sent[:10]}_3Int"
        wsA = wb_20.create_sheet(title=tabA_name)
        row_dataA = [out_colsA]
        for r_i in range(len(dfA)):
            row_vals=[]
            for c in out_colsA:
                row_vals.append(dfA.loc[r_i,c])
            row_dataA.append(row_vals)
        for r_idx, rowvals in enumerate(row_dataA,1):
            for c_idx, cell_val in enumerate(rowvals,1):
                wsA.cell(row=r_idx,column=c_idx,value=cell_val)

        # B) "2NonInt" => unify rows for Fulltext, Quotation
        measure_grpB = ["Fulltext","Quotation"]
        cat_outB = unify_rows_across_measures(df, sent, measure_grpB)

        out_colsB=[]
        for cat in cat_names:
            for m in measure_grpB:
                out_colsB.append(f"{cat}_{m}")
        max_lenB = max(len(cat_outB[c]) for c in cat_names)
        data_B = {}
        for coln in out_colsB:
            data_B[coln]=[]
        for cat in cat_names:
            for m in measure_grpB:
                col_key = f"{cat}_{m}"
                col_vals=[]
                row_list = cat_outB[cat]
                for row_i in range(max_lenB):
                    if row_i<len(row_list):
                        val = row_list[row_i].get(m,np.nan)
                        col_vals.append(val)
                    else:
                        col_vals.append(np.nan)
                data_B[col_key] = col_vals
        dfB = pd.DataFrame(data_B, columns=out_colsB)
        tabB_name = f"{sent[:10]}_2NonInt"
        wsB = wb_20.create_sheet(title=tabB_name)
        row_dataB = [out_colsB]
        for r_i in range(len(dfB)):
            rowvals=[]
            for c in out_colsB:
                rowvals.append(dfB.loc[r_i,c])
            row_dataB.append(rowvals)
        for r_idx, rowvals in enumerate(row_dataB,1):
            for c_idx, cell_val in enumerate(rowvals,1):
                wsB.cell(row=r_idx,column=c_idx,value=cell_val)


# ---------------------------------------------------------------------------
# 6) MAIN
# ---------------------------------------------------------------------------
def main():
    print("Loading + mapping data...")
    df = load_jsonl(INPUT_JSONL_FILE)
    df = map_outlet_to_6cats(df)
    if "article_id" not in df.columns:
        df["article_id"] = df.index + 1

    print(f"Creating 50-tab scatter file => {OUTPUT_50TABS_SCATTER}")
    wb50 = Workbook()
    create_50tabs_scatter(df, wb50)
    wb50.save(OUTPUT_50TABS_SCATTER)
    print(f"Wrote => {OUTPUT_50TABS_SCATTER}")

    print(f"Creating 20-tab scatter file => {OUTPUT_20TABS_SCATTER}")
    wb20 = Workbook()
    create_20tabs_scatter(df, wb20)
    wb20.save(OUTPUT_20TABS_SCATTER)
    print(f"Wrote => {OUTPUT_20TABS_SCATTER}")

    print("All done.")

if __name__=="__main__":
    main()
