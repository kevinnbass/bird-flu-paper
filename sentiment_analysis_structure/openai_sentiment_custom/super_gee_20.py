#!/usr/bin/env python3
# super_gee_all_validations_dual_cats.py
"""
A complete Python script that:
  1) Loads data from JSONL
  2) Normalizes high_rate_2 => 'yes'/'no'
  3) Runs the entire pipeline twice:
       a) Original categories => "Yes"/"All"
       b) Collapsed categories => "Yes_collapsed"/"All_collapsed"
  => 4 total runs (plus sensitivity merges).

All references are defined, including cross_validation_gee and try_all_families_and_scales.
Now with maxiter=300 added to every .fit(...) call.
"""

import json
import os
import re
import sys
import warnings
import logging
import math
import gc
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import seaborn as sns
from tqdm import tqdm

# Statsmodels & related
import statsmodels.api as sm
from statsmodels.genmod.families import (
    Poisson, NegativeBinomial,
    Gaussian, Gamma, InverseGaussian
)
import statsmodels.genmod.families.links as ln
from statsmodels.genmod.generalized_estimating_equations import GEE
from statsmodels.genmod.cov_struct import Independence, Exchangeable
from statsmodels.stats.multitest import multipletests
from scipy.stats import pearsonr, norm

from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill

###############################################################################
# CONFIG
###############################################################################
INPUT_JSONL_FILE = "processed_all_articles_fixed_4.jsonl"

BASE_OUTPUT_DIR = "graphs_analysis"
BASE_CSV_OUTPUT_DIR = "csv_raw_scores"

OUTPUT_EXCEL_MAIN = "analysis_main.xlsx"
OUTPUT_EXCEL_RAW = "analysis_raw.xlsx"
OUTPUT_EXCEL_GEE = "analysis_gee.xlsx"
OUTPUT_EXCEL_PLOTS = "analysis_plots.xlsx"
OUTPUT_EXCEL_COMBINED = "analysis_combined.xlsx"

LOG_FILE = "analysis.log"

os.makedirs(BASE_OUTPUT_DIR, exist_ok=True)
os.makedirs(BASE_CSV_OUTPUT_DIR, exist_ok=True)

CATEGORIES = [
    "joy","sadness","anger","fear",
    "surprise","disgust","trust","anticipation",
    "negative_sentiment","positive_sentiment"
]

# Original 6-cat scheme
MEDIA_CATEGORIES_ORIG = {
    "Scientific": ["nature","sciam","stat","newscientist"],
    "Left": ["theatlantic","the daily beast","the intercept","mother jones","msnbc","slate","vox","huffpost"],
    "Lean Left": ["ap","axios","cnn","guardian","business insider","nbcnews","npr","nytimes","politico","propublica","wapo","usa today"],
    "Center": ["reuters","marketwatch","financial times","newsweek","forbes"],
    "Lean Right": ["thedispatch","epochtimes","foxbusiness","wsj","national review","washtimes"],
    "Right": ["breitbart","theblaze","daily mail","dailywire","foxnews","nypost","newsmax"],
}

# Collapsed 4-cat scheme
MEDIA_CATEGORIES_COLLAPSED = {
    "Scientific": ["nature","sciam","stat","newscientist"],
    "Left": [
        # old "Left"
        "theatlantic","the daily beast","the intercept","mother jones",
        "msnbc","slate","vox","huffpost",
        # old "Lean Left"
        "ap","axios","cnn","guardian","business insider","nbcnews",
        "npr","nytimes","politico","propublica","wapo","usa today"
    ],
    "Center": [
        "reuters","marketwatch","financial times","newsweek","forbes"
    ],
    "Right": [
        # old "Lean Right"
        "thedispatch","epochtimes","foxbusiness","wsj","national review","washtimes",
        # old "Right"
        "breitbart","theblaze","daily mail","dailywire","foxnews","nypost","newsmax"
    ],
}

###############################################################################
# GLOBAL: For the "Compare_Imbalance" table
###############################################################################
COMPARE_IMBALANCE = []

###############################################################################
# LOGGING
###############################################################################
def setup_logging():
    log_format = "%(asctime)s [%(levelname)s] %(module)s - %(message)s"
    logging.basicConfig(filename=LOG_FILE, level=logging.DEBUG, format=log_format)
    console = logging.StreamHandler(sys.stdout)
    console.setLevel(logging.INFO)
    console.setFormatter(logging.Formatter(log_format))
    logging.getLogger().addHandler(console)
    logging.info("Logging initialized.")

warnings.filterwarnings(
    "ignore",
    message="QIC values obtained using scale=None are not appropriate for comparing models"
)

###############################################################################
# 1) Load + map
###############################################################################
def load_jsonl(jsonl_file):
    logging.info(f"Loading JSONL => {jsonl_file}")
    records=[]
    with open(jsonl_file,"r",encoding="utf-8") as f:
        for line in tqdm(f,desc="Loading JSONL"):
            rec=json.loads(line)
            records.append(rec)
    df=pd.DataFrame(records)
    return df

def map_media_outlet_to_category(df, cat_dict):
    cat_map={}
    for cat, outls in cat_dict.items():
        for o in outls:
            cat_map[o.lower().strip()] = cat

    df2 = df.copy()
    df2["media_outlet_clean"] = df2["media_outlet"].str.lower().str.strip()
    df2["media_category"] = df2["media_outlet_clean"].map(cat_map).fillna("Other")
    return df2

###############################################################################
# 2) chunk + stats
###############################################################################
def chunk_and_save(df, chunk_size=20000, prefix=""):
    logging.info(f"Chunking => len={len(df)}, chunk_size={chunk_size}, prefix={prefix}")
    for i in range(0,len(df),chunk_size):
        part=df.iloc[i:i+chunk_size]
        suffix=f"_{prefix}" if prefix else ""
        out_csv=os.path.join(BASE_CSV_OUTPUT_DIR,f"raw_data_part_{(i//chunk_size)+1}{suffix}.csv")
        part.to_csv(out_csv,index=False)
        print(f"Saved chunk {(i//chunk_size)+1} => {out_csv}")

def print_basic_stats(df, prefix=""):
    logging.info(f"Basic stats => total articles={len(df)}, prefix={prefix}")
    print(f"\nSummary Stats (prefix={prefix or '(None)'}) =>")
    print("Total articles:", len(df))
    if "media_outlet_clean" in df.columns:
        vc=df["media_outlet_clean"].value_counts()
        print("\nArticles per outlet (clusters):")
        print(vc)
    if "media_category" in df.columns:
        vc2=df["media_category"].value_counts()
        print("\nArticles per category:")
        print(vc2)
    print()

###############################################################################
# 3) Correlation analyses => Quotation vs Fulltext, etc.
###############################################################################
def analyze_2fields_correlation(df, left_field_pattern, right_field,
                                correlation_title, output_excel_base,
                                prefix=""):
    from scipy.stats import pearsonr
    suffix=f"_{prefix}" if prefix else ""

    records=[]
    for cat in df["media_category"].dropna().unique():
        dcat=df[df["media_category"]==cat]
        for s in CATEGORIES:
            pat=re.compile(left_field_pattern.replace("<sent>", re.escape(s)))
            matched=[c for c in dcat.columns if pat.match(c)]
            if matched:
                clp=dcat[matched].clip(lower=0)
                sum_v=clp.sum(skipna=True).sum()
                ccount=clp.count().sum()
                left_avg=sum_v/ccount if ccount>0 else np.nan
            else:
                left_avg=np.nan

            rfield=right_field.replace("<sent>", s)
            if rfield in dcat.columns:
                rv=dcat[rfield].clip(lower=0)
                r_sum=rv.sum(skipna=True)
                r_cnt=rv.count()
                right_avg=r_sum/r_cnt if r_cnt>0 else np.nan
            else:
                right_avg=np.nan

            records.append({
                "MediaCategory": cat,
                "Sentiment": s,
                "Left_Average": left_avg,
                "Right_Average": right_avg
            })

    agg_df=pd.DataFrame(records)
    cor_results=[]
    all_combo=[]
    for s in CATEGORIES:
        sub=agg_df[(agg_df["Sentiment"]==s)].dropna(subset=["Left_Average","Right_Average"])
        if len(sub)>1:
            cval,_=pearsonr(sub["Left_Average"], sub["Right_Average"])
        else:
            cval=np.nan
        cor_results.append({"Sentiment":s,"Correlation":cval})
        if not sub.empty:
            all_combo.append(sub.copy())

    cor_df=pd.DataFrame(cor_results)
    plt.figure(figsize=(8,5))
    sns.barplot(x="Sentiment", y="Correlation", data=cor_df, color="gray")
    plt.title(f"{correlation_title} - prefix={prefix}")
    plt.xticks(rotation=45,ha="right")
    plt.ylim(-1,1)
    plt.tight_layout()

    barname=f"correlation_{prefix}_{correlation_title.replace(' ','_')}_bar.png"
    barpath=os.path.join(BASE_OUTPUT_DIR, barname.lower())
    try:
        plt.savefig(barpath)
    except:
        pass
    plt.close()

    if all_combo:
        allc=pd.concat(all_combo, ignore_index=True)
        allc["Lmean"]=allc.groupby("Sentiment")["Left_Average"].transform("mean")
        allc["Lstd"]=allc.groupby("Sentiment")["Left_Average"].transform("std")
        allc["Rmean"]=allc.groupby("Sentiment")["Right_Average"].transform("mean")
        allc["Rstd"]=allc.groupby("Sentiment")["Right_Average"].transform("std")
        allc["Left_Z"]=(allc["Left_Average"]-allc["Lmean"])/allc["Lstd"]
        allc["Right_Z"]=(allc["Right_Average"]-allc["Rmean"])/allc["Rstd"]

        plt.figure(figsize=(6,5))
        sns.scatterplot(
            x="Left_Average",
            y="Right_Average",
            hue="MediaCategory",
            data=allc, s=50
        )
        plt.title(f"{correlation_title} scatter - prefix={prefix}")
        scattername=f"scatter_{prefix}_{correlation_title.replace(' ','_')}.png"
        scatterpath=os.path.join(BASE_OUTPUT_DIR, scattername.lower())
        plt.tight_layout()
        try:
            plt.savefig(scatterpath)
        except:
            pass
        plt.close()

        out_xlsx=output_excel_base.replace(".xlsx", f"{suffix}.xlsx")
        with pd.ExcelWriter(out_xlsx, engine="openpyxl") as writer:
            cor_df.to_excel(writer, sheet_name="CorrelationSummary", index=False)
            allc.to_excel(writer, sheet_name="CombinedZScatterData", index=False)

    out_csv=os.path.join(BASE_CSV_OUTPUT_DIR, f"correlation_{prefix}_{correlation_title.replace(' ','_')}.csv")
    cor_df.to_csv(out_csv, index=False)

def analyze_all_custom_correlations(df, prefix=""):
    QFTC = "quotation_fulltext_correlation.xlsx"
    QIvFI= "quotation_intensity_fulltext_intensity_correlation.xlsx"
    FIvF = "fulltext_intensity_vs_fulltext_correlation.xlsx"

    analyze_2fields_correlation(
        df,
        "^<sent>_\\d+$",
        "<sent>_fulltext",
        "Quotation_vs_Fulltext",
        QFTC,
        prefix
    )
    analyze_2fields_correlation(
        df,
        "^<sent>_\\d+_intensity$",
        "<sent>_fulltext_intensity",
        "Quotation_Intensity_vs_Fulltext_Intensity",
        QIvFI,
        prefix
    )
    analyze_2fields_correlation(
        df,
        "<sent>_fulltext_intensity",
        "<sent>_fulltext",
        "Fulltext_Intensity_vs_Fulltext",
        FIvF,
        prefix
    )

###############################################################################
# 4) Aggregation & Stats
###############################################################################
def aggregate_sentiment_scores(df, sentiments):
    recs=[]
    cats_in_df = df["media_category"].unique()
    for cat in cats_in_df:
        sub=df[df["media_category"]==cat]
        for s in sentiments:
            pat=rf"^{re.escape(s)}_\d+$"
            matched=[c for c in sub.columns if re.match(pat,c)]
            if matched:
                qsum=sub[matched].clip(lower=0).sum(skipna=True).sum()
                qcount=sub[matched].clip(lower=0).count().sum()
            else:
                qsum,qcount=(0,0)
            fcol=f"{s}_fulltext"
            if fcol in sub.columns:
                fv=sub[fcol].clip(lower=0)
                f_sum=fv.sum(skipna=True)
                f_cnt=fv.count()
            else:
                f_sum,f_cnt=(0,0)
            recs.append({
                "Media Category": cat,
                "Sentiment/Emotion": s,
                "Quotation_Sum": qsum,
                "Quotation_Count": qcount,
                "Fulltext_Sum": f_sum,
                "Fulltext_Count": f_cnt
            })
    return pd.DataFrame(recs)

def calculate_averages(agg_df):
    def sdiv(a,b): 
        return a/b if b>0 else None
    agg_df["Quotation_Average"]=agg_df.apply(
        lambda r: sdiv(r["Quotation_Sum"],r["Quotation_Count"]), axis=1
    )
    agg_df["Fulltext_Average"]=agg_df.apply(
        lambda r: sdiv(r["Fulltext_Sum"],r["Fulltext_Count"]), axis=1
    )
    return agg_df

def calculate_mean_median(agg_df):
    rows=[]
    for s in CATEGORIES:
        sub=agg_df[agg_df["Sentiment/Emotion"]==s]
        qa=sub["Quotation_Average"].dropna()
        fa=sub["Fulltext_Average"].dropna()
        rows.append({
            "Sentiment/Emotion": s,
            "Mean_Quotation_Average": qa.mean() if len(qa)>0 else None,
            "Median_Quotation_Average": qa.median() if len(qa)>0 else None,
            "Mean_Fulltext_Average": fa.mean() if len(fa)>0 else None,
            "Median_Fulltext_Average": fa.median() if len(fa)>0 else None
        })
    return pd.DataFrame(rows)

def save_aggregated_scores_to_csv(agg_df, out_dir, prefix=""):
    suffix=f"_{prefix}" if prefix else ""
    fn=os.path.join(out_dir, f"aggregated_sentiment_emotion_scores{suffix}.csv")
    agg_df.to_csv(fn,index=False)
    print(f"Aggregated => {fn}")

def plot_statistics(agg_df, out_dir, prefix=""):
    suffix=f"_{prefix}" if prefix else ""
    sns.set_style("whitegrid")
    catvals = agg_df["Media Category"].unique()
    for s in CATEGORIES:
        sub=agg_df[agg_df["Sentiment/Emotion"]==s]
        plt.figure(figsize=(10,6))
        sns.barplot(x="Media Category", y="Quotation_Average", data=sub,
                    color="steelblue", order=catvals)
        plt.title(f"Mean Quotation '{s.capitalize()}' Scores {prefix}")
        plt.xticks(rotation=45,ha="right")
        plt.tight_layout()
        out1=os.path.join(out_dir,f"quote_{s}{suffix}.png")
        try:
            plt.savefig(out1)
        except:
            pass
        plt.close()

        plt.figure(figsize=(10,6))
        sns.barplot(x="Media Category", y="Fulltext_Average", data=sub,
                    color="darkorange", order=catvals)
        plt.title(f"Mean Fulltext '{s.capitalize()}' Scores {prefix}")
        plt.xticks(rotation=45,ha="right")
        plt.tight_layout()
        out2=os.path.join(out_dir,f"fulltext_{s}{suffix}.png")
        try:
            plt.savefig(out2)
        except:
            pass
        plt.close()

###############################################################################
# Mancl & DeRouen Correction in Bootstrap
###############################################################################
def mancl_derouen_correction(gee_result):
    clusters = np.unique(gee_result.model.groups)
    M = len(clusters)
    if M < 2:
        logging.warning("Cannot apply M&D correction with <2 clusters.")
        return gee_result.cov_params()
    V_robust = gee_result.cov_params()
    factor = M/(M-1.0)
    return factor * V_robust

def bootstrap_gee_md(df, formula, group_col,
                     B=200, family=Poisson(), cov_struct=Independence()):
    cluster_ids = df[group_col].unique()
    M = len(cluster_ids)
    param_records = []
    md_se_records = []

    for _ in range(B):
        sample_ids = np.random.choice(cluster_ids, size=M, replace=True)
        pieces=[]
        for cid in sample_ids:
            pieces.append(df[df[group_col]==cid])
        boot_df=pd.concat(pieces, ignore_index=True)

        mod=GEE.from_formula(formula, groups=group_col, data=boot_df,
                             family=family, cov_struct=cov_struct)
        # add maxiter=300
        res=mod.fit(maxiter=300)
        base_cov = res.cov_params()
        factor = M/(M-1) if M>1 else 1.0
        cov_mand = factor * base_cov
        se_mand = np.sqrt(np.diag(cov_mand))

        param_records.append(res.params)
        md_se_records.append(se_mand)

    df_params = pd.DataFrame(param_records)
    df_md_se = pd.DataFrame(md_se_records, columns=df_params.columns)
    return df_params, df_md_se

###############################################################################
# coverage checks & cluster checks
###############################################################################
def coverage_simulation(n_sims=1000, n_clusters=10, cluster_size=10,
                       true_beta=0.5, alpha=0.05,
                       family=Poisson(), cov_struct=Independence(),
                       random_state=42):
    np.random.seed(random_state)
    coverage_count = 0
    for _ in range(n_sims):
        cluster_ids = np.repeat(range(n_clusters), cluster_size)
        X = np.random.binomial(1, 0.5, size=n_clusters*cluster_size)
        lin_pred = 0 + true_beta * X
        mu = np.exp(lin_pred)
        Y = np.random.poisson(mu)
        sim_df = pd.DataFrame({"Y": Y,"X": X,"cluster": cluster_ids})
        model=GEE.from_formula("Y ~ X", groups="cluster", data=sim_df,
                               family=family, cov_struct=cov_struct)
        # add maxiter=300
        res=model.fit(maxiter=300)
        est=res.params["X"]
        se=res.bse["X"]
        z_crit=norm.ppf(1 - alpha/2)
        ci_lower=est - z_crit*se
        ci_upper=est + z_crit*se
        if (true_beta >= ci_lower) and (true_beta <= ci_upper):
            coverage_count+=1
    coverage_rate=coverage_count/n_sims
    if abs(coverage_rate-0.95)<=0.05:
        concl="No major issues"
    else:
        concl="Potential over/under correction"
    return {
        "Check":"CoverageSimulation",
        "N_Sims":n_sims,
        "CoverageRate":round(coverage_rate,3),
        "Conclusion":concl
    }

def check_number_of_clusters(df, cluster_col, threshold=20):
    n_clusters=df[cluster_col].nunique()
    if n_clusters<threshold:
        msg=f"WARNING: #Clusters={n_clusters} < {threshold}. GEE robust SE may be biased."
        logging.warning(msg)
        conclusion="Potential small-sample bias"
    else:
        msg=f"#Clusters={n_clusters} >= {threshold} => likely OK for GEE asymptotics."
        conclusion="No major issues"
    print(msg)
    return {
        "Check":"NumberOfClusters",
        "Threshold":threshold,
        "Value":n_clusters,
        "Conclusion":conclusion
    }

def check_cluster_balance(df, cluster_col, imbalance_ratio=5.0):
    sizes=df[cluster_col].value_counts()
    if len(sizes)==0:
        return {
            "Check":"ClusterBalance",
            "Value":"No clusters found",
            "Conclusion":"No data"
        }
    min_size=sizes.min()
    max_size=sizes.max()
    if min_size==0:
        return {
            "Check":"ClusterBalance",
            "Value":"Smallest cluster=0",
            "Conclusion":"Potential degenerate cluster"
        }
    ratio=max_size/min_size
    if ratio>imbalance_ratio:
        msg=f"WARNING: Cluster size ratio={ratio:.1f} > {imbalance_ratio}"
        logging.warning(msg)
        conclusion="High imbalance => potential bias in SE"
    else:
        conclusion="No major imbalance"
    return {
        "Check":"ClusterBalance",
        "Value":f"{ratio:.2f}",
        "Conclusion":conclusion,
        "ImbalanceRatio":imbalance_ratio
    }

###############################################################################
# scale computations
###############################################################################
def compute_pearson_scale(y, mu, df_resid):
    if df_resid<=0:
        return np.nan
    y_arr=np.asarray(y)
    mu_arr=np.asarray(mu)
    r=(y_arr - mu_arr)/np.sqrt(mu_arr+1e-9)
    return np.sum(r**2)/df_resid

def compute_deviance_scale(y, mu, df_resid):
    if df_resid<=0:
        return np.nan
    y_arr=np.asarray(y)
    mu_arr=np.asarray(mu)
    arr=[]
    for obs,lam in zip(y_arr,mu_arr):
        if obs>0 and lam>0:
            arr.append(obs*math.log(obs/lam)-(obs-lam))
        elif obs==0:
            arr.append(-(obs-lam))
        else:
            arr.append(np.nan)
    dev=2*np.nansum(arr)
    return dev/df_resid if df_resid>0 else np.nan

def compute_ub_scale(y, mu, df_resid):
    p=compute_pearson_scale(y, mu, df_resid)
    if np.isnan(p):
        return p
    return 1.1*p

def compute_bc_scale(y, mu, df_resid):
    d=compute_deviance_scale(y, mu, df_resid)
    if np.isnan(d):
        return d
    return 0.9*d

###############################################################################
# check_residuals_and_correlation, pseudo_likelihood
###############################################################################
def check_residuals_and_correlation(final_fit):
    y_arr=np.asarray(final_fit.model.endog)
    mu_arr=np.asarray(final_fit.fittedvalues)
    pearson_res_arr=(y_arr - mu_arr)/np.sqrt(mu_arr+1e-9)
    pearson_res_arr = np.asarray(pearson_res_arr)

    mean_res=np.mean(pearson_res_arr)
    std_res=np.std(pearson_res_arr)

    clusters=final_fit.model.groups
    cluster_map={}
    for i,cid in enumerate(clusters):
        cluster_map.setdefault(cid,[]).append(pearson_res_arr[i])

    wcorr=[]
    for cid, arr in cluster_map.items():
        arr=np.array(arr)
        if len(arr)>1:
            cmat=np.corrcoef(arr)
            if cmat.ndim==2 and cmat.shape[0]>1:
                n2=len(arr)
                sum_offdiag=np.sum(cmat)-n2
                avg_off=sum_offdiag/(n2*(n2-1)) if n2>1 else 0
                wcorr.append(avg_off)

    avg_corr=np.mean(wcorr) if len(wcorr)>0 else 0.0

    deviance=getattr(final_fit,"pearson_chi2",np.nan)
    dfres=getattr(final_fit,"df_resid",1)
    overdisp=np.nan
    if dfres>0 and not np.isnan(deviance):
        overdisp=deviance/dfres

    assess=""
    if abs(mean_res)>1:
        assess+="Mean Pearson residual>1 => possible misfit. "
    if std_res>2:
        assess+="Residual std>2 => possible outliers. "
    if abs(avg_corr)>0.3:
        assess+=f"Within-cluster corr={avg_corr:.2f} => structure suspect. "
    if (not np.isnan(overdisp)) and overdisp>2:
        assess+=f"Overdisp={overdisp:.2f} => consider NB or other approach. "

    if assess=="":
        assess="No major issues from these checks."

    return {
        "mean_pearson_res": mean_res,
        "std_pearson_res": std_res,
        "avg_within_corr": avg_corr,
        "overdisp_ratio": overdisp,
        "assessment": assess
    }

def pseudo_likelihood_check(final_fit):
    from statsmodels.genmod.families.family import NegativeBinomial
    fam=final_fit.model.family
    if not isinstance(fam, sm.genmod.families.family.Poisson):
        return {
            "NB_QIC":None,
            "Poisson_QIC":None,
            "diff_QIC":None,
            "conclusion":"Non-Poisson => skip NB check"
        }
    data=final_fit.model.data.frame
    groups=final_fit.model.groups
    formula=final_fit.model.formula
    cov_struct=final_fit.model.cov_struct
    nb_model=GEE.from_formula(formula, groups=groups, data=data,
                              family=NegativeBinomial(alpha=1.0),
                              cov_struct=cov_struct)
    # add maxiter=300
    nb_res=nb_model.fit(maxiter=300)
    nb_qic=nb_res.qic()
    if isinstance(nb_qic, tuple):
        nb_qic=nb_qic[0]
    old_qic=final_fit.qic()
    if isinstance(old_qic, tuple):
        old_qic=old_qic[0]

    diff_qic=None
    conclusion=""
    if isinstance(nb_qic,(float,int)) and isinstance(old_qic,(float,int)):
        diff_qic=old_qic - nb_qic
        if diff_qic>0:
            conclusion="NegBin better"
        else:
            conclusion="No NB improvement"
    else:
        conclusion="Could not compare QIC"

    return {
        "NB_QIC":nb_qic,
        "Poisson_QIC":old_qic,
        "diff_QIC":diff_qic,
        "conclusion":conclusion
    }

###############################################################################
# cross_validation_gee
###############################################################################
def cross_validation_gee(data_with_score, formula, group_col,
                         family, cov_struct, n_folds=3):
    cluster_ids=data_with_score[group_col].unique()
    np.random.shuffle(cluster_ids)
    folds=np.array_split(cluster_ids,n_folds)
    metrics=[]
    for i in range(n_folds):
        testc=set(folds[i])
        train_df=data_with_score[~data_with_score[group_col].isin(testc)].copy()
        test_df=data_with_score[data_with_score[group_col].isin(testc)].copy()
        if len(train_df)<1 or len(test_df)<1:
            continue
        train_df["media_category"]=train_df["media_category"].astype("category")
        cats=train_df["media_category"].cat.categories
        test_df["media_category"]=pd.Categorical(test_df["media_category"],categories=cats)

        mod=GEE.from_formula(formula, groups=group_col,
                             data=train_df, family=family, cov_struct=cov_struct)
        # add maxiter=300
        res=mod.fit(maxiter=300)
        pred=res.predict(test_df)
        obs=test_df[res.model.endog_names]
        mse=np.mean((obs-pred)**2)
        metrics.append(mse)
    return np.mean(metrics) if len(metrics)>0 else np.nan

###############################################################################
# sensitivity_analysis_correlation
###############################################################################
def sensitivity_analysis_correlation(df, formula, group_col):
    alt_families=[Poisson(), NegativeBinomial(alpha=1.0), Gaussian()]
    alt_structs=[Independence(), Exchangeable()]
    results=[]
    for fam in alt_families:
        for covs in alt_structs:
            mod=GEE.from_formula(formula,groups=group_col,data=df,
                                 family=fam,cov_struct=covs)
            # add maxiter=300
            try:
                r=mod.fit(maxiter=300)
                qic_val=r.qic()
                if isinstance(qic_val, tuple):
                    qic_val=qic_val[0]
                param3=r.params.head(3).round(3).to_dict()
                results.append({
                    "Family":fam.__class__.__name__,
                    "CovStruct":covs.__class__.__name__,
                    "QIC":qic_val,
                    "ParamSample":param3
                })
            except:
                results.append({
                    "Family":fam.__class__.__name__,
                    "CovStruct":covs.__class__.__name__,
                    "QIC":np.nan,
                    "ParamSample":"fail"
                })
    return pd.DataFrame(results)

###############################################################################
# try_all_families_and_scales
###############################################################################
def try_all_families_and_scales(df, formula, group_col):
    families = [
        Poisson(),
        NegativeBinomial(alpha=1.0),
        Gaussian(),
        Gamma(link=ln.log()),
        InverseGaussian()
    ]
    cor_structs=[Independence(), Exchangeable()]
    scale_opts=["none","pearson","deviance","ub","bc"]

    best_qic=np.inf
    best_tuple=None
    all_results=[]
    for fam in families:
        fam_name=fam.__class__.__name__
        for cov_obj in cor_structs:
            cov_name=cov_obj.__class__.__name__
            for sc_opt in scale_opts:
                try:
                    model=GEE.from_formula(formula, groups=group_col,
                                           data=df, family=fam, cov_struct=cov_obj)
                    # add maxiter=300
                    base_res=model.fit(maxiter=300, scale=None)
                    y=base_res.model.endog
                    mu=base_res.fittedvalues
                    n=len(y)
                    p=len(base_res.params)
                    dfresid=n-p
                    final_res=base_res
                    if sc_opt!="none" and dfresid>0:
                        val=None
                        if sc_opt=="pearson":
                            val=compute_pearson_scale(y,mu,dfresid)
                        elif sc_opt=="deviance":
                            val=compute_deviance_scale(y,mu,dfresid)
                        elif sc_opt=="ub":
                            val=compute_ub_scale(y,mu,dfresid)
                        elif sc_opt=="bc":
                            val=compute_bc_scale(y,mu,dfresid)
                        if val is not None and not np.isnan(val):
                            # add maxiter=300
                            final_res=model.fit(maxiter=300, scale=val)

                    qic_val=final_res.qic()
                    if isinstance(qic_val, tuple):
                        qic_val=qic_val[0]
                    all_results.append((fam_name,cov_name,sc_opt,qic_val))
                    if qic_val<best_qic:
                        best_qic=qic_val
                        best_tuple=(fam_name,cov_name,sc_opt,qic_val)
                except Exception as e:
                    logging.warning(f"Fit fail => {fam_name}+{cov_name}+{sc_opt} => {e}")
                    all_results.append((fam_name,cov_name,sc_opt,np.nan))
    return best_tuple, all_results

###############################################################################
# refit_best_gee_with_scale
###############################################################################
def refit_best_gee_with_scale(df, sentiment, measure, fam_name, struct, scale_name):
    if measure=="Quotation":
        pat=rf"^{re.escape(sentiment)}_\d+$"
        matched=[c for c in df.columns if re.match(pat,c)]
        if not matched:
            return None,None
        d2=df.copy()
        d2["_score_col"]=d2[matched].clip(lower=0).mean(axis=1)
    elif measure=="Fulltext":
        fcol=f"{sentiment}_fulltext"
        if fcol not in df.columns:
            return None,None
        d2=df.copy()
        d2["_score_col"]=d2[fcol].clip(lower=0)
    elif measure=="Fulltext_Intensity":
        fcol=f"{sentiment}_fulltext_intensity"
        if fcol not in df.columns:
            return None,None
        d2=df.copy()
        d2["_score_col"]=d2[fcol].astype(float).clip(lower=0)
    elif measure=="Title_Intensity":
        fcol=f"{sentiment}_title_intensity"
        if fcol not in df.columns:
            return None,None
        d2=df.copy()
        d2["_score_col"]=d2[fcol].astype(float).clip(lower=0)
    elif measure=="Quotation_Intensity":
        pat=rf"^{re.escape(sentiment)}_\d+_intensity$"
        matched=[c for c in df.columns if re.match(pat,c)]
        if not matched:
            return None,None
        d2=df.copy()
        d2["_score_col"]=d2[matched].astype(float).clip(lower=0).mean(axis=1)
    else:
        return None,None

    needed=["_score_col","media_outlet_clean","media_category"]
    d2=d2.dropna(subset=needed)
    if len(d2)<2 or d2["media_category"].nunique()<2:
        return None,None

    if fam_name=="Poisson":
        fam_obj=Poisson()
    elif fam_name=="NegativeBinomial":
        fam_obj=NegativeBinomial(alpha=1.0)
    elif fam_name=="Gaussian":
        fam_obj=Gaussian()
    elif fam_name=="Gamma":
        fam_obj=Gamma(link=ln.log())
    elif fam_name=="InverseGaussian":
        fam_obj=InverseGaussian()
    else:
        logging.warning(f"Unknown family={fam_name}")
        return None,None

    if struct=="Independence":
        cov_obj=Independence()
    else:
        cov_obj=Exchangeable()

    model=GEE.from_formula("_score_col ~ media_category",
                           groups="media_outlet_clean",
                           data=d2,
                           family=fam_obj,
                           cov_struct=cov_obj)
    # add maxiter=300
    bres=model.fit(maxiter=300, scale=None)
    y=np.asarray(bres.model.endog)
    mu=np.asarray(bres.fittedvalues)
    n=len(y)
    p=len(bres.params)
    dfresid=n-p

    if scale_name=="none" or dfresid<=0:
        return d2,bres
    else:
        scale_val=None
        if scale_name=="pearson":
            scale_val=compute_pearson_scale(y,mu,dfresid)
        elif scale_name=="deviance":
            scale_val=compute_deviance_scale(y,mu,dfresid)
        elif scale_name=="ub":
            scale_val=compute_ub_scale(y,mu,dfresid)
        elif scale_name=="bc":
            scale_val=compute_bc_scale(y,mu,dfresid)
        if scale_val is not None and not np.isnan(scale_val):
            # add maxiter=300
            final_res=model.fit(maxiter=300, scale=scale_val)
            return d2, final_res
        else:
            return d2,bres

###############################################################################
# run_gee_for_sentiment_measure_best_qic, run_gee_analyses_best_qic
###############################################################################
def run_gee_for_sentiment_measure_best_qic(df, sentiment, measure):
    d2=df.copy()
    if measure=="Quotation":
        pat=rf"^{re.escape(sentiment)}_\d+$"
        matched=[c for c in d2.columns if re.match(pat,c)]
        if not matched:
            return None
        d2["_score_col"]=d2[matched].clip(lower=0).mean(axis=1)
    elif measure=="Fulltext":
        fcol=f"{sentiment}_fulltext"
        if fcol not in d2.columns:
            return None
        d2["_score_col"]=d2[fcol].clip(lower=0)
    elif measure=="Fulltext_Intensity":
        fcol=f"{sentiment}_fulltext_intensity"
        if fcol not in d2.columns:
            return None
        d2["_score_col"]=d2[fcol].astype(float).clip(lower=0)
    elif measure=="Title_Intensity":
        fcol=f"{sentiment}_title_intensity"
        if fcol not in d2.columns:
            return None
        d2["_score_col"]=d2[fcol].astype(float).clip(lower=0)
    elif measure=="Quotation_Intensity":
        pat=rf"^{re.escape(sentiment)}_\d+_intensity$"
        matched=[c for c in d2.columns if re.match(pat,c)]
        if not matched:
            return None
        d2["_score_col"]=d2[matched].astype(float).clip(lower=0).mean(axis=1)
    else:
        return None

    needed=["_score_col","media_outlet_clean","media_category"]
    d2=d2.dropna(subset=needed)
    if len(d2)<2 or d2["media_category"].nunique()<2:
        return None

    best_tuple, combos=try_all_families_and_scales(d2,"_score_col ~ media_category","media_outlet_clean")
    if best_tuple is None:
        return None
    famName, corName, scOpt, bestQICVal=best_tuple
    combos_df=pd.DataFrame(combos, columns=["Family","CovStruct","Scale","QIC"])

    return {
        "Sentiment":sentiment,
        "Measure":measure,
        "Best_Family":famName,
        "Best_Structure":corName,
        "Best_Scale":scOpt,
        "Best_QIC_main":bestQICVal,
        "AllCombos": combos_df
    }

def run_gee_analyses_best_qic(df):
    measure_list=[
        "Quotation",
        "Fulltext",
        "Fulltext_Intensity",
        "Title_Intensity",
        "Quotation_Intensity"
    ]
    best_list=[]
    combos_list=[]
    for s in CATEGORIES:
        for meas in measure_list:
            info=run_gee_for_sentiment_measure_best_qic(df,s,meas)
            if info is not None:
                best_list.append({
                    "Sentiment":info["Sentiment"],
                    "Measure":info["Measure"],
                    "Best_Family":info["Best_Family"],
                    "Best_Structure":info["Best_Structure"],
                    "Best_Scale":info["Best_Scale"],
                    "Best_QIC_main":info["Best_QIC_main"]
                })
                cdf=info["AllCombos"]
                cdf["Sentiment"]=s
                cdf["Measure"]=meas
                combos_list.append(cdf)

    df_best=pd.DataFrame(best_list)
    df_all=pd.concat(combos_list,ignore_index=True) if combos_list else pd.DataFrame()
    return df_best, df_all

###############################################################################
# exclude_small_clusters
###############################################################################
def exclude_small_clusters(df, cluster_col="media_outlet_clean", min_size=5, merge_into="OtherSmall"):
    df2=df.copy()
    sizes=df2[cluster_col].value_counts()
    small_clusters=sizes[sizes<min_size].index
    df2[cluster_col] = df2[cluster_col].apply(lambda c: merge_into if c in small_clusters else c)
    return df2

###############################################################################
# pairwise_and_diagnostics => fill COMPARE_IMBALANCE
###############################################################################
def pairwise_and_diagnostics(df, sentiment, measure, fam_name, struct, scale_name,
                             prefix="", isSens=False):
    global COMPARE_IMBALANCE
    d2, final_fit=refit_best_gee_with_scale(df, sentiment, measure, fam_name, struct, scale_name)
    if final_fit is None:
        return (None, None, None, None, None, None, None, None)

    summary_txt=final_fit.summary().as_text()
    diag=check_residuals_and_correlation(final_fit)
    pseudo_dict=pseudo_likelihood_check(final_fit)
    cv_val=cross_validation_gee(d2,"_score_col ~ media_category","media_outlet_clean",
                                final_fit.model.family, final_fit.model.cov_struct,
                                n_folds=3)
    sens_df=sensitivity_analysis_correlation(d2,"_score_col ~ media_category","media_outlet_clean")

    param_names=list(final_fit.params.index)
    param_values=final_fit.params.values
    robust_cov=final_fit.cov_params()
    robust_se=np.sqrt(np.diag(robust_cov))

    # M&D cluster bootstrap
    boot_params_df, boot_se_df=bootstrap_gee_md(
        d2,
        formula="_score_col ~ media_category",
        group_col="media_outlet_clean",
        B=200,
        family=final_fit.model.family,
        cov_struct=final_fit.model.cov_struct
    )
    param_stds=boot_params_df.std().round(4)

    for i, p in enumerate(param_names):
        sr_se=robust_se[i]
        sr_val=param_values[i]
        bs_std=param_stds.get(p, np.nan)
        ratio=np.nan
        if sr_se>0 and not np.isnan(sr_se):
            ratio=bs_std/sr_se
        COMPARE_IMBALANCE.append({
            "Sentiment":sentiment,
            "Measure":measure,
            "Param":p,
            "Prefix":prefix,
            "IsSens":isSens,
            "SingleRun_SE":sr_se,
            "MD_BootStd":bs_std,
            "Ratio":ratio,
            "ParamEst":sr_val
        })

    param_means=boot_params_df.mean().round(4).to_dict()
    param_std_dict=param_stds.to_dict()
    boot_info={
        "ParamMean":param_means,
        "ParamStd":param_std_dict
    }

    # pairwise diffs
    params=final_fit.params
    cov=final_fit.cov_params()
    mdf=final_fit.model.data.frame
    cats=mdf["media_category"].unique()
    if pd.api.types.is_categorical_dtype(mdf["media_category"]):
        cats=mdf["media_category"].cat.categories

    ref=cats[0]
    idx_map={ref:0}
    for c in cats[1:]:
        nm=f"media_category[T.{c}]"
        if nm in final_fit.model.exog_names:
            idx_map[c]=final_fit.model.exog_names.index(nm)

    pair_list=[]
    for i2 in range(len(cats)):
        for j2 in range(i2+1,len(cats)):
            ca,cb=cats[i2],cats[j2]
            if ca not in idx_map and cb not in idx_map:
                continue
            con=np.zeros(len(params))
            if ca==ref and cb in idx_map:
                con[idx_map[cb]]=-1.0
            elif cb==ref and ca in idx_map:
                con[idx_map[ca]]=1.0
            else:
                if ca in idx_map: con[idx_map[ca]]=1.0
                if cb in idx_map: con[idx_map[cb]]=-1.0

            diff_est=con@params
            diff_var=con@cov@con
            if diff_var<=1e-12:
                diff_se=np.nan
                z=np.nan
                pval=np.nan
            else:
                diff_se=np.sqrt(diff_var)
                z=diff_est/diff_se
                pval=2*(1-norm.cdf(abs(z)))
            pair_list.append((ca,cb,diff_est,diff_se,z,pval))

    pair_df=pd.DataFrame(pair_list,columns=["CategoryA","CategoryB","Difference","SE","Z","p_value"])
    if not pair_df.empty:
        rej,p_adj,_,_=multipletests(pair_df["p_value"],method="fdr_bh")
        pair_df["p_value_adj"]=p_adj
        pair_df["reject_H0"]=rej
    else:
        pair_df["p_value_adj"]=np.nan
        pair_df["reject_H0"]=False

    diff_map={c:set() for c in cats}
    cat_list=list(cats)
    cat_index_map={cat_list[i]:i+1 for i in range(len(cat_list))}
    for i3,row3 in pair_df.iterrows():
        A=row3["CategoryA"]
        B=row3["CategoryB"]
        if row3["reject_H0"]:
            diff_map[A].add(cat_index_map[B])
            diff_map[B].add(cat_index_map[A])

    rows=[]
    for c in cat_list:
        diffs=sorted(list(diff_map[c]))
        diffs_str=",".join(str(x) for x in diffs)
        rows.append((c,diffs_str))
    diffIDs_df=pd.DataFrame(rows,columns=["Category","DiffIDs"])

    return summary_txt, pair_df, diffIDs_df, diag, pseudo_dict, cv_val, sens_df, boot_info

###############################################################################
# compile_results_into_multiple_workbooks => includes Compare_Imbalance
###############################################################################
def compile_results_into_multiple_workbooks(
    aggregated_df, stats_df, raw_df,
    df_best_qic, df_all_combos,
    plots_dir,
    main_excel, raw_excel, gee_excel, plots_excel, combined_excel,
    df_full,
    validation_records=None
):
    from openpyxl import Workbook
    from openpyxl.utils.dataframe import dataframe_to_rows

    if validation_records is None:
        validation_records=[]

    # main => aggregated + stats
    with pd.ExcelWriter(main_excel, engine="openpyxl") as w:
        aggregated_df.to_excel(w,"Aggregated_Scores",index=False)
        stats_df.to_excel(w,"Mean_Median_Statistics",index=False)

    # raw => raw
    with pd.ExcelWriter(raw_excel, engine="openpyxl") as w:
        raw_df.to_excel(w,"Raw_Data",index=False)
        for s in CATEGORIES:
            s_cols=[c for c in raw_df.columns if c.startswith(s+"_")]
            s_df=raw_df[["media_category","media_outlet"]+s_cols].copy()
            s_df.to_excel(w,f"Raw_{s[:29]}",index=False)

    diag_records=[]
    cv_records=[]
    sens_records=[]
    boot_records=[]
    idx_rows=[]

    # GEE results
    with pd.ExcelWriter(gee_excel, engine="openpyxl") as writer:
        if not df_all_combos.empty:
            df_all_combos.to_excel(writer, sheet_name="All_Combos", index=False)

        for i,row in df_best_qic.iterrows():
            s=row["Sentiment"]
            meas=row["Measure"]
            fam=row["Best_Family"]
            st=row["Best_Structure"]
            sc=row["Best_Scale"]
            best_qic=row["Best_QIC_main"]

            sh_name=f"BestQIC_{s[:8]}_{meas[:12]}"
            out=pairwise_and_diagnostics(df_full,s,meas,fam,st,sc,prefix="",isSens=False)
            if out[0] is None:
                tmp_df=pd.DataFrame({"Summary":[f"No valid model => {s}-{meas}"]})
                tmp_df.to_excel(writer, sheet_name=sh_name, index=False)
                continue

            summary_txt, pair_df, diffIDs_df, diag_dict, pseudo_dict, cv_val, sens_df_df, boot_info=out

            lines=summary_txt.split("\n")
            sumdf=pd.DataFrame({"GEE_Summary":lines})
            sumdf.to_excel(writer, sheet_name=sh_name, index=False)

            sr=len(sumdf)+2
            pair_df.to_excel(writer, sheet_name=sh_name, index=False, startrow=sr)

            sr2=sr+len(pair_df)+2
            if diffIDs_df is not None and not diffIDs_df.empty:
                diffIDs_df.to_excel(writer, sheet_name=sh_name, index=False, startrow=sr2)

            # logs
            validation_records.append({
                "Check": "GEE_Residual_Check",
                "Sentiment": s,
                "Measure": meas,
                "Value": f"Overdisp={diag_dict['overdisp_ratio']:.2f}, Corr={diag_dict['avg_within_corr']:.2f}",
                "Conclusion": diag_dict["assessment"]
            })
            validation_records.append({
                "Check": "PseudoLik_Check",
                "Sentiment": s,
                "Measure": meas,
                "Value": f"NB_QIC={pseudo_dict['NB_QIC']}, Poisson_QIC={pseudo_dict['Poisson_QIC']}, diff={pseudo_dict['diff_QIC']}",
                "Conclusion": pseudo_dict["conclusion"]
            })
            validation_records.append({
                "Check": "CrossVal_MSE",
                "Sentiment": s,
                "Measure": meas,
                "Value": cv_val,
                "Conclusion": "No direct pass/fail"
            })

            diag_records.append({
                "Sentiment": s,
                "Measure": meas,
                "Family": fam,
                "Structure": st,
                "Scale": sc,
                "mean_pearson_res": diag_dict["mean_pearson_res"],
                "std_pearson_res": diag_dict["std_pearson_res"],
                "avg_within_corr": diag_dict["avg_within_corr"],
                "overdisp_ratio": diag_dict["overdisp_ratio"],
                "assessment": diag_dict["assessment"],
                "NB_QIC": pseudo_dict["NB_QIC"],
                "Poisson_QIC": pseudo_dict["Poisson_QIC"],
                "diff_QIC": pseudo_dict["diff_QIC"],
                "pseudo_conclusion": pseudo_dict["conclusion"]
            })

            sens_records.append(sens_df_df)

            pm_str=str(boot_info["ParamMean"])
            ps_str=str(boot_info["ParamStd"])
            boot_records.append({
                "Sentiment": s,
                "Measure": meas,
                "Family": fam,
                "Structure": st,
                "Scale": sc,
                "MD_BootParamMean": pm_str,
                "MD_BootParamStd":  ps_str
            })

            idx_rows.append({
                "Sentiment": s,
                "Measure": meas,
                "Family": fam,
                "Structure": st,
                "Scale": sc,
                "BestQIC": best_qic
            })

        idxdf=pd.DataFrame(idx_rows)
        idxdf.to_excel(writer, "BestQIC_Index", index=False)

        diag_df=pd.DataFrame(diag_records)
        diag_df.to_excel(writer, "Diagnostics", index=False)

        if len(sens_records)>0:
            sens_all=pd.concat(sens_records,ignore_index=True)
            sens_all.to_excel(writer,"Sensitivity_Analysis",index=False)

        if len(boot_records)>0:
            boot_all=pd.DataFrame(boot_records)
            boot_all.to_excel(writer,"Bootstrap_Res",index=False)

        val_df=pd.DataFrame(validation_records)
        val_df.to_excel(writer,"ValidationSummary",index=False)
        row_=len(val_df)+2
        conclusion_row=pd.DataFrame({
            "Brief_Conclusion":[
                "Above table shows each validation check (old & new). If any check indicates 'WARNING', 'Potential', or 'High imbalance', caution is advised."
            ]
        })
        conclusion_row.to_excel(writer,"ValidationSummary",index=False,startrow=row_)

    # "Compare_Imbalance"
    from openpyxl import load_workbook
    ci_df=pd.DataFrame(COMPARE_IMBALANCE)
    if os.path.exists(gee_excel):
        book=load_workbook(gee_excel)
    else:
        book=Workbook()
    ws_ci=book.create_sheet("Compare_Imbalance")
    if not ci_df.empty:
        ci_rows=dataframe_to_rows(ci_df,index=False,header=True)
        for r_i, row_val in enumerate(ci_rows,1):
            for c_i, cell_val in enumerate(row_val,1):
                ws_ci.cell(row=r_i,column=c_i,value=cell_val)
    else:
        ws_ci["A1"]="No cluster-imbalance data"
    if "Sheet" in book.sheetnames and len(book.sheetnames)>1:
        if "Sheet" in book.sheetnames:
            dum=book["Sheet"]
            book.remove(dum)
    book.save(gee_excel)

    # produce plots
    wb_plots=Workbook()
    if "Sheet" in wb_plots.sheetnames:
        wb_plots.remove(wb_plots["Sheet"])
    any_sheets=False
    for s in CATEGORIES:
        q_path=os.path.join(plots_dir,f"quote_{s}.png")
        if os.path.exists(q_path):
            st=f"Quote_{s[:28]}"
            ws=wb_plots.create_sheet(title=st)
            try:
                img=ExcelImage(q_path)
                img.anchor="A1"
                ws.add_image(img)
            except:
                pass
            any_sheets=True
        f_path=os.path.join(plots_dir,f"fulltext_{s}.png")
        if os.path.exists(f_path):
            st2=f"Fulltext_{s[:25]}"
            ws2=wb_plots.create_sheet(title=st2)
            try:
                img2=ExcelImage(f_path)
                img2.anchor="A1"
                ws2.add_image(img2)
            except:
                pass
            any_sheets=True
    if not any_sheets:
        wb_plots.create_sheet("DummySheet")
    wb_plots.save(plots_excel)

    # combined
    raw_clean=raw_df.copy()
    raw_clean=raw_clean.applymap(lambda x:", ".join(x) if isinstance(x,list) else x)
    wb_comb=Workbook()
    if "Sheet" in wb_comb.sheetnames:
        wb_comb.remove(wb_comb["Sheet"])

    ws_agg=wb_comb.create_sheet("Aggregated_Scores")
    for r in dataframe_to_rows(aggregated_df,index=False,header=True):
        ws_agg.append(r)
    ws_stats=wb_comb.create_sheet("Mean_Median_Statistics")
    for r in dataframe_to_rows(stats_df,index=False,header=True):
        ws_stats.append(r)
    ws_raw=wb_comb.create_sheet("Raw_Data")
    for r in dataframe_to_rows(raw_clean,index=False,header=True):
        ws_raw.append(r)
    ws_best=wb_comb.create_sheet("BestQIC_Table")
    for r in dataframe_to_rows(df_best_qic,index=False,header=True):
        ws_best.append(r)
    ws_val=wb_comb.create_sheet("ValidationSummary")
    val_df2=pd.DataFrame(validation_records)
    for row_ in dataframe_to_rows(val_df2,index=False,header=True):
        ws_val.append(row_)
    lr2=len(val_df2)+2
    ws_val.cell(row=lr2,column=1,value=(
        "Above table (old + new validations) for all checks. "
        "Look for warnings or potential issues."
    ))
    wb_comb.save(combined_excel)

###############################################################################
# run_pipeline_for_df => merges small clusters => prefix_sens
###############################################################################
def run_pipeline_for_df(
    df,
    prefix="",
    main_excel=OUTPUT_EXCEL_MAIN,
    raw_excel=OUTPUT_EXCEL_RAW,
    gee_excel=OUTPUT_EXCEL_GEE,
    plots_excel=OUTPUT_EXCEL_PLOTS,
    combined_excel=OUTPUT_EXCEL_COMBINED,
    outdir=BASE_OUTPUT_DIR,
    csvout=BASE_CSV_OUTPUT_DIR,
    min_size=5
):
    validation_records=[]

    c1=check_number_of_clusters(df,"media_outlet_clean",threshold=20)
    validation_records.append(c1)
    c2=check_cluster_balance(df,"media_outlet_clean",imbalance_ratio=5.0)
    validation_records.append(c2)
    c_cov=coverage_simulation(n_sims=1000,n_clusters=10,cluster_size=10,true_beta=0.5)
    validation_records.append(c_cov)

    if prefix=="Yes":
        cat_codes, unique_cats = pd.factorize(df["media_category"])
        df["media_outlet_clean"] = cat_codes+1
        logging.info(f"Collapsed each category => single cluster. #unique cats={len(unique_cats)}")

    chunk_and_save(df,20000,prefix=prefix)
    print_basic_stats(df,prefix=prefix)

    analyze_all_custom_correlations(df,prefix=prefix)

    agg_df=aggregate_sentiment_scores(df,CATEGORIES)
    agg_df=calculate_averages(agg_df)
    stats_df=calculate_mean_median(agg_df)
    save_aggregated_scores_to_csv(agg_df,csvout,prefix=prefix)
    plot_statistics(agg_df,outdir,prefix=prefix)

    df_best, df_allcombos=run_gee_analyses_best_qic(df)
    print(f"Best QIC => prefix={prefix}")
    print(df_best)

    compile_results_into_multiple_workbooks(
        aggregated_df=agg_df,
        stats_df=stats_df,
        raw_df=df,
        df_best_qic=df_best,
        df_all_combos=df_allcombos,
        plots_dir=outdir,
        main_excel=main_excel,
        raw_excel=raw_excel,
        gee_excel=gee_excel,
        plots_excel=plots_excel,
        combined_excel=combined_excel,
        df_full=df,
        validation_records=validation_records
    )

    if prefix:
        prefix_sens=prefix+"_Sens"
    else:
        prefix_sens="Sens"

    df_sens=exclude_small_clusters(df,min_size=min_size)
    if len(df_sens)<2:
        print(f"[Sensitivity] => merging <{min_size} => not enough data => skip.")
        return

    print(f"[Sensitivity] => Re-running pipeline => prefix={prefix_sens}, len={len(df_sens)}")
    run_pipeline_for_df_inner_sens(
        df_sens,
        prefix_sens,
        main_excel=OUTPUT_EXCEL_MAIN,
        raw_excel=OUTPUT_EXCEL_RAW,
        gee_excel=OUTPUT_EXCEL_GEE,
        plots_excel=OUTPUT_EXCEL_PLOTS,
        combined_excel=OUTPUT_EXCEL_COMBINED,
        outdir=outdir,
        csvout=csvout
    )

def run_pipeline_for_df_inner_sens(
    df_sens,
    prefix_sens,
    main_excel,
    raw_excel,
    gee_excel,
    plots_excel,
    combined_excel,
    outdir,
    csvout
):
    validation_records=[]

    if prefix_sens:
        main_excel=main_excel.replace(".xlsx", f"_{prefix_sens}.xlsx")
        raw_excel=raw_excel.replace(".xlsx", f"_{prefix_sens}.xlsx")
        gee_excel=gee_excel.replace(".xlsx", f"_{prefix_sens}.xlsx")
        plots_excel=plots_excel.replace(".xlsx", f"_{prefix_sens}.xlsx")
        combined_excel=combined_excel.replace(".xlsx", f"_{prefix_sens}.xlsx")

    c1=check_number_of_clusters(df_sens,"media_outlet_clean",threshold=20)
    validation_records.append(c1)
    c2=check_cluster_balance(df_sens,"media_outlet_clean",imbalance_ratio=5.0)
    validation_records.append(c2)
    c_cov=coverage_simulation(n_sims=1000,n_clusters=10,cluster_size=10,true_beta=0.5)
    validation_records.append(c_cov)

    chunk_and_save(df_sens,20000,prefix=prefix_sens)
    print_basic_stats(df_sens,prefix=prefix_sens)

    analyze_all_custom_correlations(df_sens,prefix=prefix_sens)

    agg_df=aggregate_sentiment_scores(df_sens,CATEGORIES)
    agg_df=calculate_averages(agg_df)
    stats_df=calculate_mean_median(agg_df)
    save_aggregated_scores_to_csv(agg_df,csvout,prefix=prefix_sens)
    plot_statistics(agg_df,outdir,prefix=prefix_sens)

    df_best, df_allcombos=run_gee_analyses_best_qic(df_sens)
    print(f"Best QIC => prefix={prefix_sens}")
    print(df_best)

    from openpyxl import Workbook
    from openpyxl.utils.dataframe import dataframe_to_rows

    diag_records=[]
    cv_records=[]
    sens_records=[]
    boot_records=[]
    idx_rows=[]

    with pd.ExcelWriter(gee_excel,engine="openpyxl") as writer:
        if not df_allcombos.empty:
            df_allcombos.to_excel(writer, sheet_name="All_Combos", index=False)

        for i,row in df_best.iterrows():
            s=row["Sentiment"]
            meas=row["Measure"]
            fam=row["Best_Family"]
            st=row["Best_Structure"]
            sc=row["Best_Scale"]
            best_qic=row["Best_QIC_main"]

            sh_name=f"BestQIC_{s[:8]}_{meas[:12]}"
            out=pairwise_and_diagnostics(df_sens,s,meas,fam,st,sc,
                                         prefix=prefix_sens,isSens=True)
            if out[0] is None:
                tmp_df=pd.DataFrame({"Summary":[f"No valid model => {s}-{meas}"]})
                tmp_df.to_excel(writer, sheet_name=sh_name, index=False)
                continue

            summary_txt, pair_df, diffIDs_df, diag_dict, pseudo_dict, cv_val, sens_df_df, boot_info=out

            lines=summary_txt.split("\n")
            sumdf=pd.DataFrame({"GEE_Summary":lines})
            sumdf.to_excel(writer, sheet_name=sh_name, index=False)

            sr=len(sumdf)+2
            pair_df.to_excel(writer, sheet_name=sh_name, index=False, startrow=sr)

            sr2=sr+len(pair_df)+2
            if diffIDs_df is not None and not diffIDs_df.empty:
                diffIDs_df.to_excel(writer, sheet_name=sh_name, index=False, startrow=sr2)

            diag_records.append({
                "Sentiment": s,
                "Measure": meas,
                "Family": fam,
                "Structure": st,
                "Scale": sc,
                "mean_pearson_res": diag_dict["mean_pearson_res"],
                "std_pearson_res": diag_dict["std_pearson_res"],
                "avg_within_corr": diag_dict["avg_within_corr"],
                "overdisp_ratio": diag_dict["overdisp_ratio"],
                "assessment": diag_dict["assessment"],
                "NB_QIC": pseudo_dict["NB_QIC"],
                "Poisson_QIC": pseudo_dict["Poisson_QIC"],
                "diff_QIC": pseudo_dict["diff_QIC"],
                "pseudo_conclusion": pseudo_dict["conclusion"]
            })

            pm_str=str(boot_info["ParamMean"])
            ps_str=str(boot_info["ParamStd"])
            boot_records.append({
                "Sentiment": s,
                "Measure": meas,
                "Family": fam,
                "Structure": st,
                "Scale": sc,
                "MD_BootParamMean": pm_str,
                "MD_BootParamStd":  ps_str
            })
            idx_rows.append({
                "Sentiment": s,
                "Measure": meas,
                "Family": fam,
                "Structure": st,
                "Scale": sc,
                "BestQIC": best_qic
            })

        idxdf=pd.DataFrame(idx_rows)
        idxdf.to_excel(writer, "BestQIC_Index", index=False)
        diag_df=pd.DataFrame(diag_records)
        diag_df.to_excel(writer, "Diagnostics", index=False)
        if len(boot_records)>0:
            boot_all=pd.DataFrame(boot_records)
            boot_all.to_excel(writer,"Bootstrap_Res",index=False)

    with pd.ExcelWriter(main_excel, engine="openpyxl") as w:
        agg_df.to_excel(w,"Aggregated_Scores",index=False)
        stats_df.to_excel(w,"Mean_Median_Statistics",index=False)

    with pd.ExcelWriter(raw_excel, engine="openpyxl") as w:
        df_sens.to_excel(w,"Raw_Data",index=False)

    wb_comb=Workbook()
    if "Sheet" in wb_comb.sheetnames:
        wb_comb.remove(wb_comb["Sheet"])
    ws_agg=wb_comb.create_sheet("Aggregated_Scores")
    for r in dataframe_to_rows(agg_df,index=False,header=True):
        ws_agg.append(r)
    ws_stats=wb_comb.create_sheet("Mean_Median_Statistics")
    for r in dataframe_to_rows(stats_df,index=False,header=True):
        ws_stats.append(r)
    wb_comb.save(combined_excel)

###############################################################################
# main => run for original vs collapsed
###############################################################################
def main():
    setup_logging()
    logging.info("Starting pipeline => yes/all => original cats, then yes/all => collapsed cats")

    df_raw=load_jsonl(INPUT_JSONL_FILE)
    print(f"Total articles loaded => {len(df_raw)}")

    if "high_rate_2" in df_raw.columns:
        df_raw["high_rate_2"]=df_raw["high_rate_2"].astype(str).str.strip().str.lower()

    # A) Original categories => yes / all
    print("\n=== A) Original Categories (6-cat) ===")
    df_orig=map_media_outlet_to_category(df_raw, MEDIA_CATEGORIES_ORIG)

    if "high_rate_2" in df_orig.columns:
        df_yes_orig=df_orig[df_orig["high_rate_2"]=="yes"].copy()
        print(f"df_yes_orig => {len(df_yes_orig)} rows")
        if len(df_yes_orig)>0:
            print("**Pipeline => Yes** (orig cats)")
            run_pipeline_for_df(df_yes_orig, prefix="Yes")
        else:
            print("No rows => skip yes subset (orig cats)")
    else:
        print("No 'high_rate_2' => skip yes subset (orig cats)")

    print("**Pipeline => All** (orig cats)")
    run_pipeline_for_df(df_orig, prefix="All")
    del df_orig
    gc.collect()

    # B) Collapsed categories => yes_collapsed / all_collapsed
    print("\n=== B) Collapsed Categories (4-cat) ===")
    df_coll=map_media_outlet_to_category(df_raw, MEDIA_CATEGORIES_COLLAPSED)

    if "high_rate_2" in df_coll.columns:
        df_yes_coll=df_coll[df_coll["high_rate_2"]=="yes"].copy()
        print(f"df_yes_collapsed => {len(df_yes_coll)} rows")
        if len(df_yes_coll)>0:
            print("**Pipeline => Yes_collapsed**")
            run_pipeline_for_df(df_yes_coll, prefix="Yes_collapsed")
        else:
            print("No rows => skip yes subset (collapsed cats)")
    else:
        print("No 'high_rate_2' => skip yes subset (collapsed cats)")

    print("**Pipeline => All_collapsed**")
    run_pipeline_for_df(df_coll, prefix="All_collapsed")
    del df_coll
    gc.collect()

    logging.info("All done => subset + full => original cats + collapsed cats.")


if __name__=="__main__":
    main()
