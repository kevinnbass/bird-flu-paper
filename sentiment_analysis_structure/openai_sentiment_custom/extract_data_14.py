#!/usr/bin/env python3
# extract_data_10_fixed_no_forceblank_with_aggregators.py
"""
Creates two Excel files, non-transposed, with *no forced blanking*:

File A) raw_values_6cat_60tabs_no_blanks_scatterplot.xlsx
   => 60 tabs => (previous 50 tabs) + (10 new "Title" tabs)
   => 10 sentiments × [Quotation, Fulltext, Title_Intensity, 
                       Quotation_Intensity, Fulltext_Intensity, Title]
   => We unify rows in groups [Quotation, Fulltext, Title_Intensity],
                               [Quotation_Intensity, Fulltext_Intensity, Title_Intensity],
                               [Title].
   => We do NOT forcibly blank anything if Quotation or Quotation_Intensity is missing.

   PLUS 6 aggregator tabs (one for each measure), no forced blanking:
   => "AGG_Quotation", "AGG_Fulltext", "AGG_Title_Intensity",
      "AGG_Quotation_Intensity", "AGG_Fulltext_Intensity", "AGG_Title"
   => each aggregator tab has [article_id + 10 sentiments], no category split

File B) raw_values_6cat_20tabs_intv2_no_blanks_scatterplot.xlsx
   => 20 tabs => 2 per sentiment: "3Int" and "2NonInt"
   => "3Int": [Fulltext_Intensity, Quotation_Intensity, Title_Intensity]
   => "2NonInt": [Fulltext, Quotation]
   => We do NOT forcibly blank anything if Quotation or Quotation_Intensity is missing.

Author: ChatGPT
"""

import json
import re
import numpy as np
import pandas as pd
from tqdm import tqdm
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# ------------------ GLOBAL CONFIG ------------------
INPUT_JSONL_FILE = "processed_all_articles_fixed_5.jsonl"

OUTPUT_60TABS_SCATTER = "raw_values_6cat_60tabs_no_blanks_scatterplot.xlsx"
OUTPUT_20TABS_SCATTER = "raw_values_6cat_20tabs_intv2_no_blanks_scatterplot.xlsx"

# 6 categories
MEDIA_CATEGORIES = {
    "Scientific": ["nature","sciam","stat","newscientist"],
    "Left": ["theatlantic","the daily beast","the intercept","mother jones",
             "msnbc","slate","vox","huffpost"],
    "Lean Left": ["ap","axios","cnn","guardian","business insider","nbcnews",
                  "npr","nytimes","politico","propublica","wapo","usa today"],
    "Center": ["reuters","marketwatch","financial times","newsweek","forbes"],
    "Lean Right": ["thedispatch","epochtimes","foxbusiness","wsj","national review","washtimes"],
    "Right": ["breitbart","theblaze","daily mail","dailywire","foxnews","nypost","newsmax"],
}

SENTIMENTS = [
    "joy","sadness","anger","fear",
    "surprise","disgust","trust","anticipation",
    "negative_sentiment","positive_sentiment"
]

MEASURES_5 = [
    "Fulltext",
    "Quotation",
    "Fulltext_Intensity",
    "Quotation_Intensity",
    "Title_Intensity"
]

# We'll also add "Title" as a new measure
THREE_INTENSITIES = ["Fulltext_Intensity","Quotation_Intensity","Title_Intensity"]
TWO_NONINTS       = ["Fulltext","Quotation"]

ALL_6_MEASURES = [
    "Quotation",
    "Fulltext",
    "Title_Intensity",
    "Quotation_Intensity",
    "Fulltext_Intensity",
    "Title",
]

# --------------------------------------------------------------------------------
# 1) Load + map
# --------------------------------------------------------------------------------
def load_jsonl(jsonl_file):
    records = []
    with open(jsonl_file,"r",encoding="utf-8") as f:
        for line in tqdm(f,desc="Loading JSONL"):
            rec = json.loads(line)
            records.append(rec)
    return pd.DataFrame(records)

def map_outlet_to_6cats(df):
    cat_map={}
    for cat, outls in MEDIA_CATEGORIES.items():
        for o in outls:
            cat_map[o.lower().strip()] = cat
    df2 = df.copy()
    df2["media_outlet_clean"] = df2["media_outlet"].str.lower().str.strip()
    df2["media_category_6"]   = df2["media_outlet_clean"].map(cat_map).fillna("Other")
    return df2


# --------------------------------------------------------------------------------
# 2) extract_measure_value
# --------------------------------------------------------------------------------
def extract_measure_value(row, sentiment, measure):
    """
    Quotation => average columns ^{sentiment}_\\d+$
    Quotation_Intensity => average columns ^{sentiment}_\\d+_intensity$
    Fulltext => row[f"{sentiment}_fulltext"]
    Fulltext_Intensity => row[f"{sentiment}_fulltext_intensity"]
    Title_Intensity => row[f"{sentiment}_title_intensity"]
    Title => row[f"{sentiment}_title"]
    """
    try:
        if measure=="Quotation":
            pat = rf"^{re.escape(sentiment)}_\d+$"
            matched = [c for c in row.index if re.match(pat,c)]
            if not matched:
                return np.nan
            vals = row[matched].astype(float).clip(lower=0)
            return vals.mean(skipna=True)

        elif measure=="Quotation_Intensity":
            pat = rf"^{re.escape(sentiment)}_\d+_intensity$"
            matched = [c for c in row.index if re.match(pat,c)]
            if not matched:
                return np.nan
            vals = row[matched].astype(float).clip(lower=0)
            return vals.mean(skipna=True)

        elif measure=="Fulltext":
            col = f"{sentiment}_fulltext"
            if col not in row.index:
                return np.nan
            return float(row[col])

        elif measure=="Fulltext_Intensity":
            col = f"{sentiment}_fulltext_intensity"
            if col not in row.index:
                return np.nan
            return float(row[col])

        elif measure=="Title_Intensity":
            col = f"{sentiment}_title_intensity"
            if col not in row.index:
                return np.nan
            return float(row[col])

        elif measure=="Title":
            col = f"{sentiment}_title"
            if col not in row.index:
                return np.nan
            return float(row[col])

        else:
            return np.nan
    except:
        return np.nan


# --------------------------------------------------------------------------------
# remove_default_sheet_if_present
# --------------------------------------------------------------------------------
def remove_default_sheet_if_present(wb):
    if "Sheet" in wb.sheetnames:
        default_ws = wb["Sheet"]
        wb.remove(default_ws)

# --------------------------------------------------------------------------------
# 3) unify_rows_across_measures => gather data for each measure, grouped by media category
#    * NO forced blanking logic *
# --------------------------------------------------------------------------------
def gather_article_vals_for_measure(df, sentiment, measure):
    """Return {cat: list of (article_id, measure_val)} for that measure, storing articles even if val=NaN."""
    cat_names = ["Scientific","Left","Lean Left","Center","Lean Right","Right"]
    cat_dict = {c: [] for c in cat_names}
    for i, row in df.iterrows():
        cat_6 = row["media_category_6"]
        if cat_6 not in cat_dict:
            continue
        a_id = row["article_id"]
        val  = extract_measure_value(row, sentiment, measure)
        cat_dict[cat_6].append((a_id,val))
    return cat_dict

def unify_rows_across_measures(df, sentiment, measure_group):
    """
    measure_group could be ["Quotation","Fulltext","Title_Intensity"],
    or ["Quotation_Intensity","Fulltext_Intensity","Title_Intensity"], 
    or ["Title"].
    
    No forced blanking: we simply gather values as is for each measure.
    """
    cat_names = ["Scientific","Left","Lean Left","Center","Lean Right","Right"]
    measure_data={}
    for meas in measure_group:
        measure_data[meas] = {cat:{} for cat in cat_names}

    # fill measure_data
    for meas in measure_group:
        cat_dict = gather_article_vals_for_measure(df, sentiment, meas)
        for cat in cat_names:
            for (a_id,v) in cat_dict[cat]:
                measure_data[meas][cat][a_id] = v

    # unify => cat_out
    cat_out={}
    for cat in cat_names:
        # gather all article IDs
        id_set = set()
        for meas in measure_group:
            id_set.update(measure_data[meas][cat].keys())
        sorted_ids = sorted(list(id_set))
        cat_out[cat] = []
        for a_id in sorted_ids:
            row_dict={"article_id": a_id}
            for meas in measure_group:
                val = measure_data[meas][cat].get(a_id, np.nan)
                row_dict[meas] = val
            cat_out[cat].append(row_dict)
    return cat_out

# --------------------------------------------------------------------------------
# 4) CREATE THE 60 TABS => measure_alignment_groups => produce separate tabs
# --------------------------------------------------------------------------------
def create_60tabs_scatter(df, wb):
    """
    Produce 60 tabs total (10 sentiments × 6 measures):
      Groups: 
        - [Quotation, Fulltext, Title_Intensity]
        - [Quotation_Intensity, Fulltext_Intensity, Title_Intensity]
        - [Title] (single)
    No forced blanking.
    """
    remove_default_sheet_if_present(wb)

    cat_names = ["Scientific","Left","Lean Left","Center","Lean Right","Right"]
    measure_alignment_groups = [
        ["Quotation","Fulltext","Title_Intensity"],
        ["Quotation_Intensity","Fulltext_Intensity","Title_Intensity"],
        ["Title"],
    ]
    for sent in SENTIMENTS:
        for group in measure_alignment_groups:
            cat_out = unify_rows_across_measures(df, sent, group)
            for meas in group:
                tab_name=f"{sent[:10]}_{meas[:10]}"
                if tab_name in wb.sheetnames:
                    continue

                max_len = max(len(cat_out[c]) for c in cat_names) if len(cat_names)>0 else 0
                data_dict={}
                for c in cat_names:
                    col_vals=[]
                    row_list = cat_out[c]
                    for r_i in range(max_len):
                        if r_i<len(row_list):
                            val= row_list[r_i].get(meas,np.nan)
                            col_vals.append(val)
                        else:
                            col_vals.append(np.nan)
                    data_dict[c]=col_vals
                df_tab = pd.DataFrame(data_dict, columns=cat_names)

                ws = wb.create_sheet(title=tab_name)
                # first row => col names
                row_data = [cat_names]
                for r_i in range(len(df_tab)):
                    rowvals=[]
                    for c_ in cat_names:
                        rowvals.append(df_tab.loc[r_i,c_])
                    row_data.append(rowvals)
                # write to sheet
                for R_i, rowv in enumerate(row_data,1):
                    for C_i, valv in enumerate(rowv,1):
                        ws.cell(row=R_i,column=C_i,value=valv)

# --------------------------------------------------------------------------------
# 5) CREATE THE 20 TABS => 2 tabs / sentiment => "3Int" & "2NonInt"
# --------------------------------------------------------------------------------
def create_20tabs_scatter(df, wb_20):
    """
    20 tabs => for each sentiment:
       - "3Int": [Fulltext_Intensity, Quotation_Intensity, Title_Intensity]
       - "2NonInt": [Fulltext, Quotation]
    No forced blanking.
    """
    remove_default_sheet_if_present(wb_20)

    cat_names = ["Scientific","Left","Lean Left","Center","Lean Right","Right"]

    for sent in SENTIMENTS:
        # (A) 3Int
        measure_grpA = ["Fulltext_Intensity","Quotation_Intensity","Title_Intensity"]
        cat_outA= unify_rows_across_measures(df, sent, measure_grpA)
        out_colsA=[]
        for cat in cat_names:
            for m in measure_grpA:
                out_colsA.append(f"{cat}_{m}")

        max_lenA = max(len(cat_outA[c]) for c in cat_names)
        dataA={}
        for colN in out_colsA:
            dataA[colN]=[]
        for cat in cat_names:
            row_list = cat_outA[cat]
            for m in measure_grpA:
                col_key=f"{cat}_{m}"
                col_vals=[]
                for r_i in range(max_lenA):
                    if r_i<len(row_list):
                        val = row_list[r_i].get(m,np.nan)
                        col_vals.append(val)
                    else:
                        col_vals.append(np.nan)
                dataA[col_key] = col_vals

        dfA = pd.DataFrame(dataA, columns=out_colsA)
        tabA_name= f"{sent[:10]}_3Int"
        wsA=wb_20.create_sheet(title=tabA_name)

        row_dataA = [out_colsA]
        for r_i in range(len(dfA)):
            rowvals=[]
            for c_ in out_colsA:
                rowvals.append(dfA.loc[r_i,c_])
            row_dataA.append(rowvals)

        for R_i, rowv in enumerate(row_dataA,1):
            for C_i, valv in enumerate(rowv,1):
                wsA.cell(row=R_i,column=C_i,value=valv)

        # (B) 2NonInt
        measure_grpB=["Fulltext","Quotation"]
        cat_outB=unify_rows_across_measures(df, sent, measure_grpB)
        out_colsB=[]
        for cat in cat_names:
            for m in measure_grpB:
                out_colsB.append(f"{cat}_{m}")

        max_lenB= max(len(cat_outB[c]) for c in cat_names)
        dataB={}
        for colN in out_colsB:
            dataB[colN]=[]
        for cat in cat_names:
            row_list = cat_outB[cat]
            for m in measure_grpB:
                col_key=f"{cat}_{m}"
                col_vals=[]
                for r_i in range(max_lenB):
                    if r_i< len(row_list):
                        val = row_list[r_i].get(m,np.nan)
                        col_vals.append(val)
                    else:
                        col_vals.append(np.nan)
                dataB[col_key]= col_vals

        dfB= pd.DataFrame(dataB, columns=out_colsB)
        tabB_name= f"{sent[:10]}_2NonInt"
        wsB= wb_20.create_sheet(title=tabB_name)

        row_dataB = [out_colsB]
        for r_i in range(len(dfB)):
            rowvals=[]
            for c_ in out_colsB:
                rowvals.append(dfB.loc[r_i,c_])
            row_dataB.append(rowvals)

        for R_i, rowv in enumerate(row_dataB,1):
            for C_i, valv in enumerate(rowv,1):
                wsB.cell(row=R_i,column=C_i,value=valv)

# --------------------------------------------------------------------------------
# 6) CREATE 6 “Aggregator” TABS => one for each measure, all articles, no category split
# --------------------------------------------------------------------------------
def create_aggregator_tabs_all_articles(df, wb):
    """
    For each measure in ALL_6_MEASURES, we produce a single sheet with:
        - 1 row per article
        - 1 column for article_id
        - 10 columns for the 10 sentiments
    No forced blanking logic.
    """
    for measure in ALL_6_MEASURES:
        rows_out = []
        for idx, row in df.iterrows():
            a_id = row["article_id"]
            row_dict = {"article_id": a_id}
            for s in SENTIMENTS:
                row_dict[s] = extract_measure_value(row, s, measure)
            rows_out.append(row_dict)

        col_order = ["article_id"] + SENTIMENTS
        df_tab = pd.DataFrame(rows_out, columns=col_order)

        tab_name = f"AGG_{measure[:10]}"
        ws = wb.create_sheet(title=tab_name)

        row_data = [col_order]  # header row
        for r_i in range(len(df_tab)):
            rowvals=[]
            for c_ in col_order:
                rowvals.append(df_tab.loc[r_i,c_])
            row_data.append(rowvals)

        for R_i, rowv in enumerate(row_data, 1):
            for C_i, valv in enumerate(rowv, 1):
                ws.cell(row=R_i, column=C_i, value=valv)

# --------------------------------------------------------------------------------
# 7) main
# --------------------------------------------------------------------------------
def main():
    print("Loading + mapping data...")
    df = load_jsonl(INPUT_JSONL_FILE)
    df = map_outlet_to_6cats(df)
    if "article_id" not in df.columns:
        df["article_id"] = df.index + 1

    print("Creating 60+6-tab scatter file =>", OUTPUT_60TABS_SCATTER)
    wb60 = Workbook()
    # (1) Create 60 tabs (by sentiment × measure, splitted by category)
    create_60tabs_scatter(df, wb60)
    # (2) Add 6 aggregator tabs (no category split)
    create_aggregator_tabs_all_articles(df, wb60)
    wb60.save(OUTPUT_60TABS_SCATTER)
    print(f"Wrote => {OUTPUT_60TABS_SCATTER} (should contain 66 tabs total)")

    print("Creating 20-tab scatter file =>", OUTPUT_20TABS_SCATTER)
    wb20 = Workbook()
    create_20tabs_scatter(df, wb20)
    wb20.save(OUTPUT_20TABS_SCATTER)
    print(f"Wrote => {OUTPUT_20TABS_SCATTER} (20 tabs)")

    print("All done.")

if __name__=="__main__":
    main()
