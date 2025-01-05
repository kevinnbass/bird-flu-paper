import json
import pandas as pd
import matplotlib.pyplot as plt
import seaborn as sns
import os
import logging
from tqdm import tqdm
from openpyxl.drawing.image import Image as ExcelImage
import re
import statsmodels.api as sm
from statsmodels.formula.api import mixedlm
from statsmodels.stats.multicomp import pairwise_tukeyhsd
import sys
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import numpy as np

# ------------------------------ #
#        Configuration           #
# ------------------------------ #

INPUT_JSONL_FILE = 'processed_all_articles_with_fulltext_sentiment_analysis.jsonl'
OUTPUT_DIR = 'graphs_analysis'
os.makedirs(OUTPUT_DIR, exist_ok=True)

CSV_OUTPUT_DIR = 'csv_raw_scores'
os.makedirs(CSV_OUTPUT_DIR, exist_ok=True)

# Multiple XLSX outputs
OUTPUT_EXCEL_MAIN = 'analysis_main.xlsx'
OUTPUT_EXCEL_RAW = 'analysis_raw.xlsx'
OUTPUT_EXCEL_LMM = 'analysis_lmm.xlsx'
OUTPUT_EXCEL_PLOTS = 'analysis_plots.xlsx'

# Single combined XLSX output
OUTPUT_EXCEL_COMBINED = 'analysis_combined.xlsx'

CATEGORIES = [
    'joy', 'sadness', 'anger', 'fear',
    'surprise', 'disgust', 'trust', 'anticipation',
    'negative_sentiment', 'positive_sentiment'
]

MEDIA_CATEGORIES = {
    'Scientific': ['nature', 'sciam', 'stat', 'newscientist'],
    'Left': ['theatlantic', 'the daily beast', 'the intercept', 'mother jones', 'msnbc', 'slate', 'vox', 'huffpost'],
    'Lean Left': ['ap', 'axios', 'cnn', 'guardian', 'business insider', 'nbcnews', 'npr', 'nytimes', 'politico', 'propublica', 'wapo', 'usa today'],
    'Center': ['reuters', 'marketwatch', 'financial times', 'newsweek', 'forbes'],
    'Lean Right': ['thedispatch', 'epochtimes', 'foxbusiness', 'wsj', 'national review', 'washtimes'],
    'Right': ['breitbart', 'theblaze', 'daily mail', 'dailywire', 'foxnews', 'nypost', 'newsmax'],
}


def setup_logging(log_file='analysis.log'):
    logging.basicConfig(
        filename=log_file,
        level=logging.DEBUG,
        format='%(asctime)s - %(levelname)s - %(message)s'
    )
    logging.info("Logging initialized.")


def load_jsonl(jsonl_file):
    records = []
    with open(jsonl_file, 'r', encoding='utf-8') as f:
        for line in tqdm(f, desc="Loading JSONL data"):
            try:
                record = json.loads(line)
                records.append(record)
            except json.JSONDecodeError as e:
                logging.error(f"JSON decoding error: {e}")
    df = pd.DataFrame(records)
    return df


def map_media_outlet_to_category(df, media_categories):
    outlet_to_category = {}
    for category, outlets in media_categories.items():
        for outlet in outlets:
            outlet_to_category[outlet.lower().strip()] = category

    if 'media_outlet' not in df.columns:
        logging.error("'media_outlet' column not found in DataFrame.")
        raise KeyError("'media_outlet' column not found in DataFrame.")

    df['media_outlet_clean'] = df['media_outlet'].str.lower().str.strip()
    df['media_category'] = df['media_outlet_clean'].map(outlet_to_category)
    df['media_category'] = df['media_category'].fillna('Other')
    unmapped_outlets = df[df['media_category'] == 'Other']['media_outlet'].unique()
    if len(unmapped_outlets) > 0:
        logging.warning(f"Unmapped media outlets found: {unmapped_outlets}")
        print(f"Warning: The following media outlets were not mapped and categorized as 'Other': {unmapped_outlets}")
    return df


def aggregate_sentiment_scores(df, sentiment_categories):
    aggregation = []
    for media_category in MEDIA_CATEGORIES.keys():
        mask = (df['media_category'] == media_category)
        subset = df[mask]

        for sentiment in sentiment_categories:
            pattern = f"^{re.escape(sentiment)}_\\d+$"
            matched_cols = [col for col in df.columns if re.match(pattern, col)]

            if matched_cols:
                negative_quot = subset[matched_cols] < 0
                negative_quot_count = negative_quot.sum().sum()
                if negative_quot_count > 0:
                    logging.warning(f"{negative_quot_count} negative quotation-based scores found for '{sentiment}' in '{media_category}' and set to zero.")
                quotation_sum = subset[matched_cols].clip(lower=0).sum(skipna=True).sum()
                quotation_count = subset[matched_cols].clip(lower=0).count().sum()
            else:
                quotation_sum = 0
                quotation_count = 0
                logging.warning(f"No quotation-based fields found for '{sentiment}' in '{media_category}'.")

            fulltext_col = f"{sentiment}_fulltext"
            if fulltext_col in df.columns:
                negative_full = subset[fulltext_col] < 0
                negative_full_count = negative_full.sum()
                if negative_full_count > 0:
                    logging.warning(f"{negative_full_count} negative fulltext-based scores found for '{sentiment}' in '{media_category}' and set to zero.")
                fulltext_sum = subset[fulltext_col].clip(lower=0).sum(skipna=True)
                fulltext_count = subset[fulltext_col].clip(lower=0).count()
            else:
                fulltext_sum = 0
                fulltext_count = 0
                logging.warning(f"Fulltext field '{fulltext_col}' not found in DataFrame.")

            aggregation.append({
                'Media Category': media_category,
                'Sentiment/Emotion': sentiment,
                'Quotation_Sum': quotation_sum,
                'Quotation_Count': quotation_count,
                'Fulltext_Sum': fulltext_sum,
                'Fulltext_Count': fulltext_count
            })

    aggregated_df = pd.DataFrame(aggregation)
    return aggregated_df


def calculate_averages(aggregated_df):
    aggregated_df['Quotation_Average'] = aggregated_df.apply(
        lambda row: row['Quotation_Sum'] / row['Quotation_Count'] if row['Quotation_Count'] > 0 else None, axis=1
    )
    aggregated_df['Fulltext_Average'] = aggregated_df.apply(
        lambda row: row['Fulltext_Sum'] / row['Fulltext_Count'] if row['Fulltext_Count'] > 0 else None, axis=1
    )

    negative_quot_avg = aggregated_df['Quotation_Average'] < 0
    negative_full_avg = aggregated_df['Fulltext_Average'] < 0

    if aggregated_df['Quotation_Average'][negative_quot_avg].any():
        logging.error("Negative values found in Quotation_Average. Setting them to None.")
        aggregated_df.loc[negative_quot_avg, 'Quotation_Average'] = None

    if aggregated_df['Fulltext_Average'][negative_full_avg].any():
        logging.error("Negative values found in Fulltext_Average. Setting them to None.")
        aggregated_df.loc[negative_full_avg, 'Fulltext_Average'] = None

    return aggregated_df


def calculate_mean_median(aggregated_df):
    stats = []
    for sentiment in CATEGORIES:
        sentiment_data = aggregated_df[aggregated_df['Sentiment/Emotion'] == sentiment]

        quotation_avg = sentiment_data['Quotation_Average'].dropna()
        if not quotation_avg.empty:
            mean_quotation = quotation_avg.mean()
            median_quotation = quotation_avg.median()
        else:
            mean_quotation = None
            median_quotation = None
            logging.warning(f"No quotation-based data for '{sentiment}'.")

        fulltext_avg = sentiment_data['Fulltext_Average'].dropna()
        if not fulltext_avg.empty:
            mean_fulltext = fulltext_avg.mean()
            median_fulltext = fulltext_avg.median()
        else:
            mean_fulltext = None
            median_fulltext = None
            logging.warning(f"No fulltext-based data for '{sentiment}'.")

        stats.append({
            'Sentiment/Emotion': sentiment,
            'Mean_Quotation_Average': mean_quotation,
            'Median_Quotation_Average': median_quotation,
            'Mean_Fulltext_Average': mean_fulltext,
            'Median_Fulltext_Average': median_fulltext
        })

    stats_df = pd.DataFrame(stats)
    return stats_df


def save_aggregated_scores_to_csv(aggregated_df, csv_output_dir):
    csv_file = os.path.join(csv_output_dir, "aggregated_sentiment_emotion_scores.csv")
    try:
        aggregated_df.to_csv(csv_file, index=False)
        print(f"Aggregated sentiment/emotion scores saved to '{csv_file}'.")
        logging.info(f"Aggregated sentiment/emotion scores saved to '{csv_file}'.")
    except Exception as e:
        print(f"Error saving aggregated scores to CSV: {e}")
        logging.error(f"Error saving aggregated scores to CSV: {e}")


def plot_statistics(aggregated_df, output_dir):
    sns.set_style("whitegrid")
    for sentiment in CATEGORIES:
        sentiment_data = aggregated_df[aggregated_df['Sentiment/Emotion'] == sentiment]

        # Quotation-Based Plot
        plt.figure(figsize=(10, 6))
        sns.barplot(
            x='Media Category',
            y='Quotation_Average',
            data=sentiment_data,
            color='steelblue'
        )
        plt.title(f"Mean Quotation-Based '{sentiment.capitalize()}' Scores Across Media Categories", fontsize=14)
        plt.xlabel('Media Category', fontsize=12)
        plt.ylabel('Mean Quotation-Based Average Score', fontsize=12)
        plt.xticks(rotation=45, ha='right')
        plt.tight_layout()
        quote_plot_filename = f"quote_{sentiment}.png"
        quote_plot_path = os.path.join(output_dir, quote_plot_filename)
        try:
            plt.savefig(quote_plot_path)
            plt.close()
            print(f"Quotation-Based '{sentiment}' scores plot saved to '{quote_plot_path}'.")
            logging.info(f"Quotation-Based '{sentiment}' scores plot saved to '{quote_plot_path}'.")
        except Exception as e:
            print(f"Error saving quotation-based '{sentiment}' scores plot: {e}")
            logging.error(f"Error saving quotation-based '{sentiment}' scores plot: {e}")

        # Fulltext-Based Plot
        plt.figure(figsize=(10, 6))
        sns.barplot(
            x='Media Category',
            y='Fulltext_Average',
            data=sentiment_data,
            color='darkorange'
        )
        plt.title(f"Mean Fulltext-Based '{sentiment.capitalize()}' Scores Across Media Categories", fontsize=14)
        plt.xlabel('Media Category', fontsize=12)
        plt.ylabel('Mean Fulltext-Based Average Score', fontsize=12)
        plt.xticks(rotation=45, ha='right')
        plt.tight_layout()
        fulltext_plot_filename = f"fulltext_{sentiment}.png"
        fulltext_plot_path = os.path.join(output_dir, fulltext_plot_filename)
        try:
            plt.savefig(fulltext_plot_path)
            plt.close()
            print(f"Fulltext-Based '{sentiment}' scores plot saved to '{fulltext_plot_path}'.")
            logging.info(f"Fulltext-Based '{sentiment}' scores plot saved to '{fulltext_plot_path}'.")
        except Exception as e:
            print(f"Error saving fulltext-based '{sentiment}' scores plot: {e}")
            logging.error(f"Error saving fulltext-based '{sentiment}' scores plot: {e}")


def fit_lmm_and_posthoc(df, sentiment, measure_type='Quotation'):
    logger = logging.getLogger(__name__)

    if measure_type == 'Quotation':
        pattern = f"^{re.escape(sentiment)}_\\d+$"
        matched_cols = [col for col in df.columns if re.match(pattern, col)]
        if not matched_cols:
            logger.debug(f"No matched quotation-based columns for sentiment='{sentiment}'. Skipping.")
            return None
        df = df.copy()
        df[f'{sentiment}_quotation_mean'] = df[matched_cols].clip(lower=0).mean(axis=1)
        score_col = f'{sentiment}_quotation_mean'
    else:
        fulltext_col = f"{sentiment}_fulltext"
        if fulltext_col not in df.columns:
            logger.debug(f"No fulltext column found for sentiment='{sentiment}'. Skipping.")
            return None
        df = df.copy()
        df[f'{sentiment}_fulltext_clipped'] = df[fulltext_col].clip(lower=0)
        score_col = f'{sentiment}_fulltext_clipped'

    model_df = df.dropna(subset=[score_col, 'media_category', 'media_outlet']).copy()

    row_count = len(model_df)
    unique_categories = model_df['media_category'].unique()
    category_count = len(unique_categories)
    logger.debug(f"Fitting LMM for sentiment='{sentiment}', measure_type='{measure_type}'.")
    logger.debug(f"Data subset size: {row_count} rows.")
    logger.debug(f"Unique media_category values: {unique_categories} (count={category_count}).")

    if row_count < 2 or category_count < 2:
        logger.debug("Not enough data points or categories to fit the model. Skipping.")
        return None

    cat_counts = model_df['media_category'].value_counts()
    logger.debug(f"Media category counts:\n{cat_counts.to_string()}")

    outlet_counts = model_df['media_outlet'].value_counts()
    logger.debug(f"media_outlet counts:\n{outlet_counts.to_string()}")
    # Check variation per outlet:
    outlets_with_variation_issues = []
    for outlet in outlet_counts.index:
        outlet_data = model_df.loc[model_df['media_outlet'] == outlet, score_col]
        if outlet_data.nunique() < 2:
            outlets_with_variation_issues.append((outlet, outlet_data.nunique()))
    print("Checking each media outlet for sufficient variation in the selected score column...")
    logging.info("Checking each media outlet for sufficient variation in the selected score column...")
    if outlets_with_variation_issues:
        for outlet, uniq_count in outlets_with_variation_issues:
            print(f"Outlet '{outlet}' may have insufficient variation (unique values: {uniq_count}).")
            logging.info(f"Outlet '{outlet}' may have insufficient variation (unique values: {uniq_count}).")
    else:
        print("All outlets show some variation in the response measure.")
        logging.info("All outlets show some variation in the response measure.")

    score_describe = model_df[score_col].describe()
    logger.debug(f"Score column '{score_col}' distribution:\n{score_describe.to_string()}")

    model_df.loc[:, 'media_category'] = pd.Categorical(model_df['media_category'])

    try:
        md = mixedlm(f"{score_col} ~ media_category", data=model_df, groups=model_df["media_outlet"])
        mdf = md.fit(reml=True, method='lbfgs')
    except Exception as e:
        logger.warning(f"Model failed to converge for {sentiment}-{measure_type}: {e}")
        return None

    if not mdf.converged:
        logger.warning(f"LMM did not converge for {sentiment}-{measure_type}.")

    summ_text = mdf.summary().as_text()
    if "Random effects covariance is singular" in summ_text:
        logger.warning("Random effects covariance singular message found in LMM summary.")

    cat_means = model_df.groupby('media_category', observed=True)[score_col].mean().reset_index()
    if cat_means['media_category'].nunique() < 2:
        logger.debug("Only one category after grouping. Skipping post-hoc tests.")
        return {
            'LMM_Summary': mdf.summary().as_text(),
            'PostHoc': pd.DataFrame()
        }

    try:
        tukey = pairwise_tukeyhsd(cat_means[score_col], cat_means['media_category'], alpha=0.05)
        posthoc_df = pd.DataFrame(data=tukey._results_table.data[1:], columns=tukey._results_table.data[0])
        logger.debug("Post-hoc tests completed successfully.")
    except Exception as e:
        logger.warning(f"Post-hoc failed for {sentiment}-{measure_type}: {e}")
        posthoc_df = pd.DataFrame()

    results_dict = {
        'LMM_Summary': mdf.summary().as_text(),
        'PostHoc': posthoc_df
    }

    return results_dict


def run_lmm_analyses(df):
    all_results = {}
    for sentiment in CATEGORIES:
        all_results[sentiment] = {}
        for measure_type in ['Quotation', 'Fulltext']:
            res = fit_lmm_and_posthoc(df, sentiment, measure_type=measure_type)
            if res is not None:
                all_results[sentiment][measure_type] = res
    return all_results


def compile_results_into_multiple_workbooks(aggregated_df, stats_df, raw_df, lmm_results, plots_dir):
    # MAIN WORKBOOK
    with pd.ExcelWriter(OUTPUT_EXCEL_MAIN, engine='openpyxl') as writer:
        aggregated_df.to_excel(writer, sheet_name='Aggregated_Scores', index=False)
        stats_df.to_excel(writer, sheet_name='Mean_Median_Statistics', index=False)
    logging.info(f"Main aggregated scores and stats compiled into '{OUTPUT_EXCEL_MAIN}'.")

    # RAW WORKBOOK
    with pd.ExcelWriter(OUTPUT_EXCEL_RAW, engine='openpyxl') as writer:
        raw_df.to_excel(writer, sheet_name='Raw_Data', index=False)
        for sentiment in CATEGORIES:
            sentiment_cols = [c for c in raw_df.columns if c.startswith(sentiment+'_')]
            sentiment_df = raw_df[['media_category', 'media_outlet'] + sentiment_cols].copy()
            sheet_name = f"Raw_{sentiment[:29]}"
            sentiment_df.to_excel(writer, sheet_name=sheet_name, index=False)
    logging.info(f"Raw data and raw sentiment sheets compiled into '{OUTPUT_EXCEL_RAW}'.")

    # LMM WORKBOOK
    with pd.ExcelWriter(OUTPUT_EXCEL_LMM, engine='openpyxl') as writer:
        summary_rows = []
        for sentiment in lmm_results:
            for measure_type in lmm_results[sentiment]:
                sheet_name = f"LMM_{sentiment[:20]}_{measure_type[:8]}"
                lmm_summary = lmm_results[sentiment][measure_type]['LMM_Summary']
                posthoc_df = lmm_results[sentiment][measure_type]['PostHoc']

                summary_df = pd.DataFrame({'LMM_Summary': lmm_summary.split('\n')})
                summary_df.to_excel(writer, sheet_name=sheet_name, index=False)
                startrow = len(summary_df) + 2
                posthoc_df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=startrow)

                summary_rows.append({
                    'Sentiment': sentiment,
                    'Measure': measure_type,
                    'SheetName': sheet_name
                })

        summary_index_df = pd.DataFrame(summary_rows)
        summary_index_df.to_excel(writer, sheet_name='LMM_Results_Index', index=False)
    logging.info(f"LMM results and post-hoc tests compiled into '{OUTPUT_EXCEL_LMM}'.")

    # PLOTS WORKBOOK
    wb = Workbook()
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])

    for sentiment in CATEGORIES:
        quote_plot_filename = f"quote_{sentiment}.png"
        quote_plot_path = os.path.join(plots_dir, quote_plot_filename)
        if os.path.exists(quote_plot_path):
            sheet_title = f"Quote_{sentiment[:28]}"
            worksheet = wb.create_sheet(title=sheet_title)
            try:
                img = ExcelImage(quote_plot_path)
                img.anchor = 'A1'
                worksheet.add_image(img)
            except Exception as e:
                logging.error(f"Error embedding image '{quote_plot_path}' into Excel: {e}")

        fulltext_plot_filename = f"fulltext_{sentiment}.png"
        fulltext_plot_path = os.path.join(plots_dir, fulltext_plot_filename)
        if os.path.exists(fulltext_plot_path):
            sheet_title = f"Fulltext_{sentiment[:25]}"
            worksheet = wb.create_sheet(title=sheet_title)
            try:
                img = ExcelImage(fulltext_plot_path)
                img.anchor = 'A1'
                worksheet.add_image(img)
            except Exception as e:
                logging.error(f"Error embedding image '{fulltext_plot_path}' into Excel: {e}")

    wb.save(OUTPUT_EXCEL_PLOTS)
    logging.info(f"All plots compiled into '{OUTPUT_EXCEL_PLOTS}'.")

    print(f"All statistics, raw data, LMM results, and plots have been compiled into multiple Excel files:\n"
          f" - {OUTPUT_EXCEL_MAIN}\n"
          f" - {OUTPUT_EXCEL_RAW}\n"
          f" - {OUTPUT_EXCEL_LMM}\n"
          f" - {OUTPUT_EXCEL_PLOTS}")
    logging.info("All results compiled into multiple Excel files successfully.")


def compile_into_single_combined_workbook(aggregated_df, stats_df, raw_df, lmm_results, plots_dir):
    # Create a cleaned copy of raw_df where lists are converted to strings
    raw_df_cleaned = raw_df.copy()
    raw_df_cleaned = raw_df_cleaned.applymap(lambda x: ", ".join(x) if isinstance(x, list) else x)

    wb = Workbook()
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])

    # Aggregated Scores
    ws_agg = wb.create_sheet("Aggregated_Scores")
    for r in dataframe_to_rows(aggregated_df, index=False, header=True):
        ws_agg.append(r)

    # Stats
    ws_stats = wb.create_sheet("Mean_Median_Statistics")
    for r in dataframe_to_rows(stats_df, index=False, header=True):
        ws_stats.append(r)

    # Raw Data (use cleaned version)
    ws_raw = wb.create_sheet("Raw_Data")
    for r in dataframe_to_rows(raw_df_cleaned, index=False, header=True):
        ws_raw.append(r)

    # LMM Results: Summaries in one sheet
    ws_lmm = wb.create_sheet("LMM_Summaries")
    ws_lmm.append(["Sentiment", "Measure", "LMM_Summary"])
    for sentiment in lmm_results:
        for measure_type in lmm_results[sentiment]:
            summary = lmm_results[sentiment][measure_type]['LMM_Summary']
            ws_lmm.append([sentiment, measure_type, summary])

    # Plots: Embed images
    for sentiment in CATEGORIES:
        quote_plot_path = os.path.join(plots_dir, f"quote_{sentiment}.png")
        if os.path.exists(quote_plot_path):
            stitle = f"Quote_{sentiment[:28]}"
            w = wb.create_sheet(stitle)
            try:
                img = ExcelImage(quote_plot_path)
                img.anchor = 'A1'
                w.add_image(img)
            except Exception as e:
                logging.error(f"Error embedding image {quote_plot_path} in combined workbook: {e}")

        fulltext_plot_path = os.path.join(plots_dir, f"fulltext_{sentiment}.png")
        if os.path.exists(fulltext_plot_path):
            stitle = f"Fulltext_{sentiment[:25]}"
            w = wb.create_sheet(stitle)
            try:
                img = ExcelImage(fulltext_plot_path)
                img.anchor = 'A1'
                w.add_image(img)
            except Exception as e:
                logging.error(f"Error embedding image {fulltext_plot_path} in combined workbook: {e}")

    wb.save(OUTPUT_EXCEL_COMBINED)
    logging.info(f"All results compiled into single combined Excel file '{OUTPUT_EXCEL_COMBINED}'.")


def check_collinearity(df, predictor_columns):
    if not predictor_columns:
        print("No predictor columns specified for collinearity check.")
        logging.info("No predictor columns specified for collinearity check.")
        return

    zero_var_cols = [col for col in predictor_columns if df[col].nunique() <= 1]
    if zero_var_cols:
        print(f"Predictors with zero variance: {zero_var_cols}")
        logging.info(f"Predictors with zero variance: {zero_var_cols}")

    if len(predictor_columns) > 1:
        corr = df[predictor_columns].corr().abs()
        high_corr_pairs = [(i,j) for i in corr.columns for j in corr.columns if i<j and corr.loc[i,j]>0.9]
        if high_corr_pairs:
            print("Highly collinear predictor pairs found:")
            for pair in high_corr_pairs:
                print(pair, corr.loc[pair[0], pair[1]])
            logging.info(f"Highly collinear predictor pairs: {high_corr_pairs}")


def main():
    setup_logging()
    print("Loading data from JSONL file...")
    try:
        df = load_jsonl(INPUT_JSONL_FILE)
        print(f"Total articles loaded: {len(df)}")
        logging.info(f"Total articles loaded: {len(df)}")
    except Exception as e:
        print(f"Error loading data: {e}")
        logging.error(f"Error loading data: {e}")
        return

    if df.empty:
        print("The input file is empty. Exiting.")
        logging.info("The input file is empty. Exiting.")
        return

    print("Mapping media outlets to categories...")
    try:
        df = map_media_outlet_to_category(df, MEDIA_CATEGORIES)
        print("Media outlets mapped to categories successfully.\n")
        logging.info("Media outlets mapped to categories successfully.")
    except Exception as e:
        print(f"Error mapping media outlets to categories: {e}")
        logging.error(f"Error mapping media outlets to categories: {e}")
        return

    print("Aggregating sentiment/emotion scores per media category...")
    try:
        aggregated_df = aggregate_sentiment_scores(df, CATEGORIES)
        aggregated_df = calculate_averages(aggregated_df)
        print("Aggregation of sentiment/emotion scores completed successfully.\n")
        logging.info("Aggregation of sentiment/emotion scores completed successfully.")
    except Exception as e:
        print(f"Error aggregating sentiment/emotion scores: {e}")
        logging.error(f"Error aggregating sentiment/emotion scores: {e}")
        return

    print("Calculating mean and median statistics...")
    try:
        stats_df = calculate_mean_median(aggregated_df)
        print("Mean and median statistics calculated successfully.\n")
        logging.info("Mean and median statistics calculated successfully.")
    except Exception as e:
        print(f"Error calculating statistics: {e}")
        logging.error(f"Error calculating statistics: {e}")
        return

    print("Saving aggregated scores to CSV files...")
    try:
        save_aggregated_scores_to_csv(aggregated_df, CSV_OUTPUT_DIR)
        print("Aggregated scores saved to CSV files successfully.\n")
        logging.info("Aggregated scores saved to CSV files successfully.")
    except Exception as e:
        print(f"Error saving aggregated scores to CSV: {e}")
        logging.error(f"Error saving aggregated scores to CSV: {e}")
        return

    print("Generating plots for statistics...")
    try:
        plot_statistics(aggregated_df, OUTPUT_DIR)
        print("Plots generated successfully.\n")
        logging.info("Plots generated successfully.")
    except Exception as e:
        print(f"Error generating plots: {e}")
        logging.error(f"Error generating plots: {e}")
        return

    # Filter out outlets with <2 obs
    print("Filtering outlets with insufficient observations...")
    outlet_counts = df['media_outlet'].value_counts()
    outlets_to_remove = outlet_counts[outlet_counts < 2].index.tolist()
    if outlets_to_remove:
        print("Filtering out outlets with fewer than 2 observations:")
        logging.info("Filtering out outlets with fewer than 2 observations:")
        for outlet in outlets_to_remove:
            obs_count = outlet_counts[outlet]
            print(f" - Removing {outlet}: {obs_count} observation(s)")
            logging.info(f" - Removing {outlet}: {obs_count} observation(s)")
        initial_count = len(df)
        df = df[~df['media_outlet'].isin(outlets_to_remove)]
        final_count = len(df)
        removed_count = initial_count - final_count
        print(f"Filtered out {removed_count} total observations from {len(outlets_to_remove)} outlet(s).")
        logging.info(f"Filtered out {removed_count} total observations from {len(outlets_to_remove)} outlet(s).")
    else:
        print("No outlets with fewer than 2 observations found.")
        logging.info("No outlets with fewer than 2 observations found.")

    # Optional predictor columns
    predictor_columns = []  # If you have predictors, list them here
    check_collinearity(df, predictor_columns)

    print("Fitting linear mixed models and running post-hoc tests...")
    try:
        lmm_results = run_lmm_analyses(df)
        print("LMM analysis and post-hoc tests completed successfully.\n")
        logging.info("LMM analysis and post-hoc tests completed successfully.")
    except Exception as e:
        print(f"Error in LMM analysis: {e}")
        logging.error(f"Error in LMM analysis: {e}")
        return

    print("Compiling statistics, raw data, LMM results, and plots into multiple Excel files...")
    try:
        compile_results_into_multiple_workbooks(aggregated_df, stats_df, df, lmm_results, OUTPUT_DIR)
        logging.info("All results compiled into multiple Excel files successfully.")
    except Exception as e:
        print(f"Error compiling statistics and plots into separate Excel files: {e}")
        logging.error(f"Error compiling statistics and plots into separate Excel files: {e}")
        return

    print("Now compiling everything into a single combined Excel workbook as well...")
    try:
        compile_into_single_combined_workbook(aggregated_df, stats_df, df, lmm_results, OUTPUT_DIR)
        print(f"All results compiled into a single combined Excel file: {OUTPUT_EXCEL_COMBINED}")
        logging.info(f"All results compiled into a single combined Excel file: {OUTPUT_EXCEL_COMBINED}")
    except Exception as e:
        print(f"Error compiling single combined Excel file: {e}")
        logging.error(f"Error compiling single combined Excel file: {e}")

    print("Analysis completed successfully.")
    logging.info("Analysis completed successfully.")


if __name__ == "__main__":
    main()
