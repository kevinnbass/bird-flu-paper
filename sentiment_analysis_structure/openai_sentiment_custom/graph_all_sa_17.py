import json
import pandas as pd
import matplotlib.pyplot as plt
import seaborn as sns
import os
import logging
from tqdm import tqdm
from openpyxl.drawing.image import Image as ExcelImage
import re
import statsmodels.api as sm
from statsmodels.formula.api import mixedlm
from statsmodels.stats.multicomp import pairwise_tukeyhsd
import sys
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import numpy as np

# ------------------------------ #
#        Configuration           #
# ------------------------------ #

INPUT_JSONL_FILE = 'processed_all_articles_with_fulltext_sentiment_analysis.jsonl'
OUTPUT_DIR = 'graphs_analysis'
os.makedirs(OUTPUT_DIR, exist_ok=True)

CSV_OUTPUT_DIR = 'csv_raw_scores'
os.makedirs(CSV_OUTPUT_DIR, exist_ok=True)

OUTPUT_EXCEL_MAIN = 'analysis_main.xlsx'
OUTPUT_EXCEL_RAW = 'analysis_raw.xlsx'
OUTPUT_EXCEL_LMM = 'analysis_lmm.xlsx'
OUTPUT_EXCEL_PLOTS = 'analysis_plots.xlsx'
OUTPUT_EXCEL_COMBINED = 'analysis_combined.xlsx'

CATEGORIES = [
    'joy', 'sadness', 'anger', 'fear',
    'surprise', 'disgust', 'trust', 'anticipation',
    'negative_sentiment', 'positive_sentiment'
]

MEDIA_CATEGORIES = {
    'Scientific': ['nature', 'sciam', 'stat', 'newscientist'],
    'Left': ['theatlantic', 'the daily beast', 'the intercept', 'mother jones', 'msnbc', 'slate', 'vox', 'huffpost'],
    'Lean Left': ['ap', 'axios', 'cnn', 'guardian', 'business insider', 'nbcnews', 'npr', 'nytimes', 'politico', 'propublica', 'wapo', 'usa today'],
    'Center': ['reuters', 'marketwatch', 'financial times', 'newsweek', 'forbes'],
    'Lean Right': ['thedispatch', 'epochtimes', 'foxbusiness', 'wsj', 'national review', 'washtimes'],
    'Right': ['breitbart', 'theblaze', 'daily mail', 'dailywire', 'foxnews', 'nypost', 'newsmax'],
}


def setup_logging(log_file='analysis.log'):
    logging.basicConfig(
        filename=log_file,
        level=logging.DEBUG,
        format='%(asctime)s - %(levelname)s - %(message)s'
    )
    logging.info("Logging initialized.")


def load_jsonl(jsonl_file):
    records = []
    with open(jsonl_file, 'r', encoding='utf-8') as f:
        for line in tqdm(f, desc="Loading JSONL data"):
            try:
                record = json.loads(line)
                records.append(record)
            except json.JSONDecodeError as e:
                logging.error(f"JSON decoding error: {e}")
    df = pd.DataFrame(records)
    return df


def map_media_outlet_to_category(df, media_categories):
    outlet_to_category = {}
    for category, outlets in media_categories.items():
        for outlet in outlets:
            outlet_to_category[outlet.lower().strip()] = category

    if 'media_outlet' not in df.columns:
        logging.error("'media_outlet' column not found in DataFrame.")
        raise KeyError("'media_outlet' column not found in DataFrame.")

    df['media_outlet_clean'] = df['media_outlet'].str.lower().str.strip()
    df['media_category'] = df['media_outlet_clean'].map(outlet_to_category)
    df['media_category'] = df['media_category'].fillna('Other')
    unmapped_outlets = df[df['media_category'] == 'Other']['media_outlet'].unique()
    if len(unmapped_outlets) > 0:
        logging.warning(f"Unmapped media outlets found: {unmapped_outlets}")
        print(f"Warning: The following media outlets were not mapped and categorized as 'Other': {unmapped_outlets}")
    return df


def aggregate_sentiment_scores(df, sentiment_categories):
    aggregation = []
    for media_category in MEDIA_CATEGORIES.keys():
        mask = (df['media_category'] == media_category)
        subset = df[mask]

        for sentiment in sentiment_categories:
            pattern = f"^{re.escape(sentiment)}_\\d+$"
            matched_cols = [col for col in df.columns if re.match(pattern, col)]

            if matched_cols:
                quotation_sum = subset[matched_cols].clip(lower=0).sum(skipna=True).sum()
                quotation_count = subset[matched_cols].clip(lower=0).count().sum()
            else:
                quotation_sum = 0
                quotation_count = 0
                logging.warning(f"No quotation-based fields found for '{sentiment}' in '{media_category}'.")

            fulltext_col = f"{sentiment}_fulltext"
            if fulltext_col in df.columns:
                fulltext_sum = subset[fulltext_col].clip(lower=0).sum(skipna=True)
                fulltext_count = subset[fulltext_col].clip(lower=0).count()
            else:
                fulltext_sum = 0
                fulltext_count = 0
                logging.warning(f"Fulltext field '{fulltext_col}' not found in DataFrame.")

            aggregation.append({
                'Media Category': media_category,
                'Sentiment/Emotion': sentiment,
                'Quotation_Sum': quotation_sum,
                'Quotation_Count': quotation_count,
                'Fulltext_Sum': fulltext_sum,
                'Fulltext_Count': fulltext_count
            })

    aggregated_df = pd.DataFrame(aggregation)
    return aggregated_df


def calculate_averages(aggregated_df):
    aggregated_df['Quotation_Average'] = aggregated_df.apply(
        lambda row: row['Quotation_Sum'] / row['Quotation_Count'] if row['Quotation_Count'] > 0 else None, axis=1
    )
    aggregated_df['Fulltext_Average'] = aggregated_df.apply(
        lambda row: row['Fulltext_Sum'] / row['Fulltext_Count'] if row['Fulltext_Count'] > 0 else None, axis=1
    )
    return aggregated_df


def calculate_mean_median(aggregated_df):
    stats = []
    for sentiment in CATEGORIES:
        sentiment_data = aggregated_df[aggregated_df['Sentiment/Emotion'] == sentiment]

        quotation_avg = sentiment_data['Quotation_Average'].dropna()
        mean_quotation = quotation_avg.mean() if not quotation_avg.empty else None
        median_quotation = quotation_avg.median() if not quotation_avg.empty else None

        fulltext_avg = sentiment_data['Fulltext_Average'].dropna()
        mean_fulltext = fulltext_avg.mean() if not fulltext_avg.empty else None
        median_fulltext = fulltext_avg.median() if not fulltext_avg.empty else None

        stats.append({
            'Sentiment/Emotion': sentiment,
            'Mean_Quotation_Average': mean_quotation,
            'Median_Quotation_Average': median_quotation,
            'Mean_Fulltext_Average': mean_fulltext,
            'Median_Fulltext_Average': median_fulltext
        })

    stats_df = pd.DataFrame(stats)
    return stats_df


def save_aggregated_scores_to_csv(aggregated_df, csv_output_dir):
    csv_file = os.path.join(csv_output_dir, "aggregated_sentiment_emotion_scores.csv")
    try:
        aggregated_df.to_csv(csv_file, index=False)
        print(f"Aggregated sentiment/emotion scores saved to '{csv_file}'.")
        logging.info(f"Aggregated sentiment/emotion scores saved to '{csv_file}'.")
    except Exception as e:
        print(f"Error saving aggregated scores to CSV: {e}")
        logging.error(f"Error saving aggregated scores to CSV: {e}")


def plot_statistics(aggregated_df, output_dir):
    sns.set_style("whitegrid")
    for sentiment in CATEGORIES:
        sentiment_data = aggregated_df[aggregated_df['Sentiment/Emotion'] == sentiment]

        # Quotation-Based Plot
        plt.figure(figsize=(10, 6))
        sns.barplot(
            x='Media Category',
            y='Quotation_Average',
            data=sentiment_data,
            color='steelblue'
        )
        plt.title(f"Mean Quotation-Based '{sentiment.capitalize()}' Scores Across Media Categories", fontsize=14)
        plt.xlabel('Media Category', fontsize=12)
        plt.ylabel('Mean Quotation-Based Average Score', fontsize=12)
        plt.xticks(rotation=45, ha='right')
        plt.tight_layout()
        quote_plot_filename = f"quote_{sentiment}.png"
        quote_plot_path = os.path.join(output_dir, quote_plot_filename)
        try:
            plt.savefig(quote_plot_path)
            plt.close()
            print(f"Quotation-Based '{sentiment}' scores plot saved to '{quote_plot_path}'.")
            logging.info(f"Quotation-Based '{sentiment}' scores plot saved to '{quote_plot_path}'.")
        except Exception as e:
            print(f"Error saving quotation-based '{sentiment}' scores plot: {e}")
            logging.error(f"Error saving quotation-based '{sentiment}' scores plot: {e}")

        # Fulltext-Based Plot
        plt.figure(figsize=(10, 6))
        sns.barplot(
            x='Media Category',
            y='Fulltext_Average',
            data=sentiment_data,
            color='darkorange'
        )
        plt.title(f"Mean Fulltext-Based '{sentiment.capitalize()}' Scores Across Media Categories", fontsize=14)
        plt.xlabel('Media Category', fontsize=12)
        plt.ylabel('Mean Fulltext-Based Average Score', fontsize=12)
        plt.xticks(rotation=45, ha='right')
        plt.tight_layout()
        fulltext_plot_filename = f"fulltext_{sentiment}.png"
        fulltext_plot_path = os.path.join(output_dir, fulltext_plot_filename)
        try:
            plt.savefig(fulltext_plot_path)
            plt.close()
            print(f"Fulltext-Based '{sentiment}' scores plot saved to '{fulltext_plot_path}'.")
            logging.info(f"Fulltext-Based '{sentiment}' scores plot saved to '{fulltext_plot_path}'.")
        except Exception as e:
            print(f"Error saving fulltext-based '{sentiment}' scores plot: {e}")
            logging.error(f"Error saving fulltext-based '{sentiment}' scores plot: {e}")


def fit_lmm_and_posthoc(df, sentiment, measure_type='Quotation'):
    logger = logging.getLogger(__name__)

    if measure_type == 'Quotation':
        pattern = f"^{re.escape(sentiment)}_\\d+$"
        matched_cols = [col for col in df.columns if re.match(pattern, col)]
        if not matched_cols:
            logger.debug(f"No matched quotation-based columns for sentiment='{sentiment}'. Skipping.")
            return None
        df = df.copy()
        df[f'{sentiment}_quotation_mean'] = df[matched_cols].clip(lower=0).mean(axis=1)
        score_col = f'{sentiment}_quotation_mean'
    else:
        fulltext_col = f"{sentiment}_fulltext"
        if fulltext_col not in df.columns:
            logger.debug(f"No fulltext column found for sentiment='{sentiment}'. Skipping.")
            return None
        df = df.copy()
        df[f'{sentiment}_fulltext_clipped'] = df[fulltext_col].clip(lower=0)
        score_col = f'{sentiment}_fulltext_clipped'

    model_df = df.dropna(subset=[score_col, 'media_category', 'media_outlet']).copy()

    # Check variation per outlet:
    outlet_counts = model_df['media_outlet'].value_counts()
    outlets_with_variation_issues = []
    for outlet in outlet_counts.index:
        outlet_data = model_df.loc[model_df['media_outlet'] == outlet, score_col]
        if outlet_data.nunique() < 2:
            outlets_with_variation_issues.append((outlet, outlet_data.nunique()))

    print("Checking each media outlet for sufficient variation in the selected score column...")
    logging.info("Checking each media outlet for sufficient variation in the selected score column...")
    if outlets_with_variation_issues:
        for outlet, uniq_count in outlets_with_variation_issues:
            print(f"Outlet '{outlet}' may have insufficient variation (unique values: {uniq_count}).")
            logging.info(f"Outlet '{outlet}' may have insufficient variation (unique values: {uniq_count}).")
            # Print the actual values from that outlet
            vals = model_df.loc[model_df['media_outlet'] == outlet, score_col].unique()
            print(f"[DEBUG] Outlet '{outlet}' unique values: {vals}")
    else:
        print("All outlets show some variation in the response measure.")
        logging.info("All outlets show some variation in the response measure.")

    model_df.loc[:, 'media_category'] = pd.Categorical(model_df['media_category'])

    try:
        md = mixedlm(f"{score_col} ~ media_category", data=model_df, groups=model_df["media_outlet"])
        mdf = md.fit(reml=True, method='lbfgs')
    except Exception as e:
        logger.warning(f"Model failed to converge for {sentiment}-{measure_type}: {e}")
        return None

    if not mdf.converged:
        logger.warning(f"LMM did not converge for {sentiment}-{measure_type}.")

    summ_text = mdf.summary().as_text()
    if "Random effects covariance is singular" in summ_text:
        logger.warning("Random effects covariance singular message found in LMM summary.")
        print("\n[INFO] Random effects covariance is singular. This often happens when certain outlets have no or very little variation.\n"
              "You may consider removing outlets that have only one unique value or extremely low variance.\n")
        # Print per-outlet stats:
        per_outlet_stats = model_df.groupby('media_outlet')[score_col].agg(['count','mean','std','min','max'])
        print("[DEBUG] Per-outlet stats for the response measure:")
        print(per_outlet_stats.to_string())
        no_var_outlets = per_outlet_stats[per_outlet_stats['std'].fillna(0) == 0].index.tolist()
        if no_var_outlets:
            print("\n[INFO] These outlets have zero standard deviation:")
            for o in no_var_outlets:
                print(f" - {o}")
            print("Consider removing these outlets from the dataset or merging them if they provide no variation.\n")

        low_var_threshold = 1e-6
        very_low_var_outlets = per_outlet_stats[(per_outlet_stats['std'].fillna(0) < low_var_threshold) & (per_outlet_stats['std'].fillna(0) > 0)]
        if not very_low_var_outlets.empty:
            print("[INFO] Outlets with very low std (near zero):")
            for o in very_low_var_outlets.index:
                print(f" - {o}: std={very_low_var_outlets.loc[o, 'std']}")
            print("These outlets may also contribute to the singularity issue.\n")

        print("If removing problematic outlets is not desirable, consider simplifying the random effects structure.\n")

    cat_means = model_df.groupby('media_category', observed=True)[score_col].mean().reset_index()
    if cat_means['media_category'].nunique() < 2:
        return {
            'LMM_Summary': mdf.summary().as_text(),
            'PostHoc': pd.DataFrame()
        }

    try:
        tukey = pairwise_tukeyhsd(cat_means[score_col], cat_means['media_category'], alpha=0.05)
        posthoc_df = pd.DataFrame(data=tukey._results_table.data[1:], columns=tukey._results_table.data[0])
    except Exception as e:
        logger.warning(f"Post-hoc failed for {sentiment}-{measure_type}: {e}")
        posthoc_df = pd.DataFrame()

    return {
        'LMM_Summary': mdf.summary().as_text(),
        'PostHoc': posthoc_df
    }


def run_lmm_analyses(df):
    all_results = {}
    for sentiment in CATEGORIES:
        all_results[sentiment] = {}
        for measure_type in ['Quotation', 'Fulltext']:
            res = fit_lmm_and_posthoc(df, sentiment, measure_type=measure_type)
            if res is not None:
                all_results[sentiment][measure_type] = res
    return all_results


def compile_results_into_multiple_workbooks(aggregated_df, stats_df, raw_df, lmm_results, plots_dir):
    # MAIN WORKBOOK
    with pd.ExcelWriter(OUTPUT_EXCEL_MAIN, engine='openpyxl') as writer:
        aggregated_df.to_excel(writer, sheet_name='Aggregated_Scores', index=False)
        stats_df.to_excel(writer, sheet_name='Mean_Median_Statistics', index=False)

    # RAW WORKBOOK
    with pd.ExcelWriter(OUTPUT_EXCEL_RAW, engine='openpyxl') as writer:
        raw_df.to_excel(writer, sheet_name='Raw_Data', index=False)
        for sentiment in CATEGORIES:
            sentiment_cols = [c for c in raw_df.columns if c.startswith(sentiment+'_')]
            sentiment_df = raw_df[['media_category', 'media_outlet'] + sentiment_cols].copy()
            sheet_name = f"Raw_{sentiment[:29]}"
            sentiment_df.to_excel(writer, sheet_name=sheet_name, index=False)

    # LMM WORKBOOK
    with pd.ExcelWriter(OUTPUT_EXCEL_LMM, engine='openpyxl') as writer:
        summary_rows = []
        for sentiment in lmm_results:
            for measure_type in lmm_results[sentiment]:
                sheet_name = f"LMM_{sentiment[:20]}_{measure_type[:8]}"
                lmm_summary = lmm_results[sentiment][measure_type]['LMM_Summary']
                posthoc_df = lmm_results[sentiment][measure_type]['PostHoc']

                summary_df = pd.DataFrame({'LMM_Summary': lmm_summary.split('\n')})
                summary_df.to_excel(writer, sheet_name=sheet_name, index=False)
                startrow = len(summary_df) + 2
                posthoc_df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=startrow)

                summary_rows.append({
                    'Sentiment': sentiment,
                    'Measure': measure_type,
                    'SheetName': sheet_name
                })

        summary_index_df = pd.DataFrame(summary_rows)
        summary_index_df.to_excel(writer, sheet_name='LMM_Results_Index', index=False)

    # PLOTS WORKBOOK
    wb = Workbook()
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])

    for sentiment in CATEGORIES:
        quote_plot_filename = f"quote_{sentiment}.png"
        quote_plot_path = os.path.join(plots_dir, quote_plot_filename)
        if os.path.exists(quote_plot_path):
            sheet_title = f"Quote_{sentiment[:28]}"
            worksheet = wb.create_sheet(title=sheet_title)
            try:
                img = ExcelImage(quote_plot_path)
                img.anchor = 'A1'
                worksheet.add_image(img)
            except Exception as e:
                logging.error(f"Error embedding image '{quote_plot_path}' into Excel: {e}")

        fulltext_plot_filename = f"fulltext_{sentiment}.png"
        fulltext_plot_path = os.path.join(plots_dir, fulltext_plot_filename)
        if os.path.exists(fulltext_plot_path):
            sheet_title = f"Fulltext_{sentiment[:25]}"
            worksheet = wb.create_sheet(title=sheet_title)
            try:
                img = ExcelImage(fulltext_plot_path)
                img.anchor = 'A1'
                worksheet.add_image(img)
            except Exception as e:
                logging.error(f"Error embedding image '{fulltext_plot_path}' into Excel: {e}")

    wb.save(OUTPUT_EXCEL_PLOTS)

    print(f"All statistics, raw data, LMM results, and plots have been compiled into multiple Excel files:\n"
          f" - {OUTPUT_EXCEL_MAIN}\n"
          f" - {OUTPUT_EXCEL_RAW}\n"
          f" - {OUTPUT_EXCEL_LMM}\n"
          f" - {OUTPUT_EXCEL_PLOTS}")


def compile_into_single_combined_workbook(aggregated_df, stats_df, raw_df, lmm_results, plots_dir):
    raw_df_cleaned = raw_df.copy()
    raw_df_cleaned = raw_df_cleaned.applymap(lambda x: ", ".join(x) if isinstance(x, list) else x)

    wb = Workbook()
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])

    ws_agg = wb.create_sheet("Aggregated_Scores")
    for r in dataframe_to_rows(aggregated_df, index=False, header=True):
        ws_agg.append(r)

    ws_stats = wb.create_sheet("Mean_Median_Statistics")
    for r in dataframe_to_rows(stats_df, index=False, header=True):
        ws_stats.append(r)

    ws_raw = wb.create_sheet("Raw_Data")
    for r in dataframe_to_rows(raw_df_cleaned, index=False, header=True):
        ws_raw.append(r)

    ws_lmm = wb.create_sheet("LMM_Summaries")
    ws_lmm.append(["Sentiment", "Measure", "LMM_Summary"])
    for sentiment in lmm_results:
        for measure_type in lmm_results[sentiment]:
            summary = lmm_results[sentiment][measure_type]['LMM_Summary']
            ws_lmm.append([sentiment, measure_type, summary])

    for sentiment in CATEGORIES:
        quote_plot_path = os.path.join(plots_dir, f"quote_{sentiment}.png")
        if os.path.exists(quote_plot_path):
            stitle = f"Quote_{sentiment[:28]}"
            w = wb.create_sheet(stitle)
            try:
                img = ExcelImage(quote_plot_path)
                img.anchor = 'A1'
                w.add_image(img)
            except Exception as e:
                logging.error(f"Error embedding image {quote_plot_path} in combined workbook: {e}")

        fulltext_plot_path = os.path.join(plots_dir, f"fulltext_{sentiment}.png")
        if os.path.exists(fulltext_plot_path):
            stitle = f"Fulltext_{sentiment[:25]}"
            w = wb.create_sheet(stitle)
            try:
                img = ExcelImage(fulltext_plot_path)
                img.anchor = 'A1'
                w.add_image(img)
            except Exception as e:
                logging.error(f"Error embedding image {fulltext_plot_path} in combined workbook: {e}")

    wb.save(OUTPUT_EXCEL_COMBINED)


def check_collinearity(df, predictor_columns):
    if not predictor_columns:
        print("No predictor columns specified for collinearity check.")
        return
    zero_var_cols = [col for col in predictor_columns if df[col].nunique() <= 1]
    if zero_var_cols:
        print(f"Predictors with zero variance: {zero_var_cols}")


def main():
    setup_logging()
    print("Loading data from JSONL file...")
    df = load_jsonl(INPUT_JSONL_FILE)
    print(f"Total articles loaded: {len(df)}")

    if df.empty:
        print("The input file is empty. Exiting.")
        return

    print("Mapping media outlets to categories...")
    df = map_media_outlet_to_category(df, MEDIA_CATEGORIES)
    print("Media outlets mapped to categories successfully.\n")

    print("Aggregating sentiment/emotion scores per media category...")
    aggregated_df = aggregate_sentiment_scores(df, CATEGORIES)
    aggregated_df = calculate_averages(aggregated_df)
    print("Aggregation of sentiment/emotion scores completed successfully.\n")

    print("Calculating mean and median statistics...")
    stats_df = calculate_mean_median(aggregated_df)
    print("Mean and median statistics calculated successfully.\n")

    print("Saving aggregated scores to CSV files...")
    save_aggregated_scores_to_csv(aggregated_df, CSV_OUTPUT_DIR)
    print("Aggregated scores saved to CSV files successfully.\n")

    print("Generating plots for statistics...")
    plot_statistics(aggregated_df, OUTPUT_DIR)
    print("Plots generated successfully.\n")

    print("Filtering outlets with insufficient observations...")
    outlet_counts = df['media_outlet'].value_counts()
    outlets_to_remove = outlet_counts[outlet_counts < 2].index.tolist()
    if outlets_to_remove:
        print("Filtering out outlets with fewer than 2 observations:")
        for outlet in outlets_to_remove:
            obs_count = outlet_counts[outlet]
            print(f" - Removing {outlet}: {obs_count} observation(s)")
        df = df[~df['media_outlet'].isin(outlets_to_remove)]
        removed_count = len(outlets_to_remove)
        print(f"Filtered out {removed_count} outlets.")
    else:
        print("No outlets with fewer than 2 observations found.")

    predictor_columns = []
    check_collinearity(df, predictor_columns)

    print("Fitting linear mixed models and running post-hoc tests...")
    lmm_results = run_lmm_analyses(df)
    print("LMM analysis and post-hoc tests completed successfully.\n")

    print("Compiling statistics, raw data, LMM results, and plots into multiple Excel files...")
    compile_results_into_multiple_workbooks(aggregated_df, stats_df, df, lmm_results, OUTPUT_DIR)

    print("Now compiling everything into a single combined Excel workbook as well...")
    compile_into_single_combined_workbook(aggregated_df, stats_df, df, lmm_results, OUTPUT_DIR)
    print(f"All results compiled into a single combined Excel file: {OUTPUT_EXCEL_COMBINED}")

    print("Analysis completed successfully.")


if __name__ == "__main__":
    main()
