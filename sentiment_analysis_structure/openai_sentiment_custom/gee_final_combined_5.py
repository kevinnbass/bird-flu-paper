#!/usr/bin/env python3
# gee_w_validation_10_fixed_zscatter.py
"""
Script that:
  1) Loads data, does correlation analyses (Quotation vs Fulltext)
  2) Aggregates stats
  3) Finds best QIC combination (multi-family)
  4) For each best QIC model:
     - Refit & produce GEE summary
     - Advanced checks => diagnostics, cross-validation, etc.
     - Pairwise => "DiffIDs"
  5) Writes summary, pairwise, & advanced checks into analysis_gee.xlsx
  6) Also saves analysis_main, analysis_raw, analysis_plots, analysis_combined
  7) Replicates the entire pipeline for a subset where "high_rate_2":"yes".

Now includes a fix in `cross_validation_gee()` so that "Scientific" or any
other category that appears only in test set does not cause a KeyError.
"""

import json
import os
import re
import sys
import warnings
import logging
import math
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import seaborn as sns
from tqdm import tqdm

import statsmodels.api as sm
from statsmodels.genmod.families import (
    Poisson, NegativeBinomial,
    Gaussian, Gamma, InverseGaussian
)
import statsmodels.genmod.families.links as ln
from statsmodels.genmod.generalized_estimating_equations import GEE
from statsmodels.genmod.cov_struct import Independence, Exchangeable
from statsmodels.stats.multitest import multipletests
from scipy.stats import pearsonr, norm

from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill

# ----------------------------------------------------------------------------
# Configuration
# ----------------------------------------------------------------------------

INPUT_JSONL_FILE = "processed_all_articles_fixed_4.jsonl"  # or your file
BASE_OUTPUT_DIR = "graphs_analysis"
BASE_CSV_OUTPUT_DIR = "csv_raw_scores"

OUTPUT_EXCEL_MAIN = "analysis_main.xlsx"
OUTPUT_EXCEL_RAW = "analysis_raw.xlsx"
OUTPUT_EXCEL_GEE = "analysis_gee.xlsx"
OUTPUT_EXCEL_PLOTS = "analysis_plots.xlsx"
OUTPUT_EXCEL_COMBINED = "analysis_combined.xlsx"
OUTPUT_EXCEL_QFTC = "quotation_fulltext_correlation.xlsx"

# For the "subset" run (high_rate_2 == "yes"), we add a "_Yes" suffix:
OUTPUT_EXCEL_MAIN_SUB = "analysis_main_Yes.xlsx"
OUTPUT_EXCEL_RAW_SUB = "analysis_raw_Yes.xlsx"
OUTPUT_EXCEL_GEE_SUB = "analysis_gee_Yes.xlsx"
OUTPUT_EXCEL_PLOTS_SUB = "analysis_plots_Yes.xlsx"
OUTPUT_EXCEL_COMBINED_SUB = "analysis_combined_Yes.xlsx"
OUTPUT_EXCEL_QFTC_SUB = "quotation_fulltext_correlation_Yes.xlsx"

LOG_FILE = "analysis.log"

os.makedirs(BASE_OUTPUT_DIR, exist_ok=True)
os.makedirs(BASE_CSV_OUTPUT_DIR, exist_ok=True)

# We'll do Quotation vs Fulltext for these sentiments
CATEGORIES = [
    "joy","sadness","anger","fear",
    "surprise","disgust","trust","anticipation",
    "negative_sentiment","positive_sentiment"
]

MEDIA_CATEGORIES = {
    "Scientific": ["nature","sciam","stat","newscientist"],
    "Left": ["theatlantic","the daily beast","the intercept","mother jones","msnbc","slate","vox","huffpost"],
    "Lean Left": ["ap","axios","cnn","guardian","business insider","nbcnews","npr","nytimes","politico","propublica","wapo","usa today"],
    "Center": ["reuters","marketwatch","financial times","newsweek","forbes"],
    "Lean Right": ["thedispatch","epochtimes","foxbusiness","wsj","national review","washtimes"],
    "Right": ["breitbart","theblaze","daily mail","dailywire","foxnews","nypost","newsmax"],
}

def setup_logging():
    log_format = "%(asctime)s [%(levelname)s] %(module)s - %(message)s"
    logging.basicConfig(filename=LOG_FILE, level=logging.DEBUG, format=log_format)
    console = logging.StreamHandler(sys.stdout)
    console.setLevel(logging.INFO)
    console.setFormatter(logging.Formatter(log_format))
    logging.getLogger().addHandler(console)
    logging.info("Logging initialized (file+console).")

warnings.filterwarnings(
    "ignore",
    message="QIC values obtained using scale=None are not appropriate for comparing models"
)

###############################################################################
# 1) Load, chunk, basic stats
###############################################################################
def load_jsonl(jsonl_file):
    logging.info(f"Loading JSONL from {jsonl_file}")
    records=[]
    with open(jsonl_file,"r",encoding="utf-8") as f:
        for line in tqdm(f, desc="Loading JSONL data"):
            rec=json.loads(line)
            records.append(rec)
    df=pd.DataFrame(records)
    logging.debug(f"Loaded DataFrame shape={df.shape}")
    return df

def map_media_outlet_to_category(df):
    cat_map={}
    for cat, outls in MEDIA_CATEGORIES.items():
        for o in outls:
            cat_map[o.lower().strip()] = cat

    if "media_outlet" not in df.columns:
        raise KeyError("'media_outlet' not found in data")

    df["media_outlet_clean"] = df["media_outlet"].str.lower().str.strip()
    df["media_category"] = df["media_outlet_clean"].map(cat_map).fillna("Other")

    unmapped = df[df["media_category"]=="Other"]["media_outlet"].unique()
    if len(unmapped)>0:
        logging.warning(f"Unmapped => {unmapped}")
        print(f"Warning: Not mapped => {unmapped}")
    return df

def chunk_and_save(df, chunk_size=20000, prefix=""):
    logging.info(f"Chunking DataFrame => len={len(df)}, chunk_size={chunk_size}, prefix={prefix}")
    for i in range(0,len(df),chunk_size):
        part=df.iloc[i:i+chunk_size]
        suffix = f"_{prefix}" if prefix else ""
        out_csv=os.path.join(BASE_CSV_OUTPUT_DIR,f"raw_data_part_{(i//chunk_size)+1}{suffix}.csv")
        part.to_csv(out_csv,index=False)
        print(f"Saved chunk {(i//chunk_size)+1} to {out_csv}")

def print_basic_stats(df, prefix=""):
    logging.info(f"Basic stats => total articles = {len(df)} for prefix={prefix}")
    print(f"\nSummary Statistics ({prefix or 'AllData'}):")
    print("Total number of articles:", len(df))
    if "media_outlet_clean" in df.columns:
        vc=df["media_outlet_clean"].value_counts()
        print("\nArticles per outlet:")
        print(vc)
    if "media_category" in df.columns:
        vc2=df["media_category"].value_counts()
        print("\nArticles per category:")
        print(vc2)
    print()

###############################################################################
# 2) Quotation vs. Fulltext correlation
###############################################################################
def analyze_quotation_fulltext_correlation(
    df, 
    output_dir=BASE_OUTPUT_DIR,
    csv_output_dir=BASE_CSV_OUTPUT_DIR,
    output_excel_qftc=OUTPUT_EXCEL_QFTC,
    prefix=""
):
    """
    For each (media_category, sentiment), compute Quotation_Average & Fulltext_Average
    Summarize correlation by sentiment => bar chart
    Also produce a second tab 'CombinedZScatterData'
    in output_excel_qftc (which we may suffix with prefix).
    """
    from scipy.stats import pearsonr
    suffix = f"_{prefix}" if prefix else ""

    logging.info(f"Analyzing Quotation vs Fulltext correlation => prefix={prefix}")
    records=[]
    for cat in df["media_category"].dropna().unique():
        dcat=df[df["media_category"]==cat]
        for s in CATEGORIES:
            pat=rf"^{s}_\d+$"
            matched=[c for c in dcat.columns if re.match(pat,c)]
            if matched:
                clp=dcat[matched].clip(lower=0)
                qsum=clp.sum(skipna=True).sum()
                qcount=clp.count().sum()
                qavg=qsum/qcount if qcount>0 else np.nan
            else:
                qavg=np.nan

            fcol=f"{s}_fulltext"
            if fcol in dcat.columns:
                fv=dcat[fcol].clip(lower=0)
                f_sum=fv.sum(skipna=True)
                f_cnt=fv.count()
                favg=f_sum/f_cnt if f_cnt>0 else np.nan
            else:
                favg=np.nan

            records.append({
                "MediaCategory": cat,
                "Sentiment": s,
                "Quotation_Average": qavg,
                "Fulltext_Average": favg
            })
    agg_df=pd.DataFrame(records)
    cor_results=[]
    all_combo=[]

    for s in CATEGORIES:
        sub=agg_df[(agg_df["Sentiment"]==s)].dropna(subset=["Quotation_Average","Fulltext_Average"])
        if len(sub)>1:
            cor_val,_=pearsonr(sub["Quotation_Average"],sub["Fulltext_Average"])
        else:
            cor_val=np.nan
        cor_results.append({"Sentiment":s,"Correlation":cor_val})
        if not sub.empty:
            cpy=sub.copy()
            cpy["Sentiment"]=s
            all_combo.append(cpy)

    cor_df=pd.DataFrame(cor_results)
    plt.figure(figsize=(8,5))
    sns.barplot(x="Sentiment",y="Correlation",data=cor_df,color="gray")
    plt.title(f"Correlation (Quotation vs. Fulltext) {prefix}")
    plt.xticks(rotation=45,ha="right")
    plt.ylim(-1,1)
    plt.tight_layout()

    cbar=os.path.join(output_dir,f"correlation_quotation_fulltext_bar{suffix}.png")
    try:
        plt.savefig(cbar)
    except:
        pass
    plt.close()

    if all_combo:
        allc=pd.concat(all_combo,ignore_index=True)
        allc["Qmean"]=allc.groupby("Sentiment")["Quotation_Average"].transform("mean")
        allc["Qstd"]=allc.groupby("Sentiment")["Quotation_Average"].transform("std")
        allc["Fmean"]=allc.groupby("Sentiment")["Fulltext_Average"].transform("mean")
        allc["Fstd"]=allc.groupby("Sentiment")["Fulltext_Average"].transform("std")
        allc["Quotation_Z"]=(allc["Quotation_Average"]-allc["Qmean"])/allc["Qstd"]
        allc["Fulltext_Z"]=(allc["Fulltext_Average"]-allc["Fmean"])/allc["Fstd"]

        out_qftc = output_excel_qftc.replace(".xlsx", f"{suffix}.xlsx")
        with pd.ExcelWriter(out_qftc, engine="openpyxl") as writer_xlsx:
            cor_df.to_excel(writer_xlsx,"CorrelationSummary",index=False)
            allc.to_excel(writer_xlsx,"CombinedZScatterData",index=False)

    out_csv=os.path.join(csv_output_dir,f"quotation_fulltext_correlation{suffix}.csv")
    cor_df.to_csv(out_csv,index=False)
    logging.info(f"Correlation data => {out_csv}")

###############################################################################
# 3) Aggregation & Stats
###############################################################################
def aggregate_sentiment_scores(df, sentiments):
    logging.info("Aggregating sentiment/emotion scores by category + sentiment.")
    recs=[]
    for cat in MEDIA_CATEGORIES.keys():
        sub=df[df["media_category"]==cat]
        for s in sentiments:
            pat=rf"^{re.escape(s)}_\d+$"
            matched=[c for c in sub.columns if re.match(pat,c)]
            if matched:
                qsum=sub[matched].clip(lower=0).sum(skipna=True).sum()
                qcount=sub[matched].clip(lower=0).count().sum()
            else:
                qsum,qcount=(0,0)

            fcol=f"{s}_fulltext"
            if fcol in sub.columns:
                fv=sub[fcol].clip(lower=0)
                f_sum=fv.sum(skipna=True)
                f_cnt=fv.count()
            else:
                f_sum,f_cnt=(0,0)

            recs.append({
                "Media Category":cat,
                "Sentiment/Emotion": s,
                "Quotation_Sum": qsum,
                "Quotation_Count": qcount,
                "Fulltext_Sum": f_sum,
                "Fulltext_Count": f_cnt
            })
    return pd.DataFrame(recs)

def calculate_averages(agg_df):
    logging.info("Calculating Quotation_Average & Fulltext_Average.")
    def sdiv(a,b): return a/b if b>0 else None
    agg_df["Quotation_Average"]=agg_df.apply(lambda r: sdiv(r["Quotation_Sum"],r["Quotation_Count"]),axis=1)
    agg_df["Fulltext_Average"]=agg_df.apply(lambda r: sdiv(r["Fulltext_Sum"],r["Fulltext_Count"]),axis=1)
    return agg_df

def calculate_mean_median(agg_df):
    logging.info("Computing global mean/median of Quotation/Fulltext averages.")
    rows=[]
    for s in CATEGORIES:
        sub=agg_df[agg_df["Sentiment/Emotion"]==s]
        qa=sub["Quotation_Average"].dropna()
        fa=sub["Fulltext_Average"].dropna()
        rows.append({
            "Sentiment/Emotion": s,
            "Mean_Quotation_Average": qa.mean() if len(qa)>0 else None,
            "Median_Quotation_Average": qa.median() if len(qa)>0 else None,
            "Mean_Fulltext_Average": fa.mean() if len(fa)>0 else None,
            "Median_Fulltext_Average": fa.median() if len(fa)>0 else None
        })
    return pd.DataFrame(rows)

def save_aggregated_scores_to_csv(agg_df, out_dir=BASE_CSV_OUTPUT_DIR, prefix=""):
    suffix = f"_{prefix}" if prefix else ""
    fn=os.path.join(out_dir,f"aggregated_sentiment_emotion_scores{suffix}.csv")
    agg_df.to_csv(fn,index=False)
    print(f"Aggregated sentiment/emotion scores => {fn}")
    logging.info(f"Aggregated => {fn}")

def plot_statistics(agg_df, out_dir=BASE_OUTPUT_DIR, prefix=""):
    logging.info(f"Plotting bar charts for Quotation/Fulltext across categories => prefix={prefix}")
    sns.set_style("whitegrid")
    suffix = f"_{prefix}" if prefix else ""
    for s in CATEGORIES:
        sub=agg_df[agg_df["Sentiment/Emotion"]==s]

        # Quotation
        plt.figure(figsize=(10,6))
        sns.barplot(x="Media Category", y="Quotation_Average", data=sub, color="steelblue")
        plt.title(f"Mean Quotation '{s.capitalize()}' Scores {prefix}")
        plt.xticks(rotation=45,ha="right")
        plt.tight_layout()
        out1=os.path.join(out_dir,f"quote_{s}{suffix}.png")
        try:
            plt.savefig(out1)
        except:
            pass
        plt.close()

        # Fulltext
        plt.figure(figsize=(10,6))
        sns.barplot(x="Media Category", y="Fulltext_Average", data=sub, color="darkorange")
        plt.title(f"Mean Fulltext '{s.capitalize()}' Scores {prefix}")
        plt.xticks(rotation=45,ha="right")
        plt.tight_layout()
        out2=os.path.join(out_dir,f"fulltext_{s}{suffix}.png")
        try:
            plt.savefig(out2)
        except:
            pass
        plt.close()

###############################################################################
# 4) Scale computations
###############################################################################
def compute_pearson_scale(y, mu, df_resid):
    if df_resid<=0: return np.nan
    y_arr=np.asarray(y)
    mu_arr=np.asarray(mu)
    r=(y_arr-mu_arr)/np.sqrt(mu_arr+1e-9)
    return np.sum(r**2)/df_resid

def compute_deviance_scale(y, mu, df_resid):
    if df_resid<=0: return np.nan
    y_arr=np.asarray(y)
    mu_arr=np.asarray(mu)
    arr=[]
    for obs,lam in zip(y_arr,mu_arr):
        if obs>0 and lam>0:
            arr.append(obs*math.log(obs/lam)-(obs-lam))
        elif obs==0:
            arr.append(-(obs-lam))
        else:
            arr.append(np.nan)
    dev=2*np.nansum(arr)
    return dev/df_resid if df_resid>0 else np.nan

def compute_ub_scale(y, mu, df_resid):
    p=compute_pearson_scale(y, mu, df_resid)
    if np.isnan(p): return p
    return 1.1*p

def compute_bc_scale(y, mu, df_resid):
    d=compute_deviance_scale(y, mu, df_resid)
    if np.isnan(d):
        return d
    return 0.9*d

###############################################################################
# 5) Additional checks & Validation
###############################################################################
def check_residuals_and_correlation(final_fit):
    """
    Convert residuals to numpy array => we can do numeric indexing.
    """
    y_arr=np.asarray(final_fit.model.endog)
    mu_arr=np.asarray(final_fit.fittedvalues)
    pearson_res_arr=(y_arr - mu_arr)/np.sqrt(mu_arr+1e-9)
    pearson_res_arr = np.asarray(pearson_res_arr)

    mean_res=np.mean(pearson_res_arr)
    std_res=np.std(pearson_res_arr)

    clusters=final_fit.model.groups
    clusters_arr=np.asarray(clusters)

    cluster_map={}
    for i in range(len(clusters_arr)):
        cid=clusters_arr[i]
        cluster_map.setdefault(cid,[]).append(pearson_res_arr[i])

    wcorr=[]
    for cid, arr in cluster_map.items():
        arr=np.array(arr)
        if len(arr)>1:
            cmat=np.corrcoef(arr)
            if cmat.ndim==2 and cmat.shape[0]>1:
                n2=len(arr)
                sum_offdiag=np.sum(cmat)-n2
                avg_off=sum_offdiag/(n2*(n2-1)) if n2>1 else 0
                wcorr.append(avg_off)

    avg_corr=np.mean(wcorr) if len(wcorr)>0 else 0.0

    deviance=getattr(final_fit,"pearson_chi2",np.nan)
    dfres=getattr(final_fit,"df_resid",1)
    overdisp=np.nan
    if dfres>0 and not np.isnan(deviance):
        overdisp=deviance/dfres

    assess=""
    if abs(mean_res)>1:
        assess+="Mean Pearson residual>1 => misfit? "
    if std_res>2:
        assess+="Residual std>2 => outliers? "
    if abs(avg_corr)>0.3:
        assess+=f"Within-cluster corr={avg_corr:.2f} => structure suspect. "
    if (not np.isnan(overdisp)) and overdisp>2:
        assess+=f"Overdisp={overdisp:.2f} => consider NB or other approach. "

    if assess=="":
        assess="No major issues from these checks."

    return {
        "mean_pearson_res": mean_res,
        "std_pearson_res": std_res,
        "avg_within_corr": avg_corr,
        "overdisp_ratio": overdisp,
        "assessment": assess
    }

def pseudo_likelihood_check(final_fit):
    fam=final_fit.model.family
    from statsmodels.genmod.families.family import NegativeBinomial
    if not isinstance(fam, sm.genmod.families.family.Poisson):
        return {
            "NB_QIC":None,
            "Poisson_QIC":None,
            "diff_QIC":None,
            "conclusion":"Non-Poisson => skip NB check"
        }
    data=final_fit.model.data.frame
    groups=final_fit.model.groups
    formula=final_fit.model.formula
    cov_struct=final_fit.model.cov_struct
    nb_model=GEE.from_formula(formula, groups=groups, data=data,
                              family=NegativeBinomial(alpha=1.0),
                              cov_struct=cov_struct)
    nb_res=nb_model.fit()
    nb_qic=nb_res.qic()
    if isinstance(nb_qic, tuple):
        nb_qic=nb_qic[0]
    old_qic=final_fit.qic()
    if isinstance(old_qic, tuple):
        old_qic=old_qic[0]

    diff_qic=None
    conclusion=""
    if isinstance(nb_qic,(float,int)) and isinstance(old_qic,(float,int)):
        diff_qic=old_qic-nb_qic
        conclusion="NegBin better" if diff_qic>0 else "No NB improvement"
    else:
        conclusion="Could not compare QIC"

    return {
        "NB_QIC":nb_qic,
        "Poisson_QIC":old_qic,
        "diff_QIC":diff_qic,
        "conclusion":conclusion
    }

def cross_validation_gee(data_with_score, formula, group_col, family, cov_struct, n_folds=3):
    """
    Cross-validation that ensures test set's 'media_category' has the
    same category levels as training => prevents KeyError for new categories
    like 'Scientific' if not in train.
    """
    cluster_ids = data_with_score[group_col].unique()
    np.random.shuffle(cluster_ids)
    folds = np.array_split(cluster_ids, n_folds)
    metrics = []

    for i in range(n_folds):
        testc = set(folds[i])
        train_df = data_with_score[~data_with_score[group_col].isin(testc)].copy()
        test_df = data_with_score[data_with_score[group_col].isin(testc)].copy()

        if len(train_df) < 1 or len(test_df) < 1:
            continue

        # Convert training to categorical => define levels
        train_df["media_category"] = train_df["media_category"].astype("category")
        train_cats = train_df["media_category"].cat.categories

        # Force test set to have the same category levels
        test_df["media_category"] = pd.Categorical(test_df["media_category"], categories=train_cats)

        # Fit GEE on train
        mod = GEE.from_formula(formula, groups=group_col,
                               data=train_df, family=family, cov_struct=cov_struct)
        res = mod.fit()

        # Predict on test
        pred = res.predict(test_df)
        obs = test_df[res.model.endog_names]
        mse = np.mean((obs - pred)**2)
        metrics.append(mse)

    return np.mean(metrics) if len(metrics) > 0 else np.nan

def sensitivity_analysis_correlation(df, formula, group_col):
    """
    Attempt multiple families, multiple covariance structures,
    gather QIC & sample parameters.
    """
    from statsmodels.genmod.families import Poisson, Gaussian
    from statsmodels.genmod.families.family import NegativeBinomial
    alt_families=[Poisson(), NegativeBinomial(alpha=1.0), Gaussian()]
    alt_structs=[Independence(), Exchangeable()]
    results=[]
    for fam in alt_families:
        for covs in alt_structs:
            mod=GEE.from_formula(formula, groups=group_col, data=df,
                                 family=fam, cov_struct=covs)
            try:
                r=mod.fit()
                qic_val=r.qic()
                if isinstance(qic_val, tuple):
                    qic_val=qic_val[0]
                param3=r.params.head(3).round(3).to_dict()
                results.append({
                    "Family":fam.__class__.__name__,
                    "CovStruct":covs.__class__.__name__,
                    "QIC":qic_val,
                    "ParamSample":param3
                })
            except:
                results.append({
                    "Family":fam.__class__.__name__,
                    "CovStruct":covs.__class__.__name__,
                    "QIC":np.nan,
                    "ParamSample":"fail"
                })
    return pd.DataFrame(results)

def bootstrap_gee(df, formula, group_col,
                  B=3, family=Poisson(), cov_struct=Independence()):
    """
    Quick bootstrap approach: sample clusters w/ replacement,
    fit GEE, store parameter estimates.
    """
    cluster_ids=df[group_col].unique()
    param_records=[]
    for _ in range(B):
        sample_ids=np.random.choice(cluster_ids, size=len(cluster_ids), replace=True)
        pieces=[]
        for cid in sample_ids:
            pieces.append(df[df[group_col]==cid])
        boot_df=pd.concat(pieces,ignore_index=True)
        mod=GEE.from_formula(formula, groups=group_col, data=boot_df,
                             family=family, cov_struct=cov_struct)
        res=mod.fit()
        param_records.append(res.params)
    return pd.DataFrame(param_records)

###############################################################################
# 6) Multi-Family enumerations
###############################################################################
def try_all_families_and_scales(df, formula, group_col):
    families = [
        Poisson(),
        NegativeBinomial(alpha=1.0),
        Gaussian(),
        Gamma(link=ln.log()),
        InverseGaussian()
    ]
    cor_structs=[Independence(), Exchangeable()]
    scale_opts=["none","pearson","deviance","ub","bc"]

    best_qic=np.inf
    best_tuple=None
    all_results=[]
    for fam in families:
        fam_name=fam.__class__.__name__
        for cov_obj in cor_structs:
            cov_name=cov_obj.__class__.__name__
            for sc_opt in scale_opts:
                try:
                    model=GEE.from_formula(formula, groups=group_col,
                                           data=df, family=fam, cov_struct=cov_obj)
                    base_res=model.fit(scale=None)
                    y=base_res.model.endog
                    mu=base_res.fittedvalues
                    n=len(y)
                    p=len(base_res.params)
                    dfresid=n-p
                    final_res=base_res
                    if sc_opt!="none" and dfresid>0:
                        val=None
                        if sc_opt=="pearson":
                            val=compute_pearson_scale(y,mu,dfresid)
                        elif sc_opt=="deviance":
                            val=compute_deviance_scale(y,mu,dfresid)
                        elif sc_opt=="ub":
                            val=compute_ub_scale(y,mu,dfresid)
                        elif sc_opt=="bc":
                            val=compute_bc_scale(y,mu,dfresid)
                        if val is not None and not np.isnan(val):
                            final_res=model.fit(scale=val)

                    qic_val=final_res.qic()
                    if isinstance(qic_val, tuple):
                        qic_val=qic_val[0]
                    all_results.append((fam_name,cov_name,sc_opt,qic_val))
                    if qic_val<best_qic:
                        best_qic=qic_val
                        best_tuple=(fam_name,cov_name,sc_opt,qic_val)
                except Exception as e:
                    logging.warning(f"Fit fail => {fam_name}+{cov_name}+{sc_opt} => {e}")
                    all_results.append((fam_name,cov_name,sc_opt,np.nan))
    return best_tuple, all_results

###############################################################################
# 7) Refit + advanced checks
###############################################################################
def refit_best_gee_with_scale(df, sentiment, measure, fam_name, struct, scale_name):
    from statsmodels.genmod.families import Poisson, Gaussian, Gamma, InverseGaussian
    from statsmodels.genmod.families.family import NegativeBinomial

    if measure=="Quotation":
        # average ^<sentiment>_\d+$
        pat=rf"^{re.escape(sentiment)}_\d+$"
        matched=[c for c in df.columns if re.match(pat,c)]
        if not matched:
            return None,None
        d2=df.copy()
        d2["_score_col"]=d2[matched].clip(lower=0).mean(axis=1)
    else:
        # measure="Fulltext"
        fcol=f"{sentiment}_fulltext"
        if fcol not in df.columns:
            return None,None
        d2=df.copy()
        d2["_score_col"]=d2[fcol].clip(lower=0)

    needed=["_score_col","media_outlet_clean","media_category"]
    d2=d2.dropna(subset=needed)
    if len(d2)<2 or d2["media_category"].nunique()<2:
        return None,None

    if fam_name=="Poisson":
        fam_obj=Poisson()
    elif fam_name=="NegativeBinomial":
        fam_obj=NegativeBinomial(alpha=1.0)
    elif fam_name=="Gaussian":
        fam_obj=Gaussian()
    elif fam_name=="Gamma":
        fam_obj=Gamma(link=ln.log())
    elif fam_name=="InverseGaussian":
        fam_obj=InverseGaussian()
    else:
        logging.warning(f"Unknown family={fam_name}")
        return None,None

    if struct=="Independence":
        cov_obj=Independence()
    else:
        cov_obj=Exchangeable()

    model=GEE.from_formula("_score_col ~ media_category",
                           groups="media_outlet_clean",
                           data=d2,
                           family=fam_obj,
                           cov_struct=cov_obj)
    bres=model.fit(scale=None)
    y=np.asarray(bres.model.endog)
    mu=np.asarray(bres.fittedvalues)
    n=len(y)
    p=len(bres.params)
    dfresid=n-p

    if scale_name=="none" or dfresid<=0:
        return d2,bres
    else:
        val=None
        if scale_name=="pearson":
            val=compute_pearson_scale(y,mu,dfresid)
        elif scale_name=="deviance":
            val=compute_deviance_scale(y,mu,dfresid)
        elif scale_name=="ub":
            val=compute_ub_scale(y,mu,dfresid)
        elif scale_name=="bc":
            val=compute_bc_scale(y,mu,dfresid)

        if val is not None and not np.isnan(val):
            final_res=model.fit(scale=val)
            return d2, final_res
        else:
            return d2,bres

def pairwise_and_diagnostics(df, sentiment, measure, fam_name, struct, scale_name):
    d2, final_fit=refit_best_gee_with_scale(df, sentiment, measure, fam_name, struct, scale_name)
    if final_fit is None:
        return (None, None, None, None, None, None, None, None)

    summary_txt=final_fit.summary().as_text()
    diag=check_residuals_and_correlation(final_fit)
    pseudo_dict=pseudo_likelihood_check(final_fit)
    cv_val=cross_validation_gee(d2,"_score_col ~ media_category","media_outlet_clean",
                                final_fit.model.family,
                                final_fit.model.cov_struct, n_folds=3)
    sens_df=sensitivity_analysis_correlation(d2,"_score_col ~ media_category","media_outlet_clean")
    boot_df=bootstrap_gee(d2,"_score_col ~ media_category","media_outlet_clean",
                          B=3,
                          family=final_fit.model.family,
                          cov_struct=final_fit.model.cov_struct)

    params=final_fit.params
    cov=final_fit.cov_params()
    mdf=final_fit.model.data.frame
    cats=mdf["media_category"].unique()
    if pd.api.types.is_categorical_dtype(mdf["media_category"]):
        cats = mdf["media_category"].cat.categories

    ref=cats[0]
    idx_map={ref:0}
    for c in cats[1:]:
        nm=f"media_category[T.{c}]"
        if nm not in final_fit.model.exog_names:
            continue
        idx_map[c]=final_fit.model.exog_names.index(nm)

    pair_list=[]
    for i in range(len(cats)):
        for j in range(i+1,len(cats)):
            ca,cb=cats[i],cats[j]
            if ca not in idx_map and cb not in idx_map:
                continue
            con=np.zeros(len(params))
            if ca==ref and cb in idx_map:
                con[idx_map[cb]]=-1.0
            elif cb==ref and ca in idx_map:
                con[idx_map[ca]]=1.0
            else:
                if ca in idx_map:
                    con[idx_map[ca]]=1.0
                if cb in idx_map:
                    con[idx_map[cb]]=-1.0

            diff_est=con@params
            diff_var=con@cov@con
            diff_se=np.sqrt(diff_var) if diff_var>0 else np.nan
            z=diff_est/diff_se if not np.isnan(diff_se) and diff_se>0 else np.nan
            pval=2*(1-norm.cdf(abs(z))) if not np.isnan(z) else np.nan
            pair_list.append((ca,cb,diff_est,diff_se,z,pval))

    pair_df=pd.DataFrame(pair_list,columns=["CategoryA","CategoryB","Difference","SE","Z","p_value"])
    if not pair_df.empty:
        rej, p_adj,_,_=multipletests(pair_df["p_value"],method="fdr_bh")
        pair_df["p_value_adj"]=p_adj
        pair_df["reject_H0"]=rej
    else:
        pair_df["p_value_adj"]=np.nan
        pair_df["reject_H0"]=False

    cat_list=list(cats)
    cat_index_map={cat_list[i]:i+1 for i in range(len(cat_list))}
    diff_map={c:set() for c in cat_list}
    for i,row in pair_df.iterrows():
        A=row["CategoryA"]
        B=row["CategoryB"]
        if row["reject_H0"]:
            diff_map[A].add(cat_index_map[B])
            diff_map[B].add(cat_index_map[A])

    rows=[]
    for c in cat_list:
        diffs=sorted(list(diff_map[c]))
        diffs_str=",".join(str(x) for x in diffs)
        rows.append((c,diffs_str))
    diffIDs_df=pd.DataFrame(rows, columns=["Category","DiffIDs"])

    return summary_txt, pair_df, diffIDs_df, diag, pseudo_dict, cv_val, sens_df, boot_df

###############################################################################
# 8) GEE analyses driver
###############################################################################
def run_gee_for_sentiment_measure_best_qic(df, sentiment, measure):
    # Create _score_col for "Quotation" or "Fulltext"
    if measure=="Quotation":
        pat=rf"^{re.escape(sentiment)}_\d+$"
        matched=[c for c in df.columns if re.match(pat,c)]
        if not matched:
            return None
        df["_score_col"]=df[matched].clip(lower=0).mean(axis=1)
    else:
        # measure="Fulltext"
        fcol=f"{sentiment}_fulltext"
        if fcol not in df.columns:
            return None
        df["_score_col"]=df[fcol].clip(lower=0)

    needed=["_score_col","media_outlet_clean","media_category"]
    d2=df.dropna(subset=needed).copy()
    if len(d2)<2 or d2["media_category"].nunique()<2:
        return None

    best_tuple, combos=try_all_families_and_scales(d2,"_score_col ~ media_category","media_outlet_clean")
    if best_tuple is None:
        return None
    famName, corName, scOpt, bestQICVal=best_tuple
    combos_df=pd.DataFrame(combos, columns=["Family","CovStruct","Scale","QIC"])

    return {
        "Sentiment":sentiment,
        "Measure":measure,
        "Best_Family":famName,
        "Best_Structure":corName,
        "Best_Scale":scOpt,
        "Best_QIC_main":bestQICVal,
        "AllCombos": combos_df
    }

def run_gee_analyses_best_qic(df):
    """
    Runs the best QIC approach multi-families for each sentiment–measure
    but only for Quotation, Fulltext in this example.
    """
    logging.info("Running best QIC approach multi-families for each sentiment–measure.")
    measure_list = ["Quotation","Fulltext"]
    best_list=[]
    combos_list=[]
    for s in CATEGORIES:
        for meas in measure_list:
            info=run_gee_for_sentiment_measure_best_qic(df,s,meas)
            if info is not None:
                best_list.append({
                    "Sentiment":info["Sentiment"],
                    "Measure":info["Measure"],
                    "Best_Family":info["Best_Family"],
                    "Best_Structure":info["Best_Structure"],
                    "Best_Scale":info["Best_Scale"],
                    "Best_QIC_main":info["Best_QIC_main"]
                })
                cdf=info["AllCombos"]
                cdf["Sentiment"]=s
                cdf["Measure"]=meas
                combos_list.append(cdf)
    df_best=pd.DataFrame(best_list)
    df_all=pd.concat(combos_list,ignore_index=True) if combos_list else pd.DataFrame()
    return df_best, df_all


###############################################################################
# 9) Compile results into multiple Excel
###############################################################################
def compile_results_into_multiple_workbooks(
    aggregated_df, stats_df, raw_df,
    df_best_qic, df_all_combos,
    plots_dir,
    main_excel, raw_excel, gee_excel, plots_excel, combined_excel,
    df_full
):
    with pd.ExcelWriter(main_excel, engine="openpyxl") as w:
        aggregated_df.to_excel(w,"Aggregated_Scores",index=False)
        stats_df.to_excel(w,"Mean_Median_Statistics",index=False)

    with pd.ExcelWriter(raw_excel, engine="openpyxl") as w:
        raw_df.to_excel(w,"Raw_Data",index=False)
        for s in CATEGORIES:
            s_cols=[c for c in raw_df.columns if c.startswith(s+"_")]
            s_df=raw_df[["media_category","media_outlet"]+s_cols].copy()
            s_df.to_excel(w,f"Raw_{s[:29]}",index=False)

    diag_records=[]
    cv_records=[]
    sens_records=[]
    boot_records=[]
    idx_rows=[]

    with pd.ExcelWriter(gee_excel, engine="openpyxl") as writer:
        if not df_all_combos.empty:
            df_all_combos.to_excel(writer,"All_Combos",index=False)

        for i,row in df_best_qic.iterrows():
            s=row["Sentiment"]
            meas=row["Measure"]
            fam=row["Best_Family"]
            st=row["Best_Structure"]
            sc=row["Best_Scale"]
            best_qic=row["Best_QIC_main"]

            sh_name=f"BestQIC_{s[:8]}_{meas[:8]}"
            out=pairwise_and_diagnostics(df_full, s, meas, fam, st, sc)
            if out[0] is None:
                tmp_df=pd.DataFrame({"Summary":["No valid model or not enough data."]})
                tmp_df.to_excel(writer,sh_name,index=False)
                continue

            summary_txt, pair_df, diffIDs_df, diag_dict, pseudo_dict, cv_val, sens_df_df, boot_df=out

            lines=summary_txt.split("\n")
            sumdf=pd.DataFrame({"GEE_Summary":lines})
            sumdf.to_excel(writer, sh_name, index=False)

            sr=len(sumdf)+2
            pair_df.to_excel(writer, sh_name, index=False, startrow=sr)
            ws=writer.sheets[sh_name]
            for row_idx in range(len(pair_df)):
                if pair_df.loc[row_idx,"reject_H0"]:
                    rrow=sr+1+row_idx
                    for col_idx in range(1, pair_df.shape[1]+1):
                        cell=ws.cell(row=rrow+1, column=col_idx)
                        cell.fill=PatternFill(fill_type="solid", start_color="FFFF0000", end_color="FFFF0000")

            sr2=sr+len(pair_df)+2
            if diffIDs_df is not None and not diffIDs_df.empty:
                diffIDs_df.to_excel(writer, sh_name, index=False, startrow=sr2)

            diag_records.append({
                "Sentiment":s,
                "Measure":meas,
                "Family":fam,
                "Structure":st,
                "Scale":sc,
                "mean_pearson_res":diag_dict["mean_pearson_res"],
                "std_pearson_res":diag_dict["std_pearson_res"],
                "avg_within_corr":diag_dict["avg_within_corr"],
                "overdisp_ratio":diag_dict["overdisp_ratio"],
                "assessment":diag_dict["assessment"],
                "NB_QIC":pseudo_dict["NB_QIC"],
                "Poisson_QIC":pseudo_dict["Poisson_QIC"],
                "diff_QIC":pseudo_dict["diff_QIC"],
                "pseudo_conclusion":pseudo_dict["conclusion"]
            })

            cv_records.append({
                "Sentiment":s,
                "Measure":meas,
                "Family":fam,
                "Structure":st,
                "Scale":sc,
                "CV_MSE":cv_val
            })

            tmp_sens=sens_df_df.copy()
            tmp_sens["Sentiment"]=s
            tmp_sens["Measure"]=meas
            sens_records.append(tmp_sens)

            param_means=boot_df.mean().round(4).to_dict()
            param_stds=boot_df.std().round(4).to_dict()
            boot_records.append({
                "Sentiment":s,
                "Measure":meas,
                "Family":fam,
                "Structure":st,
                "Scale":sc,
                "BootMean":str(param_means),
                "BootStd":str(param_stds)
            })

            idx_rows.append({
                "Sentiment": s,
                "Measure": meas,
                "Family": fam,
                "Structure": st,
                "Scale": sc,
                "BestQIC": best_qic
            })

        idxdf=pd.DataFrame(idx_rows)
        idxdf.to_excel(writer,"BestQIC_Index",index=False)

        diag_df=pd.DataFrame(diag_records)
        diag_df.to_excel(writer,"Diagnostics",index=False)

        cv_df=pd.DataFrame(cv_records)
        cv_df.to_excel(writer,"CrossValidation_Res",index=False)

        if len(sens_records)>0:
            sens_all=pd.concat(sens_records,ignore_index=True)
            sens_all.to_excel(writer,"Sensitivity_Analysis",index=False)

        if len(boot_records)>0:
            boot_all=pd.DataFrame(boot_records)
            boot_all.to_excel(writer,"Bootstrap_Res",index=False)

    # Attempt to embed any saved plots
    wb_plots=Workbook()
    if "Sheet" in wb_plots.sheetnames:
        wb_plots.remove(wb_plots["Sheet"])
    any_sheets=False
    for s in CATEGORIES:
        q_path=os.path.join(plots_dir,f"quote_{s}.png")
        if os.path.exists(q_path):
            st=f"Quote_{s[:28]}"
            ws=wb_plots.create_sheet(title=st)
            try:
                img=ExcelImage(q_path)
                img.anchor="A1"
                ws.add_image(img)
            except:
                pass
            any_sheets=True

        f_path=os.path.join(plots_dir,f"fulltext_{s}.png")
        if os.path.exists(f_path):
            st2=f"Fulltext_{s[:25]}"
            ws2=wb_plots.create_sheet(title=st2)
            try:
                img2=ExcelImage(f_path)
                img2.anchor="A1"
                ws2.add_image(img2)
            except:
                pass
            any_sheets=True

    # Also correlation bar
    cbar=os.path.join(plots_dir,"correlation_quotation_fulltext_bar.png")
    if os.path.exists(cbar):
        ws3=wb_plots.create_sheet("Correlation_Bar")
        try:
            ig3=ExcelImage(cbar)
            ig3.anchor="A1"
            ws3.add_image(ig3)
        except:
            pass
        any_sheets=True

    # Possibly combined scatter
    combp=os.path.join(plots_dir,"combined_normalized_scatter.png")
    if os.path.exists(combp):
        ws4=wb_plots.create_sheet("Combined_ZScatter")
        try:
            ig4=ExcelImage(combp)
            ig4.anchor="A1"
            ws4.add_image(ig4)
        except:
            pass
        any_sheets=True

    if not any_sheets:
        wb_plots.create_sheet("DummySheet")
    wb_plots.save(plots_excel)

    # Combined
    raw_clean=raw_df.copy()
    raw_clean=raw_clean.applymap(lambda x:", ".join(x) if isinstance(x,list) else x)
    wb_comb=Workbook()
    if "Sheet" in wb_comb.sheetnames:
        wb_comb.remove(wb_comb["Sheet"])

    ws_agg=wb_comb.create_sheet("Aggregated_Scores")
    for r in dataframe_to_rows(aggregated_df,index=False,header=True):
        ws_agg.append(r)

    ws_stats=wb_comb.create_sheet("Mean_Median_Statistics")
    for r in dataframe_to_rows(stats_df,index=False,header=True):
        ws_stats.append(r)

    ws_raw=wb_comb.create_sheet("Raw_Data")
    for r in dataframe_to_rows(raw_clean,index=False,header=True):
        ws_raw.append(r)

    ws_best=wb_comb.create_sheet("BestQIC_Table")
    for r in dataframe_to_rows(df_best_qic,index=False,header=True):
        ws_best.append(r)

    wb_comb.save(combined_excel)

###############################################################################
# 10) Pipeline function
###############################################################################
def run_pipeline_for_df(
    df, 
    prefix="", 
    # outputs
    main_excel=OUTPUT_EXCEL_MAIN,
    raw_excel=OUTPUT_EXCEL_RAW,
    gee_excel=OUTPUT_EXCEL_GEE,
    plots_excel=OUTPUT_EXCEL_PLOTS,
    combined_excel=OUTPUT_EXCEL_COMBINED,
    qftc_excel=OUTPUT_EXCEL_QFTC,
    outdir=BASE_OUTPUT_DIR,
    csvout=BASE_CSV_OUTPUT_DIR
):
    """Runs the entire pipeline for a given DataFrame with optional prefix."""
    # Suffix files if prefix
    if prefix:
        main_excel = main_excel.replace(".xlsx", f"_{prefix}.xlsx")
        raw_excel = raw_excel.replace(".xlsx", f"_{prefix}.xlsx")
        gee_excel = gee_excel.replace(".xlsx", f"_{prefix}.xlsx")
        plots_excel = plots_excel.replace(".xlsx", f"_{prefix}.xlsx")
        combined_excel = combined_excel.replace(".xlsx", f"_{prefix}.xlsx")
        qftc_excel = qftc_excel.replace(".xlsx", f"_{prefix}.xlsx")

    # chunk & stats
    chunk_and_save(df, 20000, prefix=prefix)
    print_basic_stats(df, prefix=prefix)

    # correlation
    analyze_quotation_fulltext_correlation(
        df, 
        output_dir=outdir,
        csv_output_dir=csvout,
        output_excel_qftc = qftc_excel,
        prefix=prefix
    )

    # Aggregation & stats
    agg_df=aggregate_sentiment_scores(df,CATEGORIES)
    agg_df=calculate_averages(agg_df)
    stats_df=calculate_mean_median(agg_df)
    save_aggregated_scores_to_csv(agg_df, out_dir=csvout, prefix=prefix)
    plot_statistics(agg_df, out_dir=outdir, prefix=prefix)

    # GEE
    df_best, df_allcombos=run_gee_analyses_best_qic(df)
    print(f"Best QIC approach completed for prefix={prefix}")
    print(df_best)

    # Compile
    compile_results_into_multiple_workbooks(
        aggregated_df=agg_df,
        stats_df=stats_df,
        raw_df=df,
        df_best_qic=df_best,
        df_all_combos=df_allcombos,
        plots_dir=outdir,
        main_excel=main_excel,
        raw_excel=raw_excel,
        gee_excel=gee_excel,
        plots_excel=plots_excel,
        combined_excel=combined_excel,
        df_full=df
    )

###############################################################################
# main
###############################################################################
def main():
    setup_logging()
    logging.info("Starting best QIC GEE approach + advanced checks, plus subset for high_rate_2=yes, with category fix in cross-validation.")

    df=load_jsonl(INPUT_JSONL_FILE)
    print(f"Total articles loaded: {len(df)}")

    df=map_media_outlet_to_category(df)

    print("\n=== Running pipeline for ALL data ===")
    run_pipeline_for_df(
        df,
        prefix="All",
        main_excel=OUTPUT_EXCEL_MAIN,
        raw_excel=OUTPUT_EXCEL_RAW,
        gee_excel=OUTPUT_EXCEL_GEE,
        plots_excel=OUTPUT_EXCEL_PLOTS,
        combined_excel=OUTPUT_EXCEL_COMBINED,
        qftc_excel=OUTPUT_EXCEL_QFTC,
        outdir=BASE_OUTPUT_DIR,
        csvout=BASE_CSV_OUTPUT_DIR
    )

    print("\n=== Now subsetting for high_rate_2 == 'yes' ===")
    df_yes = df[df["high_rate_2"] == "yes"].copy()
    print(f"Subset has {len(df_yes)} articles with high_rate_2 == 'yes'")

    if len(df_yes) < 1:
        print("No entries found with high_rate_2=yes => skipping subset analysis.")
    else:
        run_pipeline_for_df(
            df_yes,
            prefix="Yes",
            main_excel=OUTPUT_EXCEL_MAIN,
            raw_excel=OUTPUT_EXCEL_RAW,
            gee_excel=OUTPUT_EXCEL_GEE,
            plots_excel=OUTPUT_EXCEL_PLOTS,
            combined_excel=OUTPUT_EXCEL_COMBINED,
            qftc_excel=OUTPUT_EXCEL_QFTC,
            outdir=BASE_OUTPUT_DIR,
            csvout=BASE_CSV_OUTPUT_DIR
        )

    logging.info("All done. Exiting.")


if __name__=="__main__":
    main()
