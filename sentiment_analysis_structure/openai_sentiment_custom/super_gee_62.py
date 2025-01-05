import json
import os
import re
import sys
import warnings
import logging
import math
import gc
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import seaborn as sns

from scipy.stats import pearsonr, norm
from tqdm import tqdm

# Statsmodels & related
import statsmodels.api as sm
from statsmodels.genmod.families import (
    Poisson, NegativeBinomial,
    Gaussian, Gamma, InverseGaussian
)
import statsmodels.genmod.families.links as ln
from statsmodels.genmod.generalized_estimating_equations import GEE
from statsmodels.genmod.cov_struct import Independence, Exchangeable
from statsmodels.stats.multitest import multipletests

# For saving Excel files / images
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill

###############################################################################
# CONFIG
###############################################################################
INPUT_JSONL_FILE = "processed_all_articles_fixed_5.jsonl"

OUTPUT_DIR = "analysis_output"
os.makedirs(OUTPUT_DIR, exist_ok=True)

BASE_MAIN_FILE     = "analysis_main.xlsx"
BASE_RAW_FILE      = "analysis_raw.xlsx"
BASE_GEE_FILE      = "analysis_gee.xlsx"
BASE_PLOTS_FILE    = "analysis_plots.xlsx"
BASE_COMBINED_FILE = "analysis_combined.xlsx"

LOG_FILE = "analysis.log"

GRAPH_OUTPUT_DIR = "graphs_analysis"
CSV_OUTPUT_DIR   = "csv_raw_scores"
os.makedirs(GRAPH_OUTPUT_DIR, exist_ok=True)
os.makedirs(CSV_OUTPUT_DIR, exist_ok=True)

CATEGORIES = [
    "joy","sadness","anger","fear",
    "surprise","disgust","trust","anticipation",
    "negative_sentiment","positive_sentiment"
]

MEDIA_CATEGORIES_ORIG = {
    "Scientific": ["nature","sciam","stat","newscientist"],
    "Left": ["theatlantic","the daily beast","the intercept","mother jones","msnbc","slate","vox","huffpost"],
    "Lean Left": ["ap","axios","cnn","guardian","business insider","nbcnews","npr","nytimes","politico","propublica","wapo","usa today"],
    "Center": ["reuters","marketwatch","financial times","newsweek","forbes"],
    "Lean Right": ["thedispatch","epochtimes","foxbusiness","wsj","national review","washtimes"],
    "Right": ["breitbart","theblaze","daily mail","dailywire","foxnews","nypost","newsmax"],
}

MEDIA_CATEGORIES_COLLAPSED = {
    "Scientific": ["nature","sciam","stat","newscientist"],
    "Left": [
        "theatlantic","the daily beast","the intercept","mother jones",
        "msnbc","slate","vox","huffpost",
        "ap","axios","cnn","guardian","business insider","nbcnews",
        "npr","nytimes","politico","propublica","wapo","usa today"
    ],
    "Center": ["reuters","marketwatch","financial times","newsweek","forbes"],
    "Right": [
        "thedispatch","epochtimes","foxbusiness","wsj","national review","washtimes",
        "breitbart","theblaze","daily mail","dailywire","foxnews","nypost","newsmax"
    ],
}

# A global store for parameter-level comparisons
COMPARE_IMBALANCE = []

###############################################################################
# LOGGING
###############################################################################
def setup_logging():
    log_format = "%(asctime)s [%(levelname)s] %(module)s - %(message)s"
    logging.basicConfig(filename=LOG_FILE, level=logging.DEBUG, format=log_format)
    console = logging.StreamHandler(sys.stdout)
    console.setLevel(logging.INFO)
    console.setFormatter(logging.Formatter(log_format))
    logging.getLogger().addHandler(console)
    logging.info("Logging initialized.")

warnings.filterwarnings(
    "ignore",
    message="QIC values obtained using scale=None are not appropriate for comparing models"
)

###############################################################################
# rename_files_for_prefix
###############################################################################
def rename_files_for_prefix(base_name, prefix):
    if not prefix:
        return os.path.join(OUTPUT_DIR, base_name)
    root, ext = os.path.splitext(base_name)
    if prefix.endswith("_collapsed"):
        short_pre = prefix.replace("_collapsed","")
        new_fn = f"{root}_collapsed_{short_pre}{ext}"
    else:
        new_fn = f"{root}_{prefix}{ext}"
    return os.path.join(OUTPUT_DIR, new_fn)

###############################################################################
# 1) Load & map
###############################################################################
def load_jsonl(jsonl_file):
    logging.info(f"Loading JSONL => {jsonl_file}")
    records=[]
    with open(jsonl_file,"r",encoding="utf-8") as f:
        for line in tqdm(f, desc="Loading JSONL"):
            rec=json.loads(line)
            records.append(rec)
    df=pd.DataFrame(records)
    return df

def map_media_outlet_to_category(df, cat_dict):
    cat_map={}
    for cat, outls in cat_dict.items():
        for o in outls:
            cat_map[o.lower().strip()] = cat

    df2 = df.copy()
    df2["media_outlet_clean"] = df2["media_outlet"].str.lower().str.strip()
    df2["media_category"] = df2["media_outlet_clean"].map(cat_map).fillna("Other")
    return df2

###############################################################################
# 2) chunk + stats
###############################################################################
def chunk_and_save(df, chunk_size=20000, prefix=""):
    logging.info(f"Chunking => len={len(df)}, chunk_size={chunk_size}, prefix={prefix}")
    for i in range(0,len(df),chunk_size):
        part=df.iloc[i:i+chunk_size]
        suffix=f"_{prefix}" if prefix else ""
        out_csv=os.path.join(CSV_OUTPUT_DIR, f"raw_data_part_{(i//chunk_size)+1}{suffix}.csv")
        part.to_csv(out_csv,index=False)
        print(f"Saved chunk {(i//chunk_size)+1} => {out_csv}")

def print_basic_stats(df, prefix=""):
    logging.info(f"Basic stats => total articles={len(df)}, prefix={prefix}")
    print(f"\nSummary Stats (prefix={prefix or '(None)'}) =>")
    print("Total articles:", len(df))
    if "media_outlet_clean" in df.columns:
        vc=df["media_outlet_clean"].value_counts()
        print("\nArticles per outlet (clusters):")
        print(vc)
    if "media_category" in df.columns:
        vc2=df["media_category"].value_counts()
        print("\nArticles per category:")
        print(vc2)
    print()

###############################################################################
# 3) correlation analyses
###############################################################################
def analyze_2fields_correlation(df, left_field_pattern, right_field,
                                correlation_title, output_excel_base, prefix=""):
    suffix=f"_{prefix}" if prefix else ""

    records=[]
    for cat in df["media_category"].dropna().unique():
        dcat=df[df["media_category"]==cat]
        for s in CATEGORIES:
            pat=re.compile(left_field_pattern.replace("<sent>", re.escape(s)))
            matched=[c for c in dcat.columns if pat.match(c)]
            if matched:
                clp=dcat[matched].clip(lower=0)
                sum_v=clp.sum(skipna=True).sum()
                ccount=clp.count().sum()
                left_avg=sum_v/ccount if ccount>0 else np.nan
            else:
                left_avg=np.nan

            rfield=right_field.replace("<sent>", s)
            if rfield in dcat.columns:
                rv=dcat[rfield].clip(lower=0)
                r_sum=rv.sum(skipna=True)
                r_cnt=rv.count()
                right_avg=r_sum/r_cnt if r_cnt>0 else np.nan
            else:
                right_avg=np.nan

            records.append({
                "MediaCategory": cat,
                "Sentiment": s,
                "Left_Average": left_avg,
                "Right_Average": right_avg
            })

    agg_df=pd.DataFrame(records)
    cor_results=[]
    all_combo=[]
    for s in CATEGORIES:
        sub=agg_df[(agg_df["Sentiment"]==s)].dropna(subset=["Left_Average","Right_Average"])
        if len(sub)>1:
            cval,_=pearsonr(sub["Left_Average"], sub["Right_Average"])
        else:
            cval=np.nan
        cor_results.append({"Sentiment":s,"Correlation":cval})
        if not sub.empty:
            all_combo.append(sub.copy())

    cor_df=pd.DataFrame(cor_results)
    plt.figure(figsize=(8,5))
    sns.barplot(x="Sentiment", y="Correlation", data=cor_df, color="gray")
    plt.title(f"{correlation_title} - prefix={prefix}")
    plt.xticks(rotation=45,ha="right")
    plt.ylim(-1,1)
    plt.tight_layout()

    barname=f"correlation_{prefix}_{correlation_title.replace(' ','_')}_bar.png"
    barpath=os.path.join(GRAPH_OUTPUT_DIR, barname.lower())
    try:
        plt.savefig(barpath)
    except:
        pass
    plt.close()

    out_csv = os.path.join(CSV_OUTPUT_DIR, f"correlation_{prefix}_{correlation_title.replace(' ','_')}.csv")
    cor_df.to_csv(out_csv, index=False)

    if all_combo:
        allc=pd.concat(all_combo, ignore_index=True)
        allc["Lmean"]=allc.groupby("Sentiment")["Left_Average"].transform("mean")
        allc["Lstd"]=allc.groupby("Sentiment")["Left_Average"].transform("std")
        allc["Rmean"]=allc.groupby("Sentiment")["Right_Average"].transform("mean")
        allc["Rstd"]=allc.groupby("Sentiment")["Right_Average"].transform("std")
        allc["Left_Z"]=(allc["Left_Average"]-allc["Lmean"])/allc["Lstd"]
        allc["Right_Z"]=(allc["Right_Average"]-allc["Rmean"])/allc["Rstd"]

        r_val, p_val = pearsonr(allc["Left_Average"], allc["Right_Average"])

        plt.figure(figsize=(6,5))
        sns.set_style("white")
        ax = sns.regplot(
            x="Left_Average",
            y="Right_Average",
            data=allc,
            scatter_kws={"color": "black", "s": 40},
            line_kws={"color": "red"}
        )
        plt.grid(False)
        plt.text(
            0.05, 0.95,
            f"R = {r_val:.2f}, p = {p_val:.3g}",
            ha='left',
            va='top',
            transform=ax.transAxes
        )
        plt.title(f"{correlation_title} scatter - prefix={prefix}")
        scattername=f"scatter_{prefix}_{correlation_title.replace(' ','_')}.png"
        scatterpath=os.path.join(GRAPH_OUTPUT_DIR, scattername.lower())
        plt.tight_layout()
        try:
            plt.savefig(scatterpath)
        except:
            pass
        plt.close()

        out_xlsx = os.path.join(CSV_OUTPUT_DIR, f"correlation_{prefix}_{correlation_title.replace(' ','_')}.xlsx")
        with pd.ExcelWriter(out_xlsx, engine="openpyxl") as writer:
            cor_df.to_excel(writer, sheet_name="Correlations", index=False)
            allc.to_excel(writer, sheet_name="ZScorePlotData", index=False)

def analyze_all_custom_correlations(df, prefix=""):
    QFTC = "quotation_fulltext_correlation.xlsx"
    QIvFI= "quotation_intensity_fulltext_intensity_correlation.xlsx"
    FIvF = "fulltext_intensity_vs_fulltext_correlation.xlsx"

    analyze_2fields_correlation(
        df, "^<sent>_\\d+$", "<sent>_fulltext",
        "Quotation_vs_Fulltext", QFTC, prefix
    )
    analyze_2fields_correlation(
        df, "^<sent>_\\d+_intensity$", "<sent>_fulltext_intensity",
        "Quotation_Intensity_vs_Fulltext_Intensity", QIvFI, prefix
    )
    analyze_2fields_correlation(
        df, "<sent>_fulltext_intensity", "<sent>_fulltext",
        "Fulltext_Intensity_vs_Fulltext", FIvF, prefix
    )

###############################################################################
# 4) Aggregation & Stats
###############################################################################
def aggregate_sentiment_scores(df, sentiments):
    recs=[]
    cats_in_df = df["media_category"].unique()

    for cat in cats_in_df:
        sub = df[df["media_category"] == cat]

        for s in sentiments:
            # Quotation columns
            pat = rf"^{re.escape(s)}_\d+$"
            matched = [c for c in sub.columns if re.match(pat,c)]
            if matched:
                qvals = sub[matched].clip(lower=0).values.flatten()
                qsum  = np.nansum(qvals)
                qcount= np.count_nonzero(~np.isnan(qvals))
                if qcount>1:
                    q_sd  = float(np.nanstd(qvals, ddof=1))
                    q_sem = float(q_sd / np.sqrt(qcount))
                else:
                    q_sd  = None
                    q_sem = None
            else:
                qsum,qcount,q_sd,q_sem = (0,0,None,None)

            # Fulltext column
            fcol = f"{s}_fulltext"
            if fcol in sub.columns:
                fv   = sub[fcol].clip(lower=0).dropna()
                f_sum= np.nansum(fv)
                f_cnt= len(fv)
                if f_cnt>1:
                    f_sd  = float(np.nanstd(fv, ddof=1))
                    f_sem = float(f_sd / np.sqrt(f_cnt))
                else:
                    f_sd  = None
                    f_sem = None
            else:
                f_sum,f_cnt,f_sd,f_sem = (0,0,None,None)

            recs.append({
                "Media Category": cat,
                "Sentiment/Emotion": s,
                "Quotation_Sum":       qsum,
                "Quotation_Count":     qcount,
                "Quotation_SD":        q_sd,
                "Quotation_SEM":       q_sem,
                "Fulltext_Sum":        f_sum,
                "Fulltext_Count":      f_cnt,
                "Fulltext_SD":         f_sd,
                "Fulltext_SEM":        f_sem
            })
    return pd.DataFrame(recs)

def calculate_averages(agg_df):
    def sdiv(a,b):
        return a/b if b>0 else None
    agg_df["Quotation_Average"] = agg_df.apply(
        lambda r: sdiv(r["Quotation_Sum"], r["Quotation_Count"]), axis=1
    )
    agg_df["Fulltext_Average"] = agg_df.apply(
        lambda r: sdiv(r["Fulltext_Sum"], r["Fulltext_Count"]), axis=1
    )
    return agg_df

def calculate_mean_median(agg_df):
    rows=[]
    for s in CATEGORIES:
        sub=agg_df[agg_df["Sentiment/Emotion"]==s]
        qa=sub["Quotation_Average"].dropna()
        fa=sub["Fulltext_Average"].dropna()
        rows.append({
            "Sentiment/Emotion": s,
            "Mean_Quotation_Average": qa.mean() if len(qa)>0 else None,
            "Median_Quotation_Average": qa.median() if len(qa)>0 else None,
            "Mean_Fulltext_Average": fa.mean() if len(fa)>0 else None,
            "Median_Fulltext_Average": fa.median() if len(fa)>0 else None
        })
    return pd.DataFrame(rows)

def save_aggregated_scores_to_csv(agg_df, out_dir, prefix=""):
    suffix=f"_{prefix}" if prefix else ""
    fn=os.path.join(out_dir, f"aggregated_sentiment_emotion_scores{suffix}.csv")
    agg_df.to_csv(fn, index=False)
    print(f"Aggregated => {fn}")

def plot_statistics(agg_df, out_dir, prefix=""):
    sns.set_style("whitegrid")
    catvals = agg_df["Media Category"].unique()
    suffix=f"_{prefix}" if prefix else ""
    for s in CATEGORIES:
        sub=agg_df[agg_df["Sentiment/Emotion"]==s]
        plt.figure(figsize=(10,6))
        sns.barplot(x="Media Category", y="Quotation_Average", data=sub,
                    color="steelblue", order=catvals)
        plt.title(f"Mean Quotation '{s.capitalize()}' Scores {prefix}")
        plt.xticks(rotation=45,ha="right")
        plt.tight_layout()
        out1=os.path.join(out_dir,f"quote_{s}{suffix}.png")
        try:
            plt.savefig(out1)
        except:
            pass
        plt.close()

        plt.figure(figsize=(10,6))
        sns.barplot(x="Media Category", y="Fulltext_Average", data=sub,
                    color="darkorange", order=catvals)
        plt.title(f"Mean Fulltext '{s.capitalize()}' Scores {prefix}")
        plt.xticks(rotation=45,ha="right")
        plt.tight_layout()
        out2=os.path.join(out_dir,f"fulltext_{s}{suffix}.png")
        try:
            plt.savefig(out2)
        except:
            pass
        plt.close()

###############################################################################
# coverage checks & cluster checks
###############################################################################
def coverage_simulation(n_sims=1000, n_clusters=10, cluster_size=10,
                       true_beta=0.5, alpha=0.05,
                       family=Poisson(), cov_struct=Independence(),
                       random_state=42):
    np.random.seed(random_state)
    coverage_count=0
    for _ in range(n_sims):
        cluster_ids=np.repeat(range(n_clusters), cluster_size)
        X=np.random.binomial(1,0.5,size=n_clusters*cluster_size)
        lin_pred=0+true_beta*X
        mu=np.exp(lin_pred)
        Y=np.random.poisson(mu)
        sim_df=pd.DataFrame({"Y":Y,"X":X,"cluster":cluster_ids})
        model=GEE.from_formula("Y ~ X", groups="cluster", data=sim_df,
                               family=family, cov_struct=cov_struct)
        try:
            res=model.fit(maxiter=300)
        except:
            continue
        est=res.params["X"]
        se=res.bse["X"]
        z_crit=norm.ppf(1-alpha/2)
        ci_lower=est - z_crit*se
        ci_upper=est + z_crit*se
        if (true_beta>=ci_lower) and (true_beta<=ci_upper):
            coverage_count+=1
    coverage_rate=coverage_count/n_sims
    if abs(coverage_rate-0.95)<=0.05:
        concl="No major issues"
    else:
        concl="Potential over/under correction"
    return {
        "Check":"CoverageSimulation",
        "N_Sims":n_sims,
        "CoverageRate":round(coverage_rate,3),
        "Conclusion":concl
    }

def check_number_of_clusters(df, cluster_col, threshold=20):
    n_clusters=df[cluster_col].nunique()
    if n_clusters<threshold:
        msg=f"WARNING: #Clusters={n_clusters} < {threshold}. GEE robust SE may be biased."
        logging.warning(msg)
        conclusion="Potential small-sample bias"
    else:
        msg=f"#Clusters={n_clusters} >= {threshold} => likely OK for GEE asymptotics."
        conclusion="No major issues"
    print(msg)
    return {
        "Check":"NumberOfClusters",
        "Threshold":threshold,
        "Value":n_clusters,
        "Conclusion":conclusion
    }

def check_cluster_balance(df, cluster_col, imbalance_ratio=5.0):
    sizes=df[cluster_col].value_counts()
    if len(sizes)==0:
        return {
            "Check":"ClusterBalance",
            "Value":"No clusters found",
            "Conclusion":"No data"
        }
    min_size=sizes.min()
    max_size=sizes.max()
    if min_size==0:
        return {
            "Check":"ClusterBalance",
            "Value":"Smallest cluster=0",
            "Conclusion":"Potential degenerate cluster"
        }
    ratio=max_size/min_size
    if ratio>imbalance_ratio:
        msg=f"WARNING: Cluster size ratio={ratio:.1f} > {imbalance_ratio}"
        logging.warning(msg)
        conclusion="High imbalance => potential bias in SE"
    else:
        conclusion="No major imbalance"
    return {
        "Check":"ClusterBalance",
        "Value":f"{ratio:.2f}",
        "Conclusion":conclusion,
        "ImbalanceRatio":imbalance_ratio
    }

###############################################################################
# scale computations
###############################################################################
def compute_pearson_scale(y, mu, df_resid):
    if df_resid<=0:
        return np.nan
    r=(y-mu)/np.sqrt(mu+1e-9)
    return np.sum(r**2)/df_resid

def compute_deviance_scale(y, mu, df_resid):
    if df_resid<=0:
        return np.nan
    arr=[]
    for obs,lam in zip(y,mu):
        if obs>0 and lam>0:
            arr.append(obs*math.log(obs/lam)-(obs-lam))
        elif obs==0:
            arr.append(-(obs-lam))
        else:
            arr.append(np.nan)
    dev=2*np.nansum(arr)
    return dev/df_resid if df_resid>0 else np.nan

def compute_ub_scale(y, mu, df_resid):
    p=compute_pearson_scale(y, mu, df_resid)
    if np.isnan(p):
        return p
    return 1.1*p

def compute_bc_scale(y, mu, df_resid):
    d=compute_deviance_scale(y, mu, df_resid)
    if np.isnan(d):
        return d
    return 0.9*d

###############################################################################
# check_residuals_and_correlation
###############################################################################
def check_residuals_and_correlation(final_fit):
    y=final_fit.model.endog
    mu=final_fit.fittedvalues
    pearson_res_arr = np.asarray((y - mu) / np.sqrt(mu + 1e-9))
    mean_res=np.mean(pearson_res_arr)
    std_res=np.std(pearson_res_arr)

    clusters=final_fit.model.groups
    cluster_map={}
    for i,cid in enumerate(clusters):
        cluster_map.setdefault(cid,[]).append(pearson_res_arr[i])

    wcorr=[]
    for cid, arr in cluster_map.items():
        arr = np.array(arr)
        if len(arr) > 1:
            cmat = np.corrcoef(arr)
            if cmat.ndim == 2 and cmat.shape[0] > 1:
                sum_offdiag = np.sum(cmat) - len(arr)
                denom = len(arr)*(len(arr)-1)
                wcorr.append(sum_offdiag / denom if denom > 0 else 0)
    avg_corr=np.mean(wcorr) if len(wcorr)>0 else 0.0

    deviance=getattr(final_fit,"pearson_chi2",np.nan)
    dfres=getattr(final_fit,"df_resid",1)
    overdisp=np.nan
    if dfres>0 and not np.isnan(deviance):
        overdisp=deviance/dfres

    assess=""
    if abs(mean_res)>1:
        assess+="Mean Pearson residual>1 => possible misfit. "
    if std_res>2:
        assess+="Residual std>2 => possible outliers. "
    if abs(avg_corr)>0.3:
        assess+=f"Within-cluster corr={avg_corr:.2f} => structure suspect. "
    if (not np.isnan(overdisp)) and (overdisp>2):
        assess+=f"Overdisp={overdisp:.2f} => consider NB or other approach. "

    if assess=="":
        assess="No major issues from these checks."

    return {
        "mean_pearson_res": mean_res,
        "std_pearson_res": std_res,
        "avg_within_corr": avg_corr,
        "overdisp_ratio": overdisp,
        "assessment": assess
    }

###############################################################################
# pseudo_likelihood_check
###############################################################################
def pseudo_likelihood_check(final_fit):
    from statsmodels.genmod.families.family import NegativeBinomial
    fam=final_fit.model.family
    if not isinstance(fam, sm.genmod.families.family.Poisson):
        return {
            "NB_QIC":None,
            "Poisson_QIC":None,
            "diff_QIC":None,
            "conclusion":"Non-Poisson => skip NB check"
        }
    data=final_fit.model.data.frame
    groups=final_fit.model.groups
    formula=final_fit.model.formula
    cov_struct=final_fit.model.cov_struct
    nb_model=GEE.from_formula(formula, groups=groups, data=data,
                              family=NegativeBinomial(alpha=1.0),
                              cov_struct=cov_struct)
    nb_res=nb_model.fit(maxiter=300)
    nb_qic=nb_res.qic()
    if isinstance(nb_qic, tuple):
        nb_qic=nb_qic[0]
    old_qic=final_fit.qic()
    if isinstance(old_qic, tuple):
        old_qic=old_qic[0]

    diff_qic=None
    conclusion=""
    if isinstance(nb_qic,(float,int)) and isinstance(old_qic,(float,int)):
        diff_qic=old_qic - nb_qic
        if diff_qic>0:
            conclusion="NegBin better"
        else:
            conclusion="No NB improvement"
    else:
        conclusion="Could not compare QIC"

    return {
        "NB_QIC":nb_qic,
        "Poisson_QIC":old_qic,
        "diff_QIC":diff_qic,
        "conclusion":conclusion
    }

###############################################################################
# cross_validation_gee
###############################################################################
def cross_validation_gee(data_with_score, formula, group_col, family, cov_struct, n_folds=3):
    cluster_ids=data_with_score[group_col].unique()
    np.random.shuffle(cluster_ids)
    folds=np.array_split(cluster_ids,n_folds)
    metrics=[]
    for i in range(n_folds):
        testc=set(folds[i])
        train_df=data_with_score[~data_with_score[group_col].isin(testc)].copy()
        test_df=data_with_score[data_with_score[group_col].isin(testc)].copy()
        if len(train_df)<1 or len(test_df)<1:
            continue
        train_df["media_category"]=train_df["media_category"].astype("category")
        cats=train_df["media_category"].cat.categories
        test_df["media_category"]=pd.Categorical(test_df["media_category"], categories=cats)

        mod=GEE.from_formula(formula, groups=group_col, data=train_df,
                             family=family, cov_struct=cov_struct)
        res=mod.fit(maxiter=300)
        pred=res.predict(test_df)
        obs=test_df[res.model.endog_names]
        mse=np.mean((obs-pred)**2)
        metrics.append(mse)
    return np.mean(metrics) if len(metrics)>0 else np.nan

###############################################################################
# sensitivity_analysis_correlation
###############################################################################
def sensitivity_analysis_correlation(df, formula, group_col):
    from statsmodels.genmod.families import Poisson, Gaussian
    from statsmodels.genmod.families.family import NegativeBinomial
    alt_families=[Poisson(), NegativeBinomial(alpha=1.0), Gaussian()]
    alt_structs=[Independence(), Exchangeable()]
    results=[]
    for fam in alt_families:
        for covs in alt_structs:
            mod=GEE.from_formula(formula, groups=group_col, data=df,
                                 family=fam, cov_struct=covs)
            try:
                r=mod.fit(maxiter=300)
                qic_val=r.qic()
                if isinstance(qic_val, tuple):
                    qic_val=qic_val[0]
                param3=r.params.head(3).round(3).to_dict()
                results.append({
                    "Family":fam.__class__.__name__,
                    "CovStruct":covs.__class__.__name__,
                    "QIC":qic_val,
                    "ParamSample":param3
                })
            except:
                results.append({
                    "Family":fam.__class__.__name__,
                    "CovStruct":covs.__class__.__name__,
                    "QIC":np.nan,
                    "ParamSample":"fail"
                })
    return pd.DataFrame(results)

###############################################################################
# try_all_families_and_scales
###############################################################################
###############################################################################
# SCALE COMPUTATIONS (same as in new code)
###############################################################################
def compute_pearson_scale(y, mu, df_resid):
    if df_resid <= 0:
        return np.nan
    r = (y - mu) / np.sqrt(mu + 1e-9)
    return np.sum(r**2) / df_resid

def compute_deviance_scale(y, mu, df_resid):
    if df_resid <= 0:
        return np.nan
    arr = []
    for obs, lam in zip(y, mu):
        if obs > 0 and lam > 0:
            arr.append(obs * np.log(obs / lam) - (obs - lam))
        elif obs == 0:
            arr.append(-(obs - lam))
        else:
            arr.append(np.nan)
    dev = 2 * np.nansum(arr)
    return dev / df_resid if df_resid > 0 else np.nan

def compute_ub_scale(y, mu, df_resid):
    p = compute_pearson_scale(y, mu, df_resid)
    if np.isnan(p):
        return p
    return 1.1 * p

def compute_bc_scale(y, mu, df_resid):
    d = compute_deviance_scale(y, mu, df_resid)
    if np.isnan(d):
        return d
    return 0.9 * d

###############################################################################
# 1) try_all_families_and_scales (Merged):
###############################################################################
def try_all_families_and_scales(df, formula, group_col):
    """
    Re-fit for every single combination of (Family, CovStruct, Scale).
    - First pass: model.fit(scale=None)
    - If 'scale' != 'none', compute scale => second pass: model.fit(scale=val)
    - If scale is invalid/NaN, do NOT raise an error (as in the old code).
      Instead, we simply keep 'final_res = base_res' (scale=None).
    - Among all combos that succeed, pick the one with lowest QIC.
    """
    families = [
        Poisson(),
        NegativeBinomial(alpha=1.0),
        Gaussian(),
        Gamma(link=ln.log()),
        InverseGaussian()
    ]
    cor_structs = [Independence(), Exchangeable()]
    scale_opts = ["none", "pearson", "deviance", "ub", "bc"]

    best_qic = float("inf")
    best_tuple = None
    all_results = []

    for fam in families:
        fam_name = fam.__class__.__name__
        for cov_obj in cor_structs:
            cov_name = cov_obj.__class__.__name__
            for sc_opt in scale_opts:
                qic_val = float("nan")
                status_flag = "ok"
                try:
                    # 1) Build GEE with the requested family/cov_struct
                    model = GEE.from_formula(
                        formula, groups=group_col,
                        data=df, family=fam,
                        cov_struct=cov_obj
                    )
                    # 2) First pass => fit with scale=None
                    base_res = model.fit(maxiter=300, scale=None)

                    # 3) If sc_opt != "none", compute scale & re-fit
                    y = base_res.model.endog
                    mu = base_res.fittedvalues
                    n = len(y)
                    p = len(base_res.params)
                    dfresid = n - p

                    final_res = base_res  # default if sc_opt == "none"

                    if sc_opt != "none" and dfresid > 0:
                        # compute scale
                        val = None
                        if sc_opt == "pearson":
                            val = compute_pearson_scale(y, mu, dfresid)
                        elif sc_opt == "deviance":
                            val = compute_deviance_scale(y, mu, dfresid)
                        elif sc_opt == "ub":
                            val = compute_ub_scale(y, mu, dfresid)
                        elif sc_opt == "bc":
                            val = compute_bc_scale(y, mu, dfresid)

                        # If scale is valid => second pass. Otherwise, keep scale=None
                        if val is not None and not np.isnan(val):
                            final_res = model.fit(maxiter=300, scale=val)
                        # OLD-STYLE MERGE: do NOT raise error if val is invalid;
                        # simply skip re-fitting => final_res=base_res.

                    qic_val = final_res.qic()
                    # Some statsmodels versions return a tuple => QIC is first element
                    if isinstance(qic_val, tuple):
                        qic_val = qic_val[0]

                    # 4) If QIC is valid & better => update
                    if (not np.isnan(qic_val)) and (qic_val < best_qic):
                        best_qic = qic_val
                        best_tuple = (fam_name, cov_name, sc_opt, qic_val)

                except Exception as e:
                    logging.warning(f"Fit fail => {fam_name}+{cov_name}+{sc_opt} => {e}")
                    status_flag = "fail"

                # 5) Record combo => QIC + status
                all_results.append((fam_name, cov_name, sc_opt, qic_val, status_flag))

    return best_tuple, all_results

###############################################################################
# 2) refit_best_gee_with_scale (Merged):
###############################################################################
def refit_best_gee_with_scale(df, sentiment, measure, fam_name, struct, scale_name):
    """
    Reproduce old style of scale handling:
      - If scale is invalid, do not skip or fail; just keep scale=None.
    Retains all new code behaviors otherwise.
    """
    # 1) define _score_col exactly as the new code
    if measure == "Quotation":
        pat = rf"^{re.escape(sentiment)}_\d+$"
        matched = [c for c in df.columns if re.match(pat, c)]
        if not matched:
            return None, None
        d2 = df.copy()
        d2["_score_col"] = d2[matched].clip(lower=0).mean(axis=1)

    elif measure == "Fulltext":
        fcol = f"{sentiment}_fulltext"
        if fcol not in df.columns:
            return None, None
        d2 = df.copy()
        d2["_score_col"] = d2[fcol].clip(lower=0)

    elif measure == "Fulltext_Intensity":
        fcol = f"{sentiment}_fulltext_intensity"
        if fcol not in df.columns:
            return None, None
        d2 = df.copy()
        d2["_score_col"] = d2[fcol].astype(float).clip(lower=0)

    elif measure == "Title_Intensity":
        fcol = f"{sentiment}_title_intensity"
        if fcol not in df.columns:
            return None, None
        d2 = df.copy()
        d2["_score_col"] = d2[fcol].astype(float).clip(lower=0)

    elif measure == "Quotation_Intensity":
        pat = rf"^{re.escape(sentiment)}_\d+_intensity$"
        matched = [c for c in df.columns if re.match(pat, c)]
        if not matched:
            return None, None
        d2 = df.copy()
        d2["_score_col"] = d2[matched].astype(float).clip(lower=0).mean(axis=1)

    # NEW BLOCK FOR THE 6TH MEASURE:
    elif measure == "Title":
        fcol = f"{sentiment}_title"
        if fcol not in df.columns:
            return None, None
        d2 = df.copy()
        d2["_score_col"] = d2[fcol].astype(float).clip(lower=0)

    else:
        return None, None

    needed = ["_score_col", "media_outlet_clean", "media_category"]
    d2 = d2.dropna(subset=needed)
    if len(d2) < 2 or d2["media_category"].nunique() < 2:
        return None, None

    # 2) define family
    if fam_name == "Poisson":
        fam_obj = Poisson()
    elif fam_name == "NegativeBinomial":
        fam_obj = NegativeBinomial(alpha=1.0)
    elif fam_name == "Gaussian":
        fam_obj = Gaussian()
    elif fam_name == "Gamma":
        fam_obj = Gamma(link=ln.log())
    elif fam_name == "InverseGaussian":
        fam_obj = InverseGaussian()
    else:
        return None, None

    # 3) define correlation
    if struct == "Independence":
        cov_obj = Independence()
    else:
        cov_obj = Exchangeable()

    # 4) Fit with scale=None first
    model = GEE.from_formula("_score_col ~ media_category",
                             groups="media_outlet_clean",
                             data=d2,
                             family=fam_obj,
                             cov_struct=cov_obj)
    base_res = model.fit(maxiter=300, scale=None)

    y = np.asarray(base_res.model.endog)
    mu = np.asarray(base_res.fittedvalues)
    n = len(y)
    p = len(base_res.params)
    dfresid = n - p

    # 5) old-style scale re-fit: if scale_name != 'none' & valid => re-fit
    final_res = base_res
    if scale_name != "none" and dfresid > 0:
        val = None
        if scale_name == "pearson":
            val = compute_pearson_scale(y, mu, dfresid)
        elif scale_name == "deviance":
            val = compute_deviance_scale(y, mu, dfresid)
        elif scale_name == "ub":
            val = compute_ub_scale(y, mu, dfresid)
        elif scale_name == "bc":
            val = compute_bc_scale(y, mu, dfresid)

        if (val is not None) and (not np.isnan(val)):
            # If scale is good => re-fit
            final_res = model.fit(maxiter=300, scale=val)
        # If invalid => do not skip => just keep final_res=base_res

    return d2, final_res

###############################################################################
# cluster_bootstrap_gee => if you want final cluster bootstrap
###############################################################################
def cluster_bootstrap_gee(df, formula, group_col, B=200, family=None, cov_struct=None):
    """
    If you decide to run a cluster bootstrap on the final model, do it here.
    Just remove or comment out the call if you don't want it.
    """
    if family is None:
        family = Poisson()
    if cov_struct is None:
        cov_struct = Independence()

    cluster_ids = df[group_col].unique()
    M = len(cluster_ids)
    param_records=[]
    for _ in range(B):
        sample_ids = np.random.choice(cluster_ids, size=M, replace=True)
        pieces=[]
        for cid in sample_ids:
            pieces.append(df[df[group_col]==cid])
        boot_df=pd.concat(pieces, ignore_index=True)
        mod=GEE.from_formula(formula, groups=group_col, data=boot_df,
                             family=family, cov_struct=cov_struct)
        try:
            res=mod.fit(maxiter=300)
            param_records.append(res.params)
        except:
            continue

    if not param_records:
        return pd.Series(dtype=float), pd.Series(dtype=float)
    df_params=pd.DataFrame(param_records)
    means = df_params.mean(axis=0)
    sds   = df_params.std(axis=0)
    return means, sds

def jackknife_gee_md(df, formula, group_col,
                     family=Poisson(), cov_struct=Independence()):
    """
    Jackknife resampling for GEE (leave-one-cluster-out):
      - Systematically omit ONE cluster at a time
      - Fit GEE on the remaining clusters
      - Collect parameter estimates + robust SE from each fit.

    Returns:
      (df_params, df_se)
        - df_params: each row is a fitted parameter set from leaving out ONE cluster
        - df_se: each row is the robust standard errors for that run
    """
    cluster_ids = df[group_col].unique()
    param_records = []
    se_records = []
    param_index_full = None

    for cid in cluster_ids:
        # 1) remove cluster 'cid'
        jack_df = df[df[group_col] != cid]

        # 2) fit GEE
        mod = GEE.from_formula(
            formula, groups=group_col,
            data=jack_df,
            family=family,
            cov_struct=cov_struct
        )
        try:
            res = mod.fit(maxiter=300)
        except Exception as e:
            logging.warning(f"Jackknife fail => leaving out {cid}: {e}")
            continue

        # 3) confirm param alignment
        param_names = res.params.index
        if param_index_full is None:
            param_index_full = param_names
        else:
            if not np.array_equal(param_names, param_index_full):
                continue

        # 4) get robust covariance => robust SE
        robust_cov = res.cov_params()
        robust_se  = np.sqrt(np.diag(robust_cov))
        if len(robust_se) != len(param_names):
            continue

        # record this run's params + SE
        param_records.append(res.params)
        se_records.append(robust_se)

    # if no runs succeeded, return empty
    if not param_records:
        return pd.DataFrame(), pd.DataFrame()

    df_params = pd.DataFrame(param_records, columns=param_index_full)
    df_se     = pd.DataFrame(se_records, columns=df_params.columns)
    return df_params, df_se

def bootstrap_gee_md(df, formula, group_col,
                     B=200, family=Poisson(), cov_struct=Independence()):
    """
    Cluster bootstrap for GEE WITHOUT Mancl–DeRouen correction.
    - We sample clusters with replacement,
    - Fit GEE each time,
    - Return just the base robust covariance's SE for each bootstrap.

    Returns (df_params, df_boot_se):
      - df_params: each row is a set of fitted parameters from one bootstrap run
      - df_boot_se: each row is the standard errors from the base robust covariance
    """
    cluster_ids = df[group_col].unique()
    param_records = []
    se_records = []
    param_index_full = None

    for _ in range(B):
        # 1) sample clusters with replacement
        sample_ids = np.random.choice(cluster_ids, size=len(cluster_ids), replace=True)
        pieces = []
        for cid in sample_ids:
            pieces.append(df[df[group_col] == cid])
        boot_df = pd.concat(pieces, ignore_index=True)

        # 2) Fit GEE on this bootstrap sample
        mod = GEE.from_formula(
            formula, groups=group_col, data=boot_df,
            family=family, cov_struct=cov_struct
        )
        try:
            res = mod.fit(maxiter=300)
        except:
            # If it fails, skip
            continue

        # 3) Check param names
        param_names = res.params.index
        if param_index_full is None:
            param_index_full = param_names
        else:
            if not np.array_equal(param_names, param_index_full):
                continue

        # 4) Extract the base robust covariance (NOT applying M&D)
        base_cov = res.cov_params()
        se_robust = np.sqrt(np.diag(base_cov))

        if len(se_robust) != len(param_names):
            continue

        # 5) Record param estimates & SE
        param_records.append(res.params)
        se_records.append(se_robust)

    # If no valid bootstrap runs, return empty
    if not param_records:
        return pd.DataFrame(), pd.DataFrame()

    df_params = pd.DataFrame(param_records)
    df_se     = pd.DataFrame(se_records, columns=df_params.columns)
    return df_params, df_se

###############################################################################
# pairwise_and_diagnostics => fill COMPARE_IMBALANCE & create 'Trans_DiffIDs'
###############################################################################
def pairwise_and_diagnostics(
    df,
    sentiment,
    measure,
    fam_name,
    struct,
    scale_name,
    prefix="",
    isSens=False
):
    """
    A comprehensive function that:
      1) Refits the model (via refit_best_gee_with_scale).
      2) Checks residuals/correlation, pseudo-likelihood, cross-validation, sensitivity.
      3) Runs both cluster bootstrap + jackknife to gather param estimates.
      4) Computes all pairwise differences (B - A) across categories, including reference.
         (FIXED so the intercept is excluded from differences of non-reference categories.)
      5) Applies BH correction => p_value_adj.
      6) Builds 'DiffIDs' for significant pairs and 'Trans_DiffIDs' with the custom mapping.
      7) Returns an 8-tuple:

         (
             summary_txt,       # textual GEE summary
             pair_df,           # pairwise table (with p_value, p_value_adj, reject_H0)
             diffIDs_df,        # includes DiffIDs + Trans_DiffIDs
             diag_dict,         # residual checks
             pseudo_dict,       # pseudo-likelihood check
             cv_val,            # cross-validation MSE
             sens_df,           # sensitivity analysis
             boot_info          # dict with cluster + jackknife param means/std
         )
    """

    import numpy as np
    import pandas as pd
    from statsmodels.stats.multitest import multipletests
    from scipy.stats import norm

    # -------------------------------------------------------------------
    # 1) Refit final GEE
    # -------------------------------------------------------------------
    d2, final_fit = refit_best_gee_with_scale(df, sentiment, measure, fam_name, struct, scale_name)
    if final_fit is None:
        # If refitting fails or no data => skip
        return (None, None, None, None, None, None, None, None)

    summary_txt = final_fit.summary().as_text()

    # -------------------------------------------------------------------
    # 2) Checks: residuals, pseudo-likelihood, cross-validation, sensitivity
    # -------------------------------------------------------------------
    diag_dict   = check_residuals_and_correlation(final_fit)
    pseudo_dict = pseudo_likelihood_check(final_fit)
    cv_val      = cross_validation_gee(
        d2, "_score_col ~ media_category",
        "media_outlet_clean",
        final_fit.model.family,
        final_fit.model.cov_struct,
        n_folds=3
    )
    sens_df = sensitivity_analysis_correlation(
        d2, "_score_col ~ media_category",
        "media_outlet_clean"
    )

    # -------------------------------------------------------------------
    # 3) Cluster bootstrap + jackknife => param means/std
    # -------------------------------------------------------------------
    boot_params_df, boot_se_df = cluster_bootstrap_gee(
        d2, "_score_col ~ media_category",
        "media_outlet_clean",
        B=200,
        family=final_fit.model.family,
        cov_struct=final_fit.model.cov_struct
    )
    jk_params_df, jk_se_df = jackknife_gee_md(
        d2, "_score_col ~ media_category",
        "media_outlet_clean",
        family=final_fit.model.family,
        cov_struct=final_fit.model.cov_struct
    )

    # Helper: convert a Series into a 1-col DataFrame, if needed
    def ensure_dataframe(obj):
        """
        If 'obj' is a Series, convert it to a 1-column DataFrame with column name = 'value'.
        If 'obj' is already a DataFrame, return as-is.
        If empty or None => return an empty DataFrame.
        """
        if obj is None:
            return pd.DataFrame()
        if isinstance(obj, pd.Series):
            # Make a 1-col DF.  The name might be None => default to 'value'
            col_name = obj.name if obj.name else "value"
            df_out = obj.to_frame(name=col_name)
            # Because it's a single column, you'd typically have each row as each param.
            # But let's also ensure orientation is correct for the subsequent .mean(axis=0).
            # The typical shape is (#params, 1).
            # That is fine. We just treat each param as a row index.
            return df_out
        elif isinstance(obj, pd.DataFrame):
            return obj
        else:
            return pd.DataFrame()

    # Ensure DataFrames for each
    boot_params_df = ensure_dataframe(boot_params_df)
    boot_se_df     = ensure_dataframe(boot_se_df)
    jk_params_df   = ensure_dataframe(jk_params_df)
    jk_se_df       = ensure_dataframe(jk_se_df)

    # A helper to handle single-column DataFrames or empty DataFrames
    def safe_mean_to_dict(df_in):
        """
        Returns {} if df_in.empty,
        else returns mean of each column (rounded) as a dict.
        If df_in has only one column, .mean(axis=0) is still a Series with 1 entry.
        """
        if df_in.empty:
            return {}

        # mean_result is always a Series if we are sure df_in is a DataFrame
        mean_result = df_in.mean(axis=0)  # compute along columns
        # mean_result is now a Series indexed by column names
        return mean_result.round(4).to_dict()

    # Use the helper instead of direct .mean(...).to_dict()
    param_mean_boot = safe_mean_to_dict(boot_params_df)
    param_std_boot  = safe_mean_to_dict(boot_se_df)
    param_mean_jk   = safe_mean_to_dict(jk_params_df)
    param_std_jk    = safe_mean_to_dict(jk_se_df)

    # Build 'boot_info' with original + bootstrap param data
    boot_info = {
        "ParamMean_Boot": param_mean_boot,
        "ParamStd_Boot":  param_std_boot,
        "ParamMean_JK":   param_mean_jk,
        "ParamStd_JK":    param_std_jk
    }

    # Store original parameter estimates & SEs
    orig_params = final_fit.params.to_dict()
    orig_se     = final_fit.bse.to_dict()
    boot_info["OriginalParams"] = orig_params
    boot_info["OriginalSE"]     = orig_se

    # -------------------------------------------------------------------
    # 4) Build pairwise differences => param(B) - param(A)
    #    (exclude intercept in non-reference comparisons)
    # -------------------------------------------------------------------
    params = final_fit.params
    cov    = final_fit.cov_params()

    mdf  = final_fit.model.data.frame
    cats = mdf["media_category"].unique()
    if pd.api.types.is_categorical_dtype(mdf["media_category"]):
        cats = mdf["media_category"].cat.categories

    exog_names = final_fit.model.exog_names  # e.g. ["Intercept","media_category[T.Left]",...]
    idx_map = {}
    for c in cats:
        nm = f"media_category[T.{c}]"
        idx_map[c] = exog_names.index(nm) if nm in exog_names else None

    pair_list = []
    cat_list = list(cats)
    n_cats   = len(cat_list)

    for i2 in range(n_cats):
        for j2 in range(i2 + 1, n_cats):
            ca = cat_list[i2]
            cb = cat_list[j2]

            a_idx = idx_map[ca]
            b_idx = idx_map[cb]

            # param(A) = 0 if A is reference
            param_a = params[a_idx] if a_idx is not None else 0.0
            param_b = params[b_idx] if b_idx is not None else 0.0
            diff_est = param_b - param_a

            # Build contrast vector
            con = np.zeros(len(params))
            if b_idx is not None:
                con[b_idx] += 1.0
            if a_idx is not None:
                con[a_idx] -= 1.0

            diff_var = con @ cov @ con
            if diff_var <= 1e-12:
                diff_se = np.nan
                z_val   = np.nan
                p_val   = np.nan
            else:
                diff_se = float(np.sqrt(diff_var))
                z_val   = diff_est / diff_se
                p_val   = 2.0 * (1.0 - norm.cdf(abs(z_val)))

            pair_list.append((ca, cb, diff_est, diff_se, z_val, p_val))

    pair_df = pd.DataFrame(
        pair_list,
        columns=["CategoryA","CategoryB","Difference","SE","Z","p_value"]
    )

    # -------------------------------------------------------------------
    # 5) BH-correct p-values => store p_value_adj, reject_H0
    # -------------------------------------------------------------------
    if not pair_df.empty:
        rej, p_adj, _, _ = multipletests(pair_df["p_value"], method="fdr_bh")
        pair_df["p_value_adj"] = p_adj
        pair_df["reject_H0"]   = rej
    else:
        pair_df["p_value_adj"] = np.nan
        pair_df["reject_H0"]   = False

    # -------------------------------------------------------------------
    # 6) Build DiffIDs => which categories differ => plus Trans_DiffIDs
    # -------------------------------------------------------------------
    cat_index_map = {cat_list[i]: i + 1 for i in range(n_cats)}
    diff_map      = {c: set() for c in cat_list}

    for _, row3 in pair_df.iterrows():
        A = row3["CategoryA"]
        B = row3["CategoryB"]
        if row3["reject_H0"]:
            diff_map[A].add(cat_index_map[B])
            diff_map[B].add(cat_index_map[A])

    rows = []
    for c in cat_list:
        diffs = sorted(diff_map[c])
        diffs_str = ",".join(str(x) for x in diffs)
        rows.append((c, diffs_str))

    diffIDs_df = pd.DataFrame(rows, columns=["Category","DiffIDs"])

    # Mapping => 1->3, 2->1, 3->5, 4->2, 5->6, 6->4
    mapping = {1:3, 2:1, 3:5, 4:2, 5:6, 6:4}

    def transform_diff_ids(val):
        if pd.isna(val) or not val.strip():
            return ""
        parts = [p.strip() for p in val.split(",") if p.strip()]
        mapped = []
        for p in parts:
            if p.isdigit():
                old_num = int(p)
                new_num = mapping.get(old_num, old_num)
                mapped.append(new_num)
        mapped_sorted = sorted(mapped)
        return ",".join(str(x) for x in mapped_sorted)

    diffIDs_df["Trans_DiffIDs"] = diffIDs_df["DiffIDs"].apply(transform_diff_ids)

    # -------------------------------------------------------------------
    # 7) Return the usual 8-tuple
    # -------------------------------------------------------------------
    return (
        summary_txt,       # textual summary (GEE summary)
        pair_df,           # pairwise table (with p_value, p_value_adj, reject_H0)
        diffIDs_df,        # includes DiffIDs + Trans_DiffIDs
        diag_dict,         # residual checks
        pseudo_dict,       # pseudo-likelihood check
        cv_val,            # cross-validation MSE
        sens_df,           # sensitivity analysis
        boot_info          # dict with cluster + jackknife param means/std
    )


def compile_results_into_multiple_workbooks(
    aggregated_df, stats_df, raw_df,
    df_best_qic, df_all_combos,
    plots_dir,
    main_excel, raw_excel, gee_excel, plots_excel, combined_excel,
    df_full,
    validation_records=None
):
    from openpyxl import Workbook, load_workbook
    from openpyxl.utils.dataframe import dataframe_to_rows

    if validation_records is None:
        validation_records = []

    # 1) main => aggregated + stats
    with pd.ExcelWriter(main_excel, engine="openpyxl") as w:
        aggregated_df.to_excel(w, "Aggregated_Scores", index=False)
        stats_df.to_excel(w, "Mean_Median_Statistics", index=False)

    # 2) raw => raw
    with pd.ExcelWriter(raw_excel, engine="openpyxl") as w:
        raw_df.to_excel(w, "Raw_Data", index=False)
        for s in CATEGORIES:
            s_cols = [c for c in raw_df.columns if c.startswith(s + "_")]
            s_df = raw_df[["media_category", "media_outlet"] + s_cols].copy()
            s_df.to_excel(w, f"Raw_{s[:29]}", index=False)

    diag_records = []
    cv_records = []
    sens_records = []
    boot_records = []
    idx_rows = []

    # ---- Helper function to build JK/Boot comparison DF for each (sentiment, measure) ----
    def _build_jk_boot_comparison(boot_info_dict, sent, meas, fam, st, sc):
        """
        Return a DataFrame with columns:
         - Sentiment
         - Measure
         - Family
         - Structure
         - Scale
         - ParamName
         - Original_Est
         - Original_SE
         - Boot_Mean
         - Boot_SD
         - JK_Mean
         - JK_SD
         - Problem_Flag
        """
        p_boot_mean = boot_info_dict.get("ParamMean_Boot", {})
        p_boot_std  = boot_info_dict.get("ParamStd_Boot", {})
        p_jk_mean   = boot_info_dict.get("ParamMean_JK", {})
        p_jk_std    = boot_info_dict.get("ParamStd_JK", {})
        p_orig_est  = boot_info_dict.get("OriginalParams", {})
        p_orig_se   = boot_info_dict.get("OriginalSE", {})

        rows = []
        for param_name, orig_val in p_orig_est.items():
            orig_se = p_orig_se.get(param_name, float('nan'))

            bmean = p_boot_mean.get(param_name, float('nan'))
            bstd  = p_boot_std.get(param_name, float('nan'))

            jkmean = p_jk_mean.get(param_name, float('nan'))
            jkstd  = p_jk_std.get(param_name, float('nan'))

            # Evaluate potential problems
            problem_flag = ""
            if abs(orig_val) > 1e-12:
                ratio_boot = abs(bmean - orig_val) / abs(orig_val)
                ratio_jk   = abs(jkmean - orig_val)/ abs(orig_val)
                if ratio_boot > 0.5 or ratio_jk > 0.5:
                    problem_flag += "LargeShift;"
            else:
                ratio_boot = float('nan')
                ratio_jk   = float('nan')

            # sign flips
            if bmean * orig_val < 0:
                problem_flag += " SignFlipBoot;"
            if jkmean * orig_val < 0:
                problem_flag += " SignFlipJK;"

            rows.append({
                "Sentiment": sent,
                "Measure": meas,
                "Family": fam,
                "Structure": st,
                "Scale": sc,
                "ParamName": param_name,
                "Original_Est": orig_val,
                "Original_SE":  orig_se,
                "Boot_Mean":    bmean,
                "Boot_SD":      bstd,
                "JK_Mean":      jkmean,
                "JK_SD":        jkstd,
                "Problem_Flag": problem_flag.strip()
            })

        return pd.DataFrame(rows)

    # 3) GEE => best combos + final pairwise
    with pd.ExcelWriter(gee_excel, engine="openpyxl") as writer:
        if not df_all_combos.empty:
            df_all_combos.to_excel(writer, sheet_name="All_Combos", index=False)

        # We'll store all "JK_Boot_Comparison" rows in memory, then write them at the end.
        all_jk_boot_records = []

        for i, row in df_best_qic.iterrows():
            s = row["Sentiment"]
            meas = row["Measure"]
            fam = row["Best_Family"]
            st = row["Best_Structure"]
            sc = row["Best_Scale"]
            best_qic = row["Best_QIC_main"]

            sh_name = f"BestQIC_{s[:8]}_{meas[:12]}"
            out = pairwise_and_diagnostics(df_full, s, meas, fam, st, sc, prefix="", isSens=False)
            if out[0] is None:
                tmp_df = pd.DataFrame({"Summary": [f"No valid model => {s}-{meas}"]})
                tmp_df.to_excel(writer, sheet_name=sh_name, index=False)
                continue

            summary_txt, pair_df, diffIDs_df, diag_dict, pseudo_dict, cv_val, sens_df_df, boot_info = out

            lines = summary_txt.split("\n")
            sumdf = pd.DataFrame({"GEE_Summary": lines})
            sumdf.to_excel(writer, sheet_name=sh_name, index=False)

            sr = len(sumdf) + 2
            pair_df.to_excel(writer, sheet_name=sh_name, index=False, startrow=sr)

            sr2 = sr + len(pair_df) + 2
            if diffIDs_df is not None and not diffIDs_df.empty:
                diffIDs_df.to_excel(writer, sheet_name=sh_name, index=False, startrow=sr2)

            # validations
            validation_records.append({
                "Check": "GEE_Residual_Check",
                "Sentiment": s,
                "Measure": meas,
                "Value": f"Overdisp={diag_dict['overdisp_ratio']:.2f}, Corr={diag_dict['avg_within_corr']:.2f}",
                "Conclusion": diag_dict["assessment"]
            })
            validation_records.append({
                "Check": "PseudoLik_Check",
                "Sentiment": s,
                "Measure": meas,
                "Value": f"NB_QIC={pseudo_dict['NB_QIC']}, Poisson_QIC={pseudo_dict['Poisson_QIC']}, diff={pseudo_dict['diff_QIC']}",
                "Conclusion": pseudo_dict["conclusion"]
            })
            validation_records.append({
                "Check": "CrossVal_MSE",
                "Sentiment": s,
                "Measure": meas,
                "Value": cv_val,
                "Conclusion": "No direct pass/fail"
            })

            diag_records.append({
                "Sentiment": s,
                "Measure": meas,
                "Family": fam,
                "Structure": st,
                "Scale": sc,
                "mean_pearson_res": diag_dict["mean_pearson_res"],
                "std_pearson_res": diag_dict["std_pearson_res"],
                "avg_within_corr": diag_dict["avg_within_corr"],
                "overdisp_ratio": diag_dict["overdisp_ratio"],
                "assessment": diag_dict["assessment"],
                "NB_QIC": pseudo_dict["NB_QIC"],
                "Poisson_QIC": pseudo_dict["Poisson_QIC"],
                "diff_QIC": pseudo_dict["diff_QIC"],
                "pseudo_conclusion": pseudo_dict["conclusion"]
            })

            sens_records.append(sens_df_df)

            pm_str = f"Boot: {boot_info.get('ParamMean_Boot', {})} | JK: {boot_info.get('ParamMean_JK', {})}"
            ps_str = f"Boot: {boot_info.get('ParamStd_Boot', {})}  | JK: {boot_info.get('ParamStd_JK', {})}"

            boot_records.append({
                "Sentiment": s,
                "Measure": meas,
                "Family": fam,
                "Structure": st,
                "Scale": sc,
                "ParamMean": pm_str,
                "ParamStd": ps_str
            })

            idx_rows.append({
                "Sentiment": s,
                "Measure": meas,
                "Family": fam,
                "Structure": st,
                "Scale": sc,
                "BestQIC": best_qic
            })

            # ----- NEW: build the direct JK/Boot comparison DF for this run ------
            jk_boot_df = _build_jk_boot_comparison(boot_info, s, meas, fam, st, sc)
            if not jk_boot_df.empty:
                all_jk_boot_records.append(jk_boot_df)

        idxdf = pd.DataFrame(idx_rows)
        idxdf.to_excel(writer, "BestQIC_Index", index=False)

        diag_df = pd.DataFrame(diag_records)
        diag_df.to_excel(writer, "Diagnostics", index=False)

        if len(sens_records) > 0:
            sens_all = pd.concat(sens_records, ignore_index=True)
            sens_all.to_excel(writer, "Sensitivity_Analysis", index=False)

        if len(boot_records) > 0:
            boot_all = pd.DataFrame(boot_records)
            boot_all.to_excel(writer, "Bootstrap_Res", index=False)

        val_df = pd.DataFrame(validation_records)
        val_df.to_excel(writer, "ValidationSummary", index=False)
        row_ = len(val_df) + 2
        conclusion_row = pd.DataFrame({
            "Brief_Conclusion": [
                "Above table shows each validation check. If any check indicates 'WARNING', 'Potential', or 'High imbalance', caution is advised."
            ]
        })
        conclusion_row.to_excel(writer, "ValidationSummary", index=False, startrow=row_)

        # ------ Finally, create the new JK_Boot_Comparison tab! --------
        if len(all_jk_boot_records) > 0:
            big_jk_boot = pd.concat(all_jk_boot_records, ignore_index=True)
            big_jk_boot.to_excel(writer, "JK_Boot_Comparison", index=False)
        else:
            tmp = pd.DataFrame({"NoData": ["No jackknife/bootstrap data found"]})
            tmp.to_excel(writer, "JK_Boot_Comparison", index=False)

    # Compare_Imbalance => param-level
    ci_df = pd.DataFrame(COMPARE_IMBALANCE)
    if os.path.exists(gee_excel):
        book = load_workbook(gee_excel)
    else:
        book = Workbook()
        if "Sheet" in book.sheetnames:
            book.remove(book["Sheet"])
    ws_ci = book.create_sheet("Compare_Imbalance")
    if not ci_df.empty:
        ci_rows = dataframe_to_rows(ci_df, index=False, header=True)
        for r_i, row_val in enumerate(ci_rows, 1):
            for c_i, cell_val in enumerate(row_val, 1):
                ws_ci.cell(row=r_i, column=c_i, value=cell_val)
    else:
        ws_ci["A1"] = "No cluster-imbalance data"

    if "Sheet" in book.sheetnames and len(book.sheetnames) > 1:
        if "Sheet" in book.sheetnames:
            dum = book["Sheet"]
            book.remove(dum)
    book.save(gee_excel)

    # 4) produce plots => plots_excel
    wb_plots = Workbook()
    if "Sheet" in wb_plots.sheetnames:
        wb_plots.remove(wb_plots["Sheet"])
    any_sheets = False
    for s in CATEGORIES:
        q_path = os.path.join(plots_dir, f"quote_{s}.png")
        if os.path.exists(q_path):
            st = f"Quote_{s[:28]}"
            ws = wb_plots.create_sheet(title=st)
            try:
                img = ExcelImage(q_path)
                img.anchor = "A1"
                ws.add_image(img)
            except:
                pass
            any_sheets = True
        f_path = os.path.join(plots_dir, f"fulltext_{s}.png")
        if os.path.exists(f_path):
            st2 = f"Fulltext_{s[:25]}"
            ws2 = wb_plots.create_sheet(title=st2)
            try:
                img2 = ExcelImage(f_path)
                img2.anchor = "A1"
                ws2.add_image(img2)
            except:
                pass
            any_sheets = True

    if not any_sheets:
        wb_plots.create_sheet("DummySheet")
    wb_plots.save(plots_excel)

    # 5) combined => aggregator + best QIC + validations
    raw_clean = raw_df.copy()
    raw_clean = raw_clean.applymap(lambda x: ", ".join(x) if isinstance(x, list) else x)
    wb_comb = Workbook()
    if "Sheet" in wb_comb.sheetnames:
        wb_comb.remove(wb_comb["Sheet"])

    ws_agg = wb_comb.create_sheet("Aggregated_Scores")
    for r in dataframe_to_rows(aggregated_df, index=False, header=True):
        ws_agg.append(r)
    ws_stats = wb_comb.create_sheet("Mean_Median_Statistics")
    for r in dataframe_to_rows(stats_df, index=False, header=True):
        ws_stats.append(r)
    ws_raw = wb_comb.create_sheet("Raw_Data")
    for r in dataframe_to_rows(raw_clean, index=False, header=True):
        ws_raw.append(r)
    ws_best = wb_comb.create_sheet("BestQIC_Table")
    for r in dataframe_to_rows(df_best_qic, index=False, header=True):
        ws_best.append(r)
    ws_val = wb_comb.create_sheet("ValidationSummary")
    val_df2 = pd.DataFrame(validation_records)
    for row_ in dataframe_to_rows(val_df2, index=False, header=True):
        ws_val.append(row_)
    lr2 = len(val_df2) + 2
    ws_val.cell(row=lr2, column=1, value=(
        "Above table (old + new validations) for all checks. "
        "Look for warnings or potential issues."
    ))
    wb_comb.save(combined_excel)

###############################################################################
# exclude_small_clusters
###############################################################################
def exclude_small_clusters_within_category(
    df,
    cat_col="media_category",
    group_col="media_outlet_clean",
    min_size=5,
    merge_into="OtherSmall"
):
    """
    For each category in 'cat_col', find all outlets (group_col) that have < min_size
    articles in that category. Then merge only those small outlets
    into 'merge_into' for that category only.

    Example:
      If 'min_size' = 5, and for category='Left' we have multiple outlets
      but 'nytimes' appears 3 times => 'nytimes' is replaced with 'OtherSmall'
      *only for that category*.

    Returns a modified DataFrame with that new group_col merging logic.
    """
    df2 = df.copy()

    # For each unique category, figure out which outlets are small (size < min_size)
    categories = df2[cat_col].dropna().unique()
    for cat in categories:
        # Subset just this category
        mask_cat = (df2[cat_col] == cat)
        sub = df2[mask_cat]
        # Count the outlets
        counts = sub[group_col].value_counts(dropna=False)
        small_outlets = counts[counts < min_size].index

        # Now, for only those outlets, we rename them to 'OtherSmall'
        # but only within this category.
        mask_small = (df2[group_col].isin(small_outlets)) & mask_cat
        df2.loc[mask_small, group_col] = merge_into

    return df2


###############################################################################
# append_sens_to_excel
###############################################################################
def append_sens_to_excel(
    aggregated_df_sens, 
    stats_df_sens, 
    raw_df_sens,
    df_best_qic_sens, 
    df_all_combos_sens,
    prefix,
    outdir,
    main_excel,
    raw_excel,
    gee_excel,
    combined_excel,
    df_full_sens
):
    """
    Extends existing Excel files with 'SENS' results and ALSO runs pairwise comparisons
    on the 'OtherSmall' merged dataset, storing them in new sheets.

    This way, you can directly compare the normal dataset's pairwise
    vs. the SENS merged dataset's pairwise in the same workbook.
    """
    from openpyxl import load_workbook
    from openpyxl.utils.dataframe import dataframe_to_rows

    # -------------------------------------------------------------------
    # 1) Append aggregated + stats to main_excel
    # -------------------------------------------------------------------
    if os.path.exists(main_excel):
        wb=load_workbook(main_excel)
    else:
        wb=Workbook()
        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])
    ws_agg=wb.create_sheet("Aggregated_Scores_Sens")
    for r in dataframe_to_rows(aggregated_df_sens, index=False, header=True):
        ws_agg.append(r)
    ws_stats=wb.create_sheet("Mean_Median_Statistics_Sens")
    for r in dataframe_to_rows(stats_df_sens, index=False, header=True):
        ws_stats.append(r)
    wb.save(main_excel)

    # -------------------------------------------------------------------
    # 2) Append raw => Raw_Data_Sens
    # -------------------------------------------------------------------
    if os.path.exists(raw_excel):
        wb2=load_workbook(raw_excel)
    else:
        wb2=Workbook()
        if "Sheet" in wb2.sheetnames:
            wb2.remove(wb2["Sheet"])
    ws_raw=wb2.create_sheet("Raw_Data_Sens")
    raw_clean_sens=raw_df_sens.copy()
    raw_clean_sens=raw_clean_sens.applymap(lambda x:", ".join(x) if isinstance(x,list) else x)
    for r in dataframe_to_rows(raw_clean_sens, index=False, header=True):
        ws_raw.append(r)
    wb2.save(raw_excel)

    # -------------------------------------------------------------------
    # 3) GEE => combos, best qic index for Sens
    # -------------------------------------------------------------------
    if os.path.exists(gee_excel):
        wb3=load_workbook(gee_excel)
    else:
        wb3=Workbook()
        if "Sheet" in wb3.sheetnames:
            wb3.remove(wb3["Sheet"])

    # Write "All_Combos_Sens" if we have them
    if not df_all_combos_sens.empty:
        ws_comb=wb3.create_sheet("All_Combos_Sens")
        combos_rows=dataframe_to_rows(df_all_combos_sens, index=False, header=True)
        for r in combos_rows:
            ws_comb.append(r)

    # Write "BestQIC_Index_Sens"
    ws_idx=wb3.create_sheet("BestQIC_Index_Sens")
    if not df_best_qic_sens.empty:
        idx_rows=dataframe_to_rows(df_best_qic_sens, index=False, header=True)
        for r in idx_rows:
            ws_idx.append(r)
    else:
        ws_idx["A1"]="No best QIC data for SENS"

    # -------------------------------------------------------------------
    # 4) **NEW**: Run pairwise_and_diagnostics for SENS data => store
    #    in new sheets, named "BestQIC_Sens_{sent}_{meas}" etc.
    # -------------------------------------------------------------------
    # We'll replicate the logic from compile_results_into_multiple_workbooks
    # but specifically for the 'df_full_sens' dataset.
    from . import pairwise_and_diagnostics  # If you have a local module import
    # (If in the same file, no import needed.)

    for i, row in df_best_qic_sens.iterrows():
        s = row["Sentiment"]
        meas = row["Measure"]
        fam = row["Best_Family"]
        st = row["Best_Structure"]
        sc = row["Best_Scale"]
        best_qic_val = row["Best_QIC_main"]

        sens_sheet_name = f"BestQIC_Sens_{s[:8]}_{meas[:12]}"
        out = pairwise_and_diagnostics(df_full_sens, s, meas, fam, st, sc, prefix="Sens", isSens=True)
        if out[0] is None:
            tmp_df = pd.DataFrame({"Summary": [f"No valid SENS model => {s}-{meas}"]})
            tmp_df.to_excel(wb3, sheet_name=sens_sheet_name, index=False)
            continue

        summary_txt, pair_df, diffIDs_df, diag_dict, pseudo_dict, cv_val, sens_df_df, boot_info = out

        # Create that sheet
        ws_sens = wb3.create_sheet(sens_sheet_name)

        # 1) summary
        lines = summary_txt.split("\n")
        row_ctr = 1
        ws_sens.cell(row=row_ctr, column=1, value="GEE_Summary")
        row_ctr += 1
        for ln in lines:
            ws_sens.cell(row=row_ctr, column=1, value=ln)
            row_ctr += 1
        row_ctr += 1

        # 2) pair_df
        if not pair_df.empty:
            # Convert to rows
            pdf_rows = dataframe_to_rows(pair_df, index=False, header=True)
            for rdata in pdf_rows:
                for c_i, cell_val in enumerate(rdata, 1):
                    ws_sens.cell(row=row_ctr, column=c_i, value=cell_val)
                row_ctr += 1
            row_ctr += 2

        # 3) diffIDs_df
        if diffIDs_df is not None and not diffIDs_df.empty:
            did_rows = dataframe_to_rows(diffIDs_df, index=False, header=True)
            for rdata in did_rows:
                for c_i, cell_val in enumerate(rdata, 1):
                    ws_sens.cell(row=row_ctr, column=c_i, value=cell_val)
                row_ctr += 1
            row_ctr += 2

        # (Optionally, store diag_dict, pseudo_dict, etc. in some format)
        # We'll just store them in text form below
        diag_str = f"SENS Overdisp={diag_dict['overdisp_ratio']}, Corr={diag_dict['avg_within_corr']}, Assess={diag_dict['assessment']}"
        ws_sens.cell(row=row_ctr, column=1, value=diag_str)
        row_ctr += 2

        pseudo_str = f"SENS NB_QIC={pseudo_dict['NB_QIC']}, Poisson_QIC={pseudo_dict['Poisson_QIC']}, diff={pseudo_dict['diff_QIC']}, conclusion={pseudo_dict['conclusion']}"
        ws_sens.cell(row=row_ctr, column=1, value=pseudo_str)
        row_ctr += 2

        crossval_str = f"SENS CrossVal MSE={cv_val}"
        ws_sens.cell(row=row_ctr, column=1, value=crossval_str)
        row_ctr += 2

        # You can do the same for boot_info if needed.

    # End loop over df_best_qic_sens
    # Finally, save the GEE Excel with these new tabs
    wb3.save(gee_excel)

    # -------------------------------------------------------------------
    # 5) Append aggregator + stats => combined
    # -------------------------------------------------------------------
    if os.path.exists(combined_excel):
        wb4=load_workbook(combined_excel)
    else:
        wb4=Workbook()
        if "Sheet" in wb4.sheetnames:
            wb4.remove(wb4["Sheet"])

    ws_agg2=wb4.create_sheet("Aggregated_Scores_Sens")
    for r in dataframe_to_rows(aggregated_df_sens, index=False, header=True):
        ws_agg2.append(r)
    ws_stats2=wb4.create_sheet("Mean_Median_Statistics_Sens")
    for r in dataframe_to_rows(stats_df_sens, index=False, header=True):
        ws_stats2.append(r)
    wb4.save(combined_excel)

###############################################################################
# run_gee_for_sentiment_measure_best_qic
###############################################################################
def run_gee_for_sentiment_measure_best_qic(df, sentiment, measure):
    """
    Finds the best QIC combination (family, cov_struct, scale)
    for a given (sentiment, measure).
    """
    # 1) Make a copy of df
    d2 = df.copy()

    # 2) Build _score_col based on measure type
    if measure == "Quotation":
        pat = rf"^{re.escape(sentiment)}_\d+$"
        matched = [c for c in d2.columns if re.match(pat, c)]
        if not matched:
            return None
        d2["_score_col"] = d2[matched].clip(lower=0).mean(axis=1)

    elif measure == "Fulltext":
        fcol = f"{sentiment}_fulltext"
        if fcol not in d2.columns:
            return None
        d2["_score_col"] = d2[fcol].clip(lower=0)

    elif measure == "Fulltext_Intensity":
        fcol = f"{sentiment}_fulltext_intensity"
        if fcol not in d2.columns:
            return None
        d2["_score_col"] = d2[fcol].astype(float).clip(lower=0)

    elif measure == "Title_Intensity":
        fcol = f"{sentiment}_title_intensity"
        if fcol not in d2.columns:
            return None
        d2["_score_col"] = d2[fcol].astype(float).clip(lower=0)

    elif measure == "Quotation_Intensity":
        pat = rf"^{re.escape(sentiment)}_\d+_intensity$"
        matched = [c for c in d2.columns if re.match(pat, c)]
        if not matched:
            return None
        d2["_score_col"] = d2[matched].astype(float).clip(lower=0).mean(axis=1)

    # NEW BLOCK FOR THE 6TH MEASURE:
    elif measure == "Title":
        fcol = f"{sentiment}_title"
        if fcol not in d2.columns:
            return None
        d2["_score_col"] = d2[fcol].clip(lower=0)

    else:
        return None

    # 3) Ensure we have enough data
    needed = ["_score_col", "media_outlet_clean", "media_category"]
    d2 = d2.dropna(subset=needed)
    if len(d2) < 2 or d2["media_category"].nunique() < 2:
        return None

    # 4) Brute-force families/cov_struct/scale
    best_tuple, all_results = try_all_families_and_scales(
        d2,
        "_score_col ~ media_category",
        "media_outlet_clean"
    )

    if best_tuple is None:
        return None

    famName, corName, scOpt, bestQICVal = best_tuple

    # 5) Build combos_df from 'all_results' with corrected columns
    combos_df = pd.DataFrame(
        all_results,
        columns=["Family", "CovStruct", "Scale", "QIC", "StatusFlag"]
    )

    # 6) Return dictionary including combos_df
    return {
        "Sentiment":       sentiment,
        "Measure":         measure,
        "Best_Family":     famName,
        "Best_Structure":  corName,
        "Best_Scale":      scOpt,
        "Best_QIC_main":   bestQICVal,
        "AllCombos":       combos_df
    }

###############################################################################
# run_gee_analyses_best_qic
###############################################################################
def run_gee_analyses_best_qic(df):
    measure_list = [
        "Quotation",
        "Fulltext",
        "Fulltext_Intensity",
        "Title_Intensity",
        "Quotation_Intensity",
        # NEW MEASURE ADDED:
        "Title"
    ]
    best_list = []
    combos_list = []
    for s in CATEGORIES:
        for meas in measure_list:
            info = run_gee_for_sentiment_measure_best_qic(df, s, meas)
            if info is not None:
                best_list.append({
                    "Sentiment":    info["Sentiment"],
                    "Measure":      info["Measure"],
                    "Best_Family":  info["Best_Family"],
                    "Best_Structure": info["Best_Structure"],
                    "Best_Scale":   info["Best_Scale"],
                    "Best_QIC_main":info["Best_QIC_main"]
                })
                cdf = info["AllCombos"]
                cdf["Sentiment"] = s
                cdf["Measure"]   = meas
                combos_list.append(cdf)
    df_best = pd.DataFrame(best_list)
    df_all  = pd.concat(combos_list, ignore_index=True) if combos_list else pd.DataFrame()
    return df_best, df_all

###############################################################################
# exclude_small_clusters
###############################################################################
def exclude_small_clusters(df, cluster_col="media_outlet_clean", min_size=5, merge_into="OtherSmall"):
    sizes=df[cluster_col].value_counts()
    small_clusters=sizes[sizes<min_size].index
    df2=df.copy()
    df2[cluster_col]=df2[cluster_col].apply(lambda c: merge_into if c in small_clusters else c)
    return df2

###############################################################################
# append_sens_to_excel
###############################################################################
def append_sens_to_excel(
    aggregated_df_sens, stats_df_sens, raw_df_sens,
    df_best_qic_sens, df_all_combos_sens,
    prefix,
    outdir,
    main_excel,
    raw_excel,
    gee_excel,
    combined_excel,
    df_full_sens
):
    from openpyxl import load_workbook
    from openpyxl.utils.dataframe import dataframe_to_rows

    if os.path.exists(main_excel):
        wb=load_workbook(main_excel)
    else:
        wb=Workbook()
        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])
    ws_agg=wb.create_sheet("Aggregated_Scores_Sens")
    for r in dataframe_to_rows(aggregated_df_sens, index=False, header=True):
        ws_agg.append(r)
    ws_stats=wb.create_sheet("Mean_Median_Statistics_Sens")
    for r in dataframe_to_rows(stats_df_sens, index=False, header=True):
        ws_stats.append(r)
    wb.save(main_excel)

    if os.path.exists(raw_excel):
        wb2=load_workbook(raw_excel)
    else:
        wb2=Workbook()
        if "Sheet" in wb2.sheetnames:
            wb2.remove(wb2["Sheet"])
    ws_raw=wb2.create_sheet("Raw_Data_Sens")
    raw_clean_sens=raw_df_sens.copy()
    raw_clean_sens=raw_clean_sens.applymap(lambda x:", ".join(x) if isinstance(x,list) else x)
    for r in dataframe_to_rows(raw_clean_sens, index=False, header=True):
        ws_raw.append(r)
    wb2.save(raw_excel)

    if os.path.exists(gee_excel):
        wb3=load_workbook(gee_excel)
    else:
        wb3=Workbook()
        if "Sheet" in wb3.sheetnames:
            wb3.remove(wb3["Sheet"])

    if not df_all_combos_sens.empty:
        ws_comb=wb3.create_sheet("All_Combos_Sens")
        combos_rows=dataframe_to_rows(df_all_combos_sens, index=False, header=True)
        for r in combos_rows:
            ws_comb.append(r)

    ws_idx=wb3.create_sheet("BestQIC_Index_Sens")
    if not df_best_qic_sens.empty:
        idx_rows=dataframe_to_rows(df_best_qic_sens, index=False, header=True)
        for r in idx_rows:
            ws_idx.append(r)
    else:
        ws_idx["A1"]="No best QIC data for SENS"
    wb3.save(gee_excel)

    if os.path.exists(combined_excel):
        wb4=load_workbook(combined_excel)
    else:
        wb4=Workbook()
        if "Sheet" in wb4.sheetnames:
            wb4.remove(wb4["Sheet"])

    ws_agg2=wb4.create_sheet("Aggregated_Scores_Sens")
    for r in dataframe_to_rows(aggregated_df_sens, index=False, header=True):
        ws_agg2.append(r)
    ws_stats2=wb4.create_sheet("Mean_Median_Statistics_Sens")
    for r in dataframe_to_rows(stats_df_sens, index=False, header=True):
        ws_stats2.append(r)
    wb4.save(combined_excel)

###############################################################################
# run_pipeline_for_df
###############################################################################
def run_pipeline_for_df(
    df,
    prefix="",
    main_excel=BASE_MAIN_FILE,
    raw_excel=BASE_RAW_FILE,
    gee_excel=BASE_GEE_FILE,
    plots_excel=BASE_PLOTS_FILE,
    combined_excel=BASE_COMBINED_FILE,
    outdir=GRAPH_OUTPUT_DIR,
    csvout=CSV_OUTPUT_DIR,
    min_size=5
):
    """
    Runs the entire pipeline on 'df' with a given 'prefix'.
    1) Does chunk_and_save, basic stats, correlation analyses, aggregator stats, plotting, etc.
    2) Finds best QIC combos => pairwise_and_diagnostics => writes to Excel.
    3) Then does a 'SENS' run where we merge small outlets (only within each category)
       into 'OtherSmall'. Compares those results side by side.
    """
    # Rename the output files for this prefix
    main_excel    = rename_files_for_prefix(main_excel,    prefix)
    raw_excel     = rename_files_for_prefix(raw_excel,     prefix)
    gee_excel     = rename_files_for_prefix(gee_excel,     prefix)
    plots_excel   = rename_files_for_prefix(plots_excel,   prefix)
    combined_excel= rename_files_for_prefix(combined_excel,prefix)

    validation_records=[]

    # 1) Some checks
    c1=check_number_of_clusters(df,"media_outlet_clean",20)
    validation_records.append(c1)
    c2=check_cluster_balance(df,"media_outlet_clean",5.0)
    validation_records.append(c2)
    c_cov=coverage_simulation(n_sims=1000,n_clusters=10,cluster_size=10,true_beta=0.5)
    validation_records.append(c_cov)

    # 2) If prefix indicates "Yes", collapse categories => single cluster
    if prefix=="Yes" or (prefix.endswith("_collapsed") and prefix.startswith("Yes_")):
        cat_codes, unique_cats = pd.factorize(df["media_category"])
        df["media_outlet_clean"] = cat_codes + 1
        logging.info(
            f"Collapsed each category => single cluster. #unique cats={len(unique_cats)}"
        )

    # 3) Chunk + basic stats
    chunk_and_save(df, 20000, prefix=prefix)
    print_basic_stats(df, prefix=prefix)

    # 4) Correlations, aggregator, stats, plots
    analyze_all_custom_correlations(df, prefix=prefix)
    agg_df = aggregate_sentiment_scores(df, CATEGORIES)
    agg_df = calculate_averages(agg_df)
    stats_df = calculate_mean_median(agg_df)
    save_aggregated_scores_to_csv(agg_df, csvout, prefix=prefix)
    plot_statistics(agg_df, outdir, prefix=prefix)

    # 5) Run GEE analyses => best combos
    df_best, df_allcombos = run_gee_analyses_best_qic(df)
    print(f"Best QIC => prefix={prefix}")
    print(df_best)

    # 6) Compile results => multiple workbooks
    compile_results_into_multiple_workbooks(
        aggregated_df=agg_df,
        stats_df=stats_df,
        raw_df=df,
        df_best_qic=df_best,
        df_all_combos=df_allcombos,
        plots_dir=outdir,
        main_excel=main_excel,
        raw_excel=raw_excel,
        gee_excel=gee_excel,
        plots_excel=plots_excel,
        combined_excel=combined_excel,
        df_full=df,
        validation_records=validation_records
    )

    # 7) Now the SENS run => merges small clusters, but only within each category
    from copy import deepcopy
    df_sens = exclude_small_clusters_within_category(
        df=deepcopy(df),
        cat_col="media_category",
        group_col="media_outlet_clean",
        min_size=min_size,
        merge_into="OtherSmall"
    )
    if len(df_sens)<2:
        print(f"[Sensitivity] => merging <{min_size} => not enough data => skip.")
        return

    print(f"[Sensitivity] => merging <{min_size}, prefix={prefix}, len={len(df_sens)} => SENS run")

    agg_sens = aggregate_sentiment_scores(df_sens, CATEGORIES)
    agg_sens = calculate_averages(agg_sens)
    stats_sens= calculate_mean_median(agg_sens)

    df_best_sens, df_allcombos_sens = run_gee_analyses_best_qic(df_sens)
    print(f"Best QIC => Sens run, prefix={prefix}")
    print(df_best_sens)

    # 8) Save aggregated => with suffix
    save_aggregated_scores_to_csv(agg_sens, csvout, prefix=f"{prefix}_Sens")

    # 9) Append 'SENS' to Excel => side by side
    append_sens_to_excel(
        aggregated_df_sens=agg_sens,
        stats_df_sens=stats_sens,
        raw_df_sens=df_sens,
        df_best_qic_sens=df_best_sens,
        df_all_combos_sens=df_allcombos_sens,
        prefix=prefix,
        outdir=outdir,
        main_excel=main_excel,
        raw_excel=raw_excel,
        gee_excel=gee_excel,
        combined_excel=combined_excel,
        df_full_sens=df_sens
    )

###############################################################################
# main
###############################################################################
def main():
    setup_logging()
    logging.info("Starting pipeline => yes/all => original cats, then yes/all => collapsed cats")

    df_raw=load_jsonl(INPUT_JSONL_FILE)
    print(f"Total articles loaded => {len(df_raw)}")

    if "high_rate_2" in df_raw.columns:
        df_raw["high_rate_2"]=df_raw["high_rate_2"].astype(str).str.strip().str.lower()

    # A) Original categories => yes / all
    print("\n=== A) Original Categories (6-cat) ===")
    df_orig=map_media_outlet_to_category(df_raw, MEDIA_CATEGORIES_ORIG)

    if "high_rate_2" in df_orig.columns:
        df_yes_orig=df_orig[df_orig["high_rate_2"]=="yes"].copy()
        print(f"df_yes_orig => {len(df_yes_orig)} rows")
        if len(df_yes_orig)>0:
            print("**Pipeline => Yes** (orig cats)")
            run_pipeline_for_df(df_yes_orig, prefix="Yes")
        else:
            print("No rows => skip yes subset (orig cats)")
    else:
        print("No 'high_rate_2' => skip yes subset (orig cats)")

    print("**Pipeline => All** (orig cats)")
    run_pipeline_for_df(df_orig, prefix="All")
    del df_orig
    gc.collect()

    # B) Collapsed categories => yes_collapsed / all_collapsed
    print("\n=== B) Collapsed Categories (4-cat) ===")
    df_coll=map_media_outlet_to_category(df_raw, MEDIA_CATEGORIES_COLLAPSED)

    if "high_rate_2" in df_coll.columns:
        df_yes_coll=df_coll[df_coll["high_rate_2"]=="yes"].copy()
        print(f"df_yes_collapsed => {len(df_yes_coll)} rows")
        if len(df_yes_coll)>0:
            print("**Pipeline => Yes_collapsed**")
            run_pipeline_for_df(df_yes_coll, prefix="Yes_collapsed")
        else:
            print("No rows => skip yes subset (collapsed cats)")
    else:
        print("No 'high_rate_2' => skip yes subset (collapsed cats)")

    print("**Pipeline => All_collapsed**")
    run_pipeline_for_df(df_coll, prefix="All_collapsed")
    del df_coll
    gc.collect()

    logging.info("All done => subset + full => original cats + collapsed cats.")

if __name__=="__main__":
    main()
